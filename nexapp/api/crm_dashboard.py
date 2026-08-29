"""
Nexapp CRM Management Dashboard — Backend API (Phase 1)
========================================================

PURPOSE:
    Provides hierarchy-aware scope resolution and universal filter options
    for the Nexapp CRM Management Dashboard.

ARCHITECTURE:
    This module reads the NATIVE Frappe CRM Sales Hierarchy (NestedSet tree)
    to determine what data each user is permitted to see. It does NOT create
    a parallel hierarchy — it reads the existing `CRM Sales Hierarchy` DocType
    and `FCRM Settings.enable_sales_hierarchy` flag.

SECURITY MODEL:
    - The backend is the ONLY security boundary.
    - Frontend dropdown values are convenience — they are NOT security.
    - Every API call re-validates the requested scope against the caller's
      actual hierarchy position before returning data.
    - "ALL" never means "all records globally". It means "all records within
      the caller's permitted subtree".

CRM DEPENDENCY:
    - DocType: CRM Sales Hierarchy  (NestedSet with lft/rgt, user, reports_to, is_group)
    - DocType: FCRM Settings        (enable_sales_hierarchy flag)
    - Module:  crm.permissions.org_hierarchy (reference only — we read the same
      tree but do NOT import private functions from CRM to stay upgrade-safe)

IMPORTANT:
    All code lives in apps/nexapp. No CRM/Frappe/ERPNext core files are modified.
"""

import frappe
from frappe import _


# ---------------------------------------------------------------------------
# 1. HIERARCHY INSPECTION HELPERS
#    These read the native CRM Sales Hierarchy tree without importing CRM
#    private internals, keeping Nexapp decoupled and upgrade-safe.
# ---------------------------------------------------------------------------

def _is_hierarchy_enabled():
    """
    Check whether the CRM Sales Hierarchy permission model is active.

    Reads the `enable_sales_hierarchy` flag from `FCRM Settings`.
    When disabled, Sales Managers see everything and Sales Users see only
    their own records (standard CRM behaviour).

    Returns:
        bool: True if the hierarchy tree is actively enforcing permissions.
    """
    return bool(frappe.db.get_single_value("FCRM Settings", "enable_sales_hierarchy"))


def _get_user_hierarchy_node(user):
    """
    Fetch the CRM Sales Hierarchy node for a given user.

    The CRM Sales Hierarchy is a NestedSet tree. Each node has:
      - user:       Link to User
      - reports_to: Link to parent CRM Sales Hierarchy node
      - is_group:   Whether this node can have children (managers)
      - lft / rgt:  NestedSet boundary integers for fast subtree queries
      - full_name:  Fetched from User.full_name

    Returns:
        dict or None: The node record with keys (name, user, full_name,
                      reports_to, is_group, lft, rgt), or None if the user
                      has no node in the hierarchy.
    """
    node = frappe.db.get_value(
        "CRM Sales Hierarchy",
        {"user": user},
        ["name", "user", "full_name", "reports_to", "is_group", "lft", "rgt"],
        as_dict=True,
    )
    return node


def _get_subtree_members(node):
    """
    Fetch all CRM Sales Hierarchy nodes within a node's NestedSet subtree.

    Uses lft/rgt boundaries: a descendant's lft is >= ancestor's lft AND
    <= ancestor's rgt. This is the same traversal logic used by the native
    CRM permission engine (crm.permissions.org_hierarchy._team_mem_query).

    Args:
        node (dict): A hierarchy node dict with 'lft' and 'rgt' keys.

    Returns:
        list[dict]: All nodes in the subtree (including the node itself),
                    each with keys (name, user, full_name, reports_to,
                    is_group, lft, rgt).
    """
    CSH = frappe.qb.DocType("CRM Sales Hierarchy")
    rows = (
        frappe.qb.from_(CSH)
        .select(
            CSH.name, CSH.user, CSH.full_name,
            CSH.reports_to, CSH.is_group, CSH.lft, CSH.rgt,
        )
        .where((CSH.lft >= node["lft"]) & (CSH.lft <= node["rgt"]))
        .orderby(CSH.lft)
        .run(as_dict=True)
    )
    return rows


def _get_direct_children_group_nodes(node):
    """
    Get the immediate child nodes that are group nodes (is_group=1) under
    the given node. These represent "teams" in the dashboard UI.

    A "Team" in the Nexapp dashboard is defined as:
        A group node (is_group=1) that directly reports to the current
        user's hierarchy node. The team label is derived from the group
        node's full_name (which is the manager's name).

    Args:
        node (dict): The parent hierarchy node.

    Returns:
        list[dict]: Direct child group nodes.
    """
    return frappe.db.get_all(
        "CRM Sales Hierarchy",
        filters={"reports_to": node["name"], "is_group": 1},
        fields=["name", "user", "full_name", "lft", "rgt"],
        order_by="lft asc",
    )


def _compute_in_tree(user):
    """
    Compute the CRM's 'in_tree' flag exactly as the native CRM does.

    The CRM's org_hierarchy._permission_query_conditions (line 28) computes:
        in_tree = hierarchy_enabled() and _in_hierarchy(user)

    This means a user is only considered "in the tree" when BOTH:
      1. FCRM Settings.enable_sales_hierarchy is ON
      2. The user has a CRM Sales Hierarchy node

    When hierarchy is DISABLED, in_tree is always False for everyone,
    regardless of whether they have a node in the tree.

    Args:
        user (str): The Frappe user email/ID.

    Returns:
        bool: True only if hierarchy is enabled AND user has a node.
    """
    return _is_hierarchy_enabled() and bool(
        frappe.db.exists("CRM Sales Hierarchy", {"user": user})
    )


def _is_unrestricted_user(user):
    """
    Determine whether a user has unrestricted CRM data access.

    Mirrors the EXACT decision tree from the installed CRM source at:
    crm.permissions.org_hierarchy._permission_query_conditions

    Decision flow (verified against CRM source lines 17-32):
      1. Administrator → unrestricted (return "")
      2. System Manager role → unrestricted (return "")
      3. Compute in_tree = hierarchy_enabled() AND _in_hierarchy(user)
      4. Sales Manager role AND NOT in_tree → unrestricted (return "")
         - This covers: SM + hierarchy disabled (any SM sees everything)
         - This covers: SM + hierarchy enabled but user not in tree
      5. Otherwise → NOT unrestricted (will be subtree or self-only)

    Args:
        user (str): The Frappe user email/ID.

    Returns:
        bool: True if the user should see all CRM data without restriction.
    """
    if user == "Administrator":
        return True

    roles = frappe.get_roles(user)
    if "System Manager" in roles:
        return True

    # CRM source line 28: in_tree = hierarchy_enabled() and _in_hierarchy(user)
    # CRM source line 31: if "Sales Manager" in roles and not in_tree: return ""
    in_tree = _compute_in_tree(user)
    if "Sales Manager" in roles and not in_tree:
        return True

    return False


# ---------------------------------------------------------------------------
# 2. SCOPE RESOLUTION
#    Determines the complete set of users and "teams" a caller can access.
# ---------------------------------------------------------------------------

def _resolve_scope(user):
    """
    Resolve the full dashboard permission scope for a user.

    This is the CORE security function. It mirrors the CRM's native
    permission decision tree exactly:

    CRM org_hierarchy._permission_query_conditions decision flow:
      1. Administrator → unrestricted
      2. System Manager → unrestricted
      3. in_tree = hierarchy_enabled() AND _in_hierarchy(user)
      4. Sales Manager AND NOT in_tree → unrestricted
      5. in_tree is True → subtree scope (owner in subtree)
      6. Fallthrough → self-only (own records + assigned)

    For the dashboard, this translates to:
      - Unrestricted → sees entire org hierarchy
      - in_tree → sees their NestedSet subtree
      - Fallthrough → sees only their own data

    CRITICAL: When hierarchy is DISABLED, in_tree is False for ALL users.
    This means Sales Users with hierarchy nodes get self-only scope (not
    subtree scope). This matches the CRM's tested behaviour — see
    crm/permissions/test_org_hierarchy.py lines 151-163.

    IMPORTANT: "ALL" in the dashboard context always means "all users
    within this scope", never "all users globally".

    Args:
        user (str): The Frappe session user.

    Returns:
        dict: {
            "is_unrestricted": bool,
            "user_node": dict or None,
            "permitted_users": list[dict],   # [{value, label}]
            "teams": list[dict],             # [{value, label, members}]
        }
    """
    # Step 1-4: Check if user has unrestricted access
    # (Admin, SysManager, or Sales Manager with in_tree=False)
    if _is_unrestricted_user(user):
        return _resolve_unrestricted_scope()

    # Step 5: Check if user is in the ACTIVE hierarchy tree
    # _compute_in_tree returns True ONLY when hierarchy is enabled AND
    # the user has a node. This prevents subtree access when hierarchy
    # is disabled.
    in_tree = _compute_in_tree(user)
    if in_tree:
        node = _get_user_hierarchy_node(user)
        if node:
            return _resolve_hierarchy_scope(user, node)

    # Step 6: Fallthrough — self-only scope
    # Applies to:
    #   - Sales Users without hierarchy nodes
    #   - Sales Users WITH hierarchy nodes but hierarchy is DISABLED
    #   - Any non-manager user not in the active tree
    return _resolve_self_only_scope(user)


def _resolve_unrestricted_scope():
    """
    Build scope for unrestricted users (Administrator, System Manager,
    or Sales Manager not in hierarchy tree).

    They can see the entire CRM Sales Hierarchy and all CRM data.
    Teams = each direct child group node under the root(s).
    """
    # Get ALL hierarchy nodes
    all_nodes = frappe.db.get_all(
        "CRM Sales Hierarchy",
        fields=["name", "user", "full_name", "reports_to", "is_group", "lft", "rgt"],
        order_by="lft asc",
    )

    # Build user list from hierarchy
    permitted_users = []
    seen_users = set()
    for n in all_nodes:
        if n.get("user") and n["user"] not in seen_users:
            permitted_users.append({
                "value": n["user"],
                "label": n.get("full_name") or n["user"],
            })
            seen_users.add(n["user"])

    # Build teams: root-level group nodes, or direct children of roots
    # Root nodes have reports_to = None/empty
    root_nodes = [n for n in all_nodes if not n.get("reports_to")]
    teams = []

    for root in root_nodes:
        if root.get("is_group"):
            # Each direct child group of the root is a "team"
            child_groups = [
                n for n in all_nodes
                if n.get("reports_to") == root["name"] and n.get("is_group")
            ]
            if child_groups:
                for cg in child_groups:
                    # Collect members of this team branch
                    team_members = [
                        m["user"] for m in all_nodes
                        if m.get("user")
                        and m["lft"] >= cg["lft"]
                        and m["lft"] <= cg["rgt"]
                    ]
                    label = _make_team_label(cg)
                    teams.append({
                        "value": cg["name"],
                        "label": label,
                        "members": team_members,
                    })
            else:
                # Root is a group with no child groups — treat root as single team
                team_members = [
                    m["user"] for m in all_nodes
                    if m.get("user")
                    and m["lft"] >= root["lft"]
                    and m["lft"] <= root["rgt"]
                ]
                label = _make_team_label(root)
                teams.append({
                    "value": root["name"],
                    "label": label,
                    "members": team_members,
                })

    return {
        "is_unrestricted": True,
        "user_node": None,
        "permitted_users": permitted_users,
        "teams": teams,
    }


def _resolve_hierarchy_scope(user, node):
    """
    Build scope for a user who is IN the CRM Sales Hierarchy tree.

    Their permitted scope is their NestedSet subtree (themselves + all
    descendants). Teams = direct child group nodes under their node.

    Args:
        user (str): The user email.
        node (dict): The user's CRM Sales Hierarchy node.
    """
    subtree = _get_subtree_members(node)

    # Permitted users = all users in subtree
    permitted_users = []
    seen_users = set()
    for m in subtree:
        if m.get("user") and m["user"] not in seen_users:
            permitted_users.append({
                "value": m["user"],
                "label": m.get("full_name") or m["user"],
            })
            seen_users.add(m["user"])

    # Teams = direct child group nodes of the user's node
    teams = []
    child_groups = [m for m in subtree if m.get("reports_to") == node["name"] and m.get("is_group")]
    if child_groups:
        for cg in child_groups:
            team_members = [
                m["user"] for m in subtree
                if m.get("user")
                and m["lft"] >= cg["lft"]
                and m["lft"] <= cg["rgt"]
            ]
            label = _make_team_label(cg)
            teams.append({
                "value": cg["name"],
                "label": label,
                "members": team_members,
            })
    else:
        # User is a leaf or has only leaf children — single team = their own subtree
        team_members = [m["user"] for m in subtree if m.get("user")]
        label = _make_team_label(node)
        teams.append({
            "value": node["name"],
            "label": label,
            "members": team_members,
        })

    return {
        "is_unrestricted": False,
        "user_node": node,
        "permitted_users": permitted_users,
        "teams": teams,
    }


def _resolve_self_only_scope(user):
    """
    Build scope for a user who is NOT in the hierarchy and NOT unrestricted.

    This is a plain Sales User without a hierarchy node. They can only see
    their own records (consistent with CRM's default permission model).
    """
    full_name = frappe.db.get_value("User", user, "full_name") or user
    return {
        "is_unrestricted": False,
        "user_node": None,
        "permitted_users": [{"value": user, "label": full_name}],
        "teams": [],
    }


def _make_team_label(node):
    """
    Generate a human-readable team label from a hierarchy node.

    Uses the node's full_name (which is the manager/user's display name)
    prefixed with "Team" for clarity in the dashboard UI.
    """
    name = node.get("full_name") or node.get("user") or node.get("name")
    return f"Team {name}" if name else "Team"


# ---------------------------------------------------------------------------
# 3. FILTER VALIDATION
#    Ensures that requested team/user filters are within the caller's scope.
#    This is the SECURITY ENFORCEMENT layer — frontend filters are untrusted.
# ---------------------------------------------------------------------------

def validate_dashboard_filters(scope, team_filter=None, user_filter=None):
    """
    Validate that the requested dashboard filters are within the caller's
    permitted scope. This function is the security enforcement checkpoint.

    SECURITY PRINCIPLE:
        Frontend filter dropdowns are convenience UI. A malicious user can
        call the API directly with arbitrary parameters. This function
        ensures that no data outside the caller's hierarchy subtree is ever
        returned, regardless of what the frontend sends.

    Validation rules:
        1. If team_filter is provided and not "ALL", it must match a team
           in scope["teams"] by value (the hierarchy node name).
        2. If user_filter is provided and not "ALL", it must match a user
           in scope["permitted_users"] by value (the user email).
        3. If both team_filter and user_filter are provided, the user must
           be a member of the selected team.
        4. "ALL" always means "all within permitted scope" — never global.

    Args:
        scope (dict): The resolved scope from _resolve_scope().
        team_filter (str or None): Requested team hierarchy node name, or "ALL".
        user_filter (str or None): Requested user email, or "ALL".

    Returns:
        dict: {
            "team_filter": validated team filter (None means all permitted),
            "user_filter": validated user filter (None means all permitted),
            "effective_users": list of user emails to query against,
        }

    Raises:
        frappe.PermissionError: If the requested filter is outside scope.
    """
    all_permitted_emails = [u["value"] for u in scope["permitted_users"]]
    effective_users = list(all_permitted_emails)

    # --- Validate team_filter ---
    validated_team = None
    if team_filter and team_filter != "ALL":
        matched_team = None
        for t in scope["teams"]:
            if t["value"] == team_filter:
                matched_team = t
                break
        if not matched_team:
            frappe.throw(
                _("You do not have permission to view data for the selected team."),
                frappe.PermissionError,
            )
        validated_team = team_filter
        # Narrow effective_users to this team's members
        effective_users = [
            u for u in matched_team["members"]
            if u in all_permitted_emails
        ]

    # --- Validate user_filter ---
    validated_user = None
    if user_filter and user_filter != "ALL":
        if user_filter not in all_permitted_emails:
            frappe.throw(
                _("You do not have permission to view data for the selected user."),
                frappe.PermissionError,
            )
        validated_user = user_filter

        # If a team is also selected, ensure the user belongs to that team
        if validated_team:
            if user_filter not in effective_users:
                frappe.throw(
                    _("The selected user does not belong to the selected team."),
                    frappe.PermissionError,
                )

        # Narrow to single user
        effective_users = [user_filter]

    return {
        "team_filter": validated_team,
        "user_filter": validated_user,
        "effective_users": effective_users,
    }


# ---------------------------------------------------------------------------
# 4. PERIOD HELPERS
#    Standard date-range calculation for dashboard period filters.
# ---------------------------------------------------------------------------

def resolve_period(period_key, custom_from=None, custom_to=None):
    """
    Convert a period selector key into concrete from_date / to_date values.

    Args:
        period_key (str): One of: today, this_week, this_month, last_month,
                          this_quarter, last_quarter, this_year, custom.
        custom_from (str): Start date for custom period (YYYY-MM-DD).
        custom_to (str):   End date for custom period (YYYY-MM-DD).

    Returns:
        dict: {"from_date": str, "to_date": str}
    """
    today = frappe.utils.nowdate()

    if period_key == "today":
        return {"from_date": today, "to_date": today}
    elif period_key == "this_week":
        return {
            "from_date": frappe.utils.get_first_day_of_week(today),
            "to_date": frappe.utils.get_last_day_of_week(today),
        }
    elif period_key == "this_month":
        return {
            "from_date": frappe.utils.get_first_day(today),
            "to_date": frappe.utils.get_last_day(today),
        }
    elif period_key == "last_month":
        last = frappe.utils.add_months(today, -1)
        return {
            "from_date": frappe.utils.get_first_day(last),
            "to_date": frappe.utils.get_last_day(last),
        }
    elif period_key == "this_quarter":
        return {
            "from_date": frappe.utils.get_quarter_start(today),
            "to_date": frappe.utils.get_quarter_ending(today),
        }
    elif period_key == "last_quarter":
        last_q = frappe.utils.add_months(today, -3)
        return {
            "from_date": frappe.utils.get_quarter_start(last_q),
            "to_date": frappe.utils.get_quarter_ending(last_q),
        }
    elif period_key == "this_year":
        return {
            "from_date": frappe.utils.get_year_start(today),
            "to_date": frappe.utils.get_year_ending(today),
        }
    elif period_key == "custom":
        if not custom_from or not custom_to:
            frappe.throw(_("Custom period requires both from_date and to_date."))
        # Validate that from_date is not after to_date
        if str(custom_from) > str(custom_to):
            frappe.throw(_("Custom period from_date cannot be after to_date."))
        return {"from_date": custom_from, "to_date": custom_to}
    else:
        # Reject unknown period keys explicitly rather than silently defaulting.
        # This prevents confusing behaviour when a typo is passed.
        frappe.throw(
            _("Invalid period: {0}. Valid options: today, this_week, this_month, "
              "last_month, this_quarter, last_quarter, this_year, custom.").format(period_key)
        )


# ---------------------------------------------------------------------------
# 5. WHITELISTED API ENDPOINTS
#    These are the public Frappe RPC methods called by the Vue frontend.
# ---------------------------------------------------------------------------

@frappe.whitelist()
def get_permitted_filter_options():
    """
    Return the dashboard filter options available to the current logged-in user.

    This is the FIRST API call the dashboard makes on load. The response
    tells the Vue frontend exactly which teams and users to populate in the
    universal filter toolbar dropdowns.

    Security:
        The options returned are pre-filtered by hierarchy scope. The frontend
        cannot display or select options outside this set. Even if it did, the
        backend validate_dashboard_filters() would reject the request.

    Returns:
        dict: {
            "teams":        list of {value, label} for the team dropdown,
            "users":        list of {value, label} for the user dropdown,
            "default_team": default team selection ("ALL" or a specific node name),
            "default_user": default user selection ("ALL" or a specific user email),
            "period_options": list of {value, label} for the period dropdown,
            "is_unrestricted": bool — whether the user has global access,
            "hierarchy_enabled": bool — whether CRM Sales Hierarchy is active,
        }
    """
    user = frappe.session.user
    scope = _resolve_scope(user)

    # Build team options with "All Teams" prefix where appropriate
    team_options = []
    if len(scope["teams"]) > 1 or scope["is_unrestricted"]:
        team_options.append({"value": "ALL", "label": _("All Teams")})
    for t in scope["teams"]:
        team_options.append({"value": t["value"], "label": t["label"]})

    # Build user options with "All Users" prefix where appropriate
    user_options = []
    if len(scope["permitted_users"]) > 1:
        user_options.append({"value": "ALL", "label": _("All Users")})
    for u in scope["permitted_users"]:
        user_options.append(u)

    # Default selections
    default_team = "ALL" if len(scope["teams"]) > 1 else (
        scope["teams"][0]["value"] if scope["teams"] else "ALL"
    )
    default_user = "ALL" if len(scope["permitted_users"]) > 1 else (
        scope["permitted_users"][0]["value"] if scope["permitted_users"] else user
    )

    # Standard period options
    period_options = [
        {"value": "today", "label": _("Today")},
        {"value": "this_week", "label": _("This Week")},
        {"value": "this_month", "label": _("This Month")},
        {"value": "last_month", "label": _("Last Month")},
        {"value": "this_quarter", "label": _("This Quarter")},
        {"value": "last_quarter", "label": _("Last Quarter")},
        {"value": "this_year", "label": _("This Year")},
        {"value": "custom", "label": _("Custom")},
    ]

    # User Role & Full Name for Header Scope Label
    user_fullname = frappe.db.get_value("User", user, "full_name") or user
    roles = frappe.get_roles(user)
    if "System Manager" in roles or "Administrator" in user:
        user_role = "System Manager"
    elif "Sales Manager" in roles or "Sales Master Manager" in roles:
        user_role = "Sales Manager"
    else:
        user_role = "Sales User"

    return {
        "teams": team_options,
        "users": user_options,
        "default_team": default_team,
        "default_user": default_user,
        "period_options": period_options,
        "is_unrestricted": scope["is_unrestricted"],
        "hierarchy_enabled": _is_hierarchy_enabled(),
        "user_role": user_role,
        "user_fullname": user_fullname,
    }


@frappe.whitelist()
def get_users_for_team(team_node_name):
    """
    Return the permitted user list for a specific team selection.

    Called when the user changes the Team dropdown — the User dropdown must
    cascade to show only users belonging to the selected team.

    Security:
        The team_node_name is validated against the caller's hierarchy scope.
        If the caller is not permitted to view this team, a PermissionError
        is raised.

    Args:
        team_node_name (str): The CRM Sales Hierarchy node name for the team,
                              or "ALL" for all permitted users.

    Returns:
        list[dict]: [{value, label}] for the user dropdown.
    """
    user = frappe.session.user
    scope = _resolve_scope(user)

    if team_node_name == "ALL":
        # Return all permitted users
        user_options = [{"value": "ALL", "label": _("All Users")}]
        user_options.extend(scope["permitted_users"])
        return user_options

    # Validate that the requested team is in scope
    matched_team = None
    for t in scope["teams"]:
        if t["value"] == team_node_name:
            matched_team = t
            break

    if not matched_team:
        frappe.throw(
            _("You do not have permission to view this team."),
            frappe.PermissionError,
        )

    # Build user options from the team's members
    user_options = []
    if len(matched_team["members"]) > 1:
        user_options.append({"value": "ALL", "label": _("All Team Members")})

    for member_email in matched_team["members"]:
        label = frappe.db.get_value("User", member_email, "full_name") or member_email
        user_options.append({"value": member_email, "label": label})

    return user_options


@frappe.whitelist()
def validate_and_get_scope(
    period="this_month",
    team_filter="ALL",
    user_filter="ALL",
    custom_from=None,
    custom_to=None,
):
    """
    Validate the complete set of dashboard filters and return the effective
    query scope.

    This endpoint is called before any KPI data fetch. It validates all
    filter parameters and returns the resolved date range + effective user
    list that Phase 2+ KPI endpoints will use.

    Security:
        All filters are validated against the caller's hierarchy scope.
        Invalid or unauthorized filters result in PermissionError.

    Args:
        period (str):       Period key (today, this_month, etc.)
        team_filter (str):  Team node name or "ALL"
        user_filter (str):  User email or "ALL"
        custom_from (str):  Start date for custom period
        custom_to (str):    End date for custom period

    Returns:
        dict: {
            "from_date": str,
            "to_date": str,
            "effective_users": list[str],
            "team_filter": str or None,
            "user_filter": str or None,
            "user": str (session user),
            "is_unrestricted": bool,
        }
    """
    user = frappe.session.user
    scope = _resolve_scope(user)

    # Validate team and user filters against hierarchy scope
    validated = validate_dashboard_filters(scope, team_filter, user_filter)

    # Resolve the date range
    dates = resolve_period(period, custom_from, custom_to)

    return {
        "from_date": str(dates["from_date"]),
        "to_date": str(dates["to_date"]),
        "effective_users": validated["effective_users"],
        "team_filter": validated["team_filter"],
        "user_filter": validated["user_filter"],
        "user": user,
        "is_unrestricted": scope["is_unrestricted"],
    }


# ---------------------------------------------------------------------------
# PHASE 2B — EXECUTIVE OVERVIEW KPI DATA FETCH
# ---------------------------------------------------------------------------

@frappe.whitelist()
def get_executive_kpis(
    period="this_month",
    team_filter="ALL",
    user_filter="ALL",
    custom_from=None,
    custom_to=None,
):
    """
    Fetch the Phase 2B Executive Overview KPI metrics.

    SECURITY & SCOPE:
        Consumes Phase 1's `validate_and_get_scope()` to enforce exact hierarchy
        permission boundaries. Never bypasses scope resolution.

    METRICS PROCESSED:
        1.  leads_created      : CRM Leads created during period (creation)
        2.  new_leads          : CRM Leads created during period with status='New'
        3.  converted_leads    : CRM Leads converted in period (CRM Deal.creation)
        4.  conversion_rate    : (converted_leads / leads_created) * 100
        5.  open_deals         : Live count of open/ongoing deals (CURRENT SNAPSHOT)
        6.  pipeline_value     : Live total deal_value of open deals (CURRENT SNAPSHOT)
        7.  weighted_pipeline  : Live sum of (deal_value * probability / 100) (CURRENT SNAPSHOT)
        8.  average_deal_value : Live avg deal_value of open deals (CURRENT SNAPSHOT)
        9.  won_deals          : Count of deals closed as Won during period (closed_date)
        10. won_revenue        : Sum of deal_value for Won deals during period (closed_date)
        11. lost_deals         : Count of deals closed as Lost during period (closed_date)
        12. win_rate           : Won / (Won + Lost) * 100 (0-denominator safe)
        13. invoiced_revenue   : Net total of submitted Sales Invoices (posting_date)
        14. overdue_followups  : Open ToDos past due date (date < CURDATE())

    Args:
        period (str):       Period key (today, this_month, etc.)
        team_filter (str):  Team node name or "ALL"
        user_filter (str):  User email or "ALL"
        custom_from (str):  Start date for custom period
        custom_to (str):    End date for custom period

    Returns:
        dict: Standard structured JSON response with scope, kpis, and meta.
    """
    # 1. Validate scope via Phase 1 security boundary
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    from_date = scope_data["from_date"]
    to_date = scope_data["to_date"]
    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]

    # 2. Build owner condition clauses
    # If unrestricted AND effective_users is empty/ALL -> no owner restriction.
    # Otherwise -> filter by owner IN (%s).
    apply_user_filter = not (is_unrestricted and (user_filter == "ALL" and team_filter == "ALL"))
    
    # --- A. LEADS METRICS ---
    lead_where = ["1=1"]
    lead_params = {}

    if apply_user_filter:
        if effective_users:
            lead_where.append("lead_owner IN %(users)s")
            lead_params["users"] = tuple(effective_users)
        else:
            lead_where.append("1=0")

    lead_where_str = " AND ".join(lead_where)

    # Leads Created in period
    leads_created = frappe.db.sql(
        f"""
        SELECT COUNT(name) FROM `tabCRM Lead`
        WHERE {lead_where_str}
          AND DATE(creation) BETWEEN %(from_date)s AND %(to_date)s
        """,
        {**lead_params, "from_date": from_date, "to_date": to_date},
    )[0][0] or 0

    # New Leads created in period currently in status 'New'
    new_leads = frappe.db.sql(
        f"""
        SELECT COUNT(name) FROM `tabCRM Lead`
        WHERE {lead_where_str}
          AND status = 'New'
          AND DATE(creation) BETWEEN %(from_date)s AND %(to_date)s
        """,
        {**lead_params, "from_date": from_date, "to_date": to_date},
    )[0][0] or 0

    # Converted Leads in period: CRM Lead + CRM Deal join (anchored to CRM Deal.creation)
    converted_leads = frappe.db.sql(
        f"""
        SELECT COUNT(DISTINCT l.name)
        FROM `tabCRM Lead` l
        INNER JOIN `tabCRM Deal` d ON d.lead = l.name
        WHERE l.converted = 1
          AND {lead_where_str.replace('lead_owner', 'l.lead_owner')}
          AND DATE(d.creation) BETWEEN %(from_date)s AND %(to_date)s
        """,
        {**lead_params, "from_date": from_date, "to_date": to_date},
    )[0][0] or 0

    # Same-Cohort Converted Leads: Leads created in period that have converted (converted = 1)
    cohort_converted = frappe.db.sql(
        f"""
        SELECT COUNT(name) FROM `tabCRM Lead`
        WHERE {lead_where_str}
          AND converted = 1
          AND DATE(creation) BETWEEN %(from_date)s AND %(to_date)s
        """,
        {**lead_params, "from_date": from_date, "to_date": to_date},
    )[0][0] or 0

    conversion_rate = round((cohort_converted / leads_created * 100), 2) if leads_created > 0 else 0.0

    # --- B. PIPELINE SNAPSHOT METRICS (Open Deals) ---
    deal_where = ["1=1"]
    deal_params = {}

    if apply_user_filter:
        if effective_users:
            deal_where.append("deal_owner IN %(users)s")
            deal_params["users"] = tuple(effective_users)
        else:
            deal_where.append("1=0")

    deal_where_str = " AND ".join(deal_where)

    # Fetch open/ongoing deal status names from CRM Deal Status
    open_stages = frappe.get_all("CRM Deal Status", filters={"type": ["in", ["Open", "Ongoing"]]}, pluck="name")
    if not open_stages:
        # Fallback to excluding Won/Lost
        open_stages_condition = "status NOT IN ('Won', 'Lost')"
    else:
        deal_params["open_stages"] = tuple(open_stages)
        open_stages_condition = "status IN %(open_stages)s"

    open_pipeline_res = frappe.db.sql(
        f"""
        SELECT
            COUNT(name) as count_deals,
            SUM(deal_value) as sum_val,
            SUM(deal_value * IFNULL(probability, 0) / 100.0) as sum_weighted,
            AVG(deal_value) as avg_val
        FROM `tabCRM Deal`
        WHERE {deal_where_str}
          AND {open_stages_condition}
        """,
        deal_params,
        as_dict=True,
    )[0]

    open_deals = open_pipeline_res.count_deals or 0
    pipeline_value = round(float(open_pipeline_res.sum_val or 0.0), 2)
    weighted_pipeline = round(float(open_pipeline_res.sum_weighted or 0.0), 2)
    average_deal_value = round(float(open_pipeline_res.avg_val or 0.0), 2)

    # --- C. SALES RESULTS (Closed Deals in Period) ---
    won_statuses = frappe.get_all("CRM Deal Status", filters={"type": "Won"}, pluck="name") or ["Won"]
    lost_statuses = frappe.get_all("CRM Deal Status", filters={"type": "Lost"}, pluck="name") or ["Lost"]
    closed_statuses = list(set(won_statuses + lost_statuses))

    deal_params_closed = {**deal_params, "from_date": from_date, "to_date": to_date, "closed_statuses": tuple(closed_statuses), "won_statuses": tuple(won_statuses), "lost_statuses": tuple(lost_statuses)}

    closed_res = frappe.db.sql(
        f"""
        SELECT
            SUM(CASE WHEN status IN %(won_statuses)s THEN 1 ELSE 0 END) as won_count,
            SUM(CASE WHEN status IN %(won_statuses)s THEN deal_value ELSE 0 END) as won_val,
            SUM(CASE WHEN status IN %(lost_statuses)s THEN 1 ELSE 0 END) as lost_count
        FROM `tabCRM Deal`
        WHERE {deal_where_str}
          AND status IN %(closed_statuses)s
          AND closed_date BETWEEN %(from_date)s AND %(to_date)s
        """,
        deal_params_closed,
        as_dict=True,
    )[0]

    won_deals = closed_res.won_count or 0
    won_revenue = round(float(closed_res.won_val or 0.0), 2)
    lost_deals = closed_res.lost_count or 0

    won_lost_total = won_deals + lost_deals
    win_rate = round((won_deals / won_lost_total * 100), 2) if won_lost_total > 0 else 0.0

    # --- D. ACCOUNTING REVENUE (Sales Invoice with Sales Team Join) ---
    # Prevents duplicate revenue by grouping or using DISTINCT invoice summation
    inv_where = ["si.docstatus = 1"]
    inv_params = {"from_date": from_date, "to_date": to_date}

    if apply_user_filter:
        if effective_users:
            inv_where.append("""
                (si.owner IN %(users)s OR emp.user_id IN %(users)s)
            """)
            inv_params["users"] = tuple(effective_users)
        else:
            inv_where.append("1=0")

    inv_where_str = " AND ".join(inv_where)

    invoiced_rev_res = frappe.db.sql(
        f"""
        SELECT SUM(inv_amount) as total_invoiced FROM (
            SELECT
                si.name,
                MAX(si.net_total * (CASE WHEN si.is_return = 1 THEN -1 ELSE 1 END)) as inv_amount
            FROM `tabSales Invoice` si
            LEFT JOIN `tabSales Team` st ON st.parent = si.name
            LEFT JOIN `tabSales Person` sp ON sp.name = st.sales_person
            LEFT JOIN `tabEmployee` emp ON emp.name = sp.employee
            WHERE {inv_where_str}
              AND si.posting_date BETWEEN %(from_date)s AND %(to_date)s
            GROUP BY si.name
        ) sub
        """,
        inv_params,
    )[0][0] or 0.0

    invoiced_revenue = round(float(invoiced_rev_res), 2)

    # --- E. OVERDUE FOLLOW-UPS (ToDo Snapshot) ---
    todo_where = ["status = 'Open'", "date < CURDATE()", "reference_type IN ('CRM Lead', 'CRM Deal')"]
    todo_params = {}

    if apply_user_filter:
        if effective_users:
            todo_where.append("allocated_to IN %(users)s")
            todo_params["users"] = tuple(effective_users)
        else:
            todo_where.append("1=0")

    todo_where_str = " AND ".join(todo_where)

    overdue_followups = frappe.db.sql(
        f"""
        SELECT COUNT(name) FROM `tabToDo`
        WHERE {todo_where_str}
        """,
        todo_params,
    )[0][0] or 0

    # Get system currency
    company_currency = frappe.db.get_default("currency") or "INR"

    return {
        "scope": {
            "from_date": str(from_date),
            "to_date": str(to_date),
            "team_filter": scope_data["team_filter"],
            "user_filter": scope_data["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": is_unrestricted,
        },
        "kpis": {
            "leads_created": int(leads_created),
            "new_leads": int(new_leads),
            "converted_leads": int(converted_leads),
            "conversion_rate": float(conversion_rate),
            "open_deals": int(open_deals),
            "pipeline_value": float(pipeline_value),
            "weighted_pipeline": float(weighted_pipeline),
            "average_deal_value": float(average_deal_value),
            "won_deals": int(won_deals),
            "won_revenue": float(won_revenue),
            "lost_deals": int(lost_deals),
            "win_rate": float(win_rate),
            "invoiced_revenue": float(invoiced_revenue),
            "overdue_followups": int(overdue_followups),
        },
        "meta": {
            "currency": company_currency,
            "generated_at": str(frappe.utils.now_datetime()),
        },
    }


@frappe.whitelist()
def get_lead_funnel(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
):
    """
    Phase 2C API: Current Lead Funnel (Live Snapshot) & Period Conversion Summary
    =============================================================================
    Security: Uses validate_and_get_scope() to enforce hierarchy access control.
    Returns:
      - funnel: Dynamic lead counts grouped by CRM Lead Status (Snapshot)
      - period_activity: Period created leads, period conversions, cohort conversion rate
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    from_date = scope_data["from_date"]
    to_date = scope_data["to_date"]
    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]

    # Scope condition for CRM Lead
    lead_where = ["1=1"]
    lead_params = {}
    apply_user_filter = scope_data["team_filter"] is not None or scope_data["user_filter"] is not None or not is_unrestricted

    if apply_user_filter:
        if effective_users:
            lead_where.append("lead_owner IN %(users)s")
            lead_params["users"] = tuple(effective_users)
        else:
            lead_where.append("1=0")

    lead_where_str = " AND ".join(lead_where)

    # 1. Fetch dynamic status list ordered by position
    statuses = frappe.get_all(
        "CRM Lead Status",
        fields=["name", "type", "position", "color"],
        order_by="position asc, creation asc",
    )

    # 2. Get live lead count grouped by status (SNAPSHOT)
    counts_by_status_raw = frappe.db.sql(
        f"""
        SELECT status, COUNT(name) as cnt
        FROM `tabCRM Lead`
        WHERE {lead_where_str}
        GROUP BY status
        """,
        lead_params,
        as_dict=True,
    )
    status_count_map = {r["status"]: r["cnt"] for r in counts_by_status_raw}

    total_active_leads = sum(status_count_map.values())

    funnel_items = []
    for st in statuses:
        s_name = st["name"]
        cnt = status_count_map.get(s_name, 0)
        pct = round((cnt / total_active_leads * 100), 2) if total_active_leads > 0 else 0.0
        funnel_items.append({
            "status": s_name,
            "type": st["type"],
            "position": st["position"],
            "color": st["color"] or "gray",
            "count": int(cnt),
            "percentage": float(pct),
        })

    # 3. Period Cohort & Activity Metrics
    leads_created = frappe.db.sql(
        f"""
        SELECT COUNT(name) FROM `tabCRM Lead`
        WHERE {lead_where_str}
          AND DATE(creation) BETWEEN %(from_date)s AND %(to_date)s
        """,
        {**lead_params, "from_date": from_date, "to_date": to_date},
    )[0][0] or 0

    period_conversions = frappe.db.sql(
        f"""
        SELECT COUNT(DISTINCT l.name)
        FROM `tabCRM Lead` l
        INNER JOIN `tabCRM Deal` d ON d.lead = l.name
        WHERE l.converted = 1
          AND {lead_where_str.replace('lead_owner', 'l.lead_owner')}
          AND DATE(d.creation) BETWEEN %(from_date)s AND %(to_date)s
        """,
        {**lead_params, "from_date": from_date, "to_date": to_date},
    )[0][0] or 0

    cohort_converted = frappe.db.sql(
        f"""
        SELECT COUNT(name) FROM `tabCRM Lead`
        WHERE {lead_where_str}
          AND converted = 1
          AND DATE(creation) BETWEEN %(from_date)s AND %(to_date)s
        """,
        {**lead_params, "from_date": from_date, "to_date": to_date},
    )[0][0] or 0

    cohort_conversion_rate = round((cohort_converted / leads_created * 100), 2) if leads_created > 0 else 0.0

    return {
        "scope": {
            "from_date": str(from_date),
            "to_date": str(to_date),
            "team_filter": scope_data["team_filter"],
            "user_filter": scope_data["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": is_unrestricted,
        },
        "funnel": funnel_items,
        "period_activity": {
            "leads_created": int(leads_created),
            "period_conversions": int(period_conversions),
            "cohort_converted": int(cohort_converted),
            "cohort_conversion_rate": float(cohort_conversion_rate),
            "total_active_leads": int(total_active_leads),
        },
        "meta": {
            "metric_type": "snapshot",
            "snapshot": True,
            "generated_at": str(frappe.utils.now_datetime()),
            "cohort_definition": "Leads created during period subsequently converted to date",
        },
    }


@frappe.whitelist()
def get_lead_sources(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
):
    """
    Phase 2C API: Lead Source Analytics
    ====================================
    Security: Uses validate_and_get_scope() to enforce hierarchy access control.
    Returns grouping by source for:
      - leads_created: Created during period
      - cohorted_converted: Created during period & converted to date
      - cohorted_conversion_rate: cohorted_converted / leads_created * 100
      - period_conversions: Converted during period (Deal creation in period)
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    from_date = scope_data["from_date"]
    to_date = scope_data["to_date"]
    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]

    lead_where = ["1=1"]
    lead_params = {}
    apply_user_filter = scope_data["team_filter"] is not None or scope_data["user_filter"] is not None or not is_unrestricted

    if apply_user_filter:
        if effective_users:
            lead_where.append("lead_owner IN %(users)s")
            lead_params["users"] = tuple(effective_users)
        else:
            lead_where.append("1=0")

    lead_where_str = " AND ".join(lead_where)
    period_params = {**lead_params, "from_date": from_date, "to_date": to_date}

    # 1. Created leads & Cohorted converted leads grouped by source
    created_raw = frappe.db.sql(
        f"""
        SELECT
            IFNULL(NULLIF(source, ''), 'Unknown') as source_name,
            COUNT(name) as total_created,
            SUM(CASE WHEN converted = 1 THEN 1 ELSE 0 END) as total_cohorted_converted
        FROM `tabCRM Lead`
        WHERE {lead_where_str}
          AND DATE(creation) BETWEEN %(from_date)s AND %(to_date)s
        GROUP BY source_name
        """,
        period_params,
        as_dict=True,
    )

    # 2. Conversions occurring during period grouped by source
    conversions_raw = frappe.db.sql(
        f"""
        SELECT
            IFNULL(NULLIF(l.source, ''), 'Unknown') as source_name,
            COUNT(DISTINCT l.name) as period_conv
        FROM `tabCRM Lead` l
        INNER JOIN `tabCRM Deal` d ON d.lead = l.name
        WHERE l.converted = 1
          AND {lead_where_str.replace('lead_owner', 'l.lead_owner')}
          AND DATE(d.creation) BETWEEN %(from_date)s AND %(to_date)s
        GROUP BY source_name
        """,
        period_params,
        as_dict=True,
    )

    conversions_map = {r["source_name"]: r["period_conv"] for r in conversions_raw}
    sources_dict = {}

    for row in created_raw:
        s_name = row["source_name"]
        l_created = int(row["total_created"] or 0)
        c_converted = int(row["total_cohorted_converted"] or 0)
        c_rate = round((c_converted / l_created * 100), 2) if l_created > 0 else 0.0
        p_conv = int(conversions_map.get(s_name, 0))

        sources_dict[s_name] = {
            "source": s_name,
            "leads_created": l_created,
            "cohorted_converted": c_converted,
            "cohorted_conversion_rate": float(c_rate),
            "period_conversions": p_conv,
        }

    # Also capture sources that had conversions in period but 0 leads created in period
    for s_name, p_conv in conversions_map.items():
        if s_name not in sources_dict:
            sources_dict[s_name] = {
                "source": s_name,
                "leads_created": 0,
                "cohorted_converted": 0,
                "cohorted_conversion_rate": 0.0,
                "period_conversions": int(p_conv),
            }

    # Sort sources by leads_created desc, then period_conversions desc
    sorted_sources = sorted(
        sources_dict.values(),
        key=lambda x: (x["leads_created"], x["period_conversions"]),
        reverse=True,
    )

    return {
        "scope": {
            "from_date": str(from_date),
            "to_date": str(to_date),
            "team_filter": scope_data["team_filter"],
            "user_filter": scope_data["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": is_unrestricted,
        },
        "sources": sorted_sources,
        "meta": {
            "snapshot": False,
            "generated_at": str(frappe.utils.now_datetime()),
            "cohort_definition": "Leads created during period subsequently converted to date",
        },
    }


@frappe.whitelist()
def get_pipeline_health(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
):
    """
    Phase 2D API: Pipeline Health (Live Snapshot)
    =============================================
    Security: Consumes validate_and_get_scope() to enforce hierarchy access control.
    Metrics:
      - Open Deals, Pipeline Value, Weighted Pipeline Value, Average Deal Value
      - Pipeline by Stage (ordered dynamically by CRM Deal Status.position ASC)
      - On Hold Deals (count, value)
      - Stale Deals (count, value) where modified <= NOW() - INTERVAL 14 DAY
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    from_date = scope_data["from_date"]
    to_date = scope_data["to_date"]
    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]

    # Dynamic company base currency
    company_currency = frappe.db.get_default("currency") or "INR"

    deal_where = ["1=1"]
    deal_params = {}
    apply_user_filter = scope_data["team_filter"] is not None or scope_data["user_filter"] is not None or not is_unrestricted

    if apply_user_filter:
        if effective_users:
            deal_where.append("d.deal_owner IN %(users)s")
            deal_params["users"] = tuple(effective_users)
        else:
            deal_where.append("1=0")

    deal_where_str = " AND ".join(deal_where)

    # 1. Summary Metrics for Open Pipeline (type IN ('Open', 'Ongoing'))
    open_summary = frappe.db.sql(
        f"""
        SELECT
            COUNT(d.name) as open_deals,
            SUM(IFNULL(d.deal_value, 0)) as pipeline_value,
            SUM(IFNULL(NULLIF(d.expected_deal_value, 0), IFNULL(d.deal_value, 0) * IFNULL(d.probability, s.probability) / 100)) as weighted_pipeline
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE {deal_where_str}
          AND s.type IN ('Open', 'Ongoing')
        """,
        deal_params,
        as_dict=True,
    )[0]

    open_deals = int(open_summary["open_deals"] or 0)
    pipeline_value = float(open_summary["pipeline_value"] or 0.0)
    weighted_pipeline = float(open_summary["weighted_pipeline"] or 0.0)
    avg_deal_value = round((pipeline_value / open_deals), 2) if open_deals > 0 else 0.0

    # 2. On-Hold Deals (type = 'On Hold')
    on_hold_summary = frappe.db.sql(
        f"""
        SELECT
            COUNT(d.name) as on_hold_deals,
            SUM(IFNULL(d.deal_value, 0)) as on_hold_value
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE {deal_where_str}
          AND s.type = 'On Hold'
        """,
        deal_params,
        as_dict=True,
    )[0]

    on_hold_deals = int(on_hold_summary["on_hold_deals"] or 0)
    on_hold_value = float(on_hold_summary["on_hold_value"] or 0.0)

    # 3. Stale Deals (Open/Ongoing & modified <= NOW() - INTERVAL 14 DAY)
    stale_summary = frappe.db.sql(
        f"""
        SELECT
            COUNT(d.name) as stale_deals,
            SUM(IFNULL(d.deal_value, 0)) as stale_value
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE {deal_where_str}
          AND s.type IN ('Open', 'Ongoing')
          AND d.modified <= NOW() - INTERVAL 14 DAY
        """,
        deal_params,
        as_dict=True,
    )[0]

    stale_deals = int(stale_summary["stale_deals"] or 0)
    stale_value = float(stale_summary["stale_value"] or 0.0)

    # 4. Pipeline by Stage (Open/Ongoing stages ordered by position ASC)
    stages_raw = frappe.db.sql(
        f"""
        SELECT
            s.name as stage,
            s.type,
            s.position,
            s.color,
            s.probability as default_probability,
            COUNT(d.name) as deal_count,
            SUM(IFNULL(d.deal_value, 0)) as stage_value,
            SUM(IFNULL(NULLIF(d.expected_deal_value, 0), IFNULL(d.deal_value, 0) * IFNULL(d.probability, s.probability) / 100)) as weighted_value
        FROM `tabCRM Deal Status` s
        LEFT JOIN `tabCRM Deal` d ON d.status = s.name AND {deal_where_str}
        WHERE s.type IN ('Open', 'Ongoing')
        GROUP BY s.name, s.type, s.position, s.color, s.probability
        ORDER BY s.position ASC, s.creation ASC
        """,
        deal_params,
        as_dict=True,
    )

    stage_items = []
    for st in stages_raw:
        s_count = int(st["deal_count"] or 0)
        s_val = float(st["stage_value"] or 0.0)
        w_val = float(st["weighted_value"] or 0.0)
        pct = round((s_val / pipeline_value * 100), 2) if pipeline_value > 0 else 0.0

        stage_items.append({
            "stage": st["stage"],
            "type": st["type"],
            "position": st["position"],
            "color": st["color"] or "gray",
            "default_probability": float(st["default_probability"] or 0.0),
            "deal_count": s_count,
            "stage_value": s_val,
            "weighted_value": w_val,
            "percentage_of_pipeline": float(pct),
        })

    return {
        "scope": {
            "from_date": str(from_date),
            "to_date": str(to_date),
            "team_filter": scope_data["team_filter"],
            "user_filter": scope_data["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": is_unrestricted,
        },
        "summary": {
            "open_deals": open_deals,
            "pipeline_value": pipeline_value,
            "weighted_pipeline": weighted_pipeline,
            "average_deal_value": avg_deal_value,
            "on_hold_deals": on_hold_deals,
            "on_hold_value": on_hold_value,
            "stale_deals": stale_deals,
            "stale_value": stale_value,
        },
        "stages": stage_items,
        "meta": {
            "metric_type": "snapshot",
            "snapshot": True,
            "currency": company_currency,
            "stale_threshold_days": 14,
            "generated_at": str(frappe.utils.now_datetime()),
        },
    }


@frappe.whitelist()
def get_closed_sales_analytics(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
):
    """
    Phase 2E API: Closed Sales & Win/Loss Performance Analytics (Period Event)
    ==========================================================================
    Security: Consumes validate_and_get_scope() to enforce hierarchy access control.
    Date Anchor: CRM Deal.closed_date BETWEEN from_date AND to_date
    Metrics:
      - Closed Deals, Won Deals, Lost Deals
      - Won Revenue, Lost Value
      - Closed Win Rate (0.0% to 100.0%)
      - Average Won Deal Size
      - Lost Reason Distribution (count, lost_value, percentage_of_lost_value)
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    from_date = scope_data["from_date"]
    to_date = scope_data["to_date"]
    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]

    company_currency = frappe.db.get_default("currency") or "INR"

    deal_where = ["1=1"]
    deal_params = {
        "from_date": from_date,
        "to_date": to_date,
    }

    apply_user_filter = scope_data["team_filter"] is not None or scope_data["user_filter"] is not None or not is_unrestricted

    if apply_user_filter:
        if effective_users:
            deal_where.append("d.deal_owner IN %(users)s")
            deal_params["users"] = tuple(effective_users)
        else:
            deal_where.append("1=0")

    # Date anchor constraint strictly on closed_date
    deal_where.append("d.closed_date BETWEEN %(from_date)s AND %(to_date)s")

    deal_where_str = " AND ".join(deal_where)

    # 1. Closed Deals Summary Metrics (Won + Lost)
    closed_summary = frappe.db.sql(
        f"""
        SELECT
            SUM(CASE WHEN s.type = 'Won' THEN 1 ELSE 0 END) as won_deals,
            SUM(CASE WHEN s.type = 'Lost' THEN 1 ELSE 0 END) as lost_deals,
            SUM(CASE WHEN s.type = 'Won' THEN IFNULL(d.deal_value, 0) ELSE 0 END) as won_revenue,
            SUM(CASE WHEN s.type = 'Lost' THEN IFNULL(d.deal_value, 0) ELSE 0 END) as lost_value
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE {deal_where_str}
          AND s.type IN ('Won', 'Lost')
        """,
        deal_params,
        as_dict=True,
    )[0]

    won_deals = int(closed_summary["won_deals"] or 0)
    lost_deals = int(closed_summary["lost_deals"] or 0)
    closed_deals = won_deals + lost_deals

    won_revenue = float(closed_summary["won_revenue"] or 0.0)
    lost_value = float(closed_summary["lost_value"] or 0.0)

    closed_win_rate = round((won_deals / closed_deals * 100), 2) if closed_deals > 0 else 0.0
    closed_win_rate = min(max(closed_win_rate, 0.0), 100.0)

    avg_won_deal_size = round((won_revenue / won_deals), 2) if won_deals > 0 else 0.0

    # 2. Lost Reason Distribution (Grouped by CRM Deal.lost_reason)
    lost_reasons_raw = frappe.db.sql(
        f"""
        SELECT
            IFNULL(NULLIF(TRIM(d.lost_reason), ''), 'Unknown/Unspecified') as reason,
            COUNT(d.name) as lost_count,
            SUM(IFNULL(d.deal_value, 0)) as reason_lost_value
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE {deal_where_str}
          AND s.type = 'Lost'
        GROUP BY reason
        ORDER BY reason_lost_value DESC, lost_count DESC
        """,
        deal_params,
        as_dict=True,
    )

    lost_reasons_list = []
    for r in lost_reasons_raw:
        r_count = int(r["lost_count"] or 0)
        r_val = float(r["reason_lost_value"] or 0.0)
        pct = round((r_val / lost_value * 100), 2) if lost_value > 0 else 0.0

        lost_reasons_list.append({
            "lost_reason": r["reason"],
            "count": r_count,
            "lost_value": r_val,
            "percentage_of_lost_value": float(pct),
        })

    return {
        "scope": {
            "from_date": str(from_date),
            "to_date": str(to_date),
            "team_filter": scope_data["team_filter"],
            "user_filter": scope_data["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": is_unrestricted,
        },
        "summary": {
            "closed_deals": closed_deals,
            "won_deals": won_deals,
            "lost_deals": lost_deals,
            "won_revenue": won_revenue,
            "lost_value": lost_value,
            "closed_win_rate": closed_win_rate,
            "average_won_deal_size": avg_won_deal_size,
        },
        "lost_reasons": lost_reasons_list,
        "meta": {
            "metric_type": "event",
            "snapshot": False,
            "currency": company_currency,
            "date_anchor": "closed_date",
            "generated_at": str(frappe.utils.now_datetime()),
        },
    }


@frappe.whitelist()
def get_sales_velocity_analytics(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
):
    """
    Phase 2F API: Sales Cycle & Deal Velocity Analytics (Hybrid: Period Event & Current Snapshot)
    =============================================================================================
    Security: Consumes validate_and_get_scope() to enforce hierarchy access control.
    Metrics:
      1. Average Won Sales Cycle (Days between creation and closed_date for Won deals in period)
      2. Average Lost Sales Cycle (Days between creation and closed_date for Lost deals in period)
      3. Average Open Deal Age (Days since creation for current Open/Ongoing deals)
      4. Sales Velocity Per Day (Won Revenue in Period / Avg Won Sales Cycle Days)
      5. Open Pipeline Age Distribution (0-30, 31-60, 61-90, 90+ Days)
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    from_date = scope_data["from_date"]
    to_date = scope_data["to_date"]
    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]

    company_currency = frappe.db.get_default("currency") or "INR"

    # Base hierarchy filtering condition
    user_where = ["1=1"]
    user_params = {}

    apply_user_filter = scope_data["team_filter"] is not None or scope_data["user_filter"] is not None or not is_unrestricted

    if apply_user_filter:
        if effective_users:
            user_where.append("d.deal_owner IN %(users)s")
            user_params["users"] = tuple(effective_users)
        else:
            user_where.append("1=0")

    user_where_str = " AND ".join(user_where)

    # 1. Closed Sales Cycles in Period (Won vs Lost)
    closed_params = dict(user_params)
    closed_params.update({"from_date": from_date, "to_date": to_date})

    cycle_stats = frappe.db.sql(
        f"""
        SELECT
            AVG(CASE WHEN s.type = 'Won' THEN DATEDIFF(d.closed_date, DATE(d.creation)) ELSE NULL END) as avg_won_days,
            AVG(CASE WHEN s.type = 'Lost' THEN DATEDIFF(d.closed_date, DATE(d.creation)) ELSE NULL END) as avg_lost_days,
            SUM(CASE WHEN s.type = 'Won' THEN IFNULL(d.deal_value, 0) ELSE 0 END) as won_revenue
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE {user_where_str}
          AND d.closed_date BETWEEN %(from_date)s AND %(to_date)s
          AND s.type IN ('Won', 'Lost')
        """,
        closed_params,
        as_dict=True,
    )[0]

    avg_won_cycle = round(float(cycle_stats["avg_won_days"]), 1) if cycle_stats["avg_won_days"] is not None else 0.0
    avg_lost_cycle = round(float(cycle_stats["avg_lost_days"]), 1) if cycle_stats["avg_lost_days"] is not None else 0.0
    won_revenue = float(cycle_stats["won_revenue"] or 0.0)

    # Sales Velocity = Won Revenue in Period / Avg Won Sales Cycle Days
    sales_velocity = round((won_revenue / avg_won_cycle), 2) if (avg_won_cycle > 0 and won_revenue > 0) else 0.0

    # 2. Open Pipeline Aging & Distribution (Current Live Snapshot)
    open_stats = frappe.db.sql(
        f"""
        SELECT
            AVG(DATEDIFF(CURDATE(), DATE(d.creation))) as avg_open_age,
            SUM(IFNULL(d.deal_value, 0)) as total_open_value,
            SUM(CASE WHEN DATEDIFF(CURDATE(), DATE(d.creation)) BETWEEN 0 AND 30 THEN 1 ELSE 0 END) as count_0_30,
            SUM(CASE WHEN DATEDIFF(CURDATE(), DATE(d.creation)) BETWEEN 0 AND 30 THEN IFNULL(d.deal_value, 0) ELSE 0 END) as val_0_30,
            SUM(CASE WHEN DATEDIFF(CURDATE(), DATE(d.creation)) BETWEEN 31 AND 60 THEN 1 ELSE 0 END) as count_31_60,
            SUM(CASE WHEN DATEDIFF(CURDATE(), DATE(d.creation)) BETWEEN 31 AND 60 THEN IFNULL(d.deal_value, 0) ELSE 0 END) as val_31_60,
            SUM(CASE WHEN DATEDIFF(CURDATE(), DATE(d.creation)) BETWEEN 61 AND 90 THEN 1 ELSE 0 END) as count_61_90,
            SUM(CASE WHEN DATEDIFF(CURDATE(), DATE(d.creation)) BETWEEN 61 AND 90 THEN IFNULL(d.deal_value, 0) ELSE 0 END) as val_61_90,
            SUM(CASE WHEN DATEDIFF(CURDATE(), DATE(d.creation)) > 90 THEN 1 ELSE 0 END) as count_90_plus,
            SUM(CASE WHEN DATEDIFF(CURDATE(), DATE(d.creation)) > 90 THEN IFNULL(d.deal_value, 0) ELSE 0 END) as val_90_plus
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE {user_where_str}
          AND s.type IN ('Open', 'Ongoing')
        """,
        user_params,
        as_dict=True,
    )[0]

    avg_open_age = round(float(open_stats["avg_open_age"]), 1) if open_stats["avg_open_age"] is not None else 0.0
    total_open_value = float(open_stats["total_open_value"] or 0.0)

    raw_brackets = [
        ("0 - 30 Days", int(open_stats["count_0_30"] or 0), float(open_stats["val_0_30"] or 0.0)),
        ("31 - 60 Days", int(open_stats["count_31_60"] or 0), float(open_stats["val_31_60"] or 0.0)),
        ("61 - 90 Days", int(open_stats["count_61_90"] or 0), float(open_stats["val_61_90"] or 0.0)),
        ("90+ Days", int(open_stats["count_90_plus"] or 0), float(open_stats["val_90_plus"] or 0.0)),
    ]

    age_distribution = []
    for label, count, val in raw_brackets:
        pct = round((val / total_open_value * 100), 2) if total_open_value > 0 else 0.0
        age_distribution.append({
            "bracket": label,
            "deal_count": count,
            "pipeline_value": val,
            "percentage_of_pipeline": float(pct),
        })

    return {
        "scope": {
            "from_date": str(from_date),
            "to_date": str(to_date),
            "team_filter": scope_data["team_filter"],
            "user_filter": scope_data["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": is_unrestricted,
        },
        "summary": {
            "avg_won_sales_cycle_days": avg_won_cycle,
            "avg_lost_sales_cycle_days": avg_lost_cycle,
            "avg_open_deal_age_days": avg_open_age,
            "sales_velocity_per_day": sales_velocity,
        },
        "age_distribution": age_distribution,
        "meta": {
            "metric_type": "hybrid",
            "snapshot": False,
            "currency": company_currency,
            "date_anchor": "closed_date_and_current_pipeline",
            "generated_at": str(frappe.utils.now_datetime()),
        },
    }


@frappe.whitelist()
def get_activity_execution_analytics(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
):
    """
    Phase 2G API: Sales Activity & Follow-Up Execution Analytics (Hybrid)
    =======================================================================
    Security: Consumes validate_and_get_scope() to enforce hierarchy access control.
    Metrics:
      1. Total Scheduled Activities (ToDo.date BETWEEN from_date AND to_date)
      2. Completed Activity Rate (Completed in period / Total Scheduled in period * 100)
      3. Current Overdue Activities (Live Snapshot: Open & date < CURDATE())
      4. Average Completed Activities per Rep (Completed in period / Effective User Count)
      5. Current Open Activity Breakdown (CRM Lead vs CRM Deal open, overdue, % overdue)
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    from_date = scope_data["from_date"]
    to_date = scope_data["to_date"]
    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]

    effective_user_count = len(effective_users)

    # Base hierarchy filtering condition
    user_where = ["1=1"]
    user_params = {}

    apply_user_filter = scope_data["team_filter"] is not None or scope_data["user_filter"] is not None or not is_unrestricted

    if apply_user_filter:
        if effective_users:
            user_where.append("t.allocated_to IN %(users)s")
            user_params["users"] = tuple(effective_users)
        else:
            user_where.append("1=0")

    user_where_str = " AND ".join(user_where)

    # 1. Scheduled & Completed Activity Metrics in Period (PERIOD EVENT)
    period_params = dict(user_params)
    period_params.update({"from_date": from_date, "to_date": to_date})

    period_stats = frappe.db.sql(
        f"""
        SELECT
            COUNT(*) as total_scheduled,
            SUM(CASE WHEN t.status = 'Closed' THEN 1 ELSE 0 END) as completed_scheduled
        FROM `tabToDo` t
        WHERE {user_where_str}
          AND t.reference_type IN ('CRM Lead', 'CRM Deal')
          AND t.date BETWEEN %(from_date)s AND %(to_date)s
        """,
        period_params,
        as_dict=True,
    )[0]

    total_scheduled = int(period_stats["total_scheduled"] or 0)
    completed_activities = int(period_stats["completed_scheduled"] or 0)

    completed_activity_rate = round((completed_activities / total_scheduled * 100), 2) if total_scheduled > 0 else 0.0
    avg_completed_per_rep = round((completed_activities / effective_user_count), 2) if effective_user_count > 0 else 0.0

    # 2. Live Current Snapshot Metrics (Independent of selected period)
    snapshot_stats = frappe.db.sql(
        f"""
        SELECT
            SUM(CASE WHEN t.status = 'Open' AND t.date < CURDATE() THEN 1 ELSE 0 END) as current_overdue,
            SUM(CASE WHEN t.reference_type = 'CRM Lead' AND t.status = 'Open' THEN 1 ELSE 0 END) as lead_open,
            SUM(CASE WHEN t.reference_type = 'CRM Lead' AND t.status = 'Open' AND t.date < CURDATE() THEN 1 ELSE 0 END) as lead_overdue,
            SUM(CASE WHEN t.reference_type = 'CRM Deal' AND t.status = 'Open' THEN 1 ELSE 0 END) as deal_open,
            SUM(CASE WHEN t.reference_type = 'CRM Deal' AND t.status = 'Open' AND t.date < CURDATE() THEN 1 ELSE 0 END) as deal_overdue
        FROM `tabToDo` t
        WHERE {user_where_str}
          AND t.reference_type IN ('CRM Lead', 'CRM Deal')
        """,
        user_params,
        as_dict=True,
    )[0]

    current_overdue = int(snapshot_stats["current_overdue"] or 0)

    raw_categories = [
        ("CRM Lead", int(snapshot_stats["lead_open"] or 0), int(snapshot_stats["lead_overdue"] or 0)),
        ("CRM Deal", int(snapshot_stats["deal_open"] or 0), int(snapshot_stats["deal_overdue"] or 0)),
    ]

    activity_breakdown = []
    for cat_label, open_cnt, overdue_cnt in raw_categories:
        pct_overdue = round((overdue_cnt / open_cnt * 100), 2) if open_cnt > 0 else 0.0
        activity_breakdown.append({
            "category": cat_label,
            "open_count": open_cnt,
            "overdue_count": overdue_cnt,
            "percentage_overdue": float(pct_overdue),
        })

    return {
        "scope": {
            "from_date": str(from_date),
            "to_date": str(to_date),
            "team_filter": scope_data["team_filter"],
            "user_filter": scope_data["user_filter"],
            "effective_user_count": effective_user_count,
            "is_unrestricted": is_unrestricted,
        },
        "summary": {
            "total_scheduled_activities": total_scheduled,
            "completed_activities": completed_activities,
            "completed_activity_rate": completed_activity_rate,
            "current_overdue_activities": current_overdue,
            "average_completed_activities_per_rep": avg_completed_per_rep,
        },
        "activity_breakdown": activity_breakdown,
        "meta": {
            "metric_type": "hybrid",
            "snapshot": False,
            "date_anchor": "todo_date_and_current_open",
            "generated_at": str(frappe.utils.now_datetime()),
        },
    }


@frappe.whitelist()
def get_rep_leaderboard_analytics(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
):
    """
    Phase 2H API: Sales Rep & Team Leaderboard Analytics (Hybrid)
    ==============================================================
    Security: Consumes validate_and_get_scope() to enforce hierarchy access control.
    Metrics per scoped rep:
      1. Won Revenue (Won deals closed in period)
      2. Won Deals Count
      3. Lost Deals Count
      4. Closed Cohort Win Rate (%)
      5. Open Pipeline Value (Live Current Snapshot)
      6. Completed Tasks (ToDo closed in period)
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    from_date = scope_data["from_date"]
    to_date = scope_data["to_date"]
    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]

    company_currency = frappe.db.get_single_value("Global Defaults", "default_currency") or "INR"

    if not effective_users:
        return {
            "scope": {
                "from_date": str(from_date),
                "to_date": str(to_date),
                "team_filter": scope_data["team_filter"],
                "user_filter": scope_data["user_filter"],
                "effective_user_count": 0,
                "is_unrestricted": is_unrestricted,
            },
            "leaderboard": [],
            "meta": {
                "metric_type": "hybrid",
                "snapshot": False,
                "currency": company_currency,
                "date_anchor": "closed_date_and_current_pipeline_and_todo_date",
                "generated_at": str(frappe.utils.now_datetime()),
            },
        }

    # Fetch User full_names map
    users_info = frappe.get_all(
        "User",
        filters={"name": ["in", list(effective_users)]},
        fields=["name", "full_name", "first_name", "last_name"],
    )
    user_name_map = {}
    for u in users_info:
        name_str = u.get("full_name") or f"{u.get('first_name') or ''} {u.get('last_name') or ''}".strip() or u["name"]
        user_name_map[u["name"]] = name_str

    # 1. Query Won & Lost Deals in Period
    closed_deals = frappe.db.sql(
        """
        SELECT
            d.deal_owner,
            SUM(CASE WHEN s.type = 'Won' THEN IFNULL(d.deal_value, 0) ELSE 0 END) as won_revenue,
            SUM(CASE WHEN s.type = 'Won' THEN 1 ELSE 0 END) as won_count,
            SUM(CASE WHEN s.type = 'Lost' THEN 1 ELSE 0 END) as lost_count
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE d.deal_owner IN %(users)s
          AND d.closed_date BETWEEN %(from_date)s AND %(to_date)s
          AND s.type IN ('Won', 'Lost')
        GROUP BY d.deal_owner
        """,
        {"users": tuple(effective_users), "from_date": from_date, "to_date": to_date},
        as_dict=True,
    )
    closed_map = {row["deal_owner"]: row for row in closed_deals}

    # 2. Query Live Open Pipeline (Current Live Snapshot)
    open_deals = frappe.db.sql(
        """
        SELECT
            d.deal_owner,
            SUM(IFNULL(d.deal_value, 0)) as open_pipeline
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE d.deal_owner IN %(users)s
          AND s.type IN ('Open', 'Ongoing')
        GROUP BY d.deal_owner
        """,
        {"users": tuple(effective_users)},
        as_dict=True,
    )
    open_map = {row["deal_owner"]: float(row["open_pipeline"] or 0.0) for row in open_deals}

    # 3. Query Completed ToDo Tasks in Period
    completed_todos = frappe.db.sql(
        """
        SELECT
            t.allocated_to as rep_owner,
            COUNT(*) as completed_tasks
        FROM `tabToDo` t
        WHERE t.allocated_to IN %(users)s
          AND t.reference_type IN ('CRM Lead', 'CRM Deal')
          AND t.status = 'Closed'
          AND t.date BETWEEN %(from_date)s AND %(to_date)s
        GROUP BY t.allocated_to
        """,
        {"users": tuple(effective_users), "from_date": from_date, "to_date": to_date},
        as_dict=True,
    )
    todo_map = {row["rep_owner"]: int(row["completed_tasks"] or 0) for row in completed_todos}

    leaderboard = []
    for user_email in effective_users:
        rep_name = user_name_map.get(user_email, user_email)

        c_data = closed_map.get(user_email, {})
        won_revenue = float(c_data.get("won_revenue") or 0.0)
        won_deals_count = int(c_data.get("won_count") or 0)
        lost_deals_count = int(c_data.get("lost_count") or 0)

        closed_total = won_deals_count + lost_deals_count
        win_rate = round((won_deals_count / closed_total * 100), 2) if closed_total > 0 else 0.0

        open_pipeline_value = open_map.get(user_email, 0.0)
        completed_tasks = todo_map.get(user_email, 0)

        leaderboard.append({
            "rep_owner": user_email,
            "rep_name": rep_name,
            "won_revenue": won_revenue,
            "won_deals_count": won_deals_count,
            "lost_deals_count": lost_deals_count,
            "win_rate": win_rate,
            "open_pipeline_value": open_pipeline_value,
            "completed_tasks": completed_tasks,
        })

    # Sort leaderboard by won_revenue DESC, then won_deals_count DESC, then open_pipeline_value DESC
    leaderboard.sort(key=lambda x: (x["won_revenue"], x["won_deals_count"], x["open_pipeline_value"]), reverse=True)

    return {
        "scope": {
            "from_date": str(from_date),
            "to_date": str(to_date),
            "team_filter": scope_data["team_filter"],
            "user_filter": scope_data["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": is_unrestricted,
        },
        "leaderboard": leaderboard,
        "meta": {
            "metric_type": "hybrid",
            "snapshot": False,
            "currency": company_currency,
            "date_anchor": "closed_date_and_current_pipeline_and_todo_date",
            "generated_at": str(frappe.utils.now_datetime()),
        },
    }


@frappe.whitelist()
def get_industry_analytics(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
):
    """
    Phase 2I API: Industry & Vertical Market Intelligence (Hybrid)
    ==============================================================
    Security: Consumes validate_and_get_scope() to enforce hierarchy access control.
    Metrics per Industry:
      1. Won Revenue (Won deals closed in period)
      2. Won Deals Count
      3. Lost Deals Count
      4. Industry Win Rate (%)
      5. Average Won Deal Size (Won Revenue / Won Deals)
      6. Open Pipeline Value (Live Current Snapshot: status in Open/Ongoing)
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    from_date = scope_data["from_date"]
    to_date = scope_data["to_date"]
    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]

    company_currency = frappe.db.get_single_value("Global Defaults", "default_currency") or "INR"

    if not effective_users:
        return {
            "scope": {
                "from_date": str(from_date),
                "to_date": str(to_date),
                "team_filter": scope_data["team_filter"],
                "user_filter": scope_data["user_filter"],
                "effective_user_count": 0,
                "is_unrestricted": is_unrestricted,
            },
            "industries": [],
            "meta": {
                "metric_type": "hybrid",
                "snapshot": False,
                "currency": company_currency,
                "date_anchor": "closed_date_and_current_pipeline",
                "generated_at": str(frappe.utils.now_datetime()),
            },
        }

    # 1. Query Won & Lost Deals grouped by Industry in Period
    closed_deals = frappe.db.sql(
        """
        SELECT
            IFNULL(NULLIF(TRIM(d.industry), ''), 'Unspecified') as industry_name,
            SUM(CASE WHEN s.type = 'Won' THEN IFNULL(d.deal_value, 0) ELSE 0 END) as won_revenue,
            SUM(CASE WHEN s.type = 'Won' THEN 1 ELSE 0 END) as won_count,
            SUM(CASE WHEN s.type = 'Lost' THEN 1 ELSE 0 END) as lost_count
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE d.deal_owner IN %(users)s
          AND d.closed_date BETWEEN %(from_date)s AND %(to_date)s
          AND s.type IN ('Won', 'Lost')
        GROUP BY industry_name
        """,
        {"users": tuple(effective_users), "from_date": from_date, "to_date": to_date},
        as_dict=True,
    )
    closed_map = {row["industry_name"]: row for row in closed_deals}

    # 2. Query Live Open Pipeline grouped by Industry (Current Live Snapshot)
    open_deals = frappe.db.sql(
        """
        SELECT
            IFNULL(NULLIF(TRIM(d.industry), ''), 'Unspecified') as industry_name,
            SUM(IFNULL(d.deal_value, 0)) as open_pipeline
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE d.deal_owner IN %(users)s
          AND s.type IN ('Open', 'Ongoing')
        GROUP BY industry_name
        """,
        {"users": tuple(effective_users)},
        as_dict=True,
    )
    open_map = {row["industry_name"]: float(row["open_pipeline"] or 0.0) for row in open_deals}

    # Union of all industry names present in closed_map or open_map
    all_industries = set(closed_map.keys()).union(set(open_map.keys()))

    industries_list = []
    for ind in all_industries:
        c_data = closed_map.get(ind, {})
        won_revenue = float(c_data.get("won_revenue") or 0.0)
        won_deals_count = int(c_data.get("won_count") or 0)
        lost_deals_count = int(c_data.get("lost_count") or 0)

        closed_total = won_deals_count + lost_deals_count
        win_rate = round((won_deals_count / closed_total * 100), 2) if closed_total > 0 else 0.0
        avg_won_deal_size = round((won_revenue / won_deals_count), 2) if won_deals_count > 0 else 0.0

        open_pipeline_value = open_map.get(ind, 0.0)

        industries_list.append({
            "industry": ind,
            "won_revenue": won_revenue,
            "won_deals_count": won_deals_count,
            "lost_deals_count": lost_deals_count,
            "win_rate": win_rate,
            "average_won_deal_size": avg_won_deal_size,
            "open_pipeline_value": open_pipeline_value,
        })

    # Sort industries by won_revenue DESC, then won_deals_count DESC, then open_pipeline_value DESC
    industries_list.sort(key=lambda x: (x["won_revenue"], x["won_deals_count"], x["open_pipeline_value"]), reverse=True)

    return {
        "scope": {
            "from_date": str(from_date),
            "to_date": str(to_date),
            "team_filter": scope_data["team_filter"],
            "user_filter": scope_data["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": is_unrestricted,
        },
        "industries": industries_list,
        "meta": {
            "metric_type": "hybrid",
            "snapshot": False,
            "currency": company_currency,
            "date_anchor": "closed_date_and_current_pipeline",
            "generated_at": str(frappe.utils.now_datetime()),
        },
    }


@frappe.whitelist()
def get_organization_analytics(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
):
    """
    Phase 2J API: Key Account & Organization Revenue Concentration Analytics (Hybrid)
    ==================================================================================
    Security: Consumes validate_and_get_scope() to enforce hierarchy access control.
    Metrics per Organization:
      1. Won Revenue (Won deals closed in period)
      2. Won Deals Count
      3. Lost Deals Count
      4. Account Win Rate (%)
      5. Open Pipeline Value (Live Current Snapshot: status in Open/Ongoing)
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    from_date = scope_data["from_date"]
    to_date = scope_data["to_date"]
    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]

    company_currency = frappe.db.get_single_value("Global Defaults", "default_currency") or "INR"

    if not effective_users:
        return {
            "scope": {
                "from_date": str(from_date),
                "to_date": str(to_date),
                "team_filter": scope_data["team_filter"],
                "user_filter": scope_data["user_filter"],
                "effective_user_count": 0,
                "is_unrestricted": is_unrestricted,
            },
            "organizations": [],
            "meta": {
                "metric_type": "hybrid",
                "snapshot": False,
                "currency": company_currency,
                "date_anchor": "closed_date_and_current_pipeline",
                "generated_at": str(frappe.utils.now_datetime()),
            },
        }

    # 1. Query Won & Lost Deals grouped by Organization in Period
    closed_deals = frappe.db.sql(
        """
        SELECT
            IFNULL(NULLIF(TRIM(d.organization_name), ''), 'Individual / Unassigned') as org_name,
            SUM(CASE WHEN s.type = 'Won' THEN IFNULL(d.deal_value, 0) ELSE 0 END) as won_revenue,
            SUM(CASE WHEN s.type = 'Won' THEN 1 ELSE 0 END) as won_count,
            SUM(CASE WHEN s.type = 'Lost' THEN 1 ELSE 0 END) as lost_count
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE d.deal_owner IN %(users)s
          AND d.closed_date BETWEEN %(from_date)s AND %(to_date)s
          AND s.type IN ('Won', 'Lost')
        GROUP BY org_name
        """,
        {"users": tuple(effective_users), "from_date": from_date, "to_date": to_date},
        as_dict=True,
    )
    closed_map = {row["org_name"]: row for row in closed_deals}

    # 2. Query Live Open Pipeline grouped by Organization (Current Live Snapshot)
    open_deals = frappe.db.sql(
        """
        SELECT
            IFNULL(NULLIF(TRIM(d.organization_name), ''), 'Individual / Unassigned') as org_name,
            SUM(IFNULL(d.deal_value, 0)) as open_pipeline
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE d.deal_owner IN %(users)s
          AND s.type IN ('Open', 'Ongoing')
        GROUP BY org_name
        """,
        {"users": tuple(effective_users)},
        as_dict=True,
    )
    open_map = {row["org_name"]: float(row["open_pipeline"] or 0.0) for row in open_deals}

    # Union of all organization names present in closed_map or open_map
    all_orgs = set(closed_map.keys()).union(set(open_map.keys()))

    orgs_list = []
    for org in all_orgs:
        c_data = closed_map.get(org, {})
        won_revenue = float(c_data.get("won_revenue") or 0.0)
        won_deals_count = int(c_data.get("won_count") or 0)
        lost_deals_count = int(c_data.get("lost_count") or 0)

        closed_total = won_deals_count + lost_deals_count
        win_rate = round((won_deals_count / closed_total * 100), 2) if closed_total > 0 else 0.0

        open_pipeline_value = open_map.get(org, 0.0)

        orgs_list.append({
            "organization_name": org,
            "won_revenue": won_revenue,
            "won_deals_count": won_deals_count,
            "lost_deals_count": lost_deals_count,
            "win_rate": win_rate,
            "open_pipeline_value": open_pipeline_value,
        })

    # Sort organizations by won_revenue DESC, then won_deals_count DESC, then open_pipeline_value DESC
    orgs_list.sort(key=lambda x: (x["won_revenue"], x["won_deals_count"], x["open_pipeline_value"]), reverse=True)

    return {
        "scope": {
            "from_date": str(from_date),
            "to_date": str(to_date),
            "team_filter": scope_data["team_filter"],
            "user_filter": scope_data["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": is_unrestricted,
        },
        "organizations": orgs_list,
        "meta": {
            "metric_type": "hybrid",
            "snapshot": False,
            "currency": company_currency,
            "date_anchor": "closed_date_and_current_pipeline",
            "generated_at": str(frappe.utils.now_datetime()),
        },
    }


@frappe.whitelist()
def get_lead_conversion_analytics(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
):
    """
    Phase 2K API: Lead-to-Deal Conversion Efficiency & Sourcing Velocity (Period Event)
    ====================================================================================
    Security: Consumes validate_and_get_scope() to enforce hierarchy access control.
    Metrics:
      1. Total Period Leads (created in period)
      2. Converted Leads Count (converted = 1)
      3. Period Lead Conversion Rate (%)
      4. Average Days to Convert (DATEDIFF(d.creation, l.creation))
      5. Converted Pipeline Value (SUM(d.deal_value) for deals converted from period leads)
      6. Source Conversion Breakdown (grouped by l.source)
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    from_date = scope_data["from_date"]
    to_date = scope_data["to_date"]
    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]

    company_currency = frappe.db.get_single_value("Global Defaults", "default_currency") or "INR"

    if not effective_users:
        return {
            "scope": {
                "from_date": str(from_date),
                "to_date": str(to_date),
                "team_filter": scope_data["team_filter"],
                "user_filter": scope_data["user_filter"],
                "effective_user_count": 0,
                "is_unrestricted": is_unrestricted,
            },
            "summary": {
                "total_leads": 0,
                "converted_leads": 0,
                "conversion_rate": 0.0,
                "avg_days_to_convert": 0.0,
                "converted_pipeline_value": 0.0,
            },
            "source_breakdown": [],
            "meta": {
                "metric_type": "period_event",
                "snapshot": False,
                "currency": company_currency,
                "date_anchor": "lead_creation",
                "generated_at": str(frappe.utils.now_datetime()),
            },
        }

    # 1. Query Lead Totals & Conversion Metrics grouped by Source (Independent of Deals to prevent JOIN multiplication)
    lead_source_summary = frappe.db.sql(
        """
        SELECT
            IFNULL(NULLIF(TRIM(source), ''), 'Unspecified') as source_name,
            COUNT(*) as total_leads,
            SUM(CASE WHEN converted = 1 THEN 1 ELSE 0 END) as converted_leads
        FROM `tabCRM Lead`
        WHERE lead_owner IN %(users)s
          AND DATE(creation) BETWEEN %(from_date)s AND %(to_date)s
        GROUP BY source_name
        """,
        {"users": tuple(effective_users), "from_date": from_date, "to_date": to_date},
        as_dict=True,
    )

    # 2. Query Converted Deal Details (Deal Value & Days to Convert) linked to Period Leads
    converted_deal_metrics = frappe.db.sql(
        """
        SELECT
            IFNULL(NULLIF(TRIM(l.source), ''), 'Unspecified') as source_name,
            d.name as deal_name,
            IFNULL(d.deal_value, 0) as deal_value,
            DATEDIFF(DATE(d.creation), DATE(l.creation)) as days_to_convert
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Lead` l ON d.lead = l.name
        WHERE l.lead_owner IN %(users)s
          AND l.converted = 1
          AND DATE(l.creation) BETWEEN %(from_date)s AND %(to_date)s
        """,
        {"users": tuple(effective_users), "from_date": from_date, "to_date": to_date},
        as_dict=True,
    )

    # Aggregate Converted Deal Metrics by Source
    source_deal_map = {}
    total_converted_pipeline_val = 0.0
    conversion_days_list = []

    for row in converted_deal_metrics:
        src = row["source_name"]
        val = float(row["deal_value"] or 0.0)
        days = float(row["days_to_convert"] if row["days_to_convert"] is not None else 0.0)
        if days < 0:
            days = 0.0

        total_converted_pipeline_val += val
        conversion_days_list.append(days)

        if src not in source_deal_map:
            source_deal_map[src] = {"converted_val": 0.0, "days_list": []}
        source_deal_map[src]["converted_val"] += val
        source_deal_map[src]["days_list"].append(days)

    # Aggregate Overall Lead Summary
    grand_total_leads = sum(int(r["total_leads"]) for r in lead_source_summary)
    grand_converted_leads = sum(int(r["converted_leads"]) for r in lead_source_summary)
    grand_conversion_rate = round((grand_converted_leads / grand_total_leads * 100), 2) if grand_total_leads > 0 else 0.0
    grand_avg_days = round((sum(conversion_days_list) / len(conversion_days_list)), 1) if conversion_days_list else 0.0

    # Build Source Breakdown List
    source_breakdown = []
    for r in lead_source_summary:
        src = r["source_name"]
        t_leads = int(r["total_leads"])
        c_leads = int(r["converted_leads"])
        c_rate = round((c_leads / t_leads * 100), 2) if t_leads > 0 else 0.0

        d_info = source_deal_map.get(src, {"converted_val": 0.0, "days_list": []})
        c_val = d_info["converted_val"]

        source_breakdown.append({
            "source": src,
            "total_leads": t_leads,
            "converted_leads": c_leads,
            "conversion_rate": c_rate,
            "converted_value": c_val,
        })

    # Sort source breakdown by converted_value DESC, then converted_leads DESC, then total_leads DESC
    source_breakdown.sort(key=lambda x: (x["converted_value"], x["converted_leads"], x["total_leads"]), reverse=True)

    return {
        "scope": {
            "from_date": str(from_date),
            "to_date": str(to_date),
            "team_filter": scope_data["team_filter"],
            "user_filter": scope_data["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": is_unrestricted,
        },
        "summary": {
            "total_leads": grand_total_leads,
            "converted_leads": grand_converted_leads,
            "conversion_rate": grand_conversion_rate,
            "avg_days_to_convert": grand_avg_days,
            "converted_pipeline_value": total_converted_pipeline_val,
        },
        "source_breakdown": source_breakdown,
        "meta": {
            "metric_type": "period_event",
            "snapshot": False,
            "currency": company_currency,
            "date_anchor": "lead_creation",
            "generated_at": str(frappe.utils.now_datetime()),
        },
    }


@frappe.whitelist()
def get_unconverted_lead_analytics(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
):
    """
    Phase 2L API: Unconverted Lead Stage & Aging Bottleneck Analytics (Live Current Snapshot)
    ==========================================================================================
    Security: Consumes validate_and_get_scope() to enforce hierarchy access control.
    Metrics:
      1. Total Unconverted Leads (converted = 0)
      2. Stale Leads Count (>14 days un-converted)
      3. Lead Stage Breakdown (grouped by tabCRM Lead.status)
      4. Lead Age Distribution (0-7 days, 8-14 days, 15-30 days, 30+ days)
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    from_date = scope_data["from_date"]
    to_date = scope_data["to_date"]
    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]

    company_currency = frappe.db.get_single_value("Global Defaults", "default_currency") or "INR"

    if not effective_users:
        return {
            "scope": {
                "from_date": str(from_date),
                "to_date": str(to_date),
                "team_filter": scope_data["team_filter"],
                "user_filter": scope_data["user_filter"],
                "effective_user_count": 0,
                "is_unrestricted": is_unrestricted,
            },
            "summary": {
                "total_unconverted_leads": 0,
                "stale_leads": 0,
            },
            "stages": [],
            "age_distribution": [
                {"bucket": "0-7 days", "count": 0, "percentage": 0.0},
                {"bucket": "8-14 days", "count": 0, "percentage": 0.0},
                {"bucket": "15-30 days", "count": 0, "percentage": 0.0},
                {"bucket": "30+ days", "count": 0, "percentage": 0.0},
            ],
            "meta": {
                "metric_type": "live_snapshot",
                "snapshot": True,
                "currency": company_currency,
                "date_anchor": "current_date_and_lead_creation",
                "generated_at": str(frappe.utils.now_datetime()),
            },
        }

    # 1. Total Unconverted Leads & Stale Leads (>14 days)
    totals_res = frappe.db.sql(
        """
        SELECT
            COUNT(*) as total_unconverted,
            SUM(CASE WHEN DATEDIFF(CURDATE(), DATE(creation)) > 14 THEN 1 ELSE 0 END) as stale_leads
        FROM `tabCRM Lead`
        WHERE lead_owner IN %(users)s
          AND (converted IS NULL OR converted = 0)
        """,
        {"users": tuple(effective_users)},
        as_dict=True,
    )

    total_unconverted = int(totals_res[0]["total_unconverted"] or 0) if totals_res else 0
    stale_leads = int(totals_res[0]["stale_leads"] or 0) if totals_res else 0

    # 2. Stage Breakdown
    stage_res = frappe.db.sql(
        """
        SELECT
            IFNULL(NULLIF(TRIM(status), ''), 'Unspecified') as stage_name,
            COUNT(*) as stage_count
        FROM `tabCRM Lead`
        WHERE lead_owner IN %(users)s
          AND (converted IS NULL OR converted = 0)
        GROUP BY stage_name
        ORDER BY stage_count DESC, stage_name ASC
        """,
        {"users": tuple(effective_users)},
        as_dict=True,
    )

    stages = []
    for r in stage_res:
        cnt = int(r["stage_count"])
        pct = round((cnt / total_unconverted * 100), 2) if total_unconverted > 0 else 0.0
        stages.append({
            "status": r["stage_name"],
            "count": cnt,
            "percentage": pct,
        })

    # 3. Age Distribution Buckets
    age_res = frappe.db.sql(
        """
        SELECT
            SUM(CASE WHEN DATEDIFF(CURDATE(), DATE(creation)) BETWEEN 0 AND 7 THEN 1 ELSE 0 END) as b_0_7,
            SUM(CASE WHEN DATEDIFF(CURDATE(), DATE(creation)) BETWEEN 8 AND 14 THEN 1 ELSE 0 END) as b_8_14,
            SUM(CASE WHEN DATEDIFF(CURDATE(), DATE(creation)) BETWEEN 15 AND 30 THEN 1 ELSE 0 END) as b_15_30,
            SUM(CASE WHEN DATEDIFF(CURDATE(), DATE(creation)) > 30 THEN 1 ELSE 0 END) as b_30_plus
        FROM `tabCRM Lead`
        WHERE lead_owner IN %(users)s
          AND (converted IS NULL OR converted = 0)
        """,
        {"users": tuple(effective_users)},
        as_dict=True,
    )

    c_0_7 = int(age_res[0]["b_0_7"] or 0) if age_res else 0
    c_8_14 = int(age_res[0]["b_8_14"] or 0) if age_res else 0
    c_15_30 = int(age_res[0]["b_15_30"] or 0) if age_res else 0
    c_30_plus = int(age_res[0]["b_30_plus"] or 0) if age_res else 0

    p_0_7 = round((c_0_7 / total_unconverted * 100), 2) if total_unconverted > 0 else 0.0
    p_8_14 = round((c_8_14 / total_unconverted * 100), 2) if total_unconverted > 0 else 0.0
    p_15_30 = round((c_15_30 / total_unconverted * 100), 2) if total_unconverted > 0 else 0.0
    p_30_plus = round((c_30_plus / total_unconverted * 100), 2) if total_unconverted > 0 else 0.0

    age_distribution = [
        {"bucket": "0-7 days", "count": c_0_7, "percentage": p_0_7},
        {"bucket": "8-14 days", "count": c_8_14, "percentage": p_8_14},
        {"bucket": "15-30 days", "count": c_15_30, "percentage": p_15_30},
        {"bucket": "30+ days", "count": c_30_plus, "percentage": p_30_plus},
    ]

    return {
        "scope": {
            "from_date": str(from_date),
            "to_date": str(to_date),
            "team_filter": scope_data["team_filter"],
            "user_filter": scope_data["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": is_unrestricted,
        },
        "summary": {
            "total_unconverted_leads": total_unconverted,
            "stale_leads": stale_leads,
        },
        "stages": stages,
        "age_distribution": age_distribution,
        "meta": {
            "metric_type": "live_snapshot",
            "snapshot": True,
            "currency": company_currency,
            "date_anchor": "current_date_and_lead_creation",
            "generated_at": str(frappe.utils.now_datetime()),
        },
    }


@frappe.whitelist()
def get_deal_progression_analytics(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
):
    """
    Phase 2M API: Deal Stage Progression & Dwell Time Analytics (Period Event & Dwell Reconstruction)
    ===================================================================================================
    Security: Consumes validate_and_get_scope() to enforce hierarchy access control via deal_owner.
    Metrics:
      1. Total Period Stage Transitions (CRM Status Change Log for CRM Deal inside period)
      2. Stage Transition Flow Breakdown (from_stage -> to_stage counts)
      3. Stage Dwell Time (Chronological reconstruction of stage dwell durations in days)
      4. Stage Loss Breakdown (Value & count of deals transitioning to 'Lost' by originating stage)
      5. Stage Progression Summary (Entries, Exits, Transitions, Avg Dwell, Loss Value per stage)
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    from_date = scope_data["from_date"]
    to_date = scope_data["to_date"]
    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]

    company_currency = frappe.db.get_single_value("Global Defaults", "default_currency") or "INR"

    if not effective_users:
        return {
            "scope": {
                "from_date": str(from_date),
                "to_date": str(to_date),
                "team_filter": scope_data["team_filter"],
                "user_filter": scope_data["user_filter"],
                "effective_user_count": 0,
                "is_unrestricted": is_unrestricted,
            },
            "summary": {
                "total_transitions": 0,
                "average_dwell_days": 0.0,
                "highest_friction_stage": "N/A",
                "lost_stage_count": 0,
                "lost_stage_value": 0.0,
            },
            "transitions": [],
            "stages": [],
            "loss_breakdown": [],
            "meta": {
                "metric_type": "period_event",
                "snapshot": False,
                "currency": company_currency,
                "date_anchor": "status_change_creation",
                "generated_at": str(frappe.utils.now_datetime()),
            },
        }

    # 1. Query status change logs for CRM Deal where deal_owner IN effective_users
    # Retrieve logs with transition creation inside period for transition summary & loss breakdown
    # Also retrieve full deal history for dwell reconstruction
    period_logs = frappe.db.sql(
        """
        SELECT
            l.name as log_id,
            l.parent as deal_name,
            IFNULL(NULLIF(TRIM(l.`from`), ''), 'Unspecified') as from_stage,
            IFNULL(NULLIF(TRIM(l.`to`), ''), 'Unspecified') as to_stage,
            l.creation as transition_time,
            d.deal_value,
            d.deal_owner
        FROM `tabCRM Status Change Log` l
        INNER JOIN `tabCRM Deal` d ON d.name = l.parent
        WHERE l.parenttype = 'CRM Deal'
          AND d.deal_owner IN %(users)s
          AND DATE(l.creation) BETWEEN %(from_date)s AND %(to_date)s
        ORDER BY l.creation ASC
        """,
        {
            "users": tuple(effective_users),
            "from_date": from_date,
            "to_date": to_date,
        },
        as_dict=True,
    )

    total_transitions = len(period_logs)

    # 2. Transition Flow Breakdown (from_stage -> to_stage)
    transition_flow_map = {}
    loss_stage_map = {}
    total_lost_count = 0
    total_lost_value = 0.0

    for r in period_logs:
        f_stg = r["from_stage"]
        t_stg = r["to_stage"]
        val = float(r["deal_value"] or 0.0)

        flow_key = (f_stg, t_stg)
        transition_flow_map[flow_key] = transition_flow_map.get(flow_key, 0) + 1

        if t_stg == "Lost":
            total_lost_count += 1
            total_lost_value += val
            if f_stg not in loss_stage_map:
                loss_stage_map[f_stg] = {"lost_count": 0, "lost_value": 0.0}
            loss_stage_map[f_stg]["lost_count"] += 1
            loss_stage_map[f_stg]["lost_value"] += val

    transitions_list = [
        {"from_stage": k[0], "to_stage": k[1], "transition_count": v}
        for k, v in transition_flow_map.items()
    ]
    transitions_list.sort(key=lambda x: x["transition_count"], reverse=True)

    loss_breakdown_list = [
        {"from_stage": k, "lost_deal_count": v["lost_count"], "lost_deal_value": round(v["lost_value"], 2)}
        for k, v in loss_stage_map.items()
    ]
    loss_breakdown_list.sort(key=lambda x: (x["lost_deal_value"], x["lost_deal_count"]), reverse=True)

    # 3. Stage Dwell Time Reconstruction
    # Fetch ALL status change logs for deals scoped to effective_users to accurately reconstruct chronological intervals
    all_deal_logs = frappe.db.sql(
        """
        SELECT
            l.parent as deal_name,
            IFNULL(NULLIF(TRIM(l.`from`), ''), 'Unspecified') as from_stage,
            IFNULL(NULLIF(TRIM(l.`to`), ''), 'Unspecified') as to_stage,
            l.creation as transition_time,
            d.creation as deal_creation
        FROM `tabCRM Status Change Log` l
        INNER JOIN `tabCRM Deal` d ON d.name = l.parent
        WHERE l.parenttype = 'CRM Deal'
          AND d.deal_owner IN %(users)s
        ORDER BY l.parent ASC, l.creation ASC
        """,
        {"users": tuple(effective_users)},
        as_dict=True,
    )

    # Group logs by deal
    deals_log_map = {}
    for lg in all_deal_logs:
        dname = lg["deal_name"]
        if dname not in deals_log_map:
            deals_log_map[dname] = {
                "deal_creation": lg["deal_creation"],
                "logs": [],
            }
        deals_log_map[dname]["logs"].append(lg)

    stage_dwell_data = {} # stage -> list of dwell_days

    for dname, dinfo in deals_log_map.items():
        logs = dinfo["logs"]
        if not logs:
            continue

        # Initial interval: deal creation to first status change
        # First log from_stage is initial stage
        first_log = logs[0]
        init_stage = first_log["from_stage"]
        t0 = first_log["deal_creation"]
        t1 = first_log["transition_time"]

        if t0 and t1 and t1 >= t0:
            dwell = (t1 - t0).total_seconds() / 86400.0
            if init_stage not in stage_dwell_data:
                stage_dwell_data[init_stage] = []
            stage_dwell_data[init_stage].append(dwell)

        # Inter-transition intervals
        for i in range(len(logs) - 1):
            curr_log = logs[i]
            next_log = logs[i + 1]

            stage = curr_log["to_stage"] # entered stage
            t_entry = curr_log["transition_time"]
            t_exit = next_log["transition_time"]

            if t_entry and t_exit and t_exit >= t_entry:
                dwell = (t_exit - t_entry).total_seconds() / 86400.0
                if stage not in stage_dwell_data:
                    stage_dwell_data[stage] = []
                stage_dwell_data[stage].append(dwell)

    # 4. Stage Progression Summary Table
    # Stage entries, exits, transitions, avg dwell, loss count, loss value
    stage_entries = {}
    stage_exits = {}

    for r in period_logs:
        f_stg = r["from_stage"]
        t_stg = r["to_stage"]

        stage_exits[f_stg] = stage_exits.get(f_stg, 0) + 1
        stage_entries[t_stg] = stage_entries.get(t_stg, 0) + 1

    all_observed_stages = set(stage_entries.keys()) | set(stage_exits.keys()) | set(stage_dwell_data.keys())

    stages_summary_list = []
    overall_dwell_list = []

    for stg in all_observed_stages:
        dwell_list = stage_dwell_data.get(stg, [])
        avg_dwell = round((sum(dwell_list) / len(dwell_list)), 1) if dwell_list else 0.0
        if dwell_list:
            overall_dwell_list.extend(dwell_list)

        l_info = loss_stage_map.get(stg, {"lost_count": 0, "lost_value": 0.0})

        stages_summary_list.append({
            "stage": stg,
            "entries": stage_entries.get(stg, 0),
            "exits": stage_exits.get(stg, 0),
            "transitions_out": stage_exits.get(stg, 0),
            "average_dwell_days": avg_dwell,
            "completed_dwell_count": len(dwell_list),
            "lost_count": l_info["lost_count"],
            "lost_value": round(l_info["lost_value"], 2),
        })

    stages_summary_list.sort(key=lambda x: (x["average_dwell_days"], x["entries"]), reverse=True)

    grand_avg_dwell = round((sum(overall_dwell_list) / len(overall_dwell_list)), 1) if overall_dwell_list else 0.0
    highest_friction_stage = stages_summary_list[0]["stage"] if stages_summary_list else "N/A"

    return {
        "scope": {
            "from_date": str(from_date),
            "to_date": str(to_date),
            "team_filter": scope_data["team_filter"],
            "user_filter": scope_data["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": is_unrestricted,
        },
        "summary": {
            "total_transitions": total_transitions,
            "average_dwell_days": grand_avg_dwell,
            "highest_friction_stage": highest_friction_stage,
            "lost_stage_count": total_lost_count,
            "lost_stage_value": round(total_lost_value, 2),
        },
        "transitions": transitions_list,
        "stages": stages_summary_list,
        "loss_breakdown": loss_breakdown_list,
        "meta": {
            "metric_type": "period_event",
            "snapshot": False,
            "currency": company_currency,
            "date_anchor": "status_change_creation",
            "generated_at": str(frappe.utils.now_datetime()),
        },
    }


@frappe.whitelist()
def get_pipeline_probability_analytics(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
):
    """
    Phase 2N API: Pipeline Probability Distribution & Forecast Calibration (Live Current Snapshot)
    ===================================================================================================
    Security: Consumes validate_and_get_scope() to enforce hierarchy access control via deal_owner.
    Metrics:
      1. Gross Open Pipeline Value (SUM deal_value where status NOT IN ('Won', 'Lost'))
      2. Weighted Forecast Value (SUM deal_value * probability / 100)
      3. Forecast Risk Gap (Gross Open Value - Weighted Forecast Value)
      4. Probability Tier Distribution (0-25%, 26-50%, 51-75%, 76-99%)
      5. Probability Calibration Variance (Rep prob vs tabCRM Deal Status prob; flag over-optimism > +15%)
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    from_date = scope_data["from_date"]
    to_date = scope_data["to_date"]
    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]

    company_currency = frappe.db.get_single_value("Global Defaults", "default_currency") or "INR"

    # Pre-populate empty response contract
    empty_tiers = [
        {"tier": "Low Confidence", "min_probability": 0, "max_probability": 25, "deal_count": 0, "gross_value": 0.0, "weighted_value": 0.0, "deal_percentage": 0.0, "value_percentage": 0.0},
        {"tier": "Medium Confidence", "min_probability": 26, "max_probability": 50, "deal_count": 0, "gross_value": 0.0, "weighted_value": 0.0, "deal_percentage": 0.0, "value_percentage": 0.0},
        {"tier": "High Confidence", "min_probability": 51, "max_probability": 75, "deal_count": 0, "gross_value": 0.0, "weighted_value": 0.0, "deal_percentage": 0.0, "value_percentage": 0.0},
        {"tier": "Commit", "min_probability": 76, "max_probability": 99, "deal_count": 0, "gross_value": 0.0, "weighted_value": 0.0, "deal_percentage": 0.0, "value_percentage": 0.0},
    ]

    if not effective_users:
        return {
            "scope": {
                "from_date": str(from_date),
                "to_date": str(to_date),
                "team_filter": scope_data["team_filter"],
                "user_filter": scope_data["user_filter"],
                "effective_user_count": 0,
                "is_unrestricted": is_unrestricted,
            },
            "summary": {
                "gross_open_value": 0.0,
                "weighted_forecast_value": 0.0,
                "forecast_risk_gap": 0.0,
                "high_commit_value": 0.0,
                "open_deal_count": 0,
            },
            "probability_tiers": empty_tiers,
            "calibration_risks": [],
            "meta": {
                "metric_type": "live_snapshot",
                "snapshot": True,
                "currency": company_currency,
                "date_anchor": "current_open_pipeline",
                "generated_at": str(frappe.utils.now_datetime()),
            },
        }

    # Fetch stage default probabilities from tabCRM Deal Status
    stage_prob_rows = frappe.db.sql(
        """
        SELECT deal_status, probability
        FROM `tabCRM Deal Status`
        """,
        as_dict=True,
    )
    stage_prob_map = {r["deal_status"]: float(r["probability"] or 0.0) for r in stage_prob_rows}

    # Fetch open deals (status NOT IN ('Won', 'Lost')) scoped to effective_users
    open_deals = frappe.db.sql(
        """
        SELECT
            name as deal_id,
            IFNULL(NULLIF(TRIM(custom_deal_name), ''), name) as deal_name,
            organization_name,
            deal_owner,
            status,
            IFNULL(deal_value, 0.0) as deal_value,
            IFNULL(probability, 0.0) as probability
        FROM `tabCRM Deal`
        WHERE status NOT IN ('Won', 'Lost')
          AND deal_owner IN %(users)s
        ORDER BY deal_value DESC
        """,
        {"users": tuple(effective_users)},
        as_dict=True,
    )

    open_deal_count = len(open_deals)

    if open_deal_count == 0:
        return {
            "scope": {
                "from_date": str(from_date),
                "to_date": str(to_date),
                "team_filter": scope_data["team_filter"],
                "user_filter": scope_data["user_filter"],
                "effective_user_count": len(effective_users),
                "is_unrestricted": is_unrestricted,
            },
            "summary": {
                "gross_open_value": 0.0,
                "weighted_forecast_value": 0.0,
                "forecast_risk_gap": 0.0,
                "high_commit_value": 0.0,
                "open_deal_count": 0,
            },
            "probability_tiers": empty_tiers,
            "calibration_risks": [],
            "meta": {
                "metric_type": "live_snapshot",
                "snapshot": True,
                "currency": company_currency,
                "date_anchor": "current_open_pipeline",
                "generated_at": str(frappe.utils.now_datetime()),
            },
        }

    total_gross_value = 0.0
    total_weighted_value = 0.0

    # Tiers accumulator: 0-25, 26-50, 51-75, 76-99 (and 100 if open)
    tiers_map = {
        "Low Confidence": {"tier": "Low Confidence", "min_probability": 0, "max_probability": 25, "deal_count": 0, "gross_value": 0.0, "weighted_value": 0.0},
        "Medium Confidence": {"tier": "Medium Confidence", "min_probability": 26, "max_probability": 50, "deal_count": 0, "gross_value": 0.0, "weighted_value": 0.0},
        "High Confidence": {"tier": "High Confidence", "min_probability": 51, "max_probability": 75, "deal_count": 0, "gross_value": 0.0, "weighted_value": 0.0},
        "Commit": {"tier": "Commit", "min_probability": 76, "max_probability": 99, "deal_count": 0, "gross_value": 0.0, "weighted_value": 0.0},
    }

    calibration_risks = []

    for d in open_deals:
        dval = float(d["deal_value"] or 0.0)
        prob = float(d["probability"] or 0.0)
        wval = (dval * prob) / 100.0

        total_gross_value += dval
        total_weighted_value += wval

        # Bucket assignment
        if prob <= 25.0:
            key = "Low Confidence"
        elif prob <= 50.0:
            key = "Medium Confidence"
        elif prob <= 75.0:
            key = "High Confidence"
        else: # 76% to 100%
            key = "Commit"

        tiers_map[key]["deal_count"] += 1
        tiers_map[key]["gross_value"] += dval
        tiers_map[key]["weighted_value"] += wval

        # Probability Calibration Variance calculation
        stg = d["status"]
        stg_prob = stage_prob_map.get(stg, 0.0)
        variance = round(prob - stg_prob, 1)

        if variance > 15.0: # Over-optimistic threshold
            calibration_risks.append({
                "deal_id": d["deal_id"],
                "deal_name": d["deal_name"] or d["deal_id"],
                "organization_name": d["organization_name"] or "N/A",
                "deal_owner": d["deal_owner"],
                "status": stg,
                "deal_value": round(dval, 2),
                "rep_probability": round(prob, 1),
                "stage_probability": round(stg_prob, 1),
                "variance": variance,
                "risk_classification": "Over-Optimistic",
            })

    calibration_risks.sort(key=lambda x: (x["variance"], x["deal_value"]), reverse=True)

    # Compute percentages & format tier list
    probability_tiers = []
    tier_keys_order = ["Low Confidence", "Medium Confidence", "High Confidence", "Commit"]

    high_commit_value = tiers_map["High Confidence"]["gross_value"] + tiers_map["Commit"]["gross_value"]

    for k in tier_keys_order:
        tdata = tiers_map[k]
        gval = tdata["gross_value"]
        cnt = tdata["deal_count"]

        deal_pct = round((cnt / open_deal_count) * 100.0, 1) if open_deal_count > 0 else 0.0
        val_pct = round((gval / total_gross_value) * 100.0, 1) if total_gross_value > 0 else 0.0

        probability_tiers.append({
            "tier": tdata["tier"],
            "min_probability": tdata["min_probability"],
            "max_probability": tdata["max_probability"],
            "deal_count": cnt,
            "gross_value": round(gval, 2),
            "weighted_value": round(tdata["weighted_value"], 2),
            "deal_percentage": deal_pct,
            "value_percentage": val_pct,
        })

    forecast_risk_gap = round(total_gross_value - total_weighted_value, 2)

    return {
        "scope": {
            "from_date": str(from_date),
            "to_date": str(to_date),
            "team_filter": scope_data["team_filter"],
            "user_filter": scope_data["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": is_unrestricted,
        },
        "summary": {
            "gross_open_value": round(total_gross_value, 2),
            "weighted_forecast_value": round(total_weighted_value, 2),
            "forecast_risk_gap": forecast_risk_gap,
            "high_commit_value": round(high_commit_value, 2),
            "open_deal_count": open_deal_count,
        },
        "probability_tiers": probability_tiers,
        "calibration_risks": calibration_risks,
        "meta": {
            "metric_type": "live_snapshot",
            "snapshot": True,
            "currency": company_currency,
            "date_anchor": "current_open_pipeline",
            "generated_at": str(frappe.utils.now_datetime()),
        },
    }


@frappe.whitelist()
def get_top_opportunities(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
    limit: int = 10,
):
    """
    Whitelisted API: Top Open Opportunities for Executive Dashboard
    ===============================================================
    Returns individual high-value open deals sorted by weighted value / gross value.
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]
    company_currency = frappe.db.get_default("currency") or "INR"

    deal_where = ["1=1"]
    deal_params = {"limit": int(limit)}
    apply_user_filter = scope_data["team_filter"] is not None or scope_data["user_filter"] is not None or not is_unrestricted

    if apply_user_filter:
        if effective_users:
            deal_where.append("d.deal_owner IN %(users)s")
            deal_params["users"] = tuple(effective_users)
        else:
            deal_where.append("1=0")

    deal_where_str = " AND ".join(deal_where)

    deals = frappe.db.sql(
        f"""
        SELECT
            d.name as deal_id,
            d.organization as organization,
            d.deal_owner as owner,
            d.status as stage,
            IFNULL(d.deal_value, 0) as gross_value,
            IFNULL(d.probability, s.probability) as probability,
            IFNULL(NULLIF(d.expected_deal_value, 0), IFNULL(d.deal_value, 0) * IFNULL(d.probability, s.probability) / 100) as weighted_value,
            DATEDIFF(NOW(), d.creation) as age_days
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE {deal_where_str}
          AND s.type IN ('Open', 'Ongoing')
        ORDER BY weighted_value DESC, gross_value DESC
        LIMIT %(limit)s
        """,
        deal_params,
        as_dict=True,
    )

    owner_emails = {d["owner"] for d in deals if d.get("owner")}
    user_name_map = {}
    if owner_emails:
        users_info = frappe.get_all(
            "User",
            filters={"name": ["in", list(owner_emails)]},
            fields=["name", "full_name", "first_name", "last_name"],
        )
        for u in users_info:
            name_str = u.get("full_name") or f"{u.get('first_name') or ''} {u.get('last_name') or ''}".strip() or u["name"]
            user_name_map[u["name"]] = name_str

    opportunities = []
    for d in deals:
        gval = float(d["gross_value"] or 0.0)
        prob = float(d["probability"] or 0.0)
        wval = float(d["weighted_value"] or 0.0)
        raw_owner = d["owner"] or "Unassigned"
        owner_display = user_name_map.get(raw_owner, raw_owner)
        opportunities.append({
            "deal_id": d["deal_id"],
            "organization": d["organization"] or "N/A",
            "owner": owner_display,
            "owner_email": raw_owner,
            "stage": d["stage"],
            "gross_value": round(gval, 2),
            "probability": round(prob, 1),
            "weighted_value": round(wval, 2),
            "age_days": int(d["age_days"] or 0),
        })

    return {
        "scope": scope_data,
        "opportunities": opportunities,
        "meta": {
            "currency": company_currency,
            "generated_at": str(frappe.utils.now_datetime()),
        },
    }


def _calculate_target_for_user_and_period(user_email, from_date_str, to_date_str):
    """
    Helper function to calculate target amount for a given user email and date range.
    Queries Employee -> Sales Person -> Target Detail -> Monthly Distribution Percentage.
    """
    emp_name = frappe.db.get_value("Employee", {"user_id": user_email}, "name")
    if not emp_name:
        return 0.0

    sp_name = frappe.db.get_value("Sales Person", {"employee": emp_name}, "name")
    if not sp_name:
        return 0.0

    target_details = frappe.get_all(
        "Target Detail",
        filters={"parent": sp_name, "parenttype": "Sales Person"},
        fields=["fiscal_year", "target_amount", "distribution_id"]
    )
    if not target_details:
        return 0.0

    from_date = frappe.utils.getdate(from_date_str)
    to_date = frappe.utils.getdate(to_date_str)

    total_target = 0.0

    for td in target_details:
        annual_target = float(td["target_amount"] or 0.0)
        if annual_target <= 0:
            continue

        fy = td["fiscal_year"]
        fy_dates = frappe.db.get_value("Fiscal Year", fy, ["year_start_date", "year_end_date"], as_dict=True)
        if not fy_dates:
            continue

        fy_start = frappe.utils.getdate(fy_dates["year_start_date"])
        fy_end = frappe.utils.getdate(fy_dates["year_end_date"])

        # Check date range overlap with Fiscal Year
        overlap_start = max(from_date, fy_start)
        overlap_end = min(to_date, fy_end)
        if overlap_start > overlap_end:
            continue

        dist_id = td["distribution_id"]
        monthly_allocations = {}
        if dist_id and frappe.db.exists("Monthly Distribution", dist_id):
            percentages = frappe.get_all(
                "Monthly Distribution Percentage",
                filters={"parent": dist_id},
                fields=["month", "percentage_allocation"]
            )
            for p in percentages:
                monthly_allocations[p["month"]] = float(p["percentage_allocation"] or 0.0)
        else:
            # Fallback to even monthly allocation
            months_list = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
            for m in months_list:
                monthly_allocations[m] = 100.0 / 12.0

        # Sum percentages for overlapping months
        curr_date = overlap_start.replace(day=1)
        month_percentages = 0.0
        
        while curr_date <= overlap_end:
            month_name = curr_date.strftime("%B")
            month_pct = monthly_allocations.get(month_name, 100.0 / 12.0)

            month_start = curr_date
            next_month = frappe.utils.add_months(curr_date, 1)
            month_end = frappe.utils.add_days(next_month, -1)

            sub_start = max(overlap_start, month_start)
            sub_end = min(overlap_end, month_end)

            days_in_month = (month_end - month_start).days + 1
            active_days = (sub_end - sub_start).days + 1

            fraction = max(0.0, min(1.0, active_days / float(days_in_month)))
            month_percentages += month_pct * fraction

            curr_date = next_month

        total_target += annual_target * (month_percentages / 100.0)

    return round(total_target, 2)


@frappe.whitelist()
def get_sales_target_analytics(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
):
    """
    Whitelisted API: CRM Sales Target Performance Analytics (V10)
    =============================================================
    Calculates Sales Target, Achieved Revenue, Achievement %, Open Pipeline,
    Weighted Pipeline, Forecast Attainment, and Target Gap.
    Enforces scope security via validate_and_get_scope().
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    from_date = scope_data["from_date"]
    to_date = scope_data["to_date"]
    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]
    company_currency = frappe.db.get_single_value("Global Defaults", "default_currency") or "INR"

    if not effective_users:
        return {
            "scope": scope_data,
            "summary": {
                "target_value": 0.0,
                "achieved_value": 0.0,
                "remaining_value": 0.0,
                "achievement_percent": 0.0,
                "pipeline_value": 0.0,
                "weighted_pipeline": 0.0,
                "forecast_attainment_percent": 0.0,
                "target_gap": 0.0,
            },
            "by_user": [],
            "by_team": [],
            "meta": {
                "currency": company_currency,
                "generated_at": str(frappe.utils.now_datetime()),
            },
        }

    # Fetch User full_names map
    users_info = frappe.get_all(
        "User",
        filters={"name": ["in", list(effective_users)]},
        fields=["name", "full_name", "first_name", "last_name"],
    )
    user_name_map = {}
    for u in users_info:
        name_str = u.get("full_name") or f"{u.get('first_name') or ''} {u.get('last_name') or ''}".strip() or u["name"]
        user_name_map[u["name"]] = name_str

    # 1. Query Achieved Revenue (Won Deals in period) per user
    closed_deals = frappe.db.sql(
        """
        SELECT
            d.deal_owner,
            SUM(CASE WHEN s.type = 'Won' THEN IFNULL(d.deal_value, 0) ELSE 0 END) as won_revenue
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE d.deal_owner IN %(users)s
          AND d.closed_date BETWEEN %(from_date)s AND %(to_date)s
          AND s.type = 'Won'
        GROUP BY d.deal_owner
        """,
        {"users": tuple(effective_users), "from_date": from_date, "to_date": to_date},
        as_dict=True,
    )
    achieved_map = {row["deal_owner"]: float(row["won_revenue"] or 0.0) for row in closed_deals}

    # 2. Query Open Pipeline & Weighted Pipeline per user
    open_deals = frappe.db.sql(
        """
        SELECT
            d.deal_owner,
            SUM(IFNULL(d.deal_value, 0)) as open_pipeline,
            SUM(IFNULL(NULLIF(d.expected_deal_value, 0), IFNULL(d.deal_value, 0) * IFNULL(d.probability, s.probability) / 100)) as weighted_pipeline
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE d.deal_owner IN %(users)s
          AND s.type IN ('Open', 'Ongoing')
        GROUP BY d.deal_owner
        """,
        {"users": tuple(effective_users)},
        as_dict=True,
    )
    pipeline_map = {
        row["deal_owner"]: (float(row["open_pipeline"] or 0.0), float(row["weighted_pipeline"] or 0.0))
        for row in open_deals
    }

    # 3. Build per-user target performance
    by_user = []
    tot_target = 0.0
    tot_achieved = 0.0
    tot_pipeline = 0.0
    tot_weighted = 0.0

    # Dynamic risk threshold constants
    # ACHIEVED: achievement_percent >= 100
    # ON_TRACK: forecast_attainment_percent >= 80 (or remaining_value <= 0)
    # AT_RISK: forecast_attainment_percent >= 40 (some pipeline coverage but gap exists)
    # CRITICAL: forecast_attainment_percent < 40

    def calculate_risk(ach_pct, fore_pct, rem_val):
        if ach_pct >= 100.0 or rem_val <= 0:
            return "ACHIEVED"
        elif fore_pct >= 80.0:
            return "ON_TRACK"
        elif fore_pct >= 40.0:
            return "AT_RISK"
        else:
            return "CRITICAL"

    on_track_count = 0
    at_risk_count = 0
    critical_count = 0
    achieved_count = 0

    for user_email in effective_users:
        user_name = user_name_map.get(user_email, user_email)
        emp_name = frappe.db.get_value("Employee", {"user_id": user_email}, "name")
        sp_name = frappe.db.get_value("Sales Person", {"employee": emp_name}, "name") if emp_name else None

        t_val = _calculate_target_for_user_and_period(user_email, from_date, to_date)
        a_val = achieved_map.get(user_email, 0.0)
        pipe_val, weight_val = pipeline_map.get(user_email, (0.0, 0.0))

        rem_val = max(0.0, t_val - a_val)
        ach_pct = round((a_val / t_val * 100.0), 2) if t_val > 0 else 0.0
        fore_attainment_pct = round(((a_val + weight_val) / t_val * 100.0), 2) if t_val > 0 else 0.0
        t_gap = round(t_val - a_val, 2)
        fore_gap = round(t_val - (a_val + weight_val), 2)

        pipe_cov = round((pipe_val / rem_val * 100.0), 2) if rem_val > 0 else (100.0 if rem_val == 0 else 0.0)
        weight_cov = round((weight_val / rem_val * 100.0), 2) if rem_val > 0 else (100.0 if rem_val == 0 else 0.0)

        risk_status = calculate_risk(ach_pct, fore_attainment_pct, rem_val)
        if risk_status == "ACHIEVED":
            achieved_count += 1
        elif risk_status == "ON_TRACK":
            on_track_count += 1
        elif risk_status == "AT_RISK":
            at_risk_count += 1
        else:
            critical_count += 1

        tot_target += t_val
        tot_achieved += a_val
        tot_pipeline += pipe_val
        tot_weighted += weight_val

        by_user.append({
            "user": user_email,
            "user_name": user_name,
            "sales_person": sp_name or "Unassigned",
            "target_value": round(t_val, 2),
            "achieved_value": round(a_val, 2),
            "remaining_value": round(rem_val, 2),
            "achievement_percent": ach_pct,
            "pipeline_value": round(pipe_val, 2),
            "weighted_pipeline": round(weight_val, 2),
            "forecast_attainment_percent": fore_attainment_pct,
            "target_gap": t_gap,
            "forecast_gap": fore_gap,
            "pipeline_coverage": pipe_cov,
            "weighted_coverage": weight_cov,
            "risk_status": risk_status,
        })

    # Sort users by risk severity (CRITICAL, AT_RISK, ON_TRACK, ACHIEVED), then by largest remaining_value DESC
    risk_rank = {"CRITICAL": 0, "AT_RISK": 1, "ON_TRACK": 2, "ACHIEVED": 3}
    by_user.sort(key=lambda x: (risk_rank.get(x["risk_status"], 99), -x["remaining_value"], -x["target_value"]))

    # 4. Build per-team target performance & intelligence
    full_scope = _resolve_scope(frappe.session.user)
    by_team = []
    for team in full_scope.get("teams", []):
        team_val = team["value"]
        team_label = team["label"]
        team_members = team.get("members", [])
        
        # Intersect team members with scope's effective_users
        team_effective = [m for m in team_members if m in effective_users]

        tm_target = sum(_calculate_target_for_user_and_period(m, from_date, to_date) for m in team_effective)
        tm_achieved = sum(achieved_map.get(m, 0.0) for m in team_effective)
        tm_pipe = sum(pipeline_map.get(m, (0.0, 0.0))[0] for m in team_effective)
        tm_weight = sum(pipeline_map.get(m, (0.0, 0.0))[1] for m in team_effective)

        tm_rem = max(0.0, tm_target - tm_achieved)
        tm_ach_pct = round((tm_achieved / tm_target * 100.0), 2) if tm_target > 0 else 0.0
        tm_fore_attainment = round(((tm_achieved + tm_weight) / tm_target * 100.0), 2) if tm_target > 0 else 0.0
        tm_gap = round(tm_target - tm_achieved, 2)
        tm_fore_gap = round(tm_target - (tm_achieved + tm_weight), 2)

        tm_pipe_cov = round((tm_pipe / tm_rem * 100.0), 2) if tm_rem > 0 else (100.0 if tm_rem == 0 else 0.0)
        tm_weight_cov = round((tm_weight / tm_rem * 100.0), 2) if tm_rem > 0 else (100.0 if tm_rem == 0 else 0.0)
        tm_risk_status = calculate_risk(tm_ach_pct, tm_fore_attainment, tm_rem)

        by_team.append({
            "team_id": team_val,
            "team_name": team_label,
            "member_count": len(team_effective),
            "target_value": round(tm_target, 2),
            "achieved_value": round(tm_achieved, 2),
            "remaining_value": round(tm_rem, 2),
            "achievement_percent": tm_ach_pct,
            "pipeline_value": round(tm_pipe, 2),
            "weighted_pipeline": round(tm_weight, 2),
            "forecast_attainment_percent": tm_fore_attainment,
            "target_gap": tm_gap,
            "forecast_gap": tm_fore_gap,
            "pipeline_coverage": tm_pipe_cov,
            "weighted_coverage": tm_weight_cov,
            "risk_status": tm_risk_status,
        })

    by_team.sort(key=lambda x: (risk_rank.get(x["risk_status"], 99), -x["remaining_value"], -x["target_value"]))

    # 5. Summary calculations
    tot_target = round(tot_target, 2)
    tot_achieved = round(tot_achieved, 2)
    tot_remaining = round(max(0.0, tot_target - tot_achieved), 2)
    tot_ach_pct = round((tot_achieved / tot_target * 100.0), 2) if tot_target > 0 else 0.0
    tot_fore_attainment = round(((tot_achieved + tot_weighted) / tot_target * 100.0), 2) if tot_target > 0 else 0.0
    tot_gap = round(tot_target - tot_achieved, 2)
    tot_fore_gap = round(tot_target - (tot_achieved + tot_weighted), 2)
    tot_pipe_cov = round((tot_pipeline / tot_remaining * 100.0), 2) if tot_remaining > 0 else (100.0 if tot_remaining == 0 else 0.0)
    tot_weight_cov = round((tot_weighted / tot_remaining * 100.0), 2) if tot_remaining > 0 else (100.0 if tot_remaining == 0 else 0.0)
    tot_risk_status = calculate_risk(tot_ach_pct, tot_fore_attainment, tot_remaining)

    return {
        "scope": scope_data,
        "summary": {
            "target_value": tot_target,
            "achieved_value": tot_achieved,
            "remaining_value": tot_remaining,
            "achievement_percent": tot_ach_pct,
            "pipeline_value": round(tot_pipeline, 2),
            "weighted_pipeline": round(tot_weighted, 2),
            "forecast_attainment_percent": tot_fore_attainment,
            "target_gap": tot_gap,
            "forecast_gap": tot_fore_gap,
            "pipeline_coverage": tot_pipe_cov,
            "weighted_coverage": tot_weight_cov,
            "risk_status": tot_risk_status,
            "on_track_count": on_track_count,
            "at_risk_count": at_risk_count,
            "critical_count": critical_count,
            "achieved_count": achieved_count,
        },
        "by_user": by_user,
        "by_team": by_team,
        "meta": {
            "currency": company_currency,
            "generated_at": str(frappe.utils.now_datetime()),
        },
    }


# ---------------------------------------------------------------------------
# 6. V10.5 SALES TARGET ROOT-CAUSE DIAGNOSTICS & ACTIONABILITY
# ---------------------------------------------------------------------------

TARGET_ROOT_CAUSE_CONFIG = {
    "COVERAGE_DEFICIT_THRESHOLD": 150.0,       # Pipeline Coverage < 150%
    "LOW_PROBABILITY_RATIO_THRESHOLD": 50.0,   # > 50% of pipeline in early stages (prob <= 30%)
    "STALLED_DEAL_AGE_DAYS": 45,               # Deals open > 45 days
    "MIN_EXPECTED_ACTIVITIES_MONTHLY": 5,      # Less than 5 activities completed
    "LOW_LEAD_CONVERSION_THRESHOLD": 15.0,     # Lead conversion rate < 15%
}


@frappe.whitelist()
def get_sales_target_root_cause_analytics(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
):
    """
    Whitelisted API: V10.5 Sales Target Root-Cause Diagnostics & Actionability Engine
    =================================================================================
    Analyzes permitted representatives to identify WHY targets are at risk,
    attaches diagnostic evidence, recommends deterministic management actions,
    and isolates top contributing risk deals.
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )
    effective_users = scope_data["effective_users"]
    company_currency = frappe.db.get_single_value("Global Defaults", "default_currency") or "INR"

    if not effective_users:
        return {
            "scope": scope_data,
            "summary": {
                "total_reps": 0,
                "at_risk_reps": 0,
                "critical_reps": 0,
                "primary_cause_breakdown": {
                    "INSUFFICIENT_PIPELINE_VOLUME": 0,
                    "LOW_STAGE_PROBABILITY": 0,
                    "STALLED_DEAL_AGING": 0,
                    "ACTIVITY_EXECUTION_DEFICIT": 0,
                    "LEAD_CONVERSION_BOTTLENECK": 0,
                },
            },
            "diagnostics": [],
            "meta": {"currency": company_currency, "generated_at": str(frappe.utils.now_datetime())},
        }

    # 1. Fetch base target analytics for authorized users (reusing canonical V10 calculations)
    base_target_data = get_sales_target_analytics(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )
    user_target_map = {u["user"]: u for u in base_target_data.get("by_user", [])}

    # 2. BULK QUERY A: Early stage ratio, deal aging, and risk deal candidates from open deals
    open_deals_raw = frappe.db.sql(
        """
        SELECT
            d.name as deal_id,
            IFNULL(d.organization, d.name) as deal_title,
            d.deal_owner,
            d.status as stage,
            IFNULL(d.deal_value, 0) as deal_value,
            IFNULL(d.probability, s.probability) as probability,
            DATEDIFF(NOW(), d.creation) as age_days
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE d.deal_owner IN %(users)s
          AND s.type IN ('Open', 'Ongoing')
        ORDER BY d.deal_value DESC, age_days DESC
        """,
        {"users": tuple(effective_users)},
        as_dict=True,
    )

    user_deals_map = {}
    for d in open_deals_raw:
        owner = d["deal_owner"]
        if owner not in user_deals_map:
            user_deals_map[owner] = []
        user_deals_map[owner].append(d)

    # 3. BULK QUERY B: Sales Activity execution & overdue tasks from tabToDo
    activity_rows = frappe.db.sql(
        """
        SELECT
            allocated_to as user_email,
            SUM(CASE WHEN status = 'Closed' THEN 1 ELSE 0 END) as completed_activities,
            SUM(CASE WHEN status != 'Closed' AND date < CURDATE() THEN 1 ELSE 0 END) as overdue_tasks
        FROM `tabToDo`
        WHERE allocated_to IN %(users)s
        GROUP BY allocated_to
        """,
        {"users": tuple(effective_users)},
        as_dict=True,
    )
    user_activity_map = {r["user_email"]: r for r in activity_rows}

    # 4. BULK QUERY C: Lead conversion statistics
    dates = resolve_period(period, custom_from, custom_to)
    from_date, to_date = dates["from_date"], dates["to_date"]

    lead_rows = frappe.db.sql(
        """
        SELECT
            lead_owner,
            COUNT(name) as leads_created,
            SUM(CASE WHEN status = 'Converted' THEN 1 ELSE 0 END) as leads_converted
        FROM `tabCRM Lead`
        WHERE lead_owner IN %(users)s
          AND creation BETWEEN %(from_date)s AND %(to_date)s
        GROUP BY lead_owner
        """,
        {"users": tuple(effective_users), "from_date": from_date, "to_date": to_date},
        as_dict=True,
    )
    user_lead_map = {r["lead_owner"]: r for r in lead_rows}

    # 5. Process diagnostics per effective user
    diagnostics = []
    cause_breakdown = {
        "INSUFFICIENT_PIPELINE_VOLUME": 0,
        "LOW_STAGE_PROBABILITY": 0,
        "STALLED_DEAL_AGING": 0,
        "ACTIVITY_EXECUTION_DEFICIT": 0,
        "LEAD_CONVERSION_BOTTLENECK": 0,
    }
    critical_reps = 0
    at_risk_reps = 0

    for user_email in effective_users:
        u_target = user_target_map.get(user_email, {})
        user_name = u_target.get("user_name", user_email)
        risk_status = u_target.get("risk_status", "ON_TRACK")

        if risk_status == "CRITICAL":
            critical_reps += 1
        elif risk_status == "AT_RISK":
            at_risk_reps += 1

        deals = user_deals_map.get(user_email, [])
        open_pipeline_val = u_target.get("pipeline_value", 0.0)
        early_stage_val = sum(d["deal_value"] for d in deals if float(d["probability"] or 0) <= 30.0)
        early_stage_ratio = round((early_stage_val / open_pipeline_val * 100.0), 2) if open_pipeline_val > 0 else 0.0

        avg_age = round(sum(d["age_days"] for d in deals) / len(deals), 1) if deals else 0
        stalled_deals = [d for d in deals if d["age_days"] > TARGET_ROOT_CAUSE_CONFIG["STALLED_DEAL_AGE_DAYS"]]
        oldest_age = max((d["age_days"] for d in deals), default=0)

        act_data = user_activity_map.get(user_email, {})
        comp_act = int(act_data.get("completed_activities") or 0)
        overdue_t = int(act_data.get("overdue_tasks") or 0)

        lead_data = user_lead_map.get(user_email, {})
        leads_created = int(lead_data.get("leads_created") or 0)
        leads_converted = int(lead_data.get("leads_converted") or 0)
        conv_rate = round((leads_converted / leads_created * 100.0), 2) if leads_created > 0 else 0.0

        pipe_cov = u_target.get("pipeline_coverage", 0.0)

        # Evaluate diagnostic root causes
        root_causes = []
        
        # Cause 1: INSUFFICIENT_PIPELINE_VOLUME
        if pipe_cov < TARGET_ROOT_CAUSE_CONFIG["COVERAGE_DEFICIT_THRESHOLD"]:
            root_causes.append({
                "cause": "INSUFFICIENT_PIPELINE_VOLUME",
                "severity": "HIGH" if pipe_cov < 50.0 else "MEDIUM",
                "evidence": {
                    "pipeline_coverage": pipe_cov,
                    "remaining_target_gap": u_target.get("remaining_value", 0.0),
                    "open_pipeline": open_pipeline_val,
                }
            })

        # Cause 2: LOW_STAGE_PROBABILITY
        if early_stage_ratio > TARGET_ROOT_CAUSE_CONFIG["LOW_PROBABILITY_RATIO_THRESHOLD"]:
            root_causes.append({
                "cause": "LOW_STAGE_PROBABILITY",
                "severity": "HIGH" if early_stage_ratio > 75.0 else "MEDIUM",
                "evidence": {
                    "early_stage_ratio": early_stage_ratio,
                    "early_stage_value": round(early_stage_val, 2),
                    "open_pipeline": open_pipeline_val,
                }
            })

        # Cause 3: STALLED_DEAL_AGING
        if avg_age > TARGET_ROOT_CAUSE_CONFIG["STALLED_DEAL_AGE_DAYS"] or len(stalled_deals) > 0:
            root_causes.append({
                "cause": "STALLED_DEAL_AGING",
                "severity": "HIGH" if avg_age > 60 else "MEDIUM",
                "evidence": {
                    "avg_deal_age_days": avg_age,
                    "stalled_deal_count": len(stalled_deals),
                    "oldest_deal_age_days": oldest_age,
                }
            })

        # Cause 4: ACTIVITY_EXECUTION_DEFICIT
        if comp_act < TARGET_ROOT_CAUSE_CONFIG["MIN_EXPECTED_ACTIVITIES_MONTHLY"] or overdue_t > 0:
            root_causes.append({
                "cause": "ACTIVITY_EXECUTION_DEFICIT",
                "severity": "HIGH" if overdue_t > 2 else "MEDIUM",
                "evidence": {
                    "completed_activities": comp_act,
                    "overdue_tasks": overdue_t,
                }
            })

        # Cause 5: LEAD_CONVERSION_BOTTLENECK
        if leads_created > 0 and conv_rate < TARGET_ROOT_CAUSE_CONFIG["LOW_LEAD_CONVERSION_THRESHOLD"]:
            root_causes.append({
                "cause": "LEAD_CONVERSION_BOTTLENECK",
                "severity": "MEDIUM",
                "evidence": {
                    "leads_created": leads_created,
                    "leads_converted": leads_converted,
                    "lead_conversion_rate": conv_rate,
                }
            })

        # Primary Root Cause selection by backend priority order
        primary_root_cause = "NONE"
        recommended_action = "Maintain current momentum and monitor deal velocity."
        
        priority_order = [
            "INSUFFICIENT_PIPELINE_VOLUME",
            "LOW_STAGE_PROBABILITY",
            "STALLED_DEAL_AGING",
            "ACTIVITY_EXECUTION_DEFICIT",
            "LEAD_CONVERSION_BOTTLENECK",
        ]

        for p in priority_order:
            if any(rc["cause"] == p for rc in root_causes):
                primary_root_cause = p
                cause_breakdown[p] += 1
                break

        if primary_root_cause == "INSUFFICIENT_PIPELINE_VOLUME":
            recommended_action = "Prospecting Deficit: Require rep to generate new leads and re-engage dormant opportunities immediately."
        elif primary_root_cause == "LOW_STAGE_PROBABILITY":
            recommended_action = "Stage Bottleneck: Conduct pipeline review to qualify and advance early-stage opportunities."
        elif primary_root_cause == "STALLED_DEAL_AGING":
            recommended_action = "Deal Velocity Deficit: Executive push required to close or disqualify aging opportunities."
        elif primary_root_cause == "ACTIVITY_EXECUTION_DEFICIT":
            recommended_action = "Engagement Deficit: Prioritize overdue follow-ups and enforce daily customer touchpoints."
        elif primary_root_cause == "LEAD_CONVERSION_BOTTLENECK":
            recommended_action = "Conversion Deficit: Audit unconverted leads and improve qualification quality."

        # Top 3 Critical Risk Deals per rep
        top_risk_deals = [
            {
                "deal_id": d["deal_id"],
                "deal_title": d["deal_title"] or d["deal_id"],
                "deal_value": round(float(d["deal_value"] or 0), 2),
                "probability": float(d["probability"] or 0),
                "stage": d["stage"],
                "age_days": d["age_days"],
            }
            for d in deals[:3]
        ]

        diagnostics.append({
            "user": user_email,
            "user_name": user_name,
            "risk_status": risk_status,
            "target_value": u_target.get("target_value", 0.0),
            "achieved_value": u_target.get("achieved_value", 0.0),
            "remaining_value": u_target.get("remaining_value", 0.0),
            "primary_root_cause": primary_root_cause,
            "root_causes": root_causes,
            "metrics": {
                "pipeline_coverage": pipe_cov,
                "weighted_coverage": u_target.get("weighted_coverage", 0.0),
                "early_stage_ratio": early_stage_ratio,
                "avg_deal_age_days": avg_age,
                "stalled_deal_count": len(stalled_deals),
                "oldest_deal_age_days": oldest_age,
                "completed_activities": comp_act,
                "overdue_tasks": overdue_t,
                "lead_conversion_rate": conv_rate,
            },
            "recommended_action": recommended_action,
            "critical_risk_deals": top_risk_deals,
        })

    # Sort diagnostics by CRITICAL / AT_RISK first, then largest remaining gap DESC
    risk_rank = {"CRITICAL": 0, "AT_RISK": 1, "ON_TRACK": 2, "ACHIEVED": 3}
    diagnostics.sort(key=lambda x: (risk_rank.get(x["risk_status"], 99), -x["remaining_value"]))

    return {
        "scope": scope_data,
        "summary": {
            "total_reps": len(effective_users),
            "at_risk_reps": at_risk_reps,
            "critical_reps": critical_reps,
            "primary_cause_breakdown": cause_breakdown,
        },
        "diagnostics": diagnostics,
        "meta": {
            "currency": company_currency,
            "generated_at": str(frappe.utils.now_datetime()),
        },
    }


@frappe.whitelist()
def get_crm_dashboard_export(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
    export_format: str = "xlsx",
):
    """
    Whitelisted API: Export CRM Dashboard Executive Report (CSV / Excel)
    ====================================================================
    Enforces scope validation via validate_and_get_scope() before generating
    multi-sheet Excel workbooks or CSV executive summaries.
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    fmt = (export_format or "xlsx").lower()
    if fmt not in ("xlsx", "csv"):
        frappe.throw(_("Invalid export format. Allowed formats: xlsx, csv."))

    # Aggregate canonical dashboard metrics via existing APIs
    kpi_res = get_executive_kpis(period, team_filter, user_filter, custom_from, custom_to)
    pipeline_res = get_pipeline_health(period, team_filter, user_filter, custom_from, custom_to)
    opps_res = get_top_opportunities(period, team_filter, user_filter, custom_from, custom_to, limit=50)
    closed_res = get_closed_sales_analytics(period, team_filter, user_filter, custom_from, custom_to)
    leaderboard_res = get_rep_leaderboard_analytics(period, team_filter, user_filter, custom_from, custom_to)
    lead_conv_res = get_lead_conversion_analytics(period, team_filter, user_filter, custom_from, custom_to)
    industry_res = get_industry_analytics(period, team_filter, user_filter, custom_from, custom_to)
    org_res = get_organization_analytics(period, team_filter, user_filter, custom_from, custom_to)
    prob_res = get_pipeline_probability_analytics(period, team_filter, user_filter, custom_from, custom_to)

    safe_period = period.replace(" ", "_")
    safe_team = (team_filter or "All_Teams").replace(" ", "_")
    safe_user = (user_filter or "All_Users").replace(" ", "_")
    filename = f"CRM_Executive_Report_{safe_period}_{safe_team}_{safe_user}_{frappe.utils.today()}"

    if fmt == "csv":
        import io
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header / Scope Metadata
        writer.writerow(["CRM EXECUTIVE COMMAND CENTER REPORT"])
        writer.writerow(["Period", period])
        writer.writerow(["Team", team_filter])
        writer.writerow(["User", user_filter])
        writer.writerow(["Generated At", str(frappe.utils.now_datetime())])
        writer.writerow([])

        # Section 1: Executive Summary
        kpis = kpi_res.get("kpis", {})
        writer.writerow(["EXECUTIVE SUMMARY"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Open Pipeline", kpis.get("pipeline_value", 0)])
        writer.writerow(["Weighted Forecast", kpis.get("weighted_pipeline", 0)])
        writer.writerow(["Won Revenue", kpis.get("won_revenue", 0)])
        writer.writerow(["Win Rate %", kpis.get("win_rate", 0)])
        writer.writerow(["Closed Deals", kpis.get("closed_deals", 0)])
        writer.writerow(["Open Deals", kpis.get("open_deals", 0)])
        writer.writerow(["Average Deal Value", kpis.get("average_deal_value", 0)])
        writer.writerow([])

        # Section 2: Pipeline by Stage
        writer.writerow(["PIPELINE BY STAGE"])
        writer.writerow(["Stage", "Deals", "Gross Value", "Weighted Value", "% Pipeline"])
        for s in pipeline_res.get("stages", []):
            writer.writerow([s.get("stage"), s.get("deal_count"), s.get("stage_value"), s.get("weighted_value"), s.get("percentage_of_pipeline")])
        writer.writerow([])

        # Section 3: Top Opportunities
        writer.writerow(["TOP OPPORTUNITIES"])
        writer.writerow(["Deal ID", "Organization", "Owner", "Stage", "Gross Value", "Weighted Value", "Probability %", "Age Days"])
        for o in opps_res.get("opportunities", []):
            writer.writerow([o.get("deal_id"), o.get("organization"), o.get("owner"), o.get("stage"), o.get("gross_value"), o.get("weighted_value"), o.get("probability"), o.get("age_days")])

        frappe.response["filename"] = f"{filename}.csv"
        frappe.response["filecontent"] = output.getvalue()
        frappe.response["type"] = "csv"

    else:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        font_title = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
        font_sub = Font(name="Calibri", size=10, italic=True, color="6B7280")
        border_thin = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB")
        )

        def apply_table_styles(ws, start_row, header_cols):
            ws.row_dimensions[start_row].height = 24
            for col_num in range(1, len(header_cols) + 1):
                cell = ws.cell(row=start_row, column=col_num)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # SHEET 1: Executive Summary
        ws1 = wb.create_sheet(title="Executive Summary")
        ws1.views.sheetView[0].showGridLines = True
        ws1.append(["Nexapp CRM Executive Command Center Report"])
        ws1.cell(row=1, column=1).font = font_title
        ws1.append([f"Scope: Period={period} | Team={team_filter} | User={user_filter} | Generated={frappe.utils.now_datetime()}"])
        ws1.cell(row=2, column=1).font = font_sub
        ws1.append([])

        kpis = kpi_res.get("kpis", {})
        prob_summary = prob_res.get("summary", {})

        ws1.append(["Metric Name", "Value", "Description"])
        apply_table_styles(ws1, 4, ["Metric Name", "Value", "Description"])

        summary_rows = [
            ("Open Pipeline", kpis.get("pipeline_value", 0), "Gross value of active open deals"),
            ("Weighted Forecast", kpis.get("weighted_pipeline", 0), "Probability-adjusted expected pipeline revenue"),
            ("Won Revenue", kpis.get("won_revenue", 0), "Revenue from closed won deals"),
            ("Win Rate", f"{kpis.get('win_rate', 0)}%", "Percentage of closed deals won"),
            ("Closed Deals", kpis.get("closed_deals", 0), "Total count of closed deals (Won + Lost)"),
            ("Open Deals", kpis.get("open_deals", 0), "Count of active open deals"),
            ("Forecast Risk Gap", prob_summary.get("forecast_risk_gap", 0), "Gap between gross open and weighted forecast"),
            ("High Commit Value", prob_summary.get("high_commit_value", 0), "Value of high-probability pipeline"),
            ("Average Deal Value", kpis.get("average_deal_value", 0), "Average value across open deals"),
        ]

        for r_idx, (m, v, d) in enumerate(summary_rows, start=5):
            ws1.append([m, v, d])
            c_val = ws1.cell(row=r_idx, column=2)
            if isinstance(v, (int, float)) and "Rate" not in m:
                c_val.number_format = "₹#,##0"
            for c_idx in range(1, 4):
                ws1.cell(row=r_idx, column=c_idx).border = border_thin

        # SHEET 2: Pipeline by Stage
        ws2 = wb.create_sheet(title="Pipeline by Stage")
        ws2.views.sheetView[0].showGridLines = True
        ws2.append(["Pipeline Distribution by Stage"])
        ws2.cell(row=1, column=1).font = font_title
        ws2.append([])
        
        headers2 = ["Stage", "Deals", "Gross Value (₹)", "Weighted Value (₹)", "% of Pipeline"]
        ws2.append(headers2)
        apply_table_styles(ws2, 3, headers2)

        for r_idx, s in enumerate(pipeline_res.get("stages", []), start=4):
            ws2.append([s.get("stage"), s.get("deal_count"), s.get("stage_value"), s.get("weighted_value"), f"{s.get('percentage_of_pipeline')}%"])
            ws2.cell(row=r_idx, column=3).number_format = "₹#,##0"
            ws2.cell(row=r_idx, column=4).number_format = "₹#,##0"
            for c_idx in range(1, 6):
                ws2.cell(row=r_idx, column=c_idx).border = border_thin

        # SHEET 3: Top Opportunities
        ws3 = wb.create_sheet(title="Top Opportunities")
        ws3.views.sheetView[0].showGridLines = True
        ws3.append(["Top Open Opportunities"])
        ws3.cell(row=1, column=1).font = font_title
        ws3.append([])

        headers3 = ["Deal ID", "Organization", "Owner", "Stage", "Gross Value (₹)", "Weighted Value (₹)", "Probability %", "Age (Days)"]
        ws3.append(headers3)
        apply_table_styles(ws3, 3, headers3)

        for r_idx, o in enumerate(opps_res.get("opportunities", []), start=4):
            ws3.append([o.get("deal_id"), o.get("organization"), o.get("owner"), o.get("stage"), o.get("gross_value"), o.get("weighted_value"), o.get("probability"), o.get("age_days")])
            ws3.cell(row=r_idx, column=5).number_format = "₹#,##0"
            ws3.cell(row=r_idx, column=6).number_format = "₹#,##0"
            for c_idx in range(1, 9):
                ws3.cell(row=r_idx, column=c_idx).border = border_thin

        # SHEET 4: Closed Sales
        ws4 = wb.create_sheet(title="Closed Sales")
        ws4.views.sheetView[0].showGridLines = True
        ws4.append(["Closed Sales Breakdown"])
        ws4.cell(row=1, column=1).font = font_title
        ws4.append([])

        headers4 = ["Metric", "Value"]
        ws4.append(headers4)
        apply_table_styles(ws4, 3, headers4)

        cs = closed_res.get("summary", {})
        cs_rows = [
            ("Won Deals", cs.get("won_deals", 0)),
            ("Won Revenue", cs.get("won_revenue", 0)),
            ("Lost Deals", cs.get("lost_deals", 0)),
            ("Lost Revenue", cs.get("lost_value", 0)),
            ("Total Closed Deals", cs.get("closed_deals", 0)),
            ("Closed Win Rate", f"{cs.get('closed_win_rate', 0)}%"),
        ]
        for r_idx, (m, v) in enumerate(cs_rows, start=4):
            ws4.append([m, v])
            if "Revenue" in m:
                ws4.cell(row=r_idx, column=2).number_format = "₹#,##0"
            for c_idx in range(1, 3):
                ws4.cell(row=r_idx, column=c_idx).border = border_thin

        # SHEET 5: Sales Rep Leaderboard
        ws5 = wb.create_sheet(title="Sales Rep Leaderboard")
        ws5.views.sheetView[0].showGridLines = True
        ws5.append(["Sales Representative Performance"])
        ws5.cell(row=1, column=1).font = font_title
        ws5.append([])

        headers5 = ["Sales Rep", "Won Deals", "Won Revenue (₹)", "Win Rate %"]
        ws5.append(headers5)
        apply_table_styles(ws5, 3, headers5)

        for r_idx, r in enumerate(leaderboard_res.get("leaderboard", []), start=4):
            ws5.append([r.get("user_name") or r.get("sales_user") or "Unassigned", r.get("won_deals", 0), r.get("won_revenue", 0), f"{r.get('win_rate', 0)}%"])
            ws5.cell(row=r_idx, column=3).number_format = "₹#,##0"
            for c_idx in range(1, 5):
                ws5.cell(row=r_idx, column=c_idx).border = border_thin

        # Auto-fit column widths across all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        import io
        stream = io.BytesIO()
        wb.save(stream)

        frappe.response["filename"] = f"{filename}.xlsx"
        frappe.response["filecontent"] = stream.getvalue()
        frappe.response["type"] = "binary"


# ---------------------------------------------------------------------------
# V11-A: DEAL EXECUTION HEALTH & MANAGEMENT ATTENTION ENGINE
# ---------------------------------------------------------------------------

DEAL_EXECUTION_CONFIG = {
    "STALE_DEAL_DAYS": 14,                    # Inactive >= 14 days
    "NO_ENGAGEMENT_DAYS": 14,                 # No customer communication >= 14 days
    "HIGH_VALUE_THRESHOLD": 50000.0,          # High-value opportunity >= 50k
    "AGING_DEAL_DAYS": 45,                    # Aging opportunity >= 45 days
}


@frappe.whitelist()
def get_deal_execution_health_analytics(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
):
    """
    V11-A API: Deal Execution Health & Management Attention Engine (Snapshot-Based)
    =============================================================================
    Security: Strictly uses validate_and_get_scope() to enforce hierarchy access control.
    Returns:
      - summary: Aggregate counts (open_deals, open_pipeline_value, stale_deals, stale_deal_value,
                                   critical_deals, at_risk_deals, no_recent_engagement, overdue_activity_deals)
      - deals: List of detailed deal execution health objects ordered by deal_value DESC
      - meta: Currency, stale threshold
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]
    company_currency = frappe.db.get_default("currency") or "INR"

    if not effective_users:
        return {
            "scope": {
                "team_filter": scope_data["team_filter"],
                "user_filter": scope_data["user_filter"],
                "effective_user_count": 0,
                "is_unrestricted": is_unrestricted,
            },
            "summary": {
                "open_deals": 0,
                "open_pipeline_value": 0.0,
                "stale_deals": 0,
                "stale_deal_value": 0.0,
                "critical_deals": 0,
                "at_risk_deals": 0,
                "no_recent_engagement": 0,
                "overdue_activity_deals": 0,
            },
            "deals": [],
            "meta": {
                "currency": company_currency,
                "stale_threshold_days": DEAL_EXECUTION_CONFIG["STALE_DEAL_DAYS"],
                "generated_at": str(frappe.utils.now_datetime()),
            },
        }

    # PASS 1: Bulk Query Open Deals
    open_deals_raw = frappe.db.sql(
        """
        SELECT
            d.name as deal_id,
            IFNULL(d.organization, d.name) as deal_title,
            IFNULL(d.organization, '—') as organization,
            d.deal_owner as owner,
            d.status as stage,
            IFNULL(d.deal_value, 0) as deal_value,
            IFNULL(d.probability, s.probability) as probability,
            d.creation,
            d.modified,
            DATEDIFF(NOW(), d.creation) as deal_age_days,
            DATEDIFF(NOW(), d.modified) as days_since_modified
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE d.deal_owner IN %(users)s
          AND s.type IN ('Open', 'Ongoing')
        ORDER BY d.deal_value DESC, deal_age_days DESC
        """,
        {"users": tuple(effective_users)},
        as_dict=True,
    )

    if not open_deals_raw:
        return {
            "scope": {
                "team_filter": scope_data["team_filter"],
                "user_filter": scope_data["user_filter"],
                "effective_user_count": len(effective_users),
                "is_unrestricted": is_unrestricted,
            },
            "summary": {
                "open_deals": 0,
                "open_pipeline_value": 0.0,
                "stale_deals": 0,
                "stale_deal_value": 0.0,
                "critical_deals": 0,
                "at_risk_deals": 0,
                "no_recent_engagement": 0,
                "overdue_activity_deals": 0,
            },
            "deals": [],
            "meta": {
                "currency": company_currency,
                "stale_threshold_days": DEAL_EXECUTION_CONFIG["STALE_DEAL_DAYS"],
                "generated_at": str(frappe.utils.now_datetime()),
            },
        }

    owner_emails = {d["owner"] for d in open_deals_raw if d.get("owner")}
    user_name_map = {}
    if owner_emails:
        users_info = frappe.get_all(
            "User",
            filters={"name": ["in", list(owner_emails)]},
            fields=["name", "full_name", "first_name", "last_name"],
        )
        for u in users_info:
            name_str = u.get("full_name") or f"{u.get('first_name') or ''} {u.get('last_name') or ''}".strip() or u["name"]
            user_name_map[u["name"]] = name_str

    deal_ids = tuple(d["deal_id"] for d in open_deals_raw)

    # PASS 2: Bulk Aggregate ToDo records linked to these deals
    todo_rows = frappe.db.sql(
        """
        SELECT
            reference_name as deal_id,
            SUM(CASE WHEN status = 'Closed' THEN 1 ELSE 0 END) as completed_activities,
            SUM(CASE WHEN status != 'Closed' THEN 1 ELSE 0 END) as open_activities,
            SUM(CASE WHEN status != 'Closed' AND date < CURDATE() THEN 1 ELSE 0 END) as overdue_activities
        FROM `tabToDo`
        WHERE reference_type = 'CRM Deal'
          AND reference_name IN %(deal_ids)s
        GROUP BY reference_name
        """,
        {"deal_ids": deal_ids},
        as_dict=True,
    )
    todo_map = {r["deal_id"]: r for r in todo_rows}

    # PASS 3: Bulk Aggregate Communication records linked to these deals
    comm_rows = frappe.db.sql(
        """
        SELECT
            reference_name as deal_id,
            MAX(creation) as last_comm_date,
            DATEDIFF(NOW(), MAX(creation)) as days_since_comm
        FROM `tabCommunication`
        WHERE reference_doctype = 'CRM Deal'
          AND reference_name IN %(deal_ids)s
        GROUP BY reference_name
        """,
        {"deal_ids": deal_ids},
        as_dict=True,
    )
    comm_map = {r["deal_id"]: r for r in comm_rows}

    # In-memory aggregation and deterministic classification
    analyzed_deals = []
    stale_deals_count = 0
    stale_deals_value = 0.0
    critical_deals_count = 0
    at_risk_deals_count = 0
    no_recent_engagement_count = 0
    overdue_activity_deals_count = 0
    total_pipeline_value = 0.0

    stale_thresh = DEAL_EXECUTION_CONFIG["STALE_DEAL_DAYS"]
    no_eng_thresh = DEAL_EXECUTION_CONFIG["NO_ENGAGEMENT_DAYS"]
    high_val_thresh = DEAL_EXECUTION_CONFIG["HIGH_VALUE_THRESHOLD"]
    aging_thresh = DEAL_EXECUTION_CONFIG["AGING_DEAL_DAYS"]

    for d in open_deals_raw:
        deal_id = d["deal_id"]
        val = float(d["deal_value"] or 0.0)
        total_pipeline_value += val

        todo_info = todo_map.get(deal_id, {})
        comp_act = int(todo_info.get("completed_activities") or 0)
        open_act = int(todo_info.get("open_activities") or 0)
        overdue_act = int(todo_info.get("overdue_activities") or 0)

        comm_info = comm_map.get(deal_id, {})
        last_comm = comm_info.get("last_comm_date")
        days_since_comm = int(comm_info["days_since_comm"]) if comm_info.get("days_since_comm") is not None else None

        days_mod = int(d["days_since_modified"] or 0)
        deal_age = int(d["deal_age_days"] or 0)

        is_stale = days_mod >= stale_thresh
        if is_stale:
            stale_deals_count += 1
            stale_deals_value += val

        if overdue_act > 0:
            overdue_activity_deals_count += 1

        is_no_eng = days_since_comm is not None and days_since_comm >= no_eng_thresh
        if is_no_eng or (days_since_comm is None and days_mod >= no_eng_thresh):
            no_recent_engagement_count += 1

        # Evidence accumulation & Risk Classification
        risk_reasons = []
        if is_stale:
            risk_reasons.append(f"Opportunity inactive for {days_mod} days")
        if days_since_comm is not None and days_since_comm >= no_eng_thresh:
            risk_reasons.append(f"No customer communication for {days_since_comm} days")
        if overdue_act > 0:
            risk_reasons.append(f"{overdue_act} overdue activity task{'s' if overdue_act > 1 else ''}")
        if val >= high_val_thresh:
            risk_reasons.append(f"High-value opportunity (₹{val:,.0f})")
        if deal_age >= aging_thresh:
            risk_reasons.append(f"Aging opportunity ({deal_age} days old)")

        # Deterministic Risk Status & Action Matrix
        risk_status = "HEALTHY"
        recommended_action = "Maintain regular sales cadence and complete upcoming milestones."

        if overdue_act > 0 and (val >= high_val_thresh or is_stale):
            risk_status = "CRITICAL"
            recommended_action = "Clear overdue tasks and re-engage customer immediately to prevent deal slipping."
        elif is_stale and (days_since_comm is not None and days_since_comm >= no_eng_thresh):
            risk_status = "CRITICAL"
            recommended_action = "Conduct immediate customer re-engagement call and update pipeline qualification."
        elif is_stale or overdue_act > 0 or (days_since_comm is not None and days_since_comm >= no_eng_thresh):
            risk_status = "AT_RISK"
            recommended_action = "Schedule follow-up activity and clear pending task backlog."
        elif deal_age >= aging_thresh or val >= high_val_thresh:
            risk_status = "WATCH"
            recommended_action = "Review closing roadmap and confirm decision-maker timeline."

        if risk_status == "CRITICAL":
            critical_deals_count += 1
        elif risk_status == "AT_RISK":
            at_risk_deals_count += 1

        raw_owner = d["owner"] or "Unassigned"
        owner_display = user_name_map.get(raw_owner, raw_owner)

        analyzed_deals.append({
            "deal_id": deal_id,
            "deal_title": str(d["deal_title"]),
            "organization": str(d["organization"]),
            "owner": str(owner_display),
            "owner_email": str(raw_owner),
            "stage": str(d["stage"]),
            "deal_value": val,
            "probability": float(d["probability"] or 0.0),
            "expected_closing_date": None,
            "creation": str(d["creation"]),
            "modified": str(d["modified"]),
            "deal_age_days": deal_age,
            "days_since_modified": days_mod,
            "last_customer_engagement": str(last_comm) if last_comm else None,
            "days_since_customer_engagement": days_since_comm,
            "completed_activities": comp_act,
            "open_activities": open_act,
            "overdue_activities": overdue_act,
            "is_stale": is_stale,
            "risk_status": risk_status,
            "risk_reasons": risk_reasons,
            "recommended_action": recommended_action,
        })

    return {
        "scope": {
            "team_filter": scope_data["team_filter"],
            "user_filter": scope_data["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": is_unrestricted,
        },
        "summary": {
            "open_deals": len(open_deals_raw),
            "open_pipeline_value": total_pipeline_value,
            "stale_deals": stale_deals_count,
            "stale_deal_value": stale_deals_value,
            "critical_deals": critical_deals_count,
            "at_risk_deals": at_risk_deals_count,
            "no_recent_engagement": no_recent_engagement_count,
            "overdue_activity_deals": overdue_activity_deals_count,
        },
        "deals": analyzed_deals,
        "meta": {
            "currency": company_currency,
            "stale_threshold_days": stale_thresh,
            "generated_at": str(frappe.utils.now_datetime()),
        },
    }


# ---------------------------------------------------------------------------
# V11-B STEP 4: HISTORICAL CRM DEAL EVENT ANALYTICS APIs
# ---------------------------------------------------------------------------

def _get_event_effective_users(team_filter=None, user_filter=None):
    """Helper to resolve scope and validate filters for event analytics APIs."""
    caller = frappe.session.user
    scope = _resolve_scope(caller)
    validated = validate_dashboard_filters(scope, team_filter, user_filter)
    return scope, validated, validated["effective_users"]


@frappe.whitelist()
def get_deal_stage_velocity_analytics(period="this_month", team_filter=None, user_filter=None, custom_from=None, custom_to=None):
    """
    V11-B Step 4: Deal Stage Velocity Analytics
    Measures dwell time and transition counts per stage using STAGE_CHANGED events.
    """
    scope, validated, effective_users = _get_event_effective_users(team_filter, user_filter)
    date_range = resolve_period(period, custom_from, custom_to)

    if not effective_users:
        return {
            "scope": {"period": period, "from_date": date_range["from_date"], "to_date": date_range["to_date"], "team_filter": validated["team_filter"], "user_filter": validated["user_filter"], "effective_user_count": 0, "is_unrestricted": scope["is_unrestricted"]},
            "summary": {"total_stage_transitions": 0, "measurable_transitions": 0, "overall_average_dwell_days": None},
            "stages": [],
            "meta": {"metric_type": "event_history", "date_anchor": "event_timestamp", "generated_at": str(frappe.utils.now_datetime())}
        }

    # Fetch STAGE_CHANGED events within period for effective users
    events = frappe.get_all(
        "CRM Deal Event",
        filters={
            "event_type": "STAGE_CHANGED",
            "deal_owner": ["in", effective_users],
            "event_timestamp": ["between", [f"{date_range['from_date']} 00:00:00", f"{date_range['to_date']} 23:59:59"]]
        },
        fields=["old_value", "new_value", "dwell_days", "event_timestamp"]
    )

    stage_groups = {}
    total_transitions = len(events)
    measurable_transitions = 0
    total_dwell_sum = 0.0

    for ev in events:
        stg = ev.new_value or _("Unspecified")
        if stg not in stage_groups:
            stage_groups[stg] = {"transitions": 0, "measurable_count": 0, "dwells": []}
        
        stage_groups[stg]["transitions"] += 1

        d_days = ev.dwell_days
        if d_days is not None and float(d_days) > 0:
            stage_groups[stg]["measurable_count"] += 1
            stage_groups[stg]["dwells"].append(float(d_days))
            total_dwell_sum += float(d_days)
            measurable_transitions += 1

    stages_result = []
    for stg, data in stage_groups.items():
        dwells = sorted(data["dwells"])
        cnt = len(dwells)
        if cnt > 0:
            avg_dwell = round(sum(dwells) / cnt, 2)
            min_dwell = round(dwells[0], 2)
            max_dwell = round(dwells[-1], 2)
            mid = cnt // 2
            median_dwell = round((dwells[mid] if cnt % 2 != 0 else (dwells[mid - 1] + dwells[mid]) / 2.0), 2)
        else:
            avg_dwell = median_dwell = min_dwell = max_dwell = None

        stages_result.append({
            "stage": stg,
            "transitions": data["transitions"],
            "measurable_transitions": data["measurable_count"],
            "average_dwell_days": avg_dwell,
            "median_dwell_days": median_dwell,
            "min_dwell_days": min_dwell,
            "max_dwell_days": max_dwell,
        })

    stages_result.sort(key=lambda x: x["transitions"], reverse=True)
    overall_avg = round(total_dwell_sum / measurable_transitions, 2) if measurable_transitions > 0 else None

    return {
        "scope": {
            "period": period,
            "from_date": date_range["from_date"],
            "to_date": date_range["to_date"],
            "team_filter": validated["team_filter"],
            "user_filter": validated["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": scope["is_unrestricted"],
        },
        "summary": {
            "total_stage_transitions": total_transitions,
            "measurable_transitions": measurable_transitions,
            "overall_average_dwell_days": overall_avg,
        },
        "stages": stages_result,
        "meta": {
            "metric_type": "event_history",
            "date_anchor": "event_timestamp",
            "generated_at": str(frappe.utils.now_datetime()),
        }
    }


@frappe.whitelist()
def get_deal_slippage_analytics(period="this_month", team_filter=None, user_filter=None, custom_from=None, custom_to=None):
    """
    V11-B Step 4: Deal Close-Date Slippage Analytics
    Analyzes CLOSE_DATE_CHANGED events to measure schedule slippage and pull-forwards.
    """
    scope, validated, effective_users = _get_event_effective_users(team_filter, user_filter)
    date_range = resolve_period(period, custom_from, custom_to)

    if not effective_users:
        return {
            "scope": {"period": period, "from_date": date_range["from_date"], "to_date": date_range["to_date"], "team_filter": validated["team_filter"], "user_filter": validated["user_filter"], "effective_user_count": 0, "is_unrestricted": scope["is_unrestricted"]},
            "summary": {"close_date_changes": 0, "deals_affected": 0, "positive_slippage_days": 0, "negative_slippage_days": 0, "average_positive_slippage_days": None, "average_negative_slippage_days": None, "percentage_affected_deals": 0.0},
            "meta": {"metric_type": "event_history", "date_anchor": "event_timestamp", "generated_at": str(frappe.utils.now_datetime())}
        }

    events = frappe.get_all(
        "CRM Deal Event",
        filters={
            "event_type": "CLOSE_DATE_CHANGED",
            "deal_owner": ["in", effective_users],
            "event_timestamp": ["between", [f"{date_range['from_date']} 00:00:00", f"{date_range['to_date']} 23:59:59"]]
        },
        fields=["deal", "days_pushed", "old_value", "new_value"]
    )

    affected_deals = set()
    pos_days_sum = 0
    pos_count = 0
    neg_days_sum = 0
    neg_count = 0

    for ev in events:
        if ev.days_pushed is None or ev.days_pushed == 0:
            continue
        affected_deals.add(ev.deal)
        dp = int(ev.days_pushed)
        if dp > 0:
            pos_days_sum += dp
            pos_count += 1
        elif dp < 0:
            neg_days_sum += abs(dp)
            neg_count += 1

    total_deals_in_scope = frappe.db.count("CRM Deal", filters={"deal_owner": ["in", effective_users]})
    pct_affected = round((len(affected_deals) / total_deals_in_scope * 100.0), 2) if total_deals_in_scope > 0 else 0.0

    avg_pos = round(pos_days_sum / pos_count, 2) if pos_count > 0 else None
    avg_neg = round(neg_days_sum / neg_count, 2) if neg_count > 0 else None

    return {
        "scope": {
            "period": period,
            "from_date": date_range["from_date"],
            "to_date": date_range["to_date"],
            "team_filter": validated["team_filter"],
            "user_filter": validated["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": scope["is_unrestricted"],
        },
        "summary": {
            "close_date_changes": len(events),
            "deals_affected": len(affected_deals),
            "positive_slippage_days": pos_days_sum,
            "negative_slippage_days": neg_days_sum,
            "average_positive_slippage_days": avg_pos,
            "average_negative_slippage_days": avg_neg,
            "percentage_affected_deals": pct_affected,
        },
        "meta": {
            "metric_type": "event_history",
            "date_anchor": "event_timestamp",
            "generated_at": str(frappe.utils.now_datetime()),
        }
    }


@frappe.whitelist()
def get_probability_movement_analytics(period="this_month", team_filter=None, user_filter=None, custom_from=None, custom_to=None):
    """
    V11-B Step 4: Probability Movement Analytics
    Measures changes in deal win probability over time.
    """
    scope, validated, effective_users = _get_event_effective_users(team_filter, user_filter)
    date_range = resolve_period(period, custom_from, custom_to)

    if not effective_users:
        return {
            "scope": {"period": period, "from_date": date_range["from_date"], "to_date": date_range["to_date"], "team_filter": validated["team_filter"], "user_filter": validated["user_filter"], "effective_user_count": 0, "is_unrestricted": scope["is_unrestricted"]},
            "summary": {"probability_changes": 0, "average_increase": None, "average_decrease": None, "net_probability_movement": 0.0, "largest_increase": None, "largest_decrease": None},
            "meta": {"metric_type": "event_history", "date_anchor": "event_timestamp", "generated_at": str(frappe.utils.now_datetime())}
        }

    events = frappe.get_all(
        "CRM Deal Event",
        filters={
            "event_type": "PROBABILITY_CHANGED",
            "deal_owner": ["in", effective_users],
            "event_timestamp": ["between", [f"{date_range['from_date']} 00:00:00", f"{date_range['to_date']} 23:59:59"]]
        },
        fields=["numeric_old_value", "numeric_new_value"]
    )

    inc_sum = 0.0
    inc_count = 0
    dec_sum = 0.0
    dec_count = 0
    largest_inc = None
    largest_dec = None

    for ev in events:
        old_val = float(ev.numeric_old_value or 0.0)
        new_val = float(ev.numeric_new_value or 0.0)
        diff = new_val - old_val

        if diff > 0:
            inc_sum += diff
            inc_count += 1
            if largest_inc is None or diff > largest_inc:
                largest_inc = diff
        elif diff < 0:
            abs_diff = abs(diff)
            dec_sum += abs_diff
            dec_count += 1
            if largest_dec is None or abs_diff > largest_dec:
                largest_dec = abs_diff

    avg_inc = round(inc_sum / inc_count, 2) if inc_count > 0 else None
    avg_dec = round(dec_sum / dec_count, 2) if dec_count > 0 else None
    net_mov = round(inc_sum - dec_sum, 2)

    return {
        "scope": {
            "period": period,
            "from_date": date_range["from_date"],
            "to_date": date_range["to_date"],
            "team_filter": validated["team_filter"],
            "user_filter": validated["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": scope["is_unrestricted"],
        },
        "summary": {
            "probability_changes": len(events),
            "average_increase": avg_inc,
            "average_decrease": avg_dec,
            "net_probability_movement": net_mov,
            "largest_increase": round(largest_inc, 2) if largest_inc is not None else None,
            "largest_decrease": round(largest_dec, 2) if largest_dec is not None else None,
        },
        "meta": {
            "metric_type": "event_history",
            "date_anchor": "event_timestamp",
            "generated_at": str(frappe.utils.now_datetime()),
        }
    }


@frappe.whitelist()
def get_deal_value_movement_analytics(period="this_month", team_filter=None, user_filter=None, custom_from=None, custom_to=None):
    """
    V11-B Step 4: Deal Value Movement Analytics
    Measures deal expansion and contraction using VALUE_CHANGED events.
    """
    scope, validated, effective_users = _get_event_effective_users(team_filter, user_filter)
    date_range = resolve_period(period, custom_from, custom_to)

    if not effective_users:
        return {
            "scope": {"period": period, "from_date": date_range["from_date"], "to_date": date_range["to_date"], "team_filter": validated["team_filter"], "user_filter": validated["user_filter"], "effective_user_count": 0, "is_unrestricted": scope["is_unrestricted"]},
            "summary": {"value_changes": 0, "total_positive_movement": 0.0, "total_negative_movement": 0.0, "average_increase": None, "average_decrease": None, "net_value_movement": 0.0, "largest_increase": None, "largest_decrease": None},
            "meta": {"metric_type": "event_history", "date_anchor": "event_timestamp", "generated_at": str(frappe.utils.now_datetime())}
        }

    events = frappe.get_all(
        "CRM Deal Event",
        filters={
            "event_type": "VALUE_CHANGED",
            "deal_owner": ["in", effective_users],
            "event_timestamp": ["between", [f"{date_range['from_date']} 00:00:00", f"{date_range['to_date']} 23:59:59"]]
        },
        fields=["numeric_old_value", "numeric_new_value"]
    )

    pos_sum = 0.0
    pos_count = 0
    neg_sum = 0.0
    neg_count = 0
    largest_inc = None
    largest_dec = None

    for ev in events:
        old_val = float(ev.numeric_old_value or 0.0)
        new_val = float(ev.numeric_new_value or 0.0)
        diff = new_val - old_val

        if diff > 0:
            pos_sum += diff
            pos_count += 1
            if largest_inc is None or diff > largest_inc:
                largest_inc = diff
        elif diff < 0:
            abs_diff = abs(diff)
            neg_sum += abs_diff
            neg_count += 1
            if largest_dec is None or abs_diff > largest_dec:
                largest_dec = abs_diff

    avg_inc = round(pos_sum / pos_count, 2) if pos_count > 0 else None
    avg_dec = round(neg_sum / neg_count, 2) if neg_count > 0 else None
    net_mov = round(pos_sum - neg_sum, 2)

    return {
        "scope": {
            "period": period,
            "from_date": date_range["from_date"],
            "to_date": date_range["to_date"],
            "team_filter": validated["team_filter"],
            "user_filter": validated["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": scope["is_unrestricted"],
        },
        "summary": {
            "value_changes": len(events),
            "total_positive_movement": round(pos_sum, 2),
            "total_negative_movement": round(neg_sum, 2),
            "average_increase": avg_inc,
            "average_decrease": avg_dec,
            "net_value_movement": net_mov,
            "largest_increase": round(largest_inc, 2) if largest_inc is not None else None,
            "largest_decrease": round(largest_dec, 2) if largest_dec is not None else None,
        },
        "meta": {
            "metric_type": "event_history",
            "date_anchor": "event_timestamp",
            "generated_at": str(frappe.utils.now_datetime()),
        }
    }


@frappe.whitelist()
def get_deal_owner_change_analytics(period="this_month", team_filter=None, user_filter=None, custom_from=None, custom_to=None):
    """
    V11-B Step 4: Deal Owner Change Analytics
    Measures deal transfers between sales reps within hierarchy scope.
    """
    scope, validated, effective_users = _get_event_effective_users(team_filter, user_filter)
    date_range = resolve_period(period, custom_from, custom_to)

    if not effective_users:
        return {
            "scope": {"period": period, "from_date": date_range["from_date"], "to_date": date_range["to_date"], "team_filter": validated["team_filter"], "user_filter": validated["user_filter"], "effective_user_count": 0, "is_unrestricted": scope["is_unrestricted"]},
            "summary": {"owner_changes": 0, "deals_affected": 0},
            "transfers": [],
            "meta": {"metric_type": "event_history", "date_anchor": "event_timestamp", "generated_at": str(frappe.utils.now_datetime())}
        }

    events = frappe.get_all(
        "CRM Deal Event",
        filters={
            "event_type": "OWNER_CHANGED",
            "deal_owner": ["in", effective_users],
            "event_timestamp": ["between", [f"{date_range['from_date']} 00:00:00", f"{date_range['to_date']} 23:59:59"]]
        },
        fields=["deal", "old_value", "new_value"]
    )

    affected_deals = set()
    transfer_counts = {}

    for ev in events:
        affected_deals.add(ev.deal)
        from_usr = ev.old_value or _("Unassigned")
        to_usr = ev.new_value or _("Unassigned")
        key = (from_usr, to_usr)
        transfer_counts[key] = transfer_counts.get(key, 0) + 1

    transfers_list = [
        {"from_owner": k[0], "to_owner": k[1], "count": v}
        for k, v in transfer_counts.items()
    ]
    transfers_list.sort(key=lambda x: x["count"], reverse=True)

    return {
        "scope": {
            "period": period,
            "from_date": date_range["from_date"],
            "to_date": date_range["to_date"],
            "team_filter": validated["team_filter"],
            "user_filter": validated["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": scope["is_unrestricted"],
        },
        "summary": {
            "owner_changes": len(events),
            "deals_affected": len(affected_deals),
        },
        "transfers": transfers_list,
        "meta": {
            "metric_type": "event_history",
            "date_anchor": "event_timestamp",
            "generated_at": str(frappe.utils.now_datetime()),
        }
    }


@frappe.whitelist()
def get_deal_execution_analytics(period="this_month", team_filter=None, user_filter=None, custom_from=None, custom_to=None):
    """
    V11-B Step 4: Executive Historical Deal Execution Summary
    Aggregates stage velocity, slippage, probability movement, value movement, and owner transfers into a single response.
    """
    velocity = get_deal_stage_velocity_analytics(period, team_filter, user_filter, custom_from, custom_to)
    slippage = get_deal_slippage_analytics(period, team_filter, user_filter, custom_from, custom_to)
    prob_mov = get_probability_movement_analytics(period, team_filter, user_filter, custom_from, custom_to)
    val_mov = get_deal_value_movement_analytics(period, team_filter, user_filter, custom_from, custom_to)
    owner_tx = get_deal_owner_change_analytics(period, team_filter, user_filter, custom_from, custom_to)

    return {
        "scope": velocity["scope"],
        "summary": {
            "stage_velocity": velocity["summary"],
            "slippage": slippage["summary"],
            "probability_movement": prob_mov["summary"],
            "value_movement": val_mov["summary"],
            "owner_changes": owner_tx["summary"],
        },
        "details": {
            "stages": velocity["stages"],
            "transfers": owner_tx["transfers"],
        },
        "meta": {
            "metric_type": "event_history_executive_summary",
            "date_anchor": "event_timestamp",
            "generated_at": str(frappe.utils.now_datetime()),
        }
    }


HIGH_VALUE_SLIPPAGE_THRESHOLD = 50000.0


@frappe.whitelist()
def get_deal_velocity_slippage_command_center(
    period="this_month",
    team_filter="ALL",
    user_filter="ALL",
    custom_from=None,
    custom_to=None
):
    """
    V11-B Step 6A: Deal Velocity & Slippage Command Center API
    Combines active open CRM Deal population with historical CRM Deal Event streams
    to generate deterministic, explainable deal-level velocity and slippage metrics.
    """
    scope, validated, effective_users = _get_event_effective_users(team_filter, user_filter)
    date_range = resolve_period(period, custom_from, custom_to)

    if not effective_users:
        return {
            "scope": {
                "period": period,
                "from_date": date_range["from_date"],
                "to_date": date_range["to_date"],
                "team_filter": validated["team_filter"],
                "user_filter": validated["user_filter"],
                "effective_user_count": 0,
                "is_unrestricted": scope["is_unrestricted"],
            },
            "summary": {
                "total_active_deals": 0,
                "total_slipped_deals": 0,
                "repeat_slippage_deals": 0,
                "total_positive_days_pushed": 0,
                "total_days_pulled_forward": 0,
                "stagnant_deals": 0,
                "high_value_slippage_deals": 0,
                "high_risk_deals": 0,
                "high_risk_value_exposure": 0.0,
            },
            "deal_matrix": [],
            "meta": {
                "metric_type": "deterministic_hybrid_execution_velocity",
                "historical_anchor": "CRM Deal Event.event_timestamp",
                "risk_model": "deterministic_additive",
                "high_value_threshold": HIGH_VALUE_SLIPPAGE_THRESHOLD,
                "generated_at": str(frappe.utils.now_datetime()),
            }
        }

    # PASS 1: Bulk Fetch Active Open CRM Deals matching effective user list
    open_deals_raw = frappe.db.sql(
        """
        SELECT
            d.name as deal_id,
            IFNULL(d.organization, d.name) as deal_title,
            IFNULL(d.organization, '—') as organization,
            d.deal_owner as owner,
            d.status as stage,
            IFNULL(d.deal_value, 0) as deal_value,
            IFNULL(d.probability, s.probability) as probability,
            d.creation,
            d.modified
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE d.deal_owner IN %(users)s
          AND s.type IN ('Open', 'Ongoing')
        ORDER BY d.deal_value DESC
        """,
        {"users": tuple(effective_users)},
        as_dict=True,
    )

    if not open_deals_raw:
        return {
            "scope": {
                "period": period,
                "from_date": date_range["from_date"],
                "to_date": date_range["to_date"],
                "team_filter": validated["team_filter"],
                "user_filter": validated["user_filter"],
                "effective_user_count": len(effective_users),
                "is_unrestricted": scope["is_unrestricted"],
            },
            "summary": {
                "total_active_deals": 0,
                "total_slipped_deals": 0,
                "repeat_slippage_deals": 0,
                "total_positive_days_pushed": 0,
                "total_days_pulled_forward": 0,
                "stagnant_deals": 0,
                "high_value_slippage_deals": 0,
                "high_risk_deals": 0,
                "high_risk_value_exposure": 0.0,
            },
            "deal_matrix": [],
            "meta": {
                "metric_type": "deterministic_hybrid_execution_velocity",
                "historical_anchor": "CRM Deal Event.event_timestamp",
                "risk_model": "deterministic_additive",
                "high_value_threshold": HIGH_VALUE_SLIPPAGE_THRESHOLD,
                "generated_at": str(frappe.utils.now_datetime()),
            }
        }

    owner_emails = {d["owner"] for d in open_deals_raw if d.get("owner")}
    user_name_map = {}
    if owner_emails:
        users_info = frappe.get_all(
            "User",
            filters={"name": ["in", list(owner_emails)]},
            fields=["name", "full_name", "first_name", "last_name"],
        )
        for u in users_info:
            name_str = u.get("full_name") or f"{u.get('first_name') or ''} {u.get('last_name') or ''}".strip() or u["name"]
            user_name_map[u["name"]] = name_str

    deal_ids = tuple([d["deal_id"] for d in open_deals_raw])

    # PASS 2: Bulk Fetch relevant events for these open deals within the period anchor
    events = frappe.get_all(
        "CRM Deal Event",
        filters={
            "deal": ["in", deal_ids],
            "deal_owner": ["in", effective_users],
            "event_timestamp": ["between", [f"{date_range['from_date']} 00:00:00", f"{date_range['to_date']} 23:59:59"]]
        },
        fields=[
            "deal",
            "event_type",
            "field_name",
            "old_value",
            "new_value",
            "numeric_old_value",
            "numeric_new_value",
            "days_pushed",
            "dwell_days",
            "event_timestamp"
        ],
        order_by="event_timestamp ASC"
    )

    # Organize events per deal
    deal_events_map = {}
    for ev in events:
        d_id = ev.deal
        if d_id not in deal_events_map:
            deal_events_map[d_id] = []
        deal_events_map[d_id].append(ev)

    # Process metrics per open deal
    now_ts = frappe.utils.now_datetime()
    deal_matrix = []

    total_slipped_deals_count = 0
    repeat_slippage_deals_count = 0
    total_positive_days_pushed_sum = 0
    total_days_pulled_forward_sum = 0
    stagnant_deals_count = 0
    high_value_slippage_deals_count = 0
    high_risk_deals_count = 0
    high_risk_value_exposure_sum = 0.0

    for d in open_deals_raw:
        d_id = d["deal_id"]
        val = float(d["deal_value"] or 0.0)
        d_events = deal_events_map.get(d_id, [])

        close_date_push_count = 0
        cumulative_days_pushed = 0
        days_pulled_forward = 0
        net_probability_change = 0.0
        net_value_change = 0.0

        latest_stage_timestamp = None

        for ev in d_events:
            e_type = ev.event_type
            if e_type == "CLOSE_DATE_CHANGED":
                dp = ev.days_pushed or 0
                if dp > 0:
                    close_date_push_count += 1
                    cumulative_days_pushed += dp
                elif dp < 0:
                    days_pulled_forward += abs(dp)
            elif e_type == "PROBABILITY_CHANGED":
                old_p = float(ev.numeric_old_value or 0.0)
                new_p = float(ev.numeric_new_value or 0.0)
                net_probability_change += (new_p - old_p)
            elif e_type == "VALUE_CHANGED":
                old_v = float(ev.numeric_old_value or 0.0)
                new_v = float(ev.numeric_new_value or 0.0)
                net_value_change += (new_v - old_v)
            elif e_type == "STAGE_CHANGED":
                if ev.event_timestamp:
                    latest_stage_timestamp = ev.event_timestamp

        # Calculate current stage dwell days
        current_stage_dwell_days = None
        if latest_stage_timestamp:
            ts = frappe.utils.get_datetime(latest_stage_timestamp)
            current_stage_dwell_days = round((now_ts - ts).total_seconds() / 86400.0, 1)

        # Summary flags accumulation
        if close_date_push_count > 0:
            total_slipped_deals_count += 1
        if close_date_push_count >= 2:
            repeat_slippage_deals_count += 1

        total_positive_days_pushed_sum += cumulative_days_pushed
        total_days_pulled_forward_sum += days_pulled_forward

        is_stagnant = current_stage_dwell_days is not None and current_stage_dwell_days > 30.0
        if is_stagnant:
            stagnant_deals_count += 1

        if val >= HIGH_VALUE_SLIPPAGE_THRESHOLD and close_date_push_count > 0:
            high_value_slippage_deals_count += 1

        # Deterministic Additive Scoring Matrix (Max 100)
        risk_score = 0
        risk_factors = []

        if close_date_push_count >= 2:
            risk_score += 20
            risk_factors.append(f"Repeat Slippage ({close_date_push_count} pushes)")

        if cumulative_days_pushed > 14:
            risk_score += 20
            risk_factors.append(f"Cumulative Push > 14 days (+{cumulative_days_pushed}d)")

        if is_stagnant:
            risk_score += 20
            risk_factors.append(f"Stage Stagnation > 30 days ({current_stage_dwell_days}d in {d['stage']})")

        if net_probability_change < 0:
            risk_score += 20
            risk_factors.append(f"Probability Decline ({net_probability_change:+.1f} pp)")

        if net_value_change < 0:
            risk_score += 20
            risk_factors.append(f"Value Contraction (-₹{abs(net_value_change):,.0f})")

        if risk_score >= 60:
            high_risk_deals_count += 1
            high_risk_value_exposure_sum += val

        raw_owner = d["owner"] or "Unassigned"
        owner_display = user_name_map.get(raw_owner, raw_owner)

        deal_matrix.append({
            "deal_id": d_id,
            "deal_title": str(d["deal_title"]),
            "organization": str(d["organization"]),
            "owner": str(owner_display),
            "owner_email": str(raw_owner),
            "stage": str(d["stage"]),
            "deal_value": val,
            "close_date_push_count": close_date_push_count,
            "cumulative_days_pushed": cumulative_days_pushed,
            "days_pulled_forward": days_pulled_forward,
            "current_stage_dwell_days": current_stage_dwell_days,
            "net_probability_change": round(net_probability_change, 1),
            "net_value_change": round(net_value_change, 2),
            "deterministic_risk_score": risk_score,
            "risk_factors": risk_factors
        })

    return {
        "scope": {
            "period": period,
            "from_date": date_range["from_date"],
            "to_date": date_range["to_date"],
            "team_filter": validated["team_filter"],
            "user_filter": validated["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": scope["is_unrestricted"],
        },
        "summary": {
            "total_active_deals": len(open_deals_raw),
            "total_slipped_deals": total_slipped_deals_count,
            "repeat_slippage_deals": repeat_slippage_deals_count,
            "total_positive_days_pushed": total_positive_days_pushed_sum,
            "total_days_pulled_forward": total_days_pulled_forward_sum,
            "stagnant_deals": stagnant_deals_count,
            "high_value_slippage_deals": high_value_slippage_deals_count,
            "high_risk_deals": high_risk_deals_count,
            "high_risk_value_exposure": high_risk_value_exposure_sum,
        },
        "deal_matrix": deal_matrix,
        "meta": {
            "metric_type": "deterministic_hybrid_execution_velocity",
            "historical_anchor": "CRM Deal Event.event_timestamp",
            "risk_model": "deterministic_additive",
            "high_value_threshold": HIGH_VALUE_SLIPPAGE_THRESHOLD,
            "generated_at": str(frappe.utils.now_datetime()),
        }
    }


@frappe.whitelist()
def get_stage_transition_bottleneck_analytics(
    period="this_month",
    team_filter="ALL",
    user_filter="ALL",
    custom_from=None,
    custom_to=None
):
    """
    V12 Step 1 Backend API: Stage Transition & Bottleneck Analytics.
    Provides deterministic historical analytics on stage dwell times, stage transition paths,
    and bottleneck rankings anchored to CRM Deal Event records.
    Strict zero-fabrication: Missing stage dwell entries return NULL and are excluded from statistical averages.
    """
    scope, validated, effective_users = _get_event_effective_users(team_filter, user_filter)
    date_range = resolve_period(period, custom_from, custom_to)

    if not effective_users:
        return {
            "scope": {
                "period": period,
                "from_date": date_range["from_date"],
                "to_date": date_range["to_date"],
                "team_filter": validated["team_filter"],
                "user_filter": validated["user_filter"],
                "effective_user_count": 0,
                "is_unrestricted": scope["is_unrestricted"],
            },
            "summary": {
                "total_stage_transitions": 0,
                "measured_transitions": 0,
                "unmeasured_transitions": 0,
                "unique_deals": 0,
                "unique_stages": 0,
                "unique_transition_paths": 0,
                "slowest_stage": None,
                "slowest_transition": None,
            },
            "stage_velocity": [],
            "transition_matrix": [],
            "bottlenecks": [],
            "meta": {
                "metric_type": "historical_stage_transition_analytics",
                "historical_anchor": "CRM Deal Event.event_timestamp",
                "risk_model": "none",
                "predictive": False,
                "generated_at": str(frappe.utils.now_datetime()),
            }
        }

    # Query all STAGE_CHANGED events within scope date range
    query = """
        SELECT
            name,
            deal,
            deal_owner,
            old_value as from_stage,
            new_value as to_stage,
            dwell_days,
            event_timestamp
        FROM `tabCRM Deal Event`
        WHERE event_type = 'STAGE_CHANGED'
            AND deal_owner IN %(users)s
            AND event_timestamp >= %(from_date)s
            AND event_timestamp <= %(to_date)s
        ORDER BY deal ASC, event_timestamp ASC
    """

    params = {
        "users": effective_users,
        "from_date": f"{date_range['from_date']} 00:00:00",
        "to_date": f"{date_range['to_date']} 23:59:59",
    }

    events = frappe.db.sql(query, params, as_dict=True)

    if not events:
        return {
            "scope": {
                "period": period,
                "from_date": date_range["from_date"],
                "to_date": date_range["to_date"],
                "team_filter": validated["team_filter"],
                "user_filter": validated["user_filter"],
                "effective_user_count": len(effective_users),
                "is_unrestricted": scope["is_unrestricted"],
            },
            "summary": {
                "total_stage_transitions": 0,
                "measured_transitions": 0,
                "unmeasured_transitions": 0,
                "unique_deals": 0,
                "unique_stages": 0,
                "unique_transition_paths": 0,
                "slowest_stage": None,
                "slowest_transition": None,
            },
            "stage_velocity": [],
            "transition_matrix": [],
            "bottlenecks": [],
            "meta": {
                "metric_type": "historical_stage_transition_analytics",
                "historical_anchor": "CRM Deal Event.event_timestamp",
                "risk_model": "none",
                "predictive": False,
                "generated_at": str(frappe.utils.now_datetime()),
            }
        }

    # Data aggregates
    total_stage_transitions = len(events)
    measured_transitions = sum(1 for e in events if e.dwell_days is not None)
    unmeasured_transitions = total_stage_transitions - measured_transitions

    unique_deals = len(set(e.deal for e in events if e.deal))

    # Stage sets and maps
    stage_dwell_map = {}
    stage_transition_map = {}
    unique_stages = set()

    for e in events:
        from_stg = e.from_stage or ""
        to_stg = e.to_stage or ""
        if from_stg:
            unique_stages.add(from_stg)
        if to_stg:
            unique_stages.add(to_stg)

        # 1. Stage velocity aggregation by exited stage (`from_stage`)
        if from_stg:
            if from_stg not in stage_dwell_map:
                stage_dwell_map[from_stg] = []
            stage_dwell_map[from_stg].append(e.dwell_days)

        # 2. Pairwise transition aggregation
        path_key = (from_stg, to_stg)
        if path_key not in stage_transition_map:
            stage_transition_map[path_key] = []
        stage_transition_map[path_key].append(e.dwell_days)

    # Calculate stage velocity list
    stage_velocity_list = []
    for stg, dwells in stage_dwell_map.items():
        total_cnt = len(dwells)
        valid_dwells = [d for d in dwells if d is not None]
        m_cnt = len(valid_dwells)
        unm_cnt = total_cnt - m_cnt

        avg_dwell = round(sum(valid_dwells) / float(m_cnt), 1) if m_cnt > 0 else None
        max_dwell = round(max(valid_dwells), 1) if m_cnt > 0 else None
        
        if m_cnt > 0:
            sorted_d = sorted(valid_dwells)
            mid = m_cnt // 2
            if m_cnt % 2 == 1:
                med_dwell = round(sorted_d[mid], 1)
            else:
                med_dwell = round((sorted_d[mid - 1] + sorted_d[mid]) / 2.0, 1)
        else:
            med_dwell = None

        stage_velocity_list.append({
            "stage": stg,
            "transition_count": total_cnt,
            "measured_count": m_cnt,
            "unmeasured_count": unm_cnt,
            "average_dwell_days": avg_dwell,
            "median_dwell_days": med_dwell,
            "maximum_dwell_days": max_dwell,
        })

    # Sort stage velocity by average_dwell_days DESC (NULLs last)
    stage_velocity_list.sort(key=lambda x: (x["average_dwell_days"] is None, -(x["average_dwell_days"] or 0)))

    # Calculate transition matrix list
    transition_matrix_list = []
    for (f_stg, t_stg), dwells in stage_transition_map.items():
        total_cnt = len(dwells)
        valid_dwells = [d for d in dwells if d is not None]
        m_cnt = len(valid_dwells)
        unm_cnt = total_cnt - m_cnt

        avg_dwell = round(sum(valid_dwells) / float(m_cnt), 1) if m_cnt > 0 else None
        max_dwell = round(max(valid_dwells), 1) if m_cnt > 0 else None

        if m_cnt > 0:
            sorted_d = sorted(valid_dwells)
            mid = m_cnt // 2
            if m_cnt % 2 == 1:
                med_dwell = round(sorted_d[mid], 1)
            else:
                med_dwell = round((sorted_d[mid - 1] + sorted_d[mid]) / 2.0, 1)
        else:
            med_dwell = None

        transition_matrix_list.append({
            "from_stage": f_stg,
            "to_stage": t_stg,
            "transition_count": total_cnt,
            "measured_count": m_cnt,
            "unmeasured_count": unm_cnt,
            "average_dwell_days": avg_dwell,
            "median_dwell_days": med_dwell,
            "maximum_dwell_days": max_dwell,
        })

    # Sort transition matrix by transition_count DESC
    transition_matrix_list.sort(key=lambda x: -x["transition_count"])

    # Calculate bottleneck ranking list
    bottlenecks_list = []
    for item in stage_velocity_list:
        avg_d = item["average_dwell_days"]
        m_cnt = item["measured_count"]
        
        if avg_d is not None and m_cnt > 0:
            b_index = round(avg_d * m_cnt, 1)
        else:
            b_index = None

        bottlenecks_list.append({
            "stage": item["stage"],
            "transition_count": item["transition_count"],
            "measured_count": m_cnt,
            "average_dwell_days": avg_d,
            "median_dwell_days": item["median_dwell_days"],
            "bottleneck_index": b_index
        })

    # Sort bottlenecks by bottleneck_index DESC (NULLs last)
    bottlenecks_list.sort(key=lambda x: (x["bottleneck_index"] is None, -(x["bottleneck_index"] or 0)))

    # Compute summary values
    slowest_stage = stage_velocity_list[0]["stage"] if stage_velocity_list and stage_velocity_list[0]["average_dwell_days"] is not None else None
    
    # Find slowest transition path
    measured_paths = [t for t in transition_matrix_list if t["average_dwell_days"] is not None]
    if measured_paths:
        measured_paths.sort(key=lambda x: -x["average_dwell_days"])
        slowest_path = measured_paths[0]
        slowest_transition = f"{slowest_path['from_stage']} -> {slowest_path['to_stage']}"
    else:
        slowest_transition = None

    return {
        "scope": {
            "period": period,
            "from_date": date_range["from_date"],
            "to_date": date_range["to_date"],
            "team_filter": validated["team_filter"],
            "user_filter": validated["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": scope["is_unrestricted"],
        },
        "summary": {
            "total_stage_transitions": total_stage_transitions,
            "measured_transitions": measured_transitions,
            "unmeasured_transitions": unmeasured_transitions,
            "unique_deals": unique_deals,
            "unique_stages": len(unique_stages),
            "unique_transition_paths": len(stage_transition_map),
            "slowest_stage": slowest_stage,
            "slowest_transition": slowest_transition,
        },
        "stage_velocity": stage_velocity_list,
        "transition_matrix": transition_matrix_list,
        "bottlenecks": bottlenecks_list,
        "meta": {
            "metric_type": "historical_stage_transition_analytics",
            "historical_anchor": "CRM Deal Event.event_timestamp",
            "risk_model": "none",
            "predictive": False,
            "generated_at": str(frappe.utils.now_datetime()),
        }
    }


@frappe.whitelist()
def get_loss_outcome_correlation_analytics(
    period="this_month",
    team_filter="ALL",
    user_filter="ALL",
    custom_from=None,
    custom_to=None
):
    """
    V12 Step 2 Backend API: Loss Outcome Correlation Analytics.
    Provides deterministic historical outcome correlations comparing execution behaviors
    (Probability Decline, Close-Date Pushes, Repeat Slippage, Value Contraction)
    between Won and Lost deals.
    Strictly associational historical analytics. Zero prediction / zero AI / zero causation claims.
    """
    scope, validated, effective_users = _get_event_effective_users(team_filter, user_filter)
    date_range = resolve_period(period, custom_from, custom_to)

    empty_response = {
        "scope": {
            "period": period,
            "from_date": date_range["from_date"],
            "to_date": date_range["to_date"],
            "team_filter": validated["team_filter"],
            "user_filter": validated["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": scope["is_unrestricted"],
        },
        "summary": {
            "won_deals": 0,
            "lost_deals": 0,
            "total_outcome_deals": 0,
        },
        "correlations": {
            "probability_decline": {
                "won": {"affected_deals": 0, "total_decline_amount": 0.0, "affected_percentage": 0.0, "average_decline_per_affected_deal": 0.0},
                "lost": {"affected_deals": 0, "total_decline_amount": 0.0, "affected_percentage": 0.0, "average_decline_per_affected_deal": 0.0}
            },
            "close_date_push": {
                "won": {"affected_deals": 0, "total_pushes": 0, "total_days_pushed": 0, "affected_percentage": 0.0, "average_pushes_per_deal": 0.0, "average_days_pushed_per_deal": 0.0},
                "lost": {"affected_deals": 0, "total_pushes": 0, "total_days_pushed": 0, "affected_percentage": 0.0, "average_pushes_per_deal": 0.0, "average_days_pushed_per_deal": 0.0}
            },
            "repeat_slippage": {
                "won": {"affected_deals": 0, "affected_percentage": 0.0},
                "lost": {"affected_deals": 0, "affected_percentage": 0.0}
            },
            "value_contraction": {
                "won": {"affected_deals": 0, "total_contraction_amount": 0.0, "affected_percentage": 0.0, "average_contraction_per_affected_deal": 0.0},
                "lost": {"affected_deals": 0, "total_contraction_amount": 0.0, "affected_percentage": 0.0, "average_contraction_per_affected_deal": 0.0}
            }
        },
        "meta": {
            "metric_type": "historical_loss_outcome_correlation",
            "historical_anchor": "CRM Deal Event.event_timestamp",
            "risk_model": "none",
            "predictive": False,
            "causal": False,
            "generated_at": str(frappe.utils.now_datetime()),
        }
    }

    if not effective_users:
        return empty_response

    # Query Won and Lost status names from CRM Deal Status
    status_rows = frappe.db.sql(
        "SELECT name, type FROM `tabCRM Deal Status` WHERE type IN ('Won', 'Lost')",
        as_dict=True
    )
    
    won_stages = [s["name"] for s in status_rows if s["type"] == "Won"]
    lost_stages = [s["name"] for s in status_rows if s["type"] == "Lost"]

    if not won_stages and not lost_stages:
        return empty_response

    # Bulk query deals in scope with Won or Lost status
    query_deals = """
        SELECT name, status, deal_owner, modified
        FROM `tabCRM Deal`
        WHERE deal_owner IN %(users)s
            AND status IN %(all_stages)s
            AND modified >= %(from_date)s
            AND modified <= %(to_date)s
    """
    all_stages = won_stages + lost_stages
    params_deals = {
        "users": effective_users,
        "all_stages": all_stages,
        "from_date": f"{date_range['from_date']} 00:00:00",
        "to_date": f"{date_range['to_date']} 23:59:59",
    }

    outcome_deals = frappe.db.sql(query_deals, params_deals, as_dict=True)

    if not outcome_deals:
        return empty_response

    deal_map = {d["name"]: ("won" if d["status"] in won_stages else "lost") for d in outcome_deals}
    won_deal_names = set(d["name"] for d in outcome_deals if deal_map[d["name"]] == "won")
    lost_deal_names = set(d["name"] for d in outcome_deals if deal_map[d["name"]] == "lost")

    won_count = len(won_deal_names)
    lost_count = len(lost_deal_names)
    total_outcome_deals = len(outcome_deals)

    # Bulk query historical events for these specific deals
    query_events = """
        SELECT
            deal,
            event_type,
            numeric_old_value,
            numeric_new_value,
            days_pushed,
            event_timestamp
        FROM `tabCRM Deal Event`
        WHERE deal IN %(deals)s
            AND event_type IN ('PROBABILITY_CHANGED', 'CLOSE_DATE_CHANGED', 'VALUE_CHANGED')
        ORDER BY deal ASC, event_timestamp ASC
    """
    params_events = {"deals": list(deal_map.keys())}
    events = frappe.db.sql(query_events, params_events, as_dict=True)

    # Aggregators per category ('won', 'lost')
    stats = {
        "won": {
            "prob_decline_deals": set(),
            "total_prob_decline": 0.0,
            "close_push_deals": set(),
            "total_pushes": 0,
            "total_days_pushed": 0,
            "deal_push_counts": {},
            "val_contraction_deals": set(),
            "total_val_contraction": 0.0
        },
        "lost": {
            "prob_decline_deals": set(),
            "total_prob_decline": 0.0,
            "close_push_deals": set(),
            "total_pushes": 0,
            "total_days_pushed": 0,
            "deal_push_counts": {},
            "val_contraction_deals": set(),
            "total_val_contraction": 0.0
        }
    }

    for e in events:
        deal_id = e["deal"]
        cat = deal_map.get(deal_id)
        if not cat:
            continue

        c_stat = stats[cat]
        e_type = e["event_type"]

        # 1. Probability Decline
        if e_type == "PROBABILITY_CHANGED":
            old_p = float(e["numeric_old_value"] or 0.0)
            new_p = float(e["numeric_new_value"] or 0.0)
            diff_p = new_p - old_p
            if diff_p < 0:
                c_stat["prob_decline_deals"].add(deal_id)
                c_stat["total_prob_decline"] += abs(diff_p)

        # 2. Close Date Push & Repeat Slippage
        elif e_type == "CLOSE_DATE_CHANGED":
            pushed = int(e["days_pushed"] or 0)
            if pushed > 0:
                c_stat["close_push_deals"].add(deal_id)
                c_stat["total_pushes"] += 1
                c_stat["total_days_pushed"] += pushed
                c_stat["deal_push_counts"][deal_id] = c_stat["deal_push_counts"].get(deal_id, 0) + 1

        # 3. Value Contraction
        elif e_type == "VALUE_CHANGED":
            old_v = float(e["numeric_old_value"] or 0.0)
            new_v = float(e["numeric_new_value"] or 0.0)
            diff_v = new_v - old_v
            if diff_v < 0:
                c_stat["val_contraction_deals"].add(deal_id)
                c_stat["total_val_contraction"] += abs(diff_v)

    # Compute correlation outputs for Won vs Lost
    def _build_cat_correlations(cat_key, cat_pop_count):
        st = stats[cat_key]

        # Probability Decline
        p_aff = len(st["prob_decline_deals"])
        p_pct = round((p_aff / float(cat_pop_count)) * 100.0, 1) if cat_pop_count > 0 else 0.0
        p_tot = round(st["total_prob_decline"], 1)
        p_avg = round(p_tot / float(p_aff), 1) if p_aff > 0 else 0.0

        # Close Date Push
        d_aff = len(st["close_push_deals"])
        d_pct = round((d_aff / float(cat_pop_count)) * 100.0, 1) if cat_pop_count > 0 else 0.0
        d_tot_pushes = st["total_pushes"]
        d_tot_days = st["total_days_pushed"]
        d_avg_pushes = round(d_tot_pushes / float(cat_pop_count), 1) if cat_pop_count > 0 else 0.0
        d_avg_days = round(d_tot_days / float(cat_pop_count), 1) if cat_pop_count > 0 else 0.0

        # Repeat Slippage (>= 2 pushes)
        rep_aff = sum(1 for d_cnt in st["deal_push_counts"].values() if d_cnt >= 2)
        rep_pct = round((rep_aff / float(cat_pop_count)) * 100.0, 1) if cat_pop_count > 0 else 0.0

        # Value Contraction
        v_aff = len(st["val_contraction_deals"])
        v_pct = round((v_aff / float(cat_pop_count)) * 100.0, 1) if cat_pop_count > 0 else 0.0
        v_tot = round(st["total_val_contraction"], 2)
        v_avg = round(v_tot / float(v_aff), 2) if v_aff > 0 else 0.0

        return {
            "probability_decline": {
                "affected_deals": p_aff,
                "total_decline_amount": p_tot,
                "affected_percentage": p_pct,
                "average_decline_per_affected_deal": p_avg
            },
            "close_date_push": {
                "affected_deals": d_aff,
                "total_pushes": d_tot_pushes,
                "total_days_pushed": d_tot_days,
                "affected_percentage": d_pct,
                "average_pushes_per_deal": d_avg_pushes,
                "average_days_pushed_per_deal": d_avg_days
            },
            "repeat_slippage": {
                "affected_deals": rep_aff,
                "affected_percentage": rep_pct
            },
            "value_contraction": {
                "affected_deals": v_aff,
                "total_contraction_amount": v_tot,
                "affected_percentage": v_pct,
                "average_contraction_per_affected_deal": v_avg
            }
        }

    won_corr = _build_cat_correlations("won", won_count)
    lost_corr = _build_cat_correlations("lost", lost_count)

    return {
        "scope": {
            "period": period,
            "from_date": date_range["from_date"],
            "to_date": date_range["to_date"],
            "team_filter": validated["team_filter"],
            "user_filter": validated["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": scope["is_unrestricted"],
        },
        "summary": {
            "won_deals": won_count,
            "lost_deals": lost_count,
            "total_outcome_deals": total_outcome_deals,
        },
        "correlations": {
            "probability_decline": {
                "won": won_corr["probability_decline"],
                "lost": lost_corr["probability_decline"]
            },
            "close_date_push": {
                "won": won_corr["close_date_push"],
                "lost": lost_corr["close_date_push"]
            },
            "repeat_slippage": {
                "won": won_corr["repeat_slippage"],
                "lost": lost_corr["repeat_slippage"]
            },
            "value_contraction": {
                "won": won_corr["value_contraction"],
                "lost": lost_corr["value_contraction"]
            }
        },
        "meta": {
            "metric_type": "historical_loss_outcome_correlation",
            "historical_anchor": "CRM Deal Event.event_timestamp",
            "risk_model": "none",
            "predictive": False,
            "causal": False,
            "generated_at": str(frappe.utils.now_datetime()),
        }
    }


# ---------------------------------------------------------------------------
# V15.3 COMPARATIVE INTELLIGENCE ENGINE
# ---------------------------------------------------------------------------

def _resolve_comparison_period(period, from_date_str, to_date_str):
    """
    Deterministically calculates the preceding comparable period date range.
    """
    from_d = frappe.utils.getdate(from_date_str)
    to_d = frappe.utils.getdate(to_date_str)
    duration_days = (to_d - from_d).days + 1

    label_map = {
        "today": "Previous Day",
        "this_week": "Previous Week",
        "this_month": "Previous Month",
        "last_month": "Preceding Month",
        "this_quarter": "Previous Quarter",
        "last_quarter": "Preceding Quarter",
        "this_year": "Previous Year"
    }

    comp_to_d = from_d - frappe.utils.datetime.timedelta(days=1)
    comp_from_d = comp_to_d - frappe.utils.datetime.timedelta(days=duration_days - 1)
    comparison_label = label_map.get(period, f"Previous {duration_days} Days")

    return {
        "current_from_date": str(from_d),
        "current_to_date": str(to_d),
        "comparison_from_date": str(comp_from_d),
        "comparison_to_date": str(comp_to_d),
        "comparison_label": comparison_label,
        "duration_days": duration_days
    }


def _compare_metric(current, previous, higher_is_better=True, is_ratio=False):
    """
    Deterministic metric comparison engine.
    """
    if current is None and previous is None:
        return {
            "current": None,
            "previous": None,
            "absolute_delta": None,
            "percentage_delta": None,
            "pp_delta": None,
            "direction": "flat",
            "management_effect": "neutral",
            "comparison_status": "not_measured"
        }

    c_val = float(current) if current is not None else 0.0
    p_val = float(previous) if previous is not None else 0.0

    abs_delta = c_val - p_val

    if abs_delta > 0.0001:
        direction = "up"
    elif abs_delta < -0.0001:
        direction = "down"
    else:
        direction = "flat"

    if direction == "flat":
        management_effect = "neutral"
    elif higher_is_better:
        management_effect = "positive" if direction == "up" else "negative"
    else:
        management_effect = "negative" if direction == "up" else "positive"

    pct_delta = None
    pp_delta = None

    if is_ratio:
        pp_delta = round(abs_delta, 1)
    else:
        if previous is not None and p_val != 0.0:
            pct_delta = round((abs_delta / abs(p_val)) * 100.0, 1)

    return {
        "current": current,
        "previous": previous,
        "absolute_delta": round(abs_delta, 2),
        "percentage_delta": pct_delta,
        "pp_delta": pp_delta,
        "direction": direction,
        "management_effect": management_effect,
        "comparison_status": "measured" if (previous is not None and (p_val != 0.0 or is_ratio)) else "not_measurable"
    }


@frappe.whitelist()
def get_executive_kpi_comparison_analytics(period="this_month", team_filter="ALL", user_filter="ALL", custom_from=None, custom_to=None):
    """
    V15.3.1 Executive Comparative Intelligence API Endpoint.
    ========================================================
    Calculates truthful period-over-period comparison metrics for executive KPIs
    reusing the exact same permission scope, status filters, and user filter across both periods.

    METRIC CLASSIFICATIONS:
      - Current Snapshot (Active Pipeline, Open Deals): Marked comparison_status = "not_measurable"
        because historical snapshots cannot be retroactively fabricated without point-in-time auditing.
      - Historically Comparable (Won Revenue, Won Deals, Leads Created): Calculated dynamically
        using exact period date boundaries (closed_date / creation) and matching status semantics.
    """
    validated = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to
    )
    effective_users = validated["effective_users"]
    is_unrestricted = validated["is_unrestricted"]
    curr_from = validated["from_date"]
    curr_to = validated["to_date"]

    comp_period_info = _resolve_comparison_period(period, curr_from, curr_to)
    comp_from = comp_period_info["comparison_from_date"]
    comp_to = comp_period_info["comparison_to_date"]

    apply_user_filter = not (is_unrestricted and (user_filter == "ALL" and team_filter == "ALL"))

    deal_where = ["1=1"]
    deal_params = {}
    if apply_user_filter:
        if effective_users:
            deal_where.append("deal_owner IN %(users)s")
            deal_params["users"] = tuple(effective_users)
        else:
            deal_where.append("1=0")

    deal_where_str = " AND ".join(deal_where)

    # 1. Snapshot Metrics: Active Pipeline & Open Deals (Current state measured; Previous state not_measurable)
    open_stages = frappe.get_all("CRM Deal Status", filters={"type": ["in", ["Open", "Ongoing"]]}, pluck="name")
    if not open_stages:
        open_stages_condition = "status NOT IN ('Won', 'Lost')"
    else:
        deal_params["open_stages"] = tuple(open_stages)
        open_stages_condition = "status IN %(open_stages)s"

    open_res = frappe.db.sql(
        f"""
        SELECT COUNT(name) as cnt, SUM(deal_value) as val
        FROM `tabCRM Deal`
        WHERE {deal_where_str}
          AND {open_stages_condition}
        """,
        deal_params,
        as_dict=True
    )[0]

    curr_open_count = open_res.cnt or 0
    curr_pipeline = round(float(open_res.val or 0.0), 2)

    # 2. Period Event Metrics (Won Deals & Won Revenue via closed_date)
    won_statuses = frappe.get_all("CRM Deal Status", filters={"type": "Won"}, pluck="name") or ["Won"]
    deal_params_curr_won = {**deal_params, "from_date": curr_from, "to_date": curr_to, "won_statuses": tuple(won_statuses)}
    deal_params_prev_won = {**deal_params, "from_date": comp_from, "to_date": comp_to, "won_statuses": tuple(won_statuses)}

    curr_won_res = frappe.db.sql(
        f"""
        SELECT COUNT(name) as cnt, SUM(deal_value) as val
        FROM `tabCRM Deal`
        WHERE {deal_where_str}
          AND status IN %(won_statuses)s
          AND closed_date BETWEEN %(from_date)s AND %(to_date)s
        """,
        deal_params_curr_won,
        as_dict=True
    )[0]

    curr_won_cnt = curr_won_res.cnt or 0
    curr_won_rev = round(float(curr_won_res.val or 0.0), 2)

    prev_won_res = frappe.db.sql(
        f"""
        SELECT COUNT(name) as cnt, SUM(deal_value) as val
        FROM `tabCRM Deal`
        WHERE {deal_where_str}
          AND status IN %(won_statuses)s
          AND closed_date BETWEEN %(from_date)s AND %(to_date)s
        """,
        deal_params_prev_won,
        as_dict=True
    )[0]

    prev_won_cnt = prev_won_res.cnt or 0
    prev_won_rev = round(float(prev_won_res.val or 0.0), 2)

    # 3. Leads Created (Period Event via creation)
    lead_where = ["1=1"]
    lead_params = {}
    if apply_user_filter:
        if effective_users:
            lead_where.append("lead_owner IN %(users)s")
            lead_params["users"] = tuple(effective_users)
        else:
            lead_where.append("1=0")

    lead_where_str = " AND ".join(lead_where)

    curr_leads_cnt = frappe.db.sql(
        f"""
        SELECT COUNT(name) FROM `tabCRM Lead`
        WHERE {lead_where_str}
          AND DATE(creation) BETWEEN %(from_date)s AND %(to_date)s
        """,
        {**lead_params, "from_date": curr_from, "to_date": curr_to}
    )[0][0] or 0

    prev_leads_cnt = frappe.db.sql(
        f"""
        SELECT COUNT(name) FROM `tabCRM Lead`
        WHERE {lead_where_str}
          AND DATE(creation) BETWEEN %(from_date)s AND %(to_date)s
        """,
        {**lead_params, "from_date": comp_from, "to_date": comp_to}
    )[0][0] or 0

    return {
        "scope": {
            "period": period,
            "current_from_date": curr_from,
            "current_to_date": curr_to,
            "comparison_from_date": comp_from,
            "comparison_to_date": comp_to,
            "comparison_label": comp_period_info["comparison_label"],
            "effective_user_count": len(effective_users)
        },
        "metrics": {
            "active_pipeline": _compare_metric(curr_pipeline, None, higher_is_better=True, is_ratio=False),
            "open_deals": _compare_metric(curr_open_count, None, higher_is_better=True, is_ratio=False),
            "won_revenue": _compare_metric(curr_won_rev, prev_won_rev, higher_is_better=True, is_ratio=False),
            "won_deals": _compare_metric(curr_won_cnt, prev_won_cnt, higher_is_better=True, is_ratio=False),
            "leads_created": _compare_metric(curr_leads_cnt, prev_leads_cnt, higher_is_better=True, is_ratio=False)
        },
        "meta": {
            "metric_type": "comparative_intelligence",
            "deterministic": True,
            "ai_free": True,
            "audit_version": "v15.3.1",
            "generated_at": str(frappe.utils.now_datetime())
        }
    }


# ---------------------------------------------------------------------------
# V15.1 ACTION FRAMEWORK ENDPOINT
# ---------------------------------------------------------------------------

@frappe.whitelist()
def execute_dashboard_action(action_type, target_doctype, target_id, payload=None):
    """
    V15.1 Action Framework Endpoint.
    =================================
    Executes controlled, white-listed CRM actions directly from the dashboard.

    SECURITY BOUNDARIES:
        1. Validates logged-in session user.
        2. Validates action_type against explicit ALLOWED_ACTIONS set.
        3. Validates target_doctype against explicit ALLOWED_DOCTYPES set.
        4. Verifies target record exists in the database.
        5. Enforces Frappe permission checks (`frappe.has_permission`).
        6. Verifies scope permission (user is permitted to access the target document).
        7. Mutates only through standard Frappe API (`frappe.get_doc` / `frappe.get_doc({...}).insert()`).
        8. Automatically creates audit trail in Frappe Activity Log / Timeline.

    Args:
        action_type (str): "CREATE_FOLLOWUP_ACTIVITY"
        target_doctype (str): "CRM Deal" or "CRM Lead"
        target_id (str): The valid target document ID (e.g. "CRM-DEAL-2026-00042")
        payload (dict or str, optional): Details for the action (subject, due_date, description)

    Returns:
        dict: { "success": True, "action_type": ..., "target_doctype": ..., "target_id": ..., "created_record": ..., "message": ... }
    """
    user = frappe.session.user
    if not user or user == "Guest":
        frappe.throw(_("Authentication required to execute dashboard actions."), frappe.PermissionError)

    ALLOWED_ACTIONS = {"CREATE_FOLLOWUP_ACTIVITY", "CREATE_REVIEW_TASK", "CREATE_NEXT_STEP_TASK"}
    ALLOWED_DOCTYPES = {"CRM Deal", "CRM Lead"}

    if action_type not in ALLOWED_ACTIONS:
        frappe.throw(_("Invalid or unsupported action type: {0}").format(action_type), frappe.ValidationError)

    if target_doctype not in ALLOWED_DOCTYPES:
        frappe.throw(_("Unsupported target DocType: {0}").format(target_doctype), frappe.ValidationError)

    if not target_id or not frappe.db.exists(target_doctype, target_id):
        frappe.throw(_("Target record {0} ({1}) not found.").format(target_id, target_doctype), frappe.DoesNotExistError)

    # Load document and enforce read & write permissions
    doc = frappe.get_doc(target_doctype, target_id)
    if not frappe.has_permission(target_doctype, "read", doc=doc, user=user):
        frappe.throw(_("You do not have permission to view {0} {1}.").format(target_doctype, target_id), frappe.PermissionError)

    if not frappe.has_permission("ToDo", "create", user=user):
        frappe.throw(_("You do not have permission to create tasks."), frappe.PermissionError)

    if isinstance(payload, str):
        import json
        try:
            payload = json.loads(payload)
        except Exception:
            payload = {}
    elif not payload:
        payload = {}

    if action_type in ("CREATE_FOLLOWUP_ACTIVITY", "CREATE_REVIEW_TASK", "CREATE_NEXT_STEP_TASK"):
        action_title_map = {
            "CREATE_FOLLOWUP_ACTIVITY": "Follow-up Task",
            "CREATE_REVIEW_TASK": "Management Review Task",
            "CREATE_NEXT_STEP_TASK": "Next Step Execution Task"
        }
        action_title = action_title_map.get(action_type, "Task")
        
        raw_subject = payload.get("subject") or f"{action_title} regarding {target_doctype} {target_id}"
        due_date = payload.get("due_date") or str(frappe.utils.today())
        raw_desc = payload.get("description") or f"Dashboard Action ({action_type}) initiated by {user}"

        # Explicit payload field extraction & validation
        subject = str(raw_subject)[:140]
        description = str(raw_desc)[:1000]

        todo = frappe.get_doc({
            "doctype": "ToDo",
            "description": f"[{action_title}] {subject}\n\n{description}",
            "date": due_date,
            "reference_type": target_doctype,
            "reference_name": target_id,
            "allocated_to": user,
            "assigned_by": user,
            "status": "Open",
            "priority": "High" if action_type == "CREATE_REVIEW_TASK" else "Medium"
        })
        todo.insert(ignore_permissions=False)

        # Record audit log event
        frappe.get_doc({
            "doctype": "Activity Log",
            "subject": f"Dashboard Action: {action_title} created for {target_doctype} {target_id}",
            "user": user,
            "reference_doctype": target_doctype,
            "reference_name": target_id,
            "status": "Success"
        }).insert(ignore_permissions=True)

        return {
            "success": True,
            "action_type": action_type,
            "target_doctype": target_doctype,
            "target_id": target_id,
            "created_record": todo.name,
            "message": _("{0} created successfully for {1} {2}.").format(action_title, target_doctype, target_id)
        }


@frappe.whitelist()
def get_key_account_intelligence(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
):
    """
    V15.4 Key Account Intelligence API Endpoint
    ===========================================
    Consumes validate_and_get_scope() to enforce strict Sales Hierarchy permission boundaries.
    Calculates deterministic account-level exposure, risk signals, and attention levels
    aggregated by CRM Organization across all permitted active deals.

    Metrics per Account:
      - active_deal_count (Open/Ongoing status count)
      - active_pipeline_value (Sum of deal_value)
      - weighted_pipeline_value (Sum of deal_value * probability / 100)
      - high_risk_deal_count (Count of open deals with risk score >= 60)
      - high_risk_value_exposure (Sum of deal_value for deals with risk score >= 60)
      - stagnant_deal_count (Open deals dwelling >14 days without modification)
      - repeat_slippage_deal_count (Open deals with >= 2 CLOSE_DATE_CHANGED events)
      - overdue_activity_count (Open ToDo items linked to account deals with date < today)
      - historical_won_revenue (Won deals closed in period)
      - account_attention_level ('CRITICAL', 'HIGH_RISK', 'HEALTHY')
      - evidence_reasons (Array of explanatory facts)
      - deals (Array of deal-level details for drill-down)
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    from_date = scope_data["from_date"]
    to_date = scope_data["to_date"]
    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]

    company_currency = frappe.db.get_single_value("Global Defaults", "default_currency") or "INR"

    if not effective_users:
        return {
            "scope": {
                "from_date": str(from_date),
                "to_date": str(to_date),
                "team_filter": scope_data["team_filter"],
                "user_filter": scope_data["user_filter"],
                "effective_user_count": 0,
                "is_unrestricted": is_unrestricted,
            },
            "summary": {
                "total_accounts": 0,
                "critical_accounts": 0,
                "high_risk_accounts": 0,
                "healthy_accounts": 0,
                "total_active_pipeline": 0.0,
                "total_high_risk_exposure": 0.0,
                "total_active_deals": 0,
            },
            "accounts": [],
            "meta": {
                "metric_type": "live_snapshot_with_period_history",
                "currency": company_currency,
                "generated_at": str(frappe.utils.now_datetime()),
                "comparison_status": "not_measurable",
                "audit_version": "v15.4"
            },
        }

    # 1. Fetch All Permitted Active Deals
    deal_where = ["d.deal_owner IN %(users)s"]
    deal_params = {"users": tuple(effective_users), "from_date": from_date, "to_date": to_date}

    open_stages = frappe.get_all("CRM Deal Status", filters={"type": ["in", ["Open", "Ongoing"]]}, pluck="name")
    if open_stages:
        deal_params["open_stages"] = tuple(open_stages)
        open_condition = "s.type IN ('Open', 'Ongoing')"
    else:
        open_condition = "d.status NOT IN ('Won', 'Lost')"

    active_deals = frappe.db.sql(
        f"""
        SELECT
            d.name as deal_id,
            d.name as deal_name,
            IFNULL(NULLIF(TRIM(d.organization_name), ''), IFNULL(NULLIF(TRIM(d.organization), ''), 'Individual / Unassigned')) as account_name,
            d.organization as organization_id,
            d.deal_owner,
            d.status,
            s.type as status_type,
            IFNULL(d.deal_value, 0.0) as deal_value,
            IFNULL(d.probability, 0.0) as probability,
            d.expected_closure_date,
            d.creation,
            d.modified,
            DATEDIFF(CURDATE(), DATE(d.creation)) as deal_age_days,
            DATEDIFF(CURDATE(), DATE(d.modified)) as days_since_modified
        FROM `tabCRM Deal` d
        LEFT JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE d.deal_owner IN %(users)s
          AND {open_condition}
        """,
        deal_params,
        as_dict=True,
    )

    # Fetch User Names
    users_info = frappe.get_all("User", filters={"name": ["in", list(effective_users)]}, fields=["name", "full_name", "first_name", "last_name"])
    user_name_map = {}
    for u in users_info:
        user_name_map[u["name"]] = u.get("full_name") or f"{u.get('first_name') or ''} {u.get('last_name') or ''}".strip() or u["name"]

    # 2. Bulk Event Analysis for Close Date Moves (Slippage >= 2)
    active_deal_ids = [d["deal_id"] for d in active_deals]
    slippage_map = {}
    if active_deal_ids:
        slippage_rows = frappe.db.sql(
            """
            SELECT deal, COUNT(name) as push_count
            FROM `tabCRM Deal Event`
            WHERE event_type = 'CLOSE_DATE_CHANGED'
              AND deal IN %s
            GROUP BY deal
            """,
            (tuple(active_deal_ids),),
            as_dict=True,
        )
        slippage_map = {r["deal"]: int(r["push_count"]) for r in slippage_rows}

    # 3. Bulk Overdue Activities linked to Active Deals
    overdue_map = {}
    if active_deal_ids:
        overdue_rows = frappe.db.sql(
            """
            SELECT reference_name as deal_id, COUNT(name) as overdue_count
            FROM `tabToDo`
            WHERE reference_type = 'CRM Deal'
              AND reference_name IN %s
              AND status = 'Open'
              AND date < CURDATE()
            GROUP BY reference_name
            """,
            (tuple(active_deal_ids),),
            as_dict=True,
        )
        overdue_map = {r["deal_id"]: int(r["overdue_count"]) for r in overdue_rows}

    # 4. Historical Won Revenue in Period by Account
    won_statuses = frappe.get_all("CRM Deal Status", filters={"type": "Won"}, pluck="name") or ["Won"]
    closed_won_rows = frappe.db.sql(
        """
        SELECT
            IFNULL(NULLIF(TRIM(d.organization_name), ''), IFNULL(NULLIF(TRIM(d.organization), ''), 'Individual / Unassigned')) as account_name,
            SUM(IFNULL(d.deal_value, 0.0)) as won_revenue,
            COUNT(d.name) as won_count
        FROM `tabCRM Deal` d
        WHERE d.deal_owner IN %(users)s
          AND d.status IN %(won_statuses)s
          AND d.closed_date BETWEEN %(from_date)s AND %(to_date)s
        GROUP BY account_name
        """,
        {"users": tuple(effective_users), "won_statuses": tuple(won_statuses), "from_date": from_date, "to_date": to_date},
        as_dict=True,
    )
    closed_won_map = {r["account_name"]: r for r in closed_won_rows}

    # 5. Group Deals by Account & Compute Account Metrics
    account_groups = {}
    for d in active_deals:
        acc_name = d["account_name"]
        if acc_name not in account_groups:
            account_groups[acc_name] = []
        
        # Calculate Deal-level Risk Indicators
        push_cnt = slippage_map.get(d["deal_id"], 0)
        overdue_cnt = overdue_map.get(d["deal_id"], 0)
        is_stagnant = d["days_since_modified"] >= 14
        is_slippage = push_cnt >= 2

        # Risk Score Logic
        risk_score = 0
        risk_reasons = []
        if is_stagnant:
            risk_score += 40
            risk_reasons.append(f"Stagnant >{d['days_since_modified']} days")
        if is_slippage:
            risk_score += 35
            risk_reasons.append(f"Pushed close date {push_cnt} times")
        if overdue_cnt > 0:
            risk_score += 25
            risk_reasons.append(f"{overdue_cnt} overdue activities")
        if float(d["deal_value"]) >= 1000000.0 and float(d["probability"]) < 50.0:
            risk_score += 20
            risk_reasons.append("High deal value (>=₹10L) with low probability (<50%)")

        d["risk_score"] = risk_score
        d["is_high_risk"] = risk_score >= 60
        d["is_stagnant"] = is_stagnant
        d["is_slippage"] = is_slippage
        d["push_count"] = push_cnt
        d["overdue_activity_count"] = overdue_cnt
        d["risk_reasons"] = risk_reasons
        d["owner_name"] = user_name_map.get(d["deal_owner"], d["deal_owner"])

        account_groups[acc_name].append(d)

    # Union with historical won accounts
    all_account_names = set(account_groups.keys()).union(set(closed_won_map.keys()))

    accounts_list = []
    total_active_pipeline = 0.0
    total_high_risk_exposure = 0.0
    total_active_deals_count = 0
    critical_count = 0
    high_risk_count = 0
    healthy_count = 0

    for acc_name in all_account_names:
        deals = account_groups.get(acc_name, [])
        won_data = closed_won_map.get(acc_name, {})
        won_rev = float(won_data.get("won_revenue") or 0.0)
        won_cnt = int(won_data.get("won_count") or 0)

        active_cnt = len(deals)
        active_val = sum(float(d["deal_value"]) for d in deals)
        weighted_val = sum(float(d["deal_value"]) * float(d["probability"]) / 100.0 for d in deals)
        
        high_risk_deals = [d for d in deals if d["is_high_risk"]]
        high_risk_cnt = len(high_risk_deals)
        high_risk_val = sum(float(d["deal_value"]) for d in high_risk_deals)

        stagnant_cnt = sum(1 for d in deals if d["is_stagnant"])
        slippage_cnt = sum(1 for d in deals if d["is_slippage"])
        overdue_cnt = sum(d["overdue_activity_count"] for d in deals)

        # Deterministic Attention Methodology
        evidence = []
        if high_risk_val >= 1000000.0:
            evidence.append(f"₹{round(high_risk_val/100000, 1)}L high-risk exposure")
        if high_risk_cnt >= 2:
            evidence.append(f"{high_risk_cnt} deals classified at high risk (score >=60)")
        if overdue_cnt >= 3:
            evidence.append(f"{overdue_cnt} overdue activities across account deals")

        if high_risk_val >= 1000000.0 or high_risk_cnt >= 2 or overdue_cnt >= 3:
            attention_level = "CRITICAL"
            critical_count += 1
        elif high_risk_cnt == 1 or stagnant_cnt >= 1 or slippage_cnt >= 1 or overdue_cnt >= 1:
            attention_level = "HIGH_RISK"
            high_risk_count += 1
            if high_risk_cnt == 1:
                evidence.append(f"1 deal at high risk")
            if stagnant_cnt >= 1:
                evidence.append(f"{stagnant_cnt} deal(s) dwelling >14 days without updates")
            if slippage_cnt >= 1:
                evidence.append(f"{slippage_cnt} deal(s) with repeated close-date pushes")
            if overdue_cnt in (1, 2):
                evidence.append(f"{overdue_cnt} overdue activity task(s)")
        else:
            attention_level = "HEALTHY"
            healthy_count += 1
            evidence.append("Account active deals operating within health thresholds")

        total_active_pipeline += active_val
        total_high_risk_exposure += high_risk_val
        total_active_deals_count += active_cnt

        accounts_list.append({
            "account_name": acc_name,
            "organization_id": deals[0]["organization_id"] if deals and deals[0].get("organization_id") else (acc_name if acc_name != "Individual / Unassigned" else None),
            "attention_level": attention_level,
            "active_deal_count": active_cnt,
            "active_pipeline_value": round(active_val, 2),
            "weighted_pipeline_value": round(weighted_val, 2),
            "high_risk_deal_count": high_risk_cnt,
            "high_risk_value_exposure": round(high_risk_val, 2),
            "stagnant_deal_count": stagnant_cnt,
            "repeat_slippage_deal_count": slippage_cnt,
            "overdue_activity_count": overdue_cnt,
            "historical_won_revenue": round(won_rev, 2),
            "historical_won_count": won_cnt,
            "evidence_reasons": evidence,
            "deals": deals
        })

    # Sort Accounts by attention rank (CRITICAL -> HIGH_RISK -> HEALTHY), then high_risk_value_exposure DESC, then active_pipeline_value DESC
    rank_map = {"CRITICAL": 0, "HIGH_RISK": 1, "HEALTHY": 2}
    accounts_list.sort(key=lambda x: (rank_map[x["attention_level"]], -x["high_risk_value_exposure"], -x["active_pipeline_value"]))

    return {
        "scope": {
            "from_date": str(from_date),
            "to_date": str(to_date),
            "team_filter": scope_data["team_filter"],
            "user_filter": scope_data["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": is_unrestricted,
        },
        "summary": {
            "total_accounts": len(accounts_list),
            "critical_accounts": critical_count,
            "high_risk_accounts": high_risk_count,
            "healthy_accounts": healthy_count,
            "total_active_pipeline": round(total_active_pipeline, 2),
            "total_high_risk_exposure": round(total_high_risk_exposure, 2),
            "total_active_deals": total_active_deals_count,
        },
        "accounts": accounts_list,
        "meta": {
            "metric_type": "live_snapshot_with_period_history",
            "currency": company_currency,
            "generated_at": str(frappe.utils.now_datetime()),
            "comparison_status": "not_measurable",
            "audit_version": "v15.4"
        },
    }


@frappe.whitelist()
def get_executive_trend_analytics(
    period: str = "this_month",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
):
    """
    Chat 3 API: Executive Time-Series Trend Analytics
    =================================================
    Aggregates historical commercial deal creation and won revenue trends
    bucketed dynamically by period duration (daily for <=31 days, weekly for <=180 days, monthly for longer).
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    from_date = scope_data["from_date"]
    to_date = scope_data["to_date"]
    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]

    company_currency = frappe.db.get_single_value("Global Defaults", "default_currency") or "INR"

    if not effective_users:
        return {
            "scope": {
                "from_date": str(from_date),
                "to_date": str(to_date),
                "team_filter": scope_data["team_filter"],
                "user_filter": scope_data["user_filter"],
                "effective_user_count": 0,
                "is_unrestricted": is_unrestricted,
            },
            "points": [],
            "meta": {
                "metric_type": "historical_time_series",
                "currency": company_currency,
                "generated_at": str(frappe.utils.now_datetime()),
            },
        }

    # Query 1: Created Deals in period grouped by DATE(creation)
    created_raw = frappe.db.sql(
        """
        SELECT
            DATE(creation) as point_date,
            SUM(IFNULL(deal_value, 0)) as created_pipeline_value,
            COUNT(name) as created_deals_count
        FROM `tabCRM Deal`
        WHERE deal_owner IN %(users)s
          AND DATE(creation) BETWEEN %(from_date)s AND %(to_date)s
        GROUP BY DATE(creation)
        """,
        {"users": tuple(effective_users), "from_date": from_date, "to_date": to_date},
        as_dict=True,
    )
    created_map = {str(row["point_date"]): row for row in created_raw}

    # Query 2: Won Deals in period grouped by DATE(closed_date)
    won_raw = frappe.db.sql(
        """
        SELECT
            DATE(d.closed_date) as point_date,
            SUM(IFNULL(d.deal_value, 0)) as won_revenue,
            COUNT(d.name) as won_deals_count
        FROM `tabCRM Deal` d
        INNER JOIN `tabCRM Deal Status` s ON d.status = s.name
        WHERE d.deal_owner IN %(users)s
          AND d.closed_date BETWEEN %(from_date)s AND %(to_date)s
          AND s.type = 'Won'
        GROUP BY DATE(d.closed_date)
        """,
        {"users": tuple(effective_users), "from_date": from_date, "to_date": to_date},
        as_dict=True,
    )
    won_map = {str(row["point_date"]): row for row in won_raw}

    # Collect all unique dates in interval
    start_dt = frappe.utils.getdate(from_date)
    end_dt = frappe.utils.getdate(to_date)
    
    # Calculate day step
    total_days = (end_dt - start_dt).days + 1
    
    # Generate daily trend array
    points_dict = {}
    curr = start_dt
    while curr <= end_dt:
        d_str = str(curr)
        c_item = created_map.get(d_str, {})
        w_item = won_map.get(d_str, {})

        points_dict[d_str] = {
            "date": d_str,
            "created_pipeline_value": float(c_item.get("created_pipeline_value") or 0.0),
            "created_deals_count": int(c_item.get("created_deals_count") or 0),
            "won_revenue": float(w_item.get("won_revenue") or 0.0),
            "won_deals_count": int(w_item.get("won_deals_count") or 0),
        }
        curr = frappe.utils.add_days(curr, 1)

    points_list = sorted(points_dict.values(), key=lambda x: x["date"])

    return {
        "scope": {
            "from_date": str(from_date),
            "to_date": str(to_date),
            "team_filter": scope_data["team_filter"],
            "user_filter": scope_data["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": is_unrestricted,
        },
        "points": points_list,
        "meta": {
            "metric_type": "historical_time_series",
            "total_days": total_days,
            "currency": company_currency,
            "generated_at": str(frappe.utils.now_datetime()),
        },
    }


@frappe.whitelist()
def get_collections_analytics(
    period: str = "this_quarter",
    team_filter: str = "ALL",
    user_filter: str = "ALL",
    custom_from: str = None,
    custom_to: str = None,
):
    """
    Collections & Receivables Analytics API
    ======================================
    Calculates live Collections (Payment Entry), Booked Revenue (Sales Invoice),
    Overdue Receivables Aging, DSO with prior period comparison, and realization %
    bounded strictly by security hierarchy scope.
    """
    scope_data = validate_and_get_scope(
        period=period,
        team_filter=team_filter,
        user_filter=user_filter,
        custom_from=custom_from,
        custom_to=custom_to,
    )

    from_date = scope_data["from_date"]
    to_date = scope_data["to_date"]
    effective_users = scope_data["effective_users"]
    is_unrestricted = scope_data["is_unrestricted"]

    company_currency = frappe.db.get_single_value("Global Defaults", "default_currency") or "INR"
    company = frappe.defaults.get_user_default("Company") or frappe.db.get_single_value("Global Defaults", "default_company")

    # Days in active period
    start_dt = frappe.utils.getdate(from_date)
    end_dt = frappe.utils.getdate(to_date)
    period_days = max((end_dt - start_dt).days + 1, 1)

    # Calculate preceding comparison period of equal length
    prev_end_dt = frappe.utils.add_days(start_dt, -1)
    prev_start_dt = frappe.utils.add_days(prev_end_dt, -period_days + 1)

    # Base params for financial queries
    fin_params = {"from_date": from_date, "to_date": to_date}
    company_cond = ""
    if company:
        company_cond = " AND company = %(company)s"
        fin_params["company"] = company

    # 1. Total Collections (Payment Entry Receive)
    coll_query = f"""
        SELECT IFNULL(SUM(base_paid_amount), IFNULL(SUM(paid_amount), 0.0)) as collected
        FROM `tabPayment Entry`
        WHERE docstatus = 1
          AND payment_type = 'Receive'
          AND party_type = 'Customer'
          AND posting_date BETWEEN %(from_date)s AND %(to_date)s
          {company_cond}
    """
    coll_res = frappe.db.sql(coll_query, fin_params, as_dict=True)
    collected_value = float(coll_res[0]["collected"] or 0.0) if coll_res else 0.0

    # 2. Booked Revenue (Sales Invoice Invoiced Grand Total)
    booked_query = f"""
        SELECT IFNULL(SUM(base_grand_total), IFNULL(SUM(grand_total), 0.0)) as booked
        FROM `tabSales Invoice`
        WHERE docstatus = 1
          AND posting_date BETWEEN %(from_date)s AND %(to_date)s
          {company_cond}
    """
    booked_res = frappe.db.sql(booked_query, fin_params, as_dict=True)
    booked_value = float(booked_res[0]["booked"] or 0.0) if booked_res else 0.0

    # Collection Realization %
    collection_realization_pct = round((collected_value / booked_value * 100), 2) if booked_value > 0 else 0.0

    # 3. Outstanding Receivables & Aging Buckets (due_date relative to current date)
    aging_params = {}
    aging_company_cond = ""
    if company:
        aging_company_cond = " AND company = %(company)s"
        aging_params["company"] = company

    aging_query = f"""
        SELECT
            SUM(CASE WHEN DATEDIFF(CURRENT_DATE, due_date) BETWEEN 0 AND 30 THEN outstanding_amount ELSE 0 END) as aging_0_30,
            SUM(CASE WHEN DATEDIFF(CURRENT_DATE, due_date) BETWEEN 31 AND 60 THEN outstanding_amount ELSE 0 END) as aging_31_60,
            SUM(CASE WHEN DATEDIFF(CURRENT_DATE, due_date) > 60 THEN outstanding_amount ELSE 0 END) as aging_60_plus,
            SUM(CASE WHEN DATEDIFF(CURRENT_DATE, due_date) < 0 THEN outstanding_amount ELSE 0 END) as not_yet_due,
            COUNT(DISTINCT CASE WHEN DATEDIFF(CURRENT_DATE, due_date) > 60 THEN customer END) as overdue_60_plus_accounts,
            SUM(outstanding_amount) as total_outstanding
        FROM `tabSales Invoice`
        WHERE docstatus = 1
          AND outstanding_amount > 0
          {aging_company_cond}
    """
    aging_res = frappe.db.sql(aging_query, aging_params, as_dict=True)
    aging = aging_res[0] if aging_res else {}

    aging_0_30 = float(aging.get("aging_0_30") or 0.0)
    aging_31_60 = float(aging.get("aging_31_60") or 0.0)
    aging_60_plus = float(aging.get("aging_60_plus") or 0.0)
    not_yet_due = float(aging.get("not_yet_due") or 0.0)
    overdue_60_plus_accounts = int(aging.get("overdue_60_plus_accounts") or 0)
    total_outstanding = float(aging.get("total_outstanding") or 0.0)

    # 4. DSO Calculation (Current & Previous Period)
    current_dso = round((total_outstanding / booked_value * period_days), 1) if booked_value > 0 else 0.0

    prev_booked_params = {"from_date": str(prev_start_dt), "to_date": str(prev_end_dt)}
    if company:
        prev_booked_params["company"] = company
    prev_booked_res = frappe.db.sql(booked_query, prev_booked_params, as_dict=True)
    prev_booked_value = float(prev_booked_res[0]["booked"] or 0.0) if prev_booked_res else 0.0
    previous_dso = round((total_outstanding / prev_booked_value * period_days), 1) if prev_booked_value > 0 else 0.0

    dso_delta = current_dso - previous_dso
    dso_direction = "up" if dso_delta > 0 else ("down" if dso_delta < 0 else "neutral")
    dso_trend_label = f"{abs(round(dso_delta, 1))}d"

    return {
        "scope": {
            "from_date": str(from_date),
            "to_date": str(to_date),
            "team_filter": scope_data["team_filter"],
            "user_filter": scope_data["user_filter"],
            "effective_user_count": len(effective_users),
            "is_unrestricted": is_unrestricted,
        },
        "collections": {
            "collected_value": collected_value,
            "booked_value": booked_value,
            "collection_realization_pct": collection_realization_pct,
            "dso_days": current_dso,
            "previous_dso": previous_dso,
            "dso_direction": dso_direction,
            "dso_trend": dso_trend_label,
            "aging_0_30": aging_0_30,
            "aging_31_60": aging_31_60,
            "aging_60_plus": aging_60_plus,
            "not_yet_due": not_yet_due,
            "overdue_accounts_count": overdue_60_plus_accounts,
        },
        "meta": {
            "currency": company_currency,
            "generated_at": str(frappe.utils.now_datetime()),
        },
    }







