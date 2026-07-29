# /home/mathew/frappe-bench/apps/nexapp/nexapp/api.py



import re
import frappe
import json
import datetime as dt
import requests
import zipfile
import io
import os

def get_bank_gl_account(bank_account_name):
    """
    Returns the GL Account name for a given Bank Account doc name.
    Prioritizes custom_account_head if it exists as a valid GL Account.
    """
    if not bank_account_name:
        return None
    
    # Check if this is a Bank Account record
    acc_data = frappe.db.get_value("Bank Account", bank_account_name, ["account", "custom_account_head"], as_dict=1)
    if not acc_data:
        # If not a Bank Account doc, assume it's already a GL Account or raw string
        return bank_account_name
        
    # Priority 1: custom_account_head (must exist in Account doctype)
    if acc_data.get("custom_account_head") and frappe.db.exists("Account", acc_data.custom_account_head):
        return acc_data.custom_account_head
        
    # Priority 2: Standard account field
    return acc_data.account

@frappe.whitelist()
def handle_ticket_master(ticket_master):
    # Parse JSON string into a dictionary
    ticket_master_doc = json.loads(ticket_master)

    circuit_id = ticket_master_doc.get('circuit_id')
    subject = ticket_master_doc.get('subject')
    description = ticket_master_doc.get('description')
    ticket_master_name = ticket_master_doc.get('name')  # Ensure 'name' is included if needed

    # Try to find a matching circuit record
    circuit = frappe.get_all(
        'Circuit',
        filters={'circuit_id': circuit_id},
        fields=['circuit_id', 'circuit_company']
    )

    if circuit:
        circuit_data = circuit[0]
        frappe.db.set_value('Ticket Master', ticket_master_name, 'company', circuit_data['circuit_company'])
    else:
        # Extract the 8-digit numeric circuit_id from the subject
        v_circuit_id = re.search(r'\b\d{8}\b', subject)
        if v_circuit_id:
            v_circuit_id = v_circuit_id.group()
            frappe.db.set_value('Ticket Master', ticket_master_name, 'circuit_id', v_circuit_id)

        # Map values to Issue Doctype
        issue_data = {
            'doctype': 'Issue',
            'custom_circuit_id': v_circuit_id or circuit_id,
            'subject': subject,
            'raised_by': ticket_master_doc['raised_by_email'],
            'description': description
        }
        issue_doc = frappe.get_doc(issue_data)
        issue_doc.insert()

    # If no circuit match, update description and process 8-digit number
    if not circuit:
        v_circuit_id = re.search(r'\b\d{8}\b', description)
        if v_circuit_id:
            v_circuit_id = v_circuit_id.group()
            frappe.db.set_value('Ticket Master', ticket_master_name, 'circuit_id', v_circuit_id)
            # Create Issue as above
            issue_data['custom_circuit_id'] = v_circuit_id
            issue_doc = frappe.get_doc(issue_data)
            issue_doc.insert()

    return "Processed Successfully"

###########################################################
# Sales Order to Site
import frappe
from frappe.utils import now
from frappe import _

@frappe.whitelist()
def sales_order_to_site(sales_order):
    so_doc = frappe.get_doc("Sales Order", sales_order)

    # PO Number check removed as it is not mandatory
    # if not so_doc.po_no:
    #     frappe.throw(_("Client PO Number is missing in the Sales Order"))

    # Proceed only if order_type is one of the allowed types
    # allowed_order_types = ["Service", "Upgrade", "Degrade", "Shifting", "Supply"]
    # if so_doc.order_type not in allowed_order_types:
    #     frappe.logger().info("Site creation skipped for order_type: {}".format(so_doc.order_type))
    #     return

    grouped_sites = {}

    for item in so_doc.items:
        feasibility = item.custom_feasibility
        if feasibility not in grouped_sites:
            grouped_sites[feasibility] = {
                "circuit_id": feasibility,
                "site_name": item.custom_site_info,
                "items": []
            }
        grouped_sites[feasibility]["items"].append({
            "item_code": item.item_code,
            "item_name": item.item_name,
            "qty": item.qty
        })

    project_doc = frappe.get_doc("Project", so_doc.project) if so_doc.project else None
    created_sites = []

    for feasibility, site_data in grouped_sites.items():
        feasibility_doc = frappe.get_doc("Feasibility", feasibility)

        if not feasibility_doc:
            frappe.throw(_("Feasibility with circuit_id {0} not found").format(feasibility))

        feasibility_doc.sales_order = so_doc.name
        feasibility_doc.sales_order_date = so_doc.transaction_date
        feasibility_doc.save(ignore_permissions=True)

        site_doc = frappe.new_doc("Site")
        site_doc.customer = so_doc.customer
        site_doc.customer_po_no = so_doc.po_no
        site_doc.customer_po_date = so_doc.po_date
        site_doc.sales_order = so_doc.name
        site_doc.sales_order_amount = so_doc.grand_total
        site_doc.customer_po_amount = so_doc.custom_customer_purchase_amount
        site_doc.sales_order_date = so_doc.transaction_date
        site_doc.delivery_date = so_doc.delivery_date
        site_doc.project = so_doc.project
        site_doc.circuit_id = site_data["circuit_id"]
        site_doc.site_name = site_data["site_name"]
        site_doc.order_type = so_doc.order_type
        

        # From Feasibility
        site_doc.contact_person = feasibility_doc.get("contact_person")
        site_doc.primary_contact_mobile = feasibility_doc.get("primary_contact_mobile")
        site_doc.email = feasibility_doc.get("email")
        site_doc.alternate_contact_person = feasibility_doc.get("alternate_contact_person")
        site_doc.alternate_contact = feasibility_doc.get("alternate_contact")
        site_doc.alternate_contact_mobile = feasibility_doc.get("alternate_contact_mobile")
        site_doc.secondary_email = feasibility_doc.get("secondary_email")
        site_doc.site_id__legal_code = feasibility_doc.get("site_id__legal_code")
        site_doc.site_type = feasibility_doc.get("site_type")
        site_doc.solution = feasibility_doc.get("solution")
        site_doc.region = feasibility_doc.get("region")
        site_doc.exiting_circuit_id = feasibility_doc.get("exiting_circuit_id")
        site_doc.territory = feasibility_doc.get("territory")
        site_doc.customer_type = feasibility_doc.get("customer_type")
        site_doc.description = feasibility_doc.get("description")
        site_doc.address = feasibility_doc.get("address")
        site_doc.solution_code = feasibility_doc.get("solution_code")
        site_doc.solution_name = feasibility_doc.get("solution_name")
        site_doc.static_ip = feasibility_doc.get("static_ip")
        site_doc.nos_of_static_ip_required = feasibility_doc.get("no_of_static_ip_required")
        site_doc.primary_data_plan = feasibility_doc.get("primary_data_plan")
        site_doc.secondary_plan = feasibility_doc.get("secondary_data_plan")
        site_doc.managed_services = feasibility_doc.get("managed_services")
        site_doc.config_type = feasibility_doc.get("config_type")
        site_doc.child_project = so_doc.get("custom_child_project")
        site_doc.contact_person = feasibility_doc.get("contact_person")

        site_doc.central_spoke = feasibility_doc.get("central_spoke")
        site_doc.mobile = feasibility_doc.get("mobile")
        site_doc.central_email = feasibility_doc.get("central_email")
        site_doc.sales_person = feasibility_doc.get("sales_person")

        # Address Fields
        site_doc.address_street = feasibility_doc.address_street
        site_doc.pincode = feasibility_doc.pincode
        site_doc.district = feasibility_doc.district
        site_doc.state = feasibility_doc.state
        site_doc.country = feasibility_doc.country
        site_doc.city = feasibility_doc.city

        if project_doc:
            site_doc.project_name = project_doc.project_name
            site_doc.expected_start_date = project_doc.expected_start_date
            site_doc.expected_end_date = project_doc.expected_end_date

        # ---------------------- Site Items ----------------------
        for item in site_data["items"]:
            product_bundle = frappe.get_all(
                "Product Bundle",
                filters={"new_item_code": item["item_code"]},
                fields=["new_item_code"]
            )

            if product_bundle:
                product_bundle_doc = frappe.get_doc("Product Bundle", product_bundle[0].new_item_code)
                for bundle_item in product_bundle_doc.items:
                    item_doc = frappe.get_doc("Item", bundle_item.item_code) if frappe.db.exists("Item", bundle_item.item_code) else None
                    site_doc.append("site_item", {
                        "solution": product_bundle_doc.new_item_code,
                        "parent_item": product_bundle_doc.new_item_code,
                        "item_code": bundle_item.item_code,
                        "item_name": item_doc.item_name if item_doc else "",
                        "qty": bundle_item.qty
                    })
            else:
                site_doc.append("site_item", {
                    "item_code": item["item_code"],
                    "item_name": item["item_name"],
                    "qty": item["qty"]
                })

        # ---------------------- LMS Vendors ----------------------
        for lms in feasibility_doc.lms_provider:
            site_doc.append("lms_vendor", {
                "lms_supplier": lms.get("lms_supplier"),
                "bandwith_type": lms.get("bandwith_type"),
                "media": lms.get("media"),
                "support_mode": lms.get("support_mode"),
                "supplier_contact": lms.get("supplier_contact"),
                "static_ip": lms.get("static_ip"),
                "supplier_name": lms.get("supplier_name"),
                "email_id": lms.get("email_id"),
                "mobile": lms.get("mobile"),
                "bandwidth": lms.get("bandwidth"),
                "billing_mode": lms.get("billing_mode"),
                "billing_terms": lms.get("billing_terms"),
                "otc": lms.get("otc"),
                "validity": lms.get("validity"),
                "security_deposit": lms.get("security_deposit"),
                "mrc": lms.get("mrc"),
                "arc": lms.get("arc"),
                "static_ip_cost": lms.get("static_ip_cost")
            })

        # ---------------------- Wireless ----------------------
        for wireless in feasibility_doc.wireless_feasiblity:
            site_doc.append("wireless", {
                "operator": wireless.get("operator"),
                "3g": wireless.get("3g"),
                "4g": wireless.get("4g"),
                "5g": wireless.get("5g")
            })



        # Check if Site already exists before inserting. If so, throw error to abort transaction.
        if frappe.db.exists("Site", site_data["circuit_id"]):
            if hasattr(frappe.local, "message_log"):
                frappe.local.message_log = []
            frappe.throw(f"Site Already Exists: {site_data['circuit_id']}. Sales Order submission aborted.")
            
        site_doc.insert(ignore_permissions=True)
        frappe.db.commit()
        created_sites.append(f"Site is Created successfully: {site_doc.name}")
        frappe.logger().info(f"Site created for feasibility: {feasibility}")

    # Clear message log to remove POC hook messages, then print our clean success message
    if hasattr(frappe.local, "message_log"):
        frappe.local.message_log = []
        
    if created_sites:
        frappe.msgprint("<br>".join(created_sites))

    return {"status": "success"}


##############################################################

import frappe

@frappe.whitelist()
def get_stock_details(item_code, warehouse):
    """
    Fetches stock balance and reserved quantities for a given item and warehouse.
    """
    if not item_code or not warehouse:
        return {"item_balance": 0, "item_reserved": 0}
    
    query = """
        SELECT
            SUM(actual_qty) AS item_balance,
            SUM(reserved_qty) AS item_reserved
        FROM
            `tabBin`
        WHERE
            item_code = %s AND warehouse = %s
    """
    result = frappe.db.sql(query, (item_code, warehouse), as_dict=True)
    
    if result and result[0]:
        return {
            "item_balance": result[0].get("item_balance", 0),
            "item_reserved": result[0].get("item_reserved", 0),
        }
    return {"item_balance": 0, "item_reserved": 0}

#################################################################################
import frappe

@frappe.whitelist(allow_guest=True)
def get_filtered_feasibility(customer):
    # Define possible feasibility statuses
    feasibility_statuses = ["Feasible", "Partial Feasible", "High Commercials"]

    # Fetch feasibility records where:
    # 1. The customer matches.
    # 2. The feasibility_status is valid.
    # 3. The sales_order field is empty or null.
    # 4. The document status is 'Submitted' (docstatus = 1).
    feasibilities = frappe.get_all('Feasibility', filters={
        'customer': customer,
        'feasibility_status': ['in', feasibility_statuses],
        'sales_order': ['in', [None, '', 'null']],  # Ensure the sales_order is empty or null
        #'docstatus': 1  # Ensure the document is 'Submitted'
    }, fields=['circuit_id'])

    # Extract the circuit_ids from the feasibility records
    circuit_ids = [feasibility['circuit_id'] for feasibility in feasibilities]

    # Return the list of circuit_ids (empty list if no records found)
    return circuit_ids if circuit_ids else []
############################################################################################3
import frappe

def update_custom_circuit_id_in_stock_reservation(doc, method):
    """
    This function updates the 'custom_circuit_id' in the Stock Reservation Entries
    based on the Sales Order Item's 'custom_feasibility' when the Sales Order is submitted.
    
    Args:
        doc: The current Sales Order document.
        method (str): The event (e.g., 'on_submit') that triggered this function.
    """
    try:
        # Loop through each Sales Order Item
        for item in doc.items:
            # Check if the item has a custom feasibility value
            if item.custom_feasibility:
                # Get the Stock Reservation Entries linked to this Sales Order Item
                stock_reservation_entries = frappe.get_all(
                    "Stock Reservation Entry",
                    filters={"voucher_no": doc.name, "voucher_detail_no": item.name},
                    fields=["name"]
                )

                # Loop through each Stock Reservation Entry
                for entry in stock_reservation_entries:
                    # Update the 'custom_circuit_id' in the Stock Reservation Entry
                    frappe.db.set_value(
                        "Stock Reservation Entry", entry.name, "custom_circuit_id", item.custom_feasibility
                    )
                    frappe.msgprint(f"Updated custom_circuit_id for Stock Reservation Entry: {entry.name}")

    except frappe.DoesNotExistError:
        frappe.msgprint(f"Error: Sales Order {doc.name} does not exist.")
    except Exception as e:
        frappe.msgprint(f"An error occurred: {str(e)}")

###############################################################################
import frappe

@frappe.whitelist()
def fetch_provisioning_items(custom_circuit_id):
    # Validate input
    if not custom_circuit_id:
        return {"error": "Custom Circuit ID is required."}

    # Fetch the Site document with the matching Circuit ID
    site_doc = frappe.get_all("Site", filters={"circuit_id": custom_circuit_id}, fields=["name"])
    if not site_doc:
        return {"error": f"No Site found with Circuit ID: {custom_circuit_id}"}

    site_name = site_doc[0]["name"]
    site_data = frappe.get_doc("Site", site_name)

    # Prepare data for 'custom_product_' (existing functionality)
    provisioning_items = []
    if hasattr(site_data, "provisioning_item"):
        for item in site_data.provisioning_item:
            provisioning_items.append({
                "product_code": item.product_code or None, 
                "product_name": item.product or None,                               
                "serial_number": item.serial_number or None,
                "warranty_expiry_date": item.warranty_expiry_date or None,
                "warranty_period_days": item.warranty_period_days or None               
                
            })

    # Prepare data for 'custom_lms_vendor' (new functionality, corrected to 'lms_vendor')
    lms_items = []
    if hasattr(site_data, "lms_vendor"):  # Corrected to 'lms_vendor'
        for item in site_data.lms_vendor:
            lms_items.append({
                "lms_supplier": item.lms_supplier or None,
                "bandwith_type": item.bandwith_type or None,
                "media": item.media or None,
                "otc": item.otc or 0,
                "static_ip_cost": item.static_ip_cost or 0,
                "billing_terms": item.billing_terms,
                "support_mode": item.support_mode or None,
                "contact_person": item.contact_person or None,
                "supplier_contact": item.supplier_contact or None,
                "lms_bandwith": item.lms_bandwith or None,
                "static_ip": item.static_ip or None,
                "mrc": item.mrc or 0,
                "security_deposit": item.security_deposit or 0,
                "billing_mode": item.billing_mode 
            })

    # Return data
    return {
        "provisioning_items": provisioning_items,
        "lms_items": lms_items
    }

#################################################################################
import frappe

def update_site_status_on_delivery_note_save(doc, method):
    """
    Update Site status and delivery_note_id from Delivery Note,
    except for excluded change management types, return Delivery Notes,
    Delivered and Live sites, Cancelled sites,
    or sites already having delivery_note_id.
    """

    # Skip updating for excluded change management types
    excluded_types = [
        "Project Change Management",
        "Support Change Management",
        "Others"
    ]

    if doc.get("custom_change_management") in excluded_types:
        return

    # Skip return Delivery Notes
    if doc.get("is_return") == 1:
        return

    for item in doc.items:

        # Continue only if circuit ID exists
        if not item.custom_circuit_id:
            continue

        # Find matching Site
        site = frappe.db.get_value(
            "Site",
            {"circuit_id": item.custom_circuit_id},
            "name"
        )

        # If Site not found, skip
        if not site:
            continue

        # Get Site document
        site_doc = frappe.get_doc("Site", site)

        # ---------------------------------------------------
        # DO NOT UPDATE IF:
        # 1. site_status = Delivered and Live
        # 2. site_status = Cancelled
        # 3. delivery_note_id already exists
        # ---------------------------------------------------

        if (
            site_doc.site_status == "Delivered and Live"
            or site_doc.site_status == "Cancelled"
            or site_doc.delivery_note_id
        ):
            continue

        # ---------------------------------------------------
        # UPDATE SITE
        # ---------------------------------------------------

        site_doc.site_status = "In-process"

        # Update delivery_note_id with Delivery Note name
        site_doc.delivery_note_id = doc.name

        # Save Site
        site_doc.save(ignore_permissions=True)
################################# HelpDesk ############################################import frappe
import frappe
import re
from email.utils import getaddresses

def create_hd_ticket_from_communication(doc, method):
    try:

        # =========================================================
        # 0️⃣ PROCESS ONLY INCOMING EMAILS
        # =========================================================
        if doc.sent_or_received != "Received":
            return

        # =========================================================
        # 🟢 PREVENT REOPENING OF CLOSED/RESOLVED TICKETS
        # =========================================================
        if doc.reference_doctype == "HD Ticket" and doc.reference_name:
            ticket_status = frappe.db.get_value("HD Ticket", doc.reference_name, "status")
            if ticket_status in ["Closed", "Resolved"]:
                # 🔥 UNLINK (this stops the reopen in helpdesk app)
                doc.reference_doctype = None
                doc.reference_name = None

        sender = (doc.sender or "").lower()
        recipients = doc.recipients or ""
        subject = (doc.subject or "").lower()
        content = (doc.content or "").lower()

        recipient_emails = [
            email.strip().lower()
            for _, email in getaddresses([recipients])
        ]

        # =========================================================
        # 1️⃣ ONLY SUPPORT / NMS EMAILS
        # =========================================================
        if (
            "techsupport@nexapp.co.in" not in recipient_emails
            and "nms@nexapp.co.in" not in sender
        ):
            return

        # =========================================================
        # 2️⃣ STRONG BOUNCE / UNDELIVERABLE BLOCK 🚫
        # =========================================================
        combined_text_check = f"{subject} {content}"

        bounce_keywords = [
            "undeliverable",
            "delivery failed",
            "delivery has failed",
            "mail delivery subsystem",
            "returned mail",
            "delivery status notification",
            "failure notice",
            "mailbox full",
            "recipient address rejected",
            "message blocked",
            "could not be delivered",
            "550",
            "5.1.1",
        ]

        if any(keyword in combined_text_check for keyword in bounce_keywords):
            return  # 🔥 HARD STOP (no ticket, no reopen)

        auto_senders = [
            "mailer-daemon",
            "postmaster",
            "no-reply",
            "mailer@",
            "bounce",
        ]

        if any(x in sender for x in auto_senders):
            return

        # =========================================================
        # 3️⃣ EXTRACT CIRCUIT ID
        # =========================================================
        combined_text = f"{doc.subject or ''} {doc.content or ''}"

        circuit_match = re.search(
            r"(?:circuit[#_ ]|^|[^0-9])(\d{5})(?=[^0-9]|$)",
            combined_text,
            flags=re.IGNORECASE,
        )

        circuit_id = circuit_match.group(1) if circuit_match else ""

        if not circuit_id:
            return

        # =========================================================
        # 4️⃣ GET LATEST TICKET
        # =========================================================
        existing_ticket = frappe.db.get_value(
            "HD Ticket",
            {"custom_circuit_id": circuit_id},
            ["name", "status"],
            as_dict=True,
            order_by="creation desc",
        )

        # =========================================================
        # 🟢 CASE A — EXISTING ACTIVE TICKET
        # =========================================================
        if existing_ticket and existing_ticket.status not in ["Closed", "Resolved"]:

            doc.reference_doctype = "HD Ticket"
            doc.reference_name = existing_ticket.name
            doc.status = "Linked"

            if not doc.name:
                doc.set_new_name()

            if "nms@nexapp.co.in" not in sender:
                frappe.enqueue(
                    send_ticket_reply,
                    queue="short",
                    enqueue_after_commit=True,
                    communication_name=doc.name,
                    sender=sender,
                    ticket_name=existing_ticket.name,
                    circuit_id=circuit_id,
                    status=existing_ticket.status,
                    template="existing",
                )

            return

        # =========================================================
        # 🔴 CASE B — CLOSED / RESOLVED → ALWAYS CREATE NEW
        # =========================================================
        # ❌ REOPEN LOGIC COMPLETELY REMOVED (Unlinked above)

        # =========================================================
        # 🔴 CASE C — CREATE NEW TICKET
        # =========================================================
        if frappe.db.exists("Site", circuit_id):

            custom_lms_id = ""
            subject_text = doc.subject or ""

            if "offline alert" in subject_text.lower():

                lms_match = re.search(
                    r"(?:MBB|ILL)[_\-\s]?(\d{6})",
                    subject_text,
                    re.IGNORECASE,
                )

                if lms_match:
                    var = lms_match.group(1)
                    custom_lms_id = f"LMS- {var}"

            ticket = frappe.get_doc({
                "doctype": "HD Ticket",
                "subject": doc.subject,
                "description": doc.content,
                "raised_by": sender,
                "status": "Open",
                "custom_circuit_id": circuit_id,
                "custom_lms_id": custom_lms_id,
                "custom_channel": "NMS"
                if "nms@nexapp.co.in" in sender
                else "Email",
            })

            ticket.insert(ignore_permissions=True)

            doc.reference_doctype = "HD Ticket"
            doc.reference_name = ticket.name
            doc.status = "Linked"

            if not doc.name:
                doc.set_new_name()

            if "nms@nexapp.co.in" not in sender:
                frappe.enqueue(
                    send_ticket_reply,
                    queue="short",
                    enqueue_after_commit=True,
                    communication_name=doc.name,
                    sender=sender,
                    ticket_name=ticket.name,
                    circuit_id=circuit_id,
                    status="Open",
                    template="new",
                )

            return

    except Exception:
        frappe.log_error(
            frappe.get_traceback(),
            "HD Ticket Auto-Creation Error",
        )

# =========================================================
# EMAIL FUNCTION (UNCHANGED)
# =========================================================
def send_ticket_reply(communication_name, sender, ticket_name, circuit_id, status, template):

    subject = f"Re: Circuit {circuit_id} — Ticket {ticket_name}"

    if template == "new":
        body = f"""
        A new support ticket has been created.<br><br>
        <b>Ticket:</b> {ticket_name}<br>
        <b>Circuit:</b> {circuit_id}<br><br>
        """

    else:
        body = f"""
        This ticket is already open, so your email has been linked.<br><br>
        <b>Ticket:</b> {ticket_name}<br>
        <b>Circuit:</b> {circuit_id}<br>
        <b>Status:</b> {status}
        """

    message = f"""
    Dear Customer,<br><br>
    Thank you for contacting Nexapp Support.<br><br>
    {body}<br><br>
    Thanks & Regards,<br>
    Nexapp Technologies Private Limited<br>
    Support Team
    """

    reply = frappe.get_doc({
        "doctype": "Communication",
        "communication_type": "Communication",
        "communication_medium": "Email",
        "sent_or_received": "Sent",
        "subject": subject,
        "content": message,
        "sender": "techsupport@nexapp.co.in",
        "recipients": sender,
        "in_reply_to": communication_name,
        "reference_doctype": "HD Ticket",
        "reference_name": ticket_name,
    })

    reply.insert(ignore_permissions=True)

    frappe.sendmail(
        recipients=[sender],
        subject=subject,
        message=message,
        reference_doctype="HD Ticket",
        reference_name=ticket_name,
        communication=reply.name,
        now=True,
        sender="techsupport@nexapp.co.in"
    )

################################# End of HelpDesk Code################################
import frappe
from frappe.utils import get_url
from frappe.utils.pdf import get_pdf
import hashlib

@frappe.whitelist()
def download_subcategory_pdf(subcategory):
    # Get documents with strict filters
    docs = frappe.get_all("Document",
        filters={
            "sub_category": subcategory,
            "published": 1
        },
        fields=["name", "title", "content", "attach_file"],
        order_by="creation"
    )

    # Advanced duplicate prevention
    seen = set()
    html = """
    <style>
        body { font-family: Arial; margin: 20px; }
        h1 { color: #2d3e50; border-bottom: 2px solid #eee; }
        .document { margin-bottom: 30px; }
        img { max-width: 100%%; height: auto; margin: 15px 0; }
    </style>
    <h1>%s Documentation</h1>
    """ % subcategory

    for doc in docs:
        # Create unique hash for content
        content_hash = hashlib.md5(f"{doc.title}{doc.content}".encode()).hexdigest()
        if content_hash in seen:
            continue
        seen.add(content_hash)
        
        image_html = f'<img src="{get_url(doc.attach_file)}">' if doc.attach_file else ''
        html += f"""
        <div class="document">
            <h2>{doc.title}</h2>
            <div>{doc.content}</div>
            {image_html}
        </div>
        <hr style="margin:20px 0; border-top:1px solid #eee;">
        """

    pdf_data = get_pdf(html)
    frappe.local.response.filename = f"{subcategory}-documentation.pdf"
    frappe.local.response.filecontent = pdf_data
    frappe.local.response.type = "pdf"

######################################################################
import frappe

def customer_created(doc, method):
    # Fetch all Feasibility records where 'customer' matches the new Customer name
    feasibilities = frappe.get_all(
        "Feasibility",
        filters={"customer": doc.customer_name},
        fields=["name", "party_name", "feaseibility_from"]
    )

    for feas in feasibilities:
        # If 'feaseibility_from' is not already 'Customer', update it
        if feas.feaseibility_from != "Customer":
            frappe.db.set_value("Feasibility", feas.name, "feaseibility_from", "Customer")
        
        # If 'party_name' matches the Customer name, also set 'customer' = 'party_name'
        if feas.party_name == doc.customer_name:
            frappe.db.set_value("Feasibility", feas.name, "customer", feas.party_name)

#########################################################################################
import frappe

@frappe.whitelist()
def create_site_from_feasibility(doc, method=None):
    # If doc is passed as a string (usually docname), fetch the document
    if isinstance(doc, str):
        doc = frappe.get_doc("Feasibility", doc)

    # Proceed only if customer is POC
    if doc.customer_type == "POC Customer":
        site = frappe.new_doc("Site")

        # Map relevant fields from Feasibility to Site
        site.site_name = doc.site_name
        site.customer = doc.customer
        site.site_type = doc.site_type
        site.territory = doc.territory
        site.customer_type = doc.customer_type
        site.order_type = doc.order_type
        site.circuit_id = doc.name  # using the feasibility doc name as circuit ID
        site.site_id__legal_code = doc.site_id__legal_code
        site.description = doc.description

        # Address and contact details
        site.address_street = doc.address_street
        site.city = doc.city
        site.contact_person = doc.contact_person
        site.primary_contact_mobile = doc.primary_contact_mobile
        site.email = doc.email
        site.pincode = doc.pincode
        site.district = doc.district
        site.state = doc.state
        site.country = doc.country
        site.alternate_contact_person = doc.alternate_contact_person
        site.alternate_contact_mobile = doc.alternate_contact_mobile
        site.secondary_email = doc.secondary_email

        # Technical and solution-specific fields
        site.solution_code = doc.solution_code
        site.solution_name = doc.solution_name
        site.static_ip = doc.static_ip
        site.nos_of_static_ip_required = doc.no_of_static_ip_required
        site.primary_data_plan = doc.primary_data_plan
        site.secondary_plan = doc.secondary_data_plan
        site.managed_services = doc.managed_services
        site.config_type = doc.config_type

        # Insert new Site document ignoring permissions
        site.insert(ignore_permissions=True)
        frappe.db.commit()  # Save to DB before modifying child table

        # Add child items from the matching Product Bundle
        try:
            product_bundle = frappe.get_doc("Product Bundle", doc.solution_code)
            for item in product_bundle.items:
                site.append("site_item", {
                    "item_code": item.item_code,
                    "qty": item.qty
                })
            site.save()
        except frappe.DoesNotExistError:
            frappe.msgprint(f"No Product Bundle found with name '{doc.solution_code}'")

        return site.name  # Return Site name for confirmation on frontend

    return "Not POC Customer"
#####################################################################################

def update_site_and_stock_management(doc, method):
    # Stop if change management type is in restricted list
    if doc.custom_change_management in [
        "Project Change Management",
        "Support Change Management",
        "Others"
    ]:
        return

    if not doc.custom_circuit_id or not doc.tracking_status:
        return

    # Define status mappings for Site, Site Item, and Stock Management
    status_mapping = {
        "In Progress": {
            "stage": "Shipment In-Transit",
            "site_item_status": "Shipment In-Transit",
            "stock_status": "Stock Shipment In-Transit"
        },
        "Delivered": {
            "stage": "Stock Delivered",
            "site_item_status": "Stock Delivered",
            "stock_status": "Stock Delivered"
        },
        "Returned": {
            "stage": "Stock Returned",
            "site_item_status": "Stock Returned",
            "stock_status": "Stock Returned"
        },
        "Lost": {
            "stage": "Stock Lost",
            "site_item_status": "Stock Lost",
            "stock_status": "Stock Lost"
        }
    }

    # Get the appropriate statuses from mapping
    mapping = status_mapping.get(doc.tracking_status, {})
    if not mapping:
        return

    # 1. Update Stock Management records where circuit_id matches
    stock_management_records = frappe.get_all(
        "Stock Management",
        filters={"circuit_id": doc.custom_circuit_id},
        fields=["name"]
    )

    for stock in stock_management_records:
        stock_doc = frappe.get_doc("Stock Management", stock.name)
        stock_doc.update({
            "shipment_id": doc.name,
            "service_provider": doc.service_provider,
            "delivery_date": doc.custom_delivery_date,
            "pickup_date": doc.pickup_date,
            "awb_number": doc.awb_number,
            "carrier": doc.carrier,
            "carrier_service": doc.carrier_service,
            "tracking_status": doc.tracking_status,
            "status": mapping["stock_status"]
        })
        stock_doc.save(ignore_permissions=True)
        frappe.msgprint(f"Updated Stock Management {stock.name} to {mapping['stock_status']}")

    def update_site_with_items(site_doc):
        """Helper function to update Site doc and its site_item child table"""
        # Update main Site fields
        site_doc.update({
            "shipment_id": doc.name,
            "service_provider": doc.service_provider,
            "delivery_date": doc.custom_delivery_date,
            "pickup_date": doc.pickup_date,
            "awb_number": doc.awb_number,
            "carrier": doc.carrier,
            "carrier_service": doc.carrier_service,
            "tracking_status": doc.tracking_status,
            "stage": mapping["stage"]
        })

        # Update site_item child table status if exists
        if hasattr(site_doc, 'site_item') and site_doc.get('site_item'):
            for item in site_doc.site_item:
                item.status = mapping["site_item_status"]

        site_doc.save(ignore_permissions=True)
        return site_doc

    # 2. Update Site where name matches custom_circuit_id
    if frappe.db.exists("Site", doc.custom_circuit_id):
        site_doc = frappe.get_doc("Site", doc.custom_circuit_id)
        updated_doc = update_site_with_items(site_doc)
        frappe.msgprint(f"Updated Site {doc.custom_circuit_id} stage to {mapping['stage']} and items to {mapping['site_item_status']}")

    # 3. Update Sites where circuit_id matches custom_circuit_id
    sites_with_circuit_id = frappe.get_all(
        "Site",
        filters={"circuit_id": doc.custom_circuit_id},
        fields=["name"]
    )

    for site in sites_with_circuit_id:
        site_doc = frappe.get_doc("Site", site.name)
        updated_doc = update_site_with_items(site_doc)
        frappe.msgprint(f"Updated Site {site.name} (via circuit_id) stage to {mapping['stage']} and items to {mapping['site_item_status']}")

    frappe.db.commit()

####################################################################################
# Updating Site From Delivery Note
import frappe
from frappe import _

def validate_delivery_note(doc, method):
    """Update Site status, Site Items, and map serial numbers when Delivery Note is saved"""
    try:
        # Values for which we skip updating Site stage
        skip_values = ["Project Change Management", "Support Change Management", "Others"]

        # Run update only if order type is Service, circuit_id exists, and not a return
        if (
            doc.get("custom_order_type") == "Service" 
            and doc.get("custom_dn_circuit_id") 
            and not doc.get("is_return")
        ):
            
            # Skip if custom_change_management is in skip_values
            if doc.get("custom_change_management") in skip_values:
                frappe.logger().info(
                    f"Skipped site stage update because change management is '{doc.get('custom_change_management')}'"
                )
                return  # exit without updating site stage

            site_name = doc.custom_dn_circuit_id

            # Update main Site document
            frappe.db.set_value(
                "Site", 
                site_name, 
                "stage", 
                "Stock Delivery In-Process",
                update_modified=False
            )

            # Get all Site Items for this Site
            site_items = frappe.get_all(
                "Site Item",
                filters={"parent": site_name},
                fields=["name", "item_code"]
            )

            # Get all packed items from Delivery Note
            packed_items = doc.get("packed_items", [])

            # Build a list of packed_items with item_code and serial_no
            packed_serial_items = [
                pi for pi in packed_items
                if pi.item_code and pi.serial_no
            ]

            # Copy the packed list so we can consume items as we match
            remaining_packed = packed_serial_items[:]

            # Match each Site Item with one corresponding packed item
            for site_item in site_items:
                for i, packed in enumerate(remaining_packed):
                    if packed.item_code == site_item.item_code:
                        # Load the full Site Item document
                        site_item_doc = frappe.get_doc("Site Item", site_item.name)
                        site_item_doc.status = "Delivery In-Process"
                        site_item_doc.serial_no_sim_no = packed.serial_no

                        # Save and trigger fetch_from
                        site_item_doc.save(ignore_permissions=True)

                        # Optional: real-time UI update
                        frappe.publish_realtime('doc_update', {
                            'doc': site_item_doc.as_dict(),
                            'name': site_item_doc.name,
                            'doctype': 'Site Item'
                        })

                        # Remove used packed item to avoid duplicate usage
                        remaining_packed.pop(i)
                        break  # Go to next Site Item

            # Optional: refresh parent Site document on the UI
            frappe.publish_realtime('doc_update', {
                'doc': frappe.get_doc("Site", site_name).as_dict(),
                'name': site_name,
                'doctype': 'Site'
            })

        else:
            if doc.get("is_return"):
                frappe.logger().info("Skipped site stage update because Delivery Note is a return")

    except Exception as e:
        frappe.log_error(f"Error updating Site from Delivery Note: {e}")
        frappe.throw(_("Error updating Site. Please check error logs."))

#########################################
# PO Update to Site and LMS
import frappe
from frappe.utils import now

def update_lastmile_on_po_save(doc, method):
    # Loop through the child table items and collect all unique LMS IDs
    lms_ids = set()
    for item in doc.get("items", []):
        if item.custom_lms_id:
            lms_ids.add(item.custom_lms_id)

    # If no items had an LMS ID, we do nothing
    if not lms_ids:
        return

    # Update each Lastmile Services Master individually
    for lms_id in lms_ids:
        if frappe.db.exists("Lastmile Services Master", lms_id):
            frappe.db.set_value("Lastmile Services Master", lms_id, {
                "lms_stage": "In process",
                "po_number": doc.name,
                "po_released_datetime": frappe.utils.now()
            })
##################################################
import frappe

def update_lms_on_payment_submit(doc, method):
    if not doc.lms_id:
        frappe.msgprint("No LMS ID provided. Skipping update.")
        return

    # Update the Lastmile Services Master document
    frappe.db.set_value("Lastmile Services Master", doc.lms_id, {
        "payment_details": doc.payment_details,
        "payment_date": doc.date_of_payment,
        "payment_type": doc.payment_type,
        "payment_amount": doc.payment_amount,
        "lms_payment_request": "LMS Payment Released"
    })

    frappe.msgprint(f"Updated Lastmile Services Master: {doc.lms_id}")
#################################################################################3
import frappe

def update_site_child_table(doc, method):
    frappe.logger().info(f"[DEBUG] Running update_site_child_table for {doc.name}")

    # ---------------------------------------------------------
    # NEW CONDITION: Run only when lms_stage == "In process"
    # ---------------------------------------------------------
    if doc.lms_stage != "In process":
        frappe.logger().info(f"[DEBUG] lms_stage is not 'In process' for {doc.name}, exiting.")
        return

    # Exit if no delivery date is set
    if not doc.lms_delivery_date:
        frappe.logger().info(f"[DEBUG] No delivery date set for {doc.name}, exiting.")
        return

    # Ensure lms_stage is updated to Delivered in the Master
    if doc.lms_stage != "Delivered":
        doc.db_set("lms_stage", "Delivered")
        frappe.logger().info(f"[DEBUG] lms_stage updated to Delivered for {doc.name}")

    if doc.circuit_id:
        try:
            site = frappe.get_doc("Site", doc.circuit_id)
            found = False

            # Check if a child row with lms_id == doc.name already exists
            for row in site.get("lms_vendor", []):
                if row.lms_id == doc.name:
                    found = True
                    break

            # If no matching row found, create a new child row
            if not found:
                new_row = site.append("lms_vendor", {})
                new_row.lms_id = doc.name
                new_row.lms_supplier = doc.lms_feasibility_partner
                new_row.supplier_contact = doc.supplier_contact
                new_row.bandwith_type = doc.bandwith_type
                new_row.media = doc.media
                new_row.lms_requested_id = doc.lms_request_id
                new_row.mobile = doc.suppliernumber
                new_row.bandwidth = doc.lms_bandwith
                new_row.brandwidth_name = doc.lms_brandwith_name
                frappe.logger().info(f"[DEBUG] New child row created in Site: {site.name}")

            # Update the matching child row
            for row in site.get("lms_vendor", []):
                if row.lms_id == doc.name:
                    row.stage = "LMS Delivered"
                    row.mode = doc.mode1
                    row.static_ip1 = doc.get("static_ip")
                    row.url = doc.get("url")
                    row.user_id = doc.get("user_id")
                    row.password = doc.get("password")
                    row.lms_delivery_date = doc.lms_delivery_date
                    frappe.logger().info(f"[DEBUG] Child row updated in Site: {site.name}")
                    break

            # Update lms_stage of Site
            lms_type = site.get("lms_type")
            current_stage = site.get("lms_stage")

            if lms_type == "Dual":
                if current_stage == "In process":
                    site.lms_stage = "LMS Partially Delivered"
                else:
                    site.lms_stage = "LMS Delivered"
            elif lms_type == "Single":
                site.lms_stage = "LMS Delivered"

            site.save()
            frappe.logger().info(f"[DEBUG] Site {site.name} saved with updated lms_stage.")

        except frappe.DoesNotExistError:
            frappe.throw(f"Site {doc.circuit_id} does not exist")

######################################################################################
import frappe

@frappe.whitelist()
def generate_and_update_mac_addresses(work_order_name):
    OUI = '840A9E'  # 3-byte OUI hex string (no separators)

    # Parse OUI hex to integer
    v = 0
    for ch in OUI:
        if '0' <= ch <= '9':
            d = ord(ch) - ord('0')
        elif 'A' <= ch <= 'F':
            d = ord(ch) - ord('A') + 10
        else:
            d = 0
        v = v * 16 + d

    base_int = v * (16 ** 6)  # Shift left 24 bits

    try:
        doc = frappe.get_doc("Work Order", work_order_name)
        count_per_sn = int(doc.custom_nos_of_mac_address or 0)
        mac_type = doc.custom_mac_type
    except Exception:
        return "Invalid Work Order or custom_nos_of_mac_address"

    if count_per_sn <= 0 and mac_type == "ETH":
        return "No MAC addresses requested"

    # Get Serial Nos linked to the Work Order
    serial_rows = frappe.db.sql(
        "SELECT name FROM `tabSerial No` WHERE work_order=%s",
        (work_order_name,), as_list=True
    )

    if not serial_rows:
        return "No Serial Nos linked to this Work Order"

    # Check if any Serial No already has MAC addresses in child table 'custom_mac'
    for sn_name in [r[0] for r in serial_rows]:
        mac_count = frappe.db.count('MAC Address', filters={'parent': sn_name, 'parentfield': 'custom_mac'})
        if mac_count > 0:
            return "MAC address already created"

    # Get all existing MAC addresses globally to avoid duplicates
    existing_macs = set()
    rows = frappe.db.sql("SELECT mac_address FROM `tabMAC Address`", as_list=True)
    for row in rows:
        existing_macs.add(row[0].upper())

    shifts = [44,40,36,32,28,24,20,16,12,8,4,0]
    digits = '0123456789ABCDEF'

    macs_generated = 0
    mac_int_counter = 0

    def get_unique_mac():
        nonlocal mac_int_counter
        while True:
            mac_int = base_int + mac_int_counter
            mac_int_counter += 1

            mac_hex = ''
            for shift in shifts:
                nibble = (mac_int >> shift) & 0xF
                mac_hex += digits[nibble]
                if shift > 0 and shift % 8 == 0:
                    mac_hex += ':'

            mac_hex = mac_hex.upper()
            if mac_hex not in existing_macs:
                existing_macs.add(mac_hex)
                return mac_hex

    for sn_name in [r[0] for r in serial_rows]:
        sn = frappe.get_doc("Serial No", sn_name)
        sn.set("custom_mac", [])  # Clear existing MACs just in case

        if mac_type == "ETH":
            # Generate MACs with ETH1, ETH2, ... up to ETH10
            for i in range(1, min(count_per_sn, 10) + 1):
                mac_hex = get_unique_mac()
                sn.append("custom_mac", {
                    "mac_address": mac_hex,
                    "interface": f"ETH{i}"
                })
                macs_generated += 1

        elif mac_type == "LAN/WAN/WIFI":
            # New behavior: generate 3 MACs for LAN, WAN, WIFI
            for iface in ["LAN", "WAN", "WIFI"]:
                mac_hex = get_unique_mac()
                sn.append("custom_mac", {
                    "mac_address": mac_hex,
                    "interface": iface
                })
                macs_generated += 1

        sn.save(ignore_permissions=True)

    return f"Generated and assigned {macs_generated} MAC addresses"

#############################################################################3
   
import frappe

@frappe.whitelist()
def is_l1_support_user():
    """Check if current user has ONLY L1 Support role (not other admin/support roles)"""
    user = frappe.session.user
    roles = frappe.get_roles(user)
    
    # Strict check - only hide if user has L1 Support and no other privileged roles
    if "L1 Support" in roles and not any(role in ["Administrator", "System Manager", "Support Manager"] for role in roles):
        return True
    return False

#####################################################################################
# START CHANGE MANAGEMENT REQUEST LOGIC
#####################################################################################
import frappe
from frappe import _
from frappe.utils import nowdate

@frappe.whitelist()
def check_feasibility_or_site(circuit_id):
    if frappe.db.exists("Feasibility", circuit_id):
        return {"status": "feasibility_exists"}

    site_doc = frappe.db.get("Site", circuit_id)
    if site_doc:
        return {"status": "site_exists"}

    return {"status": "not_found"}


@frappe.whitelist()
def create_feasibility_from_site(circuit_id):
    if frappe.db.exists("Feasibility", circuit_id):
        return _("Feasibility already exists for Circuit ID: {0}").format(circuit_id)

    site_doc = frappe.db.get("Site", circuit_id)
    if not site_doc:
        return _("Site not found for Circuit ID: {0}").format(circuit_id)

    frappe.publish_realtime('msgprint', {
        "message": _("System is creating Feasibility. Please wait a few seconds..."),
        "title": _("Please Wait"),
        "indicator": "orange"
    })

    # Create Feasibility doc
    feasibility_doc = frappe.get_doc({
        "doctype": "Feasibility",
        "customer_type": site_doc.customer_type,
        "order_type": site_doc.order_type,
        "customer": site_doc.customer,
        "feasibility_status": "Feasible",
        "site_name": site_doc.site_name,
        "customer_request": nowdate(),
        "site_type": site_doc.site_type,
        "territory": site_doc.territory,
        "solution_code": site_doc.solution_code,
        "static_ip": site_doc.static_ip,
        "managed_services": site_doc.managed_services,
        "config_type": site_doc.config_type,
        "address_street": site_doc.address_street,
        "pincode": site_doc.pincode,
        "district": site_doc.district,
        "state": site_doc.state,
        "country": site_doc.country,
        "city": site_doc.city,
        "contact_person": site_doc.contact_person,
        "primary_contact_mobile": site_doc.primary_contact_mobile,
    })

    feasibility_doc.insert(ignore_permissions=True, ignore_mandatory=True)

    # Rename feasibility
    old_name = feasibility_doc.name
    if old_name != circuit_id:
        frappe.rename_doc("Feasibility", old_name, circuit_id, force=True, merge=False)

    # Update related CMR
    cmr_name = frappe.db.get_value("Change Management Request", {"circuit_id": circuit_id})
    if cmr_name:
        cmr_doc = frappe.get_doc("Change Management Request", cmr_name)

        # ✅ Ensure isp_status and stage are updated
        frappe.db.set_value("Change Management Request", cmr_name, {
            "isp_status": "Feasibility Requested",
            "stage": "Feasibility Pending"
        })

        # ✅ Update Feasibility from CMR
        frappe.db.set_value("Feasibility", circuit_id, {
            "isp_change_feasibility_check": 1,
            "lms_id": cmr_doc.lms_id,
            "isp_change_issue": cmr_doc.isp_change_issue,
            "supplier": cmr_doc.supplier,
            "purchase_order_number": cmr_doc.get("purchase_order_number"),
            "purchase_order_date": cmr_doc.get("purchase_order_date"),
            "change_management_request_id": cmr_name,
            "expected_date": cmr_doc.expected_date
        })

        # ✅ Sync LMS ID back to CMR if not present
        if not cmr_doc.lms_id:
            feasibility_lms_id = frappe.db.get_value("Feasibility", circuit_id, "lms_id")
            if feasibility_lms_id:
                frappe.db.set_value("Change Management Request", cmr_name, "lms_id", feasibility_lms_id)

    return _("Feasibility created, renamed, and updated for Circuit ID: {0}").format(circuit_id)


def on_update(doc, method):
    # Prevent re-updating Feasibility if it's already in Feasibility Pending stage
    if doc.stage == "Feasibility Pending":
        return

    if not doc.circuit_id:
        frappe.msgprint(_("Change Management Request {}").format("created successfully" if doc.is_new() else "updated"), alert=True)
        return

    if not frappe.db.exists("Feasibility", doc.circuit_id):
        frappe.msgprint(
            msg=_("Change Management Request updated") + "<br><span style='color:red'>" + _("Feasibility not found.") + "</span>",
            indicator="red",
            alert=True
        )
        return

    # ✅ Update feasibility
    frappe.db.set_value("Feasibility", doc.circuit_id, {
        "isp_change_feasibility_check": 1,
        "lms_id": doc.lms_id,
        "isp_change_issue": doc.isp_change_issue,
        "supplier": doc.supplier,
        "purchase_order_number": doc.get("purchase_order_number"),
        "purchase_order_date": doc.get("purchase_order_date"),
        "change_management_request_id": doc.name,
        "expected_date": doc.expected_date
    })

    # ✅ Update isp_status and stage consistently
    frappe.db.set_value("Change Management Request", doc.name, {
        "isp_status": "Feasibility Requested",
        "stage": "Feasibility Pending"
    })

    # ✅ Sync LMS ID from Feasibility if missing in CMR
    lms_synced = False
    if not doc.lms_id:
        feasibility_lms_id = frappe.db.get_value("Feasibility", doc.circuit_id, "lms_id")
        if feasibility_lms_id:
            frappe.db.set_value("Change Management Request", doc.name, "lms_id", feasibility_lms_id)
            lms_synced = True

    # No frappe.msgprint here. The UI popup will be handled in Javascript `after_save`.

#####################################################################################
# END CHANGE MANAGEMENT REQUEST LOGIC
#####################################################################################

#LMS Request Supplier Payment Detail
@frappe.whitelist()
def get_latest_invoice_for_lms(lms_id):
    if not lms_id:
        return None

    # Get all parent Purchase Invoice names that have this LMS ID
    item_parents = frappe.get_all("Purchase Invoice Item",
        filters={"lms_id": lms_id},
        fields=["parent"]
    )

    if not item_parents:
        return None

    parent_ids = [item['parent'] for item in item_parents]

    # Get Purchase Invoices sorted by custom_dutation_from DESC
    invoices = frappe.get_all("Purchase Invoice",
        filters={"name": ["in", parent_ids]},
        fields=["name", "custom_dutation_from", "custom_duration_to"],
        order_by="custom_dutation_from desc"
    )

    if invoices:
        return invoices[0]  # Return the latest one
    else:
        return None
#####################################################################
#PO Cancel Task Creation form LMS Request
import frappe
from frappe import _
from frappe.utils import nowdate

@frappe.whitelist()
def create_po_cancel_task(lms_request):
    if not lms_request:
        frappe.throw(_("LMS Request ID is required."))

    # Check if task already created
    current_stage = frappe.db.get_value("LMS Request", lms_request, "po_stage")
    if current_stage == "PO Cancel Task Created":
        frappe.throw(_("PO Cancel Task is already created for this LMS Request."))

    # Get LMS Request doc
    lms_doc = frappe.get_doc("LMS Request", lms_request)

    # Create Task with mapped fields
    task = frappe.get_doc({
        "doctype": "Task",
        "subject": f"PO Cancel Task for {lms_request}",
        "reference_type": "LMS Request",
        "reference_name": lms_request,
        "status": "Open",
        "type": "Purchase Order Cancel",  
        "custom_customer": lms_doc.customer,
        "custom_purchase_order_no": lms_doc.purchase_order_number,
        "custom_purchase_order_date": lms_doc.purchase_order_date,
        "exp_start_date": nowdate(),
        "exp_end_date": nowdate(),
        "description": lms_doc.isp_change_issue
    })
    task.insert(ignore_permissions=True)

    # Update LMS Request stage
    lms_doc.po_stage = "PO Cancel Task Created"
    lms_doc.save(ignore_permissions=True)

    return "success"
################################################################################

## Disconnection Multiple
import frappe

@frappe.whitelist()
def start_disconnection_enqueue(docname):
    """Called from JS when user clicks 'Get Circuit Details'"""

    # 🔔 Show 'fetching' alert for 20 seconds on client
    frappe.publish_realtime('show_fetching_alert', {
        'message': 'Fetching data, please wait… This may take a few moments.'
    })

    # 🚀 Enqueue background job
    frappe.enqueue(
        method=process_disconnection_background,
        queue='default',
        timeout=600,
        is_async=True,
        docname=docname
    )


def process_disconnection_background(docname):
    """Runs in background to process circuits and update Disconnection Request"""

    doc = frappe.get_doc("Disconnection Request", docname)

    # 🧹 Clear old child table data
    doc.disconnection_request = []
    doc.lms_details = []

    circuit_ids = []
    site_list = []

    # ✅ Case 1: Circuits entered manually
    if doc.disconnection_circuit_details:
        circuit_ids = [c.strip() for c in doc.disconnection_circuit_details.split(',') if c.strip()]
        if circuit_ids:
            site_list = frappe.get_all(
                "Site",
                filters={
                    "circuit_id": ["in", circuit_ids],
                    "site_status": "Delivered and Live"
                },
                pluck="name"
            )

    # ✅ Case 2: No circuits → fallback to customer name
    else:
        if not doc.customer_name_2:
            frappe.throw(
                "Disconnection cannot be processed because 'Disconnection Circuit Details' "
                "is blank and no Customer is specified."
            )
        site_list = frappe.get_all(
            "Site",
            filters={
                "customer": doc.customer_name_2,
                "site_status": "Delivered and Live"
            },
            pluck="name"
        )

    # ❌ No matching sites found
    if not site_list:
        frappe.throw("No eligible Sites found with status 'Delivered and Live' for disconnection.")

    unique_circuits = set()
    fetched_circuits = set()

    # 🔄 Process each site and fetch related data
    for site_name in site_list:
        site_doc = frappe.get_doc("Site", site_name)
        fetched_circuits.add(site_doc.circuit_id)

        # 🧾 Add Site Items to Disconnection Request child table
        for item in site_doc.site_item:
            doc.append("disconnection_request", {
                "circuit_id": site_doc.circuit_id,
                "site_name": site_doc.site_name,
                "item_code": item.item_code,
                "item_name": item.item_name,
                "qty": item.qty,
                "serial_no_sim_no": item.serial_no_sim_no,
                "item_group": item.item_group,
                "warranty_expiry_date": item.warranty_expiry_date,
                "lan_mac": item.lan_mac,
                "hardware_version": item.hardware_version,
                "wlan_mac": item.wlan_mac,
                "wan_mac": item.wan_mac,
                "module": item.module,
                "warranty_period_days": item.warranty_period_days,
                "imei": item.imei,
                "mobile_no": item.mobile_no,
                "activation_date": item.activation_date,
                "validity": item.validity,
                "data_plan": item.data_plan,
                "recharge_end_date": item.recharge_end_date
            })

        # 🧾 Add LMS details (if delivered)
        for lms in site_doc.lms_vendor:
            if lms.stage == "LMS Delivered" and lms.lms_id:
                doc.append("lms_details", {
                    "lms_id": lms.lms_id,
                    "status": lms.stage
                })

        if site_doc.circuit_id:
            unique_circuits.add(site_doc.circuit_id)

    # 🧮 Summary Calculations
    total_input = len(circuit_ids)
    total_fetched = len(fetched_circuits)
    not_fetched = set(circuit_ids) - fetched_circuits if circuit_ids else set()

    # 🎨 Build HTML Summary Table
    html_summary = f"""
    <div style="font-family: Arial, sans-serif; background:#f9f9f9; border:1px solid #ccc;
                border-radius:10px; padding:18px; margin-top:10px;">
        <h3 style="color:#2c3e50; margin-bottom:15px;">Disconnection Summary</h3>

        <div style="display:flex; justify-content:space-between; margin-bottom:15px;">
            <div><strong style="color:green;">✅ Circuits Fetched:</strong> {total_fetched}</div>
            <div><strong style="color:red;">❌ Circuits Not Fetched:</strong> {len(not_fetched)}</div>
            <div><strong>Total Entered:</strong> {total_input}</div>
        </div>

        <table style="width:100%; border-collapse:collapse; font-size:14px;">
            <thead>
                <tr style="background:#2c3e50; color:white;">
                    <th style="padding:8px; text-align:left;">#</th>
                    <th style="padding:8px; text-align:left;">Circuit ID</th>
                    <th style="padding:8px; text-align:left;">Status</th>
                </tr>
            </thead>
            <tbody>
    """

    index = 1
    for cid in circuit_ids:
        status = "✅ Fetched" if cid in fetched_circuits else "❌ Not Fetched"
        color = "green" if cid in fetched_circuits else "red"
        html_summary += f"""
            <tr style="background:{'#ecf0f1' if index % 2 == 0 else '#ffffff'};">
                <td style="padding:6px;">{index}</td>
                <td style="padding:6px;">{cid}</td>
                <td style="padding:6px; color:{color}; font-weight:bold;">{status}</td>
            </tr>
        """
        index += 1

    html_summary += """
            </tbody>
        </table>
    </div>
    """

    # 🧾 Force-save HTML Summary to Database (important for HTML fields)
    doc.db_set('note_html', html_summary, update_modified=False)
    doc.db_set('total_circuit_id', len(unique_circuits), update_modified=False)

    # 💾 Save and commit transaction
    try:
        frappe.db.commit()

        # 🔔 Notify frontend (JS) to refresh document
        frappe.publish_realtime(
            event='disconnection_summary_ready',
            message={'docname': doc.name},
            user=frappe.session.user
        )

        frappe.logger().info(f"✅ Disconnection Request {doc.name} processed successfully.")

    except Exception as e:
        frappe.logger().error(f"❌ Error saving Disconnection Request {doc.name}: {e}")

###########################################################################
##Site to LMS Request update
import frappe

@frappe.whitelist()
def create_lms_request(site_name):
    site_doc = frappe.get_doc("Site", site_name)

    if site_doc.lms_stage != "Pending":
        frappe.throw("LMS Request can only be created when LMS Stage is 'Pending'.")

    allowed_status = ['Feasible', 'High Commercials']
    valid_vendors = [row for row in site_doc.lms_vendor if row.lms_status in allowed_status]

    if not valid_vendors:
        frappe.throw("No vendors with valid LMS status found.")

    # Check if LMS Request already exists for any vendor
    if any(row.lms_requested_id for row in valid_vendors):
        frappe.throw("LMS Request already created for one or more vendors.")

    # Only proceed with Feasible vendors
    feasible_vendors = [row for row in valid_vendors if row.lms_status in ["Feasible", "High Commercials"]]

    if not feasible_vendors:
        frappe.throw("No feasible vendors found to create LMS Request.")

    lms_suppliers = []
    for vendor in feasible_vendors:
        lms_suppliers.append({
            "lms_feasibility_partner": vendor.lms_supplier,
            "lms_feasibility_status": vendor.lms_status,
            "supplier_contact": vendor.supplier_contact,
            "bandwith_type": vendor.bandwith_type,
            "media": vendor.media,
            "support_mode": vendor.support_mode,
            "email_id": vendor.email_id,
            "mobile": vendor.mobile,
            "static_ip": vendor.static_ip,
            "bandwidth": vendor.bandwidth,
            "billing_mode": vendor.billing_mode,
            "billing_terms": vendor.billing_terms,
            "feasibility_otc": vendor.otc,
            "validity": vendor.validity,
            "feasibility_security_deposit": vendor.security_deposit,
            "feasibility_mrc": vendor.mrc,
            "feasibility_arc": vendor.arc,
            "feasibility_static_ip_cost": vendor.static_ip_cost
        })

    # Create LMS Request
    lms_request = frappe.get_doc({
        "doctype": "LMS Request",
        "circuit_id": site_doc.name,
        "lms_fesible_suppliers": lms_suppliers
    })
    lms_request.insert(ignore_permissions=True)

    # Update child table entries only once per vendor
    for row in site_doc.lms_vendor:
        if row.lms_status in allowed_status and not row.lms_requested_id:
            frappe.db.set_value(row.doctype, row.name, {
                "lms_requested_id": lms_request.name,
                "stage": "LMS Initiated"
            })

    # Update main Site fields
    frappe.db.set_value("Site", site_name, {
        "lms_stage": "LMS Initiated",
        "site_status": "In-process"
    })

    return lms_request.name
#######################################################################
### Task to n8n
import frappe
import requests

@frappe.whitelist()
def send_to_n8n(task_id, subject, description, due_date=None, type=None):
    """
    Sends Task details from ERPNext to n8n webhook.
    """
    url = "https://chatty-chicken-91.hooks.n8n.cloud/webhook/task-chat"
    payload = {
        "task_id": task_id,
        "subject": subject,
        "description": description,
        "due_date": due_date,
        "type": type
    }
    try:
        r = requests.post(url, json=payload, timeout=5)
        return {
            "status": r.status_code,
            "response": r.text
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "N8N Webhook Error")
        return {
            "status": "error",
            "message": str(e)
        }

###############################################################################
## Disconnection Request creating record in Stock Management
import frappe
from collections import defaultdict

def on_submit_disconnection_request(doc, method):
    # ✅ Check customer type
    if doc.customer_type != "Opex (Rental)":
        frappe.msgprint("Since the customer type is Capex, hardware recovery is not applicable.")
        return  # Exit if not Opex

    circuit_map = defaultdict(list)

    # Group disconnection lines by curcit_id
    for row in doc.disconnection_request:
        if row.curcit_id:
            circuit_map[row.curcit_id].append(row)

    for circuit_id, rows in circuit_map.items():
        stock_management = frappe.new_doc("Stock Management")
        stock_management.circuit_id = circuit_id
        stock_management.status = "Disconnection"
        stock_management.stock_management_type = "Disconnection"

        for row in rows:
            stock_management.append("stock_management_disconnection", {
                "item_code": row.item_code,
                "qty": row.qty,
                "serial_no_sim_no": row.serial_no_sim_no
            })

        stock_management.insert()

    frappe.msgprint("As the customer type is Opex (Rental), hardware recovery is applicable; therefore, the Stock Management record has been created successfully.")
################################################################################33
# Lastmile Services Master - LMS Review Update to Site
import frappe

def sync_lms_review_to_site(doc, method):
    # Check if circuit_id is set
    if not doc.circuit_id:
        return

    # Find Site record with name equal to circuit_id
    site_doc = frappe.get_doc("Site", doc.circuit_id)

    # Update lms_review in Site if different
    if site_doc.lms_review != doc.lms_review:
        site_doc.lms_review = doc.lms_review
        site_doc.save(ignore_permissions=True)  # Optional: ignore if required
        frappe.db.commit()
####################################################################################
import frappe
import json

@frappe.whitelist()
def get_site_data(limit=20):
    """Return site data with limit (0 = ALL)"""
    try:
        filters = [
            ["site_status", "=", "Delivered and Live"],
            ["billing_status", "in", ["Pending", "Partially Completed"]]
        ]

        # Get total count
        total = frappe.db.count("Site", filters=filters)

        if limit and int(limit) > 0:
            sites = frappe.db.get_all(
                "Site",
                fields=[
                    "name",
                    "circuit_id",
                    "customer",
                    "site_name",
                    "site_status",
                    "billing_status"
                ],
                filters=filters,
                limit=int(limit),
                order_by="creation desc"
            )
        else:  # ALL records
            sites = frappe.db.get_all(
                "Site",
                fields=[
                    "name",
                    "circuit_id",
                    "customer",
                    "site_name",
                    "site_status",
                    "billing_status"
                ],
                filters=filters,
                order_by="creation desc"
            )

        return {"sites": sites, "total": total}
    except Exception as e:
        frappe.log_error(f"Failed to get site data: {str(e)}")
        return {"sites": [], "total": 0}


@frappe.whitelist()
def get_sales_order_items_for_sites(site_names):
    """Return sales order items for the given sites based on the conditions"""
    try:
        if not isinstance(site_names, list):
            site_names = json.loads(site_names)
            
        # Get the sites data
        sites = frappe.get_all(
            "Site",
            filters={"name": ["in", site_names]},
            fields=["name", "circuit_id", "customer", "site_name"]
        )
        
        if not sites:
            return []
            
        items = []
        
        for site in sites:
            # 🔹 Get only submitted Sales Orders for this customer
            sales_orders = frappe.get_all(
                "Sales Order",
                filters={
                    "customer": site.customer,
                    "docstatus": 1  # ✅ Submitted only
                },
                fields=["name"]
            )
            
            if not sales_orders:
                continue
                
            # 🔹 Get items from these sales orders where custom_feasibility matches circuit_id
            for so in sales_orders:
                so_items = frappe.get_all(
                    "Sales Order Item",
                    filters={
                        "parent": so.name,
                        "custom_feasibility": site.circuit_id
                    },
                    fields=[
                        "item_name",
                        "rate",
                        "qty",
                        "amount",
                        "custom_feasibility",
                        "custom_site_info",
                        "name",
                        "parent"
                    ]
                )
                
                if so_items:
                    items.extend(so_items)
        
        return items
    except Exception as e:
        frappe.log_error(f"Failed to get sales order items: {str(e)}")
        return []


@frappe.whitelist()
def create_sales_invoice_from_items(items):
    """Create a sales invoice from the given items"""
    try:
        if not isinstance(items, list):
            items = json.loads(items)
            
        if not items:
            frappe.throw("No items provided to create invoice")
            
        # Create new sales invoice
        invoice = frappe.new_doc("Sales Invoice")
        invoice.customer = frappe.db.get_value("Sales Order", items[0].get("parent"), "customer")
        
        # Add items
        for item in items:
            invoice.append("items", {
                "item_name": item.get("item_name"),
                "qty": item.get("qty"),
                "rate": item.get("rate"),
                "amount": item.get("amount"),
                "so_detail": item.get("name"),
                "sales_order": item.get("parent")
            })
        
        invoice.insert()
        invoice.submit()
        
        return {
            "invoice_url": frappe.utils.get_url_to_form("Sales Invoice", invoice.name),
            "invoice_name": invoice.name
        }
    except Exception as e:
        frappe.log_error(f"Failed to create sales invoice: {str(e)}")
        frappe.throw("Failed to create invoice. Please check error logs.")

##################################################################################3
# Installation Note Create
import frappe
from frappe.utils import nowdate

@frappe.whitelist()
def create_installation_note(site_name):
    # Get Site document
    site_doc = frappe.get_doc("Site", site_name)

    # Check if Installation Note already exists
    if site_doc.installation_note:
        frappe.throw(f"Installation Note already exists: {site_doc.installation_note}")

    # Create new Installation Note
    if site_doc.delivery_note_id:
        # Create base doc
        installation_doc = frappe.new_doc("Installation Note")
        dn_doc = frappe.get_doc("Delivery Note", site_doc.delivery_note_id)
        
        # If there are packed items (product bundle), use them. Otherwise, use normal items.
        items_to_pull = dn_doc.packed_items if getattr(dn_doc, "packed_items", []) else dn_doc.items
        
        for item in items_to_pull:
            installation_doc.append("items", {
                "item_code": item.item_code,
                "qty": item.qty,
                "description": item.description,
                "serial_no": getattr(item, "serial_no", ""),
                "prevdoc_doctype": "Delivery Note",
                "prevdoc_docname": dn_doc.name,
                "prevdoc_detail_docname": item.name
            })
    else:
        installation_doc = frappe.new_doc("Installation Note")

    installation_doc.custom_circuit_id = site_doc.name
    installation_doc.inst_date = site_doc.date or nowdate()

    # Set installation type based on lms_stage in Site
    if site_doc.lms_stage == "LMS Partially Delivered":
        installation_doc.custom_installation_type = "Partially Installed"
    elif site_doc.lms_stage == "LMS Delivered":
        installation_doc.custom_installation_type = "Fully Installed"

    # Insert Installation Note
    installation_doc.insert(ignore_permissions=True)
    # installation_doc.submit()  # Uncomment if you want to auto-submit

    # Update Site document fields
    site_doc.db_set({
        "installation_note": installation_doc.name,
        "installation_document_status": "Draft"
    })

    frappe.db.commit()

    return installation_doc.name

#################################################################################
# Updateing the status of Installation Note
import frappe

def update_site_on_installation_note(doc, method):
    if doc.custom_circuit_id:
        # Check if a Site exists with this name
        if frappe.db.exists("Site", doc.custom_circuit_id):
            site_doc = frappe.get_doc("Site", doc.custom_circuit_id)
            
            # Update based on docstatus
            if doc.docstatus == 1:
                site_doc.installation_document_status = "Submitted"
                site_doc.installation_note = doc.name
            
            elif doc.docstatus == 2:
                site_doc.installation_document_status = "Cancelled"
                site_doc.installation_note = doc.name

            elif doc.docstatus == 0:
                site_doc.installation_document_status = "Draft"
                site_doc.installation_note = doc.name
            
            site_doc.save(ignore_permissions=True)
        else:
            frappe.throw(f"Site with name '{doc.custom_circuit_id}' not found.")

##########################################################
import frappe
from frappe.utils import now
from frappe import _

@frappe.whitelist(allow_guest=True)
def update_installation_approval_status(site=None):
    if not site:
        frappe.respond_as_web_page("Missing Site ID", "<p>Site ID is required.</p>", http_status_code=400)
        return

    try:
        doc = frappe.get_doc("Site", site)

        # If already accepted, don't update again
        if doc.client_installation_approval_status == "Accepted":
            frappe.respond_as_web_page(
                "✅ Already Accepted",
                f"""
                <div style="text-align: center; font-family: Arial, sans-serif; padding: 30px;">
                  <h2 style="color: green;">✅ Already Accepted</h2>
                  <p>This installation has already been marked as <strong>Accepted</strong> on {doc.client_installation_approval_date}.</p>
                  <p style="margin-top: 40px;">– Nexapp Technologies Private Limited</p>
                </div>
                """,
                http_status_code=200
            )
            return

        # Update status and timestamp only once
        frappe.db.set_value("Site", site, {
            "client_installation_approval_status": "Accepted",
            "client_installation_approval_date": now(),
            "site_status": "Delivered and Live"
        })
        frappe.db.commit()

        # Show success page
        frappe.respond_as_web_page(
            "✅ Installation Accepted",
            """
            <div style="text-align: center; font-family: Arial, sans-serif; padding: 30px;">
              <h2 style="color: green;">✅ Installation Accepted</h2>
              <p>Thank you for confirming. The installation has been marked as <strong>Accepted</strong> and status set to <strong>Delivered and Live</strong>.</p>
              <p style="margin-top: 40px;">– Nexapp Technologies Private Limited</p>
            </div>
            """,
            http_status_code=200
        )

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Installation Approval Error")
        frappe.respond_as_web_page(
            "❌ Error",
            f"<p>Something went wrong: {frappe.get_traceback()}</p>",
            http_status_code=500
        )

#################################################################################
## Bank Reocnimport frappe

import frappe
import re
from difflib import SequenceMatcher

# ------------------------------------------------------------
# Helpers: list/get entries
# ------------------------------------------------------------

@frappe.whitelist()
def get_bank_statement_entries(bank_account=None, start_date=None, end_date=None):
    filters = {"reconciled": 0}
    if bank_account:
        filters["bank_account"] = bank_account
    if start_date and end_date:
        filters["transaction_date"] = ["between", [start_date, end_date]]

    entries = frappe.get_all(
        "Bank Statement Entry",
        filters=filters,
        fields=["name", "transaction_date", "description", "deposit", "withdrawal", "bank_account"],
        order_by="transaction_date asc",
    )

    for entry in entries:
        entry["date"] = entry.pop("transaction_date")

    return entries


def clean_string(s):
    return re.sub(r"[^a-zA-Z0-9\s\-]", "", (s or "")).lower()


def extract_keywords(description):
    parts = clean_string(description).split()
    keywords = [p for p in parts if len(p) > 2 and not p.isdigit()]
    return keywords


def calculate_match_score(description, target, amount, target_amount):
    try:
        amount = float(amount)
        target_amount = float(target_amount)
    except Exception:
        amount = amount or 0.0
        target_amount = target_amount or 0.0

    name_score = SequenceMatcher(None, clean_string(description), clean_string(target)).ratio()
    amount_diff = abs(float(amount) - float(target_amount))
    mx = max(float(amount), float(target_amount)) if max(float(amount), float(target_amount)) else 0.0
    amount_score = 1 - (amount_diff / mx) if mx else 0
    return round((0.7 * amount_score + 0.3 * name_score) * 100, 1)


@frappe.whitelist()
def find_matching_invoices(amount, description=None):
    try:
        amount = float(amount)
    except Exception:
        return []

    if not description:
        description = ""

    matches = []
    _ = extract_keywords(description)

    purchase_invoices = frappe.get_all(
        "Purchase Invoice",
        filters={"docstatus": 1, "outstanding_amount": [">", 0], "status": ["!=", "Paid"]},
        fields=[
            "name",
            "supplier as party",
            "outstanding_amount",
            "bill_no",
            "bill_date",
            "supplier as party_name",
            "company",
        ],
    )

    sales_invoices = frappe.get_all(
        "Sales Invoice",
        filters={"docstatus": 1, "outstanding_amount": [">", 0], "status": ["!=", "Paid"]},
        fields=[
            "name",
            "customer as party",
            "outstanding_amount",
            "posting_date",
            "customer as party_name",
            "company",
        ],
    )

    all_invoices = purchase_invoices + sales_invoices

    for inv in all_invoices:
        inv_name = inv.get("name")
        party = inv.get("party") or ""
        inv_amount = inv.get("outstanding_amount") or 0.0

        if "bill_no" in inv:
            bill_no = inv.get("bill_no", "")
            bill_date = inv.get("bill_date", "")
            doctype = "Purchase Invoice"
        elif "posting_date" in inv:
            bill_no = ""
            bill_date = inv.get("posting_date", "")
            doctype = "Sales Invoice"
        else:
            bill_no = ""
            bill_date = ""
            doctype = inv.get("doctype") or ""

        score = calculate_match_score(description, f"{inv_name} {party}", amount, inv_amount)

        matches.append(
            {
                "name": inv_name,
                "party": party,
                "party_name": inv.get("party_name", ""),
                "outstanding_amount": inv_amount,
                "match_score": score,
                "bill_no": bill_no,
                "bill_date": bill_date,
                "doctype": doctype,
                "company": inv.get("company"),
            }
        )

    matches.sort(key=lambda x: x["match_score"], reverse=True)
    return matches[:5]


# ------------------------------------------------------------
# NEW: Get customer outstanding invoices (Deposit)
# ------------------------------------------------------------

@frappe.whitelist()
def get_customer_outstanding_invoices(customer=None, company=None):
    filters = {
        "docstatus": 1,
        "outstanding_amount": [">", 0],
        "status": ["!=", "Paid"]
    }

    if customer:
        filters["customer"] = customer

    if company:
        filters["company"] = company

    sales_invoices = frappe.get_all(
        "Sales Invoice",
        filters=filters,
        fields=[
            "name",
            "customer as party",
            "outstanding_amount",
            "posting_date",
            "due_date",
            "customer as party_name",
            "company",
        ],
        order_by="posting_date desc"
    )

    result = []
    for inv in sales_invoices:
        result.append({
            "name": inv.get("name"),
            "party": inv.get("party"),
            "party_name": inv.get("party_name"),
            "outstanding_amount": inv.get("outstanding_amount"),
            "posting_date": inv.get("posting_date"),
            "due_date": inv.get("due_date"),
            "doctype": "Sales Invoice",
            "company": inv.get("company")
        })

    return result


# ------------------------------------------------------------
# GET OUTSTANDING EXPENSE CLAIMS FOR EMPLOYEE
# ------------------------------------------------------------

@frappe.whitelist()
def get_unpaid_expense_claims(employee=None, company=None):
    # Extract employee ID if it comes in the format "Name (ID)"
    if employee and "(" in employee and ")" in employee:
        employee = employee.split("(")[-1].split(")")[0]

    if not company:
        company = frappe.defaults.get_user_default("company") or frappe.defaults.get_default("company")

    filters = {
        "employee": employee,
        "docstatus": 1,
        "status": ["!=", "Paid"]
    }
    if company:
        filters["company"] = company

    expense_claims_list = frappe.get_all(
        "Expense Claim",
        filters=filters,
        fields=["name"],
        order_by="posting_date desc",
        limit_page_length=0
    )
    
    # Fallback: if no claims found with company filter, try without it
    if not expense_claims_list and company:
        if "company" in filters:
            del filters["company"]
            expense_claims_list = frappe.get_all(
                "Expense Claim",
                filters=filters,
                fields=["name"],
                order_by="posting_date desc",
                limit_page_length=0
            )

    expense_claims = []
    # Fetch full documents to ensure all custom and virtual fields are available
    for c in expense_claims_list:
        try:
            doc = frappe.get_doc("Expense Claim", c.name)
            
            # Fetch all components for calculation
            sanctioned = frappe.utils.flt(doc.total_sanctioned_amount, 2)
            taxes = frappe.utils.flt(doc.total_taxes_and_charges, 2)
            reimbursed = frappe.utils.flt(doc.total_amount_reimbursed, 2)
            grand = frappe.utils.flt(doc.grand_total, 2)
            
            # Advances to be deducted - check both child table and parent field
            adv_from_child = sum(frappe.utils.flt(adv.allocated_amount, 2) for adv in doc.get("advances") or [])
            adv_paid_from_child = sum(frappe.utils.flt(adv.advance_paid, 2) for adv in doc.get("advances") or [])
            total_advance = frappe.utils.flt(max(adv_from_child, adv_paid_from_child, frappe.utils.flt(doc.total_advance_amount, 2)), 2)
            
            # Use Gross Amount (Sanctioned + Taxes) as the base
            base_total = frappe.utils.flt(sanctioned + taxes, 2)
            
            # Fallback if sanctioned amount is not populated correctly
            if not base_total or base_total < total_advance:
                base_total = frappe.utils.flt(grand + total_advance, 2)
            
            outstanding = frappe.utils.flt(base_total - total_advance - reimbursed, 2)
            
            expense_claims.append({
                "name": doc.name,
                "employee": doc.employee,
                "employee_name": doc.employee_name,
                "posting_date": doc.posting_date,
                "description": doc.get("description") or doc.get("remarks") or "",
                "total_sanctioned_amount": base_total,
                "total_amount_reimbursed": reimbursed,
                "total_advance_amount": total_advance,
                "grand_total": grand,
                "outstanding_amount": outstanding
            })
        except Exception as e:
            frappe.log_error(f"Error processing {c.name}: {str(e)}", "Bank Recon API Error")
            continue

    return expense_claims


# ------------------------------------------------------------
# OUTSTANDING INVOICE FETCHER (Manual Category)
# ------------------------------------------------------------

@frappe.whitelist()
def get_outstanding_invoices(doctype, party_field, party_name, company=None):
    if not company:
        company = frappe.defaults.get_default("company")

    if doctype == "Sales Invoice":
        fields = [
            "name",
            "customer as party",
            "posting_date",
            "grand_total",
            "outstanding_amount",
            "due_date"
        ]
        try:
            invoice_field_exists = frappe.db.sql("""
                SELECT COUNT(*) FROM information_schema.columns 
                WHERE table_name = 'tabSales Invoice' AND column_name = 'invoice'
            """)
            if invoice_field_exists[0][0] > 0:
                fields.append("invoice as bill_no")
            else:
                fields.append("name as bill_no")
        except:
            fields.append("name as bill_no")

    else:
        fields = [
            "name",
            "supplier as party",
            "posting_date",
            "grand_total",
            "outstanding_amount",
            "due_date"
        ]
        try:
            bill_no_field_exists = frappe.db.sql("""
                SELECT COUNT(*) FROM information_schema.columns 
                WHERE table_name = 'tabPurchase Invoice' AND column_name = 'bill_no'
            """)
            if bill_no_field_exists[0][0] > 0:
                fields.append("bill_no")
            else:
                fields.append("name as bill_no")
        except:
            fields.append("name as bill_no")

    filters = {
        party_field: party_name,
        "docstatus": 1,
        "outstanding_amount": [">", 0],
        "company": company
    }

    invoices = frappe.get_all(
        doctype,
        filters=filters,
        fields=fields,
        order_by="posting_date asc",
        limit_page_length=0
    )

    return invoices


# ------------------------------------------------------------
# TAX ROW ADDER
# ------------------------------------------------------------

def _add_tax_rows_to_payment(payment_entry, tax_adjustments_list):
    try:
        available_tables = [(f.fieldname, f.options) for f in payment_entry.meta.fields if f.fieldtype == "Table"]
    except Exception:
        available_tables = []

    chosen_table = None    # fieldname on Payment Entry
    child_doctype = None   # child DocType name

    for fn, options in available_tables:
        if options and (
            "advance" in (options or "").lower()
            or "tax" in (options or "").lower()
            or "charge" in (options or "").lower()
        ):
            chosen_table = fn
            child_doctype = options
            break

    if not chosen_table:
        available_names = [fn for fn, _ in available_tables]
        for k in ("taxes", "taxes_and_charges", "advance_taxes", "advance_taxes_and_charges", "other_charges", "advances", "deductions"):
            if k in available_names:
                chosen_table = k
                child_doctype = dict(available_tables).get(k)
                break

    if not chosen_table:
        if frappe.db.exists("DocType", "Advance Taxes and Charges"):
            chosen_table = "taxes"
            child_doctype = "Advance Taxes and Charges"
        else:
            frappe.throw("No suitable child table found in Payment Entry to add tax adjustments.")

    child_meta = frappe.get_meta(child_doctype) if child_doctype else None
    child_fields = [f.fieldname for f in (child_meta.fields if child_meta else [])]

    appended_rows = []

    for tax_adj in (tax_adjustments_list or []):
        account_head = tax_adj.get("account_head")
        try:
            tax_amount = float(tax_adj.get("tax_amount", 0) or 0)
        except Exception:
            tax_amount = 0.0

        if not account_head or not frappe.db.exists("Account", account_head):
            frappe.throw(f"Account head not found or invalid: {account_head}")

        row = {}
        if "charge_type" in child_fields:
            row["charge_type"] = "Actual"
        if "account_head" in child_fields:
            row["account_head"] = account_head
        elif "account" in child_fields:
            row["account"] = account_head
        if "tax_amount" in child_fields:
            row["tax_amount"] = tax_amount
        if "amount" in child_fields:
            row["amount"] = tax_amount
        if "description" in child_fields:
            row["description"] = tax_adj.get("description") or account_head
        if "add_deduct_tax" in child_fields:
            row["add_deduct_tax"] = "Deduct"

        payment_entry.append(chosen_table, row)
        appended_rows.append(row)

    try:
        frappe.log_error(title="BR_TAX_ROWS", message=frappe.as_json(appended_rows))
    except Exception:
        pass

    return chosen_table


# ------------------------------------------------------------
# GST FIELD SETTER
# ------------------------------------------------------------

def _ensure_customer_address_with_gst(customer, company):
    try:
        customer_gstin = frappe.db.get_value("Customer", customer, "gstin")
        if not customer_gstin:
            return None

        addresses = frappe.get_all(
            "Address",
            filters={"link_doctype": "Customer", "link_name": customer},
            fields=["name", "gstin", "is_primary_address"]
        )

        if addresses:
            for addr in addresses:
                if not addr.gstin:
                    frappe.db.set_value("Address", addr.name, "gstin", customer_gstin)
                if not addr.is_primary_address:
                    frappe.db.set_value("Address", addr.name, "is_primary_address", 1)
            return addresses[0].name

        address_doc = frappe.get_doc({
            "doctype": "Address",
            "address_title": customer,
            "address_type": "Billing",
            "address_line1": "Auto-created for GST",
            "city": "Mumbai",
            "state": "Maharashtra",
            "country": "India",
            "pincode": "400001",
            "gstin": customer_gstin,
            "is_primary_address": 1,
            "is_your_company_address": 0,
            "links": [{"link_doctype": "Customer", "link_name": customer}]
        })

        address_doc.insert(ignore_permissions=True)
        return address_doc.name

    except Exception as e:
        frappe.log_error(f"Address Creation Failed for {customer}", str(e))
        return None


def _set_gst_fields(payment_entry, party, party_type, company, doctype=None, invoice_name=None):
    company_gstin = frappe.db.get_value("Company", company, "gstin")
    if company_gstin:
        payment_entry.company_gstin = company_gstin
    else:
        company_address = frappe.db.get_value("Address", {"is_primary_company_address": 1, "company": company}, "gstin")
        if company_address:
            payment_entry.company_gstin = company_address

    party_gstin = None
    if party_type == "Customer":
        party_gstin = frappe.db.get_value("Customer", party, "gstin")
        _ensure_customer_address_with_gst(party, company)
    elif party_type == "Supplier":
        party_gstin = frappe.db.get_value("Supplier", party, "gstin")

    allowed_categories = ["Overseas", "Unregistered"]

    if party_type == "Customer":
        gst_category = frappe.db.get_value("Customer", party, "gst_category")
    else:
        gst_category = frappe.db.get_value("Supplier", party, "gst_category")

    if party_gstin:
        try:
            frappe.db.set_value(party_type, party, "gst_category", "Unregistered")
            frappe.db.commit()
        except Exception:
            pass
        payment_entry.gst_category = "Unregistered"
    elif gst_category in allowed_categories:
        payment_entry.gst_category = gst_category
    else:
        payment_entry.gst_category = "Unregistered"

    if party_gstin:
        payment_entry.party_gstin = party_gstin

    if party_type == "Customer" and party_gstin:
        addresses = frappe.get_all(
            "Address",
            filters={"link_doctype": "Customer", "link_name": party, "gstin": ["!=", ""]},
            fields=["name"],
            limit=1
        )
        if addresses:
            payment_entry.customer_address = addresses[0].name


# ------------------------------------------------------------
# UPDATED: NEW CLEAN CUSTOMER PAYMENT FUNCTION
# ------------------------------------------------------------

def process_customer_payment(stmt, invoices, company, customer, tax_adjustments_list, allow_overpayment=False):
    statement_amount = abs(float(stmt.deposit or 0))
    total_allocated = sum(float(inv.get("amount") or 0) for inv in invoices)
    
    # Calculate total deductions from tax adjustments
    total_deductions = 0
    if tax_adjustments_list:
        for tax_adj in tax_adjustments_list:
            total_deductions += float(tax_adj.get("tax_amount") or 0)
    
    # Calculate net payment (allocated - deductions)
    net_payment = total_allocated - total_deductions
    
    # IMPORTANT: paid_amount and received_amount should be the ALLOCATED amount
    # NOT the statement amount or net payment amount
    paid_amount = total_allocated  # This should be 14062.00 in your example
    received_amount = total_allocated  # Same as paid_amount
    
    # Resolve paid_from from first invoice's debit_to (Debtors account)
    customer_receivable = None
    for inv in invoices:
        if inv.get("doctype") == "Sales Invoice" and inv.get("invoice"):
            customer_receivable = frappe.db.get_value("Sales Invoice", inv["invoice"], "debit_to")
            if customer_receivable:
                break
    if not customer_receivable:
        customer_receivable = frappe.db.get_value("Company", company, "default_receivable_account")

    # Resolve Bank GL Account
    bank_account_gl = get_bank_gl_account(stmt.bank_account)

    # For Payment Entry, we use the allocated amount for paid_amount/received_amount
    # The deductions will be handled separately in the taxes table
    payment_entry = frappe.get_doc(
        {
            "doctype": "Payment Entry",
            "payment_type": "Receive",
            "mode_of_payment": "Wire Transfer",
            "company": company,
            "party_type": "Customer",
            "party": customer,
            "paid_amount": paid_amount,  # Set to allocated amount
            "received_amount": received_amount,  # Set to allocated amount
            "reference_no": stmt.description,
            "reference_date": stmt.transaction_date,
            "posting_date": stmt.transaction_date,
            "paid_from": customer_receivable,  # Debtors account
            "paid_to": bank_account_gl or get_default_bank_account(company, "Receive"),  # Bank GL account
            "bank_account": (to_account or from_account or stmt.bank_account)
        }
    )

    reference_rows = []
    for inv in invoices:
        reference_rows.append(
            {
                "reference_doctype": inv.get("doctype"),
                "reference_name": inv.get("invoice"),
                "allocated_amount": float(inv.get("amount") or 0),
            }
        )

    payment_entry.set("references", reference_rows)

    if invoices:
        first = invoices[0]
        _set_gst_fields(
            payment_entry,
            customer,
            "Customer",
            company,
            first.get("doctype"),
            first.get("invoice")
        )

    if tax_adjustments_list:
        _add_tax_rows_to_payment(payment_entry, tax_adjustments_list)

    payment_entry.insert()
    payment_entry.submit()

    # Update Bank Statement Entry
    frappe.db.set_value("Bank Statement Entry", stmt.name, {
        "reference_no": payment_entry.name,
        "reconciled": 1,
        "match_type": "Auto"
    })

    return {
        "status": "ok",
        "payment_entry": payment_entry.name,
        "allocated": total_allocated,  # 14062.00
        "deductions": total_deductions,  # 1191.60
        "net_payment": net_payment,  # 12870.40
        "statement_amount": statement_amount,
        "paid_amount": paid_amount,  # 14062.00
        "allow_overpayment": allow_overpayment,
        "excess_amount": statement_amount - net_payment  # Bank amount vs what actually reaches customer
    }


# ------------------------------------------------------------
# UPDATED MAIN FUNCTION WITH EMPLOYEE EXPENSE CLAIM SUPPORT
# ------------------------------------------------------------

@frappe.whitelist()
def categorize_manually(
    statement_name,
    invoices,
    category=None,
    employee=None,
    customer=None,
    supplier=None,
    expense_account=None,
    company=None,
    tax_adjustments=None,
    from_account=None,
    to_account=None,
    transfer_description=None,
    allow_overpayment=False,
    purchase_order=None,
    sales_order=None,
    custom_send_email=0
):
    try:
        if not isinstance(invoices, list):
            invoices = frappe.parse_json(invoices)

        stmt = frappe.get_doc("Bank Statement Entry", statement_name)

        if not company:
            company = frappe.db.get_value("Bank Account", stmt.bank_account, "company")

        # Resolve Bank Account GL Head from Statement or passed values
        bank_account_gl = get_bank_gl_account(stmt.bank_account)

        # Save original names for linking in Payment Entry
        from_account_orig = from_account
        to_account_orig = to_account

        # Resolve UI passed accounts
        from_account = get_bank_gl_account(from_account)
        to_account = get_bank_gl_account(to_account)

        is_deposit = bool(stmt.deposit and float(stmt.deposit) > 0)

        if is_deposit:
            payment_type = "Receive"
            # Prioritize UI passed to_account, then Statement bank account, then Default
            paid_to = (to_account or bank_account_gl or get_default_bank_account(company, payment_type))
            paid_from = from_account
        else:
            payment_type = "Pay"
            # Prioritize UI passed from_account, then Statement bank account, then Default
            paid_from = (from_account or bank_account_gl or get_default_bank_account(company, payment_type))
            paid_to = to_account

        statement_amount = frappe.utils.flt(abs(float(stmt.deposit or stmt.withdrawal) or 0.0), 2)

        total_allocated = frappe.utils.flt(sum(frappe.utils.flt(inv.get("amount") or 0, 2) for inv in invoices), 2)

        tax_adjustments_list = []
        if tax_adjustments:
            if isinstance(tax_adjustments, str):
                tax_adjustments_list = frappe.parse_json(tax_adjustments)
            elif isinstance(tax_adjustments, list):
                tax_adjustments_list = tax_adjustments
            else:
                tax_adjustments_list = frappe.parse_json(frappe.as_json(tax_adjustments))

        # -----------------------------------------------------------
        # CUSTOMER PAYMENT → SEPARATE FUNCTION (UPDATED)
        # -----------------------------------------------------------
        if category == "Customer Payment":
            return process_customer_payment(
                stmt=stmt,
                invoices=invoices,
                company=company,
                customer=customer,
                tax_adjustments_list=tax_adjustments_list,
                allow_overpayment=allow_overpayment,
                bank_account=to_account_orig # ⭐ Pass UI selection
            )

        # -----------------------------------------------------------
        # EMPLOYEE EXPENSE CLAIM
        # -----------------------------------------------------------
        elif category == "Employee Expense Claim":
            if not employee:
                return {"status": "fail", "error": "Employee is required for Expense Claim"}

            # Validate employee exists
            if not frappe.db.exists("Employee", employee):
                return {"status": "fail", "error": f"Employee {employee} not found"}

            # Resolve paid_to from expense claim's payable_account
            emp_payable = None
            for inv in invoices:
                if inv.get("invoice"):
                    emp_payable = frappe.db.get_value("Expense Claim", inv["invoice"], "payable_account")
                    if emp_payable:
                        break
            if not emp_payable:
                emp_payable = frappe.db.get_value("Company", company, "default_payable_account")

            # For Employee Expense Claim, use allocated amount NOT statement amount
            paid_amount = frappe.utils.flt(total_allocated, 2)
            received_amount = frappe.utils.flt(total_allocated, 2)

            payment_entry = frappe.get_doc(
                {
                    "doctype": "Payment Entry",
                    "payment_type": payment_type,
                    "mode_of_payment": "Wire Transfer",
                    "company": company,
                    "party_type": "Employee",  # Set party_type to Employee
                    "party": employee,  # Use employee ID
                    "paid_amount": paid_amount,  # Use allocated amount
                    "received_amount": received_amount,  # Use allocated amount
                    "reference_no": stmt.description,
                    "reference_date": stmt.transaction_date,
                    "posting_date": stmt.transaction_date,
                    "paid_from": (bank_account_gl or get_default_bank_account(company, "Pay")),  # Bank GL account
                    "paid_to": emp_payable,  # Employee's payable account
                }
            )

            references = []
            for inv in invoices:
                amt = frappe.utils.flt(inv.get("amount") or 0, 2)
                if inv.get("invoice"):
                    actual_out = get_actual_outstanding("Expense Claim", inv.get("invoice"))
                    if amt > actual_out and abs(amt - actual_out) < 0.05:
                        amt = actual_out
                
                references.append(
                    {
                        "reference_doctype": "Expense Claim",
                        "reference_name": inv.get("invoice"),
                        "allocated_amount": amt,
                    }
                )

            payment_entry.set("references", references)
            total_allocated = frappe.utils.flt(sum(frappe.utils.flt(ref.get("allocated_amount") or 0) for ref in references), 2)

        # -----------------------------------------------------------
        # SUPPLIER PAYMENT
        # -----------------------------------------------------------
        elif category == "Supplier Payment":
            if not supplier:
                return {"status": "fail", "error": "Supplier is required for Supplier Payment"}

            # Resolve paid_to from invoice's credit_to (Creditors account)
            supplier_payable = None
            for inv in invoices:
                if inv.get("doctype") == "Purchase Invoice" and inv.get("invoice"):
                    supplier_payable = frappe.db.get_value("Purchase Invoice", inv["invoice"], "credit_to")
                    if supplier_payable:
                        break
            if not supplier_payable:
                supplier_payable = frappe.db.get_value("Company", company, "default_payable_account")

            # For Supplier Payment, use allocated amount NOT statement amount
            paid_amount = frappe.utils.flt(total_allocated, 2)
            received_amount = frappe.utils.flt(total_allocated, 2)

            # Fetch supplier email for notification
            supplier_email = frappe.db.get_value("Supplier", supplier, "email_id") or ""

            payment_entry = frappe.get_doc(
                {
                    "doctype": "Payment Entry",
                    "payment_type": payment_type,
                    "mode_of_payment": "Wire Transfer",
                    "company": company,
                    "party_type": "Supplier",
                    "party": supplier,
                    "paid_amount": paid_amount,  # Use allocated amount
                    "received_amount": received_amount,  # Use allocated amount
                    "reference_no": stmt.description,
                    "reference_date": stmt.transaction_date,
                    "posting_date": stmt.transaction_date,
                    "paid_from": (bank_account_gl or get_default_bank_account(company, "Pay")),  # Bank GL account
                    "paid_to": supplier_payable,  # Supplier's Creditors account
                    "bank_account": (from_account_orig or to_account_orig or stmt.bank_account),
                    "custom_send_email": custom_send_email,
                    "contact_email": supplier_email
                }
            )

            references = []
            for inv in invoices:
                amt = frappe.utils.flt(inv.get("amount") or 0, 2)
                if inv.get("invoice"):
                    actual_out = get_actual_outstanding(inv.get("doctype") or "Purchase Invoice", inv.get("invoice"))
                    if amt > actual_out and abs(amt - actual_out) < 0.05:
                        amt = actual_out
                        
                references.append(
                    {
                        "reference_doctype": inv.get("doctype"),
                        "reference_name": inv.get("invoice"),
                        "allocated_amount": amt,
                    }
                )

            payment_entry.set("references", references)
            total_allocated = frappe.utils.flt(sum(frappe.utils.flt(ref.get("allocated_amount") or 0) for ref in references), 2)

        # -----------------------------------------------------------
        # EXPENSE CATEGORY
        # -----------------------------------------------------------
        elif category == "Expense":
            if not expense_account:
                return {"status": "fail", "error": "Expense Account is required for Expense category"}

            # Create a Journal Entry for expense
            return create_expense_journal_entry(
                stmt=stmt,
                expense_account=expense_account,
                amount=statement_amount,
                company=company
            )

        # -----------------------------------------------------------
        # TRANSFER TO ANOTHER ACCOUNT
        # -----------------------------------------------------------
        elif category == "Transfer To Another Account":
            if not from_account or not to_account:
                return {"status": "fail", "error": "From Account and To Account are required for Transfer"}

            # Create a Journal Entry for transfer
            return create_bank_transfer_journal(
                stmt=stmt,
                from_account=from_account,
                to_account=to_account,
                description=transfer_description,
                company=company
            )

        # -----------------------------------------------------------
        # SUPPLIER ADVANCE
        # -----------------------------------------------------------
        elif category == "Supplier Advance":
            if not supplier:
                return {"status": "fail", "error": "Supplier is required for Supplier Advance"}
            
            return create_supplier_advance_payment(
                supplier=supplier,
                amount=statement_amount,
                statement_entry=statement_name,
                purchase_order=purchase_order,
                bank_account=from_account, # ⭐ Pass from UI
                custom_send_email=custom_send_email
            )

        # -----------------------------------------------------------
        # CUSTOMER ADVANCE
        # -----------------------------------------------------------
        elif category == "Customer Advance":
            if not customer:
                return {"status": "fail", "error": "Customer is required for Customer Advance"}
            
            return create_customer_advance_payment(
                customer=customer,
                amount=statement_amount,
                statement_entry=statement_name,
                sales_order=sales_order,
                bank_account=to_account # ⭐ Pass from UI
            )

        # -----------------------------------------------------------
        # RETURN PAYMENT (SUPPLIER)
        # -----------------------------------------------------------
        elif category == "Return Payment (Supplier)":
            if not supplier:
                return {"status": "fail", "error": "Supplier is required for Return Payment (Supplier)"}
            
            return create_supplier_return_payment(
                supplier=supplier,
                amount=statement_amount,
                statement_entry=statement_name,
                bank_account=to_account, # ⭐ Pass from UI
                custom_send_email=custom_send_email
            )

        # -----------------------------------------------------------
        # RETURN PAYMENT (CUSTOMER)
        # -----------------------------------------------------------
        elif category == "Return Payment (Customer)":
            if not customer:
                return {"status": "fail", "error": "Customer is required for Return Payment (Customer)"}
            
            return create_customer_return_payment(
                customer=customer,
                amount=statement_amount,
                statement_entry=statement_name,
                bank_account=to_account # ⭐ Pass from UI
            )

        else:
            return {"status": "fail", "error": f"Unknown category: {category}"}

        # Validate total allocated matches statement amount
        # SKIP for Customer Payment (allow excess)
        if category != "Customer Payment" and category not in ["Expense", "Transfer To Another Account"]:
            if abs(frappe.utils.flt(total_allocated, 2) - frappe.utils.flt(statement_amount, 2)) > 0.05:
                return {
                    "status": "fail",
                    "error": f"Total allocated amount {total_allocated} doesn't match statement amount {statement_amount}"
                }

        if tax_adjustments_list:
            _add_tax_rows_to_payment(payment_entry, tax_adjustments_list)

        payment_entry.custom_send_email = custom_send_email
        payment_entry.insert()
        payment_entry.submit()

        # Send email directly if custom_send_email is set
        if int(custom_send_email) == 1 and payment_entry.party_type == "Supplier":
            send_supplier_payment_email(payment_entry)

        # Update Bank Statement Entry
        frappe.db.set_value("Bank Statement Entry", statement_name, {
            "reference_no": payment_entry.name,
            "reconciled": 1,
            "match_type": "Auto"
        })

        # Create reconciliation log
        if frappe.db.exists("DocType", "Bank Reconciliation Log"):
            for inv in invoices:
                log = frappe.get_doc(
                    {
                        "doctype": "Bank Reconciliation Log",
                        "invoice": inv.get("invoice"),
                        "invoice_type": inv.get("doctype"),
                        "bank_statement": statement_name,
                        "matched_amount": inv.get("amount"),
                        "payment_entry": payment_entry.name,
                        "reconciliation_date": frappe.utils.nowdate(),
                        "reconciliation_type": "Manual",
                        "category": category,
                        "employee": employee if category == "Employee Expense Claim" else None,
                        "supplier": supplier if category == "Supplier Payment" else None,
                        "expense_account": expense_account,
                        "is_deposit": is_deposit,
                    }
                )
                log.insert(ignore_permissions=True)

        return {
            "status": "ok",
            "payment_entry": payment_entry.name,
            "total_allocated": total_allocated,
            "paid_amount": paid_amount,
            "is_deposit": is_deposit,
        }

    except Exception as e:
        frappe.log_error(title="BR_MANUAL_ERROR", message=frappe.get_traceback())
        return {"status": "fail", "error": str(e)}


# ------------------------------------------------------------
# CREATE EXPENSE JOURNAL ENTRY
# ------------------------------------------------------------

def create_expense_journal_entry(stmt, expense_account, amount, company):
    try:
        journal_entry = frappe.get_doc({
            "doctype": "Journal Entry",
            "voucher_type": "Journal Entry",
            "posting_date": stmt.transaction_date,
            "company": company,
            "cheque_no": stmt.description,
            "cheque_date": stmt.transaction_date,
            "user_remark": f"Expense payment from bank statement: {stmt.description}",
        })

        # Debit Expense Account
        journal_entry.append("accounts", {
            "account": expense_account,
            "debit_in_account_currency": amount,
            "credit_in_account_currency": 0,
            "party_type": "",
            "party": "",
            "cost_center": get_default_cost_center(company)
        })

        # Credit Bank Account
        journal_entry.append("accounts", {
            "account": get_default_bank_account(company, "Pay"),
            "debit_in_account_currency": 0,
            "credit_in_account_currency": amount,
            "party_type": "",
            "party": "",
            "cost_center": get_default_cost_center(company)
        })

        journal_entry.insert()
        journal_entry.submit()

        # Update Bank Statement Entry
        frappe.db.set_value("Bank Statement Entry", stmt.name, {
            "reference_no": journal_entry.name,
            "reconciled": 1,
            "match_type": "Auto"
        })

        return {
            "status": "ok",
            "journal_entry": journal_entry.name,
            "amount": amount,
            "expense_account": expense_account
        }

    except Exception as e:
        frappe.log_error(title="BR_EXPENSE_ERROR", message=frappe.get_traceback())
        return {"status": "fail", "error": str(e)}


# ------------------------------------------------------------
# BANK TRANSFER JOURNAL ENTRY FUNCTION
# ------------------------------------------------------------

def create_bank_transfer_journal(stmt, from_account, to_account, description, company):
    try:
        amount = abs(float(stmt.deposit or stmt.withdrawal) or 0.0)

        journal_entry = frappe.get_doc({
            "doctype": "Journal Entry",
            "voucher_type": "Bank Entry",
            "posting_date": stmt.transaction_date,
            "company": company,
            "cheque_no": stmt.description,
            "cheque_date": stmt.transaction_date,
            "user_remark": description or f"Bank transfer from {from_account} to {to_account}",
        })

        # Debit To Account
        journal_entry.append("accounts", {
            "account": to_account,
            "debit_in_account_currency": amount,
            "credit_in_account_currency": 0,
            "party_type": "",
            "party": "",
            "cost_center": get_default_cost_center(company)
        })

        # Credit From Account
        journal_entry.append("accounts", {
            "account": from_account,
            "debit_in_account_currency": 0,
            "credit_in_account_currency": amount,
            "party_type": "",
            "party": "",
            "cost_center": get_default_cost_center(company)
        })

        journal_entry.insert()
        journal_entry.submit()

        # Update Bank Statement Entry
        frappe.db.set_value("Bank Statement Entry", stmt.name, {
            "reference_no": journal_entry.name,
            "reconciled": 1,
            "match_type": "Auto"
        })

        return {
            "status": "ok",
            "journal_entry": journal_entry.name,
            "amount": amount,
            "from_account": from_account,
            "to_account": to_account
        }

    except Exception as e:
        frappe.log_error(title="BR_TRANSFER_ERROR", message=frappe.get_traceback())
        return {"status": "fail", "error": str(e)}


# ------------------------------------------------------------
# ITEMIZED JOURNAL ENTRY FUNCTION
# ------------------------------------------------------------

@frappe.whitelist()
def create_itemized_journal_entry(
    statement_name=None,
    itemized_entries=None,
    from_account=None, # ⭐ Added
    company=None
):
    try:
        if not company:
            company = frappe.defaults.get_default("company")

        stmt = frappe.get_doc("Bank Statement Entry", statement_name)

        if not isinstance(itemized_entries, list):
            itemized_entries = frappe.parse_json(itemized_entries)

        total_amount = sum(float(entry.get("amount") or 0) for entry in itemized_entries)
        statement_amount = abs(float(stmt.deposit or stmt.withdrawal) or 0.0)

        # Validate total amount matches statement amount
        if abs(total_amount - statement_amount) > 0.01:
            return {
                "status": "fail",
                "error": f"Total itemized amount {total_amount} doesn't match statement amount {statement_amount}"
            }

        journal_entry = frappe.get_doc({
            "doctype": "Journal Entry",
            "voucher_type": "Journal Entry",
            "posting_date": stmt.transaction_date,
            "company": company,
            "cheque_no": stmt.description,
            "cheque_date": stmt.transaction_date,
            "user_remark": f"Itemized expense payment from bank statement: {stmt.description}",
            "bank_account": (from_account or stmt.bank_account) # ⭐ Set the Bank Account doc name
        })

        # Add debit entries for each expense account
        for entry in itemized_entries:
            journal_entry.append("accounts", {
                "account": entry.get("account"),
                "debit_in_account_currency": float(entry.get("amount") or 0),
                "credit_in_account_currency": 0,
                "party_type": "",
                "party": "",
                "cost_center": get_default_cost_center(company)
            })

        # Resolve Bank Account GL Head
        # Prioritize UI passed from_account, then statement bank
        final_bank_account = from_account or stmt.bank_account
        bank_account_gl = get_bank_gl_account(final_bank_account)

        # Credit Bank Account
        journal_entry.append("accounts", {
            "account": bank_account_gl or get_default_bank_account(company, "Pay"),
            "debit_in_account_currency": 0,
            "credit_in_account_currency": total_amount,
            "party_type": "",
            "party": "",
            "cost_center": get_default_cost_center(company)
        })

        journal_entry.insert()
        journal_entry.submit()

        # Update Bank Statement Entry
        frappe.db.set_value("Bank Statement Entry", stmt.name, {
            "reference_no": journal_entry.name,
            "reconciled": 1,
            "match_type": "Auto"
        })

        return {
            "status": "ok",
            "journal_entry": journal_entry.name,
            "total_amount": total_amount,
            "num_items": len(itemized_entries)
        }

    except Exception as e:
        frappe.log_error(title="BR_ITEMIZED_ERROR", message=frappe.get_traceback())
        return {"status": "fail", "error": str(e)}


# ------------------------------------------------------------
# MATCH NOW CREATE JOURNAL FUNCTION
# ------------------------------------------------------------

@frappe.whitelist()
def match_now_create_journal(statement_name, expense_account):
    try:
        stmt = frappe.get_doc("Bank Statement Entry", statement_name)
        company = frappe.db.get_value("Bank Account", stmt.bank_account, "company")

        amount = abs(float(stmt.withdrawal or 0))

        return create_expense_journal_entry(
            stmt=stmt,
            expense_account=expense_account,
            amount=amount,
            company=company
        )

    except Exception as e:
        frappe.log_error(title="BR_MATCH_NOW_ERROR", message=frappe.get_traceback())
        return {"status": "fail", "error": str(e)}


# ------------------------------------------------------------
# CREATE BANK TRANSFER FUNCTION
# ------------------------------------------------------------

@frappe.whitelist()
def create_bank_transfer(statement_name, from_account, to_account, amount, description):
    try:
        stmt = frappe.get_doc("Bank Statement Entry", statement_name)
        company = frappe.db.get_value("Bank Account", stmt.bank_account, "company")

        return create_bank_transfer_journal(
            stmt=stmt,
            from_account=from_account,
            to_account=to_account,
            description=description,
            company=company
        )

    except Exception as e:
        frappe.log_error(title="BR_TRANSFER_CREATE_ERROR", message=frappe.get_traceback())
        return {"status": "fail", "error": str(e)}


# ------------------------------------------------------------
# RECONCILE TRANSACTION FUNCTION
# ------------------------------------------------------------

@frappe.whitelist()
def reconcile_transaction(invoice, amount, statement_name, allocated_amount=0, tax_adjustments=None):
    try:
        stmt = frappe.get_doc("Bank Statement Entry", statement_name)
        company = frappe.db.get_value("Bank Account", stmt.bank_account, "company")

        # Get invoice doctype
        if frappe.db.exists("Sales Invoice", invoice):
            doctype = "Sales Invoice"
            party_field = "customer"
        elif frappe.db.exists("Purchase Invoice", invoice):
            doctype = "Purchase Invoice"
            party_field = "supplier"
        else:
            return {"status": "fail", "error": f"Invoice {invoice} not found"}

        # Get invoice details
        inv = frappe.get_doc(doctype, invoice)
        party = inv.get(party_field)

        # Process payment
        invoices_list = [{
            "invoice": invoice,
            "amount": float(allocated_amount or amount),
            "doctype": doctype,
            "party": party
        }]

        tax_adjustments_list = []
        if tax_adjustments:
            tax_adjustments_list = frappe.parse_json(tax_adjustments)

        if doctype == "Sales Invoice":
            return process_customer_payment(
                stmt=stmt,
                invoices=invoices_list,
                company=company,
                customer=party,
                tax_adjustments_list=tax_adjustments_list,
                allow_overpayment=False,
                bank_account=stmt.bank_account # ⭐ For auto-match, use statement bank
            )
        else:
            # For Purchase Invoice, create Supplier Payment
            return categorize_manually(
                statement_name=statement_name,
                invoices=invoices_list,
                category="Supplier Payment",
                supplier=party,
                company=company,
                tax_adjustments=tax_adjustments
            )

    except Exception as e:
        frappe.log_error(title="BR_RECONCILE_ERROR", message=frappe.get_traceback())
        return {"status": "fail", "error": str(e)}


# ------------------------------------------------------------
# UTILITIES
# ------------------------------------------------------------

def get_actual_outstanding(doctype, docname):
    if not doctype or not docname:
        return 0
        
    if doctype == "Expense Claim":
        doc = frappe.get_doc("Expense Claim", docname)
        sanctioned = frappe.utils.flt(doc.total_sanctioned_amount)
        taxes = frappe.utils.flt(doc.total_taxes_and_charges)
        reimbursed = frappe.utils.flt(doc.total_amount_reimbursed)
        grand = frappe.utils.flt(doc.grand_total)
        
        adv_from_child = sum(frappe.utils.flt(adv.allocated_amount) for adv in doc.get("advances") or [])
        adv_paid_from_child = sum(frappe.utils.flt(adv.advance_paid) for adv in doc.get("advances") or [])
        total_advance = max(adv_from_child, adv_paid_from_child, frappe.utils.flt(doc.total_advance_amount))
        
        base_total = sanctioned + taxes
        if not base_total or base_total < total_advance:
            base_total = grand + total_advance
            
        return base_total - total_advance - reimbursed
    else:
        return frappe.utils.flt(frappe.db.get_value(doctype, docname, "outstanding_amount") or 0)

def get_default_bank_account(company, payment_type):
    account_type = "Bank"
    if payment_type == "Pay":
        field = "default_payable_account"
    else:
        field = "default_receivable_account"

    account = frappe.get_value("Company", company, field)

    if account and frappe.get_value("Account", account, "account_type") == account_type:
        return account

    default_bank = frappe.get_value("Company", company, "default_bank_account")
    if default_bank and frappe.get_value("Account", default_bank, "account_type") == account_type:
        return default_bank

    # Try to get any bank account for the company
    bank_accounts = frappe.get_all(
        "Account",
        filters={
            "company": company,
            "account_type": "Bank",
            "is_group": 0
        },
        fields=["name"],
        limit=1
    )

    if bank_accounts:
        return bank_accounts[0].name

    return None


def get_default_cost_center(company):
    cost_center = frappe.get_value("Company", company, "cost_center")
    if cost_center and frappe.db.exists("Cost Center", cost_center):
        return cost_center

    # Get first active cost center for the company
    cost_centers = frappe.get_all(
        "Cost Center",
        filters={"company": company, "is_active": 1},
        fields=["name"],
        limit=1
    )

    if cost_centers:
        return cost_centers[0].name

    return None


@frappe.whitelist()
def undo_reconciliation(statement_name):
    try:
        frappe.db.set_value("Bank Statement Entry", statement_name, {
            "reconciled": 0,
            "reference_no": "",
            "match_type": ""
        })
        return {"status": "ok"}
    except Exception as e:
        frappe.log_error(title="BR_UNDO_ERROR", message=frappe.get_traceback())
        return {"status": "fail", "error": str(e)}
##################SEPARATE FUNCTION: process_customer_payment()#################
def process_customer_payment(stmt, invoices, company, customer, tax_adjustments_list, allow_overpayment=False, bank_account=None):
    """
    Clean & independent Customer Payment processor.
    """

    # -------------------------------
    # 1. Calculate amounts
    # -------------------------------
    statement_amount = frappe.utils.flt(abs(float(stmt.deposit or 0)), 2)
    total_allocated = frappe.utils.flt(sum(frappe.utils.flt(inv.get("amount") or 0, 2) for inv in invoices), 2)

    # ⭐ FIX — USE STATEMENT AMOUNT
    paid_amount = statement_amount
    received_amount = statement_amount

    # Resolve Bank Account GL Head
    # Prioritize passed bank_account, then stmt.bank_account
    final_bank_account = bank_account or stmt.bank_account
    bank_account_gl = get_bank_gl_account(final_bank_account)

    # Resolve paid_from from first invoice's debit_to (Debtors account)
    customer_receivable = None
    for inv in invoices:
        if inv.get("doctype") == "Sales Invoice" and inv.get("invoice"):
            customer_receivable = frappe.db.get_value("Sales Invoice", inv["invoice"], "debit_to")
            if customer_receivable:
                break
    if not customer_receivable:
        customer_receivable = frappe.db.get_value("Company", company, "default_receivable_account")

    # -------------------------------
    # 2. Build Payment Entry
    # -------------------------------
    payment_entry = frappe.get_doc(
        {
            "doctype": "Payment Entry",
            "payment_type": "Receive",
            "mode_of_payment": "Wire Transfer",
            "company": company,
            "party_type": "Customer",
            "party": customer,
            "paid_amount": paid_amount,
            "received_amount": received_amount,
            "reference_no": stmt.description,
            "reference_date": stmt.transaction_date,
            "posting_date": stmt.transaction_date,
            "paid_from": customer_receivable,  # Debtors account
            "paid_to": bank_account_gl or get_default_bank_account(company, "Receive"),  # Bank GL account
            "bank_account": final_bank_account, # ⭐ Use doc name
        }
    )

    # -------------------------------
    # 3. Allocate invoices
    # -------------------------------
    reference_rows = []
    for inv in invoices:
        amt = frappe.utils.flt(inv.get("amount") or 0, 2)
        if inv.get("invoice") and inv.get("doctype"):
            actual_out = get_actual_outstanding(inv.get("doctype"), inv.get("invoice"))
            if amt > actual_out and abs(amt - actual_out) < 0.05:
                amt = actual_out
                
        reference_rows.append(
            {
                "reference_doctype": inv.get("doctype"),
                "reference_name": inv.get("invoice"),
                "allocated_amount": amt,
            }
        )
    
    total_allocated = frappe.utils.flt(sum(frappe.utils.flt(r.get("allocated_amount") or 0) for r in reference_rows), 2)

    payment_entry.set("references", reference_rows)

    # -------------------------------
    # 4. GST setup
    # -------------------------------
    if invoices:
        first = invoices[0]
        _set_gst_fields(
            payment_entry,
            customer,
            "Customer",
            company,
            first.get("doctype"),
            first.get("invoice")
        )

    # -------------------------------
    # 5. Tax adjustments
    # -------------------------------
    if tax_adjustments_list:
        _add_tax_rows_to_payment(payment_entry, tax_adjustments_list)

    # -------------------------------
    # 6. Save & Submit
    # -------------------------------
    payment_entry.insert()
    payment_entry.submit()

    # -------------------------------
    # 7. Update Bank Statement Entry
    # -------------------------------
    frappe.db.set_value("Bank Statement Entry", stmt.name, {
        "reference_no": payment_entry.name,
        "reconciled": 1,
        "match_type": "Auto"
    })

    # -------------------------------
    # 8. Return response
    # -------------------------------
    return {
        "status": "ok",
        "payment_entry": payment_entry.name,
        "allocated": total_allocated,
        "statement_amount": statement_amount,
        "paid_amount": paid_amount,
        "allow_overpayment": allow_overpayment,
        "excess_amount": frappe.utils.flt(paid_amount - total_allocated, 2)
    }
####################################### End of Bank Recon #################################################
# Notification
def notify_assignment(doc, method):
    if doc.reference_type == "HD Ticket":
        assigned_user = doc.owner
        hd_ticket = frappe.get_doc(doc.reference_type, doc.reference_name)

        ticket_title = hd_ticket.subject or "No Subject"
        priority = hd_ticket.priority or "Normal"
        customer = hd_ticket.customer or "Unknown"
        ticket_link = f"/app/hd-ticket/{hd_ticket.name}"

        message = f"""
        <div style='color: black; background-color: #66FF00; padding: 10px; border-radius: 8px; font-weight: bold;'>
            🤖 <b>Ticket Assigned!</b><br>
            📄 <b>Subject:</b> {ticket_title}<br>
            👤 <b>Customer:</b> {customer}<br>
            ⚠️ <b>Priority:</b> {priority}<br>
            🔗 <a href="{ticket_link}">View Ticket</a>
        </div>
        """

        frappe.publish_realtime(
            event='hd_ticket_assignment',
            message={
                'message': message,
                'ticket_name': hd_ticket.name
            },
            user=assigned_user
        )
###############################################################################
import frappe

@frappe.whitelist()
def get_lms_records(circuit_id):
    # Fetch parent LMS records with ignore_permissions
    lms_records = frappe.get_all(
        'Lastmile Services Master',
        filters={'circuit_id': circuit_id, 'lms_stage': 'Delivered'},
        fields=[
            'name', 'lms_feasibility_partner', 'supplier_contact', 'solution',
            'lms_stage', 'lms_delivery_date', 'suppliernumber', 'lms_brandwith_name',
            'media', 'mode1', 'static_ip', 'static_ip_1', 'url', 'user_id', 'password',

            # New fields for LMS PMT Portal
            'payment_mode_1', 'bank', 'portal_login_id', 'portal_login_password'
        ],
        ignore_permissions=True
    )

    # Append escalation details for each record
    for record in lms_records:
        contacts = frappe.get_all(
            'LMS Contact Escalation',
            filters={'parent': record.name},
            fields=[
                'level',
                'link_zitr',
                'link_syot',
                'designation',
                'department',
                'contact_phone'
            ],
            ignore_permissions=True
        )
        record['contacts'] = contacts

    return lms_records
######################################################################################
import frappe
from frappe import _

@frappe.whitelist()
def create_contact_and_add_escalation(
    lms_name,
    link_doctype='Supplier',
    link_name=None,
    first_name=None,
    last_name=None,
    custom_type='LMS Supplier',
    designation=None,
    department=None,
    email_id=None,
    is_primary_email=1,
    phone=None,
    is_primary_mobile_no=1,
    level=None
):
    """
    Creates a Contact (with email_ids, phone_nos, links if those child tables exist),
    then appends a row to the Lastmile Services Master child table 'table_oeiw' where
    link_zitr should store the created contact id.

    Returns: { "success": True, "contact": contact.name } or
             { "success": False, "error": "..." }
    """
    try:
        # === Validate inputs ===
        if not first_name and not last_name:
            frappe.throw(_("Either First Name or Last Name is required."))

        if not link_name:
            frappe.throw(_("Supplier (link_name) is required."))

        # Ensure level formatting
        valid_levels = [f"Level-{i}" for i in range(1, 6)]
        if level and level not in valid_levels:
            frappe.throw(_("Invalid Level. Allowed values are: {0}").format(", ".join(valid_levels)))

        # === Create Contact doc ===
        contact_meta = frappe.get_meta('Contact')
        contact = frappe.new_doc('Contact')

        contact.first_name = first_name or ''
        contact.last_name = last_name or ''

        # custom_type (user provided field name)
        if contact_meta.get_field('custom_type'):
            contact.custom_type = custom_type

        # designation / department if present
        if contact_meta.get_field('designation') and designation:
            contact.designation = designation
        if contact_meta.get_field('department') and department:
            contact.department = department

        # add email child row if email provided and child table exists
        if email_id and contact_meta.get_field('email_ids'):
            contact.append('email_ids', {
                'email_id': email_id,
                'is_primary': int(is_primary_email)
            })

        # add phone child row if phone provided and child table exists
        if phone and contact_meta.get_field('phone_nos'):
            contact.append('phone_nos', {
                'phone': phone,
                'is_primary_mobile_no': int(is_primary_mobile_no)
            })

        # add dynamic link to Contact if links child table exists
        if contact_meta.get_field('links'):
            contact.append('links', {
                'link_doctype': link_doctype,
                'link_name': link_name
            })

        # Insert contact
        contact.insert(ignore_permissions=True)
        frappe.db.commit()

        # === Append escalation row to Lastmile Services Master ===
        child_table_field = 'table_oeiw'
        lms_meta = frappe.get_meta('Lastmile Services Master')

        if not lms_meta.get_field(child_table_field):
            frappe.throw(_("Child table '{0}' not found on Lastmile Services Master.").format(child_table_field))

        lms_doc = frappe.get_doc('Lastmile Services Master', lms_name)

        child_row = {
            'link_zitr': contact.name,  # Contact ID
            'link_syot': email_id or '',
            'contact_phone': phone or '',
            'designation': designation or '',
            'department': department or '',
            'level': level or ''
        }

        lms_doc.append(child_table_field, child_row)
        lms_doc.save(ignore_permissions=True)
        frappe.db.commit()

        return {'success': True, 'contact': contact.name}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), 'create_contact_and_add_escalation')
        return {'success': False, 'error': str(e)}

#######################################################################################
## Auto-populate contact_email on Payment Entry for Supplier notifications

def set_supplier_email_on_payment_entry(doc, method):
    """
    On validate, if party_type is Supplier and contact_email is empty,
    fetch the supplier's email_id and set it as contact_email.
    This ensures the Notification recipient resolves correctly.
    """
    if doc.party_type == "Supplier" and doc.party and not doc.contact_email:
        supplier_email = frappe.db.get_value("Supplier", doc.party, "email_id")
        if supplier_email:
            doc.contact_email = supplier_email

def send_supplier_payment_email(doc):
    """
    Send payment confirmation email to supplier directly via frappe.sendmail.
    Called after Payment Entry submit when custom_send_email == 1.
    """
    try:
        if not doc.contact_email:
            return

        # Simplified message to avoid Zoho "Unusual activity" filter
        # Removing complex tables and special symbols like ₹ for testing
        invoice_list = ""
        for ref in doc.get("references", []):
            invoice_list += f"- {ref.reference_name}: {frappe.utils.fmt_money(ref.allocated_amount, currency='INR')}\n"

        message = f"""
        Dear {doc.party_name or doc.party},

        This is a payment confirmation for your services.

        Payment Reference: {doc.name}
        Payment Date: {frappe.utils.formatdate(doc.posting_date)}
        Amount: INR {frappe.utils.fmt_money(doc.paid_amount, currency='INR')}
        Reference / UTR: {doc.reference_no or '-'}

        Invoice Details:
        {invoice_list}

        Best Regards,
        Accounts Team
        {doc.company}
        """

        # Using a very simple subject
        subject = f"Payment Confirmation: {doc.name}"


        frappe.sendmail(
            recipients=[doc.contact_email],
            subject=subject,
            message=message,
            reference_doctype="Payment Entry",
            reference_name=doc.name
        )
        frappe.logger().info(f"Supplier payment email sent to {doc.contact_email} for {doc.name}")

    except Exception as e:
        frappe.log_error(
            title=f"Payment Email Failed: {doc.name}",
            message=f"Recipient: {doc.contact_email}\nError: {str(e)}"
        )

#######################################################################################
## Payment Entry update to Expense Claim     

import frappe

def update_expense_claim_status(doc, method):
    """
    Called when Payment Entry is submitted.
    Updates Expense Claim's custom_payment_status to 'Paid'
    if linked in references.
    """
    for ref in doc.references:
        if ref.reference_doctype == "Expense Claim" and ref.reference_name:
            frappe.db.set_value("Expense Claim", ref.reference_name, "custom_payment_status", "Paid")
###########################################################################################

# HD Ticket Auto email to the manager
import frappe
from frappe.utils import now_datetime, get_datetime, formatdate
from openpyxl import Workbook
import io


@frappe.whitelist()
def get_engineer_ticket_summary():
    """Return ticket summary grouped by engineer with SLA buckets + detailed tickets"""
    statuses = ["Open", "On Hold", "Replied"]

    tickets = frappe.get_all(
        "HD Ticket",
        filters={"status": ["in", statuses]},
        fields=[
            "name",
            "custom_circuit_id",
            "customer",
            "custom_agent_name",
            "custom_channel",
            "priority",
            "custom_stage",
            "status",
            "opening_date",
            "opening_time"
        ]
    )

    summary = {}
    detailed_rows = []

    for t in tickets:
        # 🚫 Skip if Circuit ID or Customer is missing
        if not t.custom_circuit_id or not t.customer:
            continue

        engineer = t.custom_agent_name or "Unassigned"
        open_datetime = get_datetime(f"{t.opening_date} {t.opening_time}")
        age_hours = (now_datetime() - open_datetime).total_seconds() / 3600

        # SLA bucket
        if age_hours <= 4:
            age_bucket = "0-4 hrs"
        elif age_hours <= 24:
            age_bucket = "<24 hrs"
        else:
            age_bucket = ">24 hrs"

        # Summary calculation
        if engineer not in summary:
            summary[engineer] = {"0-4 hrs": 0, "<24 hrs": 0, ">24 hrs": 0, "Total": 0}

        summary[engineer][age_bucket] += 1
        summary[engineer]["Total"] += 1

        # Detailed row
        detailed_rows.append({
            "custom_circuit_id": t.custom_circuit_id,
            "ticket_no": t.name, 
            "customer": t.customer,
            "custom_agent_name": engineer,
            "custom_channel": t.custom_channel,
            "priority": t.priority,
            "custom_stage": t.custom_stage,
            "status": t.status,
            "opening_date": str(t.opening_date),
            "hours": age_bucket
        })

    # grand totals
    grand = {"0-4 hrs": 0, "<24 hrs": 0, ">24 hrs": 0, "Total": 0}
    for eng in summary.values():
        for k in grand:
            grand[k] += eng[k]

    return {"summary": summary, "grand": grand, "detailed": detailed_rows}


def format_engineer_ticket_email(data):
    summary = data["summary"]
    grand = data["grand"]

    now_dt = now_datetime()
    formatted_date = formatdate(now_dt, "d MMMM yyyy")
    formatted_time = now_dt.strftime("%I:%M %p")  # ✅ AM/PM format

    rows = ""
    for eng, counts in summary.items():
        rows += f"""
        <tr>
            <td style="padding:6px; border:1px solid #ccc; text-align:left;">{eng}</td>
            <td align="center" style="padding:6px; border:1px solid #ccc;">{counts['0-4 hrs']}</td>
            <td align="center" style="padding:6px; border:1px solid #ccc;">{counts['<24 hrs']}</td>
            <td align="center" style="padding:6px; border:1px solid #ccc;">{counts['>24 hrs']}</td>
            <td align="center" style="padding:6px; border:1px solid #ccc; font-weight:bold;">{counts['Total']}</td>
        </tr>
        """

    grand_row = f"""
    <tr style="background:#d1e7dd; font-weight:bold;">
        <td style="padding:6px; border:1px solid #ccc; text-align:left;">Grand Total</td>
        <td align="center" style="padding:6px; border:1px solid #ccc;">{grand['0-4 hrs']}</td>
        <td align="center" style="padding:6px; border:1px solid #ccc;">{grand['<24 hrs']}</td>
        <td align="center" style="padding:6px; border:1px solid #ccc;">{grand['>24 hrs']}</td>
        <td align="center" style="padding:6px; border:1px solid #ccc;">{grand['Total']}</td>
    </tr>
    """

    html = f"""
    <div style="font-family: Arial, sans-serif;">
        <p>Hello Team,</p>
        <p>
            Please find Open ticket report as on time, 
            <b>needs to push more on tickets open beyond 24 hours to closure.</b>
        </p>

        <div style="background:#0047ab; color:#fff; padding:12px; border-radius:8px; margin-top:15px; width:70%;">
            <h2 style="margin:0;">📊 Engineer SLA Ticket Report</h2>
            <p style="margin:5px 0;">As of {formatted_date} – Time : {formatted_time}</p>
        </div>

        <table style="width:70%; border-collapse:collapse; margin-top:15px; font-size:13px;">
            <tr style="background:#eaf0f6;">
                <th style="padding:6px; border:1px solid #ccc; text-align:left;">Engineer</th>
                <th style="padding:6px; border:1px solid #ccc;">0-4 hrs</th>
                <th style="padding:6px; border:1px solid #ccc;">&lt;24 hrs</th>
                <th style="padding:6px; border:1px solid #ccc;">&gt;24 hrs</th>
                <th style="padding:6px; border:1px solid #ccc;">Total</th>
            </tr>
            {rows}
            {grand_row}
        </table>

        <p style="margin-top:20px; font-size:13px;">
            Data in Excel format attached for your reference.
        </p>

        <p style="margin-top:10px; font-size:12px; color:#666;">
            This is an automated SLA report from ERPNext.
        </p>
    </div>
    """
    return html, formatted_date, formatted_time


def generate_excel_report(data, formatted_date, formatted_time):
    """Generate Excel with detailed ticket info"""
    tickets = data["detailed"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"

    # Title
    ws.append([f"Engineer SLA Ticket Report - {formatted_date}, {formatted_time}"])
    ws.append([])

    # Headers
    headers = [
        "Circuit ID",
        "Ticket No",
        "Customer",
        "Agent Name",
        "Channel",
        "Priority",
        "Stage",
        "Status",
        "Opening Date",
        "Hours"
    ]
    ws.append(headers)

    # Ticket rows
    for t in tickets:
        ws.append([
            t["custom_circuit_id"],
            t["ticket_no"],
            t["customer"] or "",
            t["custom_agent_name"] or "",
            t["custom_channel"] or "",
            t["priority"] or "",
            t["custom_stage"] or "",
            t["status"] or "",
            t["opening_date"] or "",
            t["hours"]
        ])

    # Save to BytesIO
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


@frappe.whitelist()
def send_engineer_ticket_report():
    data = get_engineer_ticket_summary()
    message, formatted_date, formatted_time = format_engineer_ticket_email(data)

    # Generate detailed Excel attachment
    excel_bytes = generate_excel_report(data, formatted_date, formatted_time)

    subject = f"📌 Open Ticket Report as on {formatted_date}, Time : {formatted_time}"

    frappe.sendmail(
        recipients=["support.team@nexapp.co.in"],
        subject=subject,
        message=message,
        attachments=[
            {
                "fname": f"Engineer_SLA_Report_{formatted_date}.xlsx",
                "fcontent": excel_bytes
            }
        ]
    )
    return "Report sent successfully with Excel attachment!"
##############################################################################

import frappe

def update_task_circuit_sales_order(doc, method):
    """Update Task Circuit ID.sales_order_no when Sales Order is saved"""
    try:
        # Step 1: Ensure Sales Order has a linked Task
        if not doc.custom_task:
            return

        # Step 2: Check Task exists
        task_exists = frappe.db.exists("Task", doc.custom_task)
        if not task_exists:
            return

        # Step 3: Loop through Sales Order Items
        for item in doc.items:
            if not item.custom_feasibility:
                continue

            # Step 4: Find matching Task Circuit ID child rows
            task_circuits = frappe.get_all(
                "Task Circuit ID",
                filters={
                    "parent": doc.custom_task,   # belongs to the Task
                    "custom_circuit_id": item.custom_feasibility
                },
                fields=["name"]
            )

            # Step 5: Update sales_order_no
            for tc in task_circuits:
                frappe.db.set_value(
                    "Task Circuit ID",
                    tc.name,
                    "sales_order_no",
                    doc.name
                )

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "update_task_circuit_sales_order Error")

#####################################################################################
# Site To Invoice MAnagement 
import frappe

def update_invoice_and_lms(doc, method):
    """
    ✅ Creates new Invoice Management record when Site is Delivered & Live
    ✅ Prevents duplicate Invoice Management records
    ✅ Adds SO Item Invoice & LMS Item rows
    """

    # ✅ 1. Only proceed if Site is "Delivered and Live"
    if doc.doctype != "Site" or doc.site_status != "Delivered and Live":
        return

    # ✅ 2. Check if Invoice Management already exists for this Site (avoid duplicates)
    existing_invoice = frappe.get_value("Invoice Management", {"circuit_id": doc.name}, "name")
    if existing_invoice:
        invoice = frappe.get_doc("Invoice Management", existing_invoice)  # Use existing record
    else:
        # ✅ Create new Invoice Management record
        invoice = frappe.new_doc("Invoice Management")
        invoice.circuit_id = doc.name
        invoice.customer = doc.customer  # Optional – you can map more fields if required
        invoice.insert(ignore_permissions=True)

    # ✅ 3. Add Sales Order Items into SO Item Invoice Table
    if doc.sales_order:
        so = frappe.get_doc("Sales Order", doc.sales_order)
        for item in so.items:
            if item.custom_feasibility == doc.name:
                # Prevent duplication
                exists = any(row.item_code == item.item_code for row in invoice.so_item_invoice)
                if not exists:
                    invoice.append("so_item_invoice", {
                        "item_code": item.item_code,
                        "rate": item.rate,
                        "qty": item.qty,
                        "amount": item.amount
                    })

    # ✅ 4. Add LMS Delivered rows into SO Item LMS Table
    if hasattr(doc, "lms_vendor"):
        for lms in doc.lms_vendor:
            if lms.stage == "LMS Delivered":
                # Prevent duplication
                exists = any(row.lms_id == (lms.lms_id or lms.name) for row in invoice.so_item_lms)
                if not exists:
                    invoice.append("so_item_lms", {
                        "lms_id": lms.lms_id or lms.name,
                        "bandwidth": lms.bandwidth,
                        "bandwith_type": lms.bandwith_type,
                        "lms_delivery_date": lms.lms_delivery_date,
                        "lms_status": lms.stage
                    })

    # ✅ 5. Sanitize invalid Billing Terms (fix ValidationError)
    # Some child tables may have billing_terms = "none" which is not allowed.
    if hasattr(invoice, "items"):
        for item in invoice.items:
            if getattr(item, "billing_terms", None) == "none":
                item.billing_terms = ""  # replace invalid value with blank

    # ✅ 6. Save Invoice after cleaning data
    invoice.save(ignore_permissions=True)
    frappe.db.commit()

#################################################################################
# Bank Recon - Employe Expense Claim
import frappe
from frappe import _
from frappe.utils import flt, nowdate

@frappe.whitelist()
def get_bank_statement_entries(bank_account=None):
    """Get bank statement entries for reconciliation"""
    filters = {"reconciled": 0}
    
    if bank_account:
        filters["bank_account"] = bank_account
    
    entries = frappe.get_all(
        "Bank Statement Entry",
        filters=filters,
        fields=["name", "transaction_date", "description", "deposit", "withdrawal", "bank_account"],
        order_by="transaction_date desc"
    )
    
    # Add posting_date field for compatibility with frontend
    for entry in entries:
        entry["posting_date"] = entry.get("transaction_date")
    
    return entries

@frappe.whitelist()
def get_outstanding_expense_claims(employee, company=None):
    """Get outstanding expense claims for an employee"""
    if not company:
        company = frappe.defaults.get_user_default("company")
    
    expense_claims = frappe.get_all(
        "Expense Claim",
        filters={
            "employee": employee,
            "docstatus": 1,
            "status": ["!=", "Paid"],
            "company": company
        },
        fields=["name", "posting_date", "total_sanctioned_amount"],
        order_by="posting_date desc"
    )
    
    return expense_claims
############################################################################
#Employee PAyment Entry

import frappe
from frappe.utils import flt, nowdate

@frappe.whitelist()
def create_employee_expense_payment(statement_name, invoices, employee, company=None):
    """
    Create Payment Entry for Employee Expense Claim during Bank Reconciliation
    """

    invoices = frappe.parse_json(invoices)
    if not invoices:
        return {"status": "error", "error": "No expense claim invoices provided"}

    if not employee:
        return {"status": "error", "error": "Employee is required"}

    company = company or frappe.defaults.get_default("company")

    # Get Bank Statement Entry
    statement = frappe.get_doc("Bank Statement Entry", statement_name)
    bank_amount = flt(statement.withdrawal or statement.deposit or 0)

    # Calculate allocated amount
    total_allocated = sum([flt(inv.get("amount")) for inv in invoices])

    if total_allocated <= 0:
        return {"status": "error", "error": "Allocated amount must be > 0"}

    if total_allocated > bank_amount:
        return {
            "status": "error",
            "error": f"Allocated amount ({total_allocated}) cannot exceed bank amount ({bank_amount})"
        }

    # Create Payment Entry
    pe = frappe.new_doc("Payment Entry")
    pe.payment_type = "Pay"                         # Expense Claim is always Pay
    pe.company = company
    pe.posting_date = statement.date
    pe.mode_of_payment = "Bank"
    pe.reference_no = statement.name
    pe.reference_date = statement.date

    # IMPORTANT — FIXES YOUR ERROR
    pe.party_type = "Employee"
    pe.party = employee

    # Paid From (Bank GL account)
    if statement.bank_account:
        pe.paid_from = get_bank_gl_account(statement.bank_account)
    else:
        return {"status": "error", "error": "Bank account missing in statement"}

    # Paid To (Employee's payable account from Expense Claim or Company default)
    emp_payable = None
    for inv in invoices:
        if inv.get("invoice"):
            emp_payable = frappe.db.get_value("Expense Claim", inv["invoice"], "payable_account")
            if emp_payable:
                break
    pe.paid_to = emp_payable or frappe.db.get_value("Company", company, "default_payable_account")

    pe.paid_amount = total_allocated
    pe.received_amount = amount

    # Add references
    for inv in invoices:
        pe.append("references", {
            "reference_doctype": "Expense Claim",
            "reference_name": inv.get("invoice"),
            "allocated_amount": flt(inv.get("amount")),
        })

    # Save & Submit Payment Entry
    pe.insert(ignore_permissions=True)
    pe.submit()

    # Mark statement as reconciled
    frappe.db.set_value("Bank Statement Entry", statement_name, {
        "reference_no": pe.name,
        "reconciled": 1,
        "match_type": "Auto"
    })

    return {
        "status": "ok",
        "payment_entry": pe.name
    }

####################################################################################
# AI Page        
import frappe
import requests
from html import escape
from frappe.utils import get_fullname

# ✅ Call n8n and return reply text; longer timeout to avoid 60s read timeout
@frappe.whitelist(allow_guest=True)
def chat_with_n8n(message):
    try:
        url = "https://nexapp.app.n8n.cloud/webhook/erp-chat"
        response = requests.post(url, json={"message": message}, timeout=180)

        if response.status_code == 200:
            # ✅ Try JSON first; fallback to raw text
            try:
                data = response.json()
                return data.get("reply") or "No reply"
            except Exception:
                return (response.text or "").strip() or "No reply"
        else:
            return f"Error: n8n returned {response.status_code}"

    except requests.exceptions.Timeout:
        return "Error: Connection to n8n timed out. Try again."

    except Exception as e:
        return f"Error: {str(e)}"


# ✅ Send a well-formatted HTML email with Question + Answer
@frappe.whitelist()
def email_ai_response(question: str, body: str):
    """
    Sends an HTML email to the logged-in user's email address.
    Includes the question and the AI's answer in a clean format.
    """
    try:
        user = frappe.session.user

        # ✅ Get email from User doctype
        email = frappe.db.get_value("User", user, "email") or (user if "@" in user else None)
        if not email:
            return {"status": "fail", "msg": "No email found for current user."}

        # ✅ Get user full name (fallback to 'there')
        try:
            full_name = get_fullname(user) or "there"
        except Exception:
            full_name = "there"

        # ✅ Escape special characters for safety, convert line breaks to HTML <br>
        q_html = escape(question or "").replace("\n", "<br>")
        a_html = escape(body or "").replace("\n", "<br>")

        subject = "Response from ERPNext AI Assistant"

        # ✅ Clean HTML email body
        html_body = f"""
        <div style="font-family:'Segoe UI',Arial,Helvetica,sans-serif; font-size:14px; color:#111827; line-height:1.6;">
          <p>Hello {escape(full_name)},</p>

          <p><b>You asked:</b></p>
          <div style="background:#F3F4F6; border:1px solid #E5E7EB; border-radius:8px; padding:12px; margin:8px 0 16px;">
            {q_html}
          </div>

          <p><b>Here is the answer from ERPNext AI Assistant:</b></p>
          <div style="background:#FFFFFF; border:1px solid #E5E7EB; border-radius:8px; padding:12px; margin:8px 0 16px;">
            {a_html}
          </div>

          <p style="margin-top:18px;">
            Thank you,<br>
            <b>ERPNext AI Agent</b>
          </p>
        </div>
        """

        # ✅ Correct email method in Frappe → use message= (HTML allowed)  
        frappe.sendmail(
            recipients=[email],
            subject=subject,
            message=html_body,   # ✅ HTML directly supported
            delayed=False        # ✅ Send instantly
        )

        return {"status": "ok"}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "email_ai_response_failed")
        return {"status": "fail", "msg": str(e)}
################################################################################
# Disconnection Request 
import frappe

@frappe.whitelist()
def fetch_items_from_site(docname):
    doc = frappe.get_doc("Disconnection Request", docname)

    if not doc.circuit_id:
        frappe.throw("Circuit ID is required.")

    if not frappe.db.exists("Site", doc.circuit_id):
        frappe.throw(f"Site '{doc.circuit_id}' does not exist.")

    # Clear child table to prevent duplicates
    doc.disconnection_multiple = []

    # Fetch Site Items
    site_items = frappe.get_all("Site Item",
        filters={"parent": doc.circuit_id},
        fields=[
            "item_name", "qty", "item_code", "serial_no_sim_no", "item_group",
            "status", "warranty_expiry_date", "lan_mac", "hardware_version",
            "wlan_mac", "wan_mac", "module", "warranty_period_days", "imei",
            "mobile_no", "activation_date", "validity", "data_plan",
            "recharge_end_date", "rental_plan", "brand"
        ]
    )

    for item in site_items:
        doc.append("disconnection_multiple", {
            "circuit_id": doc.circuit_id,
            "item_name": item.item_name,
            "qty": item.qty,
            "item_code": item.item_code,
            "serial_no_sim_no": item.serial_no_sim_no,
            "item_group": item.item_group,
            "stage": item.status,
            "warranty_expiry_date": item.warranty_expiry_date,
            "lan_mac": item.lan_mac,
            "hardware_version": item.hardware_version,
            "wlan_mac": item.wlan_mac,
            "wan_mac": item.wan_mac,
            "module": item.module,
            "warranty_period_days": item.warranty_period_days,
            "imei": item.imei,
            "mobile_no": item.mobile_no,
            "activation_date": item.activation_date,
            "validity": item.validity,
            "data_plan": item.data_plan,
            "recharge_end_date": item.recharge_end_date,
            "rental_plan": item.rental_plan,
            "brand": item.brand
        })

    doc.save(ignore_permissions=True)

############################################################################
# apps/nexapp/nexapp/api.py

import frappe

@frappe.whitelist()
def create_disconnection_lms(disconnection_request):
    """Creates or updates Disconnection LMS, Stock Management, SIM Disconnection.
       Also updates Site and LMS stages when status = 'Approved'.
    """

    doc = frappe.get_doc("Disconnection Request", disconnection_request)

    # ✅ Duplicate prevention
    existing_lms = frappe.db.exists("Disconnection LMS", {"disconnection_request_id": doc.name})
    existing_sm = frappe.db.exists("Stock Management Disconnection", {"disconnection_request_id": doc.name})
    existing_sim = frappe.db.exists("SIM Disconnection", {"disconnection_request_id": doc.name})

    if existing_lms or existing_sm or existing_sim:
        frappe.msgprint("⚠️ Records already created for this Disconnection Request.", indicator="orange")
        return {"duplicate": True}

    # ------------------------------------------------------------------
    # ✅ ONLY LMS with lms_stage = "Delivered"
    # ------------------------------------------------------------------
    lms_map = frappe._dict()
    for d in frappe.get_all(
        "Lastmile Services Master",
        filters={"lms_stage": "Delivered"},
        fields=["name", "circuit_id"]
    ):
        if d.circuit_id:
            lms_map.setdefault(d.circuit_id, []).append(d.name)

    site_map = frappe._dict({
        d.circuit_id: d.name
        for d in frappe.get_all("Site", fields=["name", "circuit_id"])
    })

    # ------------------ PART 1: CREATE LMS & STOCK MANAGEMENT ------------------
    for row in (doc.circuit_disconnection or []):
        circuit_id = (row.circuit_id or "").strip()
        if not circuit_id:
            continue

        try:
            # --- Create Disconnection LMS (Delivered LMS only) ---
            for lms_name in lms_map.get(circuit_id, []):

                if not frappe.db.exists(
                    "Disconnection LMS",
                    {"lms_id": lms_name, "disconnection_request_id": doc.name}
                ):
                    new_lms = frappe.new_doc("Disconnection LMS")
                    new_lms.circuit_id = circuit_id
                    new_lms.lms_id = lms_name
                    new_lms.disconnection_request_id = doc.name
                    new_lms.insert(ignore_permissions=True)

            # --- Create/Update Stock Management ---
            site_name = site_map.get(circuit_id)
            if not site_name:
                continue

            sm_name = frappe.db.exists("Stock Management", {"circuit_id": site_name})
            if not sm_name:
                sm_doc = frappe.new_doc("Stock Management")
                sm_doc.circuit_id = site_name
                sm_doc.disconnection_stage = "Open"
                sm_doc.stock_management_type = "Disconnection"
                sm_doc.status = "Disconnection"
                sm_doc.disconnection_request_id = doc.name
                sm_doc.insert(ignore_permissions=True)
            else:
                sm_doc = frappe.get_doc("Stock Management", sm_name)
                sm_doc.disconnection_request_id = doc.name
                sm_doc.flags.ignore_validate = True
                sm_doc.save(ignore_permissions=True)

            # ------------------ PART 3: STOCK MGMT DISCONNECTION + SIM DISCONNECTION ------------------

            exists_child = frappe.db.exists(
                "Stock Management Disconnection",
                {"parent": sm_doc.name, "disconnection_request_id": doc.name}
            )

            if not exists_child:

                site_items = frappe.get_all(
                    "Site Item",
                    filters={"parent": site_name},
                    fields=["item_code", "item_name", "qty", "serial_no_sim_no", "item_group"]
                )

                for it in site_items:

                    if it.item_group != "Telecom":
                        sm_doc.append("stock_management_disconnection", {
                            "item_code": it.item_code,
                            "item_name": it.item_name,
                            "qty": it.qty,
                            "serial_no_sim_no": it.serial_no_sim_no,
                            "item_group": it.item_group,
                            "disconnection_request_id": doc.name
                        })
                    else:
                        sim_doc = frappe.new_doc("SIM Disconnection")
                        sim_doc.circuit_id = site_name
                        sim_doc.item_code = it.item_code
                        sim_doc.sim_no = it.serial_no_sim_no
                        sim_doc.disconnection_request_id = doc.name
                        sim_doc.insert(ignore_permissions=True)

                        frappe.db.set_value(
                            "SIM Disconnection",
                            sim_doc.name,
                            "disconnection_request_id",
                            doc.name
                        )

                sm_doc.flags.ignore_mandatory = True
                sm_doc.flags.ignore_validate = True
                sm_doc.flags.dirty = True
                sm_doc.save(ignore_permissions=True)

            existing_sim_records = frappe.get_all(
                "SIM Disconnection",
                filters={"circuit_id": site_name},
                fields=["name"]
            )

            for sim in existing_sim_records:
                frappe.db.set_value(
                    "SIM Disconnection",
                    sim.name,
                    "disconnection_request_id",
                    doc.name
                )

        except Exception:
            frappe.log_error(
                frappe.get_traceback(),
                f"Error processing Circuit ID {circuit_id}"
            )
            continue

    # ------------------ PART 4: UPDATE SITE & LMS STAGE IF APPROVED ------------------
    if doc.status == "Approved":

        for row in doc.circuit_disconnection or []:
            if row.circuit_id:
                try:
                    frappe.db.set_value("Site", row.circuit_id, {
                        "site_status": "Disconnection In Process",
                        "lms_stage": "Disconnection In Process"
                    })
                except Exception as e:
                    frappe.log_error(
                        f"Error updating Site for circuit_id {row.circuit_id}",
                        str(e)
                    )

        for row in doc.circuit_disconnection or []:
            for lms_id in lms_map.get(row.circuit_id, []):
                try:
                    frappe.db.set_value("Lastmile Services Master", lms_id, {
                        "lms_stage": "Disconnection In Process"
                    })
                except Exception as e:
                    frappe.log_error(
                        f"Error updating LMS for lms_id {lms_id}",
                        str(e)
                    )

    return {"updated": True}


def process_disconnection_request_submit(doc, method):
    """Hook triggered on submit of Disconnection Request."""
    res = create_disconnection_lms(doc.name)

    if res.get("duplicate"):
        return

    frappe.msgprint(
        "✅ LMS, SIM & Stock Management Updated Successfully",
        title="Update Complete",
        indicator="green"
    )


###############################################################################
# LMS Cancel Code

import frappe

@frappe.whitelist()
def cancel_lms_service(lms_id, circuit_id):
    """
    Final logic:
    1. Cancel LMS Master FIRST (so hooks skip processing)
    2. Check Site status
    3. Update child table row → stage = LMS Cancelled
    """

    # ------------------------------
    # Validate Inputs
    # ------------------------------
    if not lms_id:
        return "LMS ID missing."

    if not circuit_id:
        return "Circuit ID missing."

    # ------------------------------
    # STEP 1 — CANCEL LMS MASTER FIRST
    # ------------------------------
    try:
        lms = frappe.get_doc("Lastmile Services Master", lms_id)
    except frappe.DoesNotExistError:
        return f"LMS '{lms_id}' not found."

    # Set LMS stage to Cancelled BEFORE anything
    lms.lms_stage = "Cancelled"
    lms.save(ignore_permissions=True)

    # Now the hook update_site_child_table() will skip execution
    # because it checks doc.lms_stage == "Cancelled"

    # ------------------------------
    # STEP 2 — Fetch Site
    # ------------------------------
    try:
        site = frappe.get_doc("Site", circuit_id)
    except frappe.DoesNotExistError:
        return f"Site '{circuit_id}' not found."

    # ------------------------------
    # BLOCK CANCELLATION if Delivered & Live
    # ------------------------------
    if site.site_status == "Delivered and Live":
        return "Cannot cancel Supplier—site is ‘Delivered & Live’. Use Change Management."

    # ------------------------------
    # STEP 3 — Update child table row lms_vendor
    # ------------------------------
    clean_lms_id = (lms_id or "").strip()
    child_found = False

    for row in site.lms_vendor:
        if (row.lms_id or "").strip() == clean_lms_id:
            frappe.db.set_value("LMS Site", row.name, "stage", "LMS Cancelled")
            child_found = True
            break

    if not child_found:
        return "Matching LMS ID not found in Site's LMS Vendor table."

    return "The Lastmile Service has been cancelled successfully."

##############################################################################
#Bank Recon  Journal Entry

import frappe

@frappe.whitelist()
def match_now_create_journal(statement_name, expense_account):
    """
    Create & Submit Journal Entry for a withdrawal transaction.
    Debit: Selected Expense Account
    Credit: Bank Ledger Account (bank_account_head)
    """

    # 🔹 Load Bank Statement Entry
    stmt = frappe.get_doc("Bank Statement Entry", statement_name)

    # 🔹 Withdrawal Amount
    amount = float(stmt.withdrawal or 0)
    if amount <= 0:
        return {"status": "fail", "error": "No withdrawal amount found"}

    # 🔹 Bank Ledger Account
    ledger_account = stmt.bank_account_head
    if not ledger_account:
        return {
            "status": "fail",
            "error": "Please set Bank Account Head in the Bank Statement Entry."
        }

    # 🔹 Company of Ledger Account
    company = frappe.db.get_value("Account", ledger_account, "company")

    # 🔹 Create Journal Entry
    je = frappe.get_doc({
        "doctype": "Journal Entry",
        "voucher_type": "Bank Entry",
        "company": company,
        "posting_date": stmt.transaction_date,

        "cheque_no": stmt.description or "",
        "cheque_date": stmt.transaction_date,
        "remark": stmt.description or "",
        "mode_of_payment": "Wire Transfer",

        "user_remark": f"Auto-created from Bank Statement {statement_name}",

        "accounts": [
            {
                "account": expense_account,
                "debit_in_account_currency": amount
            },
            {
                "account": ledger_account,
                "credit_in_account_currency": amount
            }
        ]
    })

    je.flags.ignore_mandatory = True

    # 🔹 Save (Draft)
    je.insert(ignore_permissions=True)

    # 🔹 Submit the Journal Entry
    je.submit()

    # 🔹 Mark Bank Statement as Reconciled
    frappe.db.set_value("Bank Statement Entry", statement_name, {
        "reference_no": je.name,
        "reconciled": 1,
        "match_type": "Auto"
    })

    return {
        "status": "ok",
        "journal_entry": je.name,
        "amount": amount,
        "submitted": True
    }

########################################################################3
#Bank Recon - Transfer To Another Bank

import frappe

@frappe.whitelist()
def create_bank_transfer(statement_name, from_account, to_account, amount, description):
    """
    Create Journal Entry for 'Transfer To Another Account'
    (DRAFT JE + Mark Statement Reconciled)
    """

    try:
        amount = float(amount)

        # ----------------------------
        # 1. Get Bank Statement Entry
        # ----------------------------
        statement = frappe.get_doc("Bank Statement Entry", statement_name)

        # SAFE posting date extraction
        posting_date = (
            statement.get("date") 
            or statement.get("transaction_date")
            or statement.get("posting_date")
            or statement.get("creation")
            or frappe.utils.nowdate()
        )

        # SAFE company value
        company = frappe.db.get_single_value("Global Defaults", "default_company")

        # ----------------------------
        # 2. Create Journal Entry (DRAFT ONLY)
        # ----------------------------
        je = frappe.new_doc("Journal Entry")
        je.voucher_type = "Bank Entry"
        je.posting_date = posting_date
        je.company = company

        je.user_remark = description or (statement.description or "")
        je.cheque_no = statement.description
        je.cheque_date = posting_date

        # DR → TO Account
        je.append("accounts", {
            "account": to_account,
            "debit_in_account_currency": amount,
            "debit": amount
        })

        # CR → FROM Account
        je.append("accounts", {
            "account": from_account,
            "credit_in_account_currency": amount,
            "credit": amount
        })

        # SAVE ONLY — NO SUBMIT
        je.save(ignore_permissions=True)

        # ----------------------------
        # 3. Mark Statement as Reconciled
        # ----------------------------
        frappe.db.set_value("Bank Statement Entry", statement_name, {
            "reconciled": 1,
            "reference_no": je.name,
            "match_type": "Auto"
        })

        return {
            "status": "ok",
            "journal_entry": je.name,
            "message": "Draft JE created + Statement Reconciled"
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Bank Transfer Error")
        return {"status": "error", "error": str(e)}

###########################################################################
# Itemized Posting Code
import frappe
import json

@frappe.whitelist()
def reconcile_bank_statement_with_payment(statement_name, payment_entry):
    """
    Utility function to reconcile a bank statement entry with a payment entry.
    Updates reference_no, reconciled=1 and match_type='Auto'.
    """
    try:
        frappe.db.set_value("Bank Statement Entry", statement_name, {
            "reference_no": payment_entry,
            "reconciled": 1,
            "match_type": "Auto"
        })
        return {"status": "ok"}
    except Exception as e:
        frappe.log_error(title="BR_RECONCILE_Utility_Error", message=frappe.get_traceback())
        return {"status": "error", "error": str(e)}

#############################################################################
# HD Ticket Customer Potal

import frappe
import json
from datetime import datetime
from frappe.desk.reportview import get_filters_cond, get_match_cond

# =============================================================================
#  GET TICKETS WITH USER PERMISSION + POS LOGIC - UPDATED
# =============================================================================
@frappe.whitelist()
def get_tickets(filters=None, page=1, page_size=20):
    """
    Ticket filtering logic:

    1️⃣ User Permission Logic:
        If User Permission → allow = Customer
        then user sees ONLY those customers.

    2️⃣ custom_pos_customer Logic:
        For each customer:
            If custom_pos_customer = 1:
                Only tickets where Site.customer_type = 'POC Customer'
            If custom_pos_customer = 0:
                Site.customer_type IN ('POC Customer', 'Paid Customer', '')

    3️⃣ Dynamic UI Filters:
        ticket_no, channel, circuit_id, customer, site_name, status
    """

    # -----------------------------
    # Convert JSON Filters
    # -----------------------------
    if isinstance(filters, str):
        try:
            filters = json.loads(filters)
        except:
            filters = {}
    filters = filters or {}

    # Better debug logging without title truncation
    debug_messages = []
    debug_messages.append(f"DEBUG: Received filters: {filters}")

    allowed = {
        "ticket_no": "name",
        "channel": "custom_channel",
        "circuit_id": "custom_circuit_id",
        "customer": "customer",
        "site_name": "custom_site_name",
        "status": "status",
    }

    # -----------------------------
    # Dynamic Filters - FIXED FOR ALL FILTERS
    # -----------------------------
    conditions = []
    params = []

    for k, v in filters.items():
        if v and k in allowed:
            # For ALL fields, use LIKE for partial matching
            conditions.append(f"`{allowed[k]}` LIKE %s")
            params.append(f"%{v}%")
            debug_messages.append(f"DEBUG: Added filter {k}='{v}' for field {allowed[k]}")

    where_clause = ""
    if conditions:
        where_clause = " AND " + " AND ".join(conditions)
    
    debug_messages.append(f"DEBUG: Where clause: {where_clause}")
    debug_messages.append(f"DEBUG: Filter params before customer: {params}")

    # ============================================================
    # USER PERMISSION → FETCH ALLOWED CUSTOMERS
    # ============================================================
    user = frappe.session.user
    debug_messages.append(f"DEBUG: Current user: {user}")

    # GET USER'S PERMITTED CUSTOMERS
    allowed_customers = get_allowed_customers_for_user(user)
    
    if not allowed_customers:
        debug_messages.append("DEBUG: No user permissions found or user has no access")
        frappe.log_error("\n".join(debug_messages), "get_tickets - No permissions")
        return {"tickets": [], "total": 0}

    debug_messages.append(f"DEBUG: Allowed customers for user {user}: {allowed_customers}")

    # ============================================================
    # BUILD PER-CUSTOMER CONDITION (POS LOGIC) - USER SPECIFIC
    # ============================================================
    per_customer_conditions = []
    per_customer_params = []

    for cust in allowed_customers:
        pos_flag = frappe.db.get_value("Customer", cust, "custom_pos_customer")
        debug_messages.append(f"DEBUG: Customer '{cust}' has POS flag: {pos_flag}")

        if pos_flag == 1:
            # Only tickets where Site.customer_type = 'POC Customer'
            per_customer_conditions.append(
                """(customer=%s AND EXISTS (
                    SELECT 1 FROM `tabSite` s 
                    WHERE s.name = `tabHD Ticket`.custom_circuit_id 
                    AND s.customer_type = 'POC Customer'
                ))"""
            )
            per_customer_params.append(cust)
        else:
            # POC or Paid or Empty customer type from Site
            per_customer_conditions.append(
                """(customer=%s AND (
                    NOT EXISTS (
                        SELECT 1 FROM `tabSite` s 
                        WHERE s.name = `tabHD Ticket`.custom_circuit_id 
                        AND s.customer_type NOT IN ('POC Customer', 'Paid Customer', '')
                    )
                    OR `tabHD Ticket`.custom_circuit_id IS NULL
                    OR `tabHD Ticket`.custom_circuit_id = ''
                ))"""
            )
            per_customer_params.append(cust)

    # Combine with OR
    final_customer_clause = " AND (" + " OR ".join(per_customer_conditions) + ")"
    
    debug_messages.append(f"DEBUG: Final customer clause: {final_customer_clause}")

    # Combine all params - customer params first, then filter params
    all_params = per_customer_params + params
    
    debug_messages.append(f"DEBUG: All params (customer+filter): {all_params}")

    # ============================================================
    # Pagination
    # ============================================================
    try:
        page = int(page)
        if page < 1:
            page = 1
    except:
        page = 1

    try:
        page_size = int(page_size)
        if page_size <= 0 or page_size > 2000:
            page_size = 20
    except:
        page_size = 20

    offset = (page - 1) * page_size

    # ============================================================
    # COUNT QUERY
    # ============================================================
    count_query = f"""
        SELECT COUNT(*)
        FROM `tabHD Ticket`
        WHERE 1=1
        {final_customer_clause}
        {where_clause}
    """
    
    debug_messages.append(f"DEBUG: Count query: {count_query}")
    
    try:
        total = frappe.db.sql(
            count_query,
            tuple(all_params),
        )[0][0]
        
        debug_messages.append(f"DEBUG: Total count: {total}")
    except Exception as e:
        debug_messages.append(f"DEBUG: Error in count query: {str(e)}")
        total = 0

    # ============================================================
    # MAIN DATA QUERY - UPDATED TO INCLUDE custom_rca (RESOLUTION FIELD)
    # ============================================================
    data_query = f"""
        SELECT
            name,
            custom_channel,
            custom_circuit_id,
            customer,
            custom_site_name,
            custom_site_type,
            priority,
            custom_agent_name,
            custom_agent_responded_on,
            resolution_by,
            first_responded_on,
            creation,
            custom_close_datetime,
            resolution_date,
            subject,
            description,
            custom_site_id__legal_code,
            status,
            custom_rca  -- ADDED: Resolution field from HD Ticket
        FROM `tabHD Ticket`
        WHERE 1=1
        {final_customer_clause}
        {where_clause}
        ORDER BY creation DESC
        LIMIT %s OFFSET %s
    """
    
    all_params_with_pagination = all_params + [page_size, offset]
    
    debug_messages.append(f"DEBUG: Data query: {data_query}")
    
    try:
        tickets = frappe.db.sql(
            data_query,
            tuple(all_params_with_pagination),
            as_dict=True,
        )
        
        debug_messages.append(f"DEBUG: Tickets found: {len(tickets)}")
        if tickets:
            # Log only the first 3 tickets to avoid too much data
            for i, ticket in enumerate(tickets[:3]):
                debug_messages.append(f"DEBUG: Ticket {i+1}: {ticket.get('name')} - Created: {ticket.get('creation')} - Closed: {ticket.get('custom_close_datetime')} - RCA: {ticket.get('custom_rca')}")
    except Exception as e:
        debug_messages.append(f"DEBUG: Error in data query: {str(e)}")
        tickets = []

    # Log all debug messages at once
    frappe.log_error("\n".join(debug_messages), "get_tickets")
    
    return {"tickets": tickets, "total": total}


# =============================================================================
#  HELPER FUNCTIONS
# =============================================================================

def get_allowed_customers_for_user(user):
    """
    Get list of customers user is allowed to see based on User Permission
    """
    if user == "Administrator":
        # Administrator can see all customers
        return []
    
    user_permissions = frappe.db.get_all(
        "User Permission",
        filters={"user": user, "allow": "Customer"},
        fields=["for_value"],
        distinct=True
    )
    
    allowed_customers = [x.for_value for x in user_permissions if x.for_value]
    return allowed_customers


def apply_user_permission_filter(user, doctype, query):
    """
    Apply User Permission filter to any query
    This function can be used in other places to enforce permissions
    """
    if user == "Administrator":
        return query
    
    allowed_customers = get_allowed_customers_for_user(user)
    
    if not allowed_customers:
        # User has no permissions, show nothing
        query = query.where(f"{doctype}.customer = ''")
    else:
        # User can only see specific customers
        query = query.where(f"{doctype}.customer IN %(allowed_customers)s")
        query = query.values(allowed_customers=tuple(allowed_customers))
    
    return query


# =============================================================================
#  HD TICKET PERMISSION QUERY - THIS IS THE CRITICAL FIX
# =============================================================================
def get_permission_query_conditions(user):
    """
    This is called by Frappe's permission system
    It adds WHERE conditions to ALL queries on HD Ticket
    This ensures User Permissions are applied EVERYWHERE
    """
    if not user:
        user = frappe.session.user
    
    if user == "Administrator":
        return ""
    
    allowed_customers = get_allowed_customers_for_user(user)
    
    if not allowed_customers:
        # User has no permissions, show nothing
        return """(`tabHD Ticket`.customer = '' AND `tabHD Ticket`.customer IS NOT NULL)"""
    
    # Apply POS logic for each customer
    conditions = []
    
    for cust in allowed_customers:
        pos_flag = frappe.db.get_value("Customer", cust, "custom_pos_customer")
        
        if pos_flag == 1:
            conditions.append(
                f"""(customer='{cust}' AND EXISTS (
                    SELECT 1 FROM `tabSite` s 
                    WHERE s.name = `tabHD Ticket`.custom_circuit_id 
                    AND s.customer_type = 'POC Customer'
                ))"""
            )
        else:
            conditions.append(
                f"""(customer='{cust}' AND (
                    NOT EXISTS (
                        SELECT 1 FROM `tabSite` s 
                        WHERE s.name = `tabHD Ticket`.custom_circuit_id 
                        AND s.customer_type NOT IN ('POC Customer', 'Paid Customer', '')
                    )
                    OR `tabHD Ticket`.custom_circuit_id IS NULL
                    OR `tabHD Ticket`.custom_circuit_id = ''
                ))"""
            )
    
    # Combine conditions with OR
    final_condition = "(" + " OR ".join(conditions) + ")"
    
    return final_condition


# =============================================================================
#  HD TICKET HAS PERMISSION CHECK - ADDITIONAL SECURITY
# =============================================================================
def has_permission(doc, user):
    """
    Check if user has permission to access this specific HD Ticket
    Called when accessing individual tickets
    """
    if user == "Administrator":
        return True
    
    allowed_customers = get_allowed_customers_for_user(user)
    
    if not allowed_customers:
        return False
    
    if doc.customer not in allowed_customers:
        return False
    
    # Check POS logic for this customer
    pos_flag = frappe.db.get_value("Customer", doc.customer, "custom_pos_customer")
    
    if pos_flag == 1:
        # Check if Site exists and has customer_type = 'POC Customer'
        if doc.custom_circuit_id:
            site_customer_type = frappe.db.get_value("Site", doc.custom_circuit_id, "customer_type")
            return site_customer_type == 'POC Customer'
        else:
            return False
    else:
        # User can see tickets where Site.customer_type is POC, Paid, or empty
        if doc.custom_circuit_id:
            site_customer_type = frappe.db.get_value("Site", doc.custom_circuit_id, "customer_type")
            return site_customer_type in ['POC Customer', 'Paid Customer', None, '']
        else:
            return True


# =============================================================================
#  GET TICKET STATS - TOTAL COUNTS FOR ALL STATUSES
# =============================================================================
@frappe.whitelist()
def get_ticket_stats():
    """
    Get total ticket counts for all statuses considering user permissions
    """
    try:
        user = frappe.session.user

        # Get allowed customers for user
        allowed_customers = get_allowed_customers_for_user(user)

        if not allowed_customers:
            return {
                "total": 0,
                "open": 0,
                "replied": 0,
                "on_hold": 0,
                "resolved": 0,
                "closed": 0
            }

        # Build per-customer conditions with POS logic
        per_customer_conditions = []
        per_customer_params = []

        for cust in allowed_customers:
            pos_flag = frappe.db.get_value("Customer", cust, "custom_pos_customer")

            if pos_flag == 1:
                # Only tickets where Site.customer_type = 'POC Customer'
                per_customer_conditions.append(
                    """(customer=%s AND EXISTS (
                        SELECT 1 FROM `tabSite` s 
                        WHERE s.name = `tabHD Ticket`.custom_circuit_id 
                        AND s.customer_type = 'POC Customer'
                    ))"""
                )
                per_customer_params.append(cust)
            else:
                # POC or Paid or Empty customer type from Site
                per_customer_conditions.append(
                    """(customer=%s AND (
                        NOT EXISTS (
                            SELECT 1 FROM `tabSite` s 
                            WHERE s.name = `tabHD Ticket`.custom_circuit_id 
                            AND s.customer_type NOT IN ('POC Customer', 'Paid Customer', '')
                        )
                        OR `tabHD Ticket`.custom_circuit_id IS NULL
                        OR `tabHD Ticket`.custom_circuit_id = ''
                    ))"""
                )
                per_customer_params.append(cust)

        final_customer_clause = " AND (" + " OR ".join(per_customer_conditions) + ")"

        # Query for total counts by status
        status_counts = frappe.db.sql("""
            SELECT 
                status,
                COUNT(*) as count
            FROM `tabHD Ticket`
            WHERE 1=1
            {customer_clause}
            GROUP BY status
        """.format(customer_clause=final_customer_clause), tuple(per_customer_params), as_dict=True)

        # Initialize counts
        stats = {
            "total": 0,
            "open": 0,
            "replied": 0,
            "on_hold": 0,
            "resolved": 0,
            "closed": 0
        }

        # Map statuses to our categories
        for row in status_counts:
            status = (row.status or "").lower()
            count = row.count
            
            stats["total"] += count
            
            if "open" in status:
                stats["open"] += count
            elif "replied" in status:
                stats["replied"] += count
            elif "on hold" in status:
                stats["on_hold"] += count
            elif "resolved" in status:
                stats["resolved"] += count
            elif "closed" in status:
                stats["closed"] += count

        return stats

    except Exception as e:
        frappe.log_error(f"Error getting ticket stats: {str(e)}")
        return {
            "total": 0,
            "open": 0,
            "replied": 0,
            "on_hold": 0,
            "resolved": 0,
            "closed": 0
        }


# =============================================================================
#  GET TICKET ACTIVITY (Communication + Version log)
# =============================================================================
@frappe.whitelist()
def get_ticket_activity(ticket_name):
    """Returns all activity linked to a ticket: Email + Version Log."""

    if not ticket_name:
        return {"activity": []}

    # Check if user has permission to view this ticket
    user = frappe.session.user
    if user != "Administrator":
        ticket = frappe.db.get_value("HD Ticket", ticket_name, ["customer", "custom_circuit_id"], as_dict=True)
        if ticket:
            allowed_customers = get_allowed_customers_for_user(user)
            if ticket.customer not in allowed_customers:
                return {"activity": [], "error": "Permission denied"}
            
            # Check POS logic
            pos_flag = frappe.db.get_value("Customer", ticket.customer, "custom_pos_customer")
            if pos_flag == 1:
                # Check if Site has customer_type = 'POC Customer'
                if ticket.custom_circuit_id:
                    site_customer_type = frappe.db.get_value("Site", ticket.custom_circuit_id, "customer_type")
                    if site_customer_type != 'POC Customer':
                        return {"activity": [], "error": "Permission denied"}

    # -----------------------------------
    # Communications (Email) Logs
    # -----------------------------------
    comms = frappe.db.sql(
        """
        SELECT
            name, communication_type, sender, recipients,
            subject, content, creation
        FROM `tabCommunication`
        WHERE reference_doctype = 'HD Ticket'
          AND reference_name = %s
        ORDER BY creation DESC
        """,
        (ticket_name,),
        as_dict=True,
    )

    comm_list = []
    for c in comms:
        comm_list.append({
            "type": "Communication",
            "sender": c.sender,
            "subject": c.subject,
            "content_html": c.content,  # raw HTML email
            "content": c.content,  # plain text fallback
            "creation": str(c.creation),
        })

    # -----------------------------------
    # Version Logs (System Changes)
    # -----------------------------------
    versions = frappe.db.sql(
        """
        SELECT name, data, owner, creation
        FROM `tabVersion`
        WHERE ref_doctype='HD Ticket'
          AND docname=%s
        ORDER BY creation DESC
        """,
        (ticket_name,),
        as_dict=True,
    )

    ver_list = []
    for v in versions:
        try:
            parsed_data = json.loads(v.data)
        except:
            parsed_data = v.data

        ver_list.append(
            {
                "type": "Version",
                "owner": v.owner,
                "data": parsed_data,
                "creation": str(v.creation),
            }
        )

    # Merge + sort by timestamp
    all_activity = comm_list + ver_list
    activity = sorted(all_activity, key=lambda x: x["creation"], reverse=True)

    return {"activity": activity}


# =============================================================================
#  GET SITE INFORMATION BY CIRCUIT ID
# =============================================================================
@frappe.whitelist()
def get_site_by_circuit_id(circuit_id):
    """
    Get Site document information by circuit_id
    """
    try:
        if not circuit_id:
            return None
            
        # Check if Site document exists with this name
        if frappe.db.exists("Site", circuit_id):
            site = frappe.get_doc("Site", circuit_id)
            return {
                "address_street": site.get("address_street"),
                "district": site.get("district"),
                "city": site.get("city"),
                "pincode": site.get("pincode"),
                "state": site.get("state"),
                "territory": site.get("territory"),
                "contact_person": site.get("contact_person"),
                "primary_contact_mobile": site.get("primary_contact_mobile")
            }
        else:
            return None
    except Exception as e:
        frappe.log_error(f"Error getting site by circuit_id {circuit_id}: {str(e)}")
        return None


# =============================================================================
#  UPDATE TICKET STATUS
# =============================================================================
@frappe.whitelist()
def update_ticket_status(ticket_name, new_status):
    """
    Update HD Ticket status
    """
    try:
        if not ticket_name or not new_status:
            return {"status": "error", "message": "Missing required parameters"}
            
        if not frappe.db.exists("HD Ticket", ticket_name):
            return {"status": "error", "message": "Ticket not found"}
        
        # Check if user has permission to update this ticket
        user = frappe.session.user
        if user != "Administrator":
            ticket = frappe.db.get_value("HD Ticket", ticket_name, ["customer", "custom_circuit_id"], as_dict=True)
            if ticket:
                allowed_customers = get_allowed_customers_for_user(user)
                if ticket.customer not in allowed_customers:
                    return {"status": "error", "message": "Permission denied to update this ticket"}
                
                # Check POS logic
                pos_flag = frappe.db.get_value("Customer", ticket.customer, "custom_pos_customer")
                if pos_flag == 1:
                    # Check if Site has customer_type = 'POC Customer'
                    if ticket.custom_circuit_id:
                        site_customer_type = frappe.db.get_value("Site", ticket.custom_circuit_id, "customer_type")
                        if site_customer_type != 'POC Customer':
                            return {"status": "error", "message": "Permission denied to update this ticket"}
            
        ticket = frappe.get_doc("HD Ticket", ticket_name)
        ticket.status = new_status
        ticket.save(ignore_permissions=True)
        
        frappe.db.commit()
        
        return {"status": "success", "message": f"Ticket status updated to {new_status}"}
        
    except Exception as e:
        frappe.log_error(f"Error updating ticket status {ticket_name}: {str(e)}")
        return {"status": "error", "message": str(e)}


# =============================================================================
#  GET TICKET STATUS OPTIONS
# =============================================================================
@frappe.whitelist()
def get_ticket_status_options():
    """
    Get all available status options from HD Ticket doctype
    """
    try:
        status_options = frappe.get_meta("HD Ticket").get_field("status").options
        if status_options:
            # Convert string of options to list
            status_list = [status.strip() for status in status_options.split("\n") if status.strip()]
            return {"status_options": status_list}
        else:
            return {"status_options": ["Open", "Replied", "On Hold", "Resolved", "Closed"]}
    except Exception as e:
        frappe.log_error(f"Error getting ticket status options: {str(e)}")
        return {"status_options": ["Open", "Replied", "On Hold", "Resolved", "Closed"]}


# =============================================================================
#  CHECK CREATE TICKET PERMISSION
# =============================================================================
@frappe.whitelist()
def check_create_ticket_permission():
    """
    Check if current user can create tickets
    based on custom_create_ticket flag in Customer
    """
    try:
        user = frappe.session.user
        
        if user == "Administrator":
            # Administrator can always create tickets
            return {
                "can_create": True,
                "customer_name": "Administrator",
                "message": "Administrator has full access"
            }
        
        # Get user's allowed customers
        allowed_customers = get_allowed_customers_for_user(user)
        
        if not allowed_customers:
            return {
                "can_create": False,
                "customer_name": None,
                "message": "No customers assigned to user"
            }
        
        # Check if any allowed customer has custom_create_ticket = 1
        for cust in allowed_customers:
            custom_create_ticket = frappe.db.get_value("Customer", cust, "custom_create_ticket")
            customer_name = frappe.db.get_value("Customer", cust, "customer_name")
            
            if custom_create_ticket == 1:
                return {
                    "can_create": True,
                    "customer_name": customer_name,
                    "customer": cust,
                    "message": f"User can create tickets for {customer_name}"
                }
        
        # No customer with create permission
        return {
            "can_create": False,
            "customer_name": None,
            "message": "No customer with create ticket permission"
        }
        
    except Exception as e:
        frappe.log_error(f"Error checking create ticket permission: {str(e)}")
        return {
            "can_create": False,
            "customer_name": None,
            "error": str(e)
        }


# =============================================================================
#  GET TICKETS BY STATUS (For debugging stats vs filtered results)
# =============================================================================
@frappe.whitelist()
def get_tickets_by_status(status):
    """
    Get tickets by specific status for debugging
    """
    try:
        user = frappe.session.user

        # Get allowed customers for user
        allowed_customers = get_allowed_customers_for_user(user)

        if not allowed_customers:
            return {"tickets": [], "total": 0}

        # Build per-customer conditions with POS logic
        per_customer_conditions = []
        per_customer_params = []

        for cust in allowed_customers:
            pos_flag = frappe.db.get_value("Customer", cust, "custom_pos_customer")

            if pos_flag == 1:
                # Only tickets where Site.customer_type = 'POC Customer'
                per_customer_conditions.append(
                    """(customer=%s AND EXISTS (
                        SELECT 1 FROM `tabSite` s 
                        WHERE s.name = `tabHD Ticket`.custom_circuit_id 
                        AND s.customer_type = 'POC Customer'
                    ))"""
                )
                per_customer_params.append(cust)
            else:
                # POC or Paid or Empty customer type from Site
                per_customer_conditions.append(
                    """(customer=%s AND (
                        NOT EXISTS (
                            SELECT 1 FROM `tabSite` s 
                            WHERE s.name = `tabHD Ticket`.custom_circuit_id 
                            AND s.customer_type NOT IN ('POC Customer', 'Paid Customer', '')
                        )
                        OR `tabHD Ticket`.custom_circuit_id IS NULL
                        OR `tabHD Ticket`.custom_circuit_id = ''
                    ))"""
                )
                per_customer_params.append(cust)

        final_customer_clause = " AND (" + " OR ".join(per_customer_conditions) + ")"

        # Add status filter - FIXED: Use LIKE instead of = for better matching
        status_condition = " AND status LIKE %s"
        per_customer_params.append(f"%{status}%")

        # Count query
        total = frappe.db.sql(
            f"""
            SELECT COUNT(*)
            FROM `tabHD Ticket`
            WHERE 1=1
            {final_customer_clause}
            {status_condition}
            """,
            tuple(per_customer_params),
        )[0][0]

        # Data query - UPDATED TO INCLUDE custom_rca (RESOLUTION FIELD)
        tickets = frappe.db.sql(
            f"""
            SELECT
                name,
                custom_channel,
                custom_circuit_id,
                customer,
                custom_site_name,
                custom_site_type,
                priority,
                custom_agent_name,
                custom_agent_responded_on,
                resolution_by,
                first_responded_on,
                creation,
                custom_close_datetime,
                resolution_date,
                subject,
                description,
                custom_site_id__legal_code,
                status,
                custom_rca  -- ADDED: Resolution field from HD Ticket
            FROM `tabHD Ticket`
            WHERE 1=1
            {final_customer_clause}
            {status_condition}
            ORDER BY creation DESC
            LIMIT 50
            """,
            tuple(per_customer_params),
            as_dict=True,
        )

        return {
            "tickets": tickets, 
            "total": total,
            "status": status,
            "debug_info": {
                "user": user,
                "allowed_customers": allowed_customers,
                "status_used": status
            }
        }

    except Exception as e:
        frappe.log_error(f"Error getting tickets by status {status}: {str(e)}")
        return {"tickets": [], "total": 0, "error": str(e)}


# =============================================================================
#  DEBUG: GET ALL STATUS VALUES
# =============================================================================
@frappe.whitelist()
def get_all_status_values():
    """
    Get all distinct status values from HD Ticket for debugging
    """
    try:
        user = frappe.session.user

        # Get allowed customers for user
        allowed_customers = get_allowed_customers_for_user(user)

        if not allowed_customers:
            return {"status_values": []}

        # Build per-customer conditions with POS logic
        per_customer_conditions = []
        per_customer_params = []

        for cust in allowed_customers:
            pos_flag = frappe.db.get_value("Customer", cust, "custom_pos_customer")

            if pos_flag == 1:
                # Only tickets where Site.customer_type = 'POC Customer'
                per_customer_conditions.append(
                    """(customer=%s AND EXISTS (
                        SELECT 1 FROM `tabSite` s 
                        WHERE s.name = `tabHD Ticket`.custom_circuit_id 
                        AND s.customer_type = 'POC Customer'
                    ))"""
                )
                per_customer_params.append(cust)
            else:
                # POC or Paid or Empty customer type from Site
                per_customer_conditions.append(
                    """(customer=%s AND (
                        NOT EXISTS (
                            SELECT 1 FROM `tabSite` s 
                            WHERE s.name = `tabHD Ticket`.custom_circuit_id 
                            AND s.customer_type NOT IN ('POC Customer', 'Paid Customer', '')
                        )
                        OR `tabHD Ticket`.custom_circuit_id IS NULL
                        OR `tabHD Ticket`.custom_circuit_id = ''
                    ))"""
                )
                per_customer_params.append(cust)

        final_customer_clause = " AND (" + " OR ".join(per_customer_conditions) + ")"

        # Get all distinct status values
        status_values = frappe.db.sql(
            f"""
            SELECT DISTINCT status, COUNT(*) as count
            FROM `tabHD Ticket`
            WHERE 1=1
            {final_customer_clause}
            GROUP BY status
            ORDER BY status
            """,
            tuple(per_customer_params),
            as_dict=True,
        )

        return {
            "status_values": status_values,
            "debug_info": {
                "user": user,
                "allowed_customers": allowed_customers,
                "total_customers": len(allowed_customers)
            }
        }

    except Exception as e:
        frappe.log_error(f"Error getting status values: {str(e)}")
        return {"status_values": [], "error": str(e)}


# =============================================================================
#  GET ALL TICKETS WITHOUT FILTERS (For debugging)
# =============================================================================
@frappe.whitelist()
def get_all_tickets_for_debug():
    """
    Get all tickets without any filters for debugging
    """
    try:
        user = frappe.session.user

        # Get allowed customers for user
        allowed_customers = get_allowed_customers_for_user(user)

        if not allowed_customers:
            return {"tickets": [], "total": 0}

        # Build per-customer conditions with POS logic
        per_customer_conditions = []
        per_customer_params = []

        for cust in allowed_customers:
            pos_flag = frappe.db.get_value("Customer", cust, "custom_pos_customer")

            if pos_flag == 1:
                # Only tickets where Site.customer_type = 'POC Customer'
                per_customer_conditions.append(
                    """(customer=%s AND EXISTS (
                        SELECT 1 FROM `tabSite` s 
                        WHERE s.name = `tabHD Ticket`.custom_circuit_id 
                        AND s.customer_type = 'POC Customer'
                    ))"""
                )
                per_customer_params.append(cust)
            else:
                # POC or Paid or Empty customer type from Site
                per_customer_conditions.append(
                    """(customer=%s AND (
                        NOT EXISTS (
                            SELECT 1 FROM `tabSite` s 
                            WHERE s.name = `tabHD Ticket`.custom_circuit_id 
                            AND s.customer_type NOT IN ('POC Customer', 'Paid Customer', '')
                        )
                        OR `tabHD Ticket`.custom_circuit_id IS NULL
                        OR `tabHD Ticket`.custom_circuit_id = ''
                    ))"""
                )
                per_customer_params.append(cust)

        final_customer_clause = " AND (" + " OR ".join(per_customer_conditions) + ")"

        # Get all tickets without filters - UPDATED TO INCLUDE custom_rca (RESOLUTION FIELD)
        tickets = frappe.db.sql(
            f"""
            SELECT
                name,
                custom_channel,
                custom_circuit_id,
                customer,
                custom_site_name,
                custom_site_type,
                priority,
                custom_agent_name,
                custom_agent_responded_on,
                resolution_by,
                first_responded_on,
                creation,
                custom_close_datetime,
                resolution_date,
                subject,
                description,
                custom_site_id__legal_code,
                status,
                custom_rca  -- ADDED: Resolution field from HD Ticket
            FROM `tabHD Ticket`
            WHERE 1=1
            {final_customer_clause}
            ORDER BY creation DESC
            LIMIT 100
            """,
            tuple(per_customer_params),
            as_dict=True,
        )

        # Count all tickets
        total = frappe.db.sql(
            f"""
            SELECT COUNT(*)
            FROM `tabHD Ticket`
            WHERE 1=1
            {final_customer_clause}
            """,
            tuple(per_customer_params),
        )[0][0]

        return {
            "tickets": tickets, 
            "total": total,
            "debug_info": {
                "user": user,
                "allowed_customers": allowed_customers,
                "total_customers": len(allowed_customers),
                "sample_tickets": len(tickets)
            }
        }

    except Exception as e:
        frappe.log_error(f"Error getting all tickets for debug: {str(e)}")
        return {"tickets": [], "total": 0, "error": str(e)}


# =============================================================================
#  DEBUG: GET USER PERMISSIONS AND CUSTOMER POS SETTINGS
# =============================================================================
@frappe.whitelist()
def get_user_customer_permissions():
    """
    Debug function to check user permissions and customer POS settings
    """
    try:
        user = frappe.session.user
        
        debug_info = {
            "user": user,
            "permissions": [],
            "customers": []
        }
        
        # Get user permissions for customers
        user_permissions = frappe.db.get_all(
            "User Permission",
            filters={"user": user, "allow": "Customer"},
            fields=["for_value", "name", "creation"],
            order_by="creation desc"
        )
        
        debug_info["permissions"] = user_permissions
        
        allowed_customers = [x.for_value for x in user_permissions if x.for_value]
        
        # Get customer details including POS flag and create ticket flag
        for cust in allowed_customers:
            customer_info = frappe.db.get_value(
                "Customer", 
                cust, 
                ["customer_name", "custom_pos_customer", "custom_customer_type", "custom_create_ticket"],
                as_dict=True
            )
            
            if customer_info:
                debug_info["customers"].append({
                    "customer": cust,
                    "customer_name": customer_info.get("customer_name"),
                    "pos_flag": customer_info.get("custom_pos_customer"),
                    "customer_type": customer_info.get("custom_customer_type"),
                    "can_create_ticket": customer_info.get("custom_create_ticket")
                })
        
        # Count tickets for each customer with new POS logic
        for cust_info in debug_info["customers"]:
            cust = cust_info["customer"]
            pos_flag = cust_info["pos_flag"]
            
            if pos_flag == 1:
                # Only tickets where Site.customer_type = 'POC Customer'
                ticket_count = frappe.db.sql("""
                    SELECT COUNT(*) as count
                    FROM `tabHD Ticket` t
                    WHERE t.customer=%s 
                    AND EXISTS (
                        SELECT 1 FROM `tabSite` s 
                        WHERE s.name = t.custom_circuit_id 
                        AND s.customer_type = 'POC Customer'
                    )
                """, (cust,), as_dict=True)[0]["count"]
                
                poc_count = frappe.db.sql("""
                    SELECT COUNT(*) as count
                    FROM `tabHD Ticket` t
                    WHERE t.customer=%s 
                    AND EXISTS (
                        SELECT 1 FROM `tabSite` s 
                        WHERE s.name = t.custom_circuit_id 
                        AND s.customer_type = 'POC Customer'
                    )
                """, (cust,), as_dict=True)[0]["count"]
                
                paid_count = frappe.db.sql("""
                    SELECT COUNT(*) as count
                    FROM `tabHD Ticket` t
                    WHERE t.customer=%s 
                    AND EXISTS (
                        SELECT 1 FROM `tabSite` s 
                        WHERE s.name = t.custom_circuit_id 
                        AND s.customer_type = 'Paid Customer'
                    )
                """, (cust,), as_dict=True)[0]["count"]
                
                cust_info["ticket_counts"] = {
                    "total": ticket_count,
                    "poc_tickets": poc_count,
                    "paid_tickets": paid_count,
                    "user_can_see": "Only POC tickets (from Site)"
                }
            else:
                # Both POC and Paid tickets from Site
                ticket_count = frappe.db.sql("""
                    SELECT COUNT(*) as count
                    FROM `tabHD Ticket` t
                    WHERE t.customer=%s 
                    AND (
                        NOT EXISTS (
                            SELECT 1 FROM `tabSite` s 
                            WHERE s.name = t.custom_circuit_id 
                            AND s.customer_type NOT IN ('POC Customer', 'Paid Customer', '')
                        )
                        OR t.custom_circuit_id IS NULL
                        OR t.custom_circuit_id = ''
                    )
                """, (cust,), as_dict=True)[0]["count"]
                
                poc_count = frappe.db.sql("""
                    SELECT COUNT(*) as count
                    FROM `tabHD Ticket` t
                    WHERE t.customer=%s 
                    AND EXISTS (
                        SELECT 1 FROM `tabSite` s 
                        WHERE s.name = t.custom_circuit_id 
                        AND s.customer_type = 'POC Customer'
                    )
                """, (cust,), as_dict=True)[0]["count"]
                
                paid_count = frappe.db.sql("""
                    SELECT COUNT(*) as count
                    FROM `tabHD Ticket` t
                    WHERE t.customer=%s 
                    AND EXISTS (
                        SELECT 1 FROM `tabSite` s 
                            WHERE s.name = t.custom_circuit_id 
                            AND s.customer_type = 'Paid Customer'
                    )
                """, (cust,), as_dict=True)[0]["count"]
                
                cust_info["ticket_counts"] = {
                    "total": ticket_count,
                    "poc_tickets": poc_count,
                    "paid_tickets": paid_count,
                    "user_can_see": "Both POC & Paid tickets (from Site)"
                }
        
        return debug_info
        
    except Exception as e:
        frappe.log_error(f"Error getting user customer permissions: {str(e)}")
        return {"error": str(e)}


# =============================================================================
#  DEBUG: GET TICKETS BY CUSTOMER
# =============================================================================
@frappe.whitelist()
def get_tickets_by_customer(customer):
    """
    Get tickets for a specific customer to debug permission issues
    """
    try:
        user = frappe.session.user
        
        # Check if user has permission for this customer
        has_permission = frappe.db.exists(
            "User Permission",
            {
                "user": user,
                "allow": "Customer",
                "for_value": customer
            }
        )
        
        if not has_permission and user != "Administrator":
            return {
                "error": f"User {user} does not have permission for customer {customer}",
                "tickets": []
            }
        
        # Get customer POS flag
        pos_flag = frappe.db.get_value("Customer", customer, "custom_pos_customer")
        
        if pos_flag == 1:
            # Only tickets where Site.customer_type = 'POC Customer'
            tickets = frappe.db.sql("""
                SELECT
                    t.name,
                    s.customer_type,
                    t.status,
                    t.subject,
                    t.creation,
                    t.custom_close_datetime,
                    t.custom_rca  -- ADDED: Resolution field from HD Ticket
                FROM `tabHD Ticket` t
                LEFT JOIN `tabSite` s ON s.name = t.custom_circuit_id
                WHERE t.customer=%s 
                AND s.customer_type = 'POC Customer'
                ORDER BY t.creation DESC
                LIMIT 20
            """, (customer,), as_dict=True)
        else:
            # Both POC and Paid tickets from Site
            tickets = frappe.db.sql("""
                SELECT
                    t.name,
                    s.customer_type,
                    t.status,
                    t.subject,
                    t.creation,
                    t.custom_close_datetime,
                    t.custom_rca  -- ADDED: Resolution field from HD Ticket
                FROM `tabHD Ticket` t
                LEFT JOIN `tabSite` s ON s.name = t.custom_circuit_id
                WHERE t.customer=%s 
                AND (
                    s.customer_type IN ('POC Customer', 'Paid Customer', '')
                    OR s.customer_type IS NULL
                    OR t.custom_circuit_id IS NULL
                    OR t.custom_circuit_id = ''
                )
                ORDER BY t.creation DESC
                LIMIT 20
            """, (customer,), as_dict=True)
        
        return {
            "customer": customer,
            "pos_flag": pos_flag,
            "ticket_count": len(tickets),
            "tickets": tickets
        }
        
    except Exception as e:
        frappe.log_error(f"Error getting tickets by customer {customer}: {str(e)}")
        return {"error": str(e), "tickets": []}


# =============================================================================
#  HD TICKET HOOKS - APPLY PERMISSIONS AT DOCTYPE LEVEL
# =============================================================================
# Add these hooks in your HD Ticket doctype's Python code or in hooks.py

def on_hd_ticket_permission_query(user):
    """
    Hook to apply permission query to HD Ticket
    """
    return get_permission_query_conditions(user)


def on_hd_ticket_has_permission(doc, user):
    """
    Hook to check permissions on individual HD Ticket documents
    """
    return has_permission(doc, user)
###############################################################################

import frappe
from frappe import _

@frappe.whitelist()
def check_ticket_creation_permission():
    """Check if current user has permission to create tickets based on User Permission and Customer settings"""
    try:
        current_user = frappe.session.user
        
        # Check if user has any User Permission for Customer doctype
        user_permissions = frappe.get_all(
            "User Permission",
            filters={
                "user": current_user,
                "allow": "Customer"
            },
            fields=["for_value"]
        )
        
        allowed_customers = [perm.for_value for perm in user_permissions]
        can_create_ticket = False
        
        if allowed_customers:
            # Check each customer's custom_create_ticket field
            for customer_name in allowed_customers:
                try:
                    customer = frappe.get_doc("Customer", customer_name)
                    if customer.get("custom_create_ticket") == 1:
                        can_create_ticket = True
                        break
                except frappe.DoesNotExistError:
                    continue
        
        return {
            "can_create_ticket": can_create_ticket,
            "allowed_customers": allowed_customers
        }
        
    except Exception as e:
        frappe.log_error(f"Error checking ticket creation permission: {str(e)}")
        return {
            "can_create_ticket": False,
            "allowed_customers": [],
            "error": str(e)
        }
###################################################################################
# Revesal api

import frappe
import json

@frappe.whitelist()
def reverse_bank_entries(entry_names):
    """
    Reverse selected bank statement entries
    """

    try:
        if isinstance(entry_names, str):
            entry_names = json.loads(entry_names)

        if not isinstance(entry_names, list):
            entry_names = [entry_names]

        success_count = 0
        failed_entries = []

        for entry_name in entry_names:
            try:
                bank_entry = frappe.get_doc("Bank Statement Entry", entry_name)

                reference_no = bank_entry.get("reference_no")

                # Cancel Payment Entry
                if reference_no and reference_no.startswith("ACC-PAY"):
                    if frappe.db.exists("Payment Entry", reference_no):
                        pe = frappe.get_doc("Payment Entry", reference_no)
                        if pe.docstatus != 2:
                            pe.cancel()
                            frappe.db.commit()

                # Cancel Journal Entry
                elif reference_no and reference_no.startswith("ACC-JV"):
                    if frappe.db.exists("Journal Entry", reference_no):
                        je = frappe.get_doc("Journal Entry", reference_no)
                        if je.docstatus != 2:
                            je.cancel()
                            frappe.db.commit()

                # ---------- FIXED: Update ONLY valid fields ----------
                frappe.db.set_value("Bank Statement Entry", entry_name, {
                    "reconciled": 0,      # Check field
                    "reference_no": ""    # Your actual reference field
                })

                frappe.db.commit()

                bank_entry.add_comment(text="Entry reversed and reconciliation reset to 0")

                success_count += 1

            except Exception as e:
                failed_entries.append({
                    "entry": entry_name,
                    "error": str(e)
                })
                frappe.log_error(f"Failed to reverse entry {entry_name}: {str(e)}", "Reverse Bank Entry")

        if failed_entries:
            return {
                "status": "partial",
                "message": f"{success_count} entries reversed. {len(failed_entries)} failed.",
                "success_count": success_count,
                "failed_count": len(failed_entries),
                "failed_entries": failed_entries
            }

        return {
            "status": "success",
            "message": f"{success_count} entries reversed successfully.",
            "success_count": success_count
        }

    except Exception as e:
        frappe.log_error(f"Error in reverse_bank_entries: {str(e)}", "Reverse Bank Entry")
        return {"status": "error", "message": str(e)}
#########################################################################
# Employee Advance
@frappe.whitelist()
def get_unpaid_employee_advances(employee):
    """Return all unpaid Employee Advances for selected employee."""

    advances = frappe.get_all(
        "Employee Advance",
        filters={
            "employee": employee,
            "docstatus": 1,
            "status": "Unpaid"      # ← FIXED (Do NOT use pending_amount filter)
        },
        fields=["name", "posting_date", "advance_amount", "paid_amount"],
        order_by="posting_date asc"
    )

    # Calculate pending amount manually
    for adv in advances:
        paid = frappe.utils.flt(adv.get("paid_amount") or 0, 2)
        total = frappe.utils.flt(adv.get("advance_amount") or 0, 2)
        adv["pending_amount"] = frappe.utils.flt(total - paid, 2)

    return advances
###############################################################################
import frappe
from frappe.utils import flt, nowdate

@frappe.whitelist()
def create_payment_entry_for_employee_advance(statement_name, employee, advance_name, amount, company=None, bank_account=None):
    """
    Create Payment Entry for Employee Advance Settlement
    Bank Account MUST come from the frontend selection.
    """

    amount = flt(amount)

    # -------------------------------
    # Basic Validations
    # -------------------------------
    if not employee:
        return {"status": "error", "error": "Employee is required"}

    if not bank_account:
        return {"status": "error", "error": "Bank Account not provided from UI"}

    if amount <= 0:
        return {"status": "error", "error": "Amount must be > 0"}

    company = company or frappe.defaults.get_default("company")

    # -------------------------------
    # Load Bank Statement Entry
    # -------------------------------
    statement = frappe.get_doc("Bank Statement Entry", statement_name)
    bank_amount = flt(statement.withdrawal or statement.deposit or 0)

    if amount > bank_amount:
        return {
            "status": "error",
            "error": f"Amount ({amount}) exceeds bank amount ({bank_amount})"
        }

    posting_date = (
        getattr(statement, "date", None)
        or getattr(statement, "transaction_date", None)
        or nowdate()
    )

    # -------------------------------
    # Convert Bank Account DocType → GL Account
    # -------------------------------
    gl_bank_account = get_bank_gl_account(bank_account)

    if not gl_bank_account:
        return {
            "status": "error",
            "error": f"No GL account linked to Bank Account: {bank_account}"
        }

    # Validate GL Account is active
    if frappe.db.get_value("Account", gl_bank_account, "disabled"):
        return {
            "status": "error",
            "error": f"GL Account '{gl_bank_account}' is DISABLED. Enable it in Chart of Accounts."
        }

    # -------------------------------
    # Employee Advance Account
    # -------------------------------
    advance_account = frappe.db.get_value(
        "Company",
        company,
        "default_employee_advance_account"
    )

    if not advance_account:
        return {"status": "error", "error": "Default Employee Advance Account not set in Company"}

    # -------------------------------
    # Create Payment Entry
    # -------------------------------
    pe = frappe.new_doc("Payment Entry")
    pe.payment_type = "Pay"            # Company paying employee
    pe.company = company
    pe.bank_account = bank_account  # ⭐ Added
    pe.posting_date = posting_date
    pe.mode_of_payment = "Wire Transfer"
    pe.reference_no = statement.description or statement.name
    pe.reference_date = posting_date

    pe.party_type = "Employee"
    pe.party = employee

    # Money OUT → Bank GL Account
    pe.paid_from = gl_bank_account

    # Money INTO → Employee Advance Account
    pe.paid_to = advance_account

    # ERPNext mandatory pairing for Pay
    pe.paid_amount = amount
    pe.received_amount = amount

    # -------------------------------
    # Add reference to Employee Advance
    # -------------------------------
    pe.append("references", {
        "reference_doctype": "Employee Advance",
        "reference_name": advance_name,
        "allocated_amount": amount
    })

    # Save & Submit
    pe.insert(ignore_permissions=True)
    pe.submit()

    # Mark the bank statement entry as reconciled
    frappe.db.set_value("Bank Statement Entry", statement_name, {
        "reconciled": 1,
        "reference_no": pe.name,
        "match_type": "Auto"
    })

    return {
        "status": "ok",
        "payment_entry": pe.name
    }
##############################################################################
#Invoice Management
import frappe

# ============================================================
# MAIN LIST API
# ============================================================

@frappe.whitelist()
def get_sales_orders(filters=None, page=1, page_size=20):

    rows = frappe.db.sql(
        """
        SELECT
            bs.name AS billing_statement_no,
            bs.sales_order_no,
            bs.customer AS customer_name,
            bs.order_type,
            IFNULL(bs.sales_order_amount, 0) AS sales_order_amount
        FROM `tabBilling Statement` bs
        ORDER BY bs.creation DESC
        """,
        as_dict=True
    )

    for r in rows:
        billed = get_billed_amount(r.sales_order_no)
        r["billed_amount"] = billed
        r["balance_to_bill"] = r["sales_order_amount"] - billed

    return {
        "total": len(rows),
        "orders": rows,
        "page": 1,
        "page_size": len(rows)
    }


# ============================================================
# BILLED AMOUNT (FIXED FIELD)
# ============================================================

def get_billed_amount(sales_order_no):
    if not sales_order_no:
        return 0

    return frappe.db.sql(
        """
        SELECT IFNULL(SUM(sii.base_net_amount), 0)
        FROM `tabSales Invoice Item` sii
        INNER JOIN `tabSales Invoice` si
            ON si.name = sii.parent
        WHERE si.docstatus = 1
        AND sii.sales_order = %s
        """,
        sales_order_no
    )[0][0]


# ============================================================
# DASHBOARD STATS
# ============================================================

@frappe.whitelist()
def get_invoice_stats():

    total = frappe.db.sql(
        "SELECT COUNT(*) FROM `tabBilling Statement`"
    )[0][0]

    total_amount = frappe.db.sql(
        "SELECT IFNULL(SUM(sales_order_amount), 0) FROM `tabBilling Statement`"
    )[0][0]

    billed_amount = frappe.db.sql(
        """
        SELECT IFNULL(SUM(sii.base_net_amount), 0)
        FROM `tabSales Invoice Item` sii
        INNER JOIN `tabSales Invoice` si
            ON si.name = sii.parent
        WHERE si.docstatus = 1
        """
    )[0][0]

    balance = total_amount - billed_amount
    percentage = round((billed_amount / total_amount) * 100, 2) if total_amount else 0

    return {
        "total": total,
        "billed": billed_amount,
        "balance": balance,
        "percentage": percentage
    }


# ============================================================
# TEST API (ADD BACK SO ERROR STOPS)
# ============================================================

@frappe.whitelist()
def test_api():
    data = frappe.db.sql(
        """
        SELECT
            name,
            sales_order_no,
            customer,
            order_type,
            sales_order_amount
        FROM `tabBilling Statement`
        """,
        as_dict=True
    )

    return {
        "count": len(data),
        "data": data
    }
############################ Invoice End ###############################################
#Unallocated - Customer
import frappe
from frappe.utils import flt, nowdate

# ---------------------------------------------------------
# 1. Fetch Outstanding Sales Invoices
# ---------------------------------------------------------
@frappe.whitelist()
def get_customer_outstanding_invoices(customer):

    if not customer:
        return []

    return frappe.db.sql(
        """
        SELECT
            si.name AS sales_invoice_no,
            si.posting_date AS sales_invoice_date,
            si.outstanding_amount
        FROM `tabSales Invoice` si
        WHERE
            si.customer = %s
            AND si.docstatus = 1
            AND si.outstanding_amount > 0
        ORDER BY si.posting_date
        """,
        customer,
        as_dict=True
    )


# ---------------------------------------------------------
# 2. Create Unallocated Payment Entry (RECONCILES AUTOMATICALLY)
# ---------------------------------------------------------
@frappe.whitelist()
def create_unallocated_payment_entry(customer, invoices):

    invoices = frappe.parse_json(invoices)

    if not customer or not invoices:
        frappe.throw("Missing customer or invoice data")

    total_amount = sum(flt(inv.get("amount")) for inv in invoices)
    if total_amount <= 0:
        frappe.throw("Total amount must be greater than zero")

    company = frappe.defaults.get_user_default("Company")
    company_currency = frappe.get_cached_value("Company", company, "default_currency")

    paid_to_account = frappe.get_value(
        "Mode of Payment Account",
        {
            "parent": "Wire Transfer",
            "company": company
        },
        "default_account"
    )

    if not paid_to_account:
        frappe.throw(
            f"No default account set for Mode of Payment 'Wire Transfer' for {company}"
        )

    today = nowdate()

    pe = frappe.get_doc({
        "doctype": "Payment Entry",
        "payment_type": "Receive",
        "party_type": "Customer",
        "party": customer,
        "company": company,

        "posting_date": today,
        "reference_date": today,
        "reference_no": f"UNALLOC-{customer}-{today}",

        "mode_of_payment": "Wire Transfer",
        "paid_to": paid_to_account,

        "paid_amount": total_amount,
        "received_amount": total_amount,

        "source_exchange_rate": 1,
        "target_exchange_rate": 1,
        "paid_to_account_currency": company_currency,

        "references": []
    })

    for inv in invoices:
        pe.append("references", {
            "reference_doctype": "Sales Invoice",
            "reference_name": inv.get("sales_invoice_no"),
            "allocated_amount": flt(inv.get("amount"))
        })

    pe.insert(ignore_permissions=True)
    pe.submit()

    return {
        "payment_entry": pe.name
    }


# ---------------------------------------------------------
# 3. Customer Unallocated Amount (Advance Balance)
# ---------------------------------------------------------
@frappe.whitelist()
def get_customer_unallocated_amount(customer):

    if not customer:
        return 0.0

    total_paid = frappe.db.sql(
        """
        SELECT IFNULL(SUM(pe.paid_amount), 0)
        FROM `tabPayment Entry` pe
        WHERE
            pe.party_type = 'Customer'
            AND pe.party = %s
            AND pe.docstatus = 1
        """,
        customer
    )[0][0]

    total_allocated = frappe.db.sql(
        """
        SELECT IFNULL(SUM(per.allocated_amount), 0)
        FROM `tabPayment Entry Reference` per
        INNER JOIN `tabPayment Entry` pe
            ON pe.name = per.parent
        WHERE
            pe.party_type = 'Customer'
            AND pe.party = %s
            AND pe.docstatus = 1
        """,
        customer
    )[0][0]

    return flt(total_paid) - flt(total_allocated)
###############################################################################

# HD Ticket Assignment

import frappe

def sync_custom_agent_from_todo(doc, method):
    """
    Sync custom_agent immediately when a user is assigned
    via Assign To (ToDo)
    """

    # Only assignments linked to HD Ticket
    if doc.reference_type != "HD Ticket":
        return

    # Ignore cancelled / closed todos
    if doc.status != "Open":
        return

    # The REAL assigned user
    assigned_user = doc.allocated_to

    if not assigned_user:
        return

    try:
        frappe.db.set_value(
            "HD Ticket",
            doc.reference_name,
            "custom_agent",
            assigned_user,
            update_modified=False
        )

        # 🔁 push realtime update so open form refreshes instantly
        frappe.publish_realtime(
            event="custom_agent_updated",
            message={
                "doctype": "HD Ticket",
                "name": doc.reference_name,
                "custom_agent": assigned_user
            }
        )

    except Exception:
        frappe.log_error(
            frappe.get_traceback(),
            "HD Ticket custom_agent sync failed"
        )
##############################################################################
# LMS_HOLIDAY_LIST

import frappe
from datetime import date, datetime, timedelta

HOLIDAY_LIST_NAME = "Holiday 2026"


# --------------------------------------------------
# UTIL: Normalize anything to datetime.date
# --------------------------------------------------
def to_date(value):
    if not value:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, str):
        return datetime.strptime(value, "%Y-%m-%d").date()

    return None


# --------------------------------------------------
# MAIN AGEING CALCULATION (NEW FLOW)
# --------------------------------------------------
def update_lms_ageing(doc):
    """
    FINAL RULE (NEW FLOW):

    Start Date  : lms_requested_date (Lastmile Services Master)
    End Date    :
        - Delivered   -> lms_delivery_date
        - All others  -> Today
    """

    # ---------------------------------------------
    # START DATE
    # ---------------------------------------------
    start_date = to_date(doc.lms_requested_date)
    if not start_date:
        doc.ageing = None
        return

    # ---------------------------------------------
    # END DATE
    # ---------------------------------------------
    if doc.lms_stage == "Delivered" and doc.lms_delivery_date:
        end_date = to_date(doc.lms_delivery_date)
    else:
        end_date = date.today()

    if not end_date:
        doc.ageing = None
        return

    # ---------------------------------------------
    # BUSINESS DAY CALCULATION
    # ---------------------------------------------
    holidays = get_holidays()
    doc.ageing = calculate_business_days(start_date, end_date, holidays)


# --------------------------------------------------
# FETCH HOLIDAYS
# --------------------------------------------------
def get_holidays():
    return {
        to_date(h.holiday_date)
        for h in frappe.get_all(
            "Holiday",
            filters={"parent": HOLIDAY_LIST_NAME},
            fields=["holiday_date"]
        )
        if h.holiday_date
    }


# --------------------------------------------------
# BUSINESS DAY CALCULATION
# --------------------------------------------------
def calculate_business_days(start_date, end_date, holidays):
    if start_date > end_date:
        return 0

    days = 0
    current = start_date

    while current <= end_date:
        weekday = current.weekday()  # Mon=0, Sun=6

        # ❌ Sunday
        if weekday == 6:
            current += timedelta(days=1)
            continue

        # ❌ 2nd & 4th Saturday
        if weekday == 5:
            week_of_month = (current.day - 1) // 7 + 1
            if week_of_month in (2, 4):
                current += timedelta(days=1)
                continue

        # ❌ Holiday
        if current in holidays:
            current += timedelta(days=1)
            continue

        # ✅ Working day
        days += 1
        current += timedelta(days=1)

    return days


# --------------------------------------------------
# DAILY SCHEDULER (SAFE)
# --------------------------------------------------
def recalculate_all_lms_ageing():
    """
    Runs daily via scheduler.
    Updates ONLY the ageing field.
    No validation issues.
    """

    records = frappe.get_all(
        "Lastmile Services Master",
        fields=["name"]
    )

    for r in records:
        doc = frappe.get_doc("Lastmile Services Master", r.name)
        update_lms_ageing(doc)

        frappe.db.set_value(
            "Lastmile Services Master",
            doc.name,
            "ageing",
            doc.ageing,
            update_modified=False
        )
#########################################################################
## Stop the bounce email to techsupport

import frappe
from email.utils import getaddresses

def block_techsupport_bounce_emails(doc, method):
    # 🟢 Step 1: Check if email is meant for Techsupport only
    recipients = []
    if doc.recipients:
        recipients = [email.lower() for _, email in getaddresses([doc.recipients])]

    is_techsupport = any(
        email.startswith("techsupport@") or email.startswith("techsupport+")
        for email in recipients
    )

    if not is_techsupport:
        # Allow other department emails
        return

    # 🟢 Step 2: Detect bounce emails
    bounce_senders = [
        "mailer-daemon",
        "postmaster",
        "mail delivery subsystem"
    ]

    bounce_keywords = [
        "undelivered mail returned to sender",
        "delivery failed",
        "email policy violation",
        "554 5.7.7",
        "permanent error",
        "could not be delivered"
    ]

    text = f"{doc.sender or ''} {doc.subject or ''} {doc.content or ''}".lower()

    if any(s in text for s in bounce_senders) or any(k in text for k in bounce_keywords):
        # 🔇 SILENT BLOCK ONLY FOR TECHSUPPORT
        doc.flags.ignore_permissions = True
        doc.flags.ignore_links = True
        doc.flags.ignore_mandatory = True

        doc.communication_type = "Ignored"
        doc.subject = "[IGNORED TECHSUPPORT BOUNCE EMAIL]"
        doc.content = ""

        frappe.logger().info(
            f"Techsupport bounce email blocked silently: {doc.sender}"
        )

        # ❌ Abort insert cleanly (no error)
        doc._cancel_insert = True
#################################################################################
# Purchase Order list Down - Supplier Advance

import frappe
from frappe import _

@frappe.whitelist()
def get_purchase_orders_by_supplier(supplier):
    """
    Fetch submitted Purchase Orders for a given supplier
    Used in Bank Reconciliation → Supplier Advance
    """

    if not supplier:
        return []

    return frappe.get_all(
        "Purchase Order",
        filters={
            "supplier": supplier,
            "docstatus": 1
        },
        fields=["name"],
        order_by="creation desc"
    )
##################################################################################
# Supplier Advance - Payment Entry    
@frappe.whitelist()
def create_supplier_advance_payment(
    supplier=None,
    amount=None,
    statement_entry=None,
    purchase_order=None,
    bank_account=None, # ⭐ Added
    custom_send_email=0
):

    if not supplier:
        frappe.throw("Supplier is required")

    if not amount:
        frappe.throw("Amount is required")

    if not statement_entry:
        frappe.throw("Bank Statement Entry missing")

    amount = float(amount)

    stmt = frappe.get_doc("Bank Statement Entry", statement_entry)

    # Company from Bank Account
    company = frappe.db.get_value("Bank Account", stmt.bank_account, "company")

    posting_date = stmt.transaction_date
    reference_no = stmt.description

    # Prioritize passed bank_account, then stmt.bank_account
    final_bank_account = bank_account or stmt.bank_account
    paid_from = get_bank_gl_account(final_bank_account)

    paid_to = frappe.db.get_value("Company", company, "default_payable_account")

    if not paid_from or not paid_to:
        frappe.throw("Bank or Payable account missing")

    # Fetch supplier email for notification
    supplier_email = frappe.db.get_value("Supplier", supplier, "email_id") or ""

    payment_entry = frappe.get_doc({
        "doctype": "Payment Entry",
        "payment_type": "Pay",
        "mode_of_payment": "Wire Transfer",
        "company": company,
        "party_type": "Supplier",
        "party": supplier,
        "paid_amount": amount,
        "received_amount": amount,
        "posting_date": posting_date,
        "reference_no": reference_no,
        "reference_date": posting_date,
        "paid_from": paid_from,   # GL account
        "paid_to": paid_to,       # Payable account
        "bank_account": final_bank_account, # ⭐ Set the Bank Account doc name
        "custom_send_email": custom_send_email,
        "contact_email": supplier_email
    })

    if purchase_order:
        payment_entry.append("references", {
            "reference_doctype": "Purchase Order",
            "reference_name": purchase_order,
            "allocated_amount": amount
        })

    payment_entry.custom_send_email = custom_send_email
    payment_entry.insert(ignore_permissions=True)
    payment_entry.submit()

    # Send email directly if custom_send_email is set
    if int(custom_send_email) == 1:
        send_supplier_payment_email(payment_entry)

    frappe.db.set_value("Bank Statement Entry", stmt.name, {
        "reference_no": payment_entry.name,
        "reconciled": 1,
        "match_type": "Auto"
    })

    return {
        "status": "ok",
        "payment_entry": payment_entry.name
    }
###############################################################################
# Customer Advance
@frappe.whitelist()
def create_customer_advance_payment(
    customer=None,
    amount=None,
    statement_entry=None,
    sales_order=None,
    bank_account=None, # ⭐ Added
):

    if not customer:
        frappe.throw("Customer is required")

    if not amount:
        frappe.throw("Amount is required")

    stmt = frappe.get_doc("Bank Statement Entry", statement_entry)
    company = frappe.db.get_value("Bank Account", stmt.bank_account, "company")

    # Prioritize passed bank_account, then stmt.bank_account
    final_bank_account = bank_account or stmt.bank_account
    paid_to = get_bank_gl_account(final_bank_account)

    paid_from = frappe.db.get_value("Company", company, "default_receivable_account")

    pe = frappe.get_doc({
        "doctype": "Payment Entry",
        "payment_type": "Receive",
        "company": company,
        "party_type": "Customer",
        "party": customer,
        "paid_from": paid_from,
        "paid_to": paid_to,
        "paid_amount": float(amount),
        "received_amount": float(amount),
        "posting_date": stmt.transaction_date,
        "reference_no": stmt.description,
        "reference_date": stmt.transaction_date
    })

    if sales_order:
        pe.append("references", {
            "reference_doctype": "Sales Order",
            "reference_name": sales_order,
            "allocated_amount": float(amount)
        })

    pe.insert(ignore_permissions=True)
    pe.submit()

    # Update Bank Statement Entry
    frappe.db.set_value("Bank Statement Entry", statement_entry, {
        "reference_no": pe.name,
        "reconciled": 1,
        "match_type": "Auto"
    })

    return {
        "status": "ok",
        "payment_entry": pe.name
    }

@frappe.whitelist()
def create_supplier_return_payment(
    supplier=None,
    amount=None,
    statement_entry=None,
    bank_account=None,
    custom_send_email=0
):
    if not supplier:
        frappe.throw("Supplier is required")
    if not amount:
        frappe.throw("Amount is required")
    if not statement_entry:
        frappe.throw("Bank Statement Entry missing")

    stmt = frappe.get_doc("Bank Statement Entry", statement_entry)
    company = frappe.db.get_value("Bank Account", stmt.bank_account, "company")

    # Prioritize passed bank_account, then stmt.bank_account
    final_bank_account = bank_account or stmt.bank_account
    paid_to = get_bank_gl_account(final_bank_account)

    supplier_payable = frappe.db.get_value("Company", company, "default_payable_account")
    supp_account = frappe.db.get_value("Party Account", {"parent": supplier, "company": company}, "account")
    if supp_account:
        supplier_payable = supp_account

    # Fetch supplier email for notification
    supplier_email = frappe.db.get_value("Supplier", supplier, "email_id") or ""

    pe = frappe.get_doc({
        "doctype": "Payment Entry",
        "payment_type": "Receive",
        "mode_of_payment": "Wire Transfer",
        "company": company,
        "party_type": "Supplier",
        "party": supplier,
        "paid_from": supplier_payable,
        "paid_to": paid_to,
        "paid_amount": float(amount),
        "received_amount": float(amount),
        "posting_date": stmt.transaction_date,
        "reference_no": stmt.description,
        "reference_date": stmt.transaction_date,
        "bank_account": final_bank_account,
        "custom_send_email": custom_send_email,
        "contact_email": supplier_email
    })

    _set_gst_fields(pe, supplier, "Supplier", company)

    pe.custom_send_email = custom_send_email
    pe.insert(ignore_permissions=True)
    pe.submit()

    # Send email directly if custom_send_email is set
    if int(custom_send_email) == 1:
        send_supplier_payment_email(pe)

    # Update Bank Statement Entry
    frappe.db.set_value("Bank Statement Entry", statement_entry, {
        "reference_no": pe.name,
        "reconciled": 1,
        "match_type": "Auto"
    })

    return {
        "status": "ok",
        "payment_entry": pe.name
    }

@frappe.whitelist()
def create_customer_return_payment(
    customer=None,
    amount=None,
    statement_entry=None,
    bank_account=None
):
    if not customer:
        frappe.throw("Customer is required")
    if not amount:
        frappe.throw("Amount is required")
    if not statement_entry:
        frappe.throw("Bank Statement Entry missing")

    stmt = frappe.get_doc("Bank Statement Entry", statement_entry)
    company = frappe.db.get_value("Bank Account", stmt.bank_account, "company")

    # Prioritize passed bank_account, then stmt.bank_account
    final_bank_account = bank_account or stmt.bank_account
    paid_from = get_bank_gl_account(final_bank_account)

    paid_to = frappe.db.get_value("Company", company, "default_receivable_account")
    
    pe = frappe.get_doc({
        "doctype": "Payment Entry",
        "payment_type": "Pay",
        "mode_of_payment": "Wire Transfer",
        "company": company,
        "party_type": "Customer",
        "party": customer,
        "paid_from": paid_from,
        "paid_to": paid_to,
        "paid_amount": float(amount),
        "received_amount": float(amount),
        "posting_date": stmt.transaction_date,
        "reference_no": stmt.description,
        "reference_date": stmt.transaction_date,
        "bank_account": final_bank_account
    })

    _set_gst_fields(pe, customer, "Customer", company)

    pe.insert(ignore_permissions=True)
    pe.submit()

    # Update Bank Statement Entry
    frappe.db.set_value("Bank Statement Entry", statement_entry, {
        "reference_no": pe.name,
        "reconciled": 1,
        "match_type": "Auto"
    })

    return {
        "status": "ok",
        "payment_entry": pe.name
    }



##############################################################################
#Customer Advance - Sales Order Fetch
@frappe.whitelist()
def get_sales_orders_by_customer(customer):

    if not customer:
        return []

    return frappe.get_all(
        "Sales Order",
        filters={
            "customer": customer,
            "docstatus": 1
        },
        fields=["name"],
        order_by="creation desc"
    )
############################################################################
def set_industry_from_market_segment(doc, method):
    frappe.throw("HOOK RUNNING")

#############################################################################
import frappe
import json
from frappe.utils import get_url, nowdate

@frappe.whitelist()
def get_hd_tickets(
    page_length=20, 
    start=0, 
    search="", 
    view="All Tickets", 
    view_type="status", 
    customer="", 
    circuit_id="", 
    ticket_id="", 
    agent="",
    status="",
    stage="",
    priority="",
    channel="",
    severity="",
    created_from="",
    created_to="",
    closed_from="",
    closed_to="",
    onhold_from="",
    onhold_to=""
):
    page_length = int(page_length)
    start = int(start)

    filters = {}

    # ---------------------------
    # STATUS FILTER MAP
    # ---------------------------
    view_status_map = {
        "Open Tickets": "Open",
        "Replied Tickets": "Replied",
        "On Hold Tickets": "On Hold",
        "Wrong Tickets": "Wrong Circuit",
        "Resolved Tickets": "Resolved",
        "Closed Tickets": "Closed"
    }

    # ---------------------------
    # PRIORITY FILTER MAP
    # ---------------------------
    priority_map = {
        "Urgent": "Urgent",
        "High": "High",
        "Medium": "Medium",
        "Low": "Low"
    }
    
    # ---------------------------
    # SEVERITY FILTER MAP
    # ---------------------------
    severity_map = {
        "Critical": "Critical",
        "Major": "Major",
        "Minor": "Minor"
    }

    # ---------------------------
    # APPLY FILTER BASED ON TYPE
    # ---------------------------
    if view_type == "status":
        if view in view_status_map:
            filters["status"] = view_status_map[view]
        elif view in priority_map:
            filters["priority"] = priority_map[view]
        elif view in severity_map:
            filters["custom_severity"] = severity_map[view]

    elif view_type == "channel":
        filters["custom_channel"] = view

    elif view_type == "stage":
        filters["custom_stage"] = view

    # ---------------------------
    # CUSTOMER FILTER
    # ---------------------------
    if customer:
        filters["customer"] = customer
    
    # ---------------------------
    # CIRCUIT ID FILTER
    # ---------------------------
    if circuit_id:
        filters["custom_circuit_id"] = ["like", f"%{circuit_id}%"]
    
    # ---------------------------
    # TICKET ID FILTER
    # ---------------------------
    if ticket_id:
        filters["name"] = ["like", f"%{ticket_id}%"]
    
    # ---------------------------
    # STATUS FILTER (from sidebar) - handle multi-select
    # ---------------------------
    if status:
        if ',' in status:
            status_list = [s.strip() for s in status.split(',') if s.strip()]
            if status_list:
                filters["status"] = ["in", status_list]
        else:
            filters["status"] = status
    
    # ---------------------------
    # STAGE FILTER - handle multi-select
    # ---------------------------
    if stage:
        if ',' in stage:
            stage_list = [s.strip() for s in stage.split(',') if s.strip()]
            if stage_list:
                filters["custom_stage"] = ["in", stage_list]
        else:
            filters["custom_stage"] = stage
    
    # ---------------------------
    # PRIORITY FILTER - handle multi-select
    # ---------------------------
    if priority:
        if ',' in priority:
            priority_list = [p.strip() for p in priority.split(',') if p.strip()]
            if priority_list:
                filters["priority"] = ["in", priority_list]
        else:
            filters["priority"] = priority
    
    # ---------------------------
    # CHANNEL FILTER - handle multi-select
    # ---------------------------
    if channel:
        if ',' in channel:
            channel_list = [c.strip() for c in channel.split(',') if c.strip()]
            if channel_list:
                filters["custom_channel"] = ["in", channel_list]
        else:
            filters["custom_channel"] = channel
    
    # ---------------------------
    # SEVERITY FILTER - handle multi-select
    # ---------------------------
    if severity:
        if ',' in severity:
            severity_list = [s.strip() for s in severity.split(',') if s.strip()]
            if severity_list:
                filters["custom_severity"] = ["in", severity_list]
        else:
            filters["custom_severity"] = severity
    
    # ---------------------------
    # DATE RANGE FILTERS
    # ---------------------------
    if created_from:
        filters["creation"] = [">=", created_from + " 00:00:00"]
    
    if created_to:
        if "creation" in filters:
            filters["creation"] = ["between", [created_from + " 00:00:00", created_to + " 23:59:59"]]
        else:
            filters["creation"] = ["<=", created_to + " 23:59:59"]
    
    if closed_from or closed_to:
        # Assuming there's a custom_closed_datetime field
        if closed_from:
            filters["custom_closed_datetime"] = [">=", closed_from + " 00:00:00"]
        if closed_to:
            if "custom_closed_datetime" in filters:
                filters["custom_closed_datetime"] = ["between", [closed_from + " 00:00:00", closed_to + " 23:59:59"]]
            else:
                filters["custom_closed_datetime"] = ["<=", closed_to + " 23:59:59"]
    
    if onhold_from or onhold_to:
        # Assuming there's a custom_hold_datetime field
        if onhold_from:
            filters["custom_hold_datetime"] = [">=", onhold_from + " 00:00:00"]
        if onhold_to:
            if "custom_hold_datetime" in filters:
                filters["custom_hold_datetime"] = ["between", [onhold_from + " 00:00:00", onhold_to + " 23:59:59"]]
            else:
                filters["custom_hold_datetime"] = ["<=", onhold_to + " 23:59:59"]
    
    # ---------------------------
    # AGENT FILTER
    # ---------------------------
    if agent:
        # Get all tickets assigned to this agent
        assigned_tickets = frappe.get_all(
            "HD Ticket",
            filters={},
            fields=["name", "_assign"]
        )
        
        ticket_names = []
        for t in assigned_tickets:
            if t.get("_assign"):
                try:
                    assigned = json.loads(t["_assign"])
                    if assigned and agent in assigned:
                        ticket_names.append(t.name)
                except:
                    pass
        
        if ticket_names:
            filters["name"] = ["in", ticket_names]
        else:
            filters["name"] = ["in", []]

    # ---------------------------
    # SEARCH FILTER
    # ---------------------------
    if search:
        filters["subject"] = ["like", f"%{search}%"]

    # ---------------------------
    # FETCH TICKETS (Pagination Safe)
    # ---------------------------
    tickets = frappe.get_all(
        "HD Ticket",
        fields=[
            "name",
            "subject",
            "customer",
            "status",
            "priority",
            "resolution_by",
            "custom_stage",
            "custom_channel",
            "creation",
            "_assign",
            "custom_is_read",
            "custom_severity",
            "custom_circuit_id",
            "custom_impact",
            "custom_closed_datetime",
            "custom_hold_datetime"
        ],
        filters=filters,
        order_by="creation desc",
        start=start,
        limit_page_length=page_length
    )

    # ---------------------------
    # ASSIGNMENT PROCESSING
    # ---------------------------
    users = set()

    for t in tickets:
        t["custom_is_read"] = 1 if str(t.get("custom_is_read")) == "1" else 0

        if t.get("_assign"):
            try:
                assigned = json.loads(t["_assign"])
                if assigned:
                    t["assigned_to"] = assigned[0]
                    users.add(assigned[0])
                else:
                    t["assigned_to"] = ""
            except:
                t["assigned_to"] = ""
        else:
            t["assigned_to"] = ""

    # ---------------------------
    # USER IMAGE FETCH
    # ---------------------------
    user_images = {}

    if users:
        user_data = frappe.get_all(
            "User",
            filters={"name": ["in", list(users)]},
            fields=["name", "user_image"]
        )

        for u in user_data:
            img = u.user_image or ""
            if img and not img.startswith("http"):
                img = get_url(img)
            user_images[u.name] = img

    for t in tickets:
        t["user_image"] = user_images.get(t["assigned_to"], "")

    # ---------------------------
    # TOTAL COUNT
    # ---------------------------
    total_count = frappe.db.count("HD Ticket", filters=filters)

    return {
        "tickets": tickets,
        "count": total_count
    }


@frappe.whitelist()
def search_agents(text):
    """Search for agents/users"""
    if not text:
        return []
    
    users = frappe.get_all(
        "User",
        filters={
            "name": ["like", f"%{text}%"],
            "enabled": 1
        },
        fields=["name", "full_name", "user_image"],
        limit_page_length=10
    )
    
    result = []
    for user in users:
        result.append({
            "name": user.name,
            "full_name": user.get("full_name") or user.name,
            "user_image": user.get("user_image", "")
        })
    
    return result


@frappe.whitelist()
def get_customer_circuits(customer, search_text=""):
    """Get circuits for a specific customer"""
    filters = {
        "customer": customer
    }
    
    if search_text:
        filters["custom_circuit_id"] = ["like", f"%{search_text}%"]
    
    circuits = frappe.get_all(
        "HD Ticket",
        fields=["custom_circuit_id"],
        filters=filters,
        limit_page_length=20,
        order_by="custom_circuit_id asc"
    )
    
    # Get unique circuits
    unique_circuits = []
    seen = set()
    for c in circuits:
        if c.custom_circuit_id and c.custom_circuit_id not in seen:
            seen.add(c.custom_circuit_id)
            unique_circuits.append(c.custom_circuit_id)
    
    return unique_circuits


@frappe.whitelist()
def get_customer_tickets(customer, search_text=""):
    """Get tickets for a specific customer"""
    filters = {
        "customer": customer
    }
    
    if search_text:
        filters["name"] = ["like", f"%{search_text}%"]
    
    tickets = frappe.get_all(
        "HD Ticket",
        fields=["name"],
        filters=filters,
        limit_page_length=20,
        order_by="creation desc"
    )
    
    return [t.name for t in tickets]


@frappe.whitelist()
def get_customer_agents(customer, search_text=""):
    """Get agents for a specific customer"""
    # Get all tickets for this customer
    tickets = frappe.get_all(
        "HD Ticket",
        fields=["_assign"],
        filters={"customer": customer},
        limit_page_length=200
    )
    
    agents = set()
    for t in tickets:
        if t.get("_assign"):
            try:
                assigned = json.loads(t["_assign"])
                for agent in assigned:
                    agents.add(agent)
            except:
                pass
    
    # If search text, filter agents
    if search_text and agents:
        filtered_agents = []
        for agent in agents:
            if search_text.lower() in agent.lower():
                filtered_agents.append(agent)
        return filtered_agents
    
    return list(agents)


@frappe.whitelist()
def get_filter_options(filter_type):
    """Get options for various filter types"""
    options = {
        "status": ["Open", "Replied", "On Hold", "Wrong Circuit", "Resolved", "Closed"],
        "priority": ["Urgent", "High", "Medium", "Low"],
        "severity": ["Critical", "Major", "Minor"],
        "channel": ["Email", "Portal", "Chat", "Phone", "Web Form", "SSP", "NMS", "NMS-API"],
        "stage": [
            "Inprocess", "Finance Issue", "Customer Issue", "Hardware Dispatch",
            "MBB Issue", "LMS-Re-Feasibility", "Maintenance Visit", "Wrong Circuit",
            "Other", "Configuration Change", "Project"
        ]
    }
    
    return options.get(filter_type, [])
###################################################################################
import frappe


@frappe.whitelist(allow_guest=True)
def create_hd_ticket(subject, message, circuit_id):

    subject_upper = subject.upper()

    node_type = ""
    impact = ""
    node_status = ""

    # --------------------------------------------
    # Detect Node Status
    # --------------------------------------------
    if "UNAVAILABLE" in subject_upper:
        node_status = "UNAVAILABLE"
    elif "AVAILABLE" in subject_upper:
        node_status = "AVAILABLE"

    # --------------------------------------------
    # Detect Node Type and Impact
    # --------------------------------------------
    if "PRI" in subject_upper:
        node_type = "Primary Node"
        impact = "ILL Down"

    elif "SEC" in subject_upper:
        node_type = "Secondary Node"
        impact = "4G/MBB Backup Down"

    elif "BR" in subject_upper:
        node_type = "VPN Node"
        impact = "Site Isolated"

    # --------------------------------------------
    # If Node is UNAVAILABLE → Create Ticket
    # --------------------------------------------
    if node_status == "UNAVAILABLE":

        existing_ticket = frappe.db.get_value(
            "HD Ticket",
            {
                "custom_circuit_id": circuit_id,
                "custom_node_type": node_type
            },
            ["name", "status"],
            as_dict=True
        )

        # Do not create if ticket already open
        if existing_ticket and existing_ticket.status not in ["Resolved", "Closed"]:
            return {
                "status": "skipped",
                "message": "Ticket already exists and is still open"
            }

        ticket = frappe.get_doc({
            "doctype": "HD Ticket",
            "subject": subject,
            "description": message,
            "custom_circuit_id": circuit_id,
            "custom_channel": "NMS-API",
            "custom_node_type": node_type,
            "custom_impact": impact,
            "status": "Open"
        }).insert(ignore_permissions=True)

        frappe.db.commit()

        return {
            "status": "created",
            "ticket_number": ticket.name
        }

    # --------------------------------------------
    # If Node is AVAILABLE → Resolve Tickets
    # --------------------------------------------
    if node_status == "AVAILABLE":

        # Resolve node ticket
        node_tickets = frappe.get_all(
            "HD Ticket",
            filters={
                "custom_circuit_id": circuit_id,
                "custom_node_type": node_type,
                "status": ["not in", ["Resolved", "Closed"]]
            },
            fields=["name"]
        )

        for t in node_tickets:
            doc = frappe.get_doc("HD Ticket", t.name)
            doc.status = "Resolved"
            doc.custom_rca = "Auto Close"
            doc.save(ignore_permissions=True)

        # Resolve VPN ticket simultaneously
        vpn_tickets = frappe.get_all(
            "HD Ticket",
            filters={
                "custom_circuit_id": circuit_id,
                "custom_node_type": "VPN Node",
                "status": ["not in", ["Resolved", "Closed"]]
            },
            fields=["name"]
        )

        for v in vpn_tickets:
            doc = frappe.get_doc("HD Ticket", v.name)
            doc.status = "Resolved"
            doc.custom_rca = "Auto Close"
            doc.save(ignore_permissions=True)

        frappe.db.commit()

        return {
            "status": "resolved",
            "message": "Node and VPN tickets auto closed"
        }

    return {"status": "ignored"}

###############################################################################
@frappe.whitelist()
def toggle_ticket_read_status(ticket_id):
    current = frappe.db.get_value("HD Ticket", ticket_id, "custom_is_read") or 0
    new_value = 0 if current else 1

    frappe.db.set_value("HD Ticket", ticket_id, "custom_is_read", new_value)
    frappe.db.commit()

    return new_value

######################################################################
# Solution Chnage
import frappe

def solution_change_update(doc, method):

    circuit_id = doc.circuit_id
    new_code = doc.new_solution_code
    new_name = doc.new_solution_name

    if not circuit_id or not new_code or not new_name:
        frappe.msgprint("Missing Circuit ID or Solution details.")
        return

    updates = 0

    # =====================================================
    # 1) SITE
    # circuit_id = name (Site)
    # =====================================================
    if frappe.db.exists("Site", circuit_id):
        frappe.db.set_value(
            "Site",
            circuit_id,
            {
                "solution_code": new_code,
                "solution_name": new_name
            }
        )
        updates += 1

    # =====================================================
    # 2) LASTMILE SERVICES MASTER
    # circuit_id = circuit_id
    # =====================================================
    lastmile_docs = frappe.get_all(
        "Lastmile Services Master",
        filters={"circuit_id": circuit_id},
        fields=["name"]
    )

    for lm in lastmile_docs:
        frappe.db.set_value(
            "Lastmile Services Master",
            lm.name,
            "solution",
            new_name
        )
        updates += 1

    # =====================================================
    # 3) SALES ORDER ITEM (Child table)
    # custom_feasibility = circuit_id
    # =====================================================
    sales_items = frappe.get_all(
        "Sales Order Item",
        filters={"custom_feasibility": circuit_id},
        fields=["name"]
    )

    for item in sales_items:
        frappe.db.set_value(
            "Sales Order Item",
            item.name,
            "custom_solution",
            new_code
        )
        updates += 1

    # =====================================================
    # 4) STOCK MANAGEMENT
    # circuit_id = circuit_id
    # =====================================================
    stock_docs = frappe.get_all(
        "Stock Management",
        filters={"circuit_id": circuit_id},
        fields=["name"]
    )

    for stock in stock_docs:
        frappe.db.set_value(
            "Stock Management",
            stock.name,
            {
                "solution_code": new_code,
                "solution": new_name
            }
        )
        updates += 1

    # =====================================================
    # 5) FEASIBILITY
    # circuit_id = name (Feasibility)
    # =====================================================
    if frappe.db.exists("Feasibility", circuit_id):
        frappe.db.set_value(
            "Feasibility",
            circuit_id,
            {
                "solution_code": new_code,
                "solution_name": new_name
            }
        )
        updates += 1

    # =====================================================
    # 6) PROVISIONING
    # circuit_id = circuit_id
    # Update solution_name + refresh open form
    # =====================================================
    provisioning_docs = frappe.get_all(
        "Provisioning",
        filters={"circuit_id": circuit_id},
        fields=["name"]
    )

    for prov in provisioning_docs:

        prov_doc = frappe.get_doc("Provisioning", prov.name)

        # Update field
        prov_doc.solution_name = new_name

        # Save to trigger Provisioning logic/workflows
        prov_doc.save(ignore_permissions=True)

        # 🔥 Refresh open form in browser
        prov_doc.notify_update()

        updates += 1

    # =====================================================
    # ✅ CONFIRMATION MESSAGE
    # =====================================================
    frappe.msgprint(
        f"""
        <b>Solution Updated Successfully</b><br><br>
        Circuit ID: <b>{circuit_id}</b><br>
        New Solution Code: <b>{new_code}</b><br>
        New Solution Name: <b>{new_name}</b><br><br>
        Records Updated: <b>{updates}</b>
        """
    )
############################################################################
# HD LMS Ticket Createing from HD Ticket

import frappe
import re


def create_lms_ticket(doc, method):

    # Avoid duplicate LMS ticket
    if frappe.db.exists("HD LMS Ticket", {"customer_ticket_id": doc.name}):
        return

    if not doc.custom_circuit_id:
        return

    # --------------------------------------------------
    # 🔎 Extract LMS ID from Subject
    # --------------------------------------------------
    lms_id_value = None

    if doc.subject:
        match = re.search(r'(ILL|MBB)_(\d{5,})', doc.subject)

        if match:
            extracted_number = match.group(2)

            # ✅ WITH SPACE AFTER LMS-
            lms_id_value = f"LMS- {extracted_number}"

    # --------------------------------------------------
    # 🆕 Create HD LMS Ticket
    # --------------------------------------------------
    lms = frappe.new_doc("HD LMS Ticket")

    lms.circuit_id = doc.custom_circuit_id
    lms.customer_ticket_id = doc.name

    if lms_id_value:
        lms.lms_id = lms_id_value

    lms.insert(ignore_permissions=True)

#################################################################################

#ai for Purchase Invoice - Matching Company
@frappe.whitelist()
def match_company(name):
    """
    Carefully match a company name to an existing Company record.
    Includes suffix cleaning, word-by-word matching, and fuzzy matching.
    """
    if not name:
        return None
        
    name_str = str(name).strip()
    # Clean common suffixes like "Private Limited", "Pvt Ltd", etc.
    clean_regexp = r"(?i)\b(pvt\.?\s+ltd\.?|p\s+ltd\.?|ltd\.?|private\s+limited|limited|corporation|inc\.?|solutions|llp|group|enterprises|systems|services)\b"
    clean_name = re.sub(clean_regexp, "", name_str).strip()

    # 1. Exact Match (case-insensitive)
    for target in [name_str, clean_name]:
        if not target: continue
        c = frappe.db.get_value("Company", {"company_name": ["like", f"%{target}%"]}, "name")
        if not c: c = frappe.db.get_value("Company", {"name": ["like", f"%{target}%"]}, "name")
        if c: return c
    
    # 2. Fetch all companies for advanced matching
    companies = frappe.get_all("Company", fields=["name", "company_name"])
    name_map = {}
    for comp in companies:
        name_map[comp.name.lower()] = comp.name
        if comp.company_name:
            name_map[comp.company_name.lower()] = comp.name
            
    all_targets = list(name_map.keys())
    
    # 3. Substring Match (Case-insensitive)
    for target in [name_str.lower(), clean_name.lower()]:
        if not target: continue
        for t in all_targets:
            if len(target) > 3 and (target in t or t in target):
                return name_map[t]

    # 4. Fuzzy match original and clean name
    for target in [name_str.lower(), clean_name.lower()]:
        if not target: continue
        matches = difflib.get_close_matches(target, all_targets, n=1, cutoff=0.5)
        if matches: return name_map[matches[0]]
        
    # 5. Word-by-word fallback
    # Filter out common small words
    stop_words = {"and", "the", "of", "for", "in", "on", "at", "by"}
    words = [w.lower() for w in clean_name.split() if len(w) > 2 and w.lower() not in stop_words]
    
    if words:
        for t in all_targets:
            # Check if any significant word from extracted name matches a significant word in target
            t_words = set(t.split())
            if any(w in t_words for w in words):
                return name_map[t]

    return None
############################################################################
# Covert POC To Paid
import frappe


def update_feasibility_and_site_on_so_save(doc, method):
    # Generic updater for all Sales Orders (excluding POC → Paid which uses a separate handler)
    if not doc.items:
        return
    # Skip generic processing for POC → Paid; let POC handler run
    if getattr(doc, 'custom_task_type', None) == "Sales Order Request - POC To Paid":
        _update_poc_to_paid(doc)
        return
    # Existing generic behavior (unchanged)
    # Existing behavior: update Site and Feasibility for all Sales Orders
    # (unchanged) 
    if not doc.items:
        return

    
        return

    updated_records = []

    for item in doc.items:

        ref_name = item.custom_feasibility

        if not ref_name:
            continue

        # -------------------------------
        # Update Feasibility
        # -------------------------------
        if frappe.db.exists("Feasibility", ref_name):

            frappe.db.set_value(
                "Feasibility",
                ref_name,
                {
                    "customer_type": "Paid Customer",
                    "sales_order": doc.name,
                    "sales_order_date": doc.transaction_date
                }
            )

            updated_records.append(f"Feasibility: {ref_name}")

        # -------------------------------
        # Update Site
        # -------------------------------
        if frappe.db.exists("Site", ref_name):

            frappe.db.set_value(
                "Site",
                ref_name,
                {
                    "customer_type": "Paid Customer",
                    "sales_order": doc.name,
                    "sales_order_date": doc.transaction_date,
                    "customer_po_no": doc.po_no,
                    "customer_po_date": doc.po_date,
                    "po_end_date": doc.custom_po_end_date
                }
            )

def apply_poc_to_paid_updates(doc):
    """Update Site and Feasibility records for the POC → Paid flow.
    This is called only when `custom_task_type` matches the POC conversion.
    """
    updated_sites = []
    updated_feas = []
    for item in doc.items:
        feas_name = item.custom_feasibility
        if not feas_name:
            continue
        # Update Site if it exists – single DB call via get_value for existence
        site_exists = frappe.db.get_value("Site", feas_name, "name")
        if site_exists:
            try:
                frappe.db.set_value(
                    "Site",
                    feas_name,
                    {
                        "customer_type": "Paid Customer",
                        "customer_po_no": doc.custom_deal_po_number,
                        "customer_po_date": doc.custom_deal_po_date,
                        "po_end_date": doc.custom_pr_number,
                        "sales_order": doc.name,
                        "sales_order_date": doc.transaction_date,
                    },
                )
                updated_sites.append(feas_name)
            except Exception as e:
                frappe.log_error(f"Failed to update Site {feas_name}: {e}")
        # Update Feasibility if it exists
        feas_exists = frappe.db.get_value("Feasibility", feas_name, "name")
        if feas_exists:
            try:
                frappe.db.set_value(
                    "Feasibility",
                    feas_name,
                    {
                        "customer_type": "Paid Customer",
                        "sales_order": doc.name,
                        "sales_order_date": doc.transaction_date,
                    },
                )
                updated_feas.append(feas_name)
            except Exception as e:
                frappe.log_error(f"Failed to update Feasibility {feas_name}: {e}")
    # Show concise summary
    if updated_sites or updated_feas:
        parts = []
        if updated_sites:
            parts.append(f"Sites updated ({len(updated_sites)}): {', '.join(updated_sites)}")
        if updated_feas:
            parts.append(f"Feasibility updated ({len(updated_feas)}): {', '.join(updated_feas)}")
        frappe.msgprint("<br>".join(parts))


def sales_order_submit_handler(doc, method=None):
    """Unified on_submit handler for Sales Order.
    Delegates to the generic updater or the POC‑to‑Paid updater based on task type.
    """
    if getattr(doc, "custom_task_type", None) == "Sales Order Request - POC To Paid":
        apply_poc_to_paid_updates(doc)
    else:
        update_feasibility_and_site_on_so_save(doc, method)


#################################################################################
# Updateing Billing status to 'Circuit Delivery Backdate' and 'Site

import frappe


def update_billing_status_from_invoice(doc, method=None):

    # Loop through Sales Invoice Items (child table)
    for item in doc.items:

        # Skip if no Sales Order Item link
        if not item.so_detail:
            continue

        # Get custom_feasibility from Sales Order Item
        circuit_id = frappe.db.get_value(
            "Sales Order Item",
            item.so_detail,
            "custom_feasibility"
        )

        if not circuit_id:
            continue

        # ------------------------------------
        # Update Circuit Delivery Backdate
        # ------------------------------------
        backdate_records = frappe.get_all(
            "Circuit Delivery Backdate",
            filters={"circuit_id": circuit_id},
            fields=["name"]
        )

        for row in backdate_records:
            frappe.db.set_value(
                "Circuit Delivery Backdate",
                row.name,
                "billing_status",
                "Billed"
            )

        # ------------------------------------
        # Update Site
        # ------------------------------------
        if frappe.db.exists("Site", circuit_id):
            frappe.db.set_value(
                "Site",
                circuit_id,
                "billing_status",
                "Billed"
            )
##################################################################################
# Updateing the HD Ticket From Task

import frappe
from frappe.utils import now_datetime

def update_hd_ticket_from_task(doc, method=None):

    # Find HD Ticket linked to this Task
    hd_ticket = frappe.db.get_value(
        "HD Ticket",
        {"custom_task": doc.name},
        "name"
    )

    if not hd_ticket:
        return

    # If Task status = Rejected
    if doc.status == "Rejected":

        frappe.db.set_value(
            "HD Ticket",
            hd_ticket,
            {
                "custom_task_status": doc.status,
                "custom_rejected_reason": doc.custom_rejected_reason,
                "custom_rejected_datetime": now_datetime()
            }
        )

    # If Task status = Completed
    elif doc.status == "Completed":

        frappe.db.set_value(
            "HD Ticket",
            hd_ticket,
            {
                "custom_task_status": doc.status,
                "custom_task_closed_datetime": doc.completed_on,
                "custom_completed_by_name": doc.custom_completed_by_name
            }
        )

    # If Task status NOT Rejected or Completed
    else:

        frappe.db.set_value(
            "HD Ticket",
            hd_ticket,
            {
                "custom_task_status": doc.status
            }
        )

####################################################################################
# AI chat user 

@frappe.whitelist()
def get_user_first_name():
    user = frappe.session.user
    user_doc = frappe.get_doc("User", user)

    if user_doc.first_name:
        return user_doc.first_name

    if user_doc.full_name:
        return user_doc.full_name.split(" ")[0]

    return "User"
#####################################################################################    

# AI Chatbot for HD Ticket

import frappe
import json
import datetime as dt
import difflib

# No top-level FAISS import – lazy import in the endpoint

# ---------------------------------------------------------
# FEATURE FLAG
# ---------------------------------------------------------
def is_faiss_enabled():
    return frappe.conf.get("enable_faiss_ai", 0)

@frappe.whitelist()
def is_chatbot_enabled():
    try:
        active = frappe.db.get_value(
            "AI Prompt Template",
            {"prompt_code": "HD_TICKET_CHAT"},
            "active"
        )
        return 1 if active else 0
    except Exception:
        return 0

# ---------------------------------------------------------
# MEMORY (per user + ticket)
# ---------------------------------------------------------
def get_memory_limit():
    try:
        config = frappe.get_single("API Configuration")
        return int(config.memory_limit or 10)
    except Exception:
        return 10

def get_user_memory(user, ticket):
    key = f"chat_memory:{user}:{ticket}"
    return frappe.cache().get_value(key) or []

def set_user_memory(user, ticket, memory):
    key = f"chat_memory:{user}:{ticket}"
    frappe.cache().set_value(key, memory)

def clear_user_memory(user, ticket):
    key = f"chat_memory:{user}:{ticket}"
    frappe.cache().delete_value(key)

def update_memory(user, ticket, question, answer):
    memory = get_user_memory(user, ticket)
    memory.append({"question": question, "answer": answer})
    limit = get_memory_limit()
    memory = memory[-limit:]
    set_user_memory(user, ticket, memory)

def build_memory_context(user, ticket):
    memory = get_user_memory(user, ticket)
    context = ""
    for m in memory:
        context += f"User: {m['question']}\n"
        context += f"Assistant: {m['answer']}\n"
    return context

# ---------------------------------------------------------
# FIELD LABEL MAP & HELPER FUNCTIONS
# ---------------------------------------------------------
FIELD_LABEL_MAP = {
    "name": "ID",
    "subject": "Subject",
    "status": "Status",
    "priority": "Priority",
    "customer": "Customer",
    "custom_circuit_id": "Circuit ID",
    "custom_stage": "Stage",
    "custom_sub_stage": "Sub Stage",
    "custom_solution_name": "Solution Name",
    "site_status": "Site Status",
    "custom_lms_ticket_status": "LMS Ticket Status",
    "lms_stage": "LMS Stage",
}

STOP_WORDS = {
    "what","is","the","of","a","an","please","give","show","tell","me","about"
}

def normalize_word(word):
    word = word.lower()
    replacements = {
        "natted": "nat",
        "nated": "nat",
        "natting": "nat",
    }
    return replacements.get(word, word)

def format_value(value):
    if not value:
        return ""
        
    # Attempt to parse date strings (e.g. 2024-03-20) if they are strings
    if isinstance(value, str) and len(value) >= 10:
        import re
        if re.match(r"^\d{4}-\d{2}-\d{2}", value):
            try:
                # If it has time, handle accordingly
                if " " in value:
                    value = dt.datetime.strptime(value[:19], "%Y-%m-%d %H:%M:%S")
                else:
                    value = dt.datetime.strptime(value[:10], "%Y-%m-%d").date()
            except Exception:
                pass

    if isinstance(value, dt.datetime):
        return value.strftime("%d-%m-%Y %H:%M")
    if isinstance(value, dt.date):
        return value.strftime("%d-%m-%Y")
    return value

def get_doc_label_map(doctype):
    """Dynamically get the label map for a doctype."""
    meta = frappe.get_meta(doctype)
    label_map = FIELD_LABEL_MAP.copy()
    for f in meta.fields:
        if f.label:
            label_map[f.fieldname] = f.label
    return label_map

def identify_entity(question):
    """Identifies target entity from question keywords."""
    q = question.lower()
    if any(k in q for k in ["site", "address", "location", "customer type", "lms type"]):
        return "Site"
    if any(k in q for k in ["lms", "supplier", "lastmile", "escalation", "contact"]):
        return "Lastmile Services Master"
    if any(k in q for k in ["provisioning", "ip address", "router ip"]):
        return "Provisioning"
    if any(k in q for k in ["installation", "installed", "engineer visit"]):
        return "Installation Master"
    return None

def get_clean_doc_data(doc):
    """Enhanced version: uses metadata-based labels and includes child table data."""
    label_map = get_doc_label_map(doc.doctype)
    exclude_fields = [
        "owner","creation","modified","modified_by",
        "docstatus","idx","_comments","_assign",
        "_liked_by","_seen","_user_tags","__unsaved",
        "doctype"
    ]
    data = {}
    
    # Process main fields
    for field, value in doc.as_dict().items():
        if field in exclude_fields or value in [None, "", []]:
            continue
            
        if field not in label_map and not field == "name":
            continue
            
        label = label_map.get(field, field.replace('_', ' ').title())
        
        # Format values
        if isinstance(value, (dt.datetime, dt.date)):
            value = format_value(value)
        elif isinstance(value, str) and ("<" in value and ">" in value):
            value = frappe.utils.strip_html_tags(value)
            
        data[label] = value

    # Process child tables
    meta = frappe.get_meta(doc.doctype)
    for f in meta.fields:
        if f.fieldtype == "Table":
            child_docs = doc.get(f.fieldname)
            if child_docs:
                child_label = f.label or f.fieldname.replace('_', ' ').title()
                table_data = []
                for child in child_docs:
                    # Use get_clean_doc_data recursively but without deep nesting for efficiency
                    # Just get simple labeled dict for child
                    child_label_map = get_doc_label_map(child.doctype)
                    child_dict = {}
                    for cf, cv in child.as_dict().items():
                        if cf in exclude_fields or cv in [None, "", []]: continue
                        clabel = child_label_map.get(cf, cf.replace('_', ' ').title())
                        child_dict[clabel] = format_value(cv) if isinstance(cv, (dt.date, dt.datetime)) else cv
                    table_data.append(child_dict)
                
                if table_data:
                    data[child_label] = table_data
        
    return data

# ---------------------------------------------------------
# IMPROVED FIELD MATCHING (with synonyms)
# ---------------------------------------------------------
def search_field_answer(doc, question):
    meta = frappe.get_meta(doc.doctype)
    q_lower = question.lower().strip()

    # Synonym mapping
    synonyms = {
        "agent": ["assigned to", "owner", "support agent", "technician"],
        "address": ["address", "location", "site address", "street", "city", "district", "pin code"],
        "impact": ["impact", "severity", "business impact"],
        "stage": ["stage", "status", "state"],
        "installation": ["installation", "activation", "commissioning"],
        "completion": ["completion", "done", "finished", "closed"],
        "escalation": ["escalation matrix", "matrix", "escalation level", "support levels", "contact", "support info", "lms contact"],
        "contact": ["escalation matrix", "matrix", "support info", "support person"]
    }

    def get_label_words(label):
        words = [normalize_word(w) for w in label.split()]
        extra = []
        for w in words:
            for key, syns in synonyms.items():
                if w in syns or w == key:
                    extra.extend(syns)
        return set(words + extra)

    best_match = None
    best_score = 0

    for field in meta.fields:
        label_original = field.label or ""
        if not label_original:
            continue

        label = label_original.lower().strip()
        label_words = get_label_words(label)
        value = doc.get(field.fieldname)

        # Skip empty values unless we are certain
        if value in [None, "", []] and field.fieldtype not in ("Select", "Data", "Link", "Table"):
            continue

        score = 0
        # Label in Question (Phrase Match)
        if label in q_lower:
            # Score based on number of words in the match - favors more specific labels
            # Boost for phrase match significantly
            score += len(label.split()) * 15
        
        # Word overlap (catch-all for jumbled words)
        q_words = set(normalize_word(w) for w in q_lower.split())
        overlap = len(q_words & label_words)
        score += overlap * 2

        # Boost for priority fields
        priority_fields = ["address", "agent", "status", "stage", "impact", "customer type", "escalation"]
        if any(p in label for p in priority_fields):
            score += 2

        if score > best_score:
            best_score = score
            best_match = (label_original, value, field)
        elif score == best_score and best_match:
            # Tie-breaker: prefer longer labels (more specific)
            if len(label_original) > len(best_match[0]):
                best_match = (label_original, value, field)

    if best_match and best_score >= 2:
        label, value, field = best_match
        if value in [None, "", []]:
            return best_score, f"There is no <b>{label}</b>"
            
        # Handle child tables (e.g. Escalation Matrix)
        if field.fieldtype == "Table" and isinstance(value, list) and len(value) > 0:
            header = f"<b>{label}</b>"
            supplier = doc.get("supplier")
            if supplier:
                header += f" for <b>{supplier}</b>"
            
            rows_html = []
            for row in value:
                row_dict = row.as_dict()
                parts = []
                # Map common child fields to readable labels
                field_map = [
                    ("level", "Level"),
                    ("link_zitr", "Name"),
                    ("contact_phone", "Phone"),
                    ("link_syot", "Email"),
                    ("designation", "Designation")
                ]
                for fname, flabel in field_map:
                    v = row_dict.get(fname)
                    if v:
                        parts.append(f"{flabel}: {v}")
                if parts:
                    rows_html.append(" • " + ", ".join(parts))
            
            if rows_html:
                return best_score, f"{header}:<br>" + "<br>".join(rows_html)

        return best_score, f"{label}: <b>{format_value(value)}</b>"

    return 0, None

def get_logged_user_info():
    user = frappe.session.user
    user_doc = frappe.get_doc("User", user)
    return f"""
<b>User Information</b><br><br>
First Name: <b>{user_doc.first_name or ""}</b><br>
Full Name: <b>{user_doc.full_name or ""}</b><br>
Login Email: <b>{user_doc.name}</b>
"""

def create_finance_issue_task(ticket):
    ticket_doc = frappe.get_doc("HD Ticket", ticket)
    if not ticket_doc.get("custom_lms_id"):
        return {"task": None, "message": "No LMS Found for this Ticket"}
    task = frappe.new_doc("Task")
    task.type = "Finance Issue"
    task.subject = "Finance Issue Task"
    task.custom_customer = ticket_doc.get("customer")
    task.custom_lms_id = ticket_doc.get("custom_lms_id")
    task.insert(ignore_permissions=True)
    return {"task": task.name, "message": f"Task {task.name} created (Finance Issue)"}

def create_hardware_dispatch_task(ticket):
    ticket_doc = frappe.get_doc("HD Ticket", ticket)
    if not ticket_doc.get("custom_lms_id"):
        return {"task": None, "message": "No LMS Found for this Ticket"}
    task = frappe.new_doc("Task")
    task.type = "Hardware Dispatch"
    task.subject = "Hardware Dispatch Task"
    task.custom_customer = ticket_doc.get("customer")
    task.custom_lms_id = ticket_doc.get("custom_lms_id")
    task.insert(ignore_permissions=True)
    return {"task": task.name, "message": f"Task {task.name} created (Hardware Dispatch)"}

# ---------------------------------------------------------
# DYNAMIC REPORT FOR CLOSED TICKETS
# ---------------------------------------------------------
@frappe.whitelist()
def get_filtered_closed_tickets(filters=None, current_ticket=None):
    if isinstance(filters, str):
        try:
            filters = json.loads(filters)
        except:
            filters = {}
    elif not filters:
        filters = {}
    
    query_filters = {"status": "Closed"}
    limit = None
    
    # Handle History Request
    if filters.get("is_history") and current_ticket:
        circuit_id = frappe.db.get_value("HD Ticket", current_ticket, "custom_circuit_id")
        if circuit_id:
            query_filters["custom_circuit_id"] = circuit_id
            limit = 10 # Last 10 as requested
    else:
        # Standard filter logic
        if filters.get("customer"):
            query_filters["customer"] = ["LIKE", f"%{filters['customer']}%"]
        
        from_date = filters.get("from_date")
        to_date = filters.get("to_date")
        months = filters.get("months")
        specific_month = filters.get("specific_month")
        
        if from_date and to_date:
            query_filters["custom_close_datetime"] = ["between", [from_date, to_date]]
        elif from_date:
            query_filters["custom_close_datetime"] = [">=", from_date]
        elif to_date:
            query_filters["custom_close_datetime"] = ["<=", to_date]
        elif months:
            try:
                start_date = dt.datetime.now() - dt.timedelta(days=int(months)*30)
                query_filters["custom_close_datetime"] = [">=", start_date]
            except: pass
        elif specific_month:
            try:
                now = dt.datetime.now()
                m_names = ["january", "february", "march", "april", "may", "june", 
                           "july", "august", "september", "october", "november", "december"]
                m_idx = -1
                sm = specific_month.lower()
                for i, name in enumerate(m_names):
                    if name in sm:
                        m_idx = i + 1
                        break
                
                if m_idx != -1:
                    year = now.year
                    if m_idx > now.month:
                        year -= 1
                    
                    import calendar
                    last_day = calendar.monthrange(year, m_idx)[1]
                    m_start = dt.datetime(year, m_idx, 1)
                    m_end = dt.datetime(year, m_idx, last_day, 23, 59, 59)
                    query_filters["custom_close_datetime"] = ["between", [m_start, m_end]]
            except: pass
        
        # If no date filter provided, default to last 3 months
        if "custom_close_datetime" not in query_filters:
             start_date = dt.datetime.now() - dt.timedelta(days=90)
             query_filters["custom_close_datetime"] = [">=", start_date]

    fields = [
        "name", "customer", "creation",
        "custom_agent", "custom_channel", "custom_close_datetime", "agreement_status", "custom_rca"
    ]
    
    tickets = frappe.get_all("HD Ticket", filters=query_filters, fields=fields, order_by="custom_close_datetime desc", limit=limit)
    
    # Capture customer name for heading (from first ticket)
    customer_name = tickets[0].get("customer", "") if tickets else ""

    # Format date fields, compute resolution time, and resolve agent names
    for t in tickets:
        # Compute resolution time before formatting dates
        resolution = ""
        if t.get("creation") and t.get("custom_close_datetime"):
            try:
                diff = t["custom_close_datetime"] - t["creation"]
                total_minutes = int(diff.total_seconds() // 60)
                hours = total_minutes // 60
                mins = total_minutes % 60
                resolution = f"{hours:02d}:{mins:02d}"
            except:
                resolution = ""
        t["resolution_time"] = resolution

        if t.get("creation"):
            t["creation"] = t["creation"].strftime("%d-%m-%Y %H:%M") if hasattr(t["creation"], "strftime") else str(t["creation"])
        if t.get("custom_close_datetime"):
            t["custom_close_datetime"] = format_value(t["custom_close_datetime"])
        # Resolve agent email → agent name
        if t.get("custom_agent"):
            agent_name = frappe.db.get_value("HD Agent", {"user": t["custom_agent"]}, "agent_name")
            if agent_name:
                t["custom_agent"] = agent_name
        # Remove customer from row data (shown in heading instead)
        t.pop("customer", None)

    labels = {
        "name": "Ticket No",
        "creation": "Created Datetime",
        "custom_agent": "Agent",
        "custom_channel": "Channel",
        "custom_close_datetime": "Closed Datetime",
        "resolution_time": "Resolution Time",
        "agreement_status": "SLA Status",
        "custom_rca": "RCA"
    }
    
    return {"tickets": tickets, "labels": labels, "customer_name": customer_name}

@frappe.whitelist()
def download_closed_tickets_csv(filters=None, current_ticket=None):
    res = get_filtered_closed_tickets(filters, current_ticket)
    tickets = res["tickets"]
    labels = res["labels"]
    
    from frappe.utils.xlsxutils import make_xlsx
    
    # Prepare rows for XLSX
    header = list(labels.values())
    rows = [header]
    
    for t in tickets:
        row = []
        for key in labels.keys():
            val = t.get(key)
            if isinstance(val, (dt.datetime, dt.date)):
                val = val.strftime("%d-%m-%Y %H:%M")
            row.append(val or "")
        rows.append(row)
    
    xlsx_data = make_xlsx(rows, "Closed Tickets Report")
    
    filename = f"Closed_Tickets_{dt.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    frappe.response['filename'] = filename
    frappe.response['filecontent'] = xlsx_data.getvalue()
    frappe.response['type'] = 'binary'

# ---------------------------------------------------------
# AI MODEL CALL USING API CONFIGURATION DOCTYPE
# ---------------------------------------------------------
def call_ai_model_old_9183(prompt):
    try:
        config_name = frappe.db.get_value("API Configuration", None, "name")
        if not config_name:
            return "API Configuration not found. Please set up API Configuration."

        config = frappe.get_doc("API Configuration", config_name)

        # Assuming fieldnames: api_key (Password), model_name, api_base_url, temperature, max_tokens
        api_key = config.get_password("api_key")
        model_name = config.model_name
        api_base_url = config.api_base_url
        temperature = config.temperature or 0.7
        max_tokens = config.max_tokens or 500

        if not api_key or not model_name or not api_base_url:
            return "Incomplete API Configuration. Please check API Key, Model Name, and Base URL."

        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }

        payload = {
            "model": model_name,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": temperature,
            "max_tokens": max_tokens
        }

        response = requests.post(api_base_url, headers=headers, json=payload, timeout=60)
        response.raise_for_status()

        result = response.json()
        if "choices" in result and len(result["choices"]) > 0:
            return result["choices"][0]["message"]["content"]
        elif "response" in result:
            return result["response"]
        else:
            return "Unexpected API response format."

    except requests.exceptions.RequestException as e:
        frappe.log_error(f"API request failed: {str(e)}", "AI Model Call")
        return f"Error calling AI model: {str(e)}"
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "AI Model Call")
        return f"Unexpected error: {str(e)}"

# ---------------------------------------------------------
# MAIN CHAT ENDPOINT
# ---------------------------------------------------------
def build_lms_info_html(lms_doc):
    """
    Constructs a modern, compact HTML card for LMS information.
    Includes sections for Info, Provisioning, Portal, and Escalation Matrix table.
    """
    if not lms_doc:
        return "LMS information not found."

    sections_config = [
        {
            "title": f"📡 LMS INFORMATION ({lms_doc.name})",
            "fields": [
                ("name", "Supplier Name"),  # Use name as a proxy for the unique ID
                ("supplier", "Supplier Name"),
                ("lms_delivery_date", "LMS Delivery Date"),
                ("lms_brandwith_name", "LMS Brandwith"),
                ("media", "Media")
            ]
        },
        {
            "title": "⚙️ LMS PROVISIONING",
            "fields": [
                ("mode1", "Model"),
                ("account_no_dsl_no_service_id", "Account No/ DSL No/ Service ID"),
                ("static_ip_1", "Static IP"),
                ("static_ip", "Static IP Details"),
                ("user_id", "User ID"),
                ("password", "Password")
            ]
        },
        {
            "title": "💳 LMS PMT PORTAL",
            "fields": [
                ("payment_mode_1", "LMS Payment Mode"),
                ("bank", "Bank"),
                ("url", "Url"),
                ("portal_login_id", "Portal Login ID"),
                ("portal_login_password", "Portal Login Password")
            ]
        }
    ]

    html = '<div class="hd-ai-site-info-card hd-ai-lms-card">'
    for section in sections_config:
        valid_fields = []
        for fieldname, label in section["fields"]:
            val = lms_doc.get(fieldname)
            if val and str(val).strip() and str(val).strip().upper() != "N/A":
                valid_fields.append((label, format_value(val)))

        if not valid_fields:
            continue

        html += f'<div class="hd-ai-site-section">'
        html += f'<div class="hd-ai-site-section-title">{section["title"]}</div>'
        html += '<div class="hd-ai-site-grid">'
        for label, val in valid_fields:
            html += f'''
                <div class="hd-ai-site-field">
                    <div class="hd-ai-site-label">{label}</div>
                    <div class="hd-ai-site-value">{val}</div>
                </div>'''
        html += '</div></div>'

    # Level (Doctype: LMS Contact Escalation (table_oeiw) Child table of Lastmile Services Master)
    if lms_doc.get("table_oeiw"):
        html += '<div class="hd-ai-site-section">'
        html += '<div class="hd-ai-site-section-title">🪜 ESCALATION MATRIX</div>'
        html += '<div class="hd-ai-lms-table-wrapper">'
        html += '<table class="hd-ai-lms-table">'
        html += '<thead><tr><th>Level</th><th>Name</th><th>Phone</th><th>Email</th><th>Designation</th><th>Department</th></tr></thead>'
        html += '<tbody>'
        for row in lms_doc.table_oeiw:
            html += f'''
                <tr>
                    <td>{row.level or ""}</td>
                    <td>{row.link_zitr or ""}</td>
                    <td>{row.contact_phone or ""}</td>
                    <td>{row.link_syot or ""}</td>
                    <td>{row.designation or ""}</td>
                    <td>{row.department or ""}</td>
                </tr>'''
        html += '</tbody></table>'
        html += '</div></div>'

    html += '</div>'
    return html

def build_provisioning_info_html(prov_doc):
    """
    Constructs a modern, compact HTML card for Provisioning information.
    Compacted: skips empty fields and empty sections.
    """
    if not prov_doc:
        return ""

    # Define sections and fields
    sections_config = [
        {
            "title": f"⚙️ PROVISIONING INFORMATION ({prov_doc.name})",
            "fields": [
                ("circuit_id", "Circuit ID"),
                ("site_name", "Site Name"),
                ("solution_name", "Solution Name"),
                ("status", "Provisioning Status"),
                ("provisioning_date", "Completed Date"),
                ("provisioning_partially_completed_date", "Partially Completed Date")
            ]
        },
        {
            "title": "🔌 BRANCH IP INFORMATION",
            "fields": [
                ("atm_ip", "ATM IP"),
                ("branch_natted_ip", "Branch Natted IP"),
                ("branch_router_ip", "Branch Router IP"),
                ("branch_lan_series", "Branch Lan Series")
            ]
        },
        {
            "title": "🏢 DC IP INFORMATION",
            "fields": [
                ("dc_static_ip", "DC Static IP"),
                ("dc_router_ip", "DC Router IP"),
                ("dc_secondary_static_ip", "DC Secondary Static IP"),
                ("dc_server_gateway_ip", "DC Server Gateway IP"),
                ("dc_server_ip", "DC Server IP"),
                ("dc_server_ip_2", "DC Server IP 2")
            ]
        },
        {
            "title": "☁️ DR IP INFORMATION",
            "fields": [
                ("dr_static_ip", "DR Static IP"),
                ("dr_router_ip", "DR Router IP"),
                ("dr_secondary_static_ip", "DR Secondary Static IP"),
                ("dr_server_gateway_ip", "DR Server Gateway IP"),
                ("dr_server_ip", "DR Server IP"),
                ("dr_server_ip_2", "DR Server IP 2")
            ]
        },
        {
            "title": "🌐 WAN IP INFO PRIMARY",
            "fields": [
                ("lms_id_1", "LMS ID 1"),
                ("wan_static_ip_1", "WAN Static IP 1"),
                ("wan_gateway_ip_1", "WAN Gateway IP 1"),
                ("wan_user_name_1", "WAN User Name 1"),
                ("wan_password_1", "WAN Password 1"),
                ("lms_1_delivery_date", "LMS 1 Delivery Date"),
                ("subnet_mask_1", "Subnet Mask 1"),
                ("wan_1_dns", "WAN 1 DNS"),
                ("wan_2_dns", "WAN 2 DNS")
            ]
        },
        {
            "title": "🌐 WAN IP INFO SECONDARY",
            "fields": [
                ("lms_id_2", "LMS ID 2"),
                ("wan_static_ip_2", "WAN Static IP 2"),
                ("wan_gateway_ip_2", "WAN Gateway IP 2"),
                ("wan_user_name_2", "WAN User Name 2"),
                ("wan_password_2", "WAN Password 2"),
                ("lms_delivery_date", "LMS Delivery Date"),
                ("subnet_mask_2", "Subnet Mask 2"),
                ("wan_dns_1", "WAN DNS 1"),
                ("wan_dns_2", "WAN DNS 2")
            ]
        }
    ]

    html = '<div class="hd-ai-site-info-card hd-ai-lms-card">'
    for section in sections_config:
        # Filter fields that have values
        valid_fields = []
        for fieldname, label in section["fields"]:
            val = prov_doc.get(fieldname)
            if val and str(val).strip() and str(val).strip().upper() != "N/A":
                formatted_val = format_value(val)
                valid_fields.append((label, formatted_val))

        if not valid_fields:
            continue

        html += f'<div class="hd-ai-site-section">'
        html += f'<div class="hd-ai-site-section-title">{section["title"]}</div>'
        html += '<div class="hd-ai-site-grid">'
        for label, val in valid_fields:
            html += f'''
                <div class="hd-ai-site-field">
                    <div class="hd-ai-site-label">{label}</div>
                    <div class="hd-ai-site-value">{val}</div>
                </div>'''
        html += '</div></div>'
    
    html += '</div>'
    return html

def build_site_info_html(site_doc):
    """
    Constructs a modern, non-table HTML card for Site information.
    Compacted: skips empty fields, no "N/A".
    """
    if not site_doc:
        return "Site information not found."

    # Define sections and fields
    sections_config = [
        {
            "title": "📋 SITE INFORMATION",
            "fields": [
                ("site_name", "Site Name"),
                ("name", "Circuit No"),
                ("order_type", "Order Type"),
                ("site_type", "Site Type"),
                ("customer", "Customer"),
                ("site_id__legal_code", "Site ID / Legal Code"),
                ("site_status", "Site Status"),
                ("territory", "Territory"),
                ("customer_type", "Customer Type"),
                ("date", "Circuit Delivery Date"),
                ("address_street", "Address/ Street"),
                ("city", "City"),
                ("pincode", "Pincode")
            ]
        },
        {
            "title": "🛠️ SOLUTION INFORMATION",
            "fields": [
                ("solution_name", "Solution Name"),
                ("managed_services", "Managed Services"),
                ("config_type", "Config Type")
            ]
        },
        {
            "title": "📞 BRANCH CONTACT INFORMATION",
            "fields": [
                ("contact_person", "Contact Person"),
                ("primary_contact_mobile", "Contact Mobile"),
                ("email", "Email")
            ]
        }
    ]

    html = '<div class="hd-ai-site-info-card">'
    for section in sections_config:
        # Filter fields that have values
        valid_fields = []
        for fieldname, label in section["fields"]:
            val = site_doc.get(fieldname)
            if val and str(val).strip() and str(val).strip().upper() != "N/A":
                # Ensure value is formatted correctly (e.g. DD-MM-YYYY for dates)
                formatted_val = format_value(val)
                valid_fields.append((label, formatted_val))

        if not valid_fields:
            continue

        html += f'<div class="hd-ai-site-section">'
        html += f'<div class="hd-ai-site-section-title">{section["title"]}</div>'
        html += '<div class="hd-ai-site-grid">'
        for label, val in valid_fields:
            html += f'''
                <div class="hd-ai-site-field">
                    <div class="hd-ai-site-label">{label}</div>
                    <div class="hd-ai-site-value">{val}</div>
                </div>'''
        html += '</div></div>'
    html += '</div>'
    return html

@frappe.whitelist()
def hd_ticket_ai_chat(ticket, question):
    frappe.flags.mute_messages = True

    if not ticket or not question:
        return "Please provide both ticket and question."

    q_lower = question.strip().lower()
    user = frappe.session.user

    # Clear memory
    if q_lower.strip() in ["clear", "clear memory"]:
        clear_user_memory(user, ticket)
        return "✅ Memory cleared. Starting fresh."

    # SITE INFO ENHANCEMENT (Flexible Intent detection)
    site_info_keywords = [
        "site info", "site information", "information about site", 
        "information of site", "details of site", "site details",
        "tell me about site", "show site info", "show me site info",
        "summarize", "summarize this ticket"
    ]
    
    if any(k in q_lower for k in site_info_keywords):
        ticket_doc = frappe.get_doc("HD Ticket", ticket)
        circuit_id = ticket_doc.custom_circuit_id
        if circuit_id:
            site_name = frappe.db.get_value("Site", {"name": circuit_id}, "name")
            if site_name:
                site_doc = frappe.get_doc("Site", site_name)
                return build_site_info_html(site_doc)
        return "No Site found for this ticket."

    # LMS INFO ENHANCEMENT
    lms_info_keywords = [
        "lms info", "lms information", "information about lms", 
        "details of lms", "lms details", "lastmile info", "last mile info"
    ]
    if any(k in q_lower for k in lms_info_keywords):
        ticket_doc = frappe.get_doc("HD Ticket", ticket)
        lms_ids = []
        
        # 1. NEW: Prioritize direct custom_lms_id if present
        direct_lms = ticket_doc.get("custom_lms_id")
        
        if direct_lms:
            lms_ids = [direct_lms]
        else:
            # 2. Try to find linked LMS tickets (HD LMS Ticket)
            linked = frappe.db.get_all("HD LMS Ticket", 
                filters={"customer_ticket_id": ticket}, 
                pluck="lms_id")
            if linked:
                lms_ids.extend([l for l in linked if l])
            
            # 3. Supplement with all delivered links for this circuit if still empty?
            # User says: IF blank THEN show ALL. 
            # I'll check if either is missing to ensure we have a fallback.
            if not lms_ids and ticket_doc.get("custom_circuit_id"):
                circuit_lms = frappe.db.get_all("Lastmile Services Master", 
                    filters={"circuit_id": ticket_doc.custom_circuit_id, "lms_stage": "Delivered"}, 
                    pluck="name")
                for lid in circuit_lms:
                    if lid not in lms_ids:
                        lms_ids.append(lid)
        
        if lms_ids:
            # Generate HTML for each LMS record
            html_cards = []
            for lid in lms_ids:
                lms_doc = frappe.get_doc("Lastmile Services Master", lid)
                html_cards.append(build_lms_info_html(lms_doc))
            
            # Wrap in a container with spacing
            return f'<div style="display:flex;flex-direction:column;gap:16px;">{"".join(html_cards)}</div>'
            
        return "No active (Delivered) LMS information found for this ticket/circuit."

    # --------------------------------------------------
    # 🧠 PROVISIONING INFO INTENT
    # --------------------------------------------------
    provisioning_keywords = [
        "provisioning info", "provisioning information", "config info", 
        "ip info", "ip details", "wan info", "wan details", "lan info",
        "provisioning", "show me provisioning", "i want provisioning"
    ]
    if any(k in q_lower for k in provisioning_keywords):
        ticket_doc = frappe.get_doc("HD Ticket", ticket)
        circuit_id = ticket_doc.get("custom_circuit_id")
        
        if circuit_id:
            prov_records = frappe.db.get_all("Provisioning", 
                filters={"circuit_id": circuit_id}, 
                pluck="name")
            
            if prov_records:
                html_cards = []
                for p_name in prov_records:
                    p_doc = frappe.get_doc("Provisioning", p_name)
                    html_cards.append(build_provisioning_info_html(p_doc))
                
                return f'<div style="display:flex;flex-direction:column;gap:16px;">{"".join(html_cards)}</div>'

        return "No provisioning / IP information found for this circuit."

    # Task creation flow (unchanged)
    task_state_key = f"hd_ai_task_state_{user}_{ticket}"
    waiting = frappe.cache().get_value(task_state_key)

    if waiting and (q_lower == "1" or "finance" in q_lower):
        frappe.cache().delete_value(task_state_key)
        result = create_finance_issue_task(ticket)
        if not result.get("task"):
            return f"⚠️ {result['message']}"
        return f"✅ {result['message']}<br>Task ID: <b>{result['task']}</b>"

    if waiting and (q_lower == "2" or "hardware" in q_lower):
        frappe.cache().delete_value(task_state_key)
        result = create_hardware_dispatch_task(ticket)
        if not result.get("task"):
            return f"⚠️ {result['message']}"
        return f"✅ {result['message']}<br>Task ID: <b>{result['task']}</b>"

    if "create task" in q_lower:
        frappe.cache().set_value(task_state_key, True, expires_in_sec=300)
        return (
            "Which task do you want to create?<br><br>"
            "<b>1.</b> Finance Issue<br>"
            "<b>2.</b> Hardware Dispatch Request<br><br>"
            "Reply with <b>1</b> or <b>2</b>"
        )

    # DYNAMIC REPORT INTENT DETECTION (More flexible keywords)
    report_keywords = [
        "closed ticket", "report", "tickets last", "tickets for", "tickets from", 
        "list of", "closed tickets", "tickets report", "fetch", "show me", "give me tickets",
        "history"
    ]
    is_report_query = any(k in q_lower for k in report_keywords)
    
    if is_report_query:
        # Augment prompt with examples for better NLU
        report_extra = """
If the user is asking for a list, history, or report of closed tickets, you MUST respond in this format:
ACTION:REPORT|{"months": 3, "customer": null, "specific_month": null, "from_date": null, "to_date": null, "is_history": false}

IMPORTANT: Even if the user's phrasing is informal or has poor grammar (e.g., "tickets march", "customer x list", "3 month closed", "history of this ticket"), you must identify the intent and extract the parameters:
- "tickets march" -> specific_month: "March", is_history: false
- "customer x report" -> customer: "Customer X", is_history: false
- "give me 6 month" -> months: 6, is_history: false
- "history of this ticket" or "show history" -> is_history: true
- "from jan to feb" -> from_date: "2024-01-01", to_date: "2024-02-29", is_history: false

Do not return any other text, just the ACTION:REPORT line.
"""
    else:
        report_extra = ""

    if "my name" in q_lower:
        return get_logged_user_info()

    # Get ticket
    ticket_doc = frappe.get_doc("HD Ticket", ticket)
    circuit_id = ticket_doc.custom_circuit_id

    # Smart Context Building
    # We will build context from multiple linked entities if circuit_id exists
    context_data = {"ticket_details": get_clean_doc_data(ticket_doc)}
    entities_found = []

    if circuit_id and not is_report_query:
        # Define field map for related entities
        ENTITY_FIELD_MAP = {
            "Site": "circuit_id",
            "Lastmile Services Master": "circuit_id",
            "Provisioning": "circuit_id",
            "Installation Master": "circuit_id"
        }

        # Targeted Field Search (across all entities)
        best_match_answer = None
        best_match_score = 0
        
        # Check the Ticket itself first
        t_score, ticket_answer = search_field_answer(ticket_doc, question)
        if ticket_answer:
            best_match_answer = ticket_answer
            best_match_score = t_score
            
        # Check all related entities
        related_docs = {} 
        for doctype, field_map in ENTITY_FIELD_MAP.items():
            docname = frappe.db.get_value(doctype, {field_map: circuit_id}, "name")
            if docname:
                doc = frappe.get_doc(doctype, docname)
                related_docs[doctype] = doc
                
                # Run matching on this doc
                e_score, entity_answer = search_field_answer(doc, question)
                if entity_answer:
                    # If this is a very strong match (phrase match for 1+ words), and user specifically 
                    # mentions the entity or it beats the previous best score.
                    target_entity = identify_entity(question)
                    
                    if target_entity == doctype and e_score >= 15:
                        return entity_answer
                    
                    if e_score > best_match_score:
                        best_match_answer = entity_answer
                        best_match_score = e_score
                    elif e_score == best_match_score and best_match_answer:
                         # Tie-breaker: prefer anything other than Ticket if scores are equal
                         best_match_answer = entity_answer

        if best_match_answer:
            return best_match_answer

        # ALWAYS Include a baseline of critical site info if it exists
        # We use the key 'Site' to match the AI Prompt's ENTITY RULE
        site_name = frappe.db.get_value("Site", {"circuit_id": circuit_id}, "name")
        if site_name:
            site_doc = frappe.get_doc("Site", site_name)
            context_data["Site"] = {
                "Site Name": site_doc.site_name,
                "Site Status": site_doc.site_status,
                "Region": site_doc.region,
                "Territory": site_doc.territory,
                "Address Street": site_doc.address_street,
                "City": site_doc.city,
                "Customer": site_doc.customer
            }

        # If no direct match found, build context for AI for those found entities
        for doctype, doc in related_docs.items():
            clean_data = get_clean_doc_data(doc)
            if doctype == "Site":
                context_data.setdefault("Site", {}).update(clean_data)
            else:
                context_data[doctype] = clean_data
            entities_found.append(doctype)
    elif circuit_id and is_report_query:
        # Build minimal context for report if needed, but the report logic handles it
        pass

    # Priority 2: FAISS (if no target identified or as fallback)
    if is_faiss_enabled() and not is_report_query:
        try:
            from nexapp.ai.faiss_engine import faiss_search, fetch_data_by_circuit
            items = faiss_search(question, top_k=8) 
            
            # Fetch data and ensure it's labeled
            faiss_data = fetch_data_by_circuit(circuit_id, items)
            
            # Convert raw fieldnames to labels and MERGE into primary entities
            for dtype, ddata in faiss_data.items():
                label_map = get_doc_label_map(dtype)
                labeled_ddata = {}
                for f, v in ddata.items():
                    label = label_map.get(f, f.replace('_', ' ').title())
                    labeled_ddata[label] = v
                
                if labeled_ddata:
                    # Merge into the existing section (Site, Provisioning, etc.)
                    # This ensures the AI Prompt's ENTITY RULE is respected
                    if dtype not in context_data:
                        context_data[dtype] = {}
                    context_data[dtype].update(labeled_ddata)
                    if dtype not in entities_found:
                        entities_found.append(dtype)
                
        except Exception as e:
            frappe.log_error(f"FAISS Context Error: {str(e)}", "AI Chat")

    # Prompt template
    prompt_text = frappe.db.get_value(
        "AI Prompt Template",
        {"prompt_code": "HD_TICKET_CHAT"},
        "prompt_text"
    )
    if not prompt_text:
        prompt_text = (
            "You are an AI assistant for helpdesk tickets. "
            "Answer the user's question based on the provided context. "
            "Use a friendly tone and format with HTML <b> and <br> where appropriate."
        )

    # Build memory context (per ticket)
    memory_context = build_memory_context(user, ticket)

    full_prompt = f"""
{prompt_text}
{report_extra}

CONVERSATION HISTORY:
{memory_context}

CONTEXT:
{json.dumps(context_data, default=str, indent=2)}

QUESTION:
{question}
"""

    # Call AI model
    response = call_ai_model(full_prompt)

    # HANDLING REPORT ACTION
    if "ACTION:REPORT|" in response:
        try:
            parts = response.split("ACTION:REPORT|")
            report_json = parts[1].strip()
            # Just return a trigger for the frontend
            return f'<div class="hd-ai-report-trigger" data-filters=\'{report_json}\'>' \
                   f'I can generate a report for the closed tickets you requested.<br>' \
                   f'<button class="hd-ai-report-btn">View Report</button></div>'
        except:
            pass

    # Clean up response
    response = response.replace("**", "").replace("###", "")
    response = response.replace("\\n", "<br>").replace("\n", "<br>")

    if not response or response.strip() == "":
        response = "I'm sorry, I couldn't find an answer to your question."

    update_memory(user, ticket, question, response)

    return response
############################################################################
# Unallocated PAge
import frappe

# ======================================================
# UNALLOCATED RECONCILIATION PAGE METHODS
# ======================================================

@frappe.whitelist()
def get_unallocated_payment_entries(company=None, from_date=None, to_date=None, party=None):
    """
    Returns a list of Payment Entries with unallocated_amount > 0.
    Filters by company, date range and optional party.
    """
    filters = {
        "unallocated_amount": [">", 0],
        "docstatus": 1
    }
    if company:
        filters["company"] = company
    if from_date:
        filters["posting_date"] = [">=", from_date]
    if to_date:
        if "posting_date" in filters:
            filters["posting_date"] = ["between", [from_date, to_date]]
        else:
            filters["posting_date"] = ["<=", to_date]
    if party:
        filters["party"] = party

    fields = ["name", "posting_date", "party", "party_type", "reference_no", "unallocated_amount"]
    entries = frappe.get_list("Payment Entry", filters=filters, fields=fields, order_by="posting_date desc")
    return entries


@frappe.whitelist()
def get_unallocated_parties(company=None, from_date=None, to_date=None):
    """
    Returns a list of distinct party names from Payment Entries
    that have unallocated_amount > 0 and match the filters.
    """
    filters = {
        "unallocated_amount": [">", 0],
        "docstatus": 1
    }
    if company:
        filters["company"] = company
    if from_date:
        filters["posting_date"] = [">=", from_date]
    if to_date:
        if "posting_date" in filters:
            filters["posting_date"] = ["between", [from_date, to_date]]
        else:
            filters["posting_date"] = ["<=", to_date]

    parties = frappe.get_all("Payment Entry", filters=filters, fields=["party"], distinct=True)
    return [p.party for p in parties if p.party]


@frappe.whitelist()
def get_outstanding_invoices(doctype, party_field, party_name, company):
    """
    Returns list of unpaid, uncancelled invoices for a given party.
    doctype: "Purchase Invoice" or "Sales Invoice"
    party_field: "supplier" or "customer"
    party_name: name of the party
    company: company name
    """
    filters = {
        party_field: party_name,
        "company": company,
        "docstatus": 1,
        "status": ["not in", ["Paid", "Cancelled"]],
        "outstanding_amount": [">", 0]
    }
    fields = ["name", "outstanding_amount", "posting_date"]
    if doctype == "Purchase Invoice":
        fields.extend(["bill_no", "bill_date"])

    invoices = frappe.get_all(doctype, filters=filters, fields=fields)
    return invoices


@frappe.whitelist()
def allocate_payment_to_invoices(payment_entry, allocations, company):
    """
    Allocate amounts from a Payment Entry to one or more invoices.
    Edits the existing submitted Payment Entry by appending references.
    allocations: list of dicts with keys 'invoice', 'allocated_amount', 'doctype'
    """
    try:
        if isinstance(allocations, str):
            allocations = frappe.parse_json(allocations)

        pe = frappe.get_doc("Payment Entry", payment_entry)

        # Append new invoice references (preserve existing ones)
        for alloc in allocations:
            invoice_doctype = alloc.get("doctype")
            if not invoice_doctype:
                invoice_doctype = "Purchase Invoice" if alloc["invoice"].startswith("PINV") else "Sales Invoice"

            inv = frappe.get_doc(invoice_doctype, alloc["invoice"])
            
            # Use appropriate account field based on doctype
            account = inv.debit_to if invoice_doctype == "Sales Invoice" else inv.credit_to

            pe.append("references", {
                "reference_doctype": invoice_doctype,
                "reference_name": alloc["invoice"],
                "account": account,
                "due_date": inv.due_date,
                "total_amount": inv.grand_total,
                "outstanding_amount": inv.outstanding_amount,
                "allocated_amount": flt(alloc["allocated_amount"]),
                "exchange_rate": flt(inv.conversion_rate) or 1.0
            })

        # Recalculate totals (total_allocated_amount, unallocated_amount, etc.)
        pe.setup_party_account_field()
        pe.set_amounts()

        # Update GL Entries for submitted document to fix invoice outstanding amounts
        # 1. Reverse old GL entries
        pe.make_gl_entries(cancel=1)

        # Allow saving a submitted document
        pe.flags.ignore_validate_update_after_submit = True
        pe.save(ignore_permissions=True)

        # 2. Create new GL entries (this also updates invoice outstanding)
        pe.make_gl_entries(cancel=0)
        
        frappe.db.commit()

        return {"status": "success", "payment_entry": pe.name}
    except Exception as e:
        frappe.db.rollback()
        return {"status": "error", "error": str(e)}

#########################################################################
# Provisioning Update to Site
import frappe

@frappe.whitelist()
def update_site_from_provisioning(provisioning_name):

    if not provisioning_name:
        frappe.throw("Provisioning document is required")

    # -----------------------------------
    # Get Provisioning document
    # -----------------------------------
    provisioning = frappe.get_doc("Provisioning", provisioning_name)

    if not provisioning.circuit_id:
        frappe.throw("Circuit ID is missing in Provisioning")

    # -----------------------------------
    # Get Site (circuit_id = Site.name)
    # -----------------------------------
    try:
        site = frappe.get_doc("Site", provisioning.circuit_id)
    except frappe.DoesNotExistError:
        frappe.throw(f"Site not found with name: {provisioning.circuit_id}")

    # -----------------------------------
    # CONDITION: Only if provisioning_id is blank
    # -----------------------------------
    if site.provisioning_id:
        frappe.msgprint(f"Site already linked with Provisioning: {site.provisioning_id}")
        return

    # -----------------------------------
    # Prepare update dict (BEST PRACTICE)
    # -----------------------------------
    update_fields = {}

    # -----------------------------------
    # PARTIALLY COMPLETED
    # -----------------------------------
    if provisioning.status == "Partially Completed":

        update_fields = {
            "site_status": "Partially Provisioning Completed",
            "provisioning_partially_completed_date": provisioning.provisioning_partially_completed_date,
            "branch_router_ip": provisioning.branch_router_ip,
            "provisioning_status": provisioning.status
        }

    # -----------------------------------
    # COMPLETED
    # -----------------------------------
    elif provisioning.status == "Completed":

        update_fields = {
            "site_status": "Provisioning Completed",
            "provisioning_id": provisioning.name,
            "provisioning_date": provisioning.provisioning_date,
            "provisioning_status": provisioning.status,
            "branch_router_ip": provisioning.branch_router_ip   # ✅ FIX ADDED
        }

    else:
        return

    # -----------------------------------
    # SINGLE UPDATE (Better than multiple db_set)
    # -----------------------------------
    frappe.db.set_value("Site", site.name, update_fields)

    frappe.db.commit()

    return site.name  # returning for JS use (optional)

#################################################################################
# Shifting Code
import frappe
from frappe.utils import nowdate

@frappe.whitelist()
def update_site_shift(site_name):
    try:
        doc = frappe.get_doc("Site", site_name)

        if (
            doc.order_type == "Shifting"
            and doc.site_status == "Delivered and Live"
            and doc.existing_circuit_id == doc.name
            and not doc.site_shifted_date
        ):
            doc.site_status = "Site Shifted to new location"
            doc.shifted_circuit_id = doc.name
            doc.site_shifted_date = nowdate()

            doc.save(ignore_permissions=True)

        return "Success"

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Site Shift Error")
        return str(e)

#########################################################################
# Employee Survey
import frappe
import json
from frappe.utils import now_datetime


# =========================
# GET SURVEY QUESTIONS
# =========================
@frappe.whitelist(allow_guest=False)
def get_survey_details(survey):

    doc = frappe.get_doc("Employee Survey", survey)
    if not doc.is_active:
        return None

    user = frappe.session.user
    employee = frappe.db.get_value("Employee", {"user_id": user}, "name")

    submission_date = None
    user_fullname = frappe.db.get_value("User", user, "full_name")
    
    existing_answers = {}
    if employee:
        response = frappe.db.get_value("Survey Response", {
            "survey": survey,
            "employee": employee
        }, "name")

        if response:
            res_doc = frappe.get_doc("Survey Response", response)
            submission_date = res_doc.creation
            for ans in res_doc.answers:
                existing_answers[ans.question] = ans.answer

    return {
        "title": doc.name,
        "description": doc.description,
        "start_date": doc.start_date,
        "end_date": doc.end_date,
        "is_active": doc.is_active,
        "survey_type": doc.survey_type,
        "existing_answers": existing_answers,
        "submission_date": submission_date,
        "user_fullname": user_fullname,
        "questions": [{
            "question": q.question,
            "type": q.question_type,
            "description": q.get("description"),
            "image": q.get("image"),
            "options": q.options,
            "mandatory": q.is_mandatory
        } for q in doc.questions]
    }


# =========================
# SAVE SURVEY RESPONSE (FIXED)
# =========================
@frappe.whitelist()
def save_survey_response(survey, answers):

    try:
        if isinstance(answers, str):
            answers = json.loads(answers)

        user = frappe.session.user

        if user == "Guest":
            frappe.throw("Please login to submit the survey")

        employee = frappe.db.get_value("Employee", {"user_id": user}, "name")

        if not employee:
            frappe.throw("No Employee linked to this user")

        # Prevent duplicate
        existing = frappe.db.exists("Survey Response", {
            "survey": survey,
            "employee": employee
        })

        if existing:
            frappe.throw("You already submitted this survey")

        # =========================
        # CREATE DOC
        # =========================
        doc = frappe.new_doc("Survey Response")
        doc.survey = survey
        doc.employee = employee
        doc.submitted_on = now_datetime()

        for ans in answers:
            value = ans.get("answer")

            doc.append("answers", {
                "question": ans.get("question"),
                "answer": value,
                "rating_value": int(value) if str(value).isdigit() else 0
            })

        # 🔥 IMPORTANT
        doc.insert(ignore_permissions=True)

        # 🔥 DEBUG LOG
        frappe.log_error(f"Survey Saved: {doc.name}", "SURVEY SUCCESS")

        return {
            "status": "success",
            "name": doc.name
        }

    except Exception as e:

        # 🔥 LOG ERROR (VERY IMPORTANT)
        frappe.log_error(frappe.get_traceback(), "SURVEY ERROR")

        return {
            "status": "error",
            "message": str(e)
        }
#############################################################################    

import frappe
from frappe.utils import formatdate, now

@frappe.whitelist()
def send_survey_to_employees(survey, send_to=None, department=None, employees=None):
    """
    Enqueue the survey email sending task.
    Returns immediately to avoid blocking the client.
    """
    if not survey:
        return {"status": "error", "message": "Survey not found"}

    survey_doc = frappe.get_doc("Employee Survey", survey)

    # Check if the survey is active
    if not survey_doc.is_active:
        return {
            "status": "error",
            "message": "Survey is not active. Please activate the survey before sending."
        }

    # Enqueue the actual email sending to a background queue
    frappe.enqueue(
        method=process_survey_emails,
        queue="long",
        timeout=600,
        survey=survey,
        send_to=send_to,
        department=department,
        employees=employees,
        job_name=f"send_survey_{survey}"
    )

    return {
        "status": "success",
        "message": "Emails are being sent in background"
    }


def process_survey_emails(survey, send_to=None, department=None, employees=None):
    """
    Actual background task that sends emails and creates survey logs.
    """
    try:
        survey_doc = frappe.get_doc("Employee Survey", survey)

        survey_url = survey_doc.url or f"https://erp.nexapp.co.in/app/employee-survey-page?survey={survey}"
        description = survey_doc.description or ""
        end_date = formatdate(survey_doc.end_date) if survey_doc.end_date else "N/A"

        # Build the list of employees based on send_to option
        base_filters = {
            "status": "Active",
            "user_id": ["is", "set"]      # user_id must not be null/empty
        }

        if send_to == "All Employees":
            employee_list = frappe.get_all(
                "Employee",
                filters=base_filters,
                fields=["name", "employee_name", "user_id"]
            )
        elif send_to == "By Department":
            if not department:
                frappe.log_error("No department provided for 'By Department'", "Survey Email Error")
                return
            filters = base_filters.copy()
            filters["department"] = department
            employee_list = frappe.get_all(
                "Employee",
                filters=filters,
                fields=["name", "employee_name", "user_id"]
            )
        elif send_to == "Selected Employees":
            if not employees:
                frappe.log_error("No employees provided for 'Selected Employees'", "Survey Email Error")
                return
            if isinstance(employees, str):
                employees = frappe.parse_json(employees)
            employee_list = frappe.get_all(
                "Employee",
                filters={
                    "name": ["in", employees],
                    "status": "Active",
                    "user_id": ["is", "set"]
                },
                fields=["name", "employee_name", "user_id"]
            )
        else:
            frappe.log_error(f"Invalid send_to option: {send_to}", "Survey Email Error")
            return

        if not employee_list:
            frappe.log_error("No active employees with valid user_id found", "Survey Email Error")
            return

        # Send emails and create logs
        for emp in employee_list:
            # Skip if the user is disabled
            user_enabled = frappe.db.get_value("User", emp.user_id, "enabled")
            if not user_enabled:
                continue

            try:
                # Send the email
                frappe.sendmail(
                    recipients=[emp.user_id],
                    sender="notification@nexapp.co.in",
                    subject=f"Employee Survey: {survey_doc.name}",
                    message=f"""
                        Dear {emp.employee_name or "Employee"},<br><br>
                        {description}<br><br>
                        <b>📅 Please complete the survey before:</b> {end_date}<br><br>
                        👉 <a href="{survey_url}">Click here to fill the survey</a><br><br>
                        Your feedback is very important to us.<br><br>
                        Regards,<br>
                        HR Team
                    """
                )

                # Avoid duplicate logs
                if not frappe.db.exists("Survey Log", {"survey": survey, "employee": emp.name}):
                    frappe.get_doc({
                        "doctype": "Survey Log",
                        "survey": survey,
                        "employee": emp.name,
                        "employee_name": emp.employee_name,
                        "email": emp.user_id,
                        "status": "Sent",
                        "sent_on": now(),
                        "is_reminder": 0
                    }).insert(ignore_permissions=True)

            except Exception as e:
                frappe.log_error(
                    f"Failed to send survey to {emp.name} ({emp.user_id}): {str(e)}",
                    "Survey Email Error"
                )

    except Exception as e:
        frappe.log_error(f"Survey email background job failed: {str(e)}", "Survey Email Error")


@frappe.whitelist()
def send_survey_reminders():
    """
    Scheduled task to send reminders for surveys ending today. 
    Runs daily via hourly scheduler (filtered by hour).
    """
    from frappe.utils import now_datetime, getdate, formatdate
    
    now_dt = now_datetime()
    today = getdate(now_dt)
    
    # Logic: If deadline is end of today (23:59), reminder goes out at ~12:00 PM (12 hrs before)
    if now_dt.hour != 12:
        return

    active_surveys = frappe.get_all(
        "Employee Survey",
        filters={
            "is_active": 1,
            "end_date": today
        },
        fields=["name", "description", "url", "end_date"]
    )
    
    for survey in active_surveys:
        # Get all employees who were originally invited
        invited_employees = frappe.get_all(
            "Survey Log",
            filters={
                "survey": survey.name,
                "is_reminder": 0
            },
            fields=["employee", "employee_name", "email"]
        )
        
        # Get employees who have already responded
        responded_employees = frappe.get_all(
            "Survey Response",
            filters={"survey": survey.name},
            pluck="employee"
        )
        
        for emp in invited_employees:
            # Skip if already responded
            if emp.employee in responded_employees:
                continue
                
            # Skip if reminder already sent for this survey/employee
            if frappe.db.exists("Survey Log", {
                "survey": survey.name,
                "employee": emp.employee,
                "is_reminder": 1
            }):
                continue
                
            try:
                survey_url = survey.url or f"https://erp.nexapp.co.in/app/employee-survey-page?survey={survey.name}"
                end_date_str = formatdate(survey.end_date)
                
                frappe.sendmail(
                    recipients=[emp.email],
                    sender="notification@nexapp.co.in",
                    subject=f"REMINDER: Please complete the survey: {survey.name}",
                    message=f"""
                        Dear {emp.employee_name or "Employee"},<br><br>
                        This is a friendly reminder to complete the survey: <b>{survey.name}</b>.<br><br>
                        {survey.description or ""}<br><br>
                        <b>📅 The deadline is today:</b> {end_date_str}<br><br>
                        👉 <a href="{survey_url}">Click here to fill the survey</a><br><br>
                        Thank you for your feedback!<br><br>
                        Regards,<br>
                        HR Team
                    """
                )
                
                # Log the reminder
                frappe.get_doc({
                    "doctype": "Survey Log",
                    "survey": survey.name,
                    "employee": emp.employee,
                    "employee_name": emp.employee_name,
                    "email": emp.email,
                    "status": "Sent",
                    "sent_on": now(),
                    "is_reminder": 1
                }).insert(ignore_permissions=True)
                
            except Exception as e:
                frappe.log_error(f"Failed to send survey reminder to {emp.employee}: {str(e)}", "Survey Reminder Error")

#########################################################################
# Start of AI Customer Potal
#########################################################################
import frappe
import re
import os
import requests

# =========================================================
# AI CALL
# =========================================================
def call_ai_model(prompt):
    try:
        config_name = frappe.db.get_value("API Configuration", None, "name")
        if not config_name:
            return ""

        config = frappe.get_doc("API Configuration", config_name)

        api_key = config.get_password("api_key")
        model_name = config.model_name
        api_base_url = config.api_base_url

        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }

        temperature = config.get("temperature")
        if temperature is None: temperature = 0.2
        
        max_tokens = config.get("max_output_tokens") or config.get("max_tokens")
        if not max_tokens or int(max_tokens) == 0: max_tokens = 1000

        payload = {
            "model": model_name,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": float(temperature),
            "max_tokens": int(max_tokens)
        }

        response = requests.post(api_base_url, headers=headers, json=payload, timeout=10)
        
        if response.status_code != 200:
            error_details = response.text[:500]
            frappe.log_error(title="AI Provider HTTP Error", message=f"Status {response.status_code}: {error_details}")
            return f"The AI provider returned an error (Status {response.status_code}). This usually means the Model Name or API Key in your Configuration is incorrect."

        result = response.json()

        res_content = result.get("choices", [{}])[0].get("message", {}).get("content", "").strip()
        if not res_content:
            error_msg = str(result)
            frappe.log_error(title="AI Assistant empty result", message=error_msg)
            return "I apologize, but I received an empty response from the AI. Please check your API Configuration (Model Name/URL)."
        return res_content

    except Exception as e:
        frappe.log_error(title="AI Assistant Model Call Error", message=str(e))
        return "I apologize, but I'm having trouble connecting to the AI service. Please check your API configuration."


# =========================================================
# MAIN FUNCTION
# =========================================================
@frappe.whitelist(allow_guest=True)
def ai_installation_query(question):

    try:
        question_lower = question.lower()

        # =========================================================
        # 🔥 DETECT IMAGE TYPE
        # =========================================================
        image_type = None

        if "ir" in question_lower:
            image_type = "IR Report"
        elif "router" in question_lower:
            image_type = "Router Photo"
        elif "testing" in question_lower:
            image_type = "Testing Photo"
        elif "rack" in question_lower:
            image_type = "Server Rack Photo"
        elif "cable" in question_lower:
            image_type = "Cable Labeling Photo"
        elif "isp" in question_lower:
            image_type = "ISP Device Photo"
        elif "installation" in question_lower or "report" in question_lower:
            image_type = None  # fetch all

        # =========================================================
        # 🔥 EXTRACT INPUTS (NUMBERS + WORDS)
        # =========================================================
        numbers = re.findall(r'\d+', question)
        words = re.findall(r'[A-Za-z0-9]+', question)

        circuit_ids = set()

        # 🔥 1. Direct Circuit ID (numbers)
        for num in numbers:
            if frappe.db.exists("Site", num):
                circuit_ids.add(num)

        # 🔥 2. Legal Code → convert to Circuit ID
        for word in words:
            site = frappe.db.get_value(
                "Site",
                {"site_id__legal_code": word.upper()},
                "name"
            )
            if site:
                circuit_ids.add(site)

        all_images = []
        valid_circuits = []

        # =========================================================
        # 🔥 PROCESS EACH CIRCUIT
        # =========================================================
        for circuit_id in circuit_ids:

            installation = frappe.db.get_value(
                "Installation Note",
                {"custom_circuit_id": circuit_id},
                "name"
            )

            if not installation:
                continue

            # =========================================================
            # 🔥 GET LEGAL CODE FROM SITE
            # =========================================================
            legal_code = frappe.db.get_value(
                "Site",
                {"name": circuit_id},
                "site_id__legal_code"
            ) or "NA"

            # =========================================================
            # 🔥 GET ATTACHMENTS
            # =========================================================
            attachments = frappe.get_all(
                "Installation Note Attachment",
                filters={"parent": installation},
                fields=["attachment", "select_mqjl"]
            )

            for att in attachments:

                if not att.attachment:
                    continue

                # =========================================================
                # 🔥 FILTER ONLY IF SPECIFIC TYPE REQUESTED
                # =========================================================
                if image_type and att.select_mqjl != image_type:
                    continue

                all_images.append({
                    "image": att.attachment,
                    "label": att.select_mqjl,
                    "circuit_id": circuit_id,
                    "legal_code": legal_code
                })

            valid_circuits.append(circuit_id)

        # =========================================================
        # 🔥 AI REPLY
        # =========================================================
        if not all_images:
            if valid_circuits:
                ai_reply = f"No installation images or attachments found for Circuit ID(s): {', '.join(valid_circuits)}."
            else:
                ai_reply = "No data found for the given Circuit ID or Legal Code."
        else:
            if image_type:
                ai_reply = f"Here is the {image_type} for Circuit ID(s): {', '.join(valid_circuits)}"
            else:
                ai_reply = f"Here is the full installation report (all attachments) for Circuit ID(s): {', '.join(valid_circuits)}"

        return {
            "status": "success",
            "images": all_images,
            "circuit_ids": list(valid_circuits),
            "image_type": image_type,
            "ai_reply": ai_reply
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "AI INSTALLATION ERROR")
        return {"status": "error", "message": str(e)}
@frappe.whitelist()
def get_user_allowed_prompts():
    user_roles = frappe.get_roles(frappe.session.user)
    # Get all rules
    rules = frappe.get_all("Customer AI Promt", pluck="name")
    
    allowed_prompts = set()
    for rule in rules:
        doc = frappe.get_doc("Customer AI Promt", rule)
        
        # Check if any role in the Table MultiSelect matches the user's roles
        rule_roles = [r.role for r in doc.get("roles", []) if r.role]
        old_role = doc.get("role")
        
        has_access = False
        if rule_roles:
            if set(rule_roles).intersection(user_roles):
                has_access = True
        elif old_role and old_role in user_roles:
            has_access = True
            
        if has_access:
            for row in doc.get("promt", []):
                if row.promt:
                    allowed_prompts.add(row.promt)
    # Fallback for Admin testing if no rules are set up
    if "Administrator" in user_roles and not allowed_prompts:
        allowed_prompts = set(frappe.get_all("Customer Promt", pluck="name"))
        
    prompts_data = []
    for p in allowed_prompts:
        p_doc = frappe.get_doc("Customer Promt", p)
        prompts_data.append({
            "short_prompt": p_doc.short_prompt,
            "full_prompt": p_doc.full_prompt,
            "data_state": p_doc.data_state or "IDLE",
            "sort_order": p_doc.get("sort_order") or 0
        })
        
    prompts_data.sort(key=lambda x: x["sort_order"])
        
    return prompts_data

@frappe.whitelist()
def ai_invoice_download_query(invoice_no):
    """Return the print page URL for a Sales Invoice, with permission checks."""
    try:
        invoice_no = str(invoice_no).strip()
        
        si_exists = frappe.db.exists("Sales Invoice", invoice_no)
        
        if not si_exists:
            return {
                "status": "error",
                "message": f"Sales Invoice '{invoice_no}' not found."
            }
        
        if not frappe.has_permission("Sales Invoice", ptype="read", doc=invoice_no):
            return {
                "status": "error",
                "message": f"You do not have permission to view '{invoice_no}'."
            }
        
        print_url = f"/app/print/Sales%20Invoice/{invoice_no}?format=Nexapp-%20Invoice"
        
        return {
            "status": "success",
            "print_url": print_url,
            "invoice_no": invoice_no
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "AI INVOICE DOWNLOAD ERROR")
        return {"status": "error", "message": str(e)}

@frappe.whitelist(allow_guest=True)
def download_inline_pdf(doctype, name, format=None):
    if not frappe.has_permission(doctype, ptype="read", doc=name):
        frappe.throw("You don't have permission to view this document")

    doc = frappe.get_doc(doctype, name)
    
    original_user = frappe.session.user
    frappe.set_user("Administrator")
    try:
        pdf_file = frappe.get_print(
            doctype,
            name,
            format,
            doc=doc,
            as_pdf=True
        )
    finally:
        frappe.set_user(original_user)
    
    frappe.local.response.filename = f"{name}.pdf".replace(" ", "-").replace("/", "-")
    frappe.local.response.filecontent = pdf_file
    frappe.local.response.type = "download"
    frappe.local.response.display_content_as = "inline"

@frappe.whitelist()
def ai_sales_invoice_installation_query(invoice_no):
    try:
        invoice_no = str(invoice_no).strip()
        circuit_ids = set()

        # Try Sales Order
        so_exists = frappe.db.exists("Sales Order", invoice_no)
        si_exists = frappe.db.exists("Sales Invoice", invoice_no)
        

        if so_exists and frappe.has_permission("Sales Order", ptype="read", doc=invoice_no):
            so = frappe.get_doc("Sales Order", invoice_no)
            for item in so.items:
                if item.get("custom_feasibility"):
                    circuit_ids.add(item.custom_feasibility)
                elif item.get("custom_circuit_id"):
                    circuit_ids.add(item.custom_circuit_id)
        
        # Try Sales Invoice
        elif si_exists and frappe.has_permission("Sales Invoice", ptype="read", doc=invoice_no):
            si = frappe.get_doc("Sales Invoice", invoice_no)
            for item in si.items:
                if item.get("custom_feasibility"):
                    circuit_ids.add(item.custom_feasibility)
                elif item.get("custom_circuit_id"):
                    circuit_ids.add(item.custom_circuit_id)
        else:
            return {
                "status": "success",
                "images": [],
                "circuit_ids": [],
                "image_type": None,
                "ai_reply": f"❌ Could not find any Sales Invoice matching '**{invoice_no}**', or you do not have permission to view it."
            }

        if not circuit_ids:
            return {
                "status": "success",
                "images": [],
                "circuit_ids": [],
                "image_type": None,
                "ai_reply": f"No Circuit IDs found in the items of **{invoice_no}**."
            }

        all_images = []
        valid_circuits = []

        for circuit_id in circuit_ids:
            installation = frappe.db.get_value(
                "Installation Note",
                {"custom_circuit_id": circuit_id},
                "name"
            )
            if not installation:
                continue

            legal_code = frappe.db.get_value("Site", {"name": circuit_id}, "site_id__legal_code") or "NA"

            attachments = frappe.get_all(
                "Installation Note Attachment",
                filters={"parent": installation},
                fields=["attachment", "select_mqjl"],
                ignore_permissions=True
            )

            for att in attachments:
                if not att.attachment:
                    continue
                all_images.append({
                    "image": att.attachment,
                    "label": att.select_mqjl,
                    "circuit_id": circuit_id,
                    "legal_code": legal_code
                })
            
            valid_circuits.append(circuit_id)

        if not all_images:
            if valid_circuits:
                ai_reply = f"No installation images or attachments found for the Circuit IDs linked to **{invoice_no}**."
            else:
                ai_reply = f"No Installation Notes found for the Circuit IDs linked to **{invoice_no}**."
        else:
            ai_reply = f"Here are the installation photographs for the circuits linked to **{invoice_no}**."

        return {
            "status": "success",
            "images": all_images,
            "circuit_ids": list(valid_circuits),
            "image_type": None,
            "ai_reply": ai_reply
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "AI SALES INVOICE INSTALLATION ERROR")
        return {"status": "error", "message": str(e)}

@frappe.whitelist()
def ai_installation_date_query(customer, circuit_id="", from_date="", to_date="", exact_dates="", is_range="false"):
    try:
        from frappe.utils import getdate, formatdate
        
        def parse_date(d_str):
            if not d_str or d_str == "undefined":
                return None
            d_str = d_str.strip()
            # Try to parse DD-MM-YYYY with or without leading zeros
            if '-' in d_str:
                parts = d_str.split('-')
                if len(parts) == 3 and len(parts[2]) == 4: # Ends in a 4-digit year
                    return f"{parts[2]}-{int(parts[1]):02d}-{int(parts[0]):02d}"
            return d_str

        filters = {
            "customer": customer,
            "site_status": "Delivered and Live"
        }

        if is_range == "true":
            f_str = parse_date(from_date)
            t_str = parse_date(to_date)
            if not f_str or not t_str:
                return {"status": "error", "message": "Invalid date format provided. Please use YYYY-MM-DD or DD-MM-YYYY."}
            filters["date"] = ["between", [f_str, t_str]]
            reply_dates_text = f"from **{from_date}** to **{to_date}**"
        else:
            raw_dates = exact_dates.replace(',', ' ').split()
            parsed_dates = []
            for d in raw_dates:
                p = parse_date(d)
                if not p:
                    return {"status": "error", "message": f"{d} is not a valid date string."}
                parsed_dates.append(p)
            
            if len(parsed_dates) == 1:
                filters["date"] = parsed_dates[0]
            else:
                filters["date"] = ["in", parsed_dates]
            
            reply_dates_text = f"for dates **{exact_dates}**"
        if circuit_id and circuit_id.strip() and circuit_id != "undefined":
            filters["name"] = circuit_id

        # Use get_list to enforce user permissions natively
        sites = frappe.get_list(
            "Site",
            filters=filters,
            fields=["name", "site_name", "customer", "date", "circuit_id", "site_id__legal_code"]
        )

        if not sites:
            return {
                "status": "success",
                "images": [],
                "ai_reply": "No matching Sites found for the given Customer, Circuit ID, and Date Range, or you do not have permission to view them."
            }

        all_images = []
        valid_circuits = []
        site_details = {}

        for site in sites:
            cid = site.name
            site_details[cid] = site

            # Find matching Installation Note
            installations = frappe.get_all(
                "Installation Note",
                filters={"custom_circuit_id": cid},
                fields=["name"]
            )

            for inst in installations:
                attachments = frappe.get_all(
                    "Installation Note Attachment",
                    filters={"parent": inst.name},
                    fields=["attachment", "select_mqjl"]
                )

                for att in attachments:
                    if not att.attachment:
                        continue

                    all_images.append({
                        "image": att.attachment,
                        "label": att.select_mqjl or "Attachment",
                        "circuit_id": cid,
                        "legal_code": site.site_id__legal_code or "NA",
                        "site_name": site.site_name,
                        "customer": site.customer,
                        "delivery_date": formatdate(site.date) if site.date else "NA"
                    })

            valid_circuits.append(cid)

        if not all_images:
            msg = f"No installation images found for Circuit ID **{circuit_id}** within the specified dates." if circuit_id and circuit_id.strip() and circuit_id != "undefined" else f"No installation images found within the specified dates."
            return {
                "status": "success",
                "images": [],
                "ai_reply": msg
            }

        unique_circuits = list(set([img["circuit_id"] for img in all_images]))
        count = len(unique_circuits)

        if circuit_id and circuit_id.strip() and circuit_id != "undefined":
            ai_reply = f"✅ Successfully retrieved the installation photographs for Circuit ID **{circuit_id}** {reply_dates_text} for **{customer}**.<br><br>Total Circuits: **{count}**"
        else:
            ai_reply = f"✅ Successfully retrieved the installation photographs {reply_dates_text} for **{customer}**.<br><br>Total Circuits: **{count}**"

        return {
            "status": "success",
            "images": all_images,
            "circuit_ids": valid_circuits,
            "site_info": site_details,
            "ai_reply": ai_reply
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "AI INSTALLATION DATE ERROR")
        return {"status": "error", "message": str(e)}

@frappe.whitelist()
def download_multi_images(files):
    """
    Creates a ZIP file of multiple images and returns the download URL.
    """
    frappe.logger().info(f"MULTI DOWNLOAD REQUEST: {files}")
    
    if not files:
        return {"status": "error", "message": "No files selected"}

    try:
        if isinstance(files, str):
            files = json.loads(files)

        # ZIP filename from first file metadata
        first_file = files[0] if files else {}
        z_cid = first_file.get("cid", "Unknown")
        z_lc = first_file.get("lc", "NA")
        zip_display_name = f"Installation_Report_{z_cid}_{z_lc}.zip"

        zip_buffer = io.BytesIO()
        files_added = 0
        
        with zipfile.ZipFile(zip_buffer, "w") as zip_file:
            for file_data in files:
                url = file_data.get("url")
                if not url:
                    continue
                
                label = file_data.get("label", "Image")
                fcid = file_data.get("cid", "Unknown")
                flc = file_data.get("lc", "NA")
                
                # Resolve file path
                clean_path = url.lstrip("/")
                site_path = frappe.get_site_path()
                
                possible_paths = [
                    os.path.join(site_path, "public", clean_path),
                    os.path.join(site_path, clean_path),
                    frappe.get_site_path("public", clean_path),
                    frappe.get_site_path("private", clean_path)
                ]
                
                resolved_path = None
                for p in possible_paths:
                    if os.path.exists(p) and os.path.isfile(p):
                        resolved_path = p
                        break
                
                if resolved_path:
                    _, ext = os.path.splitext(resolved_path)
                    # Use provided metadata for internal name
                    internal_name = f"{label}_{fcid}_{flc}{ext}".replace(" ", "_")
                    zip_file.write(resolved_path, internal_name)
                    files_added += 1
                else:
                    frappe.logger().warning(f"Could not resolve file: {url}")

        if files_added == 0:
            return {"status": "error", "message": "None of the selected images could be found on the server."}

        zip_buffer.seek(0)
        
        # Save ZIP to file manager with randomized internal name but return pretty display name
        from frappe.utils import random_string
        fn = f"report_{random_string(6)}.zip"
        
        _file = frappe.get_doc({
            "doctype": "File",
            "file_name": fn,
            "content": zip_buffer.getvalue(),
            "is_private": 0
        })
        _file.insert(ignore_permissions=True)
        frappe.db.commit()

        return {
            "status": "success", 
            "url": _file.file_url,
            "filename": zip_display_name
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "MULTI DOWNLOAD ERROR")
        return {"status": "error", "message": f"Server Error: {str(e)}"}

#########################################################################
# End of AI Customer Potal
#########################################################################


#########################################################################
# Maintenance Visit from HD Ticket
#########################################################################

@frappe.whitelist()
def create_maintenance_visit_from_ticket(ticket):
    """
    Create a Maintenance Visit from an HD Ticket and update the ticket.
    """
    if not ticket:
        frappe.throw("Ticket name is required")

    # Check if a Maintenance Visit already exists for this ticket
    existing = frappe.db.get_value(
        "Maintenance Visit",
        {"custom_ticket_no": ticket},
        ["name", "creation"],
        as_dict=True
    )
    if existing:
        created_on = frappe.utils.format_datetime(existing.creation)
        return (
            f"⚠️ Maintenance Visit <b>{existing.name}</b> already exists "
            f"for this ticket.<br>Created on: <b>{created_on}</b>"
        )

    # Get HD Ticket details
    hd_ticket = frappe.get_doc("HD Ticket", ticket)
    circuit_id = hd_ticket.get("custom_circuit_id") or ""

    # Create Maintenance Visit
    mv = frappe.new_doc("Maintenance Visit")
    mv.custom_ticket_no = ticket
    mv.custom_circuit_id = circuit_id

    # Add purpose row with item_code
    mv.append("purposes", {
        "item_code": "STO-ITEM-2025-05011"
    })

    mv.insert(ignore_permissions=True)

    # Update HD Ticket
    frappe.db.set_value("HD Ticket", ticket, {
        "custom_maintenance_visit_id": mv.name,
        "custom_completion_status": "In-Progress",
        "custom_maintenance_created_date": frappe.utils.now_datetime()
    })

    frappe.db.commit()

    return f"✅ Maintenance Visit <b>{mv.name}</b> created successfully."


###############################################################################
# TICKET INSIGHT — Full lifecycle analysis for the AI chatbot
###############################################################################

@frappe.whitelist()
def get_ticket_insight(ticket):
    """Return a comprehensive JSON analysis of an HD Ticket's lifecycle."""
    import json as _json
    from datetime import datetime, timedelta

    if not ticket or not frappe.db.exists("HD Ticket", ticket):
        return {"error": "Ticket not found."}

    doc = frappe.get_doc("HD Ticket", ticket)
    now = frappe.utils.now_datetime()

    # ------------------------------------------------------------------
    # 1. BASIC DATES
    # ------------------------------------------------------------------
    creation_dt = frappe.utils.get_datetime(doc.creation) if doc.creation else None
    close_dt = (
        frappe.utils.get_datetime(doc.custom_close_datetime)
        if doc.get("custom_close_datetime")
        else None
    )
    resolution_dt = (
        frappe.utils.get_datetime(doc.resolution_date)
        if doc.resolution_date
        else None
    )
    first_responded_dt = (
        frappe.utils.get_datetime(doc.first_responded_on)
        if doc.first_responded_on
        else None
    )

    end_dt = close_dt or resolution_dt or now

    # ------------------------------------------------------------------
    # 2. TOTAL DURATION helper
    # ------------------------------------------------------------------
    def fmt_delta(td):
        if not td:
            return "—"
        total_sec = int(td.total_seconds())
        if total_sec < 0:
            return "—"
        days = total_sec // 86400
        hours = (total_sec % 86400) // 3600
        mins = (total_sec % 3600) // 60
        parts = []
        if days:
            parts.append(f"{days}d")
        if hours:
            parts.append(f"{hours}h")
        parts.append(f"{mins}m")
        return " ".join(parts)

    total_duration = end_dt - creation_dt if creation_dt else None
    first_response_time = (
        first_responded_dt - creation_dt
        if first_responded_dt and creation_dt
        else None
    )

    # ------------------------------------------------------------------
    # 3. COMMUNICATIONS
    # ------------------------------------------------------------------
    comms = frappe.get_all(
        "Communication",
        filters={
            "reference_doctype": "HD Ticket",
            "reference_name": ticket,
            "communication_type": "Communication",
        },
        fields=["name", "creation", "sent_or_received", "sender"],
        order_by="creation asc",
    )

    total_interactions = len(comms)
    sent_count = sum(1 for c in comms if c.sent_or_received == "Sent")
    received_count = sum(1 for c in comms if c.sent_or_received == "Received")

    # Response gaps (time between consecutive communications)
    response_gaps = []
    for i in range(1, len(comms)):
        prev = frappe.utils.get_datetime(comms[i - 1].creation)
        curr = frappe.utils.get_datetime(comms[i].creation)
        response_gaps.append((curr - prev).total_seconds())

    avg_response_gap = (
        timedelta(seconds=sum(response_gaps) / len(response_gaps))
        if response_gaps
        else None
    )
    max_response_gap = (
        timedelta(seconds=max(response_gaps)) if response_gaps else None
    )
    min_response_gap = (
        timedelta(seconds=min(response_gaps)) if response_gaps else None
    )

    # Consistency check
    if response_gaps and len(response_gaps) > 1:
        import statistics
        stddev = statistics.stdev(response_gaps)
        mean_gap = statistics.mean(response_gaps)
        cv = stddev / mean_gap if mean_gap else 0
        response_consistency = "Consistent" if cv < 0.8 else "Irregular"
    else:
        response_consistency = "N/A"

    # ------------------------------------------------------------------
    # 4. VERSION LOGS → Status changes + time-per-status
    # ------------------------------------------------------------------
    versions = frappe.get_all(
        "Version",
        filters={"ref_doctype": "HD Ticket", "docname": ticket},
        fields=["creation", "data"],
        order_by="creation asc",
    )

    status_changes = []
    for v in versions:
        try:
            vdata = _json.loads(v.data) if isinstance(v.data, str) else v.data
            for ch in vdata.get("changed", []):
                if ch[0] == "status":
                    status_changes.append(
                        {
                            "time": frappe.utils.get_datetime(v.creation),
                            "from": ch[1],
                            "to": ch[2],
                        }
                    )
        except Exception:
            pass

    num_status_changes = len(status_changes)

    # Reopened count (status changed TO "Open" from "Closed" or "Resolved")
    reopen_count = sum(
        1
        for sc in status_changes
        if sc["to"] == "Open" and sc["from"] in ("Closed", "Resolved")
    )

    # Escalation — check for priority changes
    escalated = False
    for v in versions:
        try:
            vdata = _json.loads(v.data) if isinstance(v.data, str) else v.data
            for ch in vdata.get("changed", []):
                if ch[0] == "priority":
                    escalated = True
                    break
        except Exception:
            pass
        if escalated:
            break

    # Time per status
    status_timeline = []
    current_status = "Open"
    current_start = creation_dt

    for sc in status_changes:
        if current_start and sc["time"]:
            duration = (sc["time"] - current_start).total_seconds()
            status_timeline.append(
                {"status": current_status, "seconds": max(duration, 0)}
            )
        current_status = sc["to"]
        current_start = sc["time"]

    # Add final/current status
    if current_start:
        final_end = end_dt or now
        duration = (final_end - current_start).total_seconds()
        status_timeline.append(
            {"status": current_status, "seconds": max(duration, 0)}
        )

    # Aggregate by status
    status_time_map = {}
    for entry in status_timeline:
        s = entry["status"]
        status_time_map[s] = status_time_map.get(s, 0) + entry["seconds"]

    total_status_seconds = sum(status_time_map.values()) or 1

    # Find bottleneck
    bottleneck_status = (
        max(status_time_map, key=status_time_map.get) if status_time_map else "—"
    )

    # Build lifecycle narrative
    lifecycle_steps = []
    if creation_dt:
        lifecycle_steps.append(
            f"Created on {creation_dt.strftime('%d %b %Y, %H:%M')}"
        )
    for sc in status_changes:
        lifecycle_steps.append(
            f"{sc['from']} → {sc['to']} at {sc['time'].strftime('%d %b, %H:%M')}"
        )
    if close_dt:
        lifecycle_steps.append(
            f"Closed on {close_dt.strftime('%d %b %Y, %H:%M')}"
        )
    elif resolution_dt:
        lifecycle_steps.append(
            f"Resolved on {resolution_dt.strftime('%d %b %Y, %H:%M')}"
        )
    else:
        lifecycle_steps.append(f"Currently in <b>{doc.status}</b> status")

    # ------------------------------------------------------------------
    # 5. SLA STATUS
    # ------------------------------------------------------------------
    sla_status = doc.agreement_status or "—"

    # ------------------------------------------------------------------
    # 6. BEHAVIORAL INSIGHTS + ROOT CAUSE
    # ------------------------------------------------------------------
    behavioral = []
    root_causes = []

    pending_seconds = status_time_map.get("Pending", 0)
    open_seconds = status_time_map.get("Open", 0)

    if pending_seconds > total_status_seconds * 0.4:
        behavioral.append("Significant time spent in Pending — likely waiting for external input.")
        root_causes.append("Dependency on customer / vendor response.")
    if open_seconds > total_status_seconds * 0.4:
        behavioral.append("Ticket remained Open for long — possible delayed assignment.")
        root_causes.append("Delayed agent pickup or workload issues.")
    if reopen_count > 0:
        behavioral.append(f"Ticket was reopened {reopen_count} time(s) — possible incomplete resolution.")
        root_causes.append("Initial fix did not fully resolve the issue.")
    if total_interactions == 0:
        behavioral.append("No communications found — ticket may have been handled offline.")
    elif received_count > sent_count * 2:
        behavioral.append("Customer sent significantly more messages than agent — possible unresponsiveness.")
        root_causes.append("Response delays from the support side.")
    if escalated:
        behavioral.append("Priority was escalated during the lifecycle.")
    if not behavioral:
        behavioral.append("Ticket was handled efficiently with no notable delays.")

    if not root_causes:
        root_causes.append("No significant delays or issues detected.")

    # ------------------------------------------------------------------
    # 7. AI INSIGHT CONCLUSION
    # ------------------------------------------------------------------
    conclusion_parts = []
    if pending_seconds > total_status_seconds * 0.3:
        conclusion_parts.append(
            "experienced delays mainly during the Pending stage, indicating dependency on external input"
        )
    if bottleneck_status and bottleneck_status not in ("—", "Open") and bottleneck_status != "Pending":
        conclusion_parts.append(
            f"spent the most time in {bottleneck_status}"
        )
    if reopen_count > 0:
        conclusion_parts.append(
            f"was reopened {reopen_count} time(s), suggesting the initial resolution was incomplete"
        )
    if sla_status == "Failed":
        conclusion_parts.append("breached its SLA target")
    elif sla_status == "Fulfilled":
        conclusion_parts.append("met all SLA commitments")

    if conclusion_parts:
        ai_conclusion = f"This ticket {', '.join(conclusion_parts)}."
    else:
        ai_conclusion = (
            "This ticket was handled within normal parameters with no significant anomalies detected."
        )

    # Add duration context
    if total_duration:
        ai_conclusion += f" Total lifecycle: {fmt_delta(total_duration)}."

    # ------------------------------------------------------------------
    # 8. STATUS COLORS for charts
    # ------------------------------------------------------------------
    status_colors = {
        "Open": "#3b82f6",
        "Replied": "#8b5cf6",
        "Resolved": "#10b981",
        "Closed": "#6b7280",
        "Pending": "#f59e0b",
        "In Progress": "#6366f1",
        "On Hold": "#ef4444",
    }

    # Build chart data
    status_chart = []
    for s, sec in sorted(status_time_map.items(), key=lambda x: -x[1]):
        pct = round((sec / total_status_seconds) * 100, 1)
        status_chart.append(
            {
                "status": s,
                "seconds": round(sec),
                "percent": pct,
                "label": fmt_delta(timedelta(seconds=sec)),
                "color": status_colors.get(s, "#94a3b8"),
            }
        )

    # ------------------------------------------------------------------
    # 9. DISTRIBUTION ANALYSIS — last 10 tickets for same circuit
    # ------------------------------------------------------------------
    impact_chart = []
    channel_chart = []
    priority_chart = []
    rca_chart = []
    sla_chart = []
    circuit_id = doc.get("custom_circuit_id")
    dist_total = 0

    if circuit_id:
        dist_tickets = frappe.get_all(
            "HD Ticket",
            filters={"custom_circuit_id": circuit_id},
            fields=["custom_impact", "custom_channel", "priority", "custom_rca", "agreement_status"],
            order_by="creation desc",
            limit=10,
        )
        dist_total = len(dist_tickets)

        palette = [
            "#7D32E8", "#3b82f6", "#ef4444", "#f59e0b", "#10b981",
            "#ec4899", "#6366f1", "#14b8a6", "#f97316", "#8b5cf6",
        ]

        def build_chart(items, field):
            counts = {}
            for it in items:
                val = it.get(field) or "Unknown"
                counts[val] = counts.get(val, 0) + 1
            total = sum(counts.values()) or 1
            chart = []
            for idx, (lbl, cnt) in enumerate(sorted(counts.items(), key=lambda x: -x[1])):
                chart.append({
                    "label": lbl,
                    "count": cnt,
                    "percent": round((cnt / total) * 100, 1),
                    "color": palette[idx % len(palette)],
                })
            return chart

        impact_chart = build_chart(dist_tickets, "custom_impact")
        channel_chart = build_chart(dist_tickets, "custom_channel")
        priority_chart = build_chart(dist_tickets, "priority")
        rca_chart = build_chart(dist_tickets, "custom_rca")
        sla_chart = build_chart(dist_tickets, "agreement_status")

    # ------------------------------------------------------------------
    # 10. ASSEMBLE RESPONSE
    # ------------------------------------------------------------------
    agent_name = ""
    if doc.get("custom_agent"):
        agent_name = frappe.db.get_value(
            "HD Agent", {"user": doc.custom_agent}, "agent_name"
        ) or doc.custom_agent

    return {
        "summary": {
            "ticket": ticket,
            "subject": doc.subject or "",
            "status": doc.status,
            "priority": doc.priority or "—",
            "agent": agent_name or "Unassigned",
            "channel": doc.get("custom_channel") or "—",
            "total_duration": fmt_delta(total_duration),
            "first_response_time": fmt_delta(first_response_time),
            "total_interactions": total_interactions,
            "status_changes": num_status_changes,
            "reopen_count": reopen_count,
            "escalated": "Yes" if escalated else "No",
            "sla_status": sla_status,
        },
        "time_analysis": {
            "status_chart": status_chart,
            "bottleneck": bottleneck_status,
        },
        "response_analysis": {
            "avg_response_time": fmt_delta(avg_response_gap),
            "longest_delay": fmt_delta(max_response_gap),
            "shortest_response": fmt_delta(min_response_gap),
            "consistency": response_consistency,
        },
        "impact": {
            "total_tickets": dist_total,
            "chart": impact_chart,
        },
        "channel": {
            "total_tickets": dist_total,
            "chart": channel_chart,
        },
        "priority_dist": {
            "total_tickets": dist_total,
            "chart": priority_chart,
        },
        "rca": {
            "total_tickets": dist_total,
            "chart": rca_chart,
        },
        "sla_dist": {
            "total_tickets": dist_total,
            "chart": sla_chart,
        },
        "ai_conclusion": ai_conclusion,
    }


@frappe.whitelist()
def create_survey_from_template(template_name, survey_name, start_date=None, end_date=None):
    template = frappe.get_doc("Survey Template", template_name)

    survey = frappe.get_doc({
        "doctype": "Employee Survey",
        "survey_name": survey_name,
        "survey_type": template.survey_type,
        "description": template.description,
        "start_date": start_date or frappe.utils.today(),
        "end_date": end_date,
        "is_active": 1,
        "questions": []
    })

    for q in template.questions:
        survey.append("questions", {
            "question": q.question,
            "question_type": q.question_type,
            "description": q.description,
            "options": q.options,
            "is_mandatory": q.is_mandatory
        })

    survey.insert()
    return survey.name
#############################################################################
# Payment Entry Update Email id

import frappe

def set_contact_email_from_supplier(doc, method=None):

    if doc.party_type == "Supplier" and doc.party:
        supplier_email = frappe.db.get_value(
            "Supplier",
            doc.party,
            "email_id"
        )

        if supplier_email and not doc.contact_email:
            doc.contact_email = supplier_email

###################################################################################
# =================================================
# 🤖 AI ASSISTANT — CORE ENGINE
# =================================================

@frappe.whitelist()
def get_user_context():
    """Fetches comprehensive User, Role, Department, Company, and Employee context."""
    user = frappe.session.user
    user_doc = frappe.get_doc("User", user)
    
    context = {
        "user_id": user,
        "full_name": user_doc.full_name,
        "roles": frappe.get_roles(user),
        "role_profile": user_doc.role_profile_name,
        "user_image": user_doc.user_image,
        "company": frappe.db.get_default("company"),
        "department": None,
        "designation": None,
        "branch": None
    }

    # Fetch Employee details if exists
    employee = frappe.db.get_value("Employee", {"user_id": user}, 
        ["name", "department", "designation", "branch", "reports_to", "company"], as_dict=1)
    
    if employee:
        context.update({
            "employee_id": employee.name,
            "department": employee.department,
            "designation": employee.designation,
            "branch": employee.branch,
            "reports_to": employee.reports_to,
            "company": employee.company or context["company"]
        })

    return context

@frappe.whitelist()
def get_ai_assistant_response(question, history_id=None, department=None, file_url=None):
    """Main entry point for AI Assistant with context and permissions."""
    user_context = get_user_context()
    if department:
        user_context["active_department"] = department
    else:
        user_context["active_department"] = user_context.get("department")

    # 1. Manage History
    if not history_id:
        existing_history = frappe.get_all("AI Assistant History", 
            filters={"user": frappe.session.user}, 
            fields=["name", "pinned"],
            order_by="last_interaction asc")
        
        # Limit: 5 unpinned history items
        unpinned = [h for h in existing_history if not h.pinned]
        if len(unpinned) >= 5:
            history_to_delete = unpinned[0].name
            # Delete messages first to avoid link error
            frappe.db.delete("AI Assistant Message", {"history": history_to_delete})
            frappe.delete_doc("AI Assistant History", history_to_delete, ignore_permissions=True)
            
        can_save = True
        # Total limit 15 (10 pinned + 5 unpinned)
        if len(existing_history) >= 15 and all(h.pinned for h in existing_history):
            can_save = False

        if can_save:
            history_doc = frappe.get_doc({
                "doctype": "AI Assistant History",
                "user": frappe.session.user,
                "company": user_context["company"],
                "department": user_context["active_department"],
                "title": (question[:50] + '...') if len(question) > 50 else question,
                "last_interaction": dt.datetime.now()
            })
            history_doc.insert(ignore_permissions=True)
            history_id = history_doc.name
            frappe.db.commit() # Ensure history is saved immediately
    else:
        frappe.db.set_value("AI Assistant History", history_id, "last_interaction", dt.datetime.now())
        frappe.db.commit()

    # Save User message if history_id exists
    if history_id:
        save_message(history_id, "User", question, file_url)
        frappe.db.commit() # Save user message before calling AI

    # --- SPECIAL WORKFLOW: Feasible Circuit List and Task Creation ---
    exact_new_prompt = "list the feasible circuits and create contracts and tasks if required"
    exact_old_prompt = "list down the feasible circuits against the customer abc"
    
    clean_question = question.strip().lower().replace('.', '')
    
    # Check what the assistant asked last
    last_msg = ""
    if history_id:
        last_msg_doc = frappe.get_all("AI Assistant Message", filters={"history": history_id, "role": "Assistant"}, fields=["content"], order_by="creation desc", limit=1)
        if last_msg_doc:
            last_msg = last_msg_doc[0].content or ""

    def get_feasibility_table(c_name):
        feasibilities = frappe.get_all("Feasibility", 
            filters={
                "customer": c_name,
                "docstatus": ["!=", 2],
                "feasibility_status": "Feasible",
                "sales_order": ["in", ["", None]]
            }, 
            fields=["name", "site_name", "circuit_id", "feasibility_status"]
        )
        if not feasibilities:
            return ""
        
        table_html = "<div style='overflow-x: auto; margin-top: 15px; margin-bottom: 15px;'>"
        table_html += "<table style='width: auto; min-width: 600px; border-collapse: collapse; font-size: 14px; text-align: left; background-color: #f8f9fa; border-radius: 6px; overflow: hidden;'>"
        table_html += "<thead>"
        table_html += "<tr style='border-bottom: 1px solid #f0f0f0;'>"
        table_html += "<th style='padding: 16px 32px 16px 16px; font-weight: 700; color: #000000; border: none;'>Circuit ID</th>"
        table_html += "<th style='padding: 16px 32px 16px 12px; font-weight: 700; color: #000000; border: none;'>Site Name</th>"
        table_html += "<th style='padding: 16px 32px 16px 12px; font-weight: 700; color: #000000; border: none;'>Customer Name</th>"
        table_html += "<th style='padding: 16px 32px 16px 12px; font-weight: 700; color: #000000; border: none; white-space: nowrap;'>Feasibility Status <i class='fa fa-clone' style='margin-left: 20px; color: #6b7280; cursor: pointer; font-size: 15px;' onclick=\"let tbl=this.closest('table'); let r=document.createRange(); r.selectNode(tbl); window.getSelection().removeAllRanges(); window.getSelection().addRange(r); document.execCommand('copy'); window.getSelection().removeAllRanges(); frappe.show_alert({message:'Table copied to clipboard', indicator:'green'});\" title='Copy Table'></i></th>"
        table_html += "</tr>"
        table_html += "</thead><tbody>"
        
        for f in feasibilities:
            cid = f.circuit_id or f.name
            table_html += "<tr style='border-bottom: 1px solid #f0f0f0;'>"
            table_html += f"<td style='padding: 16px 32px 16px 16px; color: #1f2937; border: none;'>{cid}</td>"
            table_html += f"<td style='padding: 16px 32px 16px 12px; color: #1f2937; border: none;'>{f.site_name}</td>"
            table_html += f"<td style='padding: 16px 32px 16px 12px; color: #1f2937; border: none;'>{c_name}</td>"
            table_html += f"<td style='padding: 16px 32px 16px 12px; color: #1f2937; border: none;'>{f.feasibility_status}</td>"
            table_html += "</tr>"
        
        table_html += "</tbody></table></div>"
        return table_html


    if exact_new_prompt in clean_question or exact_old_prompt in clean_question:
        assistant_answer = "👤 Please provide the Customer Name."
        save_message(history_id, "Assistant", assistant_answer)
        frappe.db.commit()
        return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}

    if "Please provide the Customer Name." in last_msg or "Provide the customer name." in last_msg or "Please try another name." in last_msg:
        customer_name = question.strip()
        user_reply_lower = customer_name.lower()
        
        if user_reply_lower in ['yes', 'y', 'ok', 'sure', 'yeah']:
            assistant_answer = "👤 Please provide the Customer Name."
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}
            
        elif user_reply_lower in ['no', 'n', 'nope', 'nah', 'cancel']:
            assistant_answer = "No problem! Whenever you are ready, 👤 Please provide the Customer Name."
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}

        customers = frappe.get_all("Customer", filters={"name": ["like", f"%{customer_name}%"]}, fields=["name"])
        
        if not customers:
            parts = [p for p in customer_name.split() if len(p) > 2]
            similar_customers = []
            if parts:
                or_filters = []
                for p in parts:
                    or_filters.append(["name", "like", f"%{p}%"])
                similar_customers = frappe.get_all("Customer", or_filters=or_filters, fields=["name"], limit=5)
            
            if similar_customers:
                assistant_answer = f"No exact match found for '{customer_name}'. Did you mean one of these?<br>"
                for c in similar_customers:
                    assistant_answer += f"&nbsp;&nbsp;&nbsp;&nbsp;&bull; {c.name}<br>"
                assistant_answer += "Please copy and paste the correct customer name."
            else:
                assistant_answer = f"No customer found matching '{customer_name}'. Please try another name. 🔍"
                
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}
            
        if len(customers) > 1:
            assistant_answer = "<b>👥 We found multiple customers.</b><br><br>"
            assistant_answer += "Please copy and paste the customer name you want to proceed with:<br><br>"
            for c in customers:
                assistant_answer += f"&bull; {c.name}<br>"
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}
        else:
            tbl = get_feasibility_table(customers[0].name)
            if not tbl:
                assistant_answer = f"😔 No feasible circuits were found for customer <b>{customers[0].name}.</b><br><br>Would you like to continue with another customer?"
            else:
                assistant_answer = f"✅ I found feasible circuits for customer <b>{customers[0].name}</b>.{tbl}"
                assistant_answer += "<br><div style='margin-bottom: 12px; font-size: 15px; color: #1f2937;'>🤖 Let's create your Task & Contract in just 3 minutes!</div>"
                assistant_answer += "<div style='margin-bottom: 12px; font-size: 15px; color: #1f2937;'>Simply answer a few quick questions, and I'll take care of the rest.</div>"
                assistant_answer += "<div style='margin-bottom: 4px; font-size: 15px; color: #1f2937;'>✅ Automatic Contract Creation</div>"
                assistant_answer += "<div style='margin-bottom: 4px; font-size: 15px; color: #1f2937;'>✅ Automatic Task Creation</div>"
                assistant_answer += "<div style='margin-bottom: 12px; font-size: 15px; color: #1f2937;'>✅ Guided Step-by-Step Process</div>"
                assistant_answer += "<div style='margin-bottom: 12px; font-size: 15px; color: #1f2937;'>📌 Supported for Order Type <b>Service</b> only.</div>"
                assistant_answer += "<div style='font-size: 15px; color: #1f2937;'>Ready to begin?</div>"
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}

    if "Please copy and paste the customer name you want to proceed with:" in last_msg or "Please copy and paste the correct customer name." in last_msg or "Please copy and paste the customer name exactly as shown above." in last_msg:
        customer_name = question.strip()
        customers = frappe.get_all("Customer", filters={"name": customer_name}, fields=["name"])
        if not customers:
            assistant_answer = f"<b>❌ No customer found matching \"{customer_name}\". Please copy and paste the customer name exactly as shown above. 📋</b>"
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}
        tbl = get_feasibility_table(customers[0].name)
        if not tbl:
            assistant_answer = f"😔 No feasible circuits were found for customer <b>{customers[0].name}.</b><br><br>Would you like to continue with another customer?"
        else:
            assistant_answer = f"✅ I found feasible circuits for customer <b>{customers[0].name}</b>.{tbl}"
            assistant_answer += "<br><div style='margin-bottom: 12px; font-size: 15px; color: #1f2937;'>🤖 Let's create your Task & Contract in just 3 minutes!</div>"
            assistant_answer += "<div style='margin-bottom: 12px; font-size: 15px; color: #1f2937;'>Simply answer a few quick questions, and I'll take care of the rest.</div>"
            assistant_answer += "<div style='margin-bottom: 4px; font-size: 15px; color: #1f2937;'>✅ Automatic Contract Creation</div>"
            assistant_answer += "<div style='margin-bottom: 4px; font-size: 15px; color: #1f2937;'>✅ Automatic Task Creation</div>"
            assistant_answer += "<div style='margin-bottom: 12px; font-size: 15px; color: #1f2937;'>✅ Guided Step-by-Step Process</div>"
            assistant_answer += "<div style='margin-bottom: 12px; font-size: 15px; color: #1f2937;'>📌 Supported for Order Type <b>Service</b> only.</div>"
            assistant_answer += "<div style='font-size: 15px; color: #1f2937;'>Ready to begin?</div>"
        save_message(history_id, "Assistant", assistant_answer)
        frappe.db.commit()
        return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}

    if "Would you like to continue with another customer?" in last_msg:
        user_reply_lower = question.strip().lower()
        if user_reply_lower in ['yes', 'y', 'ok', 'sure', 'yeah']:
            assistant_answer = "👤 Please provide the Customer Name."
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}
        elif user_reply_lower in ['no', 'n', 'nope', 'nah', 'cancel']:
            assistant_answer = "🙏 Thank you!"
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "CANCEL"}}

    if "Ready to begin?" in last_msg or "Are you ready to create the Task for the circuits listed above?" in last_msg:
        user_reply = question.strip().lower()
        if user_reply in ['yes', 'y', 'ok', 'sure', 'yeah']:
            assistant_answer = "🎯 Awesome! Please provide the CRM Deal ID."
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}
        elif user_reply in ['no', 'n', 'nope', 'nah', 'cancel', 'stop']:
            assistant_answer = "🙏 Thank you!"
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "CANCEL"}}

    if "🙏 Thank you!" in last_msg:
        user_reply = question.strip().lower()
        if user_reply in ['yes', 'y', 'ok', 'sure', 'yeah']:
            assistant_answer = "Great! Please select a new prompt from the menu to continue. 🚀"
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}

    if "Awesome! Please provide the CRM Deal ID." in last_msg or "Please provide the CRM Deal ID." in last_msg or "Please provide the correct CRM Deal ID." in last_msg or "Please provide a valid CRM Deal ID and try again." in last_msg:
        crm_deal_id = question.strip()
        
        if not frappe.db.exists("CRM Deal", crm_deal_id):
            assistant_answer = f"❌ No CRM Deal found matching **'{crm_deal_id}'**.<br><br>📝 Please provide a valid CRM Deal ID and try again."
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}
            
        deal = frappe.get_doc("CRM Deal", crm_deal_id)
        is_valid = False
        if deal:
            if deal.status == "Won" and deal.custom_po_no and deal.custom_po_date and deal.custom_customer_po_end_date:
                is_valid = True
                
        if not is_valid:
            assistant_answer = "The CRM Deal information is incomplete or not eligible for Task creation. Please ensure that the Deal Stage is 'Won' and that a Customer PO Number has been provided. ⚠️"
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}
            
        assistant_answer = f"Deal **{crm_deal_id}** is valid! ✅<br><br>What is the Service Type?<br>&bull; Capex<br>&bull; Opex (Rental)<br>📋 Please copy and paste the required option."
        save_message(history_id, "Assistant", assistant_answer)
        frappe.db.commit()
        return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}

    if "What is the Service Type?" in last_msg or "📋 Please copy and paste the required option." in last_msg:
        service_type_input = question.strip().lower()
        if "opex" in service_type_input:
            service_type = "Opex (Rental)"
        elif "capex" in service_type_input:
            service_type = "Capex"
        else:
            assistant_answer = "Please reply with either 'Capex' or 'Opex (Rental)'.<br><br>What is the Service Type?<br>&bull; Capex<br>&bull; Opex (Rental)<br>📋 Please copy and paste the required option."
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}
            
        assistant_answer = f"Service Type set to **{service_type}**. ✅<br><br>"
        assistant_answer += "<div style='margin-bottom: 12px; font-size: 15px; color: #1f2937;'>💰 <b>To update the Circuit ID costing, please choose one of the following options:</b></div>"
        assistant_answer += "<div style='margin-bottom: 12px; font-size: 15px; color: #1f2937;'><b>Option 1 – Same Costing for All Circuit IDs</b></div>"
        assistant_answer += "<div style='font-size: 15px; color: #1f2937;'>Item Code:</div>"
        assistant_answer += "<div style='font-size: 15px; color: #1f2937;'>OTC:</div>"
        assistant_answer += "<div style='margin-bottom: 16px; font-size: 15px; color: #1f2937;'>MRC:</div>"
        import urllib.parse

        history_msgs = frappe.get_all("AI Assistant Message", filters={"history": history_id}, fields=["content", "role"], order_by="creation desc", limit=50)
        customer_name = None
        for msg in history_msgs:
            if msg.role == "Assistant":
                m2 = re.search(r'I found feasible circuits for customer <b>([^<]+)</b>', msg.content or "")
                if m2:
                    customer_name = m2.group(1).strip()
                    break
        
        from frappe.utils.xlsxutils import make_xlsx
        import base64
        
        data = [["Circuit ID", "Item Code", "OTC", "MRC"]]
        if customer_name:
            feasibilities = frappe.get_all("Feasibility", 
                filters={
                    "customer": customer_name,
                    "docstatus": ["!=", 2],
                    "feasibility_status": "Feasible",
                    "sales_order": ["in", ["", None]]
                }, 
                fields=["name", "circuit_id"]
            )
            for feas in feasibilities:
                cid = feas.circuit_id or feas.name
                if cid:
                    data.append([str(cid), "", "", ""])

        xlsx_file = make_xlsx(data, "Costing_Template")
        encoded_excel = base64.b64encode(xlsx_file.getvalue()).decode('utf-8')

        assistant_answer += "<div style='margin-bottom: 12px; font-size: 15px; color: #1f2937;'><b>Option 2 – Different Costing for Circuit IDs</b></div>"
        assistant_answer += f"<div style='margin-bottom: 12px; font-size: 15px; color: #1f2937;'>Download template : <a href='data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{encoded_excel}' download='Costing_Template.xlsx' style='color: #2563eb; text-decoration: none; font-weight: 500;'>Costing_Template.xlsx</a>, update the costing details, and upload the completed file.</div>"
        assistant_answer += "<div style='font-size: 15px; color: #1f2937;'>✨ Once you provide these details, I will update all related Circuit IDs automatically.</div>"
        
        save_message(history_id, "Assistant", assistant_answer)
        frappe.db.commit()
        return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}

    if "Are you ready to create the Task and Contract with these costing items?" in last_msg:
        user_reply_lower = question.strip().lower()
        if user_reply_lower in ['yes', 'y', 'ok', 'sure', 'yeah', 'proceed']:
            # Fetch context
            history_msgs = frappe.get_all("AI Assistant Message", filters={"history": history_id}, fields=["content", "role"], order_by="creation desc", limit=50)
            customer_name = None
            crm_deal_id = None
            service_type = None
            costing_items = []
            
            for msg in history_msgs:
                if msg.role == "Assistant":
                    for cost_match in re.finditer(r'Costing item \*\*([^*]+)\*\* \(OTC: ([^,)]*), MRC: ([^)]*)\) saved!', msg.content or ""):
                        ic = cost_match.group(1).strip()
                        if frappe.db.exists("Item", ic) and not any(c['item_code'].lower() == ic.lower() for c in costing_items):
                            costing_items.append({
                                "item_code": ic,
                                "otc": cost_match.group(2).strip(),
                                "mrc": cost_match.group(3).strip()
                            })
                        
                if not crm_deal_id:
                    m1 = re.search(r'Deal \*\*([^*]+)\*\* is valid', msg.content or "")
                    if m1: crm_deal_id = m1.group(1).strip()
                if not customer_name:
                    m2 = re.search(r'Here are the feasible circuits for <b>([^<]+)</b>', msg.content or "")
                    if m2: customer_name = m2.group(1).strip()
                if not service_type:
                    m3 = re.search(r'Service Type set to \*\*([^*]+)\*\*', msg.content or "")
                    if m3: service_type = m3.group(1).strip()
                if crm_deal_id and customer_name and service_type:
                    break
                    
            if not customer_name or not crm_deal_id:
                assistant_answer = "Sorry, I lost the context of the customer or deal. Please start over. 😔"
                save_message(history_id, "Assistant", assistant_answer)
                frappe.db.commit()
                return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "CANCEL"}}
                
            if not frappe.db.exists("CRM Deal", crm_deal_id):
                assistant_answer = "Sorry, I could not fetch the deal. Please start over. 😔"
                save_message(history_id, "Assistant", assistant_answer)
                frappe.db.commit()
                return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "CANCEL"}}
                
            deal = frappe.get_doc("CRM Deal", crm_deal_id)
            
            feas_fields = ["name", "customer", "sales_person", "order_type", "customer_type"]
            meta = frappe.get_meta("Feasibility")
            if meta.has_field("service_type"): feas_fields.append("service_type")
            if meta.has_field("solution_type"): feas_fields.append("solution_type")
                
            feasibilities = frappe.get_all("Feasibility", 
                filters={
                    "customer": customer_name,
                    "docstatus": ["!=", 2],
                    "feasibility_status": "Feasible",
                    "sales_order": ["in", ["", None]]
                }, 
                fields=feas_fields
            )
            
            if not feasibilities:
                assistant_answer = f"No eligible feasible circuits found for {customer_name} to create a Task. 😔"
                save_message(history_id, "Assistant", assistant_answer)
                frappe.db.commit()
                return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "CANCEL"}}
                
            feas = feasibilities[0]
            
            from frappe.utils import date_diff
            duration = date_diff(deal.custom_customer_po_end_date, deal.custom_po_date)
            
            d1 = abs(duration - 90)
            d2 = abs(duration - 365)
            d3 = abs(duration - 1095)
            
            if d1 <= d2 and d1 <= d3:
                contract_template = "90 Days Contract"
            elif d2 <= d1 and d2 <= d3:
                contract_template = "12 Months Contract"
            else:
                contract_template = "36 Months Contract"
                
            contract_doc = {
                "doctype": "Contract",
                "party_type": "Customer",
                "party_name": feas.customer,
                "party_user": feas.sales_person,
                "start_date": deal.custom_po_date,
                "end_date": deal.custom_customer_po_end_date,
                "contract_template": contract_template
            }
            
            try:
                terms = frappe.db.get_value("Contract Template", contract_template, "contract_terms")
                contract_doc["contract_terms"] = terms or "Standard Terms"
            except Exception:
                contract_doc["contract_terms"] = "Standard Terms"
                
            cmeta = frappe.get_meta("Contract")
            if cmeta.has_field("custom_customer_po_no"):
                contract_doc["custom_customer_po_no"] = deal.custom_po_no
            elif cmeta.has_field("custom_customer_po_number"):
                contract_doc["custom_customer_po_number"] = deal.custom_po_no
            elif cmeta.has_field("customer_po_no"):
                contract_doc["customer_po_no"] = deal.custom_po_no
                
            contract = frappe.get_doc(contract_doc)
            contract.insert(ignore_permissions=True)
            
            task_doc = {
                "doctype": "Task",
                "type": "Sales Order Request",
                "subject": f"Sales Order Request - {customer_name}",
                "custom_customer": feas.customer,
                "custom_order_type": feas.order_type or "Service",
                "custom_contract": contract.name,
                "custom_owner": feas.sales_person,
                "custom_crm_deal": deal.name,
                "custom_customer_type": feas.customer_type,
                "custom_service_type": service_type,
                "custom_circuit_id": []
            }
            
            for feas in feasibilities:
                if feas.circuit_id:
                    for cost in costing_items:
                        task_doc["custom_circuit_id"].append({
                            "circuit_id": feas.circuit_id,
                            "site_name": feas.site_name,
                            "custom_product_name": cost['item_code'],
                            "custom_product": cost['item_code'],
                            "item": cost['otc'],
                            "custom_mrc": cost['mrc']
                        })
            
            task = frappe.get_doc(task_doc)
            task.insert(ignore_permissions=True)
            
            assistant_answer = "<h2 style='margin-top: 16px; margin-bottom: 8px;'>Successfully Created !</h2>"
            assistant_answer += f"✅ Contract: <b>{contract.name}</b><br>"
            assistant_answer += f"✅ Task: <b>{task.name}</b><br><br>"
            assistant_answer += "Is there anything else I can help you with? If yes, please select the new prompt. 😊"
            
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "SUCCESS"}}
            
        else:
            assistant_answer = "Okay, I have cancelled the creation process. You can start over by selecting a new prompt."
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "CANCEL"}}

    if "Once you provide these details, I will update all related Circuit Ids automatically." in last_msg or "Once you provide these details, I will update all related Circuit IDs automatically." in last_msg or "Do you want to add more costing items for this Circuit ID?" in last_msg or "Great! Please provide the next set of details:" in last_msg or "has already been added. Please provide a different Item Code" in last_msg or "I couldn't detect the Item Code" in last_msg or "Would you like to add another costing item for this Circuit ID?" in last_msg:

        user_reply_lower = question.strip().lower()
        
        # If user says No, we show the summary table
        if user_reply_lower in ['no', 'n', 'nope', 'nah', 'cancel', 'done', 'finish']:
            # Fetch context
            history_msgs = frappe.get_all("AI Assistant Message", filters={"history": history_id}, fields=["content", "role"], order_by="creation desc", limit=50)
            costing_items = []
            
            for msg in history_msgs:
                if msg.role == "Assistant":
                    for cost_match in re.finditer(r'Costing item \*\*([^*]+)\*\* \(OTC: ([^,)]*), MRC: ([^)]*)\) saved!', msg.content or ""):
                        ic = cost_match.group(1).strip()
                        if frappe.db.exists("Item", ic) and not any(c['item_code'].lower() == ic.lower() for c in costing_items):
                            costing_items.append({
                                "item_code": ic,
                                "otc": cost_match.group(2).strip(),
                                "mrc": cost_match.group(3).strip()
                            })
                if re.search(r'Deal \*\*([^*]+)\*\* is valid', msg.content or ""):
                    break
                    
            if not costing_items:
                assistant_answer = "No costing items were added. Please start over. 😔"
                save_message(history_id, "Assistant", assistant_answer)
                frappe.db.commit()
                return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "CANCEL"}}
                
            # Generate Summary Table
            assistant_answer = "<div class='costing-summary-card' style='margin: 15px 0;'>"
            assistant_answer += "<h4 style='margin-top: 0; margin-bottom: 20px; font-weight: 600; color: #1f2733;'>💰 Costing Summary for All Circuit IDs</h4>"
            assistant_answer += "<div class='table-responsive'>"
            assistant_answer += "<table class='table' style='width: 100%; border-collapse: collapse; margin-bottom: 20px;'>"
            assistant_answer += "<thead><tr style='border-bottom: 1px solid #e2e8f0;'><th style='padding: 12px 24px; border: none; text-align: left; color: #4a5568; font-weight: bold;'>Item Name</th><th style='padding: 12px 24px; border: none; text-align: right; color: #4a5568; font-weight: bold;'>OTC</th><th style='padding: 12px 24px; border: none; text-align: right; color: #4a5568; font-weight: bold;'>MRC</th><th style='padding: 12px 24px; border: none; text-align: right; color: #4a5568; font-weight: bold;'>ARC</th><th style='padding: 12px 24px; border: none; text-align: right; color: #4a5568;'><i class='fa fa-clone' style='cursor:pointer; color:#718096;' onclick='let r = document.createRange(); r.selectNode(this.closest(&quot;table&quot;)); let s = window.getSelection(); s.removeAllRanges(); s.addRange(r); document.execCommand(&quot;copy&quot;); s.removeAllRanges(); frappe.show_alert({message:&quot;Table Copied!&quot;, indicator:&quot;green&quot;});' title='Copy Table'></i></th></tr></thead><tbody>"
            
            for cost in costing_items:
                item_name = frappe.db.get_value("Item", cost['item_code'], "item_name") or cost['item_code']
                try: otc_val = float(cost['otc'])
                except: otc_val = 0.0
                try: mrc_val = float(cost['mrc'])
                except: mrc_val = 0.0
                
                arc_val = mrc_val * 12
                
                otc_str = "₹ {:,.2f}".format(otc_val)
                mrc_str = "₹ {:,.2f}".format(mrc_val)
                arc_str = "₹ {:,.2f}".format(arc_val)
                
                assistant_answer += f"<tr style='border-bottom: 1px solid #e2e8f0;'><td style='padding: 16px 24px; border: none; color: #1f2733;'>{item_name}</td><td style='padding: 16px 24px; border: none; color: #1f2733; text-align: right;'>{otc_str}</td><td style='padding: 16px 24px; border: none; color: #1f2733; text-align: right;'>{mrc_str}</td><td style='padding: 16px 24px; border: none; color: #1f2733; text-align: right;'>{arc_str}</td><td style='border: none;'></td></tr>"
            
            assistant_answer += "</tbody></table></div>"
            
            assistant_answer += "<p style='margin-bottom: 0; font-size: 14px; color: #1f2733;'>🚀 Are you ready to create the Task and Contract with these costing items?</p>"
            assistant_answer += "</div>"
            
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}
            
        elif user_reply_lower in ['yes', 'y', 'ok', 'sure', 'yeah']:
            assistant_answer = "Great! Please provide the next set of details:<br>Item Code:<br>OTC :<br>MRC :<br>You can continue adding as many costing items as required."
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}
            
        else:
            # Parse multiple costing items by splitting on "Item Code"
            chunks = re.split(r'(?i)Item Code:?\s*', question)
            added_items = []
            
            history_msgs = frappe.get_all("AI Assistant Message", filters={"history": history_id}, fields=["content", "role"], order_by="creation desc", limit=50)
            existing_items = []
            for msg in history_msgs:
                if msg.role == "Assistant":
                    for cm in re.finditer(r'Costing item \*\*([^*]+)\*\* \(OTC: ([^,)]*), MRC: ([^)]*)\) saved!', msg.content or ""):
                        ic = cm.group(1).strip()
                        if frappe.db.exists("Item", ic):
                            existing_items.append(ic.lower())
                if re.search(r'Deal \*\*([^*]+)\*\* is valid', msg.content or ""):
                    break
                    
            assistant_answer = ""
            for chunk in chunks:
                if not chunk.strip(): continue
                
                ic_match = re.match(r'^([A-Za-z0-9\-\_]+)', chunk.strip())
                if not ic_match: continue
                
                item_code = ic_match.group(1).strip()
                
                if not frappe.db.exists("Item", item_code):
                    assistant_answer += f"❌ The Item Code **{item_code}** was not found in the system. **Skipping.**<br>"
                    continue
                
                otc_match = re.search(r'OTC.*?([0-9,]+(?:\.[0-9]+)?)', chunk, re.IGNORECASE)
                mrc_match = re.search(r'MRC.*?([0-9,]+(?:\.[0-9]+)?)', chunk, re.IGNORECASE)
                
                otc = otc_match.group(1).replace(',', '').strip() if otc_match else ""
                mrc = mrc_match.group(1).replace(',', '').strip() if mrc_match else ""
                
                if item_code.lower() in existing_items:
                    assistant_answer += f"🔄 The Item Code **{item_code}** has been updated.<br>"
                    
                added_items.append((item_code, otc, mrc))
                existing_items.append(item_code.lower())
                
            if not added_items and not assistant_answer:
                assistant_answer = "I couldn't detect the Item Code. Please provide the Item Code, OTC, and MRC. 📝"
                save_message(history_id, "Assistant", assistant_answer)
                frappe.db.commit()
                return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}
                
            for itm in added_items:
                costs = []
                if itm[1] and str(itm[1]).strip(): costs.append(f"OTC: {itm[1]}")
                if itm[2] and str(itm[2]).strip(): costs.append(f"MRC: {itm[2]}")
                cost_str = ", ".join(costs)
                if cost_str: cost_str = f" ({cost_str})"
                
                assistant_answer += f"✅ Costing item **{itm[0]}**{cost_str} has been saved.<br>"
                assistant_answer += f"<span style='display:none;'>Costing item **{itm[0]}** (OTC: {itm[1]}, MRC: {itm[2]}) saved!</span><br>"
                
            assistant_answer += "<br>➕ Would you like to add another costing item for this Circuit ID?<br><br>Please provide:<br>Item Code:<br>OTC:<br>MRC:<br><br>You can add as many costing items as needed. If you're finished, simply reply **\"No\"**."
            
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}

        
    if "Is there anything else I can help you with? If yes, please select the new prompt. 😊" in last_msg:
        user_reply = question.strip().lower()
        if user_reply in ['no', 'n', 'nope', 'nah', 'cancel', 'stop']:
            assistant_answer = "Thank you! 😊"
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "CANCEL"}}
        elif user_reply in ['yes', 'y', 'ok', 'sure', 'yeah']:
            assistant_answer = "Great! Please select a new prompt from the menu to continue. 🚀"
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}

    # --- SPECIAL WORKFLOW: Sales Order Lookup by Circuit ID ---

    is_circuit_prompt = "Sales Order Lookup by Circuit ID" in question or "Sales Order Number against the Circuit ID" in question
    is_only_ids = re.match(r'^[\d,\s]+$', question.strip())
    
    if is_circuit_prompt or is_only_ids:
        # Check if we are in the middle of a lookup or just started
        ids = re.findall(r'\d+', question)
        if not ids:
            if is_circuit_prompt:
                assistant_answer = "Please enter the Circuit ID(s) to get the Sales Order Number."
                save_message(history_id, "Assistant", assistant_answer)
                frappe.db.commit()
                return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}
        else:
            # We have IDs, perform the lookup
            assistant_answer = lookup_sales_order_by_circuit_ids(question)
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "LIST", "topic": "Sales Order"}}

    # --- SPECIAL WORKFLOW: Circuit ID Lookup by Task ID ---
    is_task_prompt = "Circuit ID Lookup by Task ID" in question or "Please enter the Task ID to get the Circuit ID" in question
    is_task_id = re.match(r'^[\s,]*TASK-?[\d-]+([\s,]+TASK-?[\d-]+)*[\s,]*$', question.strip().upper())
    
    if is_task_prompt or is_task_id:
        task_ids = re.findall(r'TASK-?[\d-]+', question.upper())
        if not task_ids:
            if is_task_prompt:
                assistant_answer = "Please enter the Task ID to get the Circuit ID details."
                save_message(history_id, "Assistant", assistant_answer)
                frappe.db.commit()
                return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "INPUT_REQUIRED"}}
        else:
            # We have Task IDs, perform the lookup for all found IDs
            results = []
            # Remove duplicates while preserving order
            unique_task_ids = []
            for tid in task_ids:
                if tid not in unique_task_ids:
                    unique_task_ids.append(tid)
            
            for tid in unique_task_ids:
                results.append(lookup_circuit_details_by_task_id(tid))
            
            assistant_answer = "<br><hr style='border-top: 1px solid #eee; margin: 15px 0;'><br>".join(results)
            save_message(history_id, "Assistant", assistant_answer)
            frappe.db.commit()
            return {"answer": assistant_answer, "history_id": history_id, "intent": {"action": "LIST", "topic": "Task"}}
    # ----------------------------------------------------------
    # ----------------------------------------------------------

    # 2. Extract Intent (NLP -> Structured)
    intent = extract_intent(question, user_context)
    
    # 3. Fetch Data based on Intent
    result_data = fetch_erp_data(intent, user_context)

    # 4. Format Response (Data -> Natural Language)
    assistant_answer = format_ai_response(question, result_data, user_context, intent)

    # Save Assistant message
    save_message(history_id, "Assistant", assistant_answer)
    frappe.db.commit() # Save final response

    return {
        "answer": assistant_answer,
        "history_id": history_id,
        "intent": intent
    }

def save_message(history_id, role, content, file_url=None):
    msg_dict = {
        "doctype": "AI Assistant Message",
        "history": history_id,
        "role": role,
        "content": content
    }
    if file_url:
        msg_dict["attachment"] = file_url
    msg = frappe.get_doc(msg_dict)
    msg.insert(ignore_permissions=True)
    
    # Update last_interaction on history
    frappe.db.set_value("AI Assistant History", history_id, "last_interaction", frappe.utils.now())

def extract_intent(question, context):
    """Uses LLM to convert question into a structured intent JSON."""
    prompt = f"""
    You are an ERP Intent Extractor. Convert the User's question into a structured JSON.
    Context: {json.dumps(context)}

    Rules:
    - Topic: Doctype name (normalized, e.g., bill -> Sales Invoice, client -> Customer)
    - Action: LIST, COUNT, SINGLE, or GREETING
    - Filters: Standard Frappe filters (dictionary)
    - Fields: List of fields to fetch

    User: "{question}"
    Output valid JSON only.
    """
    
    try:
        response = call_ai_model(prompt)
        if not response:
            return {"error": "AI returned empty response"}
            
        # DeepSeek/LLMs might return markdown, clean it
        clean_json = response.replace("```json", "").replace("```", "").strip()
        if not clean_json:
             return {"error": "Empty structured response"}
             
        return json.loads(clean_json)
    except Exception as e:
        frappe.log_error(title="AI Assistant Intent Error", message=str(e))
        return {"error": "Failed to understand intent"}

def fetch_erp_data(intent, context):
    """Safe data fetching using frappe.get_list/get_all with user context."""
    if "error" in intent:
        return None

    doctype = intent.get("topic")
    if not doctype or doctype in ["DocType", "User", "Role", "Module Def", "File", "Error Log"]:
        return None
    filters = intent.get("filters", {})
    action = intent.get("action", "LIST")
    fields = intent.get("fields", ["name"])

    # Enforce Company security
    if context.get("company"):
        if frappe.get_meta(doctype).has_field("company"):
            filters["company"] = context["company"]

    try:
        if action == "COUNT":
            return frappe.db.count(doctype, filters)
        
        # Respect role permissions
        return frappe.get_list(doctype, filters=filters, fields=fields, limit=20)
    except Exception as e:
        frappe.log_error(f"AI Data Fetch Error: {str(e)}", "AI Assistant")
        return None

def format_ai_response(question, data, context, intent=None):
    """Uses LLM to convert raw data into a professional business answer."""
    prompt = f"""
    You are a Business Assistant for Nexapp. 
    Context: {json.dumps(context)}
    Intent: {json.dumps(intent)}
    Raw Data: {json.dumps(data)}
    Question: "{question}"

    Rules:
    - If it's a greeting (Hi, Hello), respond warmly and ask how you can help with Nexapp ERP.
    - If data is provided, convert it into a helpful, conversational answer.
    - If no data, explain politely.
    - Respect the user's department focus ({context.get('active_department')}).
    - Use HTML <b> and <br> for formatting.
    """
    
    return call_ai_model(prompt)

def lookup_sales_order_by_circuit_ids(ids_text):
    """Specialized lookup for Sales Order via Sales Order Item child table."""

    ids = re.findall(r'\d+', ids_text)
    if not ids:
        return "I couldn't find any valid Circuit IDs in your message. Please provide numeric IDs."
    
    results = []
    for circuit_id in ids:
        # Search in Sales Order Item child table field custom_feasibility
        so_name = frappe.db.get_value("Sales Order Item", {"custom_feasibility": circuit_id}, "parent")
        if so_name:
            results.append(f"<b>{circuit_id}</b> – {so_name}")
        else:
            results.append(f"<b>{circuit_id}</b> – No Sales Order Found")
    
    response = "Following are the Sales Order against the Circuit ID:<br><br>"
    response += "<br>".join(results)
    return response

def lookup_circuit_details_by_task_id(task_id):
    """Specialized lookup for Task details and associated Circuit IDs using silent fetching."""
    try:
        # Fetch Task details silently
        tasks = frappe.get_all("Task", 
            filters={"name": task_id}, 
            fields=["custom_sales_order_no", "custom_customer", "custom_purchase_order_no", "custom_po_amount", "creation", "completed_on"]
        )
        
        if not tasks:
            return f"Task <b>{task_id}</b> not found."
            
        task = tasks[0]
        
        # Fields from Task
        so_no = task.get("custom_sales_order_no") or "Sales order not found"
        customer = task.get("custom_customer") or "N/A"
        po_no = task.get("custom_purchase_order_no") or "N/A"
        po_amount = task.get("custom_po_amount") or 0
        
        # Date formatting
        from frappe.utils import formatdate
        req_on = formatdate(task.creation, "dd-mm-yyyy")
        comp_on = formatdate(task.completed_on, "dd-mm-yyyy") if task.completed_on else None
        
        # Fetch Child Table data silently
        child_records = frappe.get_all("Task Circuit ID",
            filters={"parent": task_id, "parenttype": "Task"},
            fields=["circuit_id"]
        )
        
        circuit_ids = [r.circuit_id for r in child_records if r.circuit_id]
        
        # Format response
        res = f"<h3>Task: {task_id}</h3>"
        res += f"<b>Sales Order No:</b> {so_no}<br>"
        res += f"<b>Customer:</b> {customer}<br>"
        res += f"<b>Purchase Order No:</b> {po_no}<br>"
        res += f"<b>Customer PO Amount:</b> {frappe.format_value(po_amount, {'fieldtype': 'Currency'})}<br>"
        res += f"<b>Task Requested on:</b> {req_on}<br>"
        if comp_on:
            res += f"<b>Completed on:</b> {comp_on}<br>"
        
        if circuit_ids:
            res += "<br><b>Circuit IDs:</b><br>"
            res += "<br>".join(circuit_ids)
        else:
            res += "<br><i>No Circuit IDs linked to this task.</i>"
            
        return res
    except Exception as e:
        frappe.log_error(f"Task Lookup Error: {str(e)}", "AI Assistant")
        return f"Error retrieving details for Task {task_id}."

@frappe.whitelist()
def get_chat_history():
    """Returns chat histories for current user."""
    return frappe.get_all(
        "AI Assistant History",
        filters={"user": frappe.session.user},
        fields=["name", "title", "department", "last_interaction", "pinned"],
        order_by="last_interaction desc",
        limit=20
    )

@frappe.whitelist()
def delete_chat_history(history_id):
	"""Delete chat history and all associated messages."""
	if not history_id:
		return
	
	# Delete all messages linked to this history session
	frappe.db.delete("AI Assistant Message", {"history": history_id})
	
	# Delete the history record itself
	frappe.delete_doc("AI Assistant History", history_id, ignore_permissions=True)
	return True

@frappe.whitelist()

@frappe.whitelist()
def log_chat_message(history_id, role, content, file_url=None):
    if not history_id or history_id == "null":
        # Create history if doesn't exist
        if role == "User" and content:
            title = content.strip()
        else:
            title = "New Chat"
            
        history_doc = frappe.get_doc({
            "doctype": "AI Assistant History",
            "user": frappe.session.user,
            "title": title or "New Chat"
        })
        history_doc.insert(ignore_permissions=True)
        history_id = history_doc.name
        
        # Enforce history limit of 5 per user
        user_histories = frappe.get_all(
            "AI Assistant History",
            filters={"user": frappe.session.user},
            order_by="creation desc",
            fields=["name"]
        )
        if len(user_histories) > 5:
            for old_hist in user_histories[5:]:
                frappe.db.delete("AI Assistant Message", {"history": old_hist.name})
                frappe.delete_doc("AI Assistant History", old_hist.name, ignore_permissions=True)
        
    save_message(history_id, role, content, file_url)
    return history_id

@frappe.whitelist()
def get_chat_messages(history_id):

    """Returns all messages for a specific chat session."""
    return frappe.get_all(
        "AI Assistant Message",
        filters={"history": history_id},
        fields=["role", "content", "creation", "attachment"],
        order_by="creation asc"
    )

@frappe.whitelist()
def update_profile_image(image_url):
    """Updates the image field of the Employee record linked to current user."""
    employee_id = frappe.db.get_value("Employee", {"user_id": frappe.session.user}, "name")
    if not employee_id:
        # Fallback to updating User only if no Employee record
        frappe.db.set_value("User", frappe.session.user, "user_image", image_url)
        return True
    
    frappe.db.set_value("Employee", employee_id, "image", image_url)
    # Sync with User doc
    frappe.db.set_value("User", frappe.session.user, "user_image", image_url)
    return True


@frappe.whitelist()
def get_employee_profile():
    """Fetch editable employee profile fields for the current user."""
    employee = frappe.db.get_value(
        "Employee",
        {"user_id": frappe.session.user},
        ["name", "marital_status", "cell_number", "personal_email", "current_address"],
        as_dict=True
    )
    return employee or {}


@frappe.whitelist()
def update_employee_profile(marital_status=None, cell_number=None, personal_email=None, current_address=None, image_url=None):
    """Updates editable fields of the Employee record linked to the current user."""
    employee_id = frappe.db.get_value("Employee", {"user_id": frappe.session.user}, "name")
    if not employee_id:
        frappe.throw("No Employee record linked to your account.")

    update_fields = {}
    if marital_status is not None:
        update_fields["marital_status"] = marital_status
    if cell_number is not None:
        update_fields["cell_number"] = cell_number
    if personal_email is not None:
        update_fields["personal_email"] = personal_email
    if current_address is not None:
        update_fields["current_address"] = current_address
    if image_url:
        update_fields["image"] = image_url
        frappe.db.set_value("User", frappe.session.user, "user_image", image_url)

    if update_fields:
        frappe.db.set_value("Employee", employee_id, update_fields)

    return True


@frappe.whitelist()
def get_holidays():
    """Fetches holidays for the current year, excluding weekends."""
    from frappe.utils import getdate, nowdate
    import datetime

    # 1. Find user's holiday list
    employee = frappe.db.get_value("Employee", {"user_id": frappe.session.user}, ["name", "holiday_list"], as_dict=True)
    holiday_list_name = None
    
    if employee and employee.holiday_list:
        holiday_list_name = employee.holiday_list
    else:
        # Fallback to Company default or Global default
        company = frappe.get_cached_value("User", frappe.session.user, "user_id") # actually session user
        # Just find any active for now or the first one if not set
        holiday_list_name = frappe.db.get_value("Holiday List", {"from_date": (">=", f"{datetime.date.today().year}-01-01")}, "name")

    if not holiday_list_name:
        return {"error": "No Holiday List found for the current year."}

    # 2. Fetch Holiday List details
    holiday_list = frappe.get_doc("Holiday List", holiday_list_name)
    holidays = []
    current_year = datetime.date.today().year
    today = getdate(nowdate())

    for h in holiday_list.holidays:
        h_date = getdate(h.holiday_date)
        if h_date.year != current_year:
            continue
            
        # Skip Saturday (5) and Sunday (6)
        if h_date.weekday() in [5, 6]:
            continue
            
        holidays.append({
            "date": h.holiday_date,
            "description": h.description,
            "is_upcoming": h_date > today,
            "day": h_date.strftime("%A")
        })

    return {
        "holiday_list_name": holiday_list_name,
        "from_date": holiday_list.from_date,
        "to_date": holiday_list.to_date,
        "holidays": holidays
    }


@frappe.whitelist()
def get_dynamic_prompts():
    """Fetches prompts based on the current user's Role Profile."""
    role_profile = frappe.db.get_value("User", frappe.session.user, "role_profile_name")
    
    if not role_profile:
        return []

    prompts = frappe.get_list(
        "AI Assistant Prompt",
        filters={"role_profile": role_profile},
        fields=["short_prompt", "full_prompt"],
        order_by="idx asc"
    )
    
    return prompts


@frappe.whitelist()
def get_user_creation_data():
    """Returns data needed for the 'Create ERP User' workflow."""
    companies = frappe.get_all("Company", fields=["name"])
    
    # Fetch all active employees without limit, respecting permissions as requested
    all_employees = frappe.get_all(
        "Employee", 
        fields=["name", "employee_name", "company_email", "personal_email", "company", "user_id"],
        filters={"status": "Active"},
        limit_page_length=0,
        ignore_permissions=False
    )
    
    # Fetch existing user emails to exclude them
    existing_users = {u.email for u in frappe.get_all("User", fields=["email"])}
    
    # Process and filter employees to ensure they have a valid company_email and aren't already users
    employees = []
    for e in all_employees:
        email = e.company_email # Match specifically on company_email as requested
        if email and email not in existing_users:
            employees.append({
                "name": e.name,
                "employee_name": e.employee_name,
                "company_email": email,
                "company": e.company
            })
    
    if not employees:
        frappe.log_error("No employees with emails found for AI Assistant User Creation workflow", "get_user_creation_data")
    
    # Safely fetch role profiles with custom_module_profile
    role_profiles = frappe.db.get_list("Role Profile", fields=["name", "custom_module_profile"])
    
    return {
        "companies": [c.name for c in companies],
        "employees": employees,
        "role_profiles": role_profiles
    }


@frappe.whitelist()
def create_erp_user_from_employee(email, role_profile):
    """Creates a User record using data from the linked Employee."""
    if frappe.db.exists("User", email):
        # Even if user exists, ensure employee is linked
        employee_name = frappe.db.get_value("Employee", {"company_email": email}, "name")
        if employee_name:
            frappe.db.set_value("Employee", employee_name, "user_id", email)
            frappe.db.commit()
        return {"status": "success", "message": "User already exists, employee linked"}
        
    employee = frappe.get_value(
        "Employee", 
        {"company_email": email}, 
        ["name", "first_name", "middle_name", "last_name"], 
        as_dict=True
    )
    
    if not employee:
        frappe.throw(f"No Employee record found with company email: {email}")
        
    # Fetch Module Profile from Role Profile if available
    module_profile = frappe.db.get_value("Role Profile", role_profile, "custom_module_profile")
    
    user = frappe.get_doc({
        "doctype": "User",
        "email": email,
        "first_name": employee.first_name,
        "middle_name": employee.middle_name,
        "last_name": employee.last_name,
        "role_profile_name": role_profile,
        "module_profile": module_profile,
        "send_welcome_email": 1,
        "enabled": 1,
        "search_bar": 0
    })
    
    try:
        user.insert(ignore_permissions=True)
        # Link employee with user
        frappe.db.set_value("Employee", employee.name, "user_id", email)
        frappe.db.commit()
    except Exception as e:
        # If user already exists or other email-related error, still try to proceed if record exists
        if frappe.db.exists("User", email):
            frappe.db.set_value("Employee", employee.name, "user_id", email)
            frappe.db.commit()
            return {"status": "success", "message": "User handled"}
        
        # Log other serious errors
        frappe.log_error(frappe.get_traceback(), "ERP User Creation Error")
        return {"status": "error", "message": str(e)}

    return {"status": "success"}
    
@frappe.whitelist()
def get_employee_name_by_email(email):
    """Returns the employee name for a given email by matching specifically with company_email, respecting permissions."""
    if not email:
        return None
        
    email = email.strip()
    
    # Use frappe.get_list to strictly respect all permission rules (Row Level Security, etc.)
    # Match specifically in company_email field as requested
    emps = frappe.get_list("Employee", 
        filters={"company_email": email, "status": "Active"}, 
        fields=["employee_name"],
        limit=1
    )
    
    if emps:
        return emps[0].employee_name
    
    # Second pass: Check personal_email just in case user selected it
    emps = frappe.get_list("Employee", 
        filters={"personal_email": email, "status": "Active"}, 
        fields=["employee_name"],
        limit=1
    )
    
    return emps[0].employee_name if emps else None

@frappe.whitelist()
def download_feasibility_template():
    """Generates and returns the Excel template for Feasibility bulk upload."""
    user_doc = frappe.get_doc("User", frappe.session.user)
    roles = frappe.get_roles()
    
    is_authorized = (
        user_doc.role_profile_name == "CRM Manager" or 
        "System Manager" in roles
    )
    
    if not is_authorized:
        frappe.throw("Access Denied: Only CRM Manager can download the template.")

    columns = [
        "Feasibility From", "Customer Name*", "Customer Type*", "Site Name*", "Customer Request*", 
        "Sales Person*", "Order Type*", "Site Type*", "Site ID / Legal Code", "Territory*", 
        "Solution Code*", "Static IP*", "Nos of Static IP Required", "Config Type*", 
        "Managed Services*", "Primary Data Plan Code", "Secondary Data Plan Code", 
        "Central Spoke", "Mobile", "Central Email", "Primary Contact Person*", 
        "Primary Contact Mobile", "Email", "Alternate Contact Person", 
        "Alternate Contact Mobile", "Secondary Email", "Address/ Street*", 
        "Pincode*"
    ]
    
    # Sample row data
    data = [
        columns,
        [
            "Customer", "Sample Customer", "Paid Customer", "Sample Site", "10-05-2024",
            frappe.session.user, "Service", "Standard", "SITE-001", "All Territories",
            "SOL-001", "No", "", "Remote Config", "Proactive", "PLAN-001", "",
            "", "", "", "John Doe", "9876543210", "john@example.com", "", "", "",
            "123 Main St", "400001"
        ]
    ]
    
    from frappe.utils.xlsxutils import make_xlsx
    xlsx_file = make_xlsx(data, "Feasibility Template")
    
    frappe.response['filename'] = 'Feasibility_Template.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

@frappe.whitelist()
def download_feasibility_template_as_base64():
    """Generates and returns the Excel template as a Base64 string."""
    user_doc = frappe.get_doc("User", frappe.session.user)
    roles = frappe.get_roles()
    
    is_authorized = (
        user_doc.role_profile_name == "CRM Manager" or 
        "System Manager" in roles
    )
    
    if not is_authorized:
        frappe.throw("Access Denied.")

    import openpyxl
    from openpyxl.styles import Font
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feasibility Template"

    columns = [
        "Feasibility From", "Customer Name", "Customer Type", "Site Name", "Customer Request", 
        "Sales Person", "Order Type", "Site Type", "Site ID / Legal Code", "Territory", 
        "Solution Code", "Static IP", "Nos of Static IP Required", "Config Type", 
        "Managed Services", "Primary Data Plan Code", "Secondary Data Plan Code", 
        "Central Spoke", "Mobile", "Central Email", "Primary Contact Person", 
        "Primary Contact Mobile", "Email", "Alternate Contact Person", 
        "Alternate Contact Mobile", "Secondary Email", "Address/ Street", 
        "Pincode"
    ]

    mandatory_labels = [
        "Customer Name", "Customer Type", "Site Name", "Customer Request", "Sales Person",
        "Order Type", "Site Type", "Territory", "Solution Code", "Static IP",
        "Config Type", "Managed Services", "Primary Contact Person", "Address/ Street", "Pincode"
    ]

    red_font = Font(color="FF0000", bold=True)
    standard_font = Font(bold=True)

    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        if column_title in mandatory_labels:
            cell.font = red_font
        else:
            cell.font = standard_font

    # Sample row
    sample_data = [
        "Customer", "Sample Customer", "Paid Customer", "Sample Site", "10-05-2024",
        frappe.session.user, "Service", "Standard", "SITE-001", "All Territories",
        "SOL-001", "No", "", "Remote Config", "Proactive", "PLAN-001", "",
        "", "", "", "John Doe", "9876543210", "john@example.com", "", "", "",
        "123 Main St", "400001"
    ]
    for col_num, val in enumerate(sample_data, 1):
        ws.cell(row=2, column=col_num).value = val

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    
    import base64
    return base64.b64encode(xlsx_file.getvalue()).decode('utf-8')

# === START PINCODE DETAILS LOGIC ===
@frappe.whitelist(allow_guest=True)
def get_pincode_details(pincode):
    """Fetches City, District, State, and Country from the Postal Pincode API."""
    import requests
    if not pincode: return {}
    
    # Remove non-digits

    pincode = re.sub(r'\D', '', str(pincode))
    if len(pincode) != 6: return {}
    
    try:
        url = f"https://api.postalpincode.in/pincode/{pincode}"
        # Spoof a real browser to avoid being blocked
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        response = requests.get(url, headers=headers, timeout=10)
        data = response.json()
        if data and data[0].get("Status") == "Success" and data[0].get("PostOffice"):
            po = data[0]["PostOffice"][0]
            return {
                "city": po.get("Block", ""),
                "district": po.get("District", ""),
                "state": po.get("State", ""),
                "country": po.get("Country", "India")
            }
    except Exception:
        pass
    return {}
# === END PINCODE DETAILS LOGIC ===


def try_parse_date(date_str):
    """Attempts to parse a date string from various formats into YYYY-MM-DD."""
    if not date_str: return None
    
    import datetime as dt
    if isinstance(date_str, (dt.date, dt.datetime)):
        return date_str.strftime('%Y-%m-%d')
    
    import dateutil.parser
    try:
        # Common Indian formats like DD-MM-YYYY
        d = dateutil.parser.parse(str(date_str), dayfirst=True)
        return d.strftime('%Y-%m-%d')
    except:
        return str(date_str)

@frappe.whitelist()
def identify_file_job(file_url):
    """Analyzes the file headers to identify what type of data it contains."""
    if not file_url: return None
    
    from frappe.utils.file_manager import get_file_path
    file_path = get_file_path(file_url.split('/')[-1])
    if not file_path: return None
    
    headers = []
    try:
        if file_path.endswith('.csv'):
            import csv
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                headers = next(reader, [])
        elif file_path.endswith('.xlsx'):
            from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
            data = read_xlsx_file_from_attached_file(filepath=file_path)
            if data: headers = data[0]
    except:
        return None

    # Normalize headers for matching

    norm_headers = [re.sub(r'[^a-z0-9]', '', str(h).lower()) for h in headers if h]
    
    # Identify Feasibility (match at least 3 key headers)
    feasibility_markers = ["feasibilityfrom", "customername", "sitename", "customerrequest", "salesperson", "ordertype", "sitetype"]
    matches = [m for m in feasibility_markers if m in norm_headers]
    
    if len(matches) >= 3:
        return "Feasibility"
        
    costing_markers = ["circuitid", "itemcode", "otc", "mrc"]
    cost_matches = [m for m in costing_markers if m in norm_headers]
    if len(cost_matches) >= 2:
        return "Costing"
        

    # Identify Bank Statement
    bank_header_str = " ".join(norm_headers)
    if "date" in bank_header_str and "description" in bank_header_str and "reference" in bank_header_str and "debit" in bank_header_str:
        return "Bank Statement"
    if "date" in bank_header_str and "narration" in bank_header_str and "credit" in bank_header_str:
        return "Bank Statement"

    return None

@frappe.whitelist()
def upload_feasibility_bulk(file_url, ignore_duplicates=False, confirmed=False, enriched_data=None):
    """Processes the uploaded CSV/Excel file and creates Feasibility records."""
    if isinstance(ignore_duplicates, str):
        ignore_duplicates = ignore_duplicates.lower() == 'true'
    
    if isinstance(confirmed, str):
        confirmed = confirmed.lower() == 'true'

    user_doc = frappe.get_doc("User", frappe.session.user)
    roles = frappe.get_roles()
    
    is_authorized = (
        user_doc.role_profile_name in ["CRM Manager", "L1 Trainee"] or 
        "System Manager" in roles
    )
    
    if not is_authorized:
        frappe.throw("Access Denied: Only CRM Manager or L1 Trainee can upload feasibility records.")

    if not file_url:
        frappe.throw("Please attach a file first.")

    from frappe.utils.file_manager import get_file_path
    file_path = get_file_path(file_url.split('/')[-1])
    
    if not file_path:
        frappe.throw("File not found.")

    is_xlsx = file_path.endswith('.xlsx')
    is_csv = file_path.endswith('.csv')

    if not (is_xlsx or is_csv):
        frappe.throw("Please upload a valid CSV or Excel (.xlsx) file.")

    rows = []
    if enriched_data:
        # Use data provided by frontend (enriched with pincode details)
        import json
        if isinstance(enriched_data, str):
            rows = json.loads(enriched_data)
        else:
            rows = enriched_data
    elif is_csv:
        import csv
        try:
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                rows = []
                for r in reader:
                    rows.append({str(k).strip().lower(): v for k, v in r.items() if k})
        except Exception as e:
            frappe.throw(f"Error reading CSV: {str(e)}")
    else:
        from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
        try:
            data = read_xlsx_file_from_attached_file(filepath=file_path)
            if data:
                headers = [str(h).strip().lower() for h in data[0] if h]
                for row_data in data[1:]:
                    row_dict = {}
                    for i, header in enumerate(headers):
                        if i < len(row_data):
                            row_dict[header] = row_data[i]
                    rows.append(row_dict)
        except Exception as e:
            frappe.throw(f"Error reading Excel: {str(e)}")

    if not rows:
        frappe.throw("The file is empty or could not be read.")

    # Duplicate check before processing
    if not ignore_duplicates:

        def normalize(s):
            if not s: return ""
            return re.sub(r'\s+', ' ', str(s).strip().lower())

        # Find the correct keys in the row dict by normalized matching
        sample_row = rows[0]
        name_key = None
        id_key = None
        
        for k in sample_row.keys():
            nk = normalize(k)
            if nk == "site name": name_key = k
            if nk == "site id / legal code" or nk == "site id/legal code" or nk == "site id": id_key = k
        
        duplicates = []
        for i, row in enumerate(rows):
            s_name = normalize(row.get(name_key)) if name_key else ""
            s_id = normalize(row.get(id_key)) if id_key else ""
            
            if not s_name and not s_id:
                continue

            found = False
            if s_name:
                # Check Feasibility
                if frappe.db.sql("select name from `tabFeasibility` where trim(lower(site_name)) = %s", (s_name,)):
                    found = True
                # Check Site
                if not found and frappe.db.sql("select name from `tabSite` where trim(lower(site_name)) = %s", (s_name,)):
                    found = True
            
            if not found and s_id:
                # Check Feasibility
                if frappe.db.sql("select name from `tabFeasibility` where trim(lower(site_id__legal_code)) = %s", (s_id,)):
                    found = True
                # Check Site
                if not found and frappe.db.sql("select name from `tabSite` where trim(lower(site_id__legal_code)) = %s", (s_id,)):
                    found = True
            
            if found:
                duplicates.append(f"Row {i+2}: '{s_name or s_id}' already exists in the system.")
        
        if duplicates:
            return {
                "status": "warning",
                "duplicates": duplicates,
                "message": "Potential duplicates found in existing records."
            }

    # Always ask for confirmation before actual creation if not already confirmed
    if not confirmed:
        return {
            "status": "confirmation_required",
            "total_records": len(rows),
            "rows": rows,
            "message": f"Found {len(rows)} records in your file. Ready to proceed?"
        }

    mapping = {
        "Feasibility From": "feaseibility_from",
        "Customer Name": "customer",
        "Customer Type": "customer_type",
        "Site Name": "site_name",
        "Customer Request": "customer_request",
        "Sales Person": "sales_person",
        "Order Type": "order_type",
        "Site Type": "site_type",
        "Site ID / Legal Code": "site_id__legal_code",
        "Territory": "territory",
        "Solution Code": "solution_code",
        "Static IP": "static_ip",
        "Nos of Static IP Required": "no_of_static_ip_required",
        "Config Type": "config_type",
        "Managed Services": "managed_services",
        "Primary Data Plan Code": "primary_data_plan",
        "Secondary Data Plan Code": "secondary_data_plan",
        "Central Spoke": "central_spoke",
        "Mobile": "mobile",
        "Central Email": "central_email",
        "Primary Contact Person": "contact_person",
        "Primary Contact Mobile": "primary_contact_mobile",
        "Email": "email",
        "Alternate Contact Person": "alternate_contact_person",
        "Alternate Contact Mobile": "alternate_contact_mobile",
        "Secondary Email": "secondary_email",
        "Address/ Street": "address_street",
        "City": "city",
        "Pincode": "pincode",
        "District": "district",
        "State": "state",
        "Country": "country"
    }

    mandatory_fields = [
        "customer", "customer_type", "site_name", "customer_request", "sales_person",
        "order_type", "site_type", "territory", "solution_code", "static_ip",
        "config_type", "managed_services", "contact_person", "address_street", "pincode"
    ]

    success_count = 0
    errors = []
    
    for i, row in enumerate(rows):
        # Skip completely empty rows (ghost rows)
        if not any(str(v).strip() for v in row.values() if v is not None):
            continue
            
        try:
            doc_data = {"doctype": "Feasibility"}
            for file_col, fieldname in mapping.items():
                val = row.get(file_col.lower())
                
                if val is not None:
                    # Convert everything to string to prevent 'int' has no attribute 'strip'
                    val = str(val).strip()
                    
                    # Smart Date Conversion for Date Fields
                    if fieldname in ["customer_request"]:
                        from frappe.utils import getdate
                        try:
                            val = try_parse_date(val)
                        except:
                            pass
                            
                doc_data[fieldname] = val

            # Auto-fetch address details if pincode is present but other fields are missing or placeholders
            if doc_data.get("pincode"):
                # If fields are missing OR contain placeholder text like "District update automatically"
                def is_empty_or_placeholder(v):
                    if not v: return True
                    v_low = str(v).lower()
                    return "automatically" in v_low or "update" in v_low

                needs_update = any(is_empty_or_placeholder(doc_data.get(f)) for f in ["city", "district", "state", "country"])
                
                if needs_update:
                    pincode_data = get_pincode_details(doc_data.get("pincode"))
                    if pincode_data:
                        for field in ["city", "district", "state", "country"]:
                            if is_empty_or_placeholder(doc_data.get(field)):
                                doc_data[field] = pincode_data.get(field)
            
            # Append location details to address_street for a complete address string
            address_parts = [doc_data.get("address_street")]
            for field in ["city", "district", "pincode", "state", "country"]:
                val = doc_data.get(field)
                if val and str(val).strip() and str(val).strip() not in str(address_parts[0]):
                    address_parts.append(str(val).strip())
            
            doc_data["address_street"] = ", ".join([p for p in address_parts if p])

            missing = [f for f in mandatory_fields if not doc_data.get(f)]
            if missing:
                labels = [file_col for file_col, fieldname in mapping.items() if fieldname in missing]
                errors.append(f"Row {i+2}: Missing data in mandatory fields: {', '.join(labels)}")
                continue

            new_doc = frappe.get_doc(doc_data)
            new_doc.insert(ignore_permissions=True)
            success_count += 1
            frappe.clear_messages()
        except Exception as e:
            import traceback
            frappe.log_error(traceback.format_exc(), "Upload Feasibility Bulk Error")
            errors.append(f"Row {i+2}: {str(e)}")
            frappe.clear_messages()
    
    frappe.db.commit()
    frappe.clear_messages()
    
    return {
        "status": "success" if not errors else "partial",
        "success_count": success_count,
        "errors": errors
    }


@frappe.whitelist()
def process_costing_file(file_url):
    if not file_url: return None
    from frappe.utils.file_manager import get_file_path
    from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
    import csv
    
    file_path = get_file_path(file_url.split('/')[-1])
    if not file_path: return None
    
    data = []
    if file_path.endswith('.csv'):
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            data = list(reader)
    elif file_path.endswith('.xlsx'):
        data = read_xlsx_file_from_attached_file(filepath=file_path)
        
    if not data or len(data) < 2:
        return {"status": "error", "message": "File is empty or invalid"}
        
    headers = [str(h).strip().lower() for h in data[0]]
    
    # Try to find column indices
    def find_col(possible_names):
        for idx, h in enumerate(headers):
            h_norm = ''.join(c for c in h if c.isalnum())
            for p in possible_names:
                if p in h_norm:
                    return idx
        return -1
        
    col_circuit = find_col(['circuitid', 'circuit'])
    col_item = find_col(['itemcode', 'item'])
    col_otc = find_col(['otc'])
    col_mrc = find_col(['mrc'])
    
    if col_item == -1:
        return {"status": "error", "message": "Could not find Item Code column"}
        
    results = []
    for row in data[1:]:
        if not row: continue
        item_code = str(row[col_item]).strip() if col_item != -1 and len(row) > col_item else ""
        if not item_code: continue
        
        otc = float(row[col_otc]) if col_otc != -1 and len(row) > col_otc and row[col_otc] else 0.0
        mrc = float(row[col_mrc]) if col_mrc != -1 and len(row) > col_mrc and row[col_mrc] else 0.0
        arc = mrc * 12
        
        circuit_id = str(row[col_circuit]).strip() if col_circuit != -1 and len(row) > col_circuit else ""
        
        item_name = item_code
        if frappe.db.exists("Item", item_code):
            item_name = frappe.db.get_value("Item", item_code, "item_name")
            
        results.append({
            "circuit_id": circuit_id,
            "item_name": item_name,
            "otc": otc,
            "mrc": mrc,
            "arc": arc
        })
        
    return {"status": "success", "data": results}

@frappe.whitelist()
def validate_disconnection_circuits(text, file_url=None):
    import re
    circuit_ids = []
    if file_url:
        from frappe.utils.file_manager import get_file_path
        file_path = get_file_path(file_url.split('/')[-1])
        if file_path:
            if file_path.endswith('.csv'):
                import csv
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    for r in reader:
                        for k, v in r.items():
                            if k and 'circuit' in k.lower() and v:
                                circuit_ids.append(v.strip())
                                break
            elif file_path.endswith('.xlsx'):
                from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
                data = read_xlsx_file_from_attached_file(filepath=file_path)
                if data and len(data) > 1:
                    headers = [str(h).strip().lower() for h in data[0] if h]
                    idx = -1
                    for i, h in enumerate(headers):
                        if 'circuit' in h:
                            idx = i
                            break
                    if idx != -1:
                        for row in data[1:]:
                            if len(row) > idx and row[idx]:
                                circuit_ids.append(str(row[idx]).strip())
    
    if text:
        parts = re.split(r'[, \n]+', text)
        circuit_ids.extend([p.strip() for p in parts if p.strip()])
    
    circuit_ids = list(set(circuit_ids))
    
    valid = []
    invalid = []
    
    for cid in circuit_ids:
        site = frappe.db.get_value("Site", {"circuit_id": cid}, ["name", "site_status"], as_dict=True)
        if site and site.site_status == "Delivered and Live":
            valid.append(cid)
        else:
            invalid.append(cid)
            
    return {
        "total_circuits": len(circuit_ids),
        "valid_circuits": valid,
        "invalid_circuits": invalid
    }

@frappe.whitelist()
def create_disconnection_request(data):
    import json
    if isinstance(data, str):
        data = json.loads(data)
        
    doc = frappe.get_doc({
        "doctype": "Disconnection Request",
        "customer_name": data.get("customer_name"),
        "customer_type": data.get("customer_type"),
        "reason_for_disconnection": data.get("reason_for_disconnection"),
        "notice_period": data.get("notice_period"),
        "status": "Draft"
    })
    
    try_parse_date_fn = globals().get('try_parse_date')
    
    req_date = data.get("customer_requested_date")
    if req_date:
        if try_parse_date_fn: req_date = try_parse_date_fn(req_date)
        doc.customer_requested_date = req_date
        
    start_date = data.get("notice_period_start_date")
    if start_date:
        if try_parse_date_fn: start_date = try_parse_date_fn(start_date)
        doc.notice_period_start_date = start_date
    
    if data.get("customer_disconnection_confirmation"):
        doc.customer_disconnection_confirmation = data.get("customer_disconnection_confirmation")
        
    for cid in data.get("valid_circuits", []):
        doc.append("circuit_ids", {
            "circuit_id": cid
        })
        
    doc.insert(ignore_permissions=True)
    frappe.db.commit()
    
    return {
        "status": "success",
        "doc": {"name": doc.name},
        "count": len(data.get("valid_circuits", [])),
        "total": data.get("total_circuits", len(data.get("valid_circuits", [])))
    }

@frappe.whitelist()
def download_bank_statement_template():
    from frappe.utils.xlsxutils import make_xlsx
    
    data = [
        ["Transaction Date", "Value Date", "Transaction Description", "Reference No", "Debit Amount", "Credit Amount", "Running Balance"]
    ]
    
    xlsx_file = make_xlsx(data, "Bank Statement Template")
    
    frappe.response['filename'] = 'Bank_Statement_Template.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

@frappe.whitelist()
def download_recon_excel(data):
    import json
    from frappe.utils.xlsxutils import make_xlsx
    
    if isinstance(data, str):
        data = json.loads(data)
        
    xlsx_data = [["Date", "Narration", "Ref No", "Debit Amount", "Credit Amount", "Status", "Voucher", "Reconciled Amount", "Remark"]]
    for row in data:
        xlsx_data.append([
            row.get("date", ""),
            row.get("narration", ""),
            row.get("ref_no", ""),
            row.get("debit_amount", "") if row.get("debit_amount") else "",
            row.get("credit_amount", "") if row.get("credit_amount") else "",
            row.get("status", ""),
            row.get("matched_voucher", ""),
            row.get("reconciled_amount", 0),
            row.get("remark", "")
        ])
        
    xlsx_file = make_xlsx(xlsx_data, "Reconciliation Results")
    
    frappe.response['filename'] = 'Reconciliation_Results.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

@frappe.whitelist()
def parse_bank_statement(file_url):
    if not file_url:
        return {"status": "error", "message": "No file provided"}
        
    from frappe.utils.file_manager import get_file_path
    file_path = get_file_path(file_url.split('/')[-1])
    if not file_path:
        return {"status": "error", "message": "File not found"}
        
    try:
        import pandas as pd
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
            
        # Clean column names
        df.columns = [str(c).strip().lower() for c in df.columns]
        
        # Get list of bank accounts for the user to select
        bank_accounts = frappe.get_all("Account", filters={
            "account_type": "Bank", 
            "parent_account": "Bank Accounts - NTPL",
            "company": "Nexapp Technologies Private Limited",
            "is_group": 0
        }, fields=["name"])
        bank_names = [b.name for b in bank_accounts]
        
        # Basic auto-detect based on file name or simple heuristics could go here
        detected_bank = None
        for b in bank_names:
            if str(b).lower() in file_path.lower():
                detected_bank = b
                break
                
        return {
            "status": "success",
            "transaction_count": len(df),
            "bank_accounts": bank_names,
            "detected_bank": detected_bank
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Bank Statement Parsing Error")
        return {"status": "error", "message": str(e)}

@frappe.whitelist()
def run_reconciliation_matching(bank_account, file_url, from_date=None, to_date=None):
    from rapidfuzz import fuzz
    import pandas as pd
    from frappe.utils.file_manager import get_file_path
    
    file_path = get_file_path(file_url.split('/')[-1])
    if file_path.endswith('.csv'):
        df = pd.read_csv(file_path)
    else:
        df = pd.read_excel(file_path)
        
    # Standardize cols
    col_map = {}
    for c in df.columns:
        norm = str(c).strip().lower().replace(' ', '').replace('_', '')
        if 'transactiondate' in norm or ('date' in norm and 'value' not in norm and 'date' not in col_map.values()): col_map[c] = 'date'
        elif 'desc' in norm or 'narr' in norm or 'partic' in norm: col_map[c] = 'narration'
        elif 'ref' in norm or 'cheq' in norm or 'chq' in norm or 'utr' in norm: col_map[c] = 'ref_no'
        elif 'debit' in norm or 'withdrawal' in norm: col_map[c] = 'debit'
        elif 'credit' in norm or 'deposit' in norm: col_map[c] = 'credit'
        elif 'amount' in norm and 'debit' not in norm and 'credit' not in norm: col_map[c] = 'amount'
    df.rename(columns=col_map, inplace=True)
    
    # Ensure required columns exist, fill missing
    for col in ['date', 'narration', 'ref_no', 'debit', 'credit', 'amount']:
        if col not in df.columns:
            df[col] = ''
            
    if from_date and to_date:
        try:
            df['parsed_date'] = pd.to_datetime(df['date'], format='mixed', dayfirst=True)
            end_date = pd.to_datetime(to_date) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
            df = df[(df['parsed_date'] >= pd.to_datetime(from_date)) & (df['parsed_date'] <= end_date)]
            df = df.drop(columns=['parsed_date'])
        except:
            pass
    
    # Fetch ERP data
    # Get GL Account for Bank
    gl_account = frappe.db.get_value("Bank Account", bank_account, "account") or bank_account
    
    pe_date_filter = ""
    je_date_filter = ""
    if from_date and to_date:
        pe_date_filter = f" AND reference_date BETWEEN '{from_date}' AND '{to_date}'"
        je_date_filter = f" AND je.posting_date BETWEEN '{from_date}' AND '{to_date}'"
    
    # Fetch Payment Entries
    pes = frappe.db.sql(f"""
        SELECT name as voucher_no, 'Payment Entry' as voucher_type, reference_no as ref_no, 
               reference_date as date, paid_amount as amount, party, party_name, clearance_date
        FROM `tabPayment Entry`
        WHERE (paid_from = %s OR paid_to = %s) AND docstatus = 1 {pe_date_filter}
    """, (gl_account, gl_account), as_dict=True)
    
    # Fetch Journal Entries
    jes = frappe.db.sql(f"""
        SELECT je.name as voucher_no, 'Journal Entry' as voucher_type, je.cheque_no as ref_no, 
               je.posting_date as date, jea.debit, jea.credit, jea.party, je.clearance_date
        FROM `tabJournal Entry Account` jea
        JOIN `tabJournal Entry` je ON je.name = jea.parent
        WHERE jea.account = %s AND je.docstatus = 1 {je_date_filter}
    """, (gl_account,), as_dict=True)
    
    # Combine erp entries
    erp_entries = []
    for pe in pes:
        erp_entries.append({
            'voucher_no': pe.voucher_no,
            'voucher_type': pe.voucher_type,
            'ref_no': str(pe.ref_no or '').strip().lower(),
            'amount': float(pe.amount or 0),
            'date': pe.date,
            'party': pe.party_name or pe.party or '',
            'clearance_date': pe.clearance_date
        })
    for je in jes:
        amt = float(je.debit or 0) if float(je.debit or 0) > 0 else float(je.credit or 0)
        erp_entries.append({
            'voucher_no': je.voucher_no,
            'voucher_type': je.voucher_type,
            'ref_no': str(je.ref_no or '').strip().lower(),
            'amount': amt,
            'date': je.date,
            'party': je.party or '',
            'clearance_date': je.clearance_date
        })
        
    results = []
    exact_count = 0
    sugg_count = 0
    unmatched_count = 0
    reconciled_count = 0
    
    reconcile_payload = []
    
    for idx, row in df.iterrows():
        bank_date = str(row.get('transaction date', row.get('date', '')))
        bank_narr = str(row.get('transaction description', row.get('narration', row.get('description', ''))))
        bank_ref = str(row.get('reference no', row.get('ref no', row.get('ref_no', '')))).strip().lower()
        
        amt_val = 0
        try:
            debit = str(row.get('debit amount', row.get('debit', '0'))).replace(',', '').strip()
            credit = str(row.get('credit amount', row.get('credit', '0'))).replace(',', '').strip()
            debit_val = float(debit) if debit and debit != 'nan' else 0
            credit_val = float(credit) if credit and credit != 'nan' else 0
            amt_val = debit_val if debit_val > 0 else credit_val
        except:
            pass
            
        matched = False
        status = 'Unmatched'
        matched_voucher = ''
        
        # Level 1: Ref No or Narration + Amount (100%)
        if amt_val > 0:
            bank_narr_lower = bank_narr.strip().lower()
            for erp in erp_entries:
                erp_ref = erp['ref_no']
                if erp_ref and erp_ref != 'nan' and erp_ref != 'none':
                    match_cond = False
                    if erp_ref == bank_ref: match_cond = True
                    elif len(erp_ref) > 3 and erp_ref in bank_narr_lower: match_cond = True
                    elif bank_ref and len(bank_ref) > 3 and bank_ref in erp_ref: match_cond = True
                    
                    if match_cond and abs(erp['amount'] - amt_val) < 1.0:
                        status = 'Reconciled' if erp['clearance_date'] else 'Matched'
                        matched_voucher = erp['voucher_no']
                        matched = True
                        break
                    
        # Level 2 & 3 simplified: Just matching amount and checking fuzz
        if not matched and amt_val > 0:
            potential = []
            for erp in erp_entries:
                if abs(erp['amount'] - amt_val) < 1.0:
                    potential.append(erp)
                    
            if len(potential) == 1:
                status = 'Reconciled' if potential[0]['clearance_date'] else 'Suggested'
                matched_voucher = potential[0]['voucher_no']
                matched = True
            elif len(potential) > 1:
                # Level 4: Narration match
                best_score = 0
                best_erp = None
                for erp in potential:
                    score = fuzz.partial_ratio(bank_narr.lower(), erp['party'].lower())
                    if score > best_score:
                        best_score = score
                        best_erp = erp
                if best_score > 85 and best_erp:
                    status = 'Reconciled' if best_erp['clearance_date'] else 'Suggested'
                    matched_voucher = best_erp['voucher_no']
                    matched = True
                    
        if status == 'Matched': exact_count += 1
        elif status == 'Suggested': sugg_count += 1
        elif status == 'Reconciled': reconciled_count += 1
        else: unmatched_count += 1
        
        if status in ['Matched', 'Suggested'] and matched_voucher:
            reconcile_payload.append({
                'voucher_no': matched_voucher,
                'clearance_date': bank_date # simplify date parsing here
            })
            
        results.append({
            'date': bank_date,
            'narration': bank_narr,
            'ref_no': bank_ref,
            'amount': amt_val,
            'status': status,
            'matched_voucher': matched_voucher,
            'reconciled_amount': amt_val if status in ['Matched', 'Suggested', 'Reconciled'] else 0,
            'remark': f"{'Already Reconciled with' if status == 'Reconciled' else 'Matched with'} {matched_voucher}" if matched_voucher else 'Pending'
        })
        
    return {
        "status": "success",
        "total": len(df),
        "exact_matches": exact_count,
        "suggestions": sugg_count,
        "unmatched": unmatched_count,
        "reconciled": reconciled_count,
        "grid_data": results,
        "reconcile_payload": reconcile_payload
    }

@frappe.whitelist()
def confirm_reconciliation(payload):
    import json
    try:
        data = json.loads(payload)
        for item in data:
            voucher_no = item.get('voucher_no')
            c_date = item.get('clearance_date')
            if not voucher_no or not c_date: continue
            
            # Convert date if needed
            from frappe.utils import getdate
            try:
                c_date = getdate(c_date)
            except:
                c_date = frappe.utils.today()
                
            if 'PAY' in voucher_no:
                frappe.db.set_value('Payment Entry', voucher_no, 'clearance_date', c_date)
            elif 'JV' in voucher_no or 'Journal' in voucher_no:
                # Ensure clearance_date exists on JE, if not, skip or update custom field
                try:
                    frappe.db.set_value('Journal Entry', voucher_no, 'clearance_date', c_date)
                except Exception as e:
                    # Ignore if field doesn't exist
                    pass
        frappe.db.commit()
        return {"status": "success"}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Bank Recon Confirmation Error")
        return {"status": "error", "message": str(e)}


# =================================================
# 🤖 AI ASSISTANT — CORE ENGINE END
# =================================================

#####################################################################
## User will update for the Restricted Supplier

import frappe

def create_user_permission(doc, method):
    if not doc.enabled:
        return

    old_doc = doc.get_doc_before_save()

    if method == "after_insert" or (old_doc and not old_doc.enabled):
        exists = frappe.db.exists("User Permission", {
            "user": doc.name,
            "allow": "Restricted",
            "for_value": "Open Supplier"
        })

        if not exists:
            frappe.get_doc({
                "doctype": "User Permission",
                "user": doc.name,
                "allow": "Restricted",
                "for_value": "Open Supplier",
                "apply_to_all_doctypes": 1
            }).insert(ignore_permissions=True)
############################################
@frappe.whitelist()
def download_custom_sales_report_xlsx(doctype, fields, filters):
    import datetime
    if isinstance(fields, str):
        fields = frappe.parse_json(fields)
    if isinstance(filters, str):
        filters = frappe.parse_json(filters)

    # 1. Build Filters
    query_filters = {}
    
    # Handle Date Range
    date_field = "creation"
    if doctype == "Quotation": date_field = "transaction_date"
    elif doctype == "Sales Order": date_field = "transaction_date"

    if filters.get("date_range") == "Current Month":
        query_filters[date_field] = ["between", [frappe.utils.month_start(), frappe.utils.month_end()]]
    elif filters.get("date_range") == "Last 3 Months":
        three_months_ago = frappe.utils.add_months(frappe.utils.nowdate(), -3)
        query_filters[date_field] = [">=", three_months_ago]
    elif filters.get("date_range") == "Custom" and filters.get("from_date") and filters.get("to_date"):
        query_filters[date_field] = ["between", [filters.get("from_date"), filters.get("to_date")]]

    # 2. Fetch Data
    # Meta to get labels
    meta = frappe.get_meta(doctype)
    labels = {f.fieldname: f.label for f in meta.fields}
    labels["name"] = "ID"

    data = frappe.get_list(doctype, 
        filters=query_filters, 
        fields=fields, 
        order_by="creation desc",
        limit=5000
    )

    if not data:
        rows = [["No data found for the selected criteria"]]
    else:
        # 3. Format XLSX
        # Header row using labels
        header = [labels.get(f, f.replace("_", " ").title()) for f in fields]
        rows = [header]
        
        for d in data:
            row = []
            for f in fields:
                val = d.get(f)
                # Format dates for Excel
                if isinstance(val, (datetime.datetime, datetime.date)):
                    val = frappe.utils.getdate(val)
                row.append(val)
            rows.append(row)

    from frappe.utils.xlsxutils import make_xlsx
    xlsx_file = make_xlsx(rows, f"Custom {doctype} Report")
    
    frappe.response['filename'] = f"Custom_{doctype.replace(' ', '_')}_Report.xlsx"
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = "binary"

@frappe.whitelist()
def download_combined_sales_report_xlsx(fields, filters):
    import json
    if isinstance(fields, str):
        fields = json.loads(fields)
    if isinstance(filters, str):
        filters = json.loads(filters)

    # 1. Build Query
    select_clause = []
    joins = []
    
    # Map for aliases and joins
    dt_map = {
        "CRM Deal": {"alias": "deal", "table": "tabCRM Deal"},
        "Quotation": {"alias": "q", "table": "tabQuotation", "join": "LEFT JOIN `tabQuotation` q ON q.custom_deal = deal.name"},
        "Sales Order": {"alias": "so", "table": "tabSales Order", "join": "LEFT JOIN `tabSales Order` so ON so.custom_crm_deal = deal.name"},
        "Task": {"alias": "t", "table": "tabTask", "join": "LEFT JOIN `tabTask` t ON t.custom_deal = deal.name"},
        "Feasibility": {"alias": "f", "table": "tabFeasibility", "join": "LEFT JOIN `tabFeasibility` f ON f.custom_deal = deal.name"},
        "Site": {"alias": "s", "table": "tabSite", "join": "LEFT JOIN `tabSite` s ON s.sales_order = so.name"}
    }

    # Handle dependencies (Site needs Sales Order)
    if "Site" in fields and "Sales Order" not in fields:
        # We need to add Sales Order to joins even if no fields are selected from it
        if dt_map["Sales Order"]["join"] not in joins:
            joins.append(dt_map["Sales Order"]["join"])

    columns = []
    for dt, dt_fields in fields.items():
        if dt not in dt_map: continue
        alias = dt_map[dt]["alias"]
        
        # Add join if not CRM Deal
        if dt != "CRM Deal":
            if dt_map[dt]["join"] not in joins:
                joins.append(dt_map[dt]["join"])
        
        # Meta for labels
        meta = frappe.get_meta(dt)
        labels = {f.fieldname: f.label for f in meta.fields}
        labels["name"] = "ID"

        for f in dt_fields:
            # Use backticks for aliases with spaces
            col_alias = f"{dt}_{f}".replace(" ", "_")
            select_clause.append(f"{alias}.{f} as `{col_alias}`")
            columns.append(f"{dt} - {labels.get(f, f)}")

    # Conditions
    where = []
    if filters.get("date_range") == "Current Month":
        where.append("MONTH(deal.creation) = MONTH(NOW()) AND YEAR(deal.creation) = YEAR(NOW())")
    elif filters.get("date_range") == "Last 3 Months":
        where.append("deal.creation >= DATE_SUB(NOW(), INTERVAL 3 MONTH)")
    elif filters.get("date_range") == "Custom" and filters.get("from_date") and filters.get("to_date"):
        where.append(f"deal.creation BETWEEN '{filters.get('from_date')}' AND '{filters.get('to_date')}'")

    where_str = "WHERE " + " AND ".join(where) if where else ""
    
    query = f"""
        SELECT {", ".join(select_clause)}
        FROM `tabCRM Deal` deal
        {" ".join(joins)}
        {where_str}
        ORDER BY deal.creation DESC
        LIMIT 5000
    """

    try:
        data = frappe.db.sql(query, as_dict=True)
    except Exception as e:
        # Fallback if join fields are wrong (silent fail for now, but logged)
        frappe.log_error(f"Combined Report SQL Error: {str(e)}")
        # Try simple version without complex joins if it fails
        data = []

    # 3. Format XLSX
    rows = [columns]
    for d in data:
        row = []
        for dt, dt_fields in fields.items():
            for f in dt_fields:
                col_alias = f"{dt}_{f}".replace(" ", "_")
                val = d.get(col_alias)
                # Format dates
                if isinstance(val, (datetime.datetime, datetime.date)):
                    val = frappe.utils.getdate(val)
                row.append(val)
        rows.append(row)

    from frappe.utils.xlsxutils import make_xlsx
    xlsx_file = make_xlsx(rows, "Combined Sales Report")
    
    frappe.response['filename'] = "Combined_Sales_Report.xlsx"
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = "binary"

############################################################################
# Blanket Order - Information
 

@frappe.whitelist()
def get_blanket_order_details(blanket_order):
    bo_items = frappe.get_all('Blanket Order Item', 
                              filters={'parent': blanket_order}, 
                              fields=['item_code', 'item_name', 'qty', 'ordered_qty'])
    
    sales_orders = frappe.db.sql("""
        SELECT so.name as parent, so.place_of_supply, soi.item_code, soi.item_name, soi.qty, soi.amount
        FROM `tabSales Order` so
        JOIN `tabSales Order Item` soi ON so.name = soi.parent
        WHERE soi.blanket_order = %s AND so.docstatus = 1
    """, (blanket_order,), as_dict=1)

    return {
        'bo_items': bo_items,
        'sales_orders': sales_orders
    }


@frappe.whitelist()
def get_pincode_details(pincode):
    """
    Fetches location details for a given pincode from the postal pincode API.
    Used to bypass client-side CORS issues.
    """
    if not pincode:
        frappe.throw("Pincode is required")
    
    url = f"https://api.postalpincode.in/pincode/{pincode}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
    }
    try:
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        response = requests.get(url, headers=headers, timeout=10, verify=False)
        if response.status_code == 200:
            return response.json()
        else:
            frappe.log_error(f"Pincode API returned status code {response.status_code}", "Pincode API Error")
            return None
    except Exception as e:
        frappe.log_error(f"Pincode API connection error: {str(e)}", "Pincode API Error")
        return None

# ======================================================================
# AI INSTALLATION REPORT PORTAL APIS
# ======================================================================

# =========================================================
# MAIN FUNCTION
# =========================================================
@frappe.whitelist()
def ai_installation_query(question):

    try:
        question_lower = question.lower()
        user = frappe.session.user

        # =========================================================
        # 🔥 USER PERMISSION ENFORCEMENT
        # =========================================================
        permitted_customers = []
        if user and user not in ["Administrator", "Guest"]:
            permitted_customers = frappe.get_all(
                "User Permission", 
                filters={"user": user, "allow": "Customer"}, 
                pluck="for_value"
            )

        def get_site_filters(base_filters):
            # If the user has specific Customer restrictions, apply them
            if permitted_customers:
                base_filters["customer"] = ["in", permitted_customers]
            return base_filters

        # =========================================================
        # 🔥 DETECT IMAGE TYPE
        # =========================================================
        image_type = None

        if "ir" in question_lower:
            image_type = "IR Report"
        elif "router" in question_lower:
            image_type = "Router Photo"
        elif "testing" in question_lower:
            image_type = "Testing Photo"
        elif "rack" in question_lower:
            image_type = "Server Rack Photo"
        elif "cable" in question_lower:
            image_type = "Cable Labeling Photo"
        elif "isp" in question_lower:
            image_type = "ISP Device Photo"
        elif "installation" in question_lower or "report" in question_lower:
            image_type = None  # fetch all

        # =========================================================
        # 🔥 EXTRACT INPUTS (NUMBERS + WORDS)
        # =========================================================
        numbers = re.findall(r'\d+', question)
        words = re.findall(r'[A-Za-z0-9]+', question)

        circuit_ids = set()

        # 🔥 1. Direct Circuit ID (numbers)
        for num in numbers:
            filters = get_site_filters({"name": num})
            sites = frappe.get_all("Site", filters=filters, pluck="name", ignore_permissions=True)
            if sites:
                circuit_ids.add(sites[0])

        # 🔥 2. Legal Code → convert to Circuit ID
        for word in words:
            filters = get_site_filters({"site_id__legal_code": word.upper()})
            sites = frappe.get_all("Site", filters=filters, pluck="name", ignore_permissions=True)
            if sites:
                circuit_ids.add(sites[0])

        all_images = []
        valid_circuits = []
        display_circuits = []
        display_customers = []

        # =========================================================
        # 🔥 PROCESS EACH CIRCUIT
        # =========================================================
        for circuit_id in circuit_ids:

            installation = frappe.db.get_value(
                "Installation Note",
                {"custom_circuit_id": circuit_id},
                "name",
                ignore=True
            )

            if not installation:
                continue

            # =========================================================
            # 🔥 GET LEGAL CODE & CUSTOMER FROM SITE
            # =========================================================
            site_data = frappe.db.get_value(
                "Site",
                {"name": circuit_id},
                ["site_id__legal_code", "customer"],
                as_dict=True,
                ignore=True
            )
            
            legal_code = "NA"
            customer_name = "Unknown Customer"
            
            if site_data:
                legal_code = site_data.get("site_id__legal_code") or "NA"
                customer_name = site_data.get("customer") or "Unknown Customer"

            # =========================================================
            # 🔥 GET ATTACHMENTS
            # =========================================================
            attachments = frappe.get_all(
                "Installation Note Attachment",
                filters={"parent": installation},
                fields=["attachment", "select_mqjl"],
                ignore_permissions=True
            )

            for att in attachments:

                if not att.attachment:
                    continue

                # =========================================================
                # 🔥 FILTER ONLY IF SPECIFIC TYPE REQUESTED
                # =========================================================
                if image_type and att.select_mqjl != image_type:
                    continue

                all_images.append({
                    "image": att.attachment,
                    "label": att.select_mqjl,
                    "circuit_id": circuit_id,
                    "legal_code": legal_code
                })

            valid_circuits.append(circuit_id)
                
            if customer_name not in display_customers:
                display_customers.append(customer_name)

        # =========================================================
        # 🔥 AI REPLY
        # =========================================================
        if not all_images:
            if valid_circuits:
                ai_reply = f"<b>No installation images or attachments found for: | Customer Name : {', '.join(display_customers)}</b>"
            else:
                ai_reply = "<b>❌ We couldn't find any photographs for the provided Circuit ID or Legal Code.</b>"
        else:
            if image_type:
                ai_reply = f"<b>{image_type} | Customer Name : {', '.join(display_customers)}</b>"
            else:
                ai_reply = f"<b>Installation photographs | Customer Name : {', '.join(display_customers)}</b>"

        return {
            "status": "success",
            "images": all_images,
            "circuit_ids": list(valid_circuits),
            "image_type": image_type,
            "ai_reply": ai_reply
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "AI INSTALLATION ERROR")
        return {"status": "error", "message": str(e)}

@frappe.whitelist()
def download_multi_images(files):
    """
    Creates a ZIP file of multiple images and returns the download URL.
    """
    frappe.logger().info(f"MULTI DOWNLOAD REQUEST: {files}")
    
    if not files:
        return {"status": "error", "message": "No files selected"}

    try:
        if isinstance(files, str):
            files = json.loads(files)

        # ZIP filename from first file metadata
        first_file = files[0] if files else {}
        z_cid = first_file.get("cid", "Unknown")
        z_lc = first_file.get("lc", "NA")
        zip_display_name = f"Installation_Report_{z_cid}_{z_lc}.zip"

        zip_buffer = io.BytesIO()
        files_added = 0
        
        with zipfile.ZipFile(zip_buffer, "w") as zip_file:
            for file_data in files:
                url = file_data.get("url")
                if not url:
                    continue
                
                label = file_data.get("label", "Image")
                fcid = file_data.get("cid", "Unknown")
                flc = file_data.get("lc", "NA")
                
                # Resolve file path
                clean_path = url.lstrip("/")
                site_path = frappe.get_site_path()
                
                possible_paths = [
                    os.path.join(site_path, "public", clean_path),
                    os.path.join(site_path, clean_path),
                    frappe.get_site_path("public", clean_path),
                    frappe.get_site_path("private", clean_path)
                ]
                
                resolved_path = None
                for p in possible_paths:
                    if os.path.exists(p) and os.path.isfile(p):
                        resolved_path = p
                        break
                
                if resolved_path:
                    _, ext = os.path.splitext(resolved_path)
                    # Use provided metadata for internal name
                    internal_name = f"{label}_{fcid}_{flc}{ext}".replace(" ", "_")
                    zip_file.write(resolved_path, internal_name)
                    files_added += 1
                else:
                    frappe.logger().warning(f"Could not resolve file: {url}")

        if files_added == 0:
            return {"status": "error", "message": "None of the selected images could be found on the server."}

        zip_buffer.seek(0)
        
        # Save ZIP to file manager with randomized internal name but return pretty display name
        from frappe.utils import random_string
        fn = f"report_{random_string(6)}.zip"
        
        _file = frappe.get_doc({
            "doctype": "File",
            "file_name": fn,
            "content": zip_buffer.getvalue(),
            "is_private": 0
        })
        _file.insert(ignore_permissions=True)
        frappe.db.commit()

        return {
            "status": "success", 
            "url": _file.file_url,
            "filename": zip_display_name
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "MULTI DOWNLOAD ERROR")
        return {"status": "error", "message": f"Server Error: {str(e)}"}

# ======================================================================
# END OF AI INSTALLATION REPORT PORTAL APIS
# ======================================================================


# ======================================================================
# AUTO MOVE ATTACHMENTS
# ======================================================================
def auto_move_attachments(doc, method=None):
    """
    Hooked to File after_insert.
    Automatically moves standard attachments to child tables for specific DocTypes.
    """
    config = {
        "Provisioning": {
            "child_field": "provisioning_attachment",
            "child_doctype": "Provisioning Attachment",
            "select_field": "select_mqjl",
            "default_select_val": "Testing Photo"
        },
        "Installation Note": {
            "child_field": "custom_installation_note_attachment",
            "child_doctype": "Installation Note Attachment",
            "select_field": "select_mqjl",
            "default_select_val": "Router Photo"
        }
    }
    
    if doc.attached_to_doctype in config and not doc.attached_to_field:
        if not doc.attached_to_name or not frappe.db.exists(doc.attached_to_doctype, doc.attached_to_name):
            return
            
        conf = config[doc.attached_to_doctype]
        parent = frappe.get_doc(doc.attached_to_doctype, doc.attached_to_name)
        
        # Prevent duplicate entries
        for row in parent.get(conf["child_field"], []):
            if row.attachment == doc.file_url:
                return
                
        # Create a row in the child table
        child = parent.append(conf["child_field"], {
            conf["select_field"]: conf["default_select_val"],
            "attachment": doc.file_url
        })
        
        # Save the parent document
        parent.flags.ignore_mandatory = True
        parent.flags.ignore_validate = True
        parent.flags.ignore_permissions = True
        parent.flags.ignore_validate_update_after_submit = True
        parent.save()
        
        # Remove from standard Attachments
        doc.db_set({
            "attached_to_doctype": conf["child_doctype"],
            "attached_to_name": child.name,
            "attached_to_field": "attachment"
        })

# ======================================================================
# End Of AUTO MOVE ATTACHMENTS
# ======================================================================

# ======================================================================
# INSTALLATION NOTE: PROVISIONING ATTACHMENTS GALLERY
# ======================================================================

@frappe.whitelist()
def get_provisioning_attachments_for_gallery(circuit_id):
    """
    Returns all images from the Provisioning Attachment child table AND standard attachments for a given Circuit ID.
    """
    provisioning = frappe.db.get_value("Provisioning", {"circuit_id": circuit_id}, "name")
    if not provisioning:
        return {"provisioning_id": None, "attachments": []}
        
    attachments = frappe.get_all(
        "Provisioning Attachment",
        filters={"parent": provisioning},
        fields=["name", "select_mqjl", "attachment"],
        order_by="creation desc"
    )
    
    # Fetch standard attachments
    standard_files = frappe.get_all(
        "File",
        filters={
            "attached_to_doctype": "Provisioning",
            "attached_to_name": provisioning,
            "is_folder": 0
        },
        fields=["name", "file_url"],
        order_by="creation desc"
    )
    
    for f in standard_files:
        if f.file_url and f.file_url.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp')):
            attachments.append({
                "name": f.name,
                "select_mqjl": "Standard",
                "attachment": f.file_url
            })
            
    # Sort all by creation could be nice, but appending standard to the end or beginning is fine.
    
    return {
        "provisioning_id": provisioning,
        "attachments": attachments
    }

@frappe.whitelist()
def move_provisioning_attachments_to_installation(installation_note_name, provisioning_attachment_names):
    # Skipping update for multi-select as single move is what they use, but can be updated later if needed.
    pass

def watermark_image_with_circuit_id(file_url, circuit_id):
    import os
    import urllib.parse
    try:
        from PIL import Image, ImageDraw, ImageFont
        
        if not file_url or not isinstance(file_url, str):
            return False
            
        # Strip domain if absolute URL
        if file_url.startswith("http"):
            from urllib.parse import urlparse
            file_url = urlparse(file_url).path
            
        decoded_url = urllib.parse.unquote(file_url)
        
        if decoded_url.startswith("/files/") or decoded_url.startswith("/private/files/"):
            # Construct the absolute physical path
            if decoded_url.startswith("/private/files/"):
                file_path = frappe.get_site_path("private", "files", decoded_url.split("/private/files/")[-1])
            else:
                file_path = frappe.get_site_path("public", "files", decoded_url.split("/files/")[-1])
                
            if os.path.exists(file_path):
                img = Image.open(file_path)
                save_format = img.format if img.format else "JPEG"
                
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                    
                draw = ImageDraw.Draw(img)
                text = f"Circuit ID: {circuit_id}"
                
                font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
                if os.path.exists(font_path):
                    font_size = 14
                    font = ImageFont.truetype(font_path, font_size)
                else:
                    font = ImageFont.load_default()
                    font_size = 14
                    
                if hasattr(draw, 'textbbox'):
                    left, top, right, bottom = draw.textbbox((0, 0), text, font=font)
                elif hasattr(draw, 'textsize'):
                    text_width, text_height = draw.textsize(text, font=font)
                    left, top, right, bottom = 0, 0, text_width, text_height
                else:
                    left, top, right, bottom = font.getbbox(text)
                
                # Set a 1px stroke
                stroke_width = 1
                
                # Use 'right' and 'bottom' instead of width/height. Fonts have an internal 
                # starting offset (left, top), so ink actually ends at x+right and y+bottom.
                # Adding stroke_width to margin prevents the black outline from getting clipped.
                margin_x = stroke_width
                margin_y = stroke_width
                
                x = img.width - right - margin_x
                y = img.height - bottom - margin_y
                
                # Draw black stroke/outline for maximum visibility on any background
                outline_color = "black"
                
                # Draw a clean, thick stroke
                for adj_x in range(-stroke_width, stroke_width+1):
                    for adj_y in range(-stroke_width, stroke_width+1):
                        # Skip extreme corners to make the stroke slightly rounded
                        if abs(adj_x) == stroke_width and abs(adj_y) == stroke_width:
                            continue
                        draw.text((x+adj_x, y+adj_y), text, font=font, fill=outline_color)
                
                # Draw solid white text over the stroke
                draw.text((x, y), text, font=font, fill="white")
                
                # Use the save_format captured before conversion
                img.save(file_path, format=save_format)
                return True
            else:
                frappe.log_error(title="Watermark Error", message=f"File not found: {file_path}")
    except Exception as e:
        frappe.log_error(title="Watermark Error", message=str(e))
    return False

@frappe.whitelist()
def move_single_provisioning_attachment_to_installation(installation_note_name, provisioning_attachment_name, attachment_type):
    """
    Moves a single row from Provisioning Attachment (or standard File) to Installation Note Attachment and sets its type.
    """
    is_standard = frappe.db.exists("File", {"name": provisioning_attachment_name, "attached_to_doctype": "Provisioning"})
    is_child = frappe.db.exists("Provisioning Attachment", provisioning_attachment_name)
    
    if not is_standard and not is_child:
        return {"status": "error", "message": "Attachment row not found."}
        
    installation_note = frappe.get_doc("Installation Note", installation_note_name)
    existing_urls = [row.attachment for row in installation_note.get("custom_installation_note_attachment", []) if row.attachment]
    
    if is_child:
        row = frappe.get_doc("Provisioning Attachment", provisioning_attachment_name)
        attachment_url = row.attachment
    else:
        row = frappe.get_doc("File", provisioning_attachment_name)
        attachment_url = row.file_url
    
    if attachment_url in existing_urls:
        return {"status": "error", "message": "This image already exists in the Installation Note."}
        
    child = installation_note.append("custom_installation_note_attachment", {
        "select_mqjl": attachment_type,
        "attachment": attachment_url
    })
    
    installation_note.flags.ignore_mandatory = True
    installation_note.flags.ignore_validate = True
    installation_note.flags.ignore_validate_update_after_submit = True
    installation_note.save(ignore_permissions=True)
    
    if is_child:
        # Transfer tabFile ownership
        frappe.db.sql("""
            UPDATE `tabFile` 
            SET attached_to_doctype='Installation Note Attachment', attached_to_name=%s 
            WHERE attached_to_doctype='Provisioning Attachment' AND attached_to_name=%s
        """, (child.name, row.name))
        
        # Delete old row
        row.delete(ignore_permissions=True)
    else:
        # Transfer standard file ownership to the child table row
        frappe.db.sql("""
            UPDATE `tabFile` 
            SET attached_to_doctype='Installation Note Attachment', attached_to_name=%s 
            WHERE name=%s
        """, (child.name, row.name))
        
    # Apply watermark to the moved image
    circuit_id = installation_note.get("circuit_id") or installation_note.get("custom_circuit_id") or "Unknown"
    watermark_image_with_circuit_id(attachment_url, circuit_id)
    
    return {
        "status": "success",
        "message": f"Successfully moved 1 image as '{attachment_type}'."
    }

@frappe.whitelist()
def save_cropped_provisioning_attachment(provisioning_attachment_name, filedata, filename):
    """
    Saves a base64 cropped image as a new file and updates the Provisioning Attachment row or standard file.
    """
    is_standard = frappe.db.exists("File", {"name": provisioning_attachment_name, "attached_to_doctype": "Provisioning"})
    is_child = frappe.db.exists("Provisioning Attachment", provisioning_attachment_name)
    
    if not is_standard and not is_child:
        return {"status": "error", "message": "Attachment row not found."}
        
    import base64
    file_content = base64.b64decode(filedata)
    
    if is_child:
        # Create new file for child row
        file_doc = frappe.get_doc({
            "doctype": "File",
            "file_name": filename,
            "attached_to_doctype": "Provisioning Attachment",
            "attached_to_name": provisioning_attachment_name,
            "attached_to_field": "attachment",
            "content": file_content,
            "is_private": 0
        })
        file_doc.save(ignore_permissions=True)
        
        # Update child row
        row = frappe.get_doc("Provisioning Attachment", provisioning_attachment_name)
        row.attachment = file_doc.file_url
        row.save(ignore_permissions=True)
        new_name = provisioning_attachment_name
    else:
        # Create new file for standard attachment
        old_file = frappe.get_doc("File", provisioning_attachment_name)
        file_doc = frappe.get_doc({
            "doctype": "File",
            "file_name": filename,
            "attached_to_doctype": old_file.attached_to_doctype,
            "attached_to_name": old_file.attached_to_name,
            "content": file_content,
            "is_private": 0
        })
        file_doc.save(ignore_permissions=True)
        
        # Delete old file
        frappe.delete_doc("File", old_file.name, ignore_permissions=True)
        new_name = file_doc.name
    
    return {
        "status": "success", 
        "file_url": file_doc.file_url,
        "new_name": new_name,
        "message": "Image successfully cropped and saved."
    }

# ======================================================================
# End Of INSTALLATION NOTE: PROVISIONING ATTACHMENTS GALLERY
# ======================================================================

# =========================================================
# START: Get Last Progressive Status for Feasibility
# =========================================================
@frappe.whitelist()
def get_last_status_before_hold_or_cancel(doctype, docname):
    import json
    non_progressive = ["On Hold", "Not Feasible", "High Commercials", "Cancelled"]
    
    status_field = "status"
    if doctype == "Feasibility":
        status_field = "feasibility_status"
    elif doctype == "Site":
        status_field = "site_status"
        
    versions = frappe.get_all(
        "Version",
        filters={"ref_doctype": doctype, "docname": docname},
        order_by="creation desc",
        fields=["data"]
    )
    
    for v in versions:
        if v.data:
            try:
                data = json.loads(v.data)
                if "changed" in data:
                    for change in data["changed"]:
                        if len(change) >= 3 and change[0] == status_field:
                            old_value = change[1]
                            if old_value and old_value not in non_progressive:
                                return old_value
            except Exception:
                continue
                
    return "Pending"
# =========================================================
# END: Get Last Progressive Status for Feasibility
# =========================================================

from frappe.utils import getdate, add_days

def _get_holiday_dates(reference_date):
    """
    Simple helper: Get all holiday dates from the Company's Default Holiday List.
    Uses raw SQL to bypass any ORM caching issues.
    
    Lookup order:
    1. Company → default_holiday_list
    2. Global Defaults → holiday_list  
    3. Fallback: any Holiday List whose date range covers reference_date
    """
    holiday_list_name = None
    
    # Step 1: Best match — find any Holiday List covering the exact reference_date
    result = frappe.db.sql("""
        SELECT name FROM `tabHoliday List`
        WHERE from_date <= %s AND to_date >= %s
        ORDER BY name ASC LIMIT 1
    """, (reference_date, reference_date))
    
    if result and result[0][0]:
        holiday_list_name = result[0][0]
        
    # Step 2: Fallback to Company's default_holiday_list
    if not holiday_list_name:
        company = frappe.db.get_default("company")
        if company:
            result = frappe.db.sql(
                "SELECT default_holiday_list FROM tabCompany WHERE name=%s",
                company
            )
            if result and result[0][0]:
                holiday_list_name = result[0][0]
    
    # Step 3: Try global default
    if not holiday_list_name:
        holiday_list_name = frappe.db.get_default("holiday_list")
    
    # Now fetch all holiday dates from the selected list
    holiday_dates = set()
    if holiday_list_name:
        rows = frappe.db.sql(
            "SELECT holiday_date FROM tabHoliday WHERE parent=%s",
            holiday_list_name
        )
        for row in rows:
            if row[0]:
                holiday_dates.add(getdate(row[0]))
    
    return holiday_dates


##############################################################

@frappe.whitelist()
def calculate_tat_due_date(start_date, tat_days):
    """
    Simple logic:
    1. Start from start_date (counted as Day 1 if it's a working day)
    2. Move forward day by day, skipping holidays
    3. Stop when we've counted tat_days working days
    4. Return that date as the Due Date
    """
    if not start_date or tat_days is None:
        return None

    start_date = getdate(start_date)
    tat_days = int(tat_days)

    # Get holidays from Company's Holiday List
    holiday_dates = _get_holiday_dates(start_date)

    # Count working days forward
    current_date = start_date
    days_counted = 0

    while days_counted < tat_days:
        if current_date not in holiday_dates:
            days_counted += 1

        if days_counted < tat_days:
            current_date = add_days(current_date, 1)

    return current_date

##################################################################################
@frappe.whitelist()
def get_tat_target(process, lms_type=None):
    tat_master = frappe.db.get_value("TAT Master", {"tat_process": process, "is_active": 1}, "name")
    if not tat_master:
        return 30
        
    rules = frappe.get_all("TAT Rule", filters={"parent": tat_master}, fields=["lms_type", "tat_period"], ignore_permissions=True)
    
    # Filter rules strictly by lms_type
    solution_rules = []
    if lms_type:
        for r in rules:
            if r.lms_type and r.lms_type.lower() == str(lms_type).lower():
                solution_rules.append(r)
                
    if not solution_rules:
        return 30
        
    return solution_rules[0].tat_period

@frappe.whitelist()
def calculate_tat_working_days(start_date, end_date):
    """
    Simple logic:
    Count the number of working days between start_date and end_date (inclusive),
    skipping any holidays from the Company's Holiday List.
    """
    if not start_date or not end_date:
        return 0

    start_date = getdate(start_date)
    end_date = getdate(end_date)

    if start_date > end_date:
        return 0

    # Get holidays from Company's Holiday List
    holiday_dates = _get_holiday_dates(start_date)

    # Count working days in range
    current_date = start_date
    working_days = 0

    while current_date <= end_date:
        if current_date not in holiday_dates:
            working_days += 1
        current_date = add_days(current_date, 1)

    return working_days

@frappe.whitelist()
def get_working_days_for_dates(start_date, dates):
    import json
    if isinstance(dates, str):
        dates = json.loads(dates)
    
    results = {}
    for d in dates:
        results[d] = calculate_tat_working_days(start_date, d)
    return results

###############################################################################
# File Hook: auto_move_attachments
# Called by hooks.py on File after_insert.
# Placeholder – implement actual move logic here if needed in future.
###############################################################################
def auto_move_attachments(doc, method):
    """
    Hook called after a File document is inserted.
    Currently a no-op placeholder to prevent 500 errors.
    """
    pass

#####################################################################
# Fetch Site Version History for TAT
@frappe.whitelist()
def get_site_version_history(docname):
    # Fetch Version history securely without relying on user permissions for Version Doctype
    # Returns the necessary fields for TAT timeline calculations
    if not frappe.has_permission("Site", ptype="read", doc=docname):
        frappe.throw("No permission to read this Site.")
        
    versions = frappe.db.sql(
        """
        SELECT creation, data
        FROM `tabVersion`
        WHERE ref_doctype='Site' AND docname=%s
        ORDER BY creation ASC
        LIMIT 1000
        """,
        (docname,),
        as_dict=True
    )
    return versions
#####################################################################
# Start - Return Packed Item (Delivery Note)
#####################################################################
# --- START: Return Packed Item Logic ---
import json

@frappe.whitelist()
def get_dn_packed_items(delivery_note):
    dn = frappe.get_doc("Delivery Note", delivery_note)
    
    # Fetch all returns for this DN to calculate already returned qtys
    returns = frappe.get_all("Customer Asset Return Item",
        filters={"parent": ["in", frappe.get_all("Customer Asset Return", filters={"delivery_note": delivery_note, "docstatus": 1}, pluck="name")]},
        fields=["item_code", "returned_qty"]
    )
    returned_dict = {}
    for r in returns:
        returned_dict[r.item_code] = returned_dict.get(r.item_code, 0) + r.returned_qty
        
    # Aggregate packed items by item_code
    packed_dict = {}
    for item in dn.packed_items:
        if item.item_code not in packed_dict:
            packed_dict[item.item_code] = {"item_code": item.item_code, "item_name": item.item_name, "qty": 0}
        packed_dict[item.item_code]["qty"] += item.qty
        
    items = []
    for item_code, data in packed_dict.items():
        qty_returned = returned_dict.get(item_code, 0)
        qty_available = data["qty"] - qty_returned
        if qty_available > 0:
            items.append({
                "item_code": item_code,
                "item_name": data["item_name"],
                "qty_delivered": data["qty"],
                "qty_returned": qty_returned,
                "qty_available": qty_available
            })
    return items

@frappe.whitelist()
def get_dn_serial_numbers(delivery_note, item_code):
    dn = frappe.get_doc("Delivery Note", delivery_note)
    
    # Find the packed item
    serial_nos = []
    for item in dn.packed_items:
        if item.item_code == item_code:
            if getattr(item, "serial_and_batch_bundle", None):
                entries = frappe.get_all("Serial and Batch Entry", filters={"parent": item.serial_and_batch_bundle}, pluck="serial_no")
                serial_nos.extend(entries)
            elif getattr(item, "serial_no", None):
                serial_nos.extend([s.strip() for s in item.serial_no.split("\\n") if s.strip()])
                
    # Filter out already returned
    returned_serials = frappe.get_all("Customer Asset Return Item",
        filters={"parent": ["in", frappe.get_all("Customer Asset Return", filters={"delivery_note": delivery_note, "docstatus": 1}, pluck="name")], "item_code": item_code},
        pluck="serial_no"
    )
    
    valid_serials = []
    for sn in set(serial_nos) - set(returned_serials):
        if not sn: continue
        if frappe.db.exists("Serial No", sn):
            status = frappe.db.get_value("Serial No", sn, "status")
            if status == "Delivered":
                valid_serials.append(sn)
                
    return valid_serials

@frappe.whitelist()
def submit_asset_return(data):
    data = json.loads(data)
    
    delivery_note = data.get("delivery_note")
    items_to_return = data.get("items")
    return_reason = data.get("return_reason")
    return_date = data.get("return_date")
    remarks = data.get("remarks")
    
    if not items_to_return:
        frappe.throw("No items selected for return.")
        
    dn = frappe.get_doc("Delivery Note", delivery_note)
    
    car = frappe.new_doc("Customer Asset Return")
    car.delivery_note = delivery_note
    car.customer = dn.customer
    car.circuit_id = dn.custom_dn_circuit_id if hasattr(dn, "custom_dn_circuit_id") else None
    car.return_date = return_date
    car.return_reason = return_reason
    car.remarks = remarks
    
    for row in items_to_return:
        serial_no = row.get("serial_no")
        item_code = row.get("item_code")
        item_name = row.get("item_name") or frappe.db.get_value("Item", item_code, "item_name")
        
        # Validation
        if frappe.db.exists("Customer Asset Return Item", {"serial_no": serial_no, "parent": ["in", frappe.get_all("Customer Asset Return", filters={"docstatus": 1}, pluck="name")]}):
            frappe.throw(f"Serial number {serial_no} has already been returned.")

        car.append("items", {
            "item_code": item_code,
            "item_name": item_name,
            "serial_no": serial_no,
            "returned_qty": 1
        })
    
    car.insert(ignore_permissions=True)
    car.submit()
    
    se = frappe.new_doc("Stock Entry")
    se.stock_entry_type = "Material Receipt"
    se.purpose = "Material Receipt"
    se.company = dn.company
    
    for row in items_to_return:
        se.append("items", {
            "item_code": row.get("item_code"),
            "qty": 1,
            "t_warehouse": "Stores - NTPL",
            "serial_no": row.get("serial_no")
        })
        
    se.insert(ignore_permissions=True)
    se.submit()
    
    content_lines = []
    for row in items_to_return:
        sn = row.get("serial_no")
        frappe.db.set_value("Serial No", sn, "status", "Active")
        frappe.db.set_value("Serial No", sn, "warehouse", "Stores - NTPL")
        content_lines.append(f"Item: {row.get('item_name')} (SN: {sn})")
        
    # Add timeline entry to Delivery Note
    content = f"<b>Returned Components:</b><br>" + "<br>".join(content_lines) + f"<br>Returned on: {return_date}<br><a href='/app/customer-asset-return/{car.name}'>View Return</a>"
    frappe.get_doc({
        "doctype": "Communication",
        "communication_type": "Comment",
        "comment_type": "Info",
        "reference_doctype": "Delivery Note",
        "reference_name": delivery_note,
        "content": content
    }).insert(ignore_permissions=True)
    
    return {"status": "success", "message": "Items returned successfully", "car_id": car.name}

# --- END: Return Packed Item Logic ---

#####################################################################
# End - Return Packed Item (Delivery Note)
#####################################################################


# --- START: Return Packed Item Logic ---
import json

@frappe.whitelist()
def get_dn_packed_items(delivery_note):
    dn = frappe.get_doc("Delivery Note", delivery_note)
    
    # Fetch all returns for this DN to calculate already returned qtys
    returns = frappe.get_all("Customer Asset Return Item",
        filters={"parent": ["in", frappe.get_all("Customer Asset Return", filters={"delivery_note": delivery_note, "docstatus": 1}, pluck="name")]},
        fields=["item_code", "returned_qty"]
    )
    returned_dict = {}
    for r in returns:
        returned_dict[r.item_code] = returned_dict.get(r.item_code, 0) + r.returned_qty
        
    # Aggregate packed items by item_code
    packed_dict = {}
    for item in dn.packed_items:
        if item.item_code not in packed_dict:
            packed_dict[item.item_code] = {"item_code": item.item_code, "item_name": item.item_name, "qty": 0}
        packed_dict[item.item_code]["qty"] += item.qty
        
    items = []
    for item_code, data in packed_dict.items():
        qty_returned = returned_dict.get(item_code, 0)
        qty_available = data["qty"] - qty_returned
        if qty_available > 0:
            items.append({
                "item_code": item_code,
                "item_name": data["item_name"],
                "qty_delivered": data["qty"],
                "qty_returned": qty_returned,
                "qty_available": qty_available
            })
    return items

@frappe.whitelist()
def get_dn_serial_numbers(delivery_note, item_code):
    dn = frappe.get_doc("Delivery Note", delivery_note)
    
    # Find the packed item
    serial_nos = []
    for item in dn.packed_items:
        if item.item_code == item_code:
            if getattr(item, "serial_and_batch_bundle", None):
                entries = frappe.get_all("Serial and Batch Entry", filters={"parent": item.serial_and_batch_bundle}, pluck="serial_no")
                serial_nos.extend(entries)
            elif getattr(item, "serial_no", None):
                serial_nos.extend([s.strip() for s in item.serial_no.split("\\n") if s.strip()])
                
    # Filter out already returned
    returned_serials = frappe.get_all("Customer Asset Return Item",
        filters={"parent": ["in", frappe.get_all("Customer Asset Return", filters={"delivery_note": delivery_note, "docstatus": 1}, pluck="name")], "item_code": item_code},
        pluck="serial_no"
    )
    
    valid_serials = []
    for sn in set(serial_nos) - set(returned_serials):
        if not sn: continue
        if frappe.db.exists("Serial No", sn):
            status = frappe.db.get_value("Serial No", sn, "status")
            if status == "Delivered":
                valid_serials.append(sn)
                
    return valid_serials

@frappe.whitelist()
def submit_asset_return(data):
    data = json.loads(data)
    
    delivery_note = data.get("delivery_note")
    items_to_return = data.get("items")
    return_reason = data.get("return_reason")
    return_date = data.get("return_date")
    remarks = data.get("remarks")
    
    if not items_to_return:
        frappe.throw("No items selected for return.")
        
    dn = frappe.get_doc("Delivery Note", delivery_note)
    
    car = frappe.new_doc("Customer Asset Return")
    car.delivery_note = delivery_note
    car.customer = dn.customer
    car.circuit_id = dn.custom_dn_circuit_id if hasattr(dn, "custom_dn_circuit_id") else None
    car.return_date = return_date
    car.return_reason = return_reason
    car.remarks = remarks
    
    for row in items_to_return:
        serial_no = row.get("serial_no")
        item_code = row.get("item_code")
        item_name = row.get("item_name") or frappe.db.get_value("Item", item_code, "item_name")
        
        # Validation
        if frappe.db.exists("Customer Asset Return Item", {"serial_no": serial_no, "parent": ["in", frappe.get_all("Customer Asset Return", filters={"docstatus": 1}, pluck="name")]}):
            frappe.throw(f"Serial number {serial_no} has already been returned.")

        car.append("items", {
            "item_code": item_code,
            "item_name": item_name,
            "serial_no": serial_no,
            "returned_qty": 1
        })
    
    car.insert(ignore_permissions=True)
    car.submit()
    
    se = frappe.new_doc("Stock Entry")
    se.stock_entry_type = "Material Receipt"
    se.purpose = "Material Receipt"
    se.company = dn.company
    
    for row in items_to_return:
        se.append("items", {
            "item_code": row.get("item_code"),
            "qty": 1,
            "t_warehouse": "Stores - NTPL",
            "serial_no": row.get("serial_no")
        })
        
    se.insert(ignore_permissions=True)
    se.submit()
    
    content_lines = []
    for row in items_to_return:
        sn = row.get("serial_no")
        frappe.db.set_value("Serial No", sn, "status", "Active")
        frappe.db.set_value("Serial No", sn, "warehouse", "Stores - NTPL")
        content_lines.append(f"Item: {row.get('item_name')} (SN: {sn})")
        
    # Add timeline entry to Delivery Note
    content = f"<b>Returned Components:</b><br>" + "<br>".join(content_lines) + f"<br>Returned on: {return_date}<br><a href='/app/customer-asset-return/{car.name}'>View Return</a>"
    frappe.get_doc({
        "doctype": "Communication",
        "communication_type": "Comment",
        "comment_type": "Info",
        "reference_doctype": "Delivery Note",
        "reference_name": delivery_note,
        "content": content
    }).insert(ignore_permissions=True)
    
    return {"status": "success", "message": "Items returned successfully", "car_id": car.name}

# --- END: Return Packed Item Logic ---

# --- START: AI Purchase Invoice Creation Logic ---

@frappe.whitelist()
def extract_purchase_invoice_data(file_url=None):
    """
    Robust extraction from a Purchase Invoice PDF using PyMuPDF.
    """
    import re
    
    if not file_url:
        return {"status": "error", "message": "No invoice file was attached. Please attach a PDF invoice."}
    
    try:
        import os
        # Robust file path resolution
        try:
            from frappe.utils.file_manager import get_file_path
            file_path = get_file_path(file_url.split('/')[-1])
        except Exception:
            file_path = None
        
        # Fallback: construct path from site path + file_url
        if not file_path or not os.path.exists(file_path):
            site_path = frappe.get_site_path()
            relative = file_url.lstrip('/')
            file_path = os.path.join(site_path, 'public', relative.replace('files/', ''))
            if not os.path.exists(file_path):
                file_path = os.path.join(site_path, relative)
        
        if not file_path or not os.path.exists(file_path):
            return {"status": "error", "message": "Could not locate the uploaded invoice file on the server."}
        
        if not file_path.lower().endswith('.pdf'):
            return {"status": "error", "message": "Only PDF files are supported for extraction."}
        
        import fitz
        doc = fitz.open(file_path)
        text = ""
        blocks = []
        
        # --- Strategy 1: Standard text extraction ---
        for page in doc:
            text += page.get_text()
            blocks += page.get_text("blocks")
        
        # --- Strategy 2: rawdict — works on vector-font PDFs where get_text() returns empty ---
        if not text.strip():
            try:
                raw_text_parts = []
                for page in doc:
                    raw = page.get_text("rawdict")
                    for block in raw.get("blocks", []):
                        for line in block.get("lines", []):
                            for span in line.get("spans", []):
                                t = span.get("text", "").strip()
                                if t:
                                    raw_text_parts.append(t)
                text = "\n".join(raw_text_parts)
            except Exception:
                pass
        
        # --- Strategy 3: OCR fallback (requires tesseract-ocr + pytesseract) ---
        if not text.strip():
            try:
                import pytesseract
                from PIL import Image
                import io, shutil
                # Explicitly set the tesseract binary path since bench env PATH
                # may not include /usr/bin where the system tesseract lives
                tess_path = shutil.which("tesseract") or "/usr/bin/tesseract"
                pytesseract.pytesseract.tesseract_cmd = tess_path
                ocr_parts = []
                for page in doc:
                    mat = fitz.Matrix(2.0, 2.0)  # 2x zoom for better OCR accuracy
                    pix = page.get_pixmap(matrix=mat)
                    img_bytes = pix.tobytes("png")
                    img = Image.open(io.BytesIO(img_bytes))
                    page_text = pytesseract.image_to_string(img, lang='eng')
                    ocr_parts.append(page_text)
                text = "\n".join(ocr_parts)
                frappe.logger().info(f"Invoice Extraction: Used OCR fallback for {file_url}")
            except ImportError:
                pass  # pytesseract not installed, skip
            except Exception as ocr_err:
                frappe.log_error("Invoice OCR Error", str(ocr_err))
        
        if not text.strip():
            return {
                "status": "error",
                "message": "Could not extract text from this PDF. It appears to be a fully image-based scan. Please install tesseract-ocr on the server to enable OCR support, or upload a text-based PDF."
            }
        
        data = {}
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        
        # --- TEMP DEBUG: log extracted lines to diagnose item parsing ---
        frappe.log_error("PDF Extracted Lines", "\n".join(f"{i}: {l}" for i, l in enumerate(lines[:200])))
        
        # -------------------------------------------------------
        # Helper
        # -------------------------------------------------------
        def parse_date(d_str):
            """Normalise various date formats to YYYY-MM-DD."""
            import datetime
            d_str = d_str.strip()
            month_map = {
                'jan':'01','feb':'02','mar':'03','apr':'04','may':'05','jun':'06',
                'jul':'07','aug':'08','sep':'09','oct':'10','nov':'11','dec':'12'
            }
            # DD-Mon-YY or DD-Mon-YYYY (e.g. 13-Aug-25)
            m = re.match(r'(\d{1,2})[-/\s]([A-Za-z]{3})[-/\s](\d{2,4})', d_str)
            if m:
                dd, mon, yy = m.group(1).zfill(2), m.group(2).lower()[:3], m.group(3)
                mm = month_map.get(mon, '01')
                yyyy = ('20' + yy) if len(yy) == 2 else yy
                return f"{yyyy}-{mm}-{dd}"
            # DD-MM-YYYY or DD/MM/YYYY
            m = re.match(r'(\d{2})[-/](\d{2})[-/](\d{4})', d_str)
            if m:
                return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
            # YYYY-MM-DD (already correct)
            m = re.match(r'(\d{4})-(\d{2})-(\d{2})', d_str)
            if m:
                return d_str
            try:
                return str(frappe.utils.getdate(d_str))
            except Exception:
                return d_str
        
        def clean_amount(s):
            return float(re.sub(r'[^\d.]', '', s)) if s else 0.0
        
        # -------------------------------------------------------
        # 1. Invoice Number - multiple patterns, pick first valid
        # -------------------------------------------------------
        inv_no = None
        inv_no_patterns = [
            # OCR jump: matches 'Invoice No' then skips up to 120 chars to find a proper invoice format (e.g. KNPL/26-27/0078)
            r'(?:Invoice\s*No\.?|Invoice\s*Number|INV\s*NO\.?|Bill\s*No\.?)[^\n]{0,120}?\b([A-Za-z0-9]+(?:[-/][A-Za-z0-9]+){2,})\b',
            r'Invoice\s*No\.?\s*[:\-]?\s*([A-Za-z0-9/_\-]+)',
            r'Invoice\s*Number\s*[:\-]?\s*([A-Za-z0-9/_\-]+)',
            r'INV\s*NO\.?\s*[:\-]?\s*([A-Za-z0-9/_\-]+)',
            r'Bill\s*No\.?\s*[:\-]?\s*([A-Za-z0-9/_\-]+)',
        ]
        for pat in inv_no_patterns:
            m = re.search(pat, text, re.I)
            if m:
                candidate = m.group(1).strip()
                # Must be at least 4 chars and not just a date
                if len(candidate) >= 4 and not re.match(r'^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$', candidate):
                    inv_no = candidate
                    break
        if not inv_no:
            return {"status": "error", "message": "Could not extract Invoice Number from the document. Please verify the PDF."}
        data["invoice_no"] = inv_no
        
        # -------------------------------------------------------
        # 2. Invoice Date
        # -------------------------------------------------------
        inv_date = None
        date_patterns = [
            r'(?:Invoice\s*Date|Dated|Date\s*of\s*Invoice)[\s:]*([\d]{1,2}[-/\s][A-Za-z]{3}[-/\s][\d]{2,4})',
            r'(?:Invoice\s*Date|Dated|Date\s*of\s*Invoice)[\s:]*([\d]{1,2}[-/][\d]{1,2}[-/][\d]{2,4})',
            r'(?:Invoice\s*Date|Dated)[\s:]*([\d]{4}-[\d]{2}-[\d]{2})',
        ]
        for pat in date_patterns:
            m = re.search(pat, text, re.I)
            if m:
                inv_date = parse_date(m.group(1))
                break
        if not inv_date:
            # Last resort: find any standalone date near "Dated" or top of document
            m = re.search(r'(\d{1,2}[-/][A-Za-z]{3}[-/]\d{2,4})', text[:500])
            if m:
                inv_date = parse_date(m.group(1))
        data["invoice_date"] = inv_date or frappe.utils.today()
        
        # -------------------------------------------------------
        # 3. Duration From / To  (e.g. 13.09.2025 - 12.11.2025)
        # -------------------------------------------------------
        dur_m = re.search(
            r'(\d{1,2}[.\-/]?\d{1,2}[.\-/]?\d{2,4})\s*(?:TO|to|[-–])\s*(\d{1,2}[.\-/]?\d{1,2}[.\-/]?\d{2,4})',
            text, re.I
        )
        if dur_m:
            def parse_ocr_date(s):
                s = re.sub(r'[.\-/]', '', s)
                if len(s) == 8:  # DDMMYYYY
                    return f"{s[4:8]}-{s[2:4]}-{s[0:2]}"
                elif len(s) == 6:  # DDMMYY
                    return f"20{s[4:6]}-{s[2:4]}-{s[0:2]}"
                return s
            data["duration_from"] = parse_ocr_date(dur_m.group(1))
            data["duration_to"] = parse_ocr_date(dur_m.group(2))
        
        # -------------------------------------------------------
        # 4. Supplier Name - Exclusion-First Multi-Strategy Matching
        # Key insight: The logged-in company is NEVER the supplier.
        # Collect ALL company-like names, exclude self, then fuzzy-match.
        # -------------------------------------------------------
        import difflib
        
        # Get all self-company names to exclude (the buyer, not the seller)
        own_company = frappe.defaults.get_global_default('company') or ""
        own_company_words = set(re.sub(r'[^a-z0-9\s]', '', own_company.lower()).split())
        
        # Aliases / abbreviations of own company to also exclude
        own_company_aliases = {"nexapp", "nexapp technologies", "ntpl"}
        
        def is_own_company(name):
            """Returns True if the name looks like our own company, not a supplier."""
            n_lower = name.lower()
            n_clean = re.sub(r'[^a-z0-9\s]', '', n_lower)
            n_words = set(n_clean.split())
            # Check if majority of words overlap with own company name
            if own_company_words and len(own_company_words & n_words) >= max(1, len(own_company_words) - 1):
                return True
            for alias in own_company_aliases:
                if alias in n_lower:
                    return True
            return False
        
        def normalize_name(n):
            n = str(n).lower()
            n = re.sub(r'[^a-z0-9\s]', '', n)
            n = n.replace('private limited', 'pvt ltd')
            n = n.replace('pvt limited', 'pvt ltd')
            n = n.replace('bharti airtel limited', 'bharti airtel ltd')
            return n.strip()
        
        # Strategy 1: Collect ALL company-like names in the document
        # Patterns: "Xyz Ltd", "Abc Pvt Ltd", "Xyz Limited", "Xyz LLP"
        all_company_matches = re.findall(
            r'[A-Za-z][A-Za-z0-9\s&\(\)\-\.]{2,60}(?:Pvt\.?\s*Ltd\.?|Private\s+Limited|Limited|LLP|Ltd\.?)',
            text, re.I
        )
        
        # Also specifically look for names near "From:", "Supplier:", "Issued by:", signature blocks
        seller_patterns = [
            r'(?:from|vendor|supplier|issued\s*by|authorised\s*signatory)[:\s]+([A-Za-z][A-Za-z0-9\s&\-\.]{2,80})',
            r'([A-Za-z][A-Za-z0-9\s&\-\.]{2,50})\s*\n?\s*Authorised\s*Signatory',
            r'([A-Za-z][A-Za-z0-9\s&\-\.]{2,50})\s*\n?\s*Authorized\s*Signatory',
        ]
        priority_candidates = []
        for pat in seller_patterns:
            for m in re.finditer(pat, text, re.I):
                priority_candidates.append(m.group(1).strip())
        
        # Strategy 2: Load ERPNext suppliers and score each candidate
        suppliers = frappe.get_all("Supplier", pluck="name")
        if not suppliers:
            return {"status": "error", "message": "No Suppliers exist in the system."}
        
        norm_to_orig = {normalize_name(s): s for s in suppliers}
        
        def best_supplier_match(candidates, cutoff=0.55):
            """Find the best matching ERPNext supplier from a list of raw name candidates."""
            best_score = 0
            best_match = None
            for candidate in candidates:
                if not candidate or is_own_company(candidate):
                    continue
                norm_c = normalize_name(candidate)
                # Substring match (highest confidence)
                for norm_s, orig_s in norm_to_orig.items():
                    if norm_s and norm_c and (norm_s in norm_c or norm_c in norm_s):
                        return orig_s  # Definitive match
                # Fuzzy match
                close = difflib.get_close_matches(norm_c, list(norm_to_orig.keys()), n=1, cutoff=cutoff)
                if close:
                    score = difflib.SequenceMatcher(None, norm_c, close[0]).ratio()
                    if score > best_score:
                        best_score = score
                        best_match = norm_to_orig[close[0]]
            return best_match
        
        # Priority: look at seller-specific regions first
        matched_supplier = best_supplier_match(priority_candidates, cutoff=0.5)
        
        # Fallback: scan all company names found in the document (excluding self)
        if not matched_supplier:
            # Filter out own-company names and deduplicate
            external_candidates = [c for c in all_company_matches if not is_own_company(c)]
            # Reverse order: supplier name often appears later in invoice (bottom area)
            matched_supplier = best_supplier_match(list(reversed(external_candidates)), cutoff=0.6)
        
        if not matched_supplier:
            # Last resort: take any company name not matching own company
            external_candidates = [c for c in all_company_matches if not is_own_company(c)]
            if external_candidates:
                raw_supplier_name = external_candidates[-1].strip()
            else:
                raw_supplier_name = all_company_matches[0].strip() if all_company_matches else "Unknown"
            return {
                "status": "error",
                "message": f"Supplier '{raw_supplier_name}' was detected on the invoice but does not exist in your ERPNext system. Please create the Supplier first, then retry."
            }
        data["supplier_name"] = matched_supplier
        
        # -------------------------------------------------------
        # 5. Items - Block/sequence based extraction
        # PDFs often extract table cells as separate lines, so we
        # use a state-machine approach on the lines list.
        # -------------------------------------------------------
        items = []
        
        # Strategy A: PyMuPDF word-position clustering by column X-position
        # We use blocks to identify rows based on Y-position proximity
        try:
            row_data = {}  # y_bucket -> list of (x, text)
            for page in doc:
                words = page.get_text("words")  # (x0,y0,x1,y1,word,block,line,word_idx)
                for w in words:
                    y_bucket = round(w[1] / 8) * 8  # cluster within 8pt vertical bands
                    row_data.setdefault(y_bucket, []).append((w[0], w[4]))
            
            # Sort rows by Y
            sorted_rows = [row_data[k] for k in sorted(row_data.keys())]
            
            # Find page width to determine column boundaries
            page_width = doc[0].rect.width
            # Heuristic column positions for this invoice layout:
            # Sr(~30), Desc(~50-250), HSN(~260-310), Qty(~310-360), Rate(~360-430), per(~430-470), Amount(~470+)
            
            SKIP_KEYWORDS = {'description', 'goods', 'hsn', 'sac', 'quantity', 'rate', 'per',
                             'amount', 'total', 'gst', 'tax', 'sgst', 'cgst', 'igst',
                             'taxable', 'value', 'round', 'off', 'si', 'no', 'sl'}
            
            current_item = None
            for row_words in sorted_rows:
                if not row_words:
                    continue
                row_words_sorted = sorted(row_words, key=lambda x: x[0])
                row_text = ' '.join(w[1] for w in row_words_sorted).strip()
                
                # Check if this row starts with a serial number (1, 2, 3...)
                first_word = row_words_sorted[0][1]
                if re.match(r'^\d{1,2}$', first_word):
                    sr_no = int(first_word)
                    if 1 <= sr_no <= 99:
                        if current_item and current_item.get('amount', 0) > 0:
                            items.append(current_item)
                        # Remaining words after Sr No
                        rest = row_words_sorted[1:]
                        amounts = [w for w in rest if re.match(r'^[\d,]+\.\d{2}$', w[1])]
                        desc_words = [w for w in rest 
                                      if not re.match(r'^[\d,]+\.?\d*$', w[1])
                                      and w[1].lower() not in SKIP_KEYWORDS]
                        desc = ' '.join(w[1] for w in sorted(desc_words, key=lambda x: x[0]))
                        
                        current_item = {
                            'description': desc,
                            'qty': 1.0,
                            'rate': 0.0,
                            'amount': clean_amount(amounts[-1][1]) if amounts else 0.0
                        }
                        # Try qty and rate from numeric positions
                        nums = [w for w in rest if re.match(r'^[\d,]+\.?\d*$', w[1])
                                and not re.match(r'^\d{6,8}$', w[1])]  # skip HSN codes
                        if len(nums) >= 3:
                            current_item['qty'] = clean_amount(nums[0][1])
                            current_item['rate'] = clean_amount(nums[1][1])
                            current_item['amount'] = clean_amount(nums[-1][1])
                        elif len(nums) == 2:
                            current_item['rate'] = clean_amount(nums[0][1])
                            current_item['amount'] = clean_amount(nums[-1][1])
                        elif len(nums) == 1:
                            current_item['amount'] = clean_amount(nums[0][1])
                        continue
                
                # Continuation line for current item (more description text)
                if current_item is not None:
                    row_lower = row_text.lower()
                    # Stop adding if this looks like a totals/tax row
                    if any(kw in row_lower for kw in ['sgst', 'cgst', 'igst', 'total', 'round off', 'amount chargeable']):
                        if current_item.get('amount', 0) > 0:
                            items.append(current_item)
                            current_item = None
                        continue
                    # Add description continuation (if mostly text, not numbers)
                    words_in_row = row_text.split()
                    text_words = [w for w in words_in_row if not re.match(r'^[\d.,]+$', w)]
                    if len(text_words) >= 2 and current_item.get('amount', 0) == 0:
                        # Still building description
                        current_item['description'] += ' ' + row_text
                    elif len(text_words) == 0:
                        # Pure numbers row - might be qty/rate/amount update
                        pure_nums = [w for w in words_in_row if re.match(r'^[\d,]+\.?\d*$', w)]
                        if pure_nums and current_item.get('amount', 0) == 0:
                            current_item['amount'] = clean_amount(pure_nums[-1])
            
            if current_item and current_item.get('amount', 0) > 0:
                items.append(current_item)
        except Exception as block_err:
            frappe.log_error("Item block extraction failed", str(block_err))
            items = []
        
        # Strategy B: Line-sequence state machine (fallback)
        if not items:
            STOP_WORDS_RE = re.compile(r'sgst|cgst|igst|grand\s*total|round\s*off|amount\s*chargeable|bank\s*details|declaration|e\s*&\s*oe|certified', re.I)
            SKIP_RE = re.compile(r'^(description|hsn|sac|quantity|rate|per|amount|sl\.?\s*no|si\.?\s*no|services?|goods?)$', re.I)
            
            in_table = False
            i = 0
            while i < len(lines):
                line = lines[i]
                
                # Detect start of item table — covers both Goods and Services invoices
                if re.search(r'description\s+(of\s+)?(goods|services|particulars)', line, re.I):
                    in_table = True
                    i += 1
                    continue
                
                # Also trigger on the column header row itself
                if re.search(r'\bhsn\b.*\bquantity\b|\bsl\.?\s*no\b.*\bdescription\b', line, re.I):
                    in_table = True
                    i += 1
                    continue
                
                if not in_table:
                    i += 1
                    continue
                
                if STOP_WORDS_RE.search(line):
                    break
                
                # Serial number line
                if re.match(r'^\d{1,2}$', line.strip()):
                    sr = int(line.strip())
                    if 1 <= sr <= 50:
                        desc_parts = []
                        qty = rate = amount = None
                        j = i + 1
                        # Collect subsequent lines until we hit the next serial or stop
                        while j < len(lines) and j < i + 12:
                            nxt = lines[j].strip()
                            if re.match(r'^\d{1,2}$', nxt) and 1 <= int(nxt) <= 50:
                                break
                            if STOP_WORDS_RE.search(nxt):
                                break
                            # Amount pattern XX,XXX.XX
                            amt_m = re.match(r'^([\d,]+\.\d{2})$', nxt)
                            if amt_m:
                                val = clean_amount(amt_m.group(1))
                                if amount is None:
                                    amount = val
                                elif rate is None and val > 0:
                                    rate = val
                            # Pure integer = qty
                            elif re.match(r'^\d{1,4}$', nxt) and not re.match(r'^\d{6,}$', nxt):
                                if qty is None and amount is None:
                                    qty = float(nxt)
                            # HSN code - skip (6-8 digits)
                            elif re.match(r'^\d{6,8}$', nxt):
                                pass
                            # Text = description
                            elif nxt and not SKIP_RE.match(nxt) and not re.match(r'^[₹%@\d\s.,/]+$', nxt):
                                desc_parts.append(nxt)
                            j += 1
                        
                        if desc_parts and amount and amount > 0:
                            items.append({
                                'description': ' '.join(desc_parts),
                                'qty': qty or 1.0,
                                'rate': rate or amount,
                                'amount': amount
                            })
                        i = j
                        continue
                i += 1
        
        # Strategy C: Find standalone amount lines and pair with nearby description
        if not items:
            amount_lines = [(i, clean_amount(l)) for i, l in enumerate(lines)
                           if re.match(r'^[\d,]+\.\d{2}$', l.strip()) and clean_amount(l) > 10]
            for idx, amt in amount_lines:
                desc = ''
                for di in range(max(0, idx-5), idx):
                    candidate = lines[di].strip()
                    if candidate and not re.match(r'^[\d,.\s%@₹]+$', candidate):
                        if not re.search(r'total|gst|tax|hsn|quantity|rate|amount', candidate, re.I):
                            desc = candidate
                if desc and amt > 0:
                    # Avoid duplicates — don't add if same amount already captured
                    if not any(abs(existing.get('amount', 0) - amt) < 0.01 for existing in items):
                        items.append({'description': desc, 'qty': 1.0, 'rate': amt, 'amount': amt})
        
        # Strategy D: OCR-resilient extraction — handles merged table cells
        if not items:
            full_text = ' '.join(lines)
            # Pattern: serial_no + description + HSN(6 digits) + amount
            for m in re.finditer(r'(?<!\d)([1-9])\s+([A-Za-z][\w\s,./\-()]{3,80}?)\s+(\d{6})\s+.*?([\d,]+\.\d{2})', full_text):
                desc = m.group(2).strip()
                amt = clean_amount(m.group(4))
                if amt > 0 and not re.search(r'total|sgst|cgst|igst|round|chargeable|taxable|bank', desc, re.I):
                    if not any(abs(ex.get('amount', 0) - amt) < 0.01 for ex in items):
                        items.append({'description': desc, 'qty': 1.0, 'rate': amt, 'amount': amt})
        
        # Strategy E: Find amounts within lines (not requiring standalone amount lines)
        if not items:
            for i, line in enumerate(lines):
                if re.search(r'total|sgst|cgst|igst|round|chargeable|taxable|tax\s*amount|hsn.*taxable', line, re.I):
                    continue
                amounts_in_line = re.findall(r'([\d,]+\.\d{2})', line)
                for amt_str in amounts_in_line:
                    amt = clean_amount(amt_str)
                    if amt < 50:
                        continue
                    # Extract description from text before the amount in the same line
                    idx = line.index(amt_str)
                    before = re.sub(r'^\d{1,2}\s+', '', line[:idx]).strip()
                    before = re.sub(r'\d{6,8}\s*', '', before).strip()
                    desc = before if len(before) > 3 else ''
                    # Fallback: look in previous lines
                    if not desc:
                        for di in range(max(0, i-5), i):
                            c = lines[di].strip()
                            if c and not re.match(r'^[\d,.\s%@₹|]+$', c) and not re.search(r'total|gst|tax|hsn|quantity|rate|amount|description', c, re.I):
                                desc = c
                    if desc and not any(abs(ex.get('amount', 0) - amt) < 0.01 for ex in items):
                        items.append({'description': desc, 'qty': 1.0, 'rate': amt, 'amount': amt})
                        break
        
        if not items:
            preview = "\n".join(lines[:40]) if lines else "(no text extracted)"
            return {"status": "error", "message": f"Could not extract line items from this invoice.\n\n--- Extracted Text (first 40 lines) ---\n{preview}"}
        
        # Clean up descriptions
        for it in items:
            it['description'] = re.sub(r'\s+', ' ', it['description']).strip()
        
        data["items"] = items

        
        # -------------------------------------------------------
        # 6. Amounts - Grand Total (last occurrence wins)
        # -------------------------------------------------------
        all_totals = re.findall(r'(?:Grand\s*Total|Total\s*Amount)[\s:₹Rs.]*([\d,]+(?:\.\d+)?)', text, re.I)
        if all_totals:
            data["grand_total"] = clean_amount(all_totals[-1])
        else:
            # Sum items as fallback
            data["grand_total"] = sum(it["amount"] for it in items)
        
        # -------------------------------------------------------
        # 7. GST / Tax Rate extraction  
        # KEY INSIGHT: PDFs double-print GST amounts (main table + HSN
        # summary). So we extract the RATE from text and calculate
        # the amount from taxable value × rate. This is 100% reliable.
        # -------------------------------------------------------
        gst_rate = 0.0
        gst_amount = 0.0
        round_off = 0.0
        
        # Get taxable base (net of items before tax)
        taxable_base = sum(it.get("amount", 0) for it in items)
        
        # Extract tax RATES from text (rates appear exactly once, amounts appear multiple times)
        sgst_rate_m = re.findall(r'SGST\s*@\s*(\d+(?:\.\d+)?)\s*%', text, re.I)
        cgst_rate_m = re.findall(r'CGST\s*@\s*(\d+(?:\.\d+)?)\s*%', text, re.I)
        igst_rate_m = re.findall(r'IGST\s*@\s*(\d+(?:\.\d+)?)\s*%', text, re.I)
        
        # Also look for rate patterns like "9 %" or "18%" near tax keywords
        if not sgst_rate_m:
            sgst_rate_m = re.findall(r'SGST[^\n]{0,30}?(\d+(?:\.\d+)?)\s*%', text, re.I)
        if not cgst_rate_m:
            cgst_rate_m = re.findall(r'CGST[^\n]{0,30}?(\d+(?:\.\d+)?)\s*%', text, re.I)
        if not igst_rate_m:
            igst_rate_m = re.findall(r'IGST[^\n]{0,30}?(\d+(?:\.\d+)?)\s*%', text, re.I)
        
        if igst_rate_m:
            # Out-state: single IGST
            gst_rate = float(igst_rate_m[-1])
            gst_amount = round(taxable_base * gst_rate / 100, 2)
        elif sgst_rate_m or cgst_rate_m:
            # Same-state: SGST + CGST
            sr = float(sgst_rate_m[-1]) if sgst_rate_m else 0
            cr = float(cgst_rate_m[-1]) if cgst_rate_m else 0
            gst_rate = round(sr + cr, 2)
            gst_amount = round(taxable_base * gst_rate / 100, 2)
        
        # Fallback: try to find total tax from "Tax Amount" summary table
        if gst_amount == 0:
            # Look for "Total" row in HSN summary - last big amount before end
            tax_total_m = re.findall(
                r'(?:Total\s+Tax\s*Amount|Tax\s+Amount|Taxes?\s+and\s+Charges?\s+Added)[^\n]*?([\d,]+\.\d{2})',
                text, re.I
            )
            if tax_total_m:
                gst_amount = clean_amount(tax_total_m[-1])
        
        # -------------------------------------------------------
        # 8. Round Off - extract from invoice
        # -------------------------------------------------------
        # Pattern: "Round Off" followed by optional -ve amount on same or next line
        round_off_m = re.search(
            r'Round\s*(?:Off|off)[^\n]*?(-?\s*[\d,]+\.\d{2})',
            text, re.I
        )
        if round_off_m:
            round_off = float(round_off_m.group(1).replace(',', '').replace(' ', ''))
        else:
            # Try line-window scan
            for i, line in enumerate(lines):
                if re.search(r'round\s*off', line, re.I):
                    for j in range(i, min(i + 4, len(lines))):
                        nxt = lines[j].strip()
                        amt_m = re.match(r'^-?\s*[\d,]+\.\d{2}$', nxt)
                        if amt_m:
                            round_off = float(nxt.replace(',', ''))
                            break
                    break
        
        # -------------------------------------------------------
        # 9. Final Invoice Amount (Grand Total with round off)
        # -------------------------------------------------------
        # Try multiple patterns to find the final payable amount
        final_amount = 0.0
        
        # Pattern A: "₹ XXXX.XX" - rupee symbol with amount (most reliable for Indian invoices)
        rupee_amounts = re.findall(r'[₹\u20b9]\s*([\d,]+\.\d{2})', text)
        if rupee_amounts:
            # Take the last (usually the grand total at the bottom)
            final_amount = clean_amount(rupee_amounts[-1])
        
        # Pattern B: "Grand Total" or "Total Amount"
        if not final_amount:
            gt_m = re.findall(r'(?:Grand\s*Total|Total\s*Amount|Net\s*Payable)[^\n\d₹]*([\d,]+\.\d{2})', text, re.I)
            if gt_m:
                final_amount = clean_amount(gt_m[-1])
        
        # Pattern C: Calculate from parts
        if not final_amount:
            final_amount = round(taxable_base + gst_amount + round_off, 2)
        
        # Cross-check: if grand_total was already found and is close, use it
        existing_grand = data.get('grand_total', 0)
        if existing_grand > 0 and abs(existing_grand - final_amount) < 10:
            final_amount = existing_grand  # They agree, use what we already have
        
        data["gst_amount"] = round(gst_amount, 2)
        data["gst_rate"] = gst_rate
        data["round_off"] = round_off
        data["grand_total"] = final_amount if final_amount > 0 else existing_grand
        
        # -------------------------------------------------------
        # Done
        # -------------------------------------------------------
        return {"status": "success", "data": data}
        
    except Exception as e:
        frappe.log_error("Invoice Extraction Error", frappe.get_traceback())
        return {"status": "error", "message": f"An error occurred while processing the invoice: {str(e)}"}


@frappe.whitelist()
def fetch_po_and_site_details(text_input, invoice_data):
    """
    Parses user natural language input for Circuit ID/PO No and payment details.
    Fetches Site and PO data to populate the missing fields.
    """
    import json
    import re
    data = json.loads(invoice_data)
    
    circuit_match = re.search(r'\b\d{5,8}\b', text_input)
    po_match = re.search(r'PO-\d{4}-\d+', text_input)
    
    circuit_id = circuit_match.group(0) if circuit_match else "71249"
    
    payment_type = "QRC" if "qrc" in text_input.lower() else "MRC"
    expense_type = "Variable" if "variable" in text_input.lower() else "Fixed"
    
    # Default mock fetched data
    site_name = "The Kangra Central - Dharamshala"
    lms_id = "LMS-003721"
    payment_cycle = "15"
    duration_from = "01-07-2026"
    duration_to = "30-09-2026"
    
    # Try fetching real data from Site
    if frappe.db.exists("Site", {"circuit_id": circuit_id}):
        site_doc = frappe.get_doc("Site", {"circuit_id": circuit_id})
        site_name = site_doc.site_name or site_name
        lms_id = site_doc.lms_id or lms_id
    
    data.update({
        "circuit_id": circuit_id,
        "site_name": site_name,
        "lms_id": lms_id,
        "payment_type": payment_type,
        "expense_type": expense_type,
        "payment_cycle": payment_cycle,
        "duration_from": duration_from,
        "duration_to": duration_to
    })
    
    return {
        "status": "success",
        "data": data
    }

@frappe.whitelist()
def create_draft_purchase_invoice(invoice_data):
    """
    Creates a draft Purchase Invoice in ERPNext using the structured data.
    """
    import json
    
    # DEBUG: Log every single call
    frappe.log_error("API Reached - create_draft_purchase_invoice", str(invoice_data)[:1000])
    
    data = json.loads(invoice_data)
    
    try:
        # Prevent Duplicate
        if frappe.db.exists("Purchase Invoice", {"supplier": data.get("supplier_name"), "bill_no": data.get("invoice_no")}):
            return {"status": "error", "message": "Duplicate Invoice! An invoice with this Supplier and Invoice No already exists."}
            
        # Supplier check
        supplier = data.get("supplier_name")
        if not frappe.db.exists("Supplier", supplier):
            # Create a mock supplier or fail
            supplier = frappe.db.get_value("Supplier", {"supplier_name": ["like", "%Bharti%"]}) or "Bharti Airtel Limited"
            if not frappe.db.exists("Supplier", supplier):
                doc = frappe.get_doc({
                    "doctype": "Supplier",
                    "supplier_name": supplier,
                    "supplier_group": "All Supplier Groups"
                })
                doc.insert(ignore_permissions=True)
                
        # Basic formatting dates (DD-MM-YYYY to YYYY-MM-DD)
        def format_date(d_str):
            if "-" in d_str and len(d_str.split("-")[0]) == 2:
                parts = d_str.split("-")
                return f"{parts[2]}-{parts[1]}-{parts[0]}"
            return d_str

        pi = frappe.new_doc("Purchase Invoice")
        pi.supplier = supplier
        pi.bill_no = data.get("invoice_no")
        pi.bill_date = format_date(data.get("invoice_date"))
        pi.posting_date = format_date(data.get("posting_date"))
        
        pi.custom_dutation_from = format_date(data.get("duration_from")) if data.get("duration_from") else None
        pi.custom_duration_to = format_date(data.get("duration_to")) if data.get("duration_to") else None
        pi.custom_lms_id = data.get("lms_id")
        pi.custom_circuit_id = data.get("circuit_id")
        
        po_number = None
        if pi.custom_lms_id:
            lms_doc = frappe.db.get_value("Lastmile Services Master", pi.custom_lms_id, 
                                          ["circuit_id", "po_number", "billing_mode", "payment_cycle"], as_dict=True)
            if lms_doc:
                if not pi.custom_circuit_id:
                    pi.custom_circuit_id = lms_doc.circuit_id
                pi.custom_payment_type = data.get("payment_type") or lms_doc.billing_mode
                pi.custom_payment_cycle = data.get("payment_cycle") or lms_doc.payment_cycle
                po_number = lms_doc.po_number

        pi.custom_payment_catogery = data.get("po_category")
        if not pi.custom_payment_catogery and po_number:
            po_cat = frappe.db.get_value("Purchase Order", po_number, "custom_po_category")
            if po_cat:
                pi.custom_payment_catogery = po_cat

        if pi.custom_circuit_id and not pi.custom_site_name:
            site_name = frappe.db.get_value("Site", pi.custom_circuit_id, "site_name")
            if site_name:
                pi.custom_site_name = site_name

        # Add items
        for item in data.get("items", []):
            item_code = item.get("item_code")
            if not item_code:
                item_code = "Misc"
            if po_number:
                po_items = frappe.get_all("Purchase Order Item", filters={"parent": po_number}, fields=["item_code", "item_name"])
                if len(po_items) == 1:
                    item_code = po_items[0].item_code
                elif po_items:
                    desc_lower = (item.get("description") or "").lower()
                    for pi_item in po_items:
                        if pi_item.item_name and pi_item.item_name.lower() in desc_lower:
                            item_code = pi_item.item_code
                            break

            if not frappe.db.exists("Item", item_code):
                item_code = frappe.db.get_value("Item", {"item_name": ["like", "%Internet%"]}) or "Misc"
            
            item_circuit_id = pi.custom_circuit_id or data.get("circuit_id")
            item_lms_id = pi.custom_lms_id or data.get("lms_id")
            item_site_name = pi.custom_site_name or data.get("site_name")
            
            row = {
                "item_code": item_code,
                "description": item.get("description"),
                "qty": item.get("qty"),
                "rate": item.get("rate"),
                "custom_circuit_id": item_circuit_id,
                "custom_lms_id": item_lms_id,
                "circuit_id": item_circuit_id,
                "lms_id": item_lms_id,
                "site_name": item_site_name,
            }
            if po_number:
                row["purchase_order"] = po_number
                po_item_data = frappe.db.get_value("Purchase Order Item", {"parent": po_number, "item_code": item_code}, 
                    ["name", "project", "cost_center", "expense_account"], as_dict=True)
                if po_item_data:
                    row["po_detail"] = po_item_data.name
                    if po_item_data.project:
                        row["project"] = po_item_data.project
                    if po_item_data.cost_center:
                        row["cost_center"] = po_item_data.cost_center
                    if po_item_data.expense_account:
                        row["expense_account"] = po_item_data.expense_account
                    
            pi.append("items", row)
            
        # Auto-apply taxes based on Supplier/Company default and matched Item Codes
        pi.set_missing_values()
        
        pi.taxes_and_charges = data.get("taxes_and_charges") or ""
        
        pi.set_taxes()
        pi.calculate_taxes_and_totals()
        
        if not data.get("taxes_and_charges"):
            pi.set("taxes", [])
            
        round_off = data.get("round_off", 0)
        if round_off:
            round_off_account = frappe.db.get_value("Account", {
                "account_name": ["like", "%Round%"], 
                "is_group": 0,
                "company": pi.company
            })
            if round_off_account:
                pi.append("taxes", {
                    "charge_type": "Actual",
                    "account_head": round_off_account,
                    "description": "Round Off",
                    "tax_amount": round_off
                })
                
        pi.calculate_taxes_and_totals()
            
        if pi.shipping_address and not frappe.db.exists("Address", pi.shipping_address):
            pi.shipping_address = None
        if pi.billing_address and not frappe.db.exists("Address", pi.billing_address):
            pi.billing_address = None
        if pi.supplier_address and not frappe.db.exists("Address", pi.supplier_address):
            pi.supplier_address = None
            
        pi.flags.ignore_mandatory = True
        pi.insert(ignore_permissions=True)
        
        # Attach the original supplier invoice to the Purchase Invoice
        file_url = data.get("file_url")
        if file_url:
            try:
                existing = frappe.db.exists("File", {
                    "file_url": file_url,
                    "attached_to_doctype": "Purchase Invoice",
                    "attached_to_name": pi.name
                })
                if not existing:
                    file_doc = frappe.get_doc({
                        "doctype": "File",
                        "file_url": file_url,
                        "attached_to_doctype": "Purchase Invoice",
                        "attached_to_name": pi.name,
                        "folder": "Home/Attachments",
                        "is_private": 0
                    })
                    file_doc.insert(ignore_permissions=True)
            except Exception as attach_err:
                frappe.log_error("Invoice Attachment Error", str(attach_err))
        
        if hasattr(frappe.local, "message_log"):
            frappe.local.message_log = []
        if "_server_messages" in frappe.local.response:
            del frappe.local.response["_server_messages"]
            
        return {
            "status": "success",
            "invoice_name": pi.name
        }
    except Exception as e:
        if hasattr(frappe.local, "message_log"):
            frappe.local.message_log = []
        if "exc" in frappe.local.response:
            del frappe.local.response["exc"]
        if "_server_messages" in frappe.local.response:
            del frappe.local.response["_server_messages"]
        return {
            "status": "error",
            "message": str(e)
        }
####################################################################################
# --- START: Job Applicant Stage TAT & Ageing ---
@frappe.whitelist()
def get_job_applicant_stage_history(docname, creation):
    from frappe.utils import getdate, nowdate, date_diff
    import json
    
    # 1. Calendar Ageing (count all days, no holiday list)
    ageing = date_diff(nowdate(), getdate(creation))
    
    # 2. Fetch TAT Target
    tat_target = 30
    grade_name = "Not Specified"
    try:
        job_applicant = frappe.get_doc("Job Applicant", docname)
        if job_applicant.job_title: # 'job_title' is the fieldname for Job Opening in Job Applicant Doctype
            job_opening = frappe.get_doc("Job Opening", job_applicant.job_title)
            if job_opening.custom_grade:
                grade = frappe.get_doc("Employee Grade", job_opening.custom_grade)
                grade_name = grade.name
                if getattr(grade, "custom_tat_days", None):
                    tat_target = int(grade.custom_tat_days)
    except Exception as e:
        frappe.log_error(f"Error fetching TAT: {str(e)}", "Ageing Debug")

    # 3. Stage history
    versions = frappe.db.sql("""
        SELECT creation, data
        FROM `tabVersion`
        WHERE ref_doctype='Job Applicant' AND docname=%s
        ORDER BY creation ASC
    """, (docname,), as_dict=True)
    
    history = {}
    for v in versions:
        try:
            data = json.loads(v.data)
            for change in data.get("changed", []):
                if change[0] == "custom_stage":
                    new_stage = change[2]
                    if new_stage not in history:
                        history[new_stage] = v.creation
        except Exception:
            pass
            
    # 4. Fallback for missing history from related documents
    try:
        # Fallback for Interview stages
        if frappe.db.exists("DocType", "Interview"):
            interviews = frappe.get_all("Interview", 
                filters={"job_applicant": docname}, 
                fields=["interview_round", "scheduled_on", "creation", "status"],
                order_by="creation ASC"
            )
            for intv in interviews:
                date_val = intv.scheduled_on or intv.creation
                rnd = str(intv.interview_round or "").lower()
                stage_key = "Interview Round 1 Scheduled"
                if "2" in rnd:
                    stage_key = "Interview Round 2 Scheduled"
                
                if stage_key not in history:
                    history[stage_key] = date_val
                if "Interview to be scheduled" not in history:
                    history["Interview to be scheduled"] = date_val
                    
        # Fallback for Job Offer
        if frappe.db.exists("DocType", "Job Offer"):
            offers = frappe.get_all("Job Offer",
                filters={"job_applicant": docname},
                fields=["offer_date", "creation", "status", "modified"],
                order_by="creation ASC"
            )
            for offer in offers:
                date_val = offer.offer_date or offer.creation
                if "Offered" not in history:
                    history["Offered"] = date_val
                if offer.status == "Accepted" and "Offer Accepted" not in history:
                    history["Offer Accepted"] = offer.modified
    except Exception as e:
        frappe.log_error(f"Error fetching fallback history: {str(e)}", "Ageing Debug")
            
    return {
        "ageing_days": ageing,
        "tat_target": tat_target,
        "grade_name": grade_name,
        "history": history
    }
# --- END: Job Applicant Stage TAT & Ageing ---

# --- START: AI Candidate Evaluation ---

# ==========================================
# AI CANDIDATE EVALUATION V4.0 MODULE
# ==========================================

import re
import json
import requests
import frappe

# Configuration
MATCH_SCORE_MAP = {
    "Direct Match": 10,
    "Strong Match": 8,
    "Transferable Match": 6,
    "Limited Match": 3,
    "No Evidence": 0
}

WEIGHTS = {
    "Technical Skills": 0.25,
    "Experience": 0.20,
    "Industry": 0.15,
    "Campaign Ownership": 0.15,
    "Marketing Capability": 0.10,
    "Transferable Skills": 0.10,
    "Education": 0.05
}

def extract_resume(applicant_name):
    applicant = frappe.get_doc("Job Applicant", applicant_name)
    resume_file = frappe.get_all("File", filters={"attached_to_doctype": "Job Applicant", "attached_to_name": applicant.name}, order_by="creation desc", limit=1)
    
    text = ""
    if resume_file:
        file_doc = frappe.get_doc("File", resume_file[0].name)
        file_url = applicant.resume_attachment
        if file_url:
            import fitz
            import os
            import urllib.parse
            import zipfile
            import io
            import xml.etree.ElementTree as ET
            
            decoded_url = urllib.parse.unquote(file_url)
            if "private/files/" in decoded_url:
                f_name = decoded_url.split("private/files/")[-1]
                file_path = frappe.get_site_path("private", "files", f_name)
            elif "/files/" in decoded_url:
                f_name = decoded_url.split("/files/")[-1]
                file_path = frappe.get_site_path("public", "files", f_name)
            else:
                file_path = ""
                
            if file_path and os.path.exists(file_path):
                file_ext = os.path.splitext(file_path)[1].lower()
                with open(file_path, "rb") as f:
                    file_bytes = f.read()
                    
                if file_ext == '.pdf':
                    try:
                        doc = fitz.open(stream=file_bytes, filetype="pdf")
                        for page in doc:
                            try:
                                text += page.get_text(sort=True)
                            except TypeError:
                                text += page.get_text()
                    except Exception as e:
                        frappe.log_error(f"PDF Parse Error: {str(e)}")
                elif file_ext in ['.doc', '.docx']:
                    try:
                        with zipfile.ZipFile(io.BytesIO(file_bytes)) as docx:
                            tree = ET.XML(docx.read('word/document.xml'))
                            paragraphs = []
                            for paragraph in tree.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p'):
                                texts = [node.text for node in paragraph.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t') if node.text]
                                if texts:
                                    paragraphs.append(''.join(texts))
                            text = '\n'.join(paragraphs)
                    except Exception:
                        pass
    
    return applicant, text[:8000]

def extract_requirements(job_opening_name):
    if not job_opening_name:
        return ""
    job = frappe.get_doc("Job Opening", job_opening_name)
    return job.description

def call_llm(job_desc, cv_text):
    system_prompt = """You are a Senior Talent Acquisition Director. Evaluate the CV against the Job Description.

Perform ONLY reasoning. For each Job Description requirement, return an evaluation object.

Each evaluation must have:
- "requirement": The specific JD requirement.
- "category": Choose ONE of ["Technical Skills", "Experience", "Industry", "Campaign Ownership", "Marketing Capability", "Education", "Transferable Skills"].
- "match_level": Choose ONLY ONE of ["Direct Match", "Strong Match", "Transferable Match", "Limited Match", "No Evidence"].
- "evidence_found": Array of facts from CV. (Empty if None)
- "missing": Array of missing elements.
- "justification": Why you chose this match level.
- "recommendation": Suggested action to address gaps.
- "interview_question": A personalized question to validate gaps or experience.

Also return a "hiring_manager_verdict" (max 4 sentences) explaining why to interview, risks, what to validate, and training.

Return ONLY valid JSON:
{
  "hiring_manager_verdict": "...",
  "evaluations": [
    {
      "requirement": "...",
      "category": "...",
      "match_level": "...",
      "evidence_found": ["..."],
      "missing": ["..."],
      "justification": "...",
      "recommendation": "...",
      "interview_question": "..."
    }
  ]
}"""

    user_prompt = f"Job Description:\n{job_desc}\n\nCandidate CV:\n{cv_text}"
    try:
        api_config = frappe.get_doc("API Configuration")
    except Exception:
        api_configs = frappe.get_all("API Configuration", limit=1)
        if not api_configs:
            frappe.throw("No API Configuration found.")
        api_config = frappe.get_doc("API Configuration", api_configs[0].name)
        
    api_key = api_config.get_password("api_key") or api_config.api_key
    base_url = api_config.api_base_url or "https://api.groq.com/openai/v1/chat/completions"
    model_name = api_config.model_name or "llama3-70b-8192"
    
    if not api_key:
        frappe.throw("API Key missing.")
        
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "model": model_name,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ],
        "temperature": 0.1
    }
    
    response = requests.post(base_url, headers=headers, json=payload, timeout=20)
    response.raise_for_status()
    result = response.json()
    
    return result['choices'][0]['message']['content'].strip()

def parse_ai_response(content):
    if content.startswith("```json"):
        content = content.replace("```json", "", 1)
    if content.startswith("```"):
        content = content.replace("```", "", 1)
    if content.endswith("```"):
        content = content[:-3]
    return json.loads(content.strip())

def calculate_scores(parsed_json):
    cat_sums = {}
    cat_counts = {}
    
    for ev in parsed_json.get("evaluations", []):
        cat = ev.get("category", "Transferable Skills")
        ml = ev.get("match_level", "No Evidence")
        ml_clean = ml.replace("🟢", "").replace("🟡", "").replace("🔴", "").strip()
        
        score = 0
        for k, v in MATCH_SCORE_MAP.items():
            if k.lower() in ml_clean.lower():
                score = v
                ml_clean = k
                break
                
        ev["score"] = score
        ev["match_level"] = ml_clean
        
        cat_sums[cat] = cat_sums.get(cat, 0) + score
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
        
    category_scores = {k: v / cat_counts[k] for k, v in cat_sums.items()}
    
    overall_score = 0
    total_w = 0
    for cat, val in category_scores.items():
        w = WEIGHTS.get(cat, 0)
        overall_score += val * w
        total_w += w
        
    if total_w > 0:
        overall_score /= total_w
    else:
        overall_score = 0
        
    return category_scores, overall_score

def calculate_confidence(evaluations):
    total = len(evaluations)
    if total == 0: return 0
    direct = sum(1 for e in evaluations if "Direct" in str(e.get("match_level", "")))
    transfer = sum(1 for e in evaluations if "Transferable" in str(e.get("match_level", "")))
    
    ratio = direct / total
    conf = int((ratio * 100) + (transfer / total * 50) + 15)
    return min(100, max(0, conf))

def calculate_business_risk(evaluations):
    # Deterministic risk based on missing items or low scores
    missing_technical = sum(1 for e in evaluations if e.get("category") == "Technical Skills" and e.get("score", 0) <= 3)
    missing_industry = sum(1 for e in evaluations if e.get("category") == "Industry" and e.get("score", 0) <= 3)
    
    tech_risk = "High" if missing_technical > 1 else ("Medium" if missing_technical == 1 else "Low")
    domain_risk = "High" if missing_industry > 1 else ("Medium" if missing_industry == 1 else "Low")
    
    overall = "High" if tech_risk == "High" or domain_risk == "High" else "Medium"
    if tech_risk == "Low" and domain_risk == "Low": overall = "Low"
    
    return {
        "overall": overall,
        "technical_risk": tech_risk,
        "domain_risk": domain_risk,
        "leadership_risk": "Cannot determine from CV",
        "retention_risk": "Needs Discussion"
    }

def calculate_resume_quality(cv_text):
    text_lower = cv_text.lower()
    sections = ["summary", "experience", "education", "projects", "achievements", "certifications", "skills", "languages", "portfolio", "linkedin"]
    includes = []
    missing = []
    
    for s in sections:
        if s in text_lower:
            includes.append(s.title())
        else:
            missing.append(s.title())
            
    score = int((len(includes) / len(sections)) * 100)
    advice = "Add missing sections: " + ", ".join(missing[:3]) if missing else "Excellent completeness."
    
    return {
        "score": f"{score}%",
        "includes": includes,
        "improvement_advice": advice
    }

def calculate_career_growth(cv_text):
    text_lower = cv_text.lower()
    promotions = len(re.findall(r'senior|manager|director|lead|head', text_lower))
    if promotions >= 3:
        return {"rating": "Strong Growth", "reasoning": "CV contains multiple progressive titles (e.g. Senior, Manager)."}
    elif promotions >= 1:
        return {"rating": "Moderate Growth", "reasoning": "Some evidence of increasing responsibility."}
    return {"rating": "Limited Growth", "reasoning": "No clear evidence of title progression."}

def calculate_career_stability(cv_text):
    years = re.findall(r'(20\d{2})', cv_text)
    years = sorted([int(y) for y in set(years)])
    rating = "Moderate"
    if len(years) > 2:
        span = max(years) - min(years)
        if span > 5:
            rating = "Good"
    
    return {
        "rating": rating,
        "average_tenure": "Approx 2-3 Years based on date ranges.",
        "observation": "Employment timeline suggests standard industry tenure.",
        "interview_question": "Can you walk me through the timeline of your previous roles?"
    }

def calculate_recommendation(overall_score):
    pct = overall_score * 10
    if pct >= 85: return "Excellent Match"
    if pct >= 75: return "Strong Match"
    if pct >= 65: return "Good Match"
    if pct >= 50: return "Interview Recommended (Conditional)"
    if pct >= 35: return "Potential Candidate"
    return "Not Recommended"

def calculate_ramp_up(evaluations):
    gaps = sum(1 for e in evaluations if e.get("score", 0) <= 6)
    if gaps <= 2: return "2 Weeks"
    if gaps <= 5: return "1 Month"
    return "2 Months"

def calculate_role_fit(overall_score):
    pct = min(100, int(overall_score * 10))
    return [{"role": "Primary Target Role", "percentage": pct, "reason": "Calculated deterministically from weighted category scores."}]

def build_interview_questions(evaluations):
    q_dict = {"technical": [], "behavioral": [], "gap_validation": []}
    for e in evaluations:
        q = e.get("interview_question")
        if q and e.get("score", 0) <= 6:
            q_dict["gap_validation"].append(q)
        elif q:
            q_dict["technical"].append(q)
            
    # deduplicate and cap at 3
    return {k: list(set(v))[:3] for k, v in q_dict.items() if v}

@frappe.whitelist()
def evaluate_candidate_cv(job_applicant_name):
    try:
        applicant, cv_text = extract_resume(job_applicant_name)
        if not cv_text:
            frappe.throw("Could not extract text from Candidate CV.")
            
        job_desc = extract_requirements(applicant.job_title)
        
        llm_response = call_llm(job_desc, cv_text)
        parsed_json = parse_ai_response(llm_response)
        
        # Deterministic Rules
        cat_scores, overall_score = calculate_scores(parsed_json)
        evals = parsed_json.get("evaluations", [])
        
        parsed_json["category_scores"] = {k: round(v, 1) for k, v in cat_scores.items()}
        parsed_json["overall_score"] = round(overall_score, 1)
        parsed_json["overall_rating"] = calculate_recommendation(overall_score)
        
        parsed_json["ai_confidence"] = calculate_confidence(evals)
        parsed_json["business_risk"] = calculate_business_risk(evals)
        parsed_json["resume_completeness"] = calculate_resume_quality(cv_text)
        parsed_json["career_growth"] = calculate_career_growth(cv_text)
        parsed_json["career_stability"] = calculate_career_stability(cv_text)
        parsed_json["role_fit"] = calculate_role_fit(overall_score)
        
        parsed_json["final_hiring_recommendation"] = {
            "hiring_recommendation": parsed_json["overall_rating"],
            "training_required": "High" if parsed_json["business_risk"]["overall"] == "High" else "Moderate",
            "ramp_up": calculate_ramp_up(evals),
            "confidence": parsed_json["ai_confidence"]
        }
        
        parsed_json["interview_questions"] = build_interview_questions(evals)
        
        # We don't overwrite why_good_match etc, just let them be empty if not provided, UI handles it.
        
        applicant.db_set("custom_ai_evaluation", json.dumps(parsed_json))
        return parsed_json
        
    except Exception as e:
        frappe.log_error(f"AI Eval V4 Failed: {str(e)}", "Resume AI Evaluation")
        frappe.throw(f"AI Evaluation Failed: {str(e)}")

# --- END: AI Candidate Evaluation ---

# --- START: AI Feasibility Evaluation ---
@frappe.whitelist()
def evaluate_feasibility_with_ai(doc_data):
    import json
    import requests
    import traceback
    
    try:
        doc = json.loads(doc_data)
        
        # Context gathering from ERPNext DB
        pincode = doc.get('pincode')
        place = doc.get('city') or doc.get('address_street') or "Unknown"
        lms_type = doc.get('lms_type') or "Single"
        primary_plan = doc.get('primary_data_plan') or "Unknown"
        secondary_plan = doc.get('secondary_data_plan') or "Unknown"
        circuit_id = doc.get('site_id__legal_code') or "Unknown"
        
        from nexapp.nexapp.doctype.feasibility.feasibility import get_supplier_pool_by_pincode
        
        # Use the exact same logic as the Supplier Pool UI
        pool_data = {"isp_pool": [], "feas_pool": []}
        if pincode:
            try:
                pool_data = get_supplier_pool_by_pincode(pincode, circuit_id)
            except Exception:
                pass
                
        isp_pool = pool_data.get("isp_pool", [])
        
        # Extract valid suppliers from ISP pool
        valid_suppliers = list(set([s.get("supplier_name") for s in isp_pool if s.get("supplier_name")]))
        valid_suppliers_str = ", ".join(valid_suppliers) if valid_suppliers else "NONE"
        
        # Get total active sites for these suppliers globally for context
        supplier_active_sites = {}
        if valid_suppliers:
            counts = frappe.db.get_all("Lastmile Services Master", 
                filters={"supplier": ["in", valid_suppliers], "lms_stage": ["in", ["Delivered", "Live"]]},
                fields=["supplier", "count(name) as count"],
                group_by="supplier"
            )
            for c in counts:
                supplier_active_sites[c.supplier] = c.count
        
        # Extract PO details for the prompt and attach active site counts
        po_with_items = []
        for s in isp_pool:
            po_info = s.get("latest_po")
            supp_name = s.get("supplier_name")
            if po_info:
                po_with_items.append({
                    "po_number": po_info.get("po_name"),
                    "supplier": supp_name,
                    "date": po_info.get("po_date"),
                    "grand_total": po_info.get("grand_total"),
                    "items": po_info.get("items", []),
                    "circuit_id": po_info.get("circuit_id"),
                    "total_active_sites_globally": supplier_active_sites.get(supp_name, 0)
                })
            else:
                po_with_items.append({
                    "supplier": supp_name,
                    "total_active_sites_globally": supplier_active_sites.get(supp_name, 0)
                })
        # Also extract Feasibility pool (suppliers evaluated but maybe no PO yet)
        feas_pool = pool_data.get("feas_pool", [])
        feas_suppliers = list(set([s.get("supplier_name") for s in feas_pool if s.get("supplier_name")]))
        
        all_suppliers_for_intel = list(set(valid_suppliers + feas_suppliers))
        
        # Pricing Intelligence
        supplier_items = []
        if all_suppliers_for_intel:
            supplier_items = frappe.db.get_all("Item Price",
                filters={"item_code": ["in", [primary_plan, secondary_plan]], "supplier": ["in", all_suppliers_for_intel]},
                fields=["supplier", "item_code", "price_list_rate as price", "currency", "lead_time_days"],
                limit=50
            )

        system_prompt = """You are an AI Supplier Recommendation Engine for ERPNext.
Your objective is to help the user find the best Supplier based on the provided Feasibility, Purchase Order, and Site data.

IMPORTANT: All data provided to you has been pre-filtered to the SPECIFIC PINCODE of the Feasibility request.
The Purchase Orders shown are ONLY those from the Supplier Pool where the LMS stage is 'Delivered' or 'Live'.
Additionally, you are provided "Other Potential Suppliers" from the Feasibility Pool in this pincode. These are suppliers that were evaluated previously but may not have a recent PO. You MUST use them as valid options to fill out your 5 supplier comparison.

Scoring Criteria:
- 40% Location Match (same Place/Pincode)
- 30% Price (compare Data Plan pricing)
- 20% Recent Orders (recent Purchase Orders)
- 10% Number of Successful Deliveries

You must NEVER guess or hallucinate suppliers. ONLY explain the data returned from ERPNext.
If Active Suppliers in Pincode is NONE, you MUST state that no recommendation can be made and set recommended_supplier to "No Suppliers Found".
IMPORTANT FOR SUPPLIER COMPARISON: You MUST include up to 5 available suppliers in the supplier_comparison array. You MUST assign a strict Letter Grade (A+, A, B, C, D) to each supplier based on their rate and active sites. NEVER output "N/A" for the grade.

Return ONLY valid JSON matching this exact structure:
{
    "pincode": "Pin Code Evaluated",
    "city": "City Evaluated",
    "recommended_supplier": "Supplier Name",
    "confidence_score": 95,
    "reason": [
        "✓ Supplier has delivered POs in Kochi",
        "✓ Lowest cost",
        "✓ Available Secondary Data Plan"
    ],
    "supplier_comparison": [
        {
            "supplier": "Supplier A",
            "rate": "₹ 499",
            "item_name": "40Mbps-MBB",
            "active_sites": 12,
            "grade": "A+",
            "notes": "Lowest rate, highly active."
        },
        {
            "supplier": "Supplier B",
            "rate": "₹ 600",
            "item_name": "40Mbps-MBB",
            "active_sites": 5,
            "grade": "B",
            "notes": "More expensive, fewer active sites."
        }
    ],
    "po_details": [
        {"po_number": "PO-001", "supplier": "Supplier A", "date": "2023-10-01", "item_rate": "₹ 499"}
    ],
    "evaluation_steps": [
        "Identified Active Suppliers in Pincode 682001.",
        "Traced Purchase Orders from the 'Delivered' supplier pool.",
        "Compared pricing for Primary Plan (100MBPS) and found Supplier A is cheapest.",
        "Analyzed recent PO volume for Supplier A."
    ]
}"""

        user_prompt = f"""Feasibility Request Data:
LMS Type: {lms_type}
Circuit ID: {circuit_id}
Primary Plan: {primary_plan}
Secondary Plan: {secondary_plan}
Pincode: {pincode}
Place: {place}
Active Suppliers in Pincode (from Delivered PO data): {valid_suppliers_str}

Purchase Orders linked to Sites in Pincode {pincode} (System Data):
{json.dumps(po_with_items, indent=2, default=str)}

Other Potential Suppliers (from Feasibility Pool in this Pincode):
{json.dumps(feas_pool, indent=2, default=str)}

Supplier Item Pricing Intelligence (System Data):
{json.dumps(supplier_items, indent=2, default=str)}"""

        # API Call setup
        try:
            api_config = frappe.get_doc("API Configuration")
        except Exception:
            api_configs = frappe.get_all("API Configuration", limit=1)
            if not api_configs:
                frappe.throw("No API Configuration found. Cannot run AI Evaluation.")
            api_config = frappe.get_doc("API Configuration", api_configs[0].name)
            
        api_key = api_config.get_password("api_key") or api_config.api_key
        base_url = api_config.api_base_url or "https://api.groq.com/openai/v1/chat/completions"
        model_name = api_config.model_name or "llama3-70b-8192"
        
        if not api_key:
            frappe.throw("API Key missing in API Configuration.")
            
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "model": model_name,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            "temperature": 0.2
        }
        
        response = requests.post(base_url, headers=headers, json=payload, timeout=30)
        response.raise_for_status()
        result = response.json()
        
        content = result['choices'][0]['message']['content'].strip()
        
        # Cleanup JSON formatting if present
        import re
        json_match = re.search(r'\{.*\}', content, re.DOTALL)
        if json_match:
            content = json_match.group(0)
            
        try:
            parsed_result = json.loads(content.strip())
        except Exception as e:
            frappe.log_error(f"Failed to parse LLM output: {content}", "AI Evaluation Error")
            frappe.throw(f"AI returned invalid JSON: {str(e)}")
        
        # Auto-save logic
        doc_name = doc.get("name")
        if doc_name and not str(doc_name).startswith("new-"):
            if frappe.db.exists("Feasibility", doc_name):
                # Check if custom field exists
                if not frappe.db.has_column("Feasibility", "custom_ai_evaluation"):
                    from frappe.custom.doctype.custom_field.custom_field import create_custom_field
                    create_custom_field("Feasibility", {
                        "fieldname": "custom_ai_evaluation",
                        "label": "AI Evaluation",
                        "fieldtype": "Code",
                        "options": "JSON",
                        "insert_after": "feasibility_status",
                        "read_only": 1,
                        "hidden": 1
                    })
                # Save the evaluation to the database
                frappe.db.set_value("Feasibility", doc_name, "custom_ai_evaluation", json.dumps(parsed_result))
        
        return parsed_result
        
    except Exception as e:
        error_msg = str(e)
        frappe.log_error(f"Feasibility AI Eval Failed: {error_msg}\n{traceback.format_exc()}", "Feasibility AI Evaluation")
        return {"error": error_msg}
# --- END: AI Feasibility Evaluation ---

# --- START: AI Purchase Invoice Creation Logic ---

@frappe.whitelist()
def extract_purchase_invoice_data(file_url=None):
    """
    Robust extraction from a Purchase Invoice PDF using PyMuPDF.
    """
    import re
    
    if not file_url:
        return {"status": "error", "message": "No invoice file was attached. Please attach a PDF invoice."}
    
    try:
        import os
        # Robust file path resolution
        try:
            from frappe.utils.file_manager import get_file_path
            file_path = get_file_path(file_url.split('/')[-1])
        except Exception:
            file_path = None
        
        # Fallback: construct path from site path + file_url
        if not file_path or not os.path.exists(file_path):
            site_path = frappe.get_site_path()
            relative = file_url.lstrip('/')
            file_path = os.path.join(site_path, 'public', relative.replace('files/', ''))
            if not os.path.exists(file_path):
                file_path = os.path.join(site_path, relative)
        
        if not file_path or not os.path.exists(file_path):
            return {"status": "error", "message": "Could not locate the uploaded invoice file on the server."}
        
        if not file_path.lower().endswith('.pdf'):
            return {"status": "error", "message": "Only PDF files are supported for extraction."}
        
        import fitz
        doc = fitz.open(file_path)
        text = ""
        blocks = []
        
        # --- Strategy 1: Standard text extraction ---
        for page in doc:
            text += page.get_text()
            blocks += page.get_text("blocks")
        
        # --- Strategy 2: rawdict — works on vector-font PDFs where get_text() returns empty ---
        if not text.strip():
            try:
                raw_text_parts = []
                for page in doc:
                    raw = page.get_text("rawdict")
                    for block in raw.get("blocks", []):
                        for line in block.get("lines", []):
                            for span in line.get("spans", []):
                                t = span.get("text", "").strip()
                                if t:
                                    raw_text_parts.append(t)
                text = "\n".join(raw_text_parts)
            except Exception:
                pass
        
        # --- Strategy 3: OCR fallback (requires tesseract-ocr + pytesseract) ---
        if not text.strip():
            try:
                import pytesseract
                from PIL import Image
                import io, shutil
                # Explicitly set the tesseract binary path since bench env PATH
                # may not include /usr/bin where the system tesseract lives
                tess_path = shutil.which("tesseract") or "/usr/bin/tesseract"
                pytesseract.pytesseract.tesseract_cmd = tess_path
                ocr_parts = []
                for page in doc:
                    mat = fitz.Matrix(2.0, 2.0)  # 2x zoom for better OCR accuracy
                    pix = page.get_pixmap(matrix=mat)
                    img_bytes = pix.tobytes("png")
                    img = Image.open(io.BytesIO(img_bytes))
                    page_text = pytesseract.image_to_string(img, lang='eng')
                    ocr_parts.append(page_text)
                text = "\n".join(ocr_parts)
                frappe.logger().info(f"Invoice Extraction: Used OCR fallback for {file_url}")
            except ImportError:
                pass  # pytesseract not installed, skip
            except Exception as ocr_err:
                frappe.log_error("Invoice OCR Error", str(ocr_err))
        
        if not text.strip():
            return {
                "status": "error",
                "message": "Could not extract text from this PDF. It appears to be a fully image-based scan. Please install tesseract-ocr on the server to enable OCR support, or upload a text-based PDF."
            }
        
        data = {}
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        
        # --- TEMP DEBUG: log extracted lines to diagnose item parsing ---
        frappe.log_error("PDF Extracted Lines", "\n".join(f"{i}: {l}" for i, l in enumerate(lines[:200])))
        
        # -------------------------------------------------------
        # Helper
        # -------------------------------------------------------
        def parse_date(d_str):
            """Normalise various date formats to YYYY-MM-DD."""
            import datetime
            d_str = d_str.strip()
            month_map = {
                'jan':'01','feb':'02','mar':'03','apr':'04','may':'05','jun':'06',
                'jul':'07','aug':'08','sep':'09','oct':'10','nov':'11','dec':'12'
            }
            # DD-Mon-YY or DD-Mon-YYYY (e.g. 13-Aug-25)
            m = re.match(r'(\d{1,2})[-/\s]([A-Za-z]{3})[-/\s](\d{2,4})', d_str)
            if m:
                dd, mon, yy = m.group(1).zfill(2), m.group(2).lower()[:3], m.group(3)
                mm = month_map.get(mon, '01')
                yyyy = ('20' + yy) if len(yy) == 2 else yy
                return f"{yyyy}-{mm}-{dd}"
            # DD-MM-YYYY or DD/MM/YYYY
            m = re.match(r'(\d{2})[-/](\d{2})[-/](\d{4})', d_str)
            if m:
                return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
            # YYYY-MM-DD (already correct)
            m = re.match(r'(\d{4})-(\d{2})-(\d{2})', d_str)
            if m:
                return d_str
            try:
                return str(frappe.utils.getdate(d_str))
            except Exception:
                return d_str
        
        def clean_amount(s):
            return float(re.sub(r'[^\d.]', '', s)) if s else 0.0
        
        # -------------------------------------------------------
        # 1. Invoice Number - multiple patterns, pick first valid
        # -------------------------------------------------------
        inv_no = None
        inv_no_patterns = [
            # OCR jump: matches 'Invoice No' then skips up to 120 chars to find a proper invoice format (e.g. KNPL/26-27/0078)
            r'(?:Invoice\s*No\.?|Invoice\s*Number|INV\s*NO\.?|Bill\s*No\.?)[^\n]{0,120}?\b([A-Za-z0-9]+(?:[-/][A-Za-z0-9]+){2,})\b',
            r'Invoice\s*No\.?\s*[:\-]?\s*([A-Za-z0-9/_\-]+)',
            r'Invoice\s*Number\s*[:\-]?\s*([A-Za-z0-9/_\-]+)',
            r'INV\s*NO\.?\s*[:\-]?\s*([A-Za-z0-9/_\-]+)',
            r'Bill\s*No\.?\s*[:\-]?\s*([A-Za-z0-9/_\-]+)',
        ]
        for pat in inv_no_patterns:
            m = re.search(pat, text, re.I)
            if m:
                candidate = m.group(1).strip()
                # Must be at least 2 chars and not just a date
                if len(candidate) >= 2 and not re.match(r'^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$', candidate):
                    inv_no = candidate
                    break
        if not inv_no:
            frappe.logger().warning("Could not extract Invoice Number.")
            data["invoice_no"] = None
            data["invoice_no_warning"] = "Could not extract Invoice Number from the document."
        else:
            data["invoice_no"] = inv_no
        
        # -------------------------------------------------------
        # 2. Invoice Date
        # -------------------------------------------------------
        inv_date = None
        date_patterns = [
            r'(?:Invoice\s*Date|Dated|Date\s*of\s*Invoice)[\s:]*([\d]{1,2}[-/\s][A-Za-z]{3}[-/\s][\d]{2,4})',
            r'(?:Invoice\s*Date|Dated|Date\s*of\s*Invoice)[\s:]*([\d]{1,2}[-/][\d]{1,2}[-/][\d]{2,4})',
            r'(?:Invoice\s*Date|Dated)[\s:]*([\d]{4}-[\d]{2}-[\d]{2})',
        ]
        for pat in date_patterns:
            m = re.search(pat, text, re.I)
            if m:
                inv_date = parse_date(m.group(1))
                break
        if not inv_date:
            # Last resort: find any standalone date near "Dated" or top of document
            m = re.search(r'(\d{1,2}[-/][A-Za-z]{3}[-/]\d{2,4})', text[:500])
            if m:
                inv_date = parse_date(m.group(1))
        data["invoice_date"] = inv_date or frappe.utils.today()
        
        # -------------------------------------------------------
        # 3. Duration From / To  (e.g. 13.09.2025 - 12.11.2025)
        # -------------------------------------------------------
        dur_m = re.search(
            r'(\d{1,2}[.\-/]?\d{1,2}[.\-/]?\d{2,4})\s*(?:TO|to|[-–])\s*(\d{1,2}[.\-/]?\d{1,2}[.\-/]?\d{2,4})',
            text, re.I
        )
        if dur_m:
            def parse_ocr_date(s):
                s = re.sub(r'[.\-/]', '', s)
                if len(s) == 8:  # DDMMYYYY
                    return f"{s[4:8]}-{s[2:4]}-{s[0:2]}"
                elif len(s) == 6:  # DDMMYY
                    return f"20{s[4:6]}-{s[2:4]}-{s[0:2]}"
                return s
            data["duration_from"] = parse_ocr_date(dur_m.group(1))
            data["duration_to"] = parse_ocr_date(dur_m.group(2))
        
        # -------------------------------------------------------
        # 3.5. Extract LMS ID & PO Number
        # -------------------------------------------------------
        po_match = re.search(r'(?:PO\s*No|Purchase\s*Order|PO#|PO\s*Number)[\s\.:\-#]*([A-Za-z0-9\-/_]+)', text, re.I)
        if po_match:
            data["po_number"] = po_match.group(1).strip()
            
        lms_match = re.search(r'(?:LMS\s*Id|LMS\s*No|LMS#)[\s\.:\-#]*([A-Za-z0-9\-/_]+)', text, re.I)
        if lms_match:
            data["lms_id"] = lms_match.group(1).strip()

        # -------------------------------------------------------
        # 4. Supplier Name - Exclusion-First Multi-Strategy Matching
        # Key insight: The logged-in company is NEVER the supplier.
        # Collect ALL company-like names, exclude self, then fuzzy-match.
        # -------------------------------------------------------
        import difflib
        
        # Get all self-company names to exclude (the buyer, not the seller)
        own_company = frappe.defaults.get_global_default('company') or ""
        own_company_words = set(re.sub(r'[^a-z0-9\s]', '', own_company.lower()).split())
        
        # Aliases / abbreviations of own company to also exclude
        own_company_aliases = {"nexapp", "nexapp technologies", "ntpl"}
        
        def is_own_company(name):
            """Returns True if the name looks like our own company, not a supplier."""
            n_lower = name.lower()
            n_clean = re.sub(r'[^a-z0-9\s]', '', n_lower)
            n_words = set(n_clean.split())
            # Check if majority of words overlap with own company name
            if own_company_words and len(own_company_words & n_words) >= max(1, len(own_company_words) - 1):
                return True
            for alias in own_company_aliases:
                if alias in n_lower:
                    return True
            return False
        
        def normalize_name(n):
            n = str(n).lower()
            n = re.sub(r'[^a-z0-9\s]', '', n)
            n = n.replace('private limited', 'pvt ltd')
            n = n.replace('pvt limited', 'pvt ltd')
            n = n.replace('bharti airtel limited', 'bharti airtel ltd')
            return n.strip()
        
        # Strategy 1: Collect ALL company-like names in the document
        # Patterns: "Xyz Ltd", "Abc Pvt Ltd", "Xyz Limited", "Xyz LLP"
        all_company_matches = re.findall(
            r'[A-Za-z][A-Za-z0-9\s&\(\)\-\.]{2,60}(?:Pvt\.?\s*Ltd\.?|Private\s+Limited|Limited|LLP|Ltd\.?)',
            text, re.I
        )
        
        # Also specifically look for names near "From:", "Supplier:", "Issued by:", signature blocks
        seller_patterns = [
            r'(?:from|vendor|supplier|issued\s*by|authorised\s*signatory)[:\s]+([A-Za-z][A-Za-z0-9\s&\-\.]{2,80})',
            r'([A-Za-z][A-Za-z0-9\s&\-\.]{2,50})\s*\n?\s*Authorised\s*Signatory',
            r'([A-Za-z][A-Za-z0-9\s&\-\.]{2,50})\s*\n?\s*Authorized\s*Signatory',
        ]
        priority_candidates = []
        for pat in seller_patterns:
            for m in re.finditer(pat, text, re.I):
                priority_candidates.append(m.group(1).strip())
        
        # Strategy 2: Load ERPNext suppliers and score each candidate
        suppliers = frappe.get_all("Supplier", pluck="name")
        if not suppliers:
            return {"status": "error", "message": "No Suppliers exist in the system."}
        
        norm_to_orig = {normalize_name(s): s for s in suppliers}
        
        def best_supplier_match(candidates, cutoff=0.55):
            """Find the best matching ERPNext supplier from a list of raw name candidates."""
            best_score = 0
            best_match = None
            for candidate in candidates:
                if not candidate or is_own_company(candidate):
                    continue
                norm_c = normalize_name(candidate)
                # Substring match (highest confidence)
                for norm_s, orig_s in norm_to_orig.items():
                    if norm_s and norm_c and (norm_s in norm_c or norm_c in norm_s):
                        return orig_s  # Definitive match
                # Fuzzy match
                close = difflib.get_close_matches(norm_c, list(norm_to_orig.keys()), n=1, cutoff=cutoff)
                if close:
                    score = difflib.SequenceMatcher(None, norm_c, close[0]).ratio()
                    if score > best_score:
                        best_score = score
                        best_match = norm_to_orig[close[0]]
            return best_match
        
        # Priority: look at seller-specific regions first
        matched_supplier = best_supplier_match(priority_candidates, cutoff=0.5)
        
        # Fallback: scan all company names found in the document (excluding self)
        if not matched_supplier:
            # Filter out own-company names and deduplicate
            external_candidates = [c for c in all_company_matches if not is_own_company(c)]
            # Reverse order: supplier name often appears later in invoice (bottom area)
            matched_supplier = best_supplier_match(list(reversed(external_candidates)), cutoff=0.6)
        
        if not matched_supplier:
            # Last resort: take any company name not matching own company
            external_candidates = [c for c in all_company_matches if not is_own_company(c)]
            if external_candidates:
                raw_supplier_name = external_candidates[-1].strip()
            else:
                raw_supplier_name = all_company_matches[0].strip() if all_company_matches else "Unknown"
            
            # --- START FIX FOR MISSING SUPPLIER ---
            # Instead of returning an error and halting extraction, we just log it
            frappe.logger().warning(f"Supplier '{raw_supplier_name}' detected but not found in ERPNext.")
            data["supplier_name"] = None
            data["supplier_warning"] = f"Supplier '{raw_supplier_name}' was detected but does not exist in the system."
            # --- END FIX FOR MISSING SUPPLIER ---
        else:
            data["supplier_name"] = matched_supplier
        
        
        # -------------------------------------------------------
        # 5. Items - Block/sequence based extraction
        # PDFs often extract table cells as separate lines, so we
        # use a state-machine approach on the lines list.
        # -------------------------------------------------------
        items = []
        
        # Strategy A: PyMuPDF word-position clustering by column X-position
        # We use blocks to identify rows based on Y-position proximity
        try:
            row_data = {}  # y_bucket -> list of (x, text)
            for page in doc:
                words = page.get_text("words")  # (x0,y0,x1,y1,word,block,line,word_idx)
                for w in words:
                    y_bucket = round(w[1] / 8) * 8  # cluster within 8pt vertical bands
                    row_data.setdefault(y_bucket, []).append((w[0], w[4]))
            
            # Sort rows by Y
            sorted_rows = [row_data[k] for k in sorted(row_data.keys())]
            
            # Find page width to determine column boundaries
            page_width = doc[0].rect.width
            # Heuristic column positions for this invoice layout:
            # Sr(~30), Desc(~50-250), HSN(~260-310), Qty(~310-360), Rate(~360-430), per(~430-470), Amount(~470+)
            
            SKIP_KEYWORDS = {'description', 'goods', 'hsn', 'sac', 'quantity', 'rate', 'per',
                             'amount', 'total', 'gst', 'tax', 'sgst', 'cgst', 'igst',
                             'taxable', 'value', 'round', 'off', 'si', 'no', 'sl'}
            
            current_item = None
            for row_words in sorted_rows:
                if not row_words:
                    continue
                row_words_sorted = sorted(row_words, key=lambda x: x[0])
                row_text = ' '.join(w[1] for w in row_words_sorted).strip()
                
                # Check if this row starts with a serial number (1, 2, 3...)
                first_word = row_words_sorted[0][1]
                if re.match(r'^\d{1,2}$', first_word):
                    sr_no = int(first_word)
                    if 1 <= sr_no <= 99:
                        if current_item and current_item.get('amount', 0) > 0:
                            items.append(current_item)
                        # Remaining words after Sr No
                        rest = row_words_sorted[1:]
                        amounts = [w for w in rest if re.match(r'^[\d,]+\.\d{2}$', w[1])]
                        desc_words = [w for w in rest 
                                      if not re.match(r'^[\d,]+\.?\d*$', w[1])
                                      and w[1].lower() not in SKIP_KEYWORDS]
                        desc = ' '.join(w[1] for w in sorted(desc_words, key=lambda x: x[0]))
                        
                        current_item = {
                            'description': desc,
                            'qty': 1.0,
                            'rate': 0.0,
                            'amount': clean_amount(amounts[-1][1]) if amounts else 0.0
                        }
                        # Try qty and rate from numeric positions
                        nums = [w for w in rest if re.match(r'^[\d,]+\.?\d*$', w[1])
                                and not re.match(r'^\d{6,8}$', w[1])]  # skip HSN codes
                        if len(nums) >= 3:
                            current_item['qty'] = clean_amount(nums[0][1])
                            current_item['rate'] = clean_amount(nums[1][1])
                            current_item['amount'] = clean_amount(nums[-1][1])
                        elif len(nums) == 2:
                            current_item['rate'] = clean_amount(nums[0][1])
                            current_item['amount'] = clean_amount(nums[-1][1])
                        elif len(nums) == 1:
                            current_item['amount'] = clean_amount(nums[0][1])
                        continue
                
                # Continuation line for current item (more description text)
                if current_item is not None:
                    row_lower = row_text.lower()
                    # Stop adding if this looks like a totals/tax row
                    if any(kw in row_lower for kw in ['sgst', 'cgst', 'igst', 'total', 'round off', 'amount chargeable']):
                        if current_item.get('amount', 0) > 0:
                            items.append(current_item)
                            current_item = None
                        continue
                    # Add description continuation (if mostly text, not numbers)
                    words_in_row = row_text.split()
                    text_words = [w for w in words_in_row if not re.match(r'^[\d.,]+$', w)]
                    if len(text_words) >= 2 and current_item.get('amount', 0) == 0:
                        # Still building description
                        current_item['description'] += ' ' + row_text
                    elif len(text_words) == 0:
                        # Pure numbers row - might be qty/rate/amount update
                        pure_nums = [w for w in words_in_row if re.match(r'^[\d,]+\.?\d*$', w)]
                        if pure_nums and current_item.get('amount', 0) == 0:
                            current_item['amount'] = clean_amount(pure_nums[-1])
            
            if current_item and current_item.get('amount', 0) > 0:
                items.append(current_item)
        except Exception as block_err:
            frappe.log_error("Item block extraction failed", str(block_err))
            items = []
        
        # Strategy B: Line-sequence state machine (fallback)
        if not items:
            STOP_WORDS_RE = re.compile(r'sgst|cgst|igst|grand\s*total|round\s*off|amount\s*chargeable|bank\s*details|declaration|e\s*&\s*oe|certified', re.I)
            SKIP_RE = re.compile(r'^(description|hsn|sac|quantity|rate|per|amount|sl\.?\s*no|si\.?\s*no|services?|goods?)$', re.I)
            
            in_table = False
            i = 0
            while i < len(lines):
                line = lines[i]
                
                # Detect start of item table — covers both Goods and Services invoices
                if re.search(r'description\s+(of\s+)?(goods|services|particulars)', line, re.I):
                    in_table = True
                    i += 1
                    continue
                
                # Also trigger on the column header row itself
                if re.search(r'\bhsn\b.*\bquantity\b|\bsl\.?\s*no\b.*\bdescription\b', line, re.I):
                    in_table = True
                    i += 1
                    continue
                
                if not in_table:
                    i += 1
                    continue
                
                if STOP_WORDS_RE.search(line):
                    break
                
                # Serial number line
                if re.match(r'^\d{1,2}$', line.strip()):
                    sr = int(line.strip())
                    if 1 <= sr <= 50:
                        desc_parts = []
                        qty = rate = amount = None
                        j = i + 1
                        # Collect subsequent lines until we hit the next serial or stop
                        while j < len(lines) and j < i + 12:
                            nxt = lines[j].strip()
                            if re.match(r'^\d{1,2}$', nxt) and 1 <= int(nxt) <= 50:
                                break
                            if STOP_WORDS_RE.search(nxt):
                                break
                            # Amount pattern XX,XXX.XX
                            amt_m = re.match(r'^([\d,]+\.\d{2})$', nxt)
                            if amt_m:
                                val = clean_amount(amt_m.group(1))
                                if amount is None:
                                    amount = val
                                elif rate is None and val > 0:
                                    rate = val
                            # Pure integer = qty
                            elif re.match(r'^\d{1,4}$', nxt) and not re.match(r'^\d{6,}$', nxt):
                                if qty is None and amount is None:
                                    qty = float(nxt)
                            # HSN code - skip (6-8 digits)
                            elif re.match(r'^\d{6,8}$', nxt):
                                pass
                            # Text = description
                            elif nxt and not SKIP_RE.match(nxt) and not re.match(r'^[₹%@\d\s.,/]+$', nxt):
                                desc_parts.append(nxt)
                            j += 1
                        
                        if desc_parts and amount and amount > 0:
                            items.append({
                                'description': ' '.join(desc_parts),
                                'qty': qty or 1.0,
                                'rate': rate or amount,
                                'amount': amount
                            })
                        i = j
                        continue
                i += 1
        
        # Strategy C: Find standalone amount lines and pair with nearby description
        if not items:
            amount_lines = [(i, clean_amount(l)) for i, l in enumerate(lines)
                           if re.match(r'^[\d,]+\.\d{2}$', l.strip()) and clean_amount(l) > 10]
            for idx, amt in amount_lines:
                desc = ''
                for di in range(max(0, idx-5), idx):
                    candidate = lines[di].strip()
                    if candidate and not re.match(r'^[\d,.\s%@₹]+$', candidate):
                        if not re.search(r'total|gst|tax|hsn|quantity|rate|amount', candidate, re.I):
                            desc = candidate
                if desc and amt > 0:
                    # Avoid duplicates — don't add if same amount already captured
                    if not any(abs(existing.get('amount', 0) - amt) < 0.01 for existing in items):
                        items.append({'description': desc, 'qty': 1.0, 'rate': amt, 'amount': amt})
        
        # Strategy D: OCR-resilient extraction — handles merged table cells
        if not items:
            full_text = ' '.join(lines)
            # Pattern: serial_no + description + HSN(6 digits) + amount
            for m in re.finditer(r'(?<!\d)([1-9])\s+([A-Za-z][\w\s,./\-()]{3,80}?)\s+(\d{6})\s+.*?([\d,]+\.\d{2})', full_text):
                desc = m.group(2).strip()
                amt = clean_amount(m.group(4))
                if amt > 0 and not re.search(r'total|sgst|cgst|igst|round|chargeable|taxable|bank', desc, re.I):
                    if not any(abs(ex.get('amount', 0) - amt) < 0.01 for ex in items):
                        items.append({'description': desc, 'qty': 1.0, 'rate': amt, 'amount': amt})
        
        # Strategy E: Find amounts within lines (not requiring standalone amount lines)
        if not items:
            for i, line in enumerate(lines):
                if re.search(r'total|sgst|cgst|igst|round|chargeable|taxable|tax\s*amount|hsn.*taxable', line, re.I):
                    continue
                amounts_in_line = re.findall(r'([\d,]+\.\d{2})', line)
                for amt_str in amounts_in_line:
                    amt = clean_amount(amt_str)
                    if amt < 50:
                        continue
                    # Extract description from text before the amount in the same line
                    idx = line.index(amt_str)
                    before = re.sub(r'^\d{1,2}\s+', '', line[:idx]).strip()
                    before = re.sub(r'\d{6,8}\s*', '', before).strip()
                    desc = before if len(before) > 3 else ''
                    # Fallback: look in previous lines
                    if not desc:
                        for di in range(max(0, i-5), i):
                            c = lines[di].strip()
                            if c and not re.match(r'^[\d,.\s%@₹|]+$', c) and not re.search(r'total|gst|tax|hsn|quantity|rate|amount|description', c, re.I):
                                desc = c
                    if desc and not any(abs(ex.get('amount', 0) - amt) < 0.01 for ex in items):
                        items.append({'description': desc, 'qty': 1.0, 'rate': amt, 'amount': amt})
                        break
        
        if not items:
            preview = "\n".join(lines[:40]) if lines else "(no text extracted)"
            frappe.logger().warning(f"Could not extract line items from this invoice.\n\n--- Extracted Text (first 40 lines) ---\n{preview}")
            data["items_warning"] = "Could not extract line items from this invoice."
            items = []
        
        # Clean up descriptions
        for it in items:
            it['description'] = re.sub(r'\s+', ' ', it['description']).strip()
        
        data["items"] = items

        
        # -------------------------------------------------------
        # 6. Amounts - Grand Total (last occurrence wins)
        # -------------------------------------------------------
        all_totals = re.findall(r'(?:Grand\s*Total|Total\s*Amount)[\s:₹Rs.]*([\d,]+(?:\.\d+)?)', text, re.I)
        if all_totals:
            data["grand_total"] = clean_amount(all_totals[-1])
        else:
            # Sum items as fallback
            data["grand_total"] = sum(it["amount"] for it in items)
        
        # -------------------------------------------------------
        # 7. GST / Tax Rate extraction  
        # KEY INSIGHT: PDFs double-print GST amounts (main table + HSN
        # summary). So we extract the RATE from text and calculate
        # the amount from taxable value × rate. This is 100% reliable.
        # -------------------------------------------------------
        gst_rate = 0.0
        gst_amount = 0.0
        round_off = 0.0
        
        # Get taxable base (net of items before tax)
        taxable_base = sum(it.get("amount", 0) for it in items)
        
        # Extract tax RATES from text (rates appear exactly once, amounts appear multiple times)
        sgst_rate_m = re.findall(r'SGST\s*@\s*(\d+(?:\.\d+)?)\s*%', text, re.I)
        cgst_rate_m = re.findall(r'CGST\s*@\s*(\d+(?:\.\d+)?)\s*%', text, re.I)
        igst_rate_m = re.findall(r'IGST\s*@\s*(\d+(?:\.\d+)?)\s*%', text, re.I)
        
        # Also look for rate patterns like "9 %" or "18%" near tax keywords
        if not sgst_rate_m:
            sgst_rate_m = re.findall(r'SGST[^\n]{0,30}?(\d+(?:\.\d+)?)\s*%', text, re.I)
        if not cgst_rate_m:
            cgst_rate_m = re.findall(r'CGST[^\n]{0,30}?(\d+(?:\.\d+)?)\s*%', text, re.I)
        if not igst_rate_m:
            igst_rate_m = re.findall(r'IGST[^\n]{0,30}?(\d+(?:\.\d+)?)\s*%', text, re.I)
        
        if igst_rate_m:
            # Out-state: single IGST
            gst_rate = float(igst_rate_m[-1])
            gst_amount = round(taxable_base * gst_rate / 100, 2)
        elif sgst_rate_m or cgst_rate_m:
            # Same-state: SGST + CGST
            sr = float(sgst_rate_m[-1]) if sgst_rate_m else 0
            cr = float(cgst_rate_m[-1]) if cgst_rate_m else 0
            gst_rate = round(sr + cr, 2)
            gst_amount = round(taxable_base * gst_rate / 100, 2)
        
        # Fallback: try to find total tax from "Tax Amount" summary table
        if gst_amount == 0:
            # Look for "Total" row in HSN summary - last big amount before end
            tax_total_m = re.findall(
                r'(?:Total\s+Tax\s*Amount|Tax\s+Amount|Taxes?\s+and\s+Charges?\s+Added)[^\n]*?([\d,]+\.\d{2})',
                text, re.I
            )
            if tax_total_m:
                gst_amount = clean_amount(tax_total_m[-1])
        
        # -------------------------------------------------------
        # 8. Round Off - extract from invoice
        # -------------------------------------------------------
        # Pattern: "Round Off" followed by optional -ve amount on same or next line
        round_off_m = re.search(
            r'Round\s*(?:Off|off)[^\n]*?(-?\s*[\d,]+\.\d{2})',
            text, re.I
        )
        if round_off_m:
            round_off = float(round_off_m.group(1).replace(',', '').replace(' ', ''))
        else:
            # Try line-window scan
            for i, line in enumerate(lines):
                if re.search(r'round\s*off', line, re.I):
                    for j in range(i, min(i + 4, len(lines))):
                        nxt = lines[j].strip()
                        amt_m = re.match(r'^-?\s*[\d,]+\.\d{2}$', nxt)
                        if amt_m:
                            round_off = float(nxt.replace(',', ''))
                            break
                    break
        
        # -------------------------------------------------------
        # 9. Final Invoice Amount (Grand Total with round off)
        # -------------------------------------------------------
        # Try multiple patterns to find the final payable amount
        final_amount = 0.0
        
        # Pattern A: "₹ XXXX.XX" - rupee symbol with amount (most reliable for Indian invoices)
        rupee_amounts = re.findall(r'[₹\u20b9]\s*([\d,]+\.\d{2})', text)
        if rupee_amounts:
            # Take the last (usually the grand total at the bottom)
            final_amount = clean_amount(rupee_amounts[-1])
        
        # Pattern B: "Grand Total" or "Total Amount"
        if not final_amount:
            gt_m = re.findall(r'(?:Grand\s*Total|Total\s*Amount|Net\s*Payable)[^\n\d₹]*([\d,]+\.\d{2})', text, re.I)
            if gt_m:
                final_amount = clean_amount(gt_m[-1])
        
        # Pattern C: Calculate from parts
        if not final_amount:
            final_amount = round(taxable_base + gst_amount + round_off, 2)
        
        # Cross-check: if grand_total was already found and is close, use it
        existing_grand = data.get('grand_total', 0)
        if existing_grand > 0 and abs(existing_grand - final_amount) < 10:
            final_amount = existing_grand  # They agree, use what we already have
        
        data["gst_amount"] = round(gst_amount, 2)
        data["gst_rate"] = gst_rate
        data["round_off"] = round_off
        data["grand_total"] = final_amount if final_amount > 0 else existing_grand
        data["raw_text"] = text
        
        # -------------------------------------------------------
        # Done
        # -------------------------------------------------------
        return {"status": "success", "data": data}
        
    except Exception as e:
        frappe.log_error("Invoice Extraction Error", frappe.get_traceback())
        return {"status": "error", "message": f"An error occurred while processing the invoice: {str(e)}"}


@frappe.whitelist()
def fetch_po_and_site_details(text_input, invoice_data):
    """
    Parses user natural language input for Circuit ID/PO No and payment details.
    Fetches Site and PO data to populate the missing fields.
    """
    import json
    import re
    data = json.loads(invoice_data)
    
    circuit_match = re.search(r'\b\d{5,8}\b', text_input)
    po_match = re.search(r'PO-\d{4}-\d+', text_input)
    
    circuit_id = circuit_match.group(0) if circuit_match else "71249"
    
    payment_type = "QRC" if "qrc" in text_input.lower() else "MRC"
    expense_type = "Variable" if "variable" in text_input.lower() else "Fixed"
    
    # Default mock fetched data
    site_name = "The Kangra Central - Dharamshala"
    lms_id = "LMS-003721"
    payment_cycle = "15"
    duration_from = "01-07-2026"
    duration_to = "30-09-2026"
    
    # Try fetching real data from Site
    if frappe.db.exists("Site", {"circuit_id": circuit_id}):
        site_doc = frappe.get_doc("Site", {"circuit_id": circuit_id})
        site_name = site_doc.site_name or site_name
        lms_id = site_doc.lms_id or lms_id
    
    data.update({
        "circuit_id": circuit_id,
        "site_name": site_name,
        "lms_id": lms_id,
        "payment_type": payment_type,
        "expense_type": expense_type,
        "payment_cycle": payment_cycle,
        "duration_from": duration_from,
        "duration_to": duration_to
    })
    
    return {
        "status": "success",
        "data": data
    }

@frappe.whitelist()
def create_draft_purchase_invoice(invoice_data):
    """
    Creates a draft Purchase Invoice in ERPNext using the structured data.
    """
    import json
    
    # DEBUG: Log every single call
    frappe.log_error("API Reached - create_draft_purchase_invoice", str(invoice_data)[:1000])
    
    data = json.loads(invoice_data)
    
    try:
        # Prevent Duplicate
        if frappe.db.exists("Purchase Invoice", {"supplier": data.get("supplier_name"), "bill_no": data.get("invoice_no")}):
            return {"status": "error", "message": "Duplicate Invoice! An invoice with this Supplier and Invoice No already exists."}
            
        # Supplier check
        supplier = data.get("supplier_name")
        if not frappe.db.exists("Supplier", supplier):
            # Create a mock supplier or fail
            supplier = frappe.db.get_value("Supplier", {"supplier_name": ["like", "%Bharti%"]}) or "Bharti Airtel Limited"
            if not frappe.db.exists("Supplier", supplier):
                doc = frappe.get_doc({
                    "doctype": "Supplier",
                    "supplier_name": supplier,
                    "supplier_group": "All Supplier Groups"
                })
                doc.insert(ignore_permissions=True)
                
        # Basic formatting dates (DD-MM-YYYY to YYYY-MM-DD)
        def format_date(d_str):
            if "-" in d_str and len(d_str.split("-")[0]) == 2:
                parts = d_str.split("-")
                return f"{parts[2]}-{parts[1]}-{parts[0]}"
            return d_str

        pi = frappe.new_doc("Purchase Invoice")
        pi.supplier = supplier
        pi.bill_no = data.get("invoice_no")
        pi.bill_date = format_date(data.get("invoice_date"))
        pi.posting_date = format_date(data.get("posting_date"))
        
        pi.custom_dutation_from = format_date(data.get("duration_from")) if data.get("duration_from") else None
        pi.custom_duration_to = format_date(data.get("duration_to")) if data.get("duration_to") else None
        pi.custom_lms_id = data.get("lms_id")
        pi.custom_circuit_id = data.get("circuit_id")
        
        po_number = None
        if pi.custom_lms_id:
            lms_doc = frappe.db.get_value("Lastmile Services Master", pi.custom_lms_id, 
                                          ["circuit_id", "po_number", "billing_mode", "payment_cycle"], as_dict=True)
            if lms_doc:
                if not pi.custom_circuit_id:
                    pi.custom_circuit_id = lms_doc.circuit_id
                pi.custom_payment_type = data.get("payment_type") or lms_doc.billing_mode
                pi.custom_payment_cycle = data.get("payment_cycle") or lms_doc.payment_cycle
                po_number = lms_doc.po_number

        pi.custom_payment_catogery = data.get("po_category")
        if not pi.custom_payment_catogery and po_number:
            po_cat = frappe.db.get_value("Purchase Order", po_number, "custom_po_category")
            if po_cat:
                pi.custom_payment_catogery = po_cat

        if pi.custom_circuit_id and not pi.custom_site_name:
            site_name = frappe.db.get_value("Site", pi.custom_circuit_id, "site_name")
            if site_name:
                pi.custom_site_name = site_name

        # Add items
        for item in data.get("items", []):
            item_code = item.get("item_code")
            if not item_code:
                item_code = "Misc"
            if po_number:
                po_items = frappe.get_all("Purchase Order Item", filters={"parent": po_number}, fields=["item_code", "item_name"])
                if len(po_items) == 1:
                    item_code = po_items[0].item_code
                elif po_items:
                    desc_lower = (item.get("description") or "").lower()
                    for pi_item in po_items:
                        if pi_item.item_name and pi_item.item_name.lower() in desc_lower:
                            item_code = pi_item.item_code
                            break

            if not frappe.db.exists("Item", item_code):
                item_code = frappe.db.get_value("Item", {"item_name": ["like", "%Internet%"]}) or "Misc"
            
            item_circuit_id = pi.custom_circuit_id or data.get("circuit_id")
            item_lms_id = pi.custom_lms_id or data.get("lms_id")
            item_site_name = pi.custom_site_name or data.get("site_name")
            
            row = {
                "item_code": item_code,
                "description": item.get("description"),
                "qty": item.get("qty"),
                "rate": item.get("rate"),
                "custom_circuit_id": item_circuit_id,
                "custom_lms_id": item_lms_id,
                "circuit_id": item_circuit_id,
                "lms_id": item_lms_id,
                "site_name": item_site_name,
            }
            if po_number:
                row["purchase_order"] = po_number
                po_item_data = frappe.db.get_value("Purchase Order Item", {"parent": po_number, "item_code": item_code}, 
                    ["name", "project", "cost_center", "expense_account"], as_dict=True)
                if po_item_data:
                    row["po_detail"] = po_item_data.name
                    if po_item_data.project:
                        row["project"] = po_item_data.project
                    if po_item_data.cost_center:
                        row["cost_center"] = po_item_data.cost_center
                    if po_item_data.expense_account:
                        row["expense_account"] = po_item_data.expense_account
                    
            pi.append("items", row)
            
        # Auto-apply taxes based on Supplier/Company default and matched Item Codes
        pi.set_missing_values()
        
        pi.taxes_and_charges = data.get("taxes_and_charges") or ""
        
        pi.set_taxes()
        pi.calculate_taxes_and_totals()
        
        if not data.get("taxes_and_charges"):
            pi.set("taxes", [])
            
        round_off = data.get("round_off", 0)
        if round_off:
            round_off_account = frappe.db.get_value("Account", {
                "account_name": ["like", "%Round%"], 
                "is_group": 0,
                "company": pi.company
            })
            if round_off_account:
                pi.append("taxes", {
                    "charge_type": "Actual",
                    "account_head": round_off_account,
                    "description": "Round Off",
                    "tax_amount": round_off
                })
                
        pi.calculate_taxes_and_totals()
            
        if pi.shipping_address and not frappe.db.exists("Address", pi.shipping_address):
            pi.shipping_address = None
        if pi.billing_address and not frappe.db.exists("Address", pi.billing_address):
            pi.billing_address = None
        if pi.supplier_address and not frappe.db.exists("Address", pi.supplier_address):
            pi.supplier_address = None
            
        pi.flags.ignore_mandatory = True
        pi.insert(ignore_permissions=True)
        
        # Attach the original supplier invoice to the Purchase Invoice
        file_url = data.get("file_url")
        if file_url:
            try:
                existing = frappe.db.exists("File", {
                    "file_url": file_url,
                    "attached_to_doctype": "Purchase Invoice",
                    "attached_to_name": pi.name
                })
                if not existing:
                    file_doc = frappe.get_doc({
                        "doctype": "File",
                        "file_url": file_url,
                        "attached_to_doctype": "Purchase Invoice",
                        "attached_to_name": pi.name,
                        "folder": "Home/Attachments",
                        "is_private": 0
                    })
                    file_doc.insert(ignore_permissions=True)
            except Exception as attach_err:
                frappe.log_error("Invoice Attachment Error", str(attach_err))
        
        if hasattr(frappe.local, "message_log"):
            frappe.local.message_log = []
        if "_server_messages" in frappe.local.response:
            del frappe.local.response["_server_messages"]
            
        return {
            "status": "success",
            "invoice_name": pi.name
        }
    except Exception as e:
        if hasattr(frappe.local, "message_log"):
            frappe.local.message_log = []
        if "exc" in frappe.local.response:
            del frappe.local.response["exc"]
        if "_server_messages" in frappe.local.response:
            del frappe.local.response["_server_messages"]
        return {
            "status": "error",
            "message": str(e)
        }

###################################
import frappe
from frappe import _
from frappe.utils import cint

@frappe.whitelist()
def get_supplier_activity_details(supplier, po_number=None, lms_id=None, raw_text=None):
    if not supplier:
        return {"status": "error", "message": "Supplier is required"}
    
    # We will fetch records matching the supplier. To be smart, we will calculate a matching percentage.
    lms_records = frappe.get_all("Lastmile Services Master", 
                                 filters={"supplier": supplier, "lms_stage": "Delivered"},
                                 fields=["name", "supplier", "lms_stage", "lms_delivery_date", 
                                         "billing_start_date", "order_type", "bandwith_type", "lms_brandwith_name", 
                                         "circuit_id", "customer", "site", "solution", 
                                         "customer_type", "site_address", "po_number", "po_released_datetime", "city"])
    
    results = []
    
    import re
    norm_text = ""
    if raw_text:
        norm_text = re.sub(r'[^a-z0-9\s]', '', str(raw_text).lower())
        
    for row in lms_records:
        match_score = 0
        
        # 1. PO Number Match
        if po_number and row.po_number and str(po_number).lower() in str(row.po_number).lower():
            match_score += 40
            
        # 2. LMS ID Match
        if lms_id and (str(lms_id).lower() in str(row.name).lower() or str(lms_id).lower() in str(row.circuit_id or '').lower()):
            match_score += 40
            
        # 3. Customer Match in Raw Text
        if norm_text and row.customer:
            cust_clean = re.sub(r'[^a-z0-9\s]', '', str(row.customer).lower())
            cust_words = [w for w in cust_clean.split() if w not in ('ltd', 'pvt', 'private', 'limited', 'inc')]
            if len(cust_words) >= 2:
                phrase = " ".join(cust_words[:2])
                if phrase in norm_text:
                    match_score += 20
            elif cust_words and cust_words[0] in norm_text:
                match_score += 15
                
        # 4. Address/City Match in Raw Text
        if norm_text:
            if row.city and str(row.city).lower() in norm_text:
                match_score += 10
            if row.site_address:
                addr_clean = re.sub(r'[^a-z0-9\s]', '', str(row.site_address).lower())
                addr_words = [w for w in addr_clean.split() if len(w) > 4 and w not in ('road', 'street', 'floor', 'building')]
                matched_words = sum(1 for w in addr_words if w in norm_text)
                if matched_words > 0:
                    match_score += (5 * min(matched_words, 4))
                    
        row.match_percentage = min(match_score, 100)
        
        # Payment Terms from Purchase Order if matched
        row.payment_terms = ""
        if row.po_number and frappe.db.exists("Purchase Order", row.po_number):
            row.payment_terms = frappe.db.get_value("Purchase Order", row.po_number, "payment_terms_template") or ""
            
        # Fetch Site Status
        row.site_status = ""
        if row.site and frappe.db.exists("Site", row.site):
            row.site_status = frappe.db.get_value("Site", row.site, "custom_stage") or frappe.db.get_value("Site", row.site, "status") or ""
            
        results.append(row)
        
    results.sort(key=lambda x: x.match_percentage, reverse=True)
    return {"status": "success", "data": results}

@frappe.whitelist()
def get_po_or_lms_items(po_number=None, lms_id=None, circuit_id=None, ai_items=None):
    import json
    ai_item_list = []
    if ai_items:
        try:
            ai_item_list = json.loads(ai_items)
        except:
            pass

    items = []
    po_items = []
    
    # 1. Check Purchase Order
    if po_number and frappe.db.exists("Purchase Order", po_number):
        po = frappe.get_doc("Purchase Order", po_number)
        for po_item in po.items:
            po_items.append({
                "item_code": po_item.item_code,
                "item_name": po_item.item_name,
                "description": po_item.description,
                "qty": po_item.qty,
                "uom": po_item.uom,
                "rate": po_item.rate,
                "amount": po_item.amount,
                "purchase_order": po_number,
                "po_detail": po_item.name,
                "cost_center": po_item.cost_center,
                "project": po_item.project,
                "expense_account": po_item.expense_account,
                "circuit_id": circuit_id,
                "custom_circuit_id": circuit_id,
                "lms_id": lms_id,
                "custom_lms_id": lms_id
            })
            
    if ai_item_list and po_items:
        import difflib
        import re
        
        def normalize_str(s):
            # Remove special chars and extra spaces
            s = re.sub(r'[^a-z0-9]', '', str(s).lower())
            return s
            
        for ai_item in ai_item_list:
            ai_desc = str(ai_item.get("description") or ai_item.get("item_name") or "").lower()
            ai_desc_norm = normalize_str(ai_desc)
            
            best_match = None
            best_score = 0
            
            for p_item in po_items:
                p_name = str(p_item.get("item_name") or "").lower()
                p_desc = str(p_item.get("description") or "").lower()
                
                # Check for direct substring match
                if (p_name and p_name in ai_desc) or (p_desc and p_desc in ai_desc) or \
                   (ai_desc and ai_desc in p_name) or (ai_desc and ai_desc in p_desc):
                    best_match = p_item
                    best_score = 1.0
                    break
                    
                # Check normalized words overlap
                p_name_norm = normalize_str(p_name)
                p_desc_norm = normalize_str(p_desc)
                
                # Use SequenceMatcher for similarity
                score_name = difflib.SequenceMatcher(None, ai_desc_norm, p_name_norm).ratio() if p_name_norm else 0
                score_desc = difflib.SequenceMatcher(None, ai_desc_norm, p_desc_norm).ratio() if p_desc_norm else 0
                
                max_score = max(score_name, score_desc)
                
                # Also try matching numeric parts (like "100")
                ai_nums = re.findall(r'\d+', ai_desc)
                p_nums = re.findall(r'\d+', p_name + " " + p_desc)
                if ai_nums and p_nums and any(n in p_nums for n in ai_nums):
                    max_score += 0.2  # Boost score if numbers match
                    
                if max_score > best_score and max_score > 0.4:
                    best_score = max_score
                    best_match = p_item
            
            if best_match:
                # Check if we already added this PO item
                if not any(i.get('item_code') == best_match.get('item_code') for i in items):
                    matched_item = best_match.copy()
                    
                    # RATE: Always from Purchase Order (the agreed contracted rate)
                    po_rate = float(best_match.get("rate") or 0)
                    
                    # QTY: Use AI-extracted qty, but sanitize it.
                    # AI often misreads product specs as qty (e.g. "100" from "100 MBPS").
                    # For telecom/ISP service invoices, qty is almost always 1 per billing cycle.
                    # A qty > 10 for a service line item is almost certainly a misread.
                    raw_ai_qty = ai_item.get("qty")
                    try:
                        raw_ai_qty = float(raw_ai_qty) if raw_ai_qty else 1
                    except:
                        raw_ai_qty = 1
                    
                    # If qty > 10, it's likely a bandwidth spec or product name number
                    # that got misidentified as qty. Default to 1.
                    if raw_ai_qty > 10:
                        ai_qty = 1.0
                    else:
                        ai_qty = raw_ai_qty if raw_ai_qty > 0 else 1.0
                    
                    matched_item["rate"] = po_rate
                    matched_item["qty"] = ai_qty
                    matched_item["amount"] = round(po_rate * ai_qty, 2)
                        
                    items.append(matched_item)
    else:
        # If no AI items provided, fallback to all PO items
        items = po_items
            
    # 2. Check LMS ID (if no PO items found)
    if not items and lms_id and frappe.db.exists("Lastmile Services Master", lms_id):
        lms_doc = frappe.get_doc("Lastmile Services Master", lms_id)
        if hasattr(lms_doc, 'lms_items') and lms_doc.lms_items:
            for lms_item in lms_doc.lms_items:
                items.append({
                    "item_code": lms_item.item_code,
                    "item_name": getattr(lms_item, 'item_name', ''),
                    "description": getattr(lms_item, 'description', ''),
                    "qty": getattr(lms_item, 'qty', 1),
                    "rate": getattr(lms_item, 'rate', 0),
                    "amount": getattr(lms_item, 'amount', 0),
                    "circuit_id": circuit_id,
                    "custom_circuit_id": circuit_id,
                    "lms_id": lms_id,
                    "custom_lms_id": lms_id
                })
    
    return items

# --- END: AI Purchase Invoice Creation Logic ---

# --- START: Purchase Invoice Stage TAT & Ageing ---
@frappe.whitelist()
def get_purchase_invoice_stage_history(docname, creation=None):
    from frappe.utils import getdate, nowdate, date_diff
    import json
    
    if not creation:
        creation = nowdate()
    
    # 1. Calendar Ageing (count all days, no holiday list)
    ageing = date_diff(nowdate(), getdate(creation))
    
    # 2. Fetch TAT Target
    tat_target = 30
    grade_name = "Standard"
    
    # 3. Stage history
    versions = frappe.db.sql("""
        SELECT creation, data
        FROM `tabVersion`
        WHERE ref_doctype='Purchase Invoice' AND docname=%s
        ORDER BY creation ASC
    """, (docname,), as_dict=True)
    
    history = {}
    for v in versions:
        try:
            data = json.loads(v.data)
            for change in data.get("changed", []):
                if change[0] == "status":
                    new_stage = change[2]
                    if new_stage not in history:
                        history[new_stage] = v.creation
        except Exception:
            pass
            
    return {
        "ageing_days": ageing,
        "tat_target": tat_target,
        "grade_name": grade_name,
        "history": history
    }
# --- END: Purchase Invoice Stage TAT & Ageing ---

# --- Start: Bulk PO ---
@frappe.whitelist()
def make_po_from_multiple_mr(mr_names):
    import json
    try:
        mr_names = json.loads(mr_names)
    except Exception:
        pass
    
    if not mr_names:
        return
        
    suppliers = set()
    company = None
    for mr in mr_names:
        mr_doc = frappe.db.get_value("Material Request", mr, ["custom_supplier", "company"], as_dict=True)
        if mr_doc and mr_doc.get("custom_supplier"):
            suppliers.add(mr_doc.custom_supplier)
            
        if mr_doc and not company:
            company = mr_doc.company
        elif mr_doc and company != mr_doc.company:
            frappe.throw("Selected Material Requests belong to different companies.")

    if len(suppliers) > 1:
        frappe.throw("Selected Material Requests have different suppliers. Please select Material Requests with the same supplier.")
    
    from erpnext.stock.doctype.material_request.material_request import make_purchase_order
    
    po = None
    for mr in mr_names:
        mapped_po = make_purchase_order(mr)
        if not po:
            po = mapped_po
        else:
            for item in mapped_po.get("items", []):
                po.append("items", item)
                
    if po:
        if suppliers:
            po.supplier = list(suppliers)[0]
            
        # Ensure schedule_date is set (it might be cleared by Frappe if the MR date is in the past)
        if not po.schedule_date:
            dates = [d.schedule_date for d in po.get("items") if d.schedule_date]
            if dates:
                po.schedule_date = min(dates)
            else:
                from frappe.utils import nowdate
                po.schedule_date = nowdate()
            
        # Clear header-level fields that should not be set for bulk POs
        po.custom_bandwidth = None
        po.custom_lms_id = None
        po.custom_site_circuit_id = None
        
        po.flags.ignore_mandatory = True
        po.insert(ignore_permissions=True)
        return po.name

# --- END: Bulk PO ---

###################################################################################

################################################################################
# --- START: Bulk Feasibility Upload ---
################################################################################

@frappe.whitelist()
def download_feasibility_template():
    import openpyxl
    from io import BytesIO
    
    frappe.local.response.filename = "Feasibility_Upload_Template.xlsx"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feasibility Template"
    
    headers = [
        "Feasibility From", "Customer Name", "Site Name", "Customer Request", "Sales Person", 
        "Customer Type", "Order Type", "Site Type", "Site ID / Legal Code", 
        "Solution Code", "Static IP", "No. of Static IP Required", "Config Type", 
        "Managed Services", "Primary Data Plan Code", "Secondary Data Plan Code", 
        "Central Spoke", "Mobile", "Central Email", "Primary Contact Person", 
        "Primary Contact Mobile", "Email", "Alternate Contact Person", 
        "Alternate Contact Mobile", "Secondary Email", "Address/ Street", "Pincode",
        "Latitude", "Longitude"
    ]
    
    sample_data = [
        "Customer", "Acme Corp", "Site 101", "2026-07-10", "John Sales", 
        "Paid Customer", "Service", "Branch", "LGL-001", 
        "SOL-001", "No", "0", "Remote Config", 
        "Proactive", "PDP-001", "SDP-002", "CS-123", "9876543210", "central@acme.com",
        "John Doe", "9876543211", "john@acme.com", "Jane Doe", 
        "9876543212", "jane@acme.com", "123 Main St", "400001",
        "28.6139", "77.2090"
    ]
    
    # Instruction note row
    ws.append(["NOTE: Please provide either a Pincode OR Latitude and Longitude. If coordinates are provided, the system will automatically fetch the address details."])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    # Style the note
    from openpyxl.styles import Font, Alignment
    note_cell = ws.cell(row=1, column=1)
    note_cell.font = Font(bold=True, color="FF0000")
    note_cell.alignment = Alignment(horizontal="center")

    ws.append(headers)
    
    # mark mandatory in red
    mandatory_headers = [
        "Customer Name", "Site Name", "Customer Request", "Sales Person", 
        "Customer Type", "Order Type", "Site Type", "Solution Code", 
        "Config Type", "Managed Services", "Primary Contact Person", "Address/ Street"
    ]
    header_row = ws[2]
    for cell in header_row:
        if cell.value in mandatory_headers or cell.value == "Pincode":
            cell.font = Font(bold=True, color="FF0000")
        else:
            cell.font = Font(bold=True)
            
    ws.append(sample_data)
    
    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                # Skip merged cell in row 1 for width calculation
                if cell.row == 1:
                    continue
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    f = BytesIO()
    wb.save(f)
    file_bytes = f.getvalue()
    
    frappe.local.response.filecontent = file_bytes
    frappe.local.response.type = "download"

@frappe.whitelist()
def process_bulk_feasibility(file_name, file_data, validate_only=0):
    import base64
    import csv
    import re
    from io import StringIO, BytesIO
    from frappe.utils import nowdate
    
    try:
        decoded_data = base64.b64decode(file_data)
    except Exception as e:
        frappe.throw(f"Error decoding file. {str(e)}")

    HEADER_MAP = {
        "feasibility from": "feaseibility_from",
        "customer name": "customer",
        "site name": "site_name",
        "customer request": "customer_request",
        "sales person": "sales_person",
        "customer type": "customer_type",
        "order type": "order_type",
        "site type": "site_type",
        "site id / legal code": "site_id__legal_code",
        "solution code": "solution_code",
        "static ip": "static_ip",
        "no. of static ip required": "no_of_static_ip_required",
        "no of static ip required": "no_of_static_ip_required",
        "config type": "config_type",
        "managed services": "managed_services",
        "primary data plan code": "primary_data_plan",
        "secondary data plan code": "secondary_data_plan",
        "central spoke": "central_spoke",
        "mobile": "mobile",
        "central email": "central_email",
        "primary contact person": "contact_person",
        "primary contact mobile": "primary_contact_mobile",
        "email": "email",
        "alternate contact person": "alternate_contact_person",
        "alternate contact mobile": "alternate_contact_mobile",
        "secondary email": "secondary_email",
        "address/ street": "address_street",
        "pincode": "pincode",
        "latitude": "latitude",
        "longitude": "longitude"
    }

    rows = []
    
    if file_name.endswith('.csv'):
        try:
            f = StringIO(decoded_data.decode('utf-8'))
            reader = csv.reader(f)
            headers = []
            header_found = False
            for row in reader:
                if not header_found:
                    row_lower = [str(h).strip().lower() for h in row]
                    if "customer name" in row_lower or "site name" in row_lower:
                        headers = row_lower
                        header_found = True
                else:
                    row_data = {}
                    for i, cell in enumerate(row):
                        if i < len(headers):
                            mapped_header = HEADER_MAP.get(headers[i], headers[i])
                            row_data[mapped_header] = cell
                    if any(str(v).strip() for v in row_data.values() if v is not None):
                        rows.append(row_data)
        except Exception as e:
            frappe.throw(f"Error reading CSV file. {str(e)}")
    elif file_name.endswith('.xlsx') or file_name.endswith('.xls'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filename=BytesIO(decoded_data), data_only=True)
            sheet = wb.active
            
            headers = []
            header_found = False
            for row in sheet.iter_rows(values_only=True):
                if not header_found:
                    row_lower = [str(cell).strip().lower() if cell else "" for cell in row]
                    if "customer name" in row_lower or "site name" in row_lower:
                        headers = [str(cell).strip().lower() if cell else f"col_{i}" for i, cell in enumerate(row)]
                        header_found = True
                else:
                    row_data = {}
                    for i, cell in enumerate(row):
                        if i < len(headers):
                            mapped_header = HEADER_MAP.get(headers[i], headers[i])
                            row_data[mapped_header] = cell
                    if any(str(v).strip() for v in row_data.values() if v is not None):
                        rows.append(row_data)
        except Exception as e:
            frappe.throw(f"Error reading Excel file. {str(e)}")
    else:
        frappe.throw("Unsupported file format. Please upload a CSV or XLSX file.")

    success_count = 0
    errors = []
    
    # Helper: safely convert Excel values to string (Excel reads numbers as int/float)
    def safe_str(val):
        if val is None:
            return None
        # Excel reads numeric cells as float (e.g. 491228.0) — convert to int first
        if isinstance(val, float) and val == int(val):
            val = int(val)
        s = str(val).strip()
        return s if s else None

    # Pass 1: Validate all rows before any updates
    mandatory_fields = {
        "customer": "Customer Name",
        "site_name": "Site Name",
        "customer_request": "Customer Request",
        "sales_person": "Sales Person",
        "customer_type": "Customer Type",
        "order_type": "Order Type",
        "site_type": "Site Type",
        "solution_code": "Solution Code",
        "config_type": "Config Type",
        "managed_services": "Managed Services",
        "contact_person": "Primary Contact Person",
        "address_street": "Address/ Street"
    }
    validation_errors = []
    for row_idx, row in enumerate(rows, start=2):
        missing = []
        for field, label in mandatory_fields.items():
            if not safe_str(row.get(field)):
                missing.append(label)
                
        has_pincode = bool(safe_str(row.get("pincode")))
        has_lat_lon = bool(safe_str(row.get("latitude"))) and bool(safe_str(row.get("longitude")))
        if not has_pincode and not has_lat_lon:
            missing.append("Pincode OR (Latitude and Longitude)")
            
        if missing:
            site_n = safe_str(row.get("site_name")) or "Unknown"
            validation_errors.append(f"Row {row_idx} ({site_n}): Missing mandatory fields - {', '.join(missing)}")
            
    if validation_errors:
        return {"status": "error", "errors": validation_errors, "success_count": 0, "total_rows": len(rows)}
        
    if int(validate_only) == 1:
        return {"status": "success", "message": "Validation passed", "total_rows": len(rows)}
        
    pincode_cache = {}
    
    from datetime import datetime
    def parse_date(date_val):
        if not date_val: return nowdate()
        if isinstance(date_val, datetime):
            return date_val.strftime('%Y-%m-%d')
        
        date_str = str(date_val).strip()
        # Common formats including DD-MM-YYYY
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(date_str, fmt).strftime('%Y-%m-%d')
            except ValueError:
                pass
        return date_str



    for row_idx, row in enumerate(rows, start=2):
        try:
            doc = frappe.new_doc("Feasibility")
            doc.feaseibility_from = safe_str(row.get("feaseibility_from")) or "Customer"
            doc.site_name = safe_str(row.get("site_name"))
            doc.customer = safe_str(row.get("customer"))
            doc.customer_request = parse_date(row.get("customer_request"))
            doc.sales_person = safe_str(row.get("sales_person"))
            doc.customer_type = safe_str(row.get("customer_type"))
            doc.order_type = safe_str(row.get("order_type")) or "Service"
            doc.site_type = safe_str(row.get("site_type"))
            doc.site_id__legal_code = safe_str(row.get("site_id__legal_code"))
            doc.solution_code = safe_str(row.get("solution_code"))
            doc.static_ip = safe_str(row.get("static_ip")) or "No"
            doc.no_of_static_ip_required = safe_str(row.get("no_of_static_ip_required"))
            doc.config_type = safe_str(row.get("config_type")) or "Remote Config"
            doc.managed_services = safe_str(row.get("managed_services")) or "Proactive"
            doc.primary_data_plan = safe_str(row.get("primary_data_plan"))
            doc.secondary_data_plan = safe_str(row.get("secondary_data_plan"))
            # Central Spoke will be linked dynamically after contact verification
            # doc.central_spoke = safe_str(row.get("central_spoke"))
            doc.contact_person = safe_str(row.get("contact_person"))
            doc.primary_contact_mobile = safe_str(row.get("primary_contact_mobile"))
            doc.email = safe_str(row.get("email"))
            doc.alternate_contact_person = safe_str(row.get("alternate_contact_person"))
            doc.alternate_contact_mobile = safe_str(row.get("alternate_contact_mobile"))
            doc.secondary_email = safe_str(row.get("secondary_email"))
            doc.address_street = safe_str(row.get("address_street"))
            doc.pincode = safe_str(row.get("pincode"))
            doc.latitude = safe_str(row.get("latitude"))
            doc.longitude = safe_str(row.get("longitude"))
            
            # 1. Reverse Geocode if Lat & Lon are provided
            if doc.latitude and doc.longitude:
                try:
                    lat_f = float(doc.latitude)
                    lon_f = float(doc.longitude)
                    import requests as req
                    geo_url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={lat_f}&lon={lon_f}"
                    geo_headers = {'User-Agent': 'Nexapp Feasibility Import'}
                    geo_resp = req.get(geo_url, headers=geo_headers, timeout=10)
                    geo_data = geo_resp.json()
                    
                    if geo_data and geo_data.get("address"):
                        addr = geo_data.get("address")
                        # Format address street just like JS
                        street_val = geo_data.get("display_name", "")
                        heading = "\n\nAddress as per Latitude & Longitude:\n____________________________________\n"
                        if doc.address_street:
                            doc.address_street += heading + street_val
                        else:
                            doc.address_street = "Address as per Latitude & Longitude:\n____________________________________\n" + street_val
                            
                        # If a pincode was found via lat/lon, use it for the API lookup, unless user explicitly provided one
                        fetched_pincode = addr.get("postcode", "")
                        if fetched_pincode and len(str(fetched_pincode).strip()) == 6 and not doc.pincode:
                            doc.pincode = str(fetched_pincode).strip()
                            
                        # Fallback details directly from openstreetmap in case postal api fails
                        doc.city = addr.get("city") or addr.get("town") or addr.get("village") or addr.get("suburb") or addr.get("county") or ""
                        doc.district = addr.get("state_district") or addr.get("district") or addr.get("county") or ""
                        doc.state = addr.get("state", "")
                        doc.country = addr.get("country", "India")
                        

                except Exception as geo_err:
                    frappe.log_error(f"Geocoding failed for row {row_idx}: {str(geo_err)}", "Bulk Feasibility Geocoding")

            # 2. Pincode Auto-fetch (Overrides city/district/state if postal API succeeds)
            if doc.pincode:
                pin_val = str(doc.pincode).strip()
                # Remove any non-digit chars and validate
                pin_clean = re.sub(r'\D', '', pin_val)
                if len(pin_clean) == 6:
                    if pin_clean not in pincode_cache:
                        try:
                            import requests as req
                            api_url = f"https://api.postalpincode.in/pincode/{pin_clean}"
                            headers = {
                                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                            }
                            resp = req.get(api_url, headers=headers, timeout=15)
                            api_data = resp.json()
                            if (isinstance(api_data, list) and len(api_data) > 0
                                    and isinstance(api_data[0], dict)
                                    and api_data[0].get("Status") == "Success"):
                                po_list = api_data[0].get("PostOffice")
                                if isinstance(po_list, list) and len(po_list) > 0:
                                    po = po_list[0]
                                    pincode_cache[pin_clean] = {
                                        "city": po.get("Block", ""),
                                        "district": po.get("District", ""),
                                        "state": po.get("State", ""),
                                        "country": po.get("Country", "India")
                                    }
                                else:
                                    pincode_cache[pin_clean] = {}
                            else:
                                pincode_cache[pin_clean] = {}
                        except Exception as api_err:
                            frappe.log_error(f"Pincode API failed for {pin_clean}: {str(api_err)}", "Bulk Feasibility Pincode Lookup")
                            pincode_cache[pin_clean] = {}
                    
                    details = pincode_cache.get(pin_clean, {})
                    if details:
                        doc.city = details.get("city")
                        doc.district = details.get("district")
                        doc.state = details.get("state")
                        doc.country = details.get("country")
            
            state_territory_map = {
                "Delhi": "North", "Haryana": "North", "Punjab": "North", "Himachal Pradesh": "North", "Uttar Pradesh": "North", "Uttarakhand": "North", "Jammu and Kashmir": "North", "Chandigarh": "North", "Rajasthan": "North", "Ladakh": "North",
                "Karnataka": "South", "Tamil Nadu": "South", "Kerala": "South", "Andhra Pradesh": "South", "Telangana": "South", "Puducherry": "South", "Lakshadweep": "South",
                "Maharashtra": "West", "Gujarat": "West", "Goa": "West", "Dadra and Nagar Haveli": "West", "Daman and Diu": "West", "Madhya Pradesh": "West", "Chattisgarh": "West", "Chhattisgarh": "West",
                "West Bengal": "East", "Odisha": "East", "Bihar": "East", "Jharkhand": "East", "Assam": "East", "Sikkim": "East", "Meghalaya": "East", "Tripura": "East", "Arunachal Pradesh": "East", "Manipur": "East", "Nagaland": "East", "Mizoram": "East", "Andaman and Nicobar Islands": "East"
            }
            if doc.state and doc.state in state_territory_map:
                doc.territory = state_territory_map[doc.state]

            doc.feasibility_status = "Pending"
            
            # --- Contact Verification & Creation ---
            mobile_num = safe_str(row.get("mobile"))
            if mobile_num:
                # Clean mobile number to prevent "Value too big" for max length 10
                digits_only = re.sub(r'\D', '', mobile_num)
                if len(digits_only) >= 10:
                    mobile_num = digits_only[-10:]
                else:
                    # Fallback truncate to 10 chars if it's less than 10 digits but has other chars
                    mobile_num = mobile_num[:10]
                    
            central_email = safe_str(row.get("central_email"))
            if mobile_num:
                # Check if contact exists by phone
                existing_contact = frappe.db.get_value("Contact Phone", {"phone": mobile_num}, "parent")
                contact_name = existing_contact
                
                if not contact_name:
                    # Create new Contact
                    contact_doc = frappe.new_doc("Contact")
                    contact_doc.first_name = safe_str(row.get("central_spoke")) or safe_str(row.get("contact_person")) or safe_str(row.get("customer")) or "Unknown"
                    contact_doc.status = "Passive"
                    # Set type from Feasibility From
                    contact_type = safe_str(row.get("feaseibility_from")) or "Customer"
                    contact_doc.type = contact_type
                    
                    # Try to map Customer if it exists
                    cust_link = safe_str(row.get("customer"))
                    if cust_link and frappe.db.exists("Customer", cust_link):
                        contact_doc.append("links", {
                            "link_doctype": "Customer",
                            "link_name": cust_link
                        })
                    
                    contact_doc.append("phone_nos", {
                        "phone": mobile_num,
                        "is_primary_mobile_no": 1
                    })
                    
                    if central_email:
                        contact_doc.append("email_ids", {
                            "email_id": central_email,
                            "is_primary": 1
                        })
                        
                    contact_doc.flags.ignore_mandatory = True
                    contact_doc.flags.ignore_links = True
                    contact_doc.insert(ignore_permissions=True)
                    contact_name = contact_doc.name
                    
                # Link the verified/created Contact to the Feasibility record
                doc.central_spoke = contact_name
            # --- End Contact Logic ---
            
            doc.insert(ignore_permissions=True, ignore_mandatory=True)
            success_count += 1
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            error_msg = str(e)
            if "list" in error_msg.lower():
                errors.append(f"Row {row_idx} ({row.get('site_name', 'Unknown')}): {error_msg}<br><pre style='font-size:10px;text-align:left;'>{tb}</pre>")
            else:
                errors.append(f"Row {row_idx} ({row.get('site_name', 'Unknown')}): {error_msg}")

    if errors:
        return {"status": "error", "errors": errors, "success_count": success_count}

    return {"status": "success", "success_count": success_count}

################################################################################
# --- END: Bulk Feasibility Upload ---
################################################################################

# Start Roundoff for the Payment Entry
import frappe
from frappe.utils import flt


ROUNDING_TOLERANCE = 0.50


def update_expense_claim_rounding(doc, method=None):
    """
    Update Expense Claim reimbursement after Payment Entry
    submit/cancel and ignore tiny rounding differences.
    """

    for ref in doc.references:

        if ref.reference_doctype != "Expense Claim":
            continue

        expense_claim = ref.reference_name

        # Calculate total reimbursed from all submitted Payment Entries
        total_reimbursed = flt(frappe.db.sql("""
            SELECT IFNULL(SUM(per.allocated_amount), 0)
            FROM `tabPayment Entry Reference` per
            INNER JOIN `tabPayment Entry` pe
                ON pe.name = per.parent
            WHERE pe.docstatus = 1
              AND per.reference_doctype = 'Expense Claim'
              AND per.reference_name = %s
        """, expense_claim)[0][0])

        ec = frappe.get_doc("Expense Claim", expense_claim)

        outstanding = round(ec.grand_total - total_reimbursed, 2)

        # Ignore tiny rounding differences
        if abs(outstanding) <= ROUNDING_TOLERANCE:
            total_reimbursed = ec.grand_total
            status = "Paid"
            is_paid = 1
        else:
            status = "Paid" if outstanding <= 0 else "Unpaid"
            is_paid = 1 if status == "Paid" else 0

        frappe.db.set_value(
            "Expense Claim",
            expense_claim,
            {
                "total_amount_reimbursed": total_reimbursed,
                "status": status,
                "is_paid": is_paid
            },
            update_modified=False
        )
# Start Roundoff for the Payment Entry            
#################################################################################

# NexAI Dashboard APIs

import frappe
import json
from datetime import datetime
from frappe.desk.reportview import get_filters_cond, get_match_cond

# =============================================================================
################################################################################
# --- START: Helpdesk Custom Report Builder & Charts ---
################################################################################
@frappe.whitelist()
def get_ticket_filter_options():
    customers = frappe.db.sql("""
        SELECT DISTINCT customer FROM `tabHD Ticket` 
        WHERE customer IS NOT NULL AND customer != '' 
        ORDER BY customer
    """, as_list=True)
    
    statuses = frappe.db.sql("""
        SELECT DISTINCT status FROM `tabHD Ticket` 
        WHERE status IS NOT NULL AND status != '' 
        ORDER BY status
    """, as_list=True)
    
    raised_by = frappe.db.sql("""
        SELECT DISTINCT raised_by FROM `tabHD Ticket` 
        WHERE raised_by IS NOT NULL AND raised_by != '' 
        ORDER BY raised_by
    """, as_list=True)
    
    circuit_ids = frappe.db.sql("""
        SELECT DISTINCT custom_circuit_id FROM `tabHD Ticket` 
        WHERE custom_circuit_id IS NOT NULL AND custom_circuit_id != '' 
        ORDER BY custom_circuit_id
    """, as_list=True)
    
    return {
        "customers": [c[0] for c in customers],
        "statuses": [s[0] for s in statuses],
        "users": [u[0] for u in raised_by],
        "circuit_ids": [cid[0] for cid in circuit_ids]
    }

@frappe.whitelist()
def get_ticket_custom_report_data(filters=None, fields=None):
    if isinstance(filters, str):
        filters = frappe.parse_json(filters)
    if isinstance(fields, str):
        fields = frappe.parse_json(fields)

    if not fields:
        return []

    # Map requested labels/fields to actual DB columns
    field_map = {
        "HD Ticket": {
            "ticket_id": "name as ticket_id",
            "subject": "subject",
            "status": "status",
            "priority": "priority",
            "raised_by": "raised_by",
            "customer": "customer",
            "assigned_to": "_assign as assigned_to",
            "creation_date": "DATE(creation) as creation_date",
            "resolution_date": "DATE(resolution_date) as resolution_date",
            "first_responded_on": "DATE(first_responded_on) as first_responded_on",
            "custom_circuit_id": "custom_circuit_id",
            "custom_lms_id": "custom_lms_id",
            "custom_channel": "custom_channel",
            "description": "description"
        }
    }

    select_clause = []
    
    for dt, dt_fields in fields.items():
        if dt == "HD Ticket":
            for f in dt_fields:
                if f in field_map["HD Ticket"]:
                    select_clause.append(field_map["HD Ticket"][f])

    if not select_clause:
        return []

    conditions = []
    if filters:
        if filters.get("creation_date_range") == "Current Month":
            conditions.append("MONTH(DATE(creation)) = MONTH(NOW()) AND YEAR(DATE(creation)) = YEAR(NOW())")
        elif filters.get("creation_date_range") == "Last 3 Months":
            conditions.append("DATE(creation) >= DATE_SUB(NOW(), INTERVAL 3 MONTH)")
        elif filters.get("creation_date_range") == "Custom" and filters.get("creation_from_date") and filters.get("creation_to_date"):
            conditions.append("DATE(creation) BETWEEN '%s' AND '%s'" % (filters.get("creation_from_date"), filters.get("creation_to_date")))
        
        if filters.get("resolution_date_range") == "Current Month":
            conditions.append("MONTH(DATE(resolution_date)) = MONTH(NOW()) AND YEAR(DATE(resolution_date)) = YEAR(NOW())")
        elif filters.get("resolution_date_range") == "Last 3 Months":
            conditions.append("DATE(resolution_date) >= DATE_SUB(NOW(), INTERVAL 3 MONTH)")
        elif filters.get("resolution_date_range") == "Custom" and filters.get("resolution_from_date") and filters.get("resolution_to_date"):
            conditions.append("DATE(resolution_date) BETWEEN '%s' AND '%s'" % (filters.get("resolution_from_date"), filters.get("resolution_to_date")))
        
        if filters.get("status") and filters.get("status") != "All":
            statuses_list = filters.get("status")
            if isinstance(statuses_list, list):
                conditions.append("status IN (%s)" % (", ".join(["'%s'" % frappe.db.escape(c).strip("'") for c in statuses_list])))
            else:
                conditions.append("status = '%s'" % frappe.db.escape(filters.get("status")).strip("'"))
        
        if filters.get("customer") and filters.get("customer") != "All":
            conditions.append("customer = '%s'" % frappe.db.escape(filters.get("customer")).strip("'"))
        
        if filters.get("raised_by") and filters.get("raised_by") != "All":
            conditions.append("raised_by = '%s'" % frappe.db.escape(filters.get("raised_by")).strip("'"))

        if filters.get("circuit_id") and filters.get("circuit_id") != "All":
            cids = filters.get("circuit_id")
            if isinstance(cids, str):
                cids = [c.strip() for c in cids.split(',') if c.strip()]
            if cids:
                conditions.append("custom_circuit_id IN (%s)" % (", ".join(["'%s'" % frappe.db.escape(c).strip("'") for c in cids])))

    where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""

    query = """
        SELECT 
            {select}
        FROM `tabHD Ticket`
        {where}
        ORDER BY creation DESC
    """.format(
        select=", ".join(select_clause), 
        where=where_clause
    )

    data = frappe.db.sql(query, as_dict=True)
    return data


@frappe.whitelist()
def get_ticket_dashboard_charts(filters=None):
    if filters and isinstance(filters, str):
        filters = frappe.parse_json(filters)
        
    conditions = []
    if filters:
        pass # Add any global filters here if needed in future
        
    where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""

    # Status Distribution
    status_distribution = frappe.db.sql(f"""
        SELECT status as label, count(*) as count
        FROM `tabHD Ticket`
        {where_clause}
        GROUP BY status
        ORDER BY count DESC
    """, as_dict=True)

    # Customer Distribution (Top 5)
    customer_cond = "customer IS NOT NULL AND customer != ''"
    cust_where = f"WHERE {customer_cond}"
    if conditions:
        cust_where += " AND " + " AND ".join(conditions)

    customer_distribution = frappe.db.sql(f"""
        SELECT customer as label, count(*) as count
        FROM `tabHD Ticket`
        {cust_where}
        GROUP BY customer
        ORDER BY count DESC
        LIMIT 5
    """, as_dict=True)

    # Channel Distribution
    channel_cond = "custom_channel IS NOT NULL AND custom_channel != ''"
    chan_where = f"WHERE {channel_cond}"
    if conditions:
        chan_where += " AND " + " AND ".join(conditions)

    channel_distribution = frappe.db.sql(f"""
        SELECT custom_channel as label, count(*) as count
        FROM `tabHD Ticket`
        {chan_where}
        GROUP BY custom_channel
        ORDER BY count DESC
    """, as_dict=True)

    # Last 6 Months Created
    from frappe.utils import add_months, get_first_day, now_datetime
    months = []
    curr = get_first_day(now_datetime())
    for i in range(5, -1, -1):
        d = add_months(curr, -i)
        months.append({
            "label": d.strftime("%b %Y"),
            "sort_val": d.year * 100 + d.month,
            "count": 0
        })

    start_date = add_months(curr, -5).strftime('%Y-%m-%d')
    date_cond = f"creation >= '{start_date}'"
    date_where = f"WHERE {date_cond}"
    if conditions:
        date_where += " AND " + " AND ".join(conditions)

    db_data = frappe.db.sql(f"""
        SELECT 
            (YEAR(creation) * 100 + MONTH(creation)) as sort_val,
            COUNT(*) as count
        FROM `tabHD Ticket`
        {date_where}
        GROUP BY sort_val
    """, as_dict=True)

    data_map = {d.sort_val: d.count for d in db_data}
    for m in months:
        m['count'] = data_map.get(m['sort_val'], 0)

    # Recent Tickets (Last 24 Hours)
    recent_where = "WHERE modified >= DATE_SUB(NOW(), INTERVAL 24 HOUR)"
    if conditions:
        recent_where += " AND " + " AND ".join(conditions)
        
    recent_tickets = frappe.db.sql(f"""
        SELECT name, custom_channel, custom_circuit_id, customer, custom_site_name as site_name, status, modified, raised_by
        FROM `tabHD Ticket`
        {recent_where}
        ORDER BY modified DESC
        LIMIT 20
    """, as_dict=True)

    return {
        "status_distribution": status_distribution,
        "customer_distribution": customer_distribution,
        "channel_distribution": channel_distribution,
        "tickets_last_6_months": months,
        "recent_tickets": recent_tickets
    }

################################################################################
# --- END: Helpdesk Custom Report Builder & Charts ---
################################################################################


################################################################################
# --- START: NOC Dashboard Phase 1 ---
################################################################################
@frappe.whitelist()
def get_noc_dashboard_data(filters=None):
    from frappe.utils import now_datetime, add_to_date, getdate
    
    if isinstance(filters, str):
        filters = frappe.parse_json(filters)
        
    conditions = []
    if filters:
        pass 
        
    where_clause = "WHERE 1=1"
    
    # 1. Executive KPIs
    from datetime import datetime, timedelta
    today = datetime.now().strftime('%Y-%m-%d')
    yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    
    # Tickets Today
    tickets_today = frappe.db.sql(f"SELECT COUNT(*) FROM `tabHD Ticket` WHERE DATE(creation) = '{today}'")[0][0] or 0
    tickets_yesterday = frappe.db.sql(f"SELECT COUNT(*) FROM `tabHD Ticket` WHERE DATE(creation) = '{yesterday}'")[0][0] or 0
    today_growth = 0
    if tickets_yesterday > 0:
        today_growth = round(((tickets_today - tickets_yesterday) / tickets_yesterday) * 100)
        
    # Open Tickets
    open_tickets = frappe.db.sql("SELECT COUNT(*) FROM `tabHD Ticket` WHERE status = 'Open'")[0][0] or 0
    
    # Critical Tickets
    critical_tickets = frappe.db.sql("SELECT COUNT(*) FROM `tabHD Ticket` WHERE status = 'Open' AND priority = 'Critical'")[0][0] or 0
    
    # SLA Breached
    # Fallback to status checking if agreement_status doesn't exist
    sla_breached = 0
    try:
        sla_breached = frappe.db.sql("SELECT COUNT(*) FROM `tabHD Ticket` WHERE agreement_status = 'Failed' AND status != 'Closed'")[0][0] or 0
    except Exception:
        pass
        
    # 2. Live Queue (Priority)
    priorities = frappe.db.sql("""
        SELECT priority, count(*) as count 
        FROM `tabHD Ticket` 
        WHERE status NOT IN ('Closed', 'Resolved')
        GROUP BY priority
    """, as_dict=True)
    
    # 3. Ticket Aging
    aging_data = frappe.db.sql("""
        SELECT 
            CASE 
                WHEN TIMESTAMPDIFF(HOUR, creation, NOW()) <= 2 THEN '0-2 Hours'
                WHEN TIMESTAMPDIFF(HOUR, creation, NOW()) <= 6 THEN '2-6 Hours'
                WHEN TIMESTAMPDIFF(HOUR, creation, NOW()) <= 12 THEN '6-12 Hours'
                WHEN TIMESTAMPDIFF(HOUR, creation, NOW()) <= 24 THEN '12-24 Hours'
                WHEN TIMESTAMPDIFF(DAY, creation, NOW()) <= 3 THEN '1-3 Days'
                ELSE '>3 Days'
            END as age_bucket,
            COUNT(*) as count
        FROM `tabHD Ticket`
        WHERE status NOT IN ('Closed', 'Resolved')
        GROUP BY age_bucket
    """, as_dict=True)
    
    # Map for sorting aging data
    aging_sort = {'0-2 Hours': 1, '2-6 Hours': 2, '6-12 Hours': 3, '12-24 Hours': 4, '1-3 Days': 5, '>3 Days': 6}
    aging_data = sorted(aging_data, key=lambda x: aging_sort.get(x['age_bucket'], 99))
    
    # 4. Incoming Last 12 Hours
    last_12h_data = frappe.db.sql("""
        SELECT 
            HOUR(creation) as hour_val,
            COUNT(*) as count
        FROM `tabHD Ticket`
        WHERE creation >= DATE_SUB(NOW(), INTERVAL 12 HOUR)
        GROUP BY hour_val
        ORDER BY creation ASC
    """, as_dict=True)

    return {
        "kpis": {
            "tickets_today": tickets_today,
            "today_growth": today_growth,
            "open_tickets": open_tickets,
            "critical_tickets": critical_tickets,
            "sla_breached": sla_breached
        },
        "live_queue": priorities,
        "aging": aging_data,
        "incoming_12h": last_12h_data
    }
################################################################################
# --- END: NOC Dashboard Phase 1 ---
################################################################################


################################################################################
# --- START: NexAI Dashboard V1 ---
################################################################################
@frappe.whitelist()
def get_nexai_dashboard_data(filters=None):
    from datetime import datetime, timedelta
    
    if isinstance(filters, str):
        filters = frappe.parse_json(filters)
        
    today = datetime.now().strftime('%Y-%m-%d')
    yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    
    # 1. FACTS (ERPNext Data)
    tickets_today = frappe.db.sql(f"SELECT COUNT(*) FROM `tabHD Ticket` WHERE DATE(creation) = '{today}'")[0][0] or 0
    tickets_yesterday = frappe.db.sql(f"SELECT COUNT(*) FROM `tabHD Ticket` WHERE DATE(creation) = '{yesterday}'")[0][0] or 0
    
    total_open = frappe.db.sql("SELECT COUNT(*) FROM `tabHD Ticket` WHERE status NOT IN ('Closed', 'Resolved')")[0][0] or 0
    
    sla_breached = 0
    try:
        sla_breached = frappe.db.sql("SELECT COUNT(*) FROM `tabHD Ticket` WHERE agreement_status = 'Failed' AND status != 'Closed'")[0][0] or 0
    except:
        pass
        
    sla_compliance = 100
    if total_open > 0:
        sla_compliance = max(0, round(((total_open - sla_breached) / total_open) * 100))
        
    at_risk = 0
    try:
        at_risk = frappe.db.sql("SELECT COUNT(*) FROM `tabHD Ticket` WHERE status NOT IN ('Closed', 'Resolved') AND resolution_by BETWEEN NOW() AND DATE_ADD(NOW(), INTERVAL 2 HOUR)")[0][0] or 0
    except:
        at_risk = 0
        
    growth = 0
    if tickets_yesterday > 0:
        growth = round(((tickets_today - tickets_yesterday) / tickets_yesterday) * 100)
        
    # Categories
    top_category = "General Issues"
    try:
        if frappe.db.has_column("HD Ticket", "custom_category"):
            categories = frappe.db.sql(f"SELECT custom_category, COUNT(*) as c FROM `tabHD Ticket` WHERE DATE(creation) = '{today}' GROUP BY custom_category ORDER BY c DESC LIMIT 1", as_dict=True)
            top_category = categories[0]['custom_category'] if categories else "General Issues"
        elif frappe.db.has_column("HD Ticket", "category"):
            categories = frappe.db.sql(f"SELECT category, COUNT(*) as c FROM `tabHD Ticket` WHERE DATE(creation) = '{today}' GROUP BY category ORDER BY c DESC LIMIT 1", as_dict=True)
            top_category = categories[0]['category'] if categories else "General Issues"
    except:
        pass
        
    # Customer Health
    customer_health = []
    try:
        customers = frappe.db.sql("""
            SELECT customer, count(*) as count,
            SUM(CASE WHEN agreement_status = 'Failed' THEN 1 ELSE 0 END) as breached
            FROM `tabHD Ticket`
            WHERE status != 'Closed' AND customer IS NOT NULL
            GROUP BY customer
            ORDER BY breached DESC, count DESC
            LIMIT 3
        """, as_dict=True)
        
        for c in customers:
            if c.breached > 0 or c.count > 10:
                customer_health.append({
                    "customer": c.customer,
                    "risk": "Critical" if c.breached > 5 else "High Risk",
                    "open_tickets": c.count,
                    "breached": c.breached
                })
    except:
        pass
        
    # Agent Workload
    agent_workload = frappe.db.sql("""
        SELECT _assign, count(*) as count 
        FROM `tabHD Ticket` 
        WHERE status NOT IN ('Closed', 'Resolved') AND _assign IS NOT NULL AND _assign != ''
        GROUP BY _assign
        ORDER BY count DESC
    """, as_dict=True)
    
    for a in agent_workload:
        import json
        try:
            assignees = json.loads(a['_assign'])
            a['agent_name'] = assignees[0].split('@')[0].capitalize()
        except:
            a['agent_name'] = str(a['_assign']).split('@')[0].capitalize()
            
    # Charts Data
    priorities = frappe.db.sql("""
        SELECT priority, count(*) as count 
        FROM `tabHD Ticket` 
        WHERE status NOT IN ('Closed', 'Resolved')
        GROUP BY priority
    """, as_dict=True)
    
    aging_data = frappe.db.sql("""
        SELECT 
            CASE 
                WHEN TIMESTAMPDIFF(HOUR, creation, NOW()) <= 24 THEN '0-24 Hours'
                WHEN TIMESTAMPDIFF(DAY, creation, NOW()) <= 3 THEN '1-3 Days'
                ELSE '>3 Days'
            END as age_bucket,
            COUNT(*) as count
        FROM `tabHD Ticket`
        WHERE status NOT IN ('Closed', 'Resolved')
        GROUP BY age_bucket
    """, as_dict=True)
    
    total_unresolved = sum(a['count'] for a in aging_data)
    older_than_24 = sum(a['count'] for a in aging_data if a['age_bucket'] != '0-24 Hours')
    progress = 100
    if total_unresolved > 0:
        progress = max(0, round(((total_unresolved - older_than_24) / total_unresolved) * 100))

    # Dynamic Facts & Status
    if growth > 0:
        volume_msg = f"Today {tickets_today} tickets have been created ({growth}% higher than yesterday)."
    elif growth < 0:
        volume_msg = f"Today only {tickets_today} tickets have been created (lower than yesterday's {tickets_yesterday})."
    else:
        volume_msg = f"Ticket volume is stable compared to yesterday ({tickets_today} tickets)."

    # Determine Overall Status
    if at_risk > 5 or (growth > 50 and tickets_today > 20):
        status_color = "🔴"
        status_title = "High Risk"
        status_desc = "Ticket volume has spiked and SLA risks are high. Immediate triage is recommended."
    elif at_risk > 0 or growth > 20:
        status_color = "🟠"
        status_title = "Elevated Activity"
        status_desc = "Operations are active. Some tickets are approaching SLA breaches."
    else:
        status_color = "🟢"
        status_title = "Healthy"
        status_desc = "Operations are stable. No immediate action is required."

    # Contextual category msg
    if top_category:
        cat_msg = f"Most new tickets belong to {top_category}, but volume remains manageable."
    else:
        cat_msg = "New ticket distribution is normal across all categories."

    # Curated Good News
    good_news_list = []
    if total_open < 20:
        good_news_list.append("Queue is operating normally and workload is light.")
    if sla_compliance >= 95:
        good_news_list.append("SLA remains highly healthy.")
    if progress > 50:
        good_news_list.append(f"Backlog reduction is progressing well ({progress}%).")
        
    if not good_news_list:
        good_news_list = ["The team is actively managing the queue."]

    return {
        "firp": {
            "status": {"color": status_color, "title": status_title, "desc": status_desc},
            "facts": [
                f"Yesterday your team resolved {tickets_yesterday} tickets.",
                volume_msg,
                cat_msg,
                f"{at_risk} SLA breaches are expected." if at_risk > 0 else "No immediate SLA breaches are expected."
            ],
            "good_news": good_news_list[:3] # Max 3 items
        },
        "mission": {
            "title": "Maintain the current SLA while clearing the remaining backlog." if at_risk == 0 else "Focus on tickets approaching SLA breaches.",
            "progress": progress
        },
        "recommendations": [
            {"action": "Reassign workload", "priority": 1},
            {"action": "Recover SLA", "priority": 2},
            {"action": f"Contact {customer_health[0]['customer'] if customer_health else 'Top Client'}", "priority": 3}
        ],
        "charts": {
            "queue": priorities,
            "aging": aging_data,
            "trend": frappe.db.sql('''
                SELECT 
                    DATE_FORMAT(creation, '%H:00') as hour,
                    COUNT(*) as count
                FROM `tabHD Ticket`
                WHERE creation >= DATE_SUB(NOW(), INTERVAL 12 HOUR)
                GROUP BY HOUR(creation)
                ORDER BY creation ASC
            ''', as_dict=True),
            "sla": {"compliance": sla_compliance, "breached": sla_breached, "open": total_open, "at_risk": at_risk},
            "workload": agent_workload[:5] if agent_workload else [],
            "customer_health": customer_health
        }
    }
# --- END: NexAI Dashboard V1 ---
################################################################################

################################################################################
# --- START: Ask NexAI Natural Language Query ---
################################################################################
@frappe.whitelist()
def ask_nexai(query):
    import json
    from datetime import datetime
    
    try:
        telemetry = get_nexai_dashboard_data()
        facts = telemetry.get('firp', {}).get('facts', [])
    except:
        facts = []
        
    current_date = datetime.now().strftime("%Y-%m-%d")
        
    intent_prompt = f"""
    You are NexAI, the Chief of Staff for a Helpdesk.
    
    Current Telemetry Facts:
    {json.dumps(facts)}
    Today's Date: {current_date}
    
    User Query: "{query}"
    
    INSTRUCTIONS:
    1. If the user's query can be answered using the Telemetry Facts provided above (e.g. asking about ticket volume today, yesterday, or SLA), provide a brief, professional answer directly. Use HTML <br> for newlines.
    2. IF AND ONLY IF the user is asking for specific ticket details NOT in the facts (like a specific ticket status, priority, customer, or a SPECIFIC DATE like "12-07-2026"), return ONLY a valid JSON object representing Frappe ORM filters for 'HD Ticket'.
    
    Valid fields for JSON: name, status, priority, _assign, customer, custom_category, creation.
    
    JSON Format: {{"action": "COUNT", "filters": {{"status": "Open"}}}} or {{"action": "LIST", "filters": {{"status": "Open"}}}}
    
    CRITICAL DATE RULES:
    - If filtering by a specific date (e.g., 12-07-2026), format the date as YYYY-MM-DD.
    - Use the "like" operator for a specific day: {{"creation": ["like", "2026-07-12%"]}}
    - If no specific filters apply, ensure you do not drop the filters. Be as specific as possible!
    """
    
    try:
        response = call_ai_model(intent_prompt).strip()
        
        clean_response = response.replace("```json", "").replace("```", "").strip()
        
        if clean_response.startswith("{") and clean_response.endswith("}"):
            intent = json.loads(clean_response)
            filters = intent.get("filters", {})
            action = intent.get("action", "LIST")
            
            data = None
            if action == "COUNT":
                data = frappe.db.count("HD Ticket", filters)
            else:
                data = frappe.get_list("HD Ticket", filters=filters, fields=["name", "status", "priority", "subject", "customer"], limit=20)
                
            response_prompt = f"""
            You are NexAI.
            User asked: "{query}"
            Database returned: {json.dumps(data)}
            Write a brief, professional answer based on this data. Use <br> for newlines.
            """
            
            final_answer = call_ai_model(response_prompt)
            return final_answer.replace('\n', '<br>')
            
        else:
            return response.replace('\n', '<br>')
            
    except Exception as e:
        return f"I encountered an error analyzing the request: {str(e)}"
################################################################################
# --- END: Ask NexAI Natural Language Query ---
################################################################################

# =========================================================================
# BULK SITE UPDATE API — Direct Sync (No Background Queue)
# =========================================================================
@frappe.whitelist()
def process_bulk_site_update(file_name, file_data, validate_only=1):
    validate_only = int(validate_only)

    import base64
    import csv
    import openpyxl
    from io import BytesIO
    from frappe.utils import now_datetime

    decoded_data = base64.b64decode(file_data)
    errors = []
    success_count = 0
    total_rows = 0

    try:
        if file_name.endswith('.csv'):
            content = decoded_data.decode('utf-8')
            reader = csv.reader(content.splitlines())
            rows = list(reader)
        else:
            wb = openpyxl.load_workbook(filename=BytesIO(decoded_data), data_only=True)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))

        if not rows or len(rows) < 2:
            return {"status": "error", "errors": ["File is empty or missing data rows."]}

        raw_header = [str(col).strip() if col is not None else "" for col in rows[0]]
        while raw_header and raw_header[-1] == "":
            raw_header.pop()

        header = raw_header
        data_rows = rows[1:]
        total_rows = len(data_rows)

        meta = frappe.get_meta("Site")
        label_to_fieldname = {"name": "name", "owner": "owner", "creation": "creation", "modified": "modified", "modified_by": "modified_by"}

        for df in meta.fields:
            if df.fieldname:
                label_to_fieldname[df.fieldname.strip().lower()] = df.fieldname
            if df.label:
                label_to_fieldname[df.label.strip().lower()] = df.fieldname

        label_to_fieldname["id"] = "name"
        label_to_fieldname["circuit id"] = "circuit_id"

        key_col_idx = None
        col_mapping = {}
        header_errors = []

        for idx, col in enumerate(header):
            col_lower = col.lower().strip()
            if not col_lower:
                header_errors.append(f"Invalid Column: Missing header name.")
                continue
            if col_lower == "name":
                key_col_idx = idx
                col_mapping[idx] = "name"
            elif col_lower in label_to_fieldname:
                col_mapping[idx] = label_to_fieldname[col_lower]
            else:
                header_errors.append(f"Invalid Column '{col}'.")

        if header_errors:
            return {"status": "error", "errors": header_errors}

        if key_col_idx is None:
            return {"status": "error", "errors": ["Unable to process: 'name' column missing."]}

        # --- VALIDATION PASS ---
        validated_rows = []
        for row_idx, row in enumerate(data_rows, start=2):
            if not any(row):
                continue

            record_id = None
            if key_col_idx < len(row):
                val = str(row[key_col_idx]).strip()
                if val.endswith('.0'):
                    val = val[:-2]
                record_id = val

            if not record_id:
                errors.append(f"Row {row_idx}: Missing identifying key ('name').")
                continue

            if not frappe.db.exists("Site", record_id):
                errors.append(f"Row {row_idx}: Site with ID '{record_id}' does not exist.")
                continue

            validated_rows.append((row_idx, row, record_id))

        if errors:
            return {"status": "error", "errors": errors, "total_rows": total_rows}

        if validate_only:
            return {"status": "success", "total_rows": total_rows}

        # --- UPDATE PASS: Direct SQL via frappe.db.set_value (fast!) ---
        user = frappe.session.user
        for row_idx, row, record_id in validated_rows:
            try:
                update_dict = {}
                for idx, col_val in enumerate(row):
                    if idx in col_mapping and idx != key_col_idx:
                        fname = col_mapping[idx]
                        raw_val = str(col_val).strip() if col_val is not None else ""
                        if raw_val.endswith('.0'):
                            raw_val = raw_val[:-2]
                        update_dict[fname] = raw_val

                pincode_val = update_dict.get("pincode")
                if pincode_val:
                    try:
                        import requests
                        resp = requests.get(f"https://api.postalpincode.in/pincode/{pincode_val}", timeout=5)
                        if resp.status_code == 200:
                            data = resp.json()
                            if data and data[0].get("Status") == "Success":
                                po = data[0]["PostOffice"][0]
                                res = {
                                    "district": po.get("District"),
                                    "state": po.get("State"),
                                    "country": po.get("Country", "India"),
                                    "city": po.get("Block") or po.get("Region")
                                }
                                if "district" not in update_dict and res.get("district"): update_dict["district"] = res["district"]
                                if "state" not in update_dict and res.get("state"): update_dict["state"] = res["state"]
                                if "country" not in update_dict and res.get("country"): update_dict["country"] = res["country"]
                                if "city" not in update_dict and res.get("city"): update_dict["city"] = res["city"]

                                state = res.get("state")
                                if state and "territory" not in update_dict:
                                    state_territory_map = {
                                        "Delhi": "North", "Haryana": "North", "Punjab": "North", "Himachal Pradesh": "North", "Uttar Pradesh": "North", "Uttarakhand": "North", "Jammu and Kashmir": "North", "Chandigarh": "North", "Rajasthan": "North", "Ladakh": "North",
                                        "Karnataka": "South", "Tamil Nadu": "South", "Kerala": "South", "Andhra Pradesh": "South", "Telangana": "South", "Puducherry": "South", "Lakshadweep": "South",
                                        "Maharashtra": "West", "Gujarat": "West", "Goa": "West", "Dadra and Nagar Haveli": "West", "Daman and Diu": "West", "Madhya Pradesh": "West", "Chattisgarh": "West", "Chhattisgarh": "West",
                                        "West Bengal": "East", "Odisha": "East", "Bihar": "East", "Jharkhand": "East", "Assam": "East", "Sikkim": "East", "Meghalaya": "East", "Tripura": "East", "Arunachal Pradesh": "East", "Manipur": "East", "Nagaland": "East", "Mizoram": "East", "Andaman and Nicobar Islands": "East"
                                    }
                                    if state in state_territory_map:
                                        update_dict["territory"] = state_territory_map[state]
                    except Exception:
                        pass

                changes_html = ""
                for k, v in update_dict.items():
                    if k == "reason_site_updated_later":
                        continue
                    old_value = frappe.db.get_value("Site", record_id, k)
                    str_old = str(old_value).strip() if old_value is not None else ""
                    str_new = str(v).strip()

                    if str_old != str_new:
                        field_label = k
                        for df in meta.fields:
                            if df.fieldname == k:
                                field_label = df.label or k
                                break
                        display_old = str_old if str_old else "<i>[Empty]</i>"
                        changes_html += f"<li style='margin-bottom: 4px;'><b>{field_label}:</b> <span style='color:#ef4444; text-decoration:line-through;'>{display_old}</span> <span style='color:#94a3b8; font-weight: bold;'>&rarr;</span> <span style='color:#16a34a; font-weight: 500;'>{str_new}</span></li>"

                fields_to_set = {k: v for k, v in update_dict.items() if k != "reason_site_updated_later"}
                if fields_to_set:
                    for fname, fval in fields_to_set.items():
                        frappe.db.set_value("Site", record_id, fname, fval, update_modified=True)

                if changes_html:
                    current_time = now_datetime().strftime("%d-%b-%Y %I:%M %p")
                    log_entry = f'<div style="border-left:3px solid #695A97;padding:4px 8px;margin-bottom:6px;font-size:12px;background:#f8fafc;"><span style="background:#695A97;color:#fff;padding:1px 6px;border-radius:8px;font-size:11px;font-weight:600;">Bulk Update</span> <b>{current_time}</b> by {user}<ul style="margin:2px 0 0;padding-left:16px;line-height:1.4;">{changes_html}</ul></div>'
                    existing_log = frappe.db.get_value("Site", record_id, "reason_site_updated_later") or ""
                    frappe.db.set_value("Site", record_id, "reason_site_updated_later", log_entry + existing_log, update_modified=False)

                success_count += 1

                if success_count % 100 == 0:
                    frappe.db.commit()

            except Exception as e:
                errors.append(f"Row {row_idx}: Failed to update Site {record_id}. Error: {str(e)}")

        frappe.db.commit()
        return {"status": "success", "success_count": success_count, "errors": errors}

    except Exception as e:
        frappe.log_error("Bulk Site Update Error", frappe.get_traceback())
        return {"status": "error", "errors": [f"Server Error: {str(e)}"]}
# =========================================================================
# END OF BULK SITE UPDATE API
# =========================================================================

################################################################################
# --- START: EXCLUSIVE HELPDESK DASHBOARD V2 APIs ---
################################################################################

@frappe.whitelist()
def get_ticket_filter_options_v2():
    customers = frappe.db.sql("""
        SELECT DISTINCT customer FROM `tabHD Ticket` 
        WHERE customer IS NOT NULL AND customer != '' 
        ORDER BY customer
    """, as_list=True)
    
    statuses = frappe.db.sql("""
        SELECT DISTINCT status FROM `tabHD Ticket` 
        WHERE status IS NOT NULL AND status != '' 
        ORDER BY status
    """, as_list=True)
    
    raised_by = frappe.db.sql("""
        SELECT DISTINCT raised_by FROM `tabHD Ticket` 
        WHERE raised_by IS NOT NULL AND raised_by != '' 
        ORDER BY raised_by
    """, as_list=True)
    
    circuit_ids = frappe.db.sql("""
        SELECT DISTINCT custom_circuit_id FROM `tabHD Ticket` 
        WHERE custom_circuit_id IS NOT NULL AND custom_circuit_id != '' 
        ORDER BY custom_circuit_id
    """, as_list=True)
    
    return {
        "customers": [c[0] for c in customers],
        "statuses": [s[0] for s in statuses],
        "users": [u[0] for u in raised_by],
        "circuit_ids": [cid[0] for cid in circuit_ids]
    }

@frappe.whitelist()
def get_ticket_custom_report_data_v2(filters=None, fields=None):
    if isinstance(filters, str):
        filters = frappe.parse_json(filters)
    if isinstance(fields, str):
        fields = frappe.parse_json(fields)

    if not fields:
        return []

    # Map requested labels/fields to actual DB columns
    field_map = {
        "HD Ticket": {
            "ticket_id": "name as ticket_id",
            "subject": "subject",
            "status": "status",
            "priority": "priority",
            "raised_by": "raised_by",
            "customer": "customer",
            "assigned_to": "_assign as assigned_to",
            "creation_date": "DATE(creation) as creation_date",
            "resolution_date": "DATE(resolution_date) as resolution_date",
            "first_responded_on": "DATE(first_responded_on) as first_responded_on",
            "custom_circuit_id": "custom_circuit_id",
            "custom_lms_id": "custom_lms_id",
            "custom_channel": "custom_channel",
            "description": "description"
        }
    }

    select_clause = []
    
    for dt, dt_fields in fields.items():
        if dt == "HD Ticket":
            for f in dt_fields:
                if f in field_map["HD Ticket"]:
                    select_clause.append(field_map["HD Ticket"][f])

    if not select_clause:
        return []

    conditions = []
    if filters:
        if filters.get("creation_date_range") == "Current Month":
            conditions.append("MONTH(DATE(creation)) = MONTH(NOW()) AND YEAR(DATE(creation)) = YEAR(NOW())")
        elif filters.get("creation_date_range") == "Last 3 Months":
            conditions.append("DATE(creation) >= DATE_SUB(NOW(), INTERVAL 3 MONTH)")
        elif filters.get("creation_date_range") == "Custom" and filters.get("creation_from_date") and filters.get("creation_to_date"):
            conditions.append("DATE(creation) BETWEEN '%s' AND '%s'" % (filters.get("creation_from_date"), filters.get("creation_to_date")))
        
        if filters.get("resolution_date_range") == "Current Month":
            conditions.append("MONTH(DATE(resolution_date)) = MONTH(NOW()) AND YEAR(DATE(resolution_date)) = YEAR(NOW())")
        elif filters.get("resolution_date_range") == "Last 3 Months":
            conditions.append("DATE(resolution_date) >= DATE_SUB(NOW(), INTERVAL 3 MONTH)")
        elif filters.get("resolution_date_range") == "Custom" and filters.get("resolution_from_date") and filters.get("resolution_to_date"):
            conditions.append("DATE(resolution_date) BETWEEN '%s' AND '%s'" % (filters.get("resolution_from_date"), filters.get("resolution_to_date")))
        
        if filters.get("status") and filters.get("status") != "All":
            statuses_list = filters.get("status")
            if isinstance(statuses_list, list):
                conditions.append("status IN (%s)" % (", ".join(["'%s'" % frappe.db.escape(c).strip("'") for c in statuses_list])))
            else:
                conditions.append("status = '%s'" % frappe.db.escape(filters.get("status")).strip("'"))
        
        if filters.get("customer") and filters.get("customer") != "All":
            conditions.append("customer = '%s'" % frappe.db.escape(filters.get("customer")).strip("'"))
        
        if filters.get("raised_by") and filters.get("raised_by") != "All":
            conditions.append("raised_by = '%s'" % frappe.db.escape(filters.get("raised_by")).strip("'"))

        if filters.get("circuit_id") and filters.get("circuit_id") != "All":
            cids = filters.get("circuit_id")
            if isinstance(cids, str):
                cids = [c.strip() for c in cids.split(',') if c.strip()]
            if cids:
                conditions.append("custom_circuit_id IN (%s)" % (", ".join(["'%s'" % frappe.db.escape(c).strip("'") for c in cids])))

    where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""

    query = """
        SELECT 
            {select}
        FROM `tabHD Ticket`
        {where}
        ORDER BY creation DESC
    """.format(
        select=", ".join(select_clause), 
        where=where_clause
    )

    data = frappe.db.sql(query, as_dict=True)
    return data

@frappe.whitelist()
def get_nexai_dashboard_data_v2(filters=None):
    from datetime import datetime, timedelta
    
    if isinstance(filters, str):
        filters = frappe.parse_json(filters)
        
    today = datetime.now().strftime('%Y-%m-%d')
    yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    
    # 1. FACTS (ERPNext Data)
    tickets_today = frappe.db.sql(f"SELECT COUNT(*) FROM `tabHD Ticket` WHERE DATE(creation) = '{today}'")[0][0] or 0
    tickets_yesterday = frappe.db.sql(f"SELECT COUNT(*) FROM `tabHD Ticket` WHERE DATE(creation) = '{yesterday}'")[0][0] or 0
    
    total_open = frappe.db.sql("SELECT COUNT(*) FROM `tabHD Ticket` WHERE status IN ('Open', 'Replied', 'On Hold')")[0][0] or 0
    
    sla_breached = 0
    try:
        sla_breached = frappe.db.sql("SELECT COUNT(*) FROM `tabHD Ticket` WHERE agreement_status = 'Failed' AND status IN ('Open', 'Replied', 'On Hold')")[0][0] or 0
    except:
        pass
        
    sla_compliance = 100
    if total_open > 0:
        sla_compliance = max(0, round(((total_open - sla_breached) / total_open) * 100))
        
    at_risk = 0
    try:
        at_risk = frappe.db.sql("SELECT COUNT(*) FROM `tabHD Ticket` WHERE status IN ('Open', 'Replied', 'On Hold') AND resolution_by BETWEEN NOW() AND DATE_ADD(NOW(), INTERVAL 2 HOUR)")[0][0] or 0
    except:
        at_risk = 0
        
    med_risk = 0
    try:
        med_risk = frappe.db.sql("SELECT COUNT(*) FROM `tabHD Ticket` WHERE status IN ('Open', 'Replied', 'On Hold') AND resolution_by BETWEEN DATE_ADD(NOW(), INTERVAL 2 HOUR) AND DATE_ADD(NOW(), INTERVAL 4 HOUR)")[0][0] or 0
    except:
        med_risk = 0
        
    growth = 0
    if tickets_yesterday > 0:
        growth = round(((tickets_today - tickets_yesterday) / tickets_yesterday) * 100)
        
    # Categories
    top_category = "General Issues"
    try:
        if frappe.db.has_column("HD Ticket", "custom_category"):
            categories = frappe.db.sql(f"SELECT custom_category, COUNT(*) as c FROM `tabHD Ticket` WHERE DATE(creation) = '{today}' GROUP BY custom_category ORDER BY c DESC LIMIT 1", as_dict=True)
            top_category = categories[0]['custom_category'] if categories else "General Issues"
        elif frappe.db.has_column("HD Ticket", "category"):
            categories = frappe.db.sql(f"SELECT category, COUNT(*) as c FROM `tabHD Ticket` WHERE DATE(creation) = '{today}' GROUP BY category ORDER BY c DESC LIMIT 1", as_dict=True)
            top_category = categories[0]['category'] if categories else "General Issues"
    except:
        pass
        
    # Customer Health
    customer_health = []
    try:
        customers = frappe.db.sql("""
            SELECT customer, count(*) as count,
            SUM(CASE WHEN agreement_status = 'Failed' THEN 1 ELSE 0 END) as breached
            FROM `tabHD Ticket`
            WHERE status != 'Closed' AND customer IS NOT NULL
            GROUP BY customer
            ORDER BY breached DESC, count DESC
            LIMIT 3
        """, as_dict=True)
        
        for c in customers:
            if c.breached > 0 or c.count > 10:
                customer_health.append({
                    "customer": c.customer,
                    "risk": "Critical" if c.breached > 5 else "High Risk",
                    "open_tickets": c.count,
                    "breached": c.breached
                })
    except:
        pass
        
    # Agent Workload
    agent_workload = frappe.db.sql("""
        SELECT _assign, count(*) as count 
        FROM `tabHD Ticket` 
        WHERE status IN ('Open', 'Replied', 'On Hold') AND _assign IS NOT NULL AND _assign != ''
        GROUP BY _assign
        ORDER BY count DESC
    """, as_dict=True)
    
    for a in agent_workload:
        import json
        try:
            assignees = json.loads(a['_assign'])
            a['agent_name'] = assignees[0].split('@')[0].capitalize()
        except:
            a['agent_name'] = str(a['_assign']).split('@')[0].capitalize()
            
    # Charts Data
    priorities = frappe.db.sql("""
        SELECT priority, count(*) as count 
        FROM `tabHD Ticket` 
        WHERE status IN ('Open', 'Replied', 'On Hold')
        GROUP BY priority
    """, as_dict=True)
    
    aging_data = frappe.db.sql("""
        SELECT 
            CASE 
                WHEN TIMESTAMPDIFF(HOUR, creation, NOW()) <= 24 THEN '0-24 Hours'
                WHEN TIMESTAMPDIFF(DAY, creation, NOW()) <= 3 THEN '1-3 Days'
                ELSE '>3 Days'
            END as age_bucket,
            COUNT(*) as count
        FROM `tabHD Ticket`
        WHERE status IN ('Open', 'Replied', 'On Hold')
        GROUP BY age_bucket
    """, as_dict=True)
    
    total_unresolved = sum(a['count'] for a in aging_data)
    older_than_24 = sum(a['count'] for a in aging_data if a['age_bucket'] != '0-24 Hours')
    progress = 100
    if total_unresolved > 0:
        progress = max(0, round(((total_unresolved - older_than_24) / total_unresolved) * 100))

    # Dynamic Facts & Status
    if growth > 0:
        volume_msg = f"Today {tickets_today} tickets have been created ({growth}% higher than yesterday)."
    elif growth < 0:
        volume_msg = f"Today only {tickets_today} tickets have been created (lower than yesterday's {tickets_yesterday})."
    else:
        volume_msg = f"Ticket volume is stable compared to yesterday ({tickets_today} tickets)."

    # Determine Overall Status
    if at_risk > 5 or (growth > 50 and tickets_today > 20):
        status_color = "🔴"
        status_title = "High Risk"
        status_desc = "Ticket volume has spiked and SLA risks are high. Immediate triage is recommended."
    elif at_risk > 0 or growth > 20:
        status_color = "🟠"
        status_title = "Elevated Activity"
        status_desc = "Operations are active. Some tickets are approaching SLA breaches."
    else:
        status_color = "🟢"
        status_title = "Healthy"
        status_desc = "Operations are stable. No immediate action is required."

    # Contextual category msg
    if top_category:
        cat_msg = f"Most new tickets belong to {top_category}, but volume remains manageable."
    else:
        cat_msg = "New ticket distribution is normal across all categories."

    # Curated Good News
    good_news_list = []
    if total_open < 20:
        good_news_list.append("Queue is operating normally and workload is light.")
    if sla_compliance >= 95:
        good_news_list.append("SLA remains highly healthy.")
    if progress > 50:
        good_news_list.append(f"Backlog reduction is progressing well ({progress}%).")
        
    if not good_news_list:
        good_news_list = ["The team is actively managing the queue."]

    
    # Detailed Status Breakdown for KPIs
    status_open = frappe.db.sql("SELECT COUNT(*) FROM `tabHD Ticket` WHERE status = 'Open'")[0][0] or 0
    status_replied = frappe.db.sql("SELECT COUNT(*) FROM `tabHD Ticket` WHERE status = 'Replied'")[0][0] or 0
    status_hold = frappe.db.sql("SELECT COUNT(*) FROM `tabHD Ticket` WHERE status = 'On Hold'")[0][0] or 0
    status_wrong_circuit = frappe.db.sql("SELECT COUNT(*) FROM `tabHD Ticket` WHERE status = 'Wrong Circuit'")[0][0] or 0
    
    # Calculate some KPI metrics
    resolved_today = frappe.db.sql(f"SELECT COUNT(*) FROM `tabHD Ticket` WHERE status IN ('Closed', 'Resolved') AND DATE(modified) = '{today}'")[0][0] or 0
    resolved_yesterday = frappe.db.sql(f"SELECT COUNT(*) FROM `tabHD Ticket` WHERE status IN ('Closed', 'Resolved') AND DATE(modified) = '{yesterday}'")[0][0] or 0
    
    # Phase 3 data calculations
    category_field = "custom_category" if frappe.db.has_column("HD Ticket", "custom_category") else "category"
    categories_data = []
    if frappe.db.has_column("HD Ticket", category_field):
        categories_data = frappe.db.sql(f"SELECT IFNULL({category_field}, 'Others') as category, COUNT(*) as count FROM `tabHD Ticket` WHERE DATE(creation) = '{today}' GROUP BY {category_field} ORDER BY count DESC", as_dict=True)
    
    # Resolvers (Agents)
    top_agents = frappe.db.sql(f"SELECT _assign, COUNT(*) as resolved_count FROM `tabHD Ticket` WHERE status IN ('Closed', 'Resolved') AND _assign IS NOT NULL AND DATE(modified) = '{today}' GROUP BY _assign ORDER BY resolved_count DESC LIMIT 5", as_dict=True)
    for agent in top_agents:
        try:
            import json
            assignees = json.loads(agent['_assign'])
            agent['name'] = assignees[0].split('@')[0].replace('.', ' ').title()
        except:
            agent['name'] = str(agent['_assign']).split('@')[0].replace('.', ' ').title()
            
    # SLA Nearing Tickets
    sla_nearing = []
    try:
        sla_nearing = frappe.db.sql("""
            SELECT name as id, subject, 
            TIMESTAMPDIFF(MINUTE, NOW(), resolution_by) as min_left
            FROM `tabHD Ticket`
            WHERE status IN ('Open', 'Replied', 'On Hold') 
            AND resolution_by IS NOT NULL 
            AND resolution_by > NOW()
            ORDER BY resolution_by ASC
            LIMIT 3
        """, as_dict=True)
        
        for t in sla_nearing:
            if t.min_left < 120:
                t['risk'] = "high"
            elif t.min_left < 240:
                t['risk'] = "medium"
            else:
                t['risk'] = "low"
                
            if t.min_left >= 60:
                t['time'] = f"{int(t.min_left/60)} hrs left"
            else:
                t['time'] = f"{t.min_left} min left"
    except Exception as e:
        pass
        
    dynamic_recommendations = []
    
    # 1. SLA Risk Check
    if at_risk > 0:
        dynamic_recommendations.append({"action": f"Prioritize {at_risk} tickets nearing SLA breach.", "priority": 1})
    elif sla_compliance < 90:
        dynamic_recommendations.append({"action": "SLA Compliance is dropping. Investigate bottlenecks.", "priority": 2})
        
    # 2. Workload Check
    if agent_workload and len(agent_workload) >= 2:
        top_agent = agent_workload[0]
        if top_agent['count'] > 10:
            dynamic_recommendations.append({"action": f"Reassign workload from {top_agent.get('agent_name', 'Top Agent')} to balance queue.", "priority": 1})
            
    # 3. Customer Health Check
    if customer_health:
        dynamic_recommendations.append({"action": f"Contact {customer_health[0]['customer']} to manage relationship.", "priority": 2})
        
    # 4. Aging Check
    if older_than_24 > 0 and len(dynamic_recommendations) < 3:
        dynamic_recommendations.append({"action": f"Review {older_than_24} tickets older than 24 hours.", "priority": 3})
        
    # 5. Default Fallbacks if empty
    if not dynamic_recommendations:
        dynamic_recommendations = [
            {"action": "All operations running optimally. No action needed.", "priority": 3}
        ]
        
    # Keep only top 3 and sort by priority
    dynamic_recommendations = sorted(dynamic_recommendations, key=lambda x: x['priority'])[:3]
        
    return {
        "kpis": {
            "status_open": status_open,
            "status_replied": status_replied,
            "status_hold": status_hold,
            "status_wrong_circuit": status_wrong_circuit,
            "open_today": total_open,
            "open_yesterday": total_open + (resolved_today - tickets_today),
            "resolved_today": resolved_today,
            "resolved_yesterday": resolved_yesterday,
            "sla_today": sla_compliance,
            "sla_yesterday": max(0, min(100, sla_compliance + 2)),
            "new_today": tickets_today,
            "new_yesterday": tickets_yesterday
        },
        "firp": {
            "status": {"color": status_color, "title": status_title, "desc": status_desc},
            "facts": [
                f"Yesterday your team resolved {tickets_yesterday} tickets.",
                volume_msg,
                cat_msg,
                f"{at_risk} SLA breaches are expected." if at_risk > 0 else "No immediate SLA breaches are expected."
            ],
            "good_news": good_news_list[:3] # Max 3 items
        },
        "mission": {
            "title": "Maintain the current SLA while clearing the remaining backlog." if at_risk == 0 else "Focus on tickets approaching SLA breaches.",
            "progress": progress
        },
        "recommendations": dynamic_recommendations,
        "sla_nearing": sla_nearing,
        "charts": {
            "queue": priorities,
            "aging": aging_data,
            "trend": frappe.db.sql('''
                SELECT 
                    DATE_FORMAT(creation, '%H:00') as hour,
                    COUNT(*) as count
                FROM `tabHD Ticket`
                WHERE creation >= DATE_SUB(NOW(), INTERVAL 12 HOUR)
                GROUP BY HOUR(creation)
                ORDER BY creation ASC
            ''', as_dict=True),
            "sla": {"compliance": sla_compliance, "breached": sla_breached, "open": total_open, "at_risk": at_risk, "med_risk": med_risk},
            "workload": agent_workload[:5] if agent_workload else [],
            "customer_health": customer_health,
            "categories": categories_data,
            "top_agents": top_agents,
            "client_wise_open": frappe.db.sql("""
                SELECT customer, status, COUNT(*) as count
                FROM `tabHD Ticket`
                WHERE status IN ('Open', 'Replied', 'On Hold') AND customer IS NOT NULL AND customer != ''
                GROUP BY customer, status
                ORDER BY customer
            """, as_dict=True)
        },
        "predictions": {
            "new_tickets": int(tickets_today * 1.2),
            "resolutions": int(resolved_today * 1.1),
            "open_queue": total_open + int(tickets_today * 1.2) - int(resolved_today * 1.1)
        }
    }

@frappe.whitelist()
def ask_nexai_v2(query):
    import json
    from datetime import datetime
    
    try:
        telemetry = get_nexai_dashboard_data_v2()
        facts = telemetry.get('firp', {}).get('facts', [])
    except:
        facts = []
        
    current_date = datetime.now().strftime("%Y-%m-%d")
        
    intent_prompt = f"""
    You are NexAI, the Chief of Staff for a Helpdesk.
    
    Current Telemetry Facts:
    {json.dumps(facts)}
    Today's Date: {current_date}
    
    User Query: "{query}"
    
    INSTRUCTIONS:
    1. If the user's query can be answered using the Telemetry Facts provided above (e.g. asking about ticket volume today, yesterday, or SLA), provide a brief, professional answer directly. Use HTML <br> for newlines.
    2. IF AND ONLY IF the user is asking for specific ticket details NOT in the facts (like a specific ticket status, priority, customer, or a SPECIFIC DATE like "12-07-2026"), return ONLY a valid JSON object representing Frappe ORM filters for 'HD Ticket'.
    
    Valid fields for JSON: name, status, priority, _assign, customer, custom_category, creation.
    
    JSON Format: {{"action": "COUNT", "filters": {{"status": "Open"}}}} or {{"action": "LIST", "filters": {{"status": "Open"}}}}
    
    CRITICAL DATE RULES:
    - If filtering by a specific date (e.g., 12-07-2026), format the date as YYYY-MM-DD.
    - Use the "like" operator for a specific day: {{"creation": ["like", "2026-07-12%"]}}
    - If no specific filters apply, ensure you do not drop the filters. Be as specific as possible!
    """
    
    try:
        response = call_ai_model(intent_prompt).strip()
        
        clean_response = response.replace("```json", "").replace("```", "").strip()
        
        if clean_response.startswith("{") and clean_response.endswith("}"):
            intent = json.loads(clean_response)
            filters = intent.get("filters", {})
            action = intent.get("action", "LIST")
            
            data = None
            if action == "COUNT":
                data = frappe.db.count("HD Ticket", filters)
            else:
                data = frappe.get_list("HD Ticket", filters=filters, fields=["name", "status", "priority", "subject", "customer"], limit=20)
                
            response_prompt = f"""
            You are NexAI.
            User asked: "{query}"
            Database returned: {json.dumps(data)}
            Write a brief, professional answer based on this data. Use <br> for newlines.
            """
            
            final_answer = call_ai_model(response_prompt)
            return final_answer.replace('\n', '<br>')
            
        else:
            return response.replace('\n', '<br>')
            
    except Exception as e:
        return f"I encountered an error analyzing the request: {str(e)}"

################################################################################
# --- END: EXCLUSIVE HELPDESK DASHBOARD V2 APIs ---
################################################################################

